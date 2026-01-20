#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
文件名称:compare_config_after_reboot.py
功能描述:设备基本参数设置，掉电重启后检查配置是否丢失
创建日期:2025-08-05
作者:胡赞
版本:v1.0
修改记录:
"""
from openpyxl import Workbook
from tools.excel_operate import *
import os
import time
from pathlib import Path
from result_table_tilte import TableTitle
from comm.modbus_rtu_tcp import ModbusRtuOrTcp
from modbus_config import modbus_config
from tools.log import Log
import struct
import tkinter as tk
from tkinter import messagebox
import logging
import ast

test_data_path = r'/Convenient_tools/basic_setting_checking/data/checking_data.xlsx'
save_filedir = os.path.join(Path(__file__).parent, f"Compare_BasicSetting_result_{time.strftime('%Y%m%d')}")
print(save_filedir)
print(Path(__file__))
if not os.path.exists(save_filedir):
    os.makedirs(save_filedir, exist_ok=True)


class CompareConfig:
    def __init__(self, slave_id=1):
        """
        初始化实例
        :param slave_id: 电表标志slave_id
        """
        self.slave_id = slave_id
        self.modbus_client = None
        self.log = None
        self.init_func()

    def init_func(self):
        """
        ModBus连接,log初始化
        :return:
        """
        self.modbus_client = ModbusRtuOrTcp(conn_mode=modbus_config['conn_mode'])
        self.log = Log(str(__file__).split("\\")[-1]).logger

    @staticmethod
    def get_save_filepath(filedir):
        """
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Compare_BasicSetting_result_{time.strftime('%Y%m%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def write_table_title_to_excel(file_path):
        """
        写入表头:compare result
        :param file_path: 文件路径
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for i in range(len(TableTitle.ResultTableTitle)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ResultTableTitle[i]}')
        wb.save(file_path)

    @staticmethod
    def write_comm_to_excel(file_path, comm_list):
        """
        写入表头:compare result
        :param file_path: 文件路径
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for i, row_data in enumerate(comm_list):
            for j, cell_value in enumerate(row_data[0:13]):  # 只写前13列，可根据需要调整
                ws.cell(row=i + 1, column=j+1, value=cell_value)
        wb.save(file_path)

    def write_result_data_to_excel(self, file_path, output_list, column_num=13):
        """
        function: 写读取的到的modbus中的值到excel中
        :param file_path: 文件路径
        :param output_list：读取到modbus值列表
        :param column_num: 存到那一列
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 从第二行开始读取寄存器值，然后存到excel表格中
        for i in range(len(output_list)):
            ws.cell(i + 2, column_num, str(output_list[i]))  # 存入到第14列或15列
        wb.save(file_path)

    @staticmethod
    def get_unit16_bytes_value(registers):
        """
        解析寄存器返回值
        :param registers: 寄存器返回值
        :return: 整数列表
        """
        bytes_value = []
        for value in registers:
            high_byte = (value & 0xff00) >> 8
            low_byte = (value & 0x00ff)
            bytes_value.extend([high_byte, low_byte])
        return bytes_value

    @staticmethod
    def get_uint8_list_from_regs(registers):
        """
        将 16-bit 寄存器列表解析为 uint8_t 列表（高字节在前）
        :param registers: [0x1234, 0xABCD]
        :return: [0x12, 0x34, 0xAB, 0xCD]
        """
        uint8_list = []

        for reg in registers:
            if not 0 <= reg <= 0xFFFF:
                raise ValueError(f"Invalid register value: {reg}")

            uint8_list.append((reg >> 8) & 0xFF)  # high byte
            uint8_list.append(reg & 0xFF)  # low byte

        return uint8_list

    @staticmethod
    def get_uint32_from_regs_struct(registers):
        buf = bytearray()
        for r in registers:
            buf.append((r >> 8) & 0xFF)
            buf.append(r & 0xFF)

        return struct.unpack('!I', buf)[0]

    def read_register_value(self, source_input_list, scale=1):
        """
        通用读取寄存器并解析为 float
        :param source_input_list: 需要处理数据列表
        :param datatype: 'uint16_t', 'uint8_t', 'uint32_t'
        :param scale  modbus地址放大的倍数
        :param reg modbus寄存器个数
        :return: float 类型值，如果读取失败返回 None
        """
        read_list = []
        if self.is_modbus_alive(self.modbus_client, self.slave_id):
            le = 0
            le = len(source_input_list)
            for j in range(1, len(source_input_list)):
                modbus_address = source_input_list[j][3]
                modbus_count = source_input_list[j][8]
                datatype = source_input_list[j][6]
                # 读取寄存器
                measure_value = self.modbus_client.read_measurement(modbus_address, modbus_count, self.slave_id)
                # 判断返回值是否有效
                if not measure_value or not isinstance(measure_value, list):
                    self.log.error(f"Modbus read failed: {measure_value}")
                    return None
                try:
                    value_list = []
                    if datatype == 'uint8_t':
                        # uint8_t 都是每个寄存器 2 字节，高字节在前
                        for reg in measure_value:
                            value_list.append(int((reg >> 8) & 0xFF))  # 保留高八位
                            value_list.append(int(reg & 0xFF))  # 保留低八位
                    elif datatype == 'uint16_t':
                        value_list.append(int(measure_value[0] / scale))

                    elif datatype == 'uint32_t':
                        # uint32_t 用两个寄存器组合成 4 字节
                        if len(measure_value) < 2:
                            measure_value += [0] * (2 - len(measure_value))
                        value_list.append(int(((measure_value[0] << 16) | measure_value[1])/scale))
                    read_list.append(value_list)
                except Exception as e:
                    self.log.error(f"Parse error: {e}, measure_value: {measure_value}")
                    read_list.append(None)
        else:
            print("modbus 链接异常")
        return read_list

    def restore_param_to_modbus(self, source_input_list):
        """
        设置电压接线方式
        :param source_input_list: 包含需要设置modbus地址列表
        :return:
        """
        for j in range(1, len(source_input_list)):
            modbus_address = source_input_list[j][3]
            modbus_count = source_input_list[j][8]
            modbus_slave = self.slave_id
            write_value = source_input_list[j][13]  # 写入读之前值
            datatype = source_input_list[j][6]
            modbus_rw = source_input_list[j][7]
            registers = []
            # 处理modbus地址
            # modbus_address = modbus_address.strip()  # 去掉字符串 write_value 首尾的空白字符
            # modbus_address_value = int(modbus_address)

            # 1️⃣ Excel 空单元格防御
            if write_value is None:
                self.log.warning(f"Skip write addr {modbus_address}: write_value is None")
                measure_value = self.modbus_client.read_measurement(modbus_address, modbus_count, modbus_slave)
                continue
            else:
                if datatype == "uint16_t" and modbus_rw == "R/W":
                    """
                       输入 write_value（可能是 str、list 或 int），返回安全的 uint16_t 寄存器列表
                       """
                    # 如果是字符串
                    if isinstance(write_value, str):
                        write_value = write_value.strip()  # 去掉字符串 write_value 首尾的空白字符
                        if write_value.lower() in ["none", ""]:
                            return None
                        try:
                            write_list = ast.literal_eval(write_value)  # "[56]" -> [56]
                        except Exception as e:
                            logging.error(f"Cannot parse write_value: {write_value}, {e}")
                            return None
                    # 如果是整数
                    elif isinstance(write_value, int):
                        write_list = [write_value]
                    # 如果是列表
                    elif isinstance(write_value, list):
                        write_list = write_value
                    else:
                        logging.error(f"Unsupported type for write_value: {type(write_value)}")
                        return None

                    # 转成整数
                    try:
                        # for x in write_list:
                        registers.append(int(write_list[0]))
                    except Exception as e:
                        logging.error(f"Cannot convert write_list to int: {write_list}, {e}")
                        return None

                elif datatype == "uint8_t" and modbus_rw == "R/W":
                    if isinstance(write_value, str):
                        write_value = write_value.strip()
                        if write_value == "" or write_value.lower() == "none":
                            continue
                        try:
                            write_list = ast.literal_eval(write_value)  # "[192,168,2,227]" → [192,168,2,227]
                        except Exception as e:
                            self.log.error(f"Cannot parse write_value: {write_value}, {e}")
                            continue
                    else:
                        write_list = write_value
                    write_list = [int(x) for x in write_list]
                    for i in range(0, len(write_list), 2):
                        high = write_list[i]
                        low = write_list[i + 1] if i + 1 < len(write_list) else 0  # 奇数补0
                        reg_val = (high << 8) | low
                        registers.append(int(reg_val))
                elif datatype == "uint32_t" and modbus_rw == "R/W":
                    # 2️⃣ 统一转成 list
                    try:
                        if isinstance(write_value, str):
                            write_value = write_value.strip()
                            if write_value == "" or write_value.lower() == "none":
                                continue
                            write_list = ast.literal_eval(write_value)  # "[56]" → [56]
                        else:
                            write_list = [write_value]
                    except Exception as e:
                        self.log.error(f"Parse uint32 failed at {modbus_address}: {write_value}, {e}")
                        continue
                    # 3️⃣ 校验格式
                    if not isinstance(write_list, (list, tuple)) or len(write_list) == 0:
                        self.log.error(f"Invalid uint32 write_list at {modbus_address}: {write_list}")
                        continue
                    # 4️⃣ 强制转 int（防 float / str）
                    try:
                        value = int(write_list[0])
                    except Exception as e:
                        self.log.error(f"Invalid uint32 value at {modbus_address}: {write_list}, {e}")
                        continue
                    # 5️⃣ 拆成两个 16-bit
                    high = (value >> 16) & 0xFFFF
                    low = value & 0xFFFF
                    registers = [high, low]
                print("restore输入值:", modbus_address, registers, modbus_slave)
                self.modbus_client.write_registers(modbus_address, registers, modbus_slave)
            measure_value = self.modbus_client.read_measurement(modbus_address, modbus_count, modbus_slave)

    def write_param_to_modbus(self, source_input_list):
        """
        设置电压接线方式
        :param source_input_list: 包含需要设置modbus地址列表
        :return:
        """
        for j in range(1, len(source_input_list)):
            modbus_address = source_input_list[j][3]
            modbus_count = source_input_list[j][8]
            modbus_slave = self.slave_id
            write_value = source_input_list[j][12]
            datatype = source_input_list[j][6]
            modbus_rw = source_input_list[j][7]
            registers = []

            # 1️⃣ Excel 空单元格防御
            if write_value is None:
                self.log.warning(f"Skip write addr {modbus_address}: write_value is None")
                # measure_value = self.modbus_client.read_measurement(modbus_address, modbus_count, modbus_slave)
                continue
            else:
                if datatype == "uint16_t" and modbus_rw == "R/W":
                    """
                       输入 write_value（可能是 str、list 或 int），返回安全的 uint16_t 寄存器列表
                       """
                    # 如果是字符串
                    if isinstance(write_value, str):
                        write_value = write_value.strip()  # 去掉字符串 write_value 首尾的空白字符
                        if write_value.lower() in ["none", ""]:
                            return None
                        try:
                            write_list = ast.literal_eval(write_value)  # "[56]" -> [56]
                        except Exception as e:
                            logging.error(f"Cannot parse write_value: {write_value}, {e}")
                            return None
                    # 如果是整数
                    elif isinstance(write_value, int):
                        write_list = [write_value]
                    # 如果是列表
                    elif isinstance(write_value, list):
                        write_list = write_value
                    else:
                        logging.error(f"Unsupported type for write_value: {type(write_value)}")
                        return None

                    # 转成整数
                    try:
                        # for x in write_list:
                        registers.append(int(write_list[0]))
                    except Exception as e:
                        logging.error(f"Cannot convert write_list to int: {write_list}, {e}")
                        return None

                elif datatype == "uint8_t" and modbus_rw == "R/W":
                    if isinstance(write_value, str):
                        write_value = write_value.strip()
                        if write_value == "" or write_value.lower() == "none":
                            continue
                        try:
                            write_list = ast.literal_eval(write_value)  # "[192,168,2,227]" → [192,168,2,227]
                        except Exception as e:
                            self.log.error(f"Cannot parse write_value: {write_value}, {e}")
                            continue
                    else:
                        write_list = write_value
                    write_list = [int(x) for x in write_list]
                    for i in range(0, len(write_list), 2):
                        high = write_list[i]
                        low = write_list[i + 1] if i + 1 < len(write_list) else 0  # 奇数补0
                        reg_val = (high << 8) | low
                        registers.append(int(reg_val))
                elif datatype == "uint32_t" and modbus_rw == "R/W":
                    # 2️⃣ 统一转成 list
                    try:
                        if isinstance(write_value, str):
                            write_value = write_value.strip()
                            if write_value == "" or write_value.lower() == "none":
                                continue
                            write_list = ast.literal_eval(write_value)  # "[56]" → [56]
                        else:
                            write_list = [write_value]
                    except Exception as e:
                        self.log.error(f"Parse uint32 failed at {modbus_address}: {write_value}, {e}")
                        continue
                    # 3️⃣ 校验格式
                    if not isinstance(write_list, (list, tuple)) or len(write_list) == 0:
                        self.log.error(f"Invalid uint32 write_list at {modbus_address}: {write_list}")
                        continue
                    # 4️⃣ 强制转 int（防 float / str）
                    try:
                        value = int(write_list[0])
                    except Exception as e:
                        self.log.error(f"Invalid uint32 value at {modbus_address}: {write_list}, {e}")
                        continue
                    # 5️⃣ 拆成两个 16-bit
                    high = (value >> 16) & 0xFFFF
                    low = value & 0xFFFF
                    registers = [high, low]
                print("修改输入值:", modbus_address, registers, modbus_slave)
                self.modbus_client.write_registers(modbus_address, registers, modbus_slave)
            # measure_value = self.modbus_client.read_measurement(modbus_address, modbus_count, modbus_slave)

    @staticmethod
    def power_off_reboot_prompt():
        root = tk.Tk()
        root.withdraw()  # 不显示主窗口

        messagebox.showwarning(
            title="设备掉电提示",
            message="请手动给设备掉电，然后上电。\n\n请重新给设备上电并完成重启后，再点击确定继续。"
        )
        root.destroy()

    @staticmethod
    def is_modbus_alive(client, slave_id=1):
        """
        function：判断modbus client是否断开
        return: False/ True
        """
        try:
            client.client.read_holding_registers(0, 1, slave=slave_id)
            return True
        except Exception:
            return False

    def ensure_modbus_connected(self, client, retry=100, interval=5):
        """
        检查 Modbus client 是否连接，断开则自动重连

        :param client: 当前 ModbusTcpClient（可为 None）
        :param retry: 重连次数
        :param interval: 每次重连间隔（秒）
        :return: 可用的 ModbusTcpClient
        :raises: ConnectionError
        """

        # 如果 client 为空，创建一个
        if client is None:
            client = ModbusRtuOrTcp(modbus_config['conn_mode'])

        # 尝试判断 TCP 是否断开
        if client and self.is_modbus_alive(client, self.slave_id):
            return client

        # 循环重连
        for attempt in range(1, retry + 1):
            logging.info(f"[Modbus] 尝试重连 ({attempt}/{retry})...")
            try:
                # 关闭旧连接
                try:
                    client.close()
                except Exception:
                    pass

                # 重新创建一个实例
                new_client = ModbusRtuOrTcp(modbus_config['conn_mode'])
                # TCP 的话再判断 socket 是否可用
                connected = getattr(new_client.client, 'is_socket_open', lambda: True)()
                if connected:
                    logging.info("[Modbus] 重连成功")
                    return new_client
            except Exception as e:
                logging.error(f"[Modbus] 重连异常: {e}")

            time.sleep(interval)

        raise ConnectionError("[Modbus] 重连失败，请检查设备是否已上电")

    @staticmethod
    def compare_param(output_list_before, output_list_after):
        """
        根据项目名称选择操作
        """
        compare_data_list = []
        for i in range(len(output_list_before)):
            if output_list_before[i] == output_list_after[i]:
                compare_data_list.append("Yes")
            else:
                compare_data_list.append("No")
        return compare_data_list

    def select_project(self, pj_name):
        """

        """
        if pj_name is None:
            raise ValueError("Parameter 'pj_name' cannot be None")
        # 4100项目
        if pj_name == 0:
            source_input_list = data_read(test_data_path, "AcuRev4100")
            # input_list = self.get_input_list(source_input_list)
            # 使用目录生成结果excel空文件
            file_path = self.get_save_filepath(save_filedir)
            # 对生成空excel文件中写入文件title
            self.write_table_title_to_excel(file_path)
            # 写已有excel中内容到报告excel中
            self.write_comm_to_excel(file_path, source_input_list)
            # 修改寄存器中的值
            time.sleep(3)
            self.write_param_to_modbus(source_input_list)
            # 读取寄存器中的值
            output_list_before = self.read_register_value(source_input_list)
            self.write_result_data_to_excel(file_path, output_list_before, 14)
            # 弹框提示用户重启设备
            self.power_off_reboot_prompt()
            # 确认重启后完成后，尝试重新连接Modbus client
            self.modbus_client = self.ensure_modbus_connected(self.modbus_client, 100, 5)
            # 读取寄存器中的值
            output_list_after = self.read_register_value(source_input_list)
            self.write_result_data_to_excel(file_path, output_list_after, 15)
            # 比较重启前后值，并写入到结果一栏
            output_list_compare = self.compare_param(output_list_before, output_list_after)
            self.write_result_data_to_excel(file_path, output_list_compare, 16)
            time.sleep(3)
            # 恢复寄存器中的值
            self.restore_param_to_modbus(source_input_list)

        # 1320项目
        if pj_name == 1:
            pass


if __name__ == '__main__':
    """
    :project_name: 0:4100, 1:1320
    """
    project_name = 0  # 假设要传递的参数
    script_start = CompareConfig()  # 实例化 CompareConfig 类
    script_start.select_project(project_name)  # 调用 select_project 方法并传入参数
