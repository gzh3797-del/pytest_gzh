import openpyxl
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk


def process_excel(file_path):
    """逐行检查所有列的failed，只输出超出精度的数据，并统计总failed数量，第8列空值时不参与比较"""
    try:
        # 加载Excel文件（保留原始格式，不自动转换公式）
        workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=False)
        sheet = workbook.active

        # 提取表头（第一行作为列名，处理空列）
        headers = []
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            cell_val = sheet.cell(row=1, column=col_idx).value
            if cell_val is not None:
                # 清理列名（去除空格、换行符）
                header = str(cell_val).strip().replace('\n', '').replace('\t', '')
                headers.append(header)
            else:
                headers.append(f"未命名列_{col_idx}")

        result = []
        total_data_rows = max(0, sheet.max_row - 1)  # 排除表头的实际数据行数
        total_failed_count = 0  # 统计总failed数量

        if total_data_rows == 0:
            workbook.close()
            return "Excel文件中无数据行（仅表头）", "处理完成"

        # 添加总体统计信息标题
        result.append("=" * 80)
        result.append("                检查结果概览")
        result.append("=" * 80 + "\n")

        # 逐行检查（从第2行开始，遍历所有数据行）
        for row_num in range(2, sheet.max_row + 1):
            # 提取当前行所有列的原始值
            current_row = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_num, column=col_idx)
                current_row.append(cell.value if cell.value is not None else "None")

            # 检查第8列（索引7）是否为空值
            col8_empty = False
            if len(current_row) > 7:  # 确保第8列存在
                col8_val = current_row[7]
                # 判断是否为空值（None或空字符串）
                col8_empty = (col8_val is None) or (str(col8_val).strip() == "")
                if col8_empty:
                    result.append(f"第{row_num}行提示：第8列值为空，该列不参与比较")

            # 1. 检查当前行所有列是否包含failed（不区分大小写），第8列空值时跳过
            failed_positions = []
            for col_idx, cell_val in enumerate(current_row):
                # 如果是第8列且为空值，则不参与检查
                if col_idx == 7 and col8_empty:
                    continue
                if str(cell_val).strip().lower() == "failed":
                    col_name = headers[col_idx]
                    failed_positions.append({"col_idx": col_idx, "col_name": col_name})

            # 累加总failed数量
            total_failed_count += len(failed_positions)

            # 2. 打印4-7列数据（索引3-6，对应第4-7列）
            result.append(f"=== 第{row_num}行检查结果 ===")
            result.append(f"本行failed数量：{len(failed_positions)}个")
            result.append("4-7列数据：")
            for col_idx in range(3, 7):  # 0开始索引，3=第4列，6=第7列
                if col_idx < len(current_row):
                    col_val = current_row[col_idx]
                    col_name = headers[col_idx] if col_idx < len(headers) else f"未命名列_{col_idx + 1}"
                    result.append(f"  {col_name}：{col_val}")
                else:
                    result.append(f"  第{col_idx + 1}列（未命名）：None")

            # 3. 根据是否有failed分情况处理
            if not failed_positions:
                result.append("状态：测试通过\n")
                continue

            # 4. 有failed时，逐个处理每个failed列（只记录超出精度的数据）
            result.append(f"状态：发现{len(failed_positions)}个failed列，开始精度检查")
            has_exceed_data = False  # 标记是否有超出精度的数据

            for idx, failed in enumerate(failed_positions, 1):
                col_idx = failed["col_idx"]
                col_name = failed["col_name"]
                current_failed_has_exceed = False  # 当前failed列是否有超出精度的数据

                result.append(f"\n> 第{idx}个failed列：{col_name}（第{col_idx + 1}列）")

                # 5. 区分：是否为"电压测试结果"列
                if "电压测试结果" in col_name:
                    # 特殊逻辑：取左边10列（1-9列数据，第10列精度）
                    if col_idx < 10:  # 确保左边有10列（索引0-9）
                        result.append(f"  警告：{col_name}列左边不足10列，无法进行电压精度检查")
                        continue

                    # 定义数据列和精度列索引
                    precision_col_idx = col_idx - 1  # 第10列（精度要求）
                    data_col_indices = [col_idx - i for i in range(2, 11)]  # 1-9列（数据）
                    data_col_indices.reverse()  # 保持1-9列顺序

                    # 处理精度列
                    if precision_col_idx >= len(current_row):
                        result.append(f"  错误：精度要求列（{headers[precision_col_idx]}）超出当前行范围")
                        continue
                    precision_val = current_row[precision_col_idx]
                    precision_col_name = headers[precision_col_idx]

                    # 解析精度值
                    try:
                        precision = float(str(precision_val).strip())
                    except (ValueError, TypeError):
                        result.append(f"  错误：{precision_col_name}（精度列）的值'{precision_val}'无法解析为数字")
                        continue

                    # 检查1-9列数据（只记录超出精度的）
                    result.append(f"  精度要求：{precision_col_name} = {precision}")
                    result.append(f"  超出精度的数据列：")

                    for data_idx, data_col_idx in enumerate(data_col_indices, 1):
                        # 如果是第8列且为空值，则跳过不检查
                        if data_col_idx == 7 and col8_empty:
                            result.append(f"    跳过第8列（{headers[data_col_idx]}）：值为空")
                            continue

                        if data_col_idx >= len(current_row):
                            continue  # 不显示超出行范围的信息

                        data_col_name = headers[data_col_idx]
                        data_val = current_row[data_col_idx]
                        data_str = str(data_val).strip()

                        # 解析列表格式（支持[xx,xx]或xx,xx）
                        if data_str.startswith('[') and data_str.endswith(']'):
                            list_content = data_str[1:-1].strip()
                        else:
                            list_content = data_str
                        elements = [e.strip() for e in list_content.split(',')]

                        if len(elements) < 2:
                            continue  # 不显示格式错误的信息

                        # 解析第二个元素（实际测试值）
                        try:
                            actual_val = float(elements[1])
                        except ValueError:
                            continue  # 不显示无法解析的信息

                        # 只记录超出精度的数据
                        if actual_val > precision:
                            result.append(f"    ❌ {data_col_name}超出精度")
                            result.append(f"      要求精度：{precision}")
                            result.append(f"      实际测试结果：{data_val}")
                            has_exceed_data = True
                            current_failed_has_exceed = True

                    # 如果当前failed列没有超出精度的数据，显示提示
                    if not current_failed_has_exceed:
                        result.append(f"    无超出精度的数据")

                else:
                    # 普通列逻辑：取左边4列（1-3列数据，第4列精度）
                    if col_idx < 4:  # 确保左边有4列（索引0-3）
                        result.append(f"  警告：{col_name}列左边不足4列，无法进行精度检查")
                        continue

                    # 定义数据列和精度列索引
                    precision_col_idx = col_idx - 1  # 第4列（精度要求）
                    data_col_indices = [col_idx - 4, col_idx - 3, col_idx - 2]  # 1-3列（数据）

                    # 处理精度列
                    if precision_col_idx >= len(current_row):
                        result.append(f"  错误：精度要求列超出当前行范围")
                        continue
                    precision_val = current_row[precision_col_idx]
                    precision_col_name = headers[precision_col_idx]

                    # 解析精度值
                    try:
                        precision = float(str(precision_val).strip())
                    except (ValueError, TypeError):
                        result.append(f"  错误：{precision_col_name}（精度列）的值'{precision_val}'无法解析为数字")
                        continue

                    # 检查1-3列数据（只记录超出精度的）
                    result.append(f"  精度要求：{precision_col_name} = {precision}")
                    result.append(f"  超出精度的数据列：")

                    for data_col_idx in data_col_indices:
                        # 如果是第8列且为空值，则跳过不检查
                        if data_col_idx == 7 and col8_empty:
                            result.append(f"    跳过第8列（{headers[data_col_idx]}）：值为空")
                            continue

                        if data_col_idx >= len(current_row):
                            continue  # 不显示超出行范围的信息

                        data_col_name = headers[data_col_idx]
                        data_val = current_row[data_col_idx]
                        data_str = str(data_val).strip()

                        # 解析列表格式
                        if data_str.startswith('[') and data_str.endswith(']'):
                            list_content = data_str[1:-1].strip()
                        else:
                            list_content = data_str
                        elements = [e.strip() for e in list_content.split(',')]

                        if len(elements) < 2:
                            continue  # 不显示格式错误的信息

                        # 解析第二个元素
                        try:
                            actual_val = float(elements[1])
                        except ValueError:
                            continue  # 不显示无法解析的信息

                        # 只记录超出精度的数据
                        if actual_val > precision:
                            result.append(f"    ❌ {data_col_name}超出精度")
                            result.append(f"      要求精度：{precision}")
                            result.append(f"      实际测试结果：{data_val}")
                            has_exceed_data = True
                            current_failed_has_exceed = True

                    # 如果当前failed列没有超出精度的数据，显示提示
                    if not current_failed_has_exceed:
                        result.append(f"    无超出精度的数据")

            # 行检查结束，添加分隔线
            result.append("\n" + "-" * 60 + "\n")

        workbook.close()

        # 添加总体统计结果
        result.append("\n" + "=" * 80)
        result.append(f"检查完成！总统计信息：")
        result.append(f"1. 共分析 {total_data_rows} 行数据")
        result.append(f"2. 共检查 {max_col} 列")
        result.append(f"3. 总failed数量：{total_failed_count} 个")
        result.append("=" * 80)

        # 清理空行，合并结果
        final_result = "\n".join([line for line in result if line.strip() != ""])
        return final_result, f"处理完成！总failed数量：{total_failed_count} 个"

    except Exception as e:
        return None, f"处理错误：{str(e)}（请确认Excel文件格式正确，且未被占用）"


def select_file():
    """打开Excel文件选择对话框"""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="选择待检查的Excel文件",
        filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")],
        initialdir="./"
    )
    return file_path


def show_results(results, status):
    """显示检查结果的窗口（带滚动条）"""
    window = tk.Tk()
    window.title("Excel全列Failed检查结果")
    window.geometry("950x650")
    window.resizable(True, True)

    # 顶部状态标签
    status_label = ttk.Label(
        window,
        text=status,
        font=("Arial", 11, "bold"),
        foreground="#2d3436"
    )
    status_label.pack(pady=8, padx=12, anchor="w")

    # 滚动文本区域
    frame = ttk.Frame(window, borderwidth=1, relief="solid")
    frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)

    # 垂直+水平滚动条
    scroll_y = ttk.Scrollbar(frame, orient="vertical")
    scroll_x = ttk.Scrollbar(frame, orient="horizontal")

    # 文本显示区（等宽字体，支持横向滚动）
    text_widget = tk.Text(
        frame,
        wrap="none",
        yscrollcommand=scroll_y.set,
        xscrollcommand=scroll_x.set,
        font=("Consolas", 10),
        state="normal",
        bg="#f8f9fa"
    )

    # 绑定滚动条
    scroll_y.config(command=text_widget.yview)
    scroll_x.config(command=text_widget.xview)

    # 布局
    scroll_y.pack(side="right", fill="y")
    scroll_x.pack(side="bottom", fill="x")
    text_widget.pack(fill=tk.BOTH, expand=True)

    # 填充结果
    if results:
        text_widget.insert(tk.END, results)
    else:
        text_widget.insert(tk.END, "无有效检查结果（文件为空或格式错误）")
    text_widget.config(state="disabled")

    # 关闭按钮
    close_btn = ttk.Button(
        window,
        text="关闭窗口",
        command=window.destroy,
        style="TButton"
    )
    close_btn.pack(pady=8)

    window.mainloop()


def main():
    """主流程：选择文件→处理→显示结果"""
    # 1. 选择文件
    file_path = select_file()
    if not file_path:
        tk.messagebox.showinfo("提示", "未选择任何文件，程序退出")
        return

    # 2. 显示加载提示
    loading_root = tk.Tk()
    loading_root.withdraw()
    tk.messagebox.showinfo("处理中", f"正在检查文件：\n{file_path}\n\n请耐心等待...")

    # 3. 核心处理
    results, status = process_excel(file_path)

    # 4. 显示结果
    if results is None:
        tk.messagebox.showerror("错误", status)
        return
    show_results(results, status)


if __name__ == "__main__":
    main()
