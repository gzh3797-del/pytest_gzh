"""
接线检查测试用例生成脚本
输出：knowledge/gateway/web2/testcase/接线检查_测试用例_WEB2_v1.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUTPUT = Path(r'C:\Users\ZihanGao\Desktop\testing-team\knowledge\gateway\web2\testcase\接线检查_测试用例_WEB2_v1.xlsx')

PROJ   = 'ACUREV4100WEB2'
MOD    = 'WRI'
MNAME  = 'Wiring Check'
PRE    = '设备已正常上电，相关服务正常启动'

COLUMNS = [
    '模块', '子模块', 'FTS编号', 'FTS名称(用例编号)',
    '用例编号', '用例标题', '预置条件',
    '测试步骤', '预期结果', '用例级别',
    '测试结果', '测试负责人', '备注', '完成自动化', '问题单来源',
]

HEADER_FILL = PatternFill('solid', fgColor='70AD47')
NORMAL_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT   = Font(name='微软雅黑', size=10, bold=True)
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True)
ALIGN_WRAP  = Alignment(wrap_text=True, vertical='center', horizontal='left')
ALIGN_CTR   = Alignment(wrap_text=True, vertical='center', horizontal='center')
THIN        = Side(style='thin')
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD_COLS   = {'模块', '子模块'}

COL_WIDTHS = {
    '模块': 16, '子模块': 22, 'FTS编号': 42,
    'FTS名称(用例编号)': 32, '用例编号': 42, '用例标题': 32,
    '预置条件': 30, '测试步骤': 64, '预期结果': 60,
    '用例级别': 10, '测试结果': 12, '测试负责人': 14,
    '备注': 14, '完成自动化': 12, '问题单来源': 14,
}

V  = 230
VL = 133   # ≈ 230/√3
I  = 1.0


# ── ID helpers ────────────────────────────────────────────────────────────────

def make_ids(sub3: str, seq: int):
    n = f'{seq:03d}'
    fts = f'FTS_{PROJ}_{MOD}_{sub3}_{n}'
    tc  = f'TestCase_{PROJ}_{MOD}_{sub3}_{n}'
    return fts, tc


def sig_v(va, qva, vb, qvb, vc, qvc):
    return f'Va={va}V@{qva}°  Vb={vb}V@{qvb}°  Vc={vc}V@{qvc}°'


def sig_i3(ia, qia, ib, qib, ic, qic):
    return f'Ia={ia:.2f}A@{qia}°  Ib={ib:.2f}A@{qib}°  Ic={ic:.2f}A@{qic}°'


def sig_i2(ia, qia, ic, qic):
    return f'Ia={ia:.2f}A@{qia}°  Ic={ic:.2f}A@{qic}°'


def sig_i1(ia, qia):
    return f'Ia={ia:.2f}A@{qia}°'


# ── 组装预期结果文本 ────────────────────────────────────────────────────────────

def fmt_exp(v_dict: dict, i_dict: dict, v_phases=('A', 'B', 'C'), i_phases=('A', 'B', 'C')) -> str:
    """
    v_dict: {'A': status, 'B': status, 'C': status, 'order': status}
    i_dict: {'A': status, 'B': status, 'C': status}  (N/A 相跳过)
    """
    lines = ['4. 接线检查完成，结果表 Wiring Status 如下：']
    for ph in v_phases:
        s = v_dict.get(ph, '')
        if s and s != 'N/A':
            lines.append(f'   电压侧 Phase {ph}：{s}')
    order = v_dict.get('order', '')
    if order and order not in ('Pass', 'N/A', ''):
        lines.append(f'   电压侧 Phase Order：{order}')
    for ph in i_phases:
        s = i_dict.get(ph, '')
        if s and s != 'N/A':
            lines.append(f'   电流侧 Phase {ph}：{s}')
    return '\n'.join(lines)


def make_steps(wtype, phase_order, signal_v, signal_i, phys_desc: str = '') -> str:
    if phys_desc:
        step2 = (
            f'2. 模拟接线故障：{phys_desc}\n'
            f'   （自动化等效控源输出：{signal_v} / {signal_i}）'
        )
    else:
        step2 = (
            f'2. 确认物理接线正常：按 {wtype} {phase_order} 规范接好全部电压线与 CT\n'
            f'   （自动化等效控源输出：{signal_v} / {signal_i}）'
        )
    return (
        f'1. 配置电表接线方式为 {wtype}，相序为 {phase_order}，额定电压 VRATE=230V\n'
        f'{step2}\n'
        f'3. 进入 WEB2 页面 Settings → Diagnostic → Wiring Check，点击 Wiring Check 按钮\n'
        f'4. 在 Confirm Nominal Voltage 弹窗确认 VRATE=230V，点击 Start Wiring Check'
    )


def make_expected(v_dict, i_dict, v_phases=('A', 'B', 'C'), i_phases=('A', 'B', 'C')) -> str:
    return (
        '3. 成功进入 Wiring Check 页面，弹出 Confirm Nominal Voltage 弹窗\n'
        + fmt_exp(v_dict, i_dict, v_phases, i_phases)
    )


def row(sub, sub3, seq, name, phase_order, sig_vol, sig_cur, v_dict, i_dict, level,
        v_phases=('A', 'B', 'C'), i_phases=('A', 'B', 'C'), phys: str = ''):
    fts, tc = make_ids(sub3, seq)
    steps   = make_steps(sub, phase_order, sig_vol, sig_cur, phys)
    exp     = make_expected(v_dict, i_dict, v_phases, i_phases)
    return {
        '模块':               MNAME,
        '子模块':             sub,
        'FTS编号':            fts,
        'FTS名称(用例编号)':   name,
        '用例编号':           tc,
        '用例标题':           name,
        '预置条件':           PRE,
        '测试步骤':           steps,
        '预期结果':           exp,
        '用例级别':           level,
        '测试结果':           '',
        '测试负责人':         '',
        '备注':              '',
        '完成自动化':         'Y',
        '问题单来源':         '',
    }


# ── Pass 电流状态（电流角跟随电压角 → PF=1 → Pass）──────────────────────────────
PI = {'A': 'Pass', 'B': 'Pass', 'C': 'Pass'}
PI_AC = {'A': 'Pass', 'C': 'Pass'}
PI_A  = {'A': 'Pass'}


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  3E4WY  31 条（WC-3E4-001 ~ 031）                                           ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _cases_3e4wy():
    sub, sub3 = '3E4WY', '3E4'
    cases = []
    n = 0

    def r(name, po, sv, si, vd, ci_d, lv, phys='', vph=('A','B','C'), iph=('A','B','C')):
        nonlocal n; n += 1
        return row(sub, sub3, n, name, po, sv, si, vd, ci_d, lv, vph, iph, phys)

    # ── ABC ──
    cases.append(r('3E4WY ABC 正常接线全通', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV0',
        phys='正常接线：L1→Va、L2→Vb、L3→Vc、N 端正确接入；CT1/CT2/CT3 分别套入 A/B/C 相导线，K 端朝向电源侧'))

    cases.append(r('3E4WY ABC 三相全缺失（条件1）', 'ABC',
        sig_v(0,0,0,0,0,0), sig_i3(I,0,I,0,I,0),
        {'A':'Va Wiring Missing','B':'Vb Wiring Missing','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开电表电压输入端全部电压线（L1、L2、L3 均不接），N 可保留'))

    cases.append(r('3E4WY ABC Vab 缺失（条件2）', 'ABC',
        sig_v(V,0,V,0,V,120), sig_i3(I,0,I,0,I,120),
        {'A':'Va Wiring Missing','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='断开 L1（Va）和 L2（Vb）电压输入线，仅保留 L3（Vc）正常连接'))

    cases.append(r('3E4WY ABC Vbc 缺失（条件3）', 'ABC',
        sig_v(V,0,V,240,V,240), sig_i3(I,0,I,240,I,240),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开 L2（Vb）和 L3（Vc）电压输入线，仅保留 L1（Va）正常连接'))

    cases.append(r('3E4WY ABC Vca 缺失（条件4）', 'ABC',
        sig_v(V,0,V,240,V,0), sig_i3(I,0,I,240,I,0),
        {'A':'Va Wiring Missing','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开 L1（Va）和 L3（Vc）电压输入线，仅保留 L2（Vb）正常连接'))

    # ── 条件2/3/4 guard 负向：guard 不成立时不误报双相缺失 ──────────────────────
    cases.append(r('3E4WY ABC 条件2 guard—Vcn偏低条件7触发', 'ABC',
        sig_v(V,0,V,0,172,120), sig_i3(I,0,I,0,I,120),
        {'A':'Pass','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV2',
        phys='L1/L2 断开（Vab≈0）且 L3（Vc）也偏低（<184V）；验证条件2 guard 正确拦截，不误报 Va&Vb Wiring Missing'))

    cases.append(r('3E4WY ABC 条件3 guard—Van偏低条件5触发', 'ABC',
        sig_v(172,0,V,240,V,240), sig_i3(I,0,I,240,I,240),
        {'A':'Va Wiring Missing','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV2',
        phys='L2/L3 断开（Vbc≈0）且 L1（Va）也偏低（<184V）；验证条件3 guard 正确拦截，不误报 Vb&Vc Wiring Missing'))

    cases.append(r('3E4WY ABC 条件4 guard—Vbn偏低条件6触发', 'ABC',
        sig_v(V,0,172,240,V,0), sig_i3(I,0,I,240,I,0),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI, 'LV2',
        phys='L1/L3 断开（Vca≈0）且 L2（Vb）也偏低（<184V）；验证条件4 guard 正确拦截，不误报 Va&Vc Wiring Missing'))

    cases.append(r('3E4WY ABC Va 欠压缺失（条件5）', 'ABC',
        sig_v(172,0,V,240,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Va Wiring Missing','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='断开或松动 L1（Va）电压输入线，使 Va 电压低于 0.8×VRATE（<184V）'))

    cases.append(r('3E4WY ABC Vb 欠压缺失（条件6）', 'ABC',
        sig_v(V,0,172,240,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='断开或松动 L2（Vb）电压输入线，使 Vb 电压低于 0.8×VRATE（<184V）'))

    cases.append(r('3E4WY ABC Vc 欠压缺失（条件7）', 'ABC',
        sig_v(V,0,V,240,172,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开或松动 L3（Vc）电压输入线，使 Vc 电压低于 0.8×VRATE（<184V）'))

    cases.append(r('3E4WY ABC Va-Vn 反接（条件8）', 'ABC',
        sig_v(V,0,400,30,400,330), sig_i3(I,0,I,30,I,330),
        {'A':'Va-Vn Reversed','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='将 L1（Va）与 N 接线在电表电压端子处对调（Va 相对 N 极性反向）'))

    cases.append(r('3E4WY ABC Vb-Vn 反接（条件9）', 'ABC',
        sig_v(400,0,V,30,400,60), sig_i3(I,0,I,30,I,60),
        {'A':'Pass','B':'Vb-Vn Reversed','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='将 L2（Vb）与 N 接线在电表电压端子处对调（Vb 相对 N 极性反向）'))

    cases.append(r('3E4WY ABC Vc-Vn 反接（条件10）', 'ABC',
        sig_v(400,0,400,300,V,330), sig_i3(I,0,I,300,I,330),
        {'A':'Pass','B':'Pass','C':'Vc-Vn Reversed','order':'Pass'}, PI, 'LV1',
        phys='将 L3（Vc）与 N 接线在电表电压端子处对调（Vc 相对 N 极性反向）'))

    cases.append(r('3E4WY ABC Vb 相位错误（条件11）', 'ABC',
        sig_v(V,0,V,180,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Phase Shift','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='将 L2 错误接至偏移 >20° 的非正常相位（如接至非标位置或接线松动导致相位偏移）'))

    cases.append(r('3E4WY ABC Vc 相位错误（条件12）', 'ABC',
        sig_v(V,0,V,240,V,60), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Phase Shift','order':'Pass'}, PI, 'LV1',
        phys='将 L3 错误接至偏移 >20° 的非正常相位（如接至非标位置或接线松动导致相位偏移）'))

    cases.append(r('3E4WY ABC 相序错误+VB/VC Phase Shift（条件11+12+13）', 'ABC',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Phase Shift','C':'Phase Shift','order':'Phase Order Error'}, PI, 'LV1',
        phys='将电压接线按 ACB 相序错误接入：L1→A 相、L2→C 相电源端、L3→B 相电源端'))

    cases.append(r('3E4WY ABC Ia 缺失（条件14）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(0.05,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Wiring Missing','B':'Pass','C':'Pass'}, 'LV1',
        phys='拔除或断开 CT1（A 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('3E4WY ABC Ib 缺失（条件15）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,0.05,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Wiring Missing','C':'Pass'}, 'LV1',
        phys='拔除或断开 CT2（B 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('3E4WY ABC Ic 缺失（条件16）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,0.05,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Wiring Missing'}, 'LV1',
        phys='拔除或断开 CT3（C 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('3E4WY ABC Ia 极性反接（条件17，PF=-1）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,180,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Polarity Reversed','B':'Pass','C':'Pass'}, 'LV1',
        phys='将 CT1（A 相）二次侧 K/L 接线对调（极性反接）'))

    cases.append(r('3E4WY ABC Ib 极性反接（条件18，PF=-1）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,60,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Polarity Reversed','C':'Pass'}, 'LV1',
        phys='将 CT2（B 相）二次侧 K/L 接线对调（极性反接）'))

    cases.append(r('3E4WY ABC Ic 极性反接（条件19，PF=-1）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,I,300),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Polarity Reversed'}, 'LV1',
        phys='将 CT3（C 相）二次侧 K/L 接线对调（极性反接）'))

    cases.append(r('3E4WY ABC Ia 相位错误（条件20，PF=0）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,90,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Phase Shift','B':'Pass','C':'Pass'}, 'LV1',
        phys='将 CT1 套入错误相位导线（如 B 相或 C 相），使 Ia 测量相位偏移 ≥90°'))

    cases.append(r('3E4WY ABC Ib 相位错误（条件21，PF=0）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,330,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Phase Shift','C':'Pass'}, 'LV1',
        phys='将 CT2 套入错误相位导线（如 A 相或 C 相），使 Ib 测量相位偏移 ≥90°'))

    cases.append(r('3E4WY ABC Ic 相位错误（条件22，PF=0）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,I,210),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Phase Shift'}, 'LV1',
        phys='将 CT3 套入错误相位导线（如 A 相或 B 相），使 Ic 测量相位偏移 ≥90°'))

    # ── ACB ──
    cases.append(r('3E4WY ACB 正常接线全通', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='正常接线 ACB 相序：L1→A 相、L2→C 相电源端、L3→B 相电源端；CT1/CT2/CT3 方向正确'))

    cases.append(r('3E4WY ACB Vb 相位错误（条件11）', 'ACB',
        sig_v(V,0,V,60,V,240), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Phase Shift','C':'Pass','order':'Pass'}, PI, 'LV3',
        phys='ACB 相序配置下，将 L2 接至偏移 >20° 的非正常相位'))

    cases.append(r('3E4WY ACB Vc 相位错误（条件12）', 'ACB',
        sig_v(V,0,V,120,V,180), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Pass','C':'Phase Shift','order':'Pass'}, PI, 'LV3',
        phys='ACB 相序配置下，将 L3 接至偏移 >20° 的非正常相位'))

    cases.append(r('3E4WY ACB 相序错误+VB/VC Phase Shift（条件11+12+13）', 'ACB',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Phase Shift','C':'Phase Shift','order':'Phase Order Error'}, PI, 'LV3',
        phys='设备配置为 ACB 相序，但接线按 ABC 相序接入（L1→A、L2→B、L3→C 电源端）'))

    cases.append(r('3E4WY ACB Ib 极性反接（条件18）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,300,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Polarity Reversed','C':'Pass'}, 'LV3',
        phys='ACB 相序配置下，将 CT2（B 相）K/L 接线对调'))

    cases.append(r('3E4WY ACB Ic 极性反接（条件19）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,120,I,60),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Polarity Reversed'}, 'LV3',
        phys='ACB 相序配置下，将 CT3（C 相）K/L 接线对调'))

    cases.append(r('3E4WY ACB Ib 相位错误（条件21）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,210,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Phase Shift','C':'Pass'}, 'LV3',
        phys='ACB 相序配置下，将 CT2 套入错误相位导线，使 Ib 测量相位偏移 ≥90°'))

    cases.append(r('3E4WY ACB Ic 相位错误（条件22）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,120,I,330),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Phase Shift'}, 'LV3',
        phys='ACB 相序配置下，将 CT3 套入错误相位导线，使 Ic 测量相位偏移 ≥90°'))

    return cases


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  2E3W Delta  21 条（WC-DLT-001 ~ 021）                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _cases_delta():
    sub, sub3 = '2E3W Delta', 'DLT'
    cases = []
    n = 0

    def r(name, po, sv, si, vd, ci_d, lv, phys=''):
        nonlocal n; n += 1
        return row(sub, sub3, n, name, po, sv, si, vd, ci_d, lv,
                   v_phases=('A','B','C'), i_phases=('A','C'), phys=phys)

    # ABC
    cases.append(r('2E3W Delta ABC 正常接线全通', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,0,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'}, PI_AC, 'LV0',
        phys='正常接线：L1/L2/L3 按 Delta 方式接入电表线电压端子；CT-A 套 A 相导线（K 朝电源侧）、CT-C 套 C 相导线'))

    cases.append(r('2E3W Delta ABC 三相全缺失（条件1）', 'ABC',
        sig_v(0,0,0,0,0,0), sig_i2(I,0,I,0),
        {'A':'Va Wiring Missing','B':'Vb Wiring Missing','C':'Vc Wiring Missing','order':'Pass'}, PI_AC, 'LV1',
        phys='断开全部电压输入线（L1、L2、L3 均不接到电表）'))

    cases.append(r('2E3W Delta ABC Vab 缺失（条件2）', 'ABC',
        sig_v(VL,0,VL,0,VL,120), sig_i2(I,0,I,120),
        {'A':'Va Wiring Missing','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI_AC, 'LV1',
        phys='断开 L1 和 L2（Vab 端子两侧均不接），仅保留 Vbc/Vca 正常'))

    cases.append(r('2E3W Delta ABC Vbc 缺失（条件3）', 'ABC',
        sig_v(VL,0,VL,240,VL,240), sig_i2(I,0,I,240),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Vc Wiring Missing','order':'Pass'}, PI_AC, 'LV1',
        phys='断开 L2 和 L3（Vbc 端子两侧均不接），仅保留 Vab/Vca 正常'))

    cases.append(r('2E3W Delta ABC Vca 缺失（条件4）', 'ABC',
        sig_v(VL,0,VL,240,VL,0), sig_i2(I,0,I,0),
        {'A':'Va Wiring Missing','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI_AC, 'LV1',
        phys='断开 L3 和 L1（Vca 端子两侧均不接），仅保留 Vab/Vbc 正常'))

    cases.append(r('2E3W Delta ABC Va 缺失—Vbc 正常（条件5）', 'ABC',
        sig_v(0,0,VL,240,VL,120), sig_i2(I,0,I,120),
        {'A':'Va Wiring Missing','B':'Pass','C':'Pass','order':'Pass'}, PI_AC, 'LV1',
        phys='仅断开 L1（Va）接线，L2/L3 保持正常连接（Vbc 正常，Vab 和 Vca 偏低）'))

    cases.append(r('2E3W Delta ABC Vb 缺失—Vca 正常（条件6）', 'ABC',
        sig_v(VL,0,0,0,VL,120), sig_i2(I,0,I,120),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI_AC, 'LV1',
        phys='仅断开 L2（Vb）接线，L1/L3 保持正常连接（Vca 正常，Vab 和 Vbc 偏低）'))

    cases.append(r('2E3W Delta ABC Vc 缺失—Vab 正常（条件7）', 'ABC',
        sig_v(VL,0,VL,240,0,0), sig_i2(I,0,I,0),
        {'A':'Pass','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI_AC, 'LV1',
        phys='仅断开 L3（Vc）接线，L1/L2 保持正常连接（Vab 正常，Vbc 和 Vca 偏低）'))

    cases.append(r('2E3W Delta ABC 相序错误（条件8，ACB→ABC）', 'ABC',
        sig_v(VL,0,VL,120,VL,240), sig_i2(I,0,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Phase Order Error'}, PI_AC, 'LV1',
        phys='设备配置为 ABC 相序，但电压接线按 ACB 顺序接入（L1→A 相、L2→C 相电源端、L3→B 相电源端）'))

    cases.append(r('2E3W Delta ABC Ia 缺失（条件9）', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(0.05,330,I,90),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Wiring Missing','C':'Pass'}, 'LV1',
        phys='拔除或断开 CT-A（A 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W Delta ABC Ic 缺失（条件10）', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,330,0.05,90),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','C':'Wiring Missing'}, 'LV1',
        phys='拔除或断开 CT-C（C 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W Delta ABC Ia 极性反接（条件11，∠IA_rel=150°）', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,180,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Polarity Reversed','C':'Pass'}, 'LV1',
        phys='将 CT-A K/L 接线对调（极性反接），CT-A 安装方向装反'))

    cases.append(r('2E3W Delta ABC Ic 极性反接（条件12，∠IC_rel=270°）', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,0,I,300),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','C':'Polarity Reversed'}, 'LV1',
        phys='将 CT-C K/L 接线对调（极性反接），CT-C 安装方向装反'))

    cases.append(r('2E3W Delta ABC Ia 相位错误（条件13，∠IA_rel=90°）', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,120,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Phase Shift','C':'Pass'}, 'LV1',
        phys='将 CT-A 套在错误相位导线上（如套在 B 相或 C 相导线），使测量相位偏移 ≥90°'))

    cases.append(r('2E3W Delta ABC Ic 相位错误（条件14，∠IC_rel=180°）', 'ABC',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,0,I,210),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','C':'Phase Shift'}, 'LV1',
        phys='将 CT-C 套在错误相位导线上（如套在 A 相或 B 相导线），使测量相位偏移 ≥90°'))

    # ACB
    cases.append(r('2E3W Delta ACB 正常接线全通', 'ACB',
        sig_v(VL,0,VL,120,VL,240), sig_i2(I,0,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'}, PI_AC, 'LV1',
        phys='ACB 相序正常接线：L1→A 相电源、L2→C 相电源、L3→B 相电源；CT-A/CT-C 方向正确'))

    cases.append(r('2E3W Delta ACB 相序错误（条件8，ABC→ACB）', 'ACB',
        sig_v(VL,0,VL,240,VL,120), sig_i2(I,0,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Phase Order Error'}, PI_AC, 'LV3',
        phys='设备配置为 ACB 相序，但电压接线按 ABC 顺序接入（L1→A、L2→B、L3→C 电源端）'))

    cases.append(r('2E3W Delta ACB Ia 极性反接（条件11，∠IA_rel=210°）', 'ACB',
        sig_v(VL,0,VL,120,VL,240), sig_i2(I,180,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Polarity Reversed','C':'Pass'}, 'LV3',
        phys='ACB 相序配置下，将 CT-A K/L 接线对调（极性反接）'))

    cases.append(r('2E3W Delta ACB Ic 极性反接（条件12，∠IC_rel=90°）', 'ACB',
        sig_v(VL,0,VL,120,VL,240), sig_i2(I,0,I,60),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','C':'Polarity Reversed'}, 'LV3',
        phys='ACB 相序配置下，将 CT-C K/L 接线对调（极性反接）'))

    cases.append(r('2E3W Delta ACB Ia 相位错误（条件13，∠IA_rel=90°）', 'ACB',
        sig_v(VL,0,VL,120,VL,240), sig_i2(I,60,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Phase Shift','C':'Pass'}, 'LV3',
        phys='ACB 相序配置下，将 CT-A 套在错误相位导线，使测量相位偏移 ≥90°'))

    cases.append(r('2E3W Delta ACB Ic 相位错误（条件14，∠IC_rel=180°）', 'ACB',
        sig_v(VL,0,VL,120,VL,240), sig_i2(I,0,I,150),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','C':'Phase Shift'}, 'LV3',
        phys='ACB 相序配置下，将 CT-C 套在错误相位导线，使测量相位偏移 ≥90°'))

    # ── 条件5/6 ||→&& 防回归（2 条，条件7无法构造单侧回归）────────────────────
    # ang_Vab ≈ 345° → 正常电流角：ia@315°, ic@75°
    cases.append(r('2E3W Delta 条件5 ||防&&回归—Vab单侧偏低Va Missing', 'ABC',
        sig_v(VL,270,VL,240,VL,120), sig_i2(I,315,I,75),
        {'A':'Va Wiring Missing','B':'Pass','C':'Pass','order':'Pass'}, PI_AC, 'LV2',
        phys='Va 相位偏至≈270°（近Vb@240°），Vab≈69V单侧偏低、Vca≈257V正常；'
             '|| 触发 Va Missing，&& 不触发（漏报Pass），防回归 && bug'))

    cases.append(r('2E3W Delta 条件6 ||防&&回归—Vbc单侧偏低Vb Missing', 'ABC',
        sig_v(VL,0,VL,150,VL,120), sig_i2(I,315,I,75),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI_AC, 'LV2',
        phys='Vb 相位偏至≈150°（近Vc@120°），Vbc≈69V单侧偏低、Vab≈257V正常；'
             '|| 触发 Vb Missing，&& 不触发（漏报Pass），防回归 && bug'))

    return cases


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  2E3W Network  31 条（WC-NET-001 ~ 031，算法同 3E4WY）                       ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _cases_network():
    sub, sub3 = '2E3W Network', 'NET'
    cases = []
    n = 0

    def r(name, po, sv, si, vd, ci_d, lv, phys='', vph=('A','B','C'), iph=('A','B','C')):
        nonlocal n; n += 1
        return row(sub, sub3, n, name, po, sv, si, vd, ci_d, lv, vph, iph, phys)

    # 与 3E4WY 完全相同的信号，接线方式标注改为 2E3W Network
    cases.append(r('2E3W Network ABC 正常接线全通', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV0',
        phys='正常接线：L1→Va、L2→Vb、L3→Vc、N 端正确接入；CT1/CT2/CT3 分别套入 A/B/C 相导线，K 端朝向电源侧'))

    cases.append(r('2E3W Network ABC 三相全缺失（条件1）', 'ABC',
        sig_v(0,0,0,0,0,0), sig_i3(I,0,I,0,I,0),
        {'A':'Va Wiring Missing','B':'Vb Wiring Missing','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开电表电压输入端全部电压线（L1、L2、L3 均不接），N 可保留'))

    cases.append(r('2E3W Network ABC Vab 缺失（条件2）', 'ABC',
        sig_v(V,0,V,0,V,120), sig_i3(I,0,I,0,I,120),
        {'A':'Va Wiring Missing','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='断开 L1（Va）和 L2（Vb）电压输入线，仅保留 L3（Vc）正常连接'))

    cases.append(r('2E3W Network ABC Vbc 缺失（条件3）', 'ABC',
        sig_v(V,0,V,240,V,240), sig_i3(I,0,I,240,I,240),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开 L2（Vb）和 L3（Vc）电压输入线，仅保留 L1（Va）正常连接'))

    cases.append(r('2E3W Network ABC Vca 缺失（条件4）', 'ABC',
        sig_v(V,0,V,240,V,0), sig_i3(I,0,I,240,I,0),
        {'A':'Va Wiring Missing','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开 L1（Va）和 L3（Vc）电压输入线，仅保留 L2（Vb）正常连接'))

    # ── 条件2/3/4 guard 负向（算法同 3E4WY）─────────────────────────────────────
    cases.append(r('2E3W Network ABC 条件2 guard—Vcn偏低条件7触发', 'ABC',
        sig_v(V,0,V,0,172,120), sig_i3(I,0,I,0,I,120),
        {'A':'Pass','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV2',
        phys='L1/L2 断开（Vab≈0）且 L3（Vc）也偏低（<184V）；验证条件2 guard 正确拦截'))

    cases.append(r('2E3W Network ABC 条件3 guard—Van偏低条件5触发', 'ABC',
        sig_v(172,0,V,240,V,240), sig_i3(I,0,I,240,I,240),
        {'A':'Va Wiring Missing','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV2',
        phys='L2/L3 断开（Vbc≈0）且 L1（Va）也偏低（<184V）；验证条件3 guard 正确拦截'))

    cases.append(r('2E3W Network ABC 条件4 guard—Vbn偏低条件6触发', 'ABC',
        sig_v(V,0,172,240,V,0), sig_i3(I,0,I,240,I,0),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI, 'LV2',
        phys='L1/L3 断开（Vca≈0）且 L2（Vb）也偏低（<184V）；验证条件4 guard 正确拦截'))

    cases.append(r('2E3W Network ABC Va 欠压缺失（条件5）', 'ABC',
        sig_v(172,0,V,240,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Va Wiring Missing','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='断开或松动 L1（Va）电压输入线，使 Va 电压低于 0.8×VRATE（<184V）'))

    cases.append(r('2E3W Network ABC Vb 欠压缺失（条件6）', 'ABC',
        sig_v(V,0,172,240,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Vb Wiring Missing','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='断开或松动 L2（Vb）电压输入线，使 Vb 电压低于 0.8×VRATE（<184V）'))

    cases.append(r('2E3W Network ABC Vc 欠压缺失（条件7）', 'ABC',
        sig_v(V,0,V,240,172,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Vc Wiring Missing','order':'Pass'}, PI, 'LV1',
        phys='断开或松动 L3（Vc）电压输入线，使 Vc 电压低于 0.8×VRATE（<184V）'))

    cases.append(r('2E3W Network ABC Va-Vn 反接（条件8）', 'ABC',
        sig_v(V,0,400,30,400,330), sig_i3(I,0,I,30,I,330),
        {'A':'Va-Vn Reversed','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='将 L1（Va）与 N 接线在电表电压端子处对调（Va 相对 N 极性反向）'))

    cases.append(r('2E3W Network ABC Vb-Vn 反接（条件9）', 'ABC',
        sig_v(400,0,V,30,400,60), sig_i3(I,0,I,30,I,60),
        {'A':'Pass','B':'Vb-Vn Reversed','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='将 L2（Vb）与 N 接线在电表电压端子处对调（Vb 相对 N 极性反向）'))

    cases.append(r('2E3W Network ABC Vc-Vn 反接（条件10）', 'ABC',
        sig_v(400,0,400,300,V,330), sig_i3(I,0,I,300,I,330),
        {'A':'Pass','B':'Pass','C':'Vc-Vn Reversed','order':'Pass'}, PI, 'LV1',
        phys='将 L3（Vc）与 N 接线在电表电压端子处对调（Vc 相对 N 极性反向）'))

    cases.append(r('2E3W Network ABC Vb 相位错误（条件11）', 'ABC',
        sig_v(V,0,V,180,V,120), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Phase Shift','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='将 L2 错误接至偏移 >20° 的非正常相位'))

    cases.append(r('2E3W Network ABC Vc 相位错误（条件12）', 'ABC',
        sig_v(V,0,V,240,V,60), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Phase Shift','order':'Pass'}, PI, 'LV1',
        phys='将 L3 错误接至偏移 >20° 的非正常相位'))

    cases.append(r('2E3W Network ABC 相序错误+VB/VC Phase Shift（条件11+12+13）', 'ABC',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,240,I,120),
        {'A':'Pass','B':'Phase Shift','C':'Phase Shift','order':'Phase Order Error'}, PI, 'LV1',
        phys='将电压接线按 ACB 相序错误接入：L1→A 相、L2→C 相电源端、L3→B 相电源端'))

    cases.append(r('2E3W Network ABC Ia 缺失（条件14）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(0.05,0,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Wiring Missing','B':'Pass','C':'Pass'}, 'LV1',
        phys='拔除或断开 CT1（A 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W Network ABC Ib 缺失（条件15）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,0.05,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Wiring Missing','C':'Pass'}, 'LV1',
        phys='拔除或断开 CT2（B 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W Network ABC Ic 缺失（条件16）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,0.05,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Wiring Missing'}, 'LV1',
        phys='拔除或断开 CT3（C 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W Network ABC Ia 极性反接（条件17）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,180,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Polarity Reversed','B':'Pass','C':'Pass'}, 'LV1',
        phys='将 CT1（A 相）二次侧 K/L 接线对调（极性反接）'))

    cases.append(r('2E3W Network ABC Ib 极性反接（条件18）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,60,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Polarity Reversed','C':'Pass'}, 'LV1',
        phys='将 CT2（B 相）二次侧 K/L 接线对调（极性反接）'))

    cases.append(r('2E3W Network ABC Ic 极性反接（条件19）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,I,300),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Polarity Reversed'}, 'LV1',
        phys='将 CT3（C 相）二次侧 K/L 接线对调（极性反接）'))

    cases.append(r('2E3W Network ABC Ia 相位错误（条件20）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,90,I,240,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Phase Shift','B':'Pass','C':'Pass'}, 'LV1',
        phys='将 CT1 套入错误相位导线（如 B 相或 C 相），使 Ia 测量相位偏移 ≥90°'))

    cases.append(r('2E3W Network ABC Ib 相位错误（条件21）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,330,I,120),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Phase Shift','C':'Pass'}, 'LV1',
        phys='将 CT2 套入错误相位导线（如 A 相或 C 相），使 Ib 测量相位偏移 ≥90°'))

    cases.append(r('2E3W Network ABC Ic 相位错误（条件22）', 'ABC',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,240,I,210),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Phase Shift'}, 'LV1',
        phys='将 CT3 套入错误相位导线（如 A 相或 B 相），使 Ic 测量相位偏移 ≥90°'))

    # ACB
    cases.append(r('2E3W Network ACB 正常接线全通', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'}, PI, 'LV1',
        phys='正常接线 ACB 相序：L1→A 相、L2→C 相电源端、L3→B 相电源端；CT1/CT2/CT3 方向正确'))

    cases.append(r('2E3W Network ACB Vb 相位错误（条件11）', 'ACB',
        sig_v(V,0,V,60,V,240), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Phase Shift','C':'Pass','order':'Pass'}, PI, 'LV3',
        phys='ACB 相序配置下，将 L2 接至偏移 >20° 的非正常相位'))

    cases.append(r('2E3W Network ACB Vc 相位错误（条件12）', 'ACB',
        sig_v(V,0,V,120,V,180), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Pass','C':'Phase Shift','order':'Pass'}, PI, 'LV3',
        phys='ACB 相序配置下，将 L3 接至偏移 >20° 的非正常相位'))

    cases.append(r('2E3W Network ACB 相序错误+VB/VC Phase Shift（条件11+12+13）', 'ACB',
        sig_v(V,0,V,240,V,120), sig_i3(I,0,I,120,I,240),
        {'A':'Pass','B':'Phase Shift','C':'Phase Shift','order':'Phase Order Error'}, PI, 'LV3',
        phys='设备配置为 ACB 相序，但接线按 ABC 相序接入（L1→A、L2→B、L3→C 电源端）'))

    cases.append(r('2E3W Network ACB Ib 极性反接（条件18）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,300,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Polarity Reversed','C':'Pass'}, 'LV3',
        phys='ACB 相序配置下，将 CT2（B 相）K/L 接线对调'))

    cases.append(r('2E3W Network ACB Ic 极性反接（条件19）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,120,I,60),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Polarity Reversed'}, 'LV3',
        phys='ACB 相序配置下，将 CT3（C 相）K/L 接线对调'))

    cases.append(r('2E3W Network ACB Ib 相位错误（条件21）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,210,I,240),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Phase Shift','C':'Pass'}, 'LV3',
        phys='ACB 相序配置下，将 CT2 套入错误相位导线，使 Ib 测量相位偏移 ≥90°'))

    cases.append(r('2E3W Network ACB Ic 相位错误（条件22）', 'ACB',
        sig_v(V,0,V,120,V,240), sig_i3(I,0,I,120,I,330),
        {'A':'Pass','B':'Pass','C':'Pass','order':'Pass'},
        {'A':'Pass','B':'Pass','C':'Phase Shift'}, 'LV3',
        phys='ACB 相序配置下，将 CT3 套入错误相位导线，使 Ic 测量相位偏移 ≥90°'))

    return cases


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  2E3W 1Phase  12 条（WC-1PH-001 ~ 012）                                     ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _cases_1phase():
    sub, sub3 = '2E3W 1Phase', '1PH'
    cases = []
    n = 0

    def r(name, sv, si, vd, ci_d, lv, phys=''):
        nonlocal n; n += 1
        return row(sub, sub3, n, name, 'ABC', sv, si, vd, ci_d, lv,
                   v_phases=('A','C'), i_phases=('A','C'), phys=phys)

    cases.append(r('2E3W 1Phase 正常接线全通', sig_v(V,0,0,0,V,180), sig_i2(I,0,I,180),
        {'A':'Pass','C':'Pass'}, PI_AC, 'LV0',
        phys='正常接线：L1→Va、L3→Vc、N 端接入；CT-A 套 A 相导线（K 朝电源侧）、CT-C 套 C 相导线'))

    cases.append(r('2E3W 1Phase Va&Vc 全缺失（条件1，Vca<0.1*VRATE）',
        sig_v(0,0,0,0,0,0), sig_i2(I,0,I,0),
        {'A':'Va Wiring Missing','C':'Vc Wiring Missing'}, PI_AC, 'LV1',
        phys='同时断开 L1（Va）和 L3（Vc）电压输入线（两路电压线均未接）'))

    cases.append(r('2E3W 1Phase Va 欠压缺失（条件2）', sig_v(172,0,0,0,V,180), sig_i2(I,0,I,180),
        {'A':'Va Wiring Missing','C':'Pass'}, PI_AC, 'LV1',
        phys='断开或松动 L1（Va）电压线，使 Van < 0.8×VRATE（<184V）；L3（Vc）正常'))

    cases.append(r('2E3W 1Phase Vc 欠压缺失（条件3）', sig_v(V,0,0,0,172,180), sig_i2(I,0,I,180),
        {'A':'Pass','C':'Vc Wiring Missing'}, PI_AC, 'LV1',
        phys='断开或松动 L3（Vc）电压线，使 Vcn < 0.8×VRATE（<184V）；L1（Va）正常'))

    cases.append(r('2E3W 1Phase Va-Vn 反接（条件4，Vc=400V）', sig_v(V,0,0,0,400,180), sig_i2(I,0,I,180),
        {'A':'Va-Vn Reversed','C':'Pass'}, PI_AC, 'LV1',
        phys='将 L1（Va）与 N 接线对调（Va 极性反接），L3（Vc）保持正常'))

    cases.append(r('2E3W 1Phase Vc-Vn 反接（条件5，Va=400V）', sig_v(400,0,0,0,V,180), sig_i2(I,0,I,180),
        {'A':'Pass','C':'Vc-Vn Reversed'}, PI_AC, 'LV1',
        phys='将 L3（Vc）与 N 接线对调（Vc 极性反接），L1（Va）保持正常'))

    cases.append(r('2E3W 1Phase Ia 缺失（条件6）', sig_v(V,0,0,0,V,180), sig_i2(0.05,0,I,180),
        {'A':'Pass','C':'Pass'}, {'A':'Wiring Missing','C':'Pass'}, 'LV1',
        phys='拔除或断开 CT-A（A 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W 1Phase Ic 缺失（条件7）', sig_v(V,0,0,0,V,180), sig_i2(I,0,0.05,180),
        {'A':'Pass','C':'Pass'}, {'A':'Pass','C':'Wiring Missing'}, 'LV1',
        phys='拔除或断开 CT-C（C 相电流互感器）二次侧 K/L 端接线'))

    cases.append(r('2E3W 1Phase Ia 极性反接（条件8，PF=-1）', sig_v(V,0,0,0,V,180), sig_i2(I,180,I,180),
        {'A':'Pass','C':'Pass'}, {'A':'Polarity Reversed','C':'Pass'}, 'LV1',
        phys='将 CT-A K/L 接线对调（A 相 CT 极性反接）'))

    cases.append(r('2E3W 1Phase Ic 极性反接（条件9，PF=-1）', sig_v(V,0,0,0,V,180), sig_i2(I,0,I,0),
        {'A':'Pass','C':'Pass'}, {'A':'Pass','C':'Polarity Reversed'}, 'LV1',
        phys='将 CT-C K/L 接线对调（C 相 CT 极性反接）'))

    cases.append(r('2E3W 1Phase Ia 相位错误（条件10，PF=0）', sig_v(V,0,0,0,V,180), sig_i2(I,90,I,180),
        {'A':'Pass','C':'Pass'}, {'A':'Phase Shift','C':'Pass'}, 'LV1',
        phys='将 CT-A 套入错误相位导线，使 A 相电流测量相位偏移 ≥90°'))

    cases.append(r('2E3W 1Phase Ic 相位错误（条件11，PF=0）', sig_v(V,0,0,0,V,180), sig_i2(I,0,I,90),
        {'A':'Pass','C':'Pass'}, {'A':'Pass','C':'Phase Shift'}, 'LV1',
        phys='将 CT-C 套入错误相位导线，使 C 相电流测量相位偏移 ≥90°'))

    return cases


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  1E2W  4 条（WC-1E2-001 ~ 004）                                              ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def _cases_1e2w():
    sub, sub3 = '1E2W', '1E2'
    cases = []
    n = 0

    def r(name, sv, si, vd, ci_d, lv, phys=''):
        nonlocal n; n += 1
        return row(sub, sub3, n, name, 'ABC', sv, si, vd, ci_d, lv,
                   v_phases=('A',), i_phases=('A',), phys=phys)

    cases.append(r('1E2W 正常接线全通', sig_v(V,0,0,0,0,0), sig_i1(I,0),
        {'A':'Pass'}, PI_A, 'LV0',
        phys='正常接线：L1→Va、N 端接入；CT-A 套 A 相导线（K 端朝向电源侧）'))

    cases.append(r('1E2W Va 缺失（Van<0.1*VRATE）', sig_v(0,0,0,0,0,0), sig_i1(I,0),
        {'A':'Va Wiring Missing'}, PI_A, 'LV1',
        phys='断开 L1（Va）电压输入线（电压端子 Va 未接线）'))

    cases.append(r('1E2W Ia 缺失（Ian<0.1A）', sig_v(V,0,0,0,0,0), sig_i1(0.05,0),
        {'A':'Pass'}, {'A':'Wiring Missing'}, 'LV1',
        phys='拔除或断开 CT-A 二次侧 K/L 端接线（A 相电流互感器未接入）'))

    cases.append(r('1E2W Ia 极性反接（PF=-1）', sig_v(V,0,0,0,0,0), sig_i1(I,180),
        {'A':'Pass'}, {'A':'Polarity Reversed'}, 'LV1',
        phys='将 CT-A K/L 接线对调（A 相 CT 极性反接，安装方向装反）'))

    return cases


# ── 生成 Excel ─────────────────────────────────────────────────────────────────

def build_excel(cases, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Wiring Check'

    # 表头
    for ci, cn in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = ALIGN_CTR
        cell.border    = ALL_BORDERS
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = COL_WIDTHS.get(cn, 16)

    for ri, case in enumerate(cases, 2):
        for ci, cn in enumerate(COLUMNS, 1):
            val  = case.get(cn, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = BOLD_FONT if cn in BOLD_COLS else NORMAL_FONT
            cell.alignment = ALIGN_WRAP
            cell.border    = ALL_BORDERS

        max_lines = max(
            str(case.get(c, '')).count('\n') + 1
            for c in ('测试步骤', '预期结果', '预置条件')
        )
        ws.row_dimensions[ri].height = max(18, max_lines * 16)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'已生成：{output_path}')


if __name__ == '__main__':
    all_cases = (
        _cases_3e4wy()     +   # 31
        _cases_delta()     +   # 21
        _cases_network()   +   # 31
        _cases_1phase()    +   # 12
        _cases_1e2w()          #  4
    )
    print(f'总计 {len(all_cases)} 条用例')
    build_excel(all_cases, OUTPUT)
