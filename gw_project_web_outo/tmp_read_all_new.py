# -*- coding: utf-8 -*-
"""Read all green+auto cases that need new test files, output to JSON."""
import sys, json, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook

EXCEL = "Manual_testcase/【AcuHMI-1-7】测试用例_foAI.xlsx"

COL_MODULE      = "模块"
COL_SUBMODULE   = "子模块"
COL_CASE_ID     = "用例编号"
COL_TITLE       = "用例标题"
COL_PRECOND     = "预置条件"
COL_STEPS       = "测试步骤"
COL_EXPECTED    = "预期结果"
COL_LEVEL       = "用例级别"
COL_AUTO        = "自动化"
COL_SEMI_AUTO   = "半自动化"
COL_CLAUDE_NOTE = "需补充信息(claude识别回填)"
COLOR_GREEN     = "006400"

# Modules that need new generation
TARGET_MODULES = {
    "About", "Maintenance", "Modbus协议", "UI界面测试", "Web Devices",
    "安全性测试", "兼容性测试", "接入设备参数设置", "接入设备日志管理",
    "模板管理", "设备管理", "用户管理"
}

# Collect existing test file case_ids
existing = set()
for f in glob.glob('tests/**/*.py', recursive=True):
    bn = os.path.basename(f)
    if bn.startswith('test_'):
        existing.add(bn[5:-3])

wb = load_workbook(EXCEL, data_only=True)
ws = wb.active
header = [str(c.value).strip() if c.value else '' for c in ws[1]]
cm = {name: i+1 for i, name in enumerate(header) if name}

def _s(v):
    return str(v).strip() if v is not None else ''

def _get(row, col_name):
    if col_name not in cm:
        return ''
    return _s(row[cm[col_name]-1])

results = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    module = _get(row, COL_MODULE)
    if module not in TARGET_MODULES:
        continue
    auto = _get(row, COL_AUTO)
    if auto != '是':
        continue
    color = None
    if COL_CLAUDE_NOTE in cm:
        cell = ws.cell(row=row_idx, column=cm[COL_CLAUDE_NOTE])
        if cell.font and cell.font.color and cell.font.color.rgb:
            color = cell.font.color.rgb[-6:]
    if color != COLOR_GREEN:
        continue
    case_id = _get(row, COL_CASE_ID)
    if case_id and case_id in existing:
        continue  # already exists, skip
    results.append({
        'row': row_idx,
        'module': module,
        'submodule': _get(row, COL_SUBMODULE),
        'case_id': case_id,
        'title': _get(row, COL_TITLE),
        'precondition': _get(row, COL_PRECOND),
        'steps': _get(row, COL_STEPS),
        'expected': _get(row, COL_EXPECTED),
        'level': _get(row, COL_LEVEL),
        'auto': auto,
        'semi_auto': _get(row, COL_SEMI_AUTO),
        'claude_note': _get(row, COL_CLAUDE_NOTE),
    })

wb.close()
print(json.dumps(results, ensure_ascii=False, indent=2))
