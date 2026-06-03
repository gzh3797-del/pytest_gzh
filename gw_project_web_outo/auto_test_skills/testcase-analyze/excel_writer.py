"""
excel_writer.py — auto-analyze skill 的 Excel 辅助工具

用法：
  python excel_writer.py --list-modules <excel_path>
  python excel_writer.py --init <excel_path>
  python excel_writer.py --read <excel_path> --module <模块名>
  python excel_writer.py --write <excel_path> --data <json_string>
"""

import argparse
import io
import json
import sys
from pathlib import Path

# 统一 stdout 为 UTF-8，避免 Windows 终端乱码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)

# ── 列名常量 ────────────────────────────────────────────────────────────────
COL_MODULE      = "模块"
COL_SUBMODULE   = "子模块"
COL_CASE_ID     = "用例编号"
COL_TITLE       = "用例标题"
COL_PRECOND     = "预置条件"
COL_STEPS       = "测试步骤"
COL_EXPECTED    = "预期结果"
COL_LEVEL       = "用例级别"
COL_SEMI_AUTO   = "半自动化"
COL_AUTO        = "自动化"
COL_DEBUG_PASS  = "自动化脚本是否调试通过(是/否)"
COL_CLAUDE_NOTE = "需补充信息(claude识别回填)"
COL_USER_REPLY  = "用户答复(基于需补充信息,澄清信息)"

COLOR_GREEN  = "006400"   # 深绿：已理解
COLOR_RED    = "FF0000"   # 红色：需澄清
COLOR_ORANGE = "FFA500"   # 橙色：半自动化

# ── 工具函数 ────────────────────────────────────────────────────────────────

def _load(path: str):
    p = Path(path)
    if not p.exists():
        _exit_error(f"文件不存在: {path}")
    return openpyxl.load_workbook(str(p))


def _normalize_col(s: str) -> str:
    """规范化列名：去除换行符和首尾空格，统一中英文逗号。"""
    return s.replace('\n', '').replace('\r', '').replace('，', ',').strip()


def _col_map(ws) -> dict:
    """返回 {规范化列名: 列索引(1-based)} 的映射。"""
    return {
        _normalize_col(cell.value): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def _exit_error(msg: str):
    print(json.dumps({"error": msg}, ensure_ascii=False))
    sys.exit(1)


def _str(v) -> str | None:
    return str(v).strip() if v is not None else None


# ── 子命令实现 ───────────────────────────────────────────────────────────────

def cmd_list_modules(excel_path: str):
    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)
    if COL_MODULE not in cm:
        _exit_error(f"列 '{COL_MODULE}' 不存在")
    col_idx = cm[COL_MODULE]
    counts: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = _str(row[col_idx - 1])
        if val:
            counts[val] = counts.get(val, 0) + 1
    result = [{"module": k, "count": v} for k, v in sorted(counts.items(), key=lambda x: -x[1])]
    print(json.dumps(result, ensure_ascii=False, indent=2))


def cmd_init(excel_path: str):
    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)
    added = {}

    for col_name in [COL_CLAUDE_NOTE, COL_USER_REPLY]:
        if col_name in cm:
            added[col_name] = cm[col_name]
        else:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            # 加粗表头
            ws.cell(row=1, column=new_col).font = Font(bold=True)
            added[col_name] = new_col
            cm[col_name] = new_col

    wb.save(excel_path)
    print(json.dumps(added, ensure_ascii=False, indent=2))


def cmd_read(excel_path: str, module: str):
    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)

    def _get(row_vals, col_name):
        if col_name not in cm:
            return None
        return _str(row_vals[cm[col_name] - 1])

    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if _str(row[cm.get(COL_MODULE, 1) - 1]) != module:
            continue

        # 读取 claude_note 单元格的字体颜色（判断是否已是绿色）
        note_color = None
        if COL_CLAUDE_NOTE in cm:
            cell = ws.cell(row=row_idx, column=cm[COL_CLAUDE_NOTE])
            if cell.font and cell.font.color and cell.font.color.rgb:
                note_color = cell.font.color.rgb[-6:]  # 去掉 alpha 前缀

        results.append({
            "row":          row_idx,
            "case_id":      _get(row, COL_CASE_ID),
            "title":        _get(row, COL_TITLE),
            "precondition": _get(row, COL_PRECOND),
            "steps":        _get(row, COL_STEPS),
            "expected":     _get(row, COL_EXPECTED),
            "level":        _get(row, COL_LEVEL),
            "auto":         _get(row, COL_AUTO),
            "semi_auto":    _get(row, COL_SEMI_AUTO),
            "claude_note":  _get(row, COL_CLAUDE_NOTE),
            "claude_color": note_color,
            "user_reply":   _get(row, COL_USER_REPLY),
        })

    print(json.dumps(results, ensure_ascii=False, indent=2))


def cmd_write(excel_path: str, data_json: str):
    try:
        data = json.loads(data_json)
    except json.JSONDecodeError as e:
        _exit_error(f"JSON 解析失败: {e}")

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)

    if COL_CLAUDE_NOTE not in cm:
        _exit_error(f"列 '{COL_CLAUDE_NOTE}' 不存在，请先运行 --init")

    col_idx = cm[COL_CLAUDE_NOTE]
    written = 0

    for item in data:
        row_num = item.get("row")
        text    = item.get("text", "")
        color   = item.get("color", COLOR_RED)

        if not row_num:
            continue

        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = text
        cell.font  = Font(color=color)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        written += 1

    wb.save(excel_path)
    print(json.dumps({"written": written, "file": excel_path}, ensure_ascii=False))


def cmd_write_p(excel_path: str, results_json: str):
    """
    回填 P 列（自动化脚本是否调试通过）。
    results_json: '[{"case_id": "TestCase_...", "value": "是"}, ...]'
    """
    try:
        results = json.loads(results_json)
    except json.JSONDecodeError as e:
        _exit_error(f"JSON 解析失败: {e}")

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)

    if COL_CASE_ID not in cm:
        _exit_error(f"列 '{COL_CASE_ID}' 不存在")

    if COL_DEBUG_PASS not in cm:
        _exit_error(f"列 '{COL_DEBUG_PASS}' 不存在，请确认 Excel 表头包含该列")

    case_id_col = cm[COL_CASE_ID]
    p_col       = cm[COL_DEBUG_PASS]

    # 构建 case_id → 行号 的索引（支持同一 case_id 多行）
    case_row_map: dict[str, list[int]] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cid = _str(row[case_id_col - 1])
        if cid:
            case_row_map.setdefault(cid, []).append(row_idx)

    written = 0
    not_found = []

    for item in results:
        case_id = item.get("case_id", "").strip()
        # Accept "value", "p_value", or "result" as the value key for robustness
        value = item.get("value") or item.get("p_value") or item.get("result") or ""

        if case_id not in case_row_map:
            not_found.append(case_id)
            continue

        for row_num in case_row_map[case_id]:
            cell = ws.cell(row=row_num, column=p_col)
            cell.value = value
            written += 1

    buf = io.BytesIO()
    wb.save(buf)
    with open(excel_path, "wb") as f:
        f.write(buf.getvalue())

    print(json.dumps({
        "written": written,
        "not_found": not_found,
        "file": excel_path,
    }, ensure_ascii=False))


def cmd_status(excel_path: str):
    """统计所有模块中 auto=是 的用例按 Q列字体颜色分布，识别剩余待处理用例"""
    from collections import defaultdict

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)

    mod_col   = cm.get(COL_MODULE)
    auto_col  = cm.get(COL_AUTO)
    note_col  = cm.get(COL_CLAUDE_NOTE)
    reply_col = cm.get(COL_USER_REPLY)

    if not all([mod_col, auto_col, note_col, reply_col]):
        _exit_error("必要列不存在，请先运行 --init")

    stats = defaultdict(lambda: {
        "green": 0, "red_with_reply": 0,
        "red_no_reply": 0, "orange": 0, "no_color": 0
    })

    for row in ws.iter_rows(min_row=2):
        module = _str(row[mod_col - 1].value)
        auto   = _str(row[auto_col - 1].value)
        if not module or auto != "是":
            continue

        cell_q = row[note_col - 1]
        reply  = _str(row[reply_col - 1].value)

        color = None
        if cell_q.font and cell_q.font.color and cell_q.font.color.type == "rgb":
            color = cell_q.font.color.rgb[-6:]  # 去掉 alpha 前缀

        s = stats[module]
        if color == COLOR_GREEN:
            s["green"] += 1
        elif color == COLOR_RED:
            if reply:
                s["red_with_reply"] += 1
            else:
                s["red_no_reply"] += 1
        elif color == COLOR_ORANGE:
            s["orange"] += 1
        else:
            s["no_color"] += 1

    result = []
    for mod, s in sorted(stats.items(),
                         key=lambda x: x[1]["red_with_reply"] + x[1]["red_no_reply"],
                         reverse=True):
        result.append({
            "module": mod,
            **s,
            "total_auto": sum(s.values()),
            "pending_reanalyze": s["red_with_reply"],
            "pending_user_input": s["red_no_reply"],
        })

    print(json.dumps(result, ensure_ascii=False, indent=2))


# ── CLI 入口 ─────────────────────────────────────────────────────────────────

def main():
    argv = sys.argv[1:]
    if not argv:
        print("用法: excel_writer.py --list-modules <path>")
        print("      excel_writer.py --init <path>")
        print("      excel_writer.py --read <path> --module <模块名>")
        print("      excel_writer.py --write <path> --data <json>")
        return

    cmd  = argv[0]
    rest = argv[1:]

    if cmd == "--list-modules":
        if not rest:
            _exit_error("缺少 excel_path 参数")
        cmd_list_modules(rest[0])

    elif cmd == "--init":
        if not rest:
            _exit_error("缺少 excel_path 参数")
        cmd_init(rest[0])

    elif cmd == "--read":
        p = argparse.ArgumentParser()
        p.add_argument("excel_path")
        p.add_argument("--module", required=True)
        a = p.parse_args(rest)
        cmd_read(a.excel_path, a.module)

    elif cmd == "--write":
        p = argparse.ArgumentParser()
        p.add_argument("excel_path")
        p.add_argument("--data", required=True)
        a = p.parse_args(rest)
        cmd_write(a.excel_path, a.data)

    elif cmd == "--status":
        if not rest:
            _exit_error("缺少 excel_path 参数")
        cmd_status(rest[0])

    elif cmd == "--write-p":
        p = argparse.ArgumentParser()
        p.add_argument("excel_path")
        p.add_argument("--results", required=False,
                       help='JSON 字符串，格式：[{"case_id":"...", "value":"是/否"}, ...]')
        p.add_argument("--results-file", required=False,
                       help='JSON 文件路径（推荐，避免 shell 编码问题）')
        a = p.parse_args(rest)
        if a.results_file:
            import pathlib
            results_json = pathlib.Path(a.results_file).read_text(encoding="utf-8")
        elif a.results:
            results_json = a.results
        else:
            _exit_error("--results 或 --results-file 必须提供一个")
        cmd_write_p(a.excel_path, results_json)

    else:
        _exit_error(f"未知命令: {cmd}。支持：--list-modules / --init / --read / --write / --write-p")


if __name__ == "__main__":
    main()
