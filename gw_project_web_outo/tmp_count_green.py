# -*- coding: utf-8 -*-
"""Count green+auto cases per module."""
import sys, json, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, '.')

from openpyxl import load_workbook

EXCEL = "Manual_testcase/【AcuHMI-1-7】测试用例_foAI.xlsx"

COL_MODULE      = "模块"
COL_AUTO        = "自动化"
COL_CLAUDE_NOTE = "需补充信息(claude识别回填)"
COLOR_GREEN     = "006400"

wb = load_workbook(EXCEL, data_only=True)
ws = wb.active

header = [str(c.value).strip() if c.value else '' for c in ws[1]]
cm = {name: i+1 for i, name in enumerate(header) if name}

def _s(v):
    return str(v).strip() if v is not None else ''

modules_order = []
seen = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    m = _s(row[cm[COL_MODULE]-1]) if COL_MODULE in cm else ''
    if m and m not in seen:
        modules_order.append(m)
        seen.add(m)

results = {}
for module in modules_order:
    total = 0
    green_auto = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        m = _s(row[cm[COL_MODULE]-1]) if COL_MODULE in cm else ''
        if m != module:
            continue
        total += 1
        auto = _s(row[cm[COL_AUTO]-1]) if COL_AUTO in cm else ''
        if auto != '是':
            continue
        if COL_CLAUDE_NOTE in cm:
            cell = ws.cell(row=row_idx, column=cm[COL_CLAUDE_NOTE])
            color = None
            if cell.font and cell.font.color and cell.font.color.rgb:
                color = cell.font.color.rgb[-6:]
            if color == COLOR_GREEN:
                green_auto += 1
    results[module] = {'total': total, 'green_auto': green_auto}

print(json.dumps(results, ensure_ascii=False, indent=2))
wb.close()
