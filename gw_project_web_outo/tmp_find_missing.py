# -*- coding: utf-8 -*-
"""Find green+auto cases that don't have existing test files."""
import sys, json, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook

EXCEL = "Manual_testcase/【AcuHMI-1-7】测试用例_foAI.xlsx"
COL_MODULE      = "模块"
COL_CASE_ID     = "用例编号"
COL_AUTO        = "自动化"
COL_SUBMODULE   = "子模块"
COL_CLAUDE_NOTE = "需补充信息(claude识别回填)"
COLOR_GREEN     = "006400"

wb = load_workbook(EXCEL, data_only=True)
ws = wb.active
header = [str(c.value).strip() if c.value else '' for c in ws[1]]
cm = {name: i+1 for i, name in enumerate(header) if name}

def _s(v):
    return str(v).strip() if v is not None else ''

# Collect all existing test file names (without path)
existing = set()
for f in glob.glob('tests/**/*.py', recursive=True):
    basename = os.path.basename(f)
    if basename.startswith('test_'):
        # Extract case_id from filename
        case_id = basename[5:-3]  # remove 'test_' prefix and '.py' suffix
        existing.add(case_id)

green_cases = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    module = _s(row[cm[COL_MODULE]-1]) if COL_MODULE in cm else ''
    case_id = _s(row[cm[COL_CASE_ID]-1]) if COL_CASE_ID in cm else ''
    auto = _s(row[cm[COL_AUTO]-1]) if COL_AUTO in cm else ''
    submodule = _s(row[cm[COL_SUBMODULE]-1]) if COL_SUBMODULE in cm else ''

    if auto != '是':
        continue

    color = None
    if COL_CLAUDE_NOTE in cm:
        cell = ws.cell(row=row_idx, column=cm[COL_CLAUDE_NOTE])
        if cell.font and cell.font.color and cell.font.color.rgb:
            color = cell.font.color.rgb[-6:]

    if color != COLOR_GREEN:
        continue

    already_exists = case_id in existing
    green_cases.append({
        'row': row_idx,
        'module': module,
        'submodule': submodule,
        'case_id': case_id,
        'exists': already_exists
    })

wb.close()

# Summarize
by_module = {}
for c in green_cases:
    m = c['module']
    if m not in by_module:
        by_module[m] = {'total': 0, 'exists': 0, 'new': 0}
    by_module[m]['total'] += 1
    if c['exists']:
        by_module[m]['exists'] += 1
    else:
        by_module[m]['new'] += 1

print("=== GREEN+AUTO CASES SUMMARY ===")
total_new = 0
for m, stats in by_module.items():
    total_new += stats['new']
    print(f"  {m}: total={stats['total']}, exists={stats['exists']}, new={stats['new']}")
print(f"\nTotal NEW to generate: {total_new}")
print(f"Total already exists: {sum(s['exists'] for s in by_module.values())}")

# List new cases
print("\n=== NEW CASES TO GENERATE ===")
for c in green_cases:
    if not c['exists']:
        print(f"  [{c['module']}] {c['case_id']} (sub: {c['submodule']})")
