"""将"需补充信息"注释写入用例 Excel 第10列"""
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

_EXCEL = Path(__file__).parent.parent / 'Manual_testcase' / 'AcuHMI-1-7_用户管理用例.xlsx'

wb = openpyxl.load_workbook(_EXCEL)
ws = wb['Sheet1']

# 写表头
ws.cell(1, 10).value = '需补充信息'
ws.cell(1, 10).font = Font(bold=True)
ws.cell(1, 10).fill = PatternFill('solid', fgColor='FFD700')

notes = {
    # ── EULA 类 ──────────────────────────────────────────────────────
    'TestCase_AcuHMI_007_01_case06_5': (
        '【UI确认】探索当前设备UI未发现EULA弹窗，请确认：\n'
        '①当前固件版本是否包含EULA功能；\n'
        '②EULA弹窗的触发条件（首次登录/出厂状态）；\n'
        '③弹窗中"不接受"按钮的确切文本'
    ),
    'TestCase_AcuHMI_007_01_case06_6': (
        '【UI确认】未发现EULA功能，请确认是否存在；\n'
        '另步骤9-12引用"none权限"，但Role Configuration仅见view/edit，请确认是否有none选项'
    ),
    'TestCase_AcuHMI_007_01_case06_7': (
        '【UI确认+高风险】依赖"Configuration Management恢复出厂"，请确认该操作路径；\n'
        '另需确认EULA功能是否存在'
    ),
    'TestCase_AcuHMI_007_01_case06_8': (
        '【UI确认】依赖固件升级触发EULA更新，请确认该用例是否纳入自动化范围；\n'
        '另需确认EULA功能是否存在'
    ),

    # ── 默认密码策略 ─────────────────────────────────────────────────
    'TestCase_ARM-XXL_002_04_case11_01': (
        '【密码格式】默认密码格式为"Admin@AABBCC（序列号后6位）"，\n'
        '测试环境当前用的是"Admin@110001"，请确认：\n'
        '①当前设备序列号后6位；\n'
        '②自动化使用固定密码还是动态读取序列号'
    ),
    'TestCase_ARM-XXL_002_04_case11_02': (
        '【高风险操作】步骤1"恢复出厂设置"会重置所有配置，请确认是否纳入自动化；\n'
        '另默认密码格式问题同case11_01'
    ),

    # ── 密码重置功能 ─────────────────────────────────────────────────
    'TestCase_ARM-XXL_002_04_case12_01': (
        '【外部依赖】步骤3"生成临时密码"需外部工具，请提供：\n'
        '①临时密码生成算法（如基于时间戳+SN号的规则）；\n'
        '②或可调用的生成接口/工具\n'
        '【UI确认】登录页是否有"Forgot password"按钮？当前截图未见，请确认位置'
    ),
    'TestCase_ARM-XXL_002_04_case12_02': (
        '【UI确认】登录页是否存在"Forgot password"按钮？当前截图未观察到，请确认UI位置'
    ),
    'TestCase_ARM-XXL_002_04_case12_03': (
        '【外部依赖+系统时间】依赖外部工具生成临时密码；\n'
        '步骤5需修改系统时间+1天，请确认：\n'
        '①临时密码生成方式；\n'
        '②系统时间是否可通过UI或API修改'
    ),
    'TestCase_ARM-XXL_002_04_case12_04': (
        '【外部依赖】同case12_01，依赖外部工具生成临时密码'
    ),
    'TestCase_ARM-XXL_002_04_case12_05': (
        '【外部依赖】同case12_01，依赖外部工具生成临时密码'
    ),
    'TestCase_ARM-XXL_002_04_case12_06': (
        '【外部依赖+系统时间】同case12_01；\n'
        '另"密码1天内不可修改"场景需结合密码策略配置+系统时间偏移，请确认实现方案'
    ),

    # ── 默认密码登录提醒 ─────────────────────────────────────────────
    'TestCase_AcuHMI_007_01_case06_9': (
        '【EULA依赖】步骤2要求"登录并接受EULA"，请确认EULA是否存在；\n'
        '如不存在步骤2是否改为"直接登录"；\n'
        '另"Yes, continue"弹窗的确切页面路径请确认'
    ),
    'TestCase_AcuHMI_007_01_case06_10': (
        '【EULA依赖+行为差异】同上；\n'
        'view用户点击"Yes, continue"预期为"退出登录"，\n'
        '与admin用户跳转到修改密码页不同，请确认view用户的具体交互流程'
    ),
    'TestCase_AcuHMI_007_01_case06_11': (
        '【高风险+EULA依赖】依赖"恢复出厂设置"，请确认是否纳入自动化；\n'
        '另步骤2依赖EULA功能，请确认'
    ),
    'TestCase_AcuHMI_007_01_case06_12': (
        '【高风险+EULA依赖+按钮文本】依赖"恢复出厂设置"；\n'
        '步骤3"点击close"与其他用例的"Cancel"名称不一致，请确认按钮文本'
    ),

    # ── 用户角色配置 ─────────────────────────────────────────────────
    'TestCase_AcuHMI_007_02_case01': (
        '【权限选项】添加角色时设置"所有权限为none"，\n'
        '但Role Configuration页面仅见view/edit选项，请确认是否有"none"选项'
    ),
    'TestCase_AcuHMI_007_02_case01_1': '【权限选项】同case01，请确认none选项是否存在',
    'TestCase_AcuHMI_007_02_case01_3': '【权限选项】同case01，请确认none选项是否存在',
    'TestCase_AcuHMI_007_02_case01_4': '【权限选项】同case01，请确认none选项是否存在',
    'TestCase_AcuHMI_007_02_case01_13': (
        '【标题与步骤不符】用例标题为"模板-视图"，\n'
        '但步骤和预期结果写的是"告警-视图"，请确认以哪个为准'
    ),
    'TestCase_AcuHMI_007_02_case01_14': (
        '【标题与步骤不符】用例标题为"模板-编辑"，\n'
        '但步骤和预期结果写的是"告警-编辑"，请确认以哪个为准'
    ),
    'TestCase_AcuHMI_007_02_case01_19': (
        '【模块名称】测试步骤包含"重新启动-视图"权限，\n'
        '但Role Configuration截图中未见该列，请确认UI中对应列名'
    ),
    'TestCase_AcuHMI_007_02_case01_20': (
        '【模块名称】同上，"重新启动-编辑"模块请确认UI中对应列名'
    ),
    'TestCase_AcuHMI_007_02_case02': (
        '【权限选项】预置条件含"none权限的用户test"，\n'
        '请确认none权限选项是否存在'
    ),
    'TestCase_AcuHMI_007_02_case02_1': '【权限选项】同case02，请确认none选项是否存在',

    # ── 用户密码策略 ─────────────────────────────────────────────────
    'TestCase_AcuHMI_007_03_case02': (
        '【预期矛盾】Password History=0 预期"保存配置失败"，\n'
        '但页面说明"0 = no restrict（不限制）"，请确认0是有效值还是无效值'
    ),
    'TestCase_AcuHMI_007_03_case07': (
        '【字段名称】步骤中出现"Maximum Failed Attempts"字段，\n'
        '当前UI截图Password Policy页面中未见该字段，请确认UI实际字段名称'
    ),
    'TestCase_AcuHMI_007_03_case07_1': (
        '【字段名称+逻辑冲突】同case07；\n'
        '另"Failed Login Attempt Window=0"与case08_02中"0=不锁定"描述有冲突，请确认'
    ),
    'TestCase_AcuHMI_007_03_case07_2': '【字段名称】同case07，请确认"Maximum Failed Attempts"的UI字段名',
    'TestCase_AcuHMI_007_03_case07_3': '【字段名称】同case07',
    'TestCase_AcuHMI_007_03_case07_4': '【字段名称】同case07',
    'TestCase_AcuHMI_007_03_case07_5': '【字段名称】同case07',
    'TestCase_AcuHMI_007_03_case08': (
        '【字段名称】步骤中"最大失败尝试次数"对应UI字段名请确认'
    ),
    'TestCase_AcuHMI_007_03_case08_02': '【字段名称】同case08',
    'TestCase_AcuHMI_007_03_case08_03': '【字段名称】同case08',

    # ── 时间依赖用例 ─────────────────────────────────────────────────
    'TestCase_AcuHMI_007_03_case03_2': (
        '【时间依赖】步骤3需等待"一天之内"验证，请确认是否通过修改系统时间偏移模拟'
    ),
    'TestCase_AcuHMI_007_03_case03_3': '【时间依赖】需等待50天，请确认是否通过系统时间偏移实现',
    'TestCase_AcuHMI_007_03_case03_4': '【时间依赖】需等待90天，请确认是否通过系统时间偏移实现',
    'TestCase_AcuHMI_007_03_case04_2': '【时间依赖】需等待1天后登录，请确认是否通过系统时间偏移实现',
    'TestCase_AcuHMI_007_03_case04_3': '【时间依赖】需等待50天，请确认是否通过系统时间偏移实现',
    'TestCase_AcuHMI_007_03_case04_4': '【时间依赖】需等待90天，请确认是否通过系统时间偏移实现',
    'TestCase_AcuHMI_007_03_case06': (
        '【时间依赖】步骤2需等待1天后验证密码到期锁定，请确认是否通过系统时间偏移实现'
    ),
    'TestCase_AcuHMI_007_03_case06_1': (
        '【时间依赖】需模拟"密码到期后1天内"和"1天后"两种场景，请确认时间偏移方案'
    ),
    'TestCase_AcuHMI_007_01_case01_1': (
        '【时间依赖】步骤4需等待1分钟（Session Timeout=1min），\n'
        '自动化可用page.wait_for_timeout(61000)实现，请确认是否可接受'
    ),
    'TestCase_AcuHMI_007_01_case01_2': (
        '【时间依赖】步骤2需等待31分钟，执行耗时较长，请确认是否纳入自动化范围'
    ),
    'TestCase_AcuHMI_007_01_case01_3': '【时间依赖】需等待61分钟，请确认是否纳入自动化范围',
    'TestCase_AcuHMI_007_01_case01_6': '【时间依赖】步骤3需等待60分钟，请确认是否纳入自动化范围',

    # ── 用例标题为空 ─────────────────────────────────────────────────
    'TestCase_AcuHMI_007_01_case02_01': (
        '【标题为空】用例标题列内容为空格，请补充标题；\n'
        '【UI字段未确认】步骤涉及"多重登录""覆盖密码策略""覆盖密码过期"选项，\n'
        '当前Add User弹框截图中未见这些字段，请确认是否存在及UI中的确切名称'
    ),
    'TestCase_AcuHMI_007_01_case02_02': (
        '【UI字段未确认】步骤涉及"覆盖密码过期""多重登录"选项，\n'
        '请确认UI中是否存在，以及字段的确切名称'
    ),
    'TestCase_AcuHMI_007_01_case05_01': (
        '【UI字段未确认】步骤涉及"多重登录"选项，\n'
        '请确认Add User/Edit User弹框中是否有该字段'
    ),
    'TestCase_AcuHMI_007_01_case05_02': (
        '【UI字段未确认+时间依赖】涉及"多重登录""覆盖密码过期"字段（请确认UI是否存在）；\n'
        '步骤3/7需修改系统时间+1/+2天，请确认时间偏移方案'
    ),

    # ── 密码配置 ─────────────────────────────────────────────────────
    'TestCase_AcuHMI_007_04_case01': (
        '【密码规则依赖】步骤2密码"123456"仅6位，当前密码策略默认长度要求需确认；\n'
        '另"密码长度20、40"的具体密码值用例中未给出，请补充示例值'
    ),
    'TestCase_AcuHMI_007_04_case01_1': (
        '【设备名称错误】预置条件写的是"ARM-XXL网页"，本项目为AcuHMI，请确认是否笔误'
    ),

    # ── 密码修改（引用ARM-XXL）───────────────────────────────────────
    'TestCase_ARM-XXL_002_04_case1_03': (
        '【设备引用】编号和预置条件引用"ARM-XXL"，请确认该用例是否适用于AcuHMI；\n'
        '另步骤2修改admin密码后"自动退出"，请确认AcuHMI是否有此行为'
    ),
    'TestCase_ARM-XXL_002_04_case1_04': (
        '【设备引用】预置条件引用ARM-XXL，请确认是否适用于AcuHMI'
    ),
    'TestCase_ARM-XXL_002_04_case1_05': (
        '【设备引用+UI字段】引用ARM-XXL；\n'
        '步骤4"修改密码是否需要输入当前用户密码"需确认Password Management编辑弹框的字段'
    ),
    'TestCase_ARM-XXL_002_04_case1_06': (
        '【设备引用】同case1_05，引用ARM-XXL请确认是否适用'
    ),
}

# 写入备注
cnt = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    case_id = row[2].value
    if case_id and case_id in notes:
        cell = row[9]
        cell.value = notes[case_id]
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.fill = PatternFill('solid', fgColor='FFF2CC')
        cnt += 1

ws.column_dimensions['J'].width = 60

buf = io.BytesIO()
wb.save(buf)
with open(_EXCEL, 'wb') as f:
    f.write(buf.getvalue())
print(f'完成，共标注 {cnt} 条用例，notes定义了 {len(notes)} 条')
