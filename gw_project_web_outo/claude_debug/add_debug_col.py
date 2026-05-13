"""
在 Excel 第 14 列新增"调试通过/不通过"，标记已调试用例的结果。
"""
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

_EXCEL = Path(__file__).parent.parent / 'Manual_testcase' / 'AcuHMI-1-7_用户管理用例.xlsx'

PASS_VAL = "调试通过"
FAIL_VAL = "调试不通过"

GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL   = PatternFill('solid', fgColor='FFC7CE')
HDR_FILL   = PatternFill('solid', fgColor='4472C4')
NO_FILL    = PatternFill('none')

# row -> result  (Excel row 2 = 用例第1条)
#   row 2  → TestCase_AcuHMI_007_04_case01      (密码长度 8/20/40 保存成功)
#   row 3  → TestCase_AcuHMI_007_04_case01_1    (密码长度 5/129 保存失败)
#   row 4  → TestCase_ARM_XXL_002_04_case1_03   (admin 修改自身密码)
#   row 5  → TestCase_ARM_XXL_002_04_case1_04   (admin 修改其他用户密码)
#   row 6  → TestCase_ARM_XXL_002_04_case1_05   (非admin修改自己密码)
#   row 7  → TestCase_ARM_XXL_002_04_case1_06   (非admin修改他人密码)
#   row 11 → TestCase_ARM_XXL_002_04_case12_02  (Forgot Password 弹框)
#   row 16 → TestCase_AcuHMI_007_01_case06_5    (EULA 首次登录 - SKIPPED)
#   row 17 → TestCase_AcuHMI_007_01_case06_6    (EULA 首次登录)
#   row 18 → TestCase_AcuHMI_007_01_case06_7    (EULA - SKIPPED)
#   row 19 → TestCase_AcuHMI_007_01_case06_8    (EULA - SKIPPED)
#   row 20 → TestCase_AcuHMI_007_01_case06_9    (工厂重置相关，无需自动化，脚本已删除)
#   row 21 → TestCase_AcuHMI_007_01_case06_10   (工厂重置相关，无需自动化，脚本已删除)
#   row 24 → TestCase_AcuHMI_007_02_case01      (角色全none，登录无权限)
#   row 25 → TestCase_AcuHMI_007_02_case01_1    (角色全view，登录视图权限)
#   row 26 → TestCase_AcuHMI_007_02_case01_2    (角色全edit，登录编辑权限)
#   row 27 → TestCase_AcuHMI_007_02_case01_3    (User=view其余none)
#   row 28 → TestCase_AcuHMI_007_02_case01_4    (User=edit其余none)
#   row 29 → TestCase_AcuHMI_007_02_case01_5    (Device=view其余none)
#   row 30 → TestCase_AcuHMI_007_02_case01_6    (Device=edit其余none)
#   row 31 → TestCase_AcuHMI_007_02_case01_7    (DataLog=view其余none — 登录成功，点About弹"No Any Permissions")
#   row 32 → TestCase_AcuHMI_007_02_case01_8    (DataLog=edit其余none)
#   row 33 → TestCase_AcuHMI_007_02_case01_9    (SystemSettings=view其余none)
#   row 34 → TestCase_AcuHMI_007_02_case01_10   (SystemSettings=edit其余none)
#   row 35 → TestCase_AcuHMI_007_02_case01_11   (Protocol=view其余none)
#   row 36 → TestCase_AcuHMI_007_02_case01_12   (Protocol=edit其余none)
#   row 37 → TestCase_AcuHMI_007_02_case01_13   (AlarmLog=view其余none)
#   row 38 → TestCase_AcuHMI_007_02_case01_14   (AlarmLog=edit其余none)
#   row 39 → TestCase_AcuHMI_007_02_case01_15   (Maintenance=view其余none)
#   row 40 → TestCase_AcuHMI_007_02_case01_16   (Maintenance=edit其余none)
#   row 41 → TestCase_AcuHMI_007_02_case01_17   (Diagnostics=view其余none)
#   row 42 → TestCase_AcuHMI_007_02_case01_18   (Diagnostics=edit其余none)
#   row 45 → TestCase_AcuHMI_007_02_case01_21   (FirmwareUpdate=edit其余none — 登录成功，Browse可见，点Devices弹"No Any Permissions")
#   row 46 → TestCase_AcuHMI_007_02_case01_22   (FirmwareUpdate=edit其余none)
#   row 47 → TestCase_AcuHMI_007_02_case01_23   (角色名称验证)
#   row 48 → TestCase_AcuHMI_007_02_case02      (编辑角色全无→全视图)
#   row 49 → TestCase_AcuHMI_007_02_case02_1    (编辑角色全无→全编辑)
#   row 50 → TestCase_AcuHMI_007_02_case02_2    (编辑角色全视图→全编辑)
#   row 51 → TestCase_AcuHMI_007_02_case02_3    (编辑角色全视图→全无)
#   row 52 → TestCase_AcuHMI_007_02_case02_4    (编辑角色全编辑→全无)
#   row 53 → TestCase_AcuHMI_007_02_case02_5    (编辑角色全编辑→全视图)
#   row 54 → TestCase_AcuHMI_007_02_case03      (删除有用户的角色失败)
#   row 55 → TestCase_AcuHMI_007_02_case03_1    (删除无用户角色成功)
#   row 56 → TestCase_AcuHMI_007_02_case03_2    (删除用户后再删角色成功)
#   row 57 → TestCase_AcuHMI_007_03_case01      (仅Upper+Lower策略，含大小写密码登录成功)
#   row 58 → TestCase_AcuHMI_007_03_case01_1    (仅Upper+Lower，无字母密码拒绝)
#   row 59 → TestCase_AcuHMI_007_03_case01_2    (仅Numbers+Letters，含数字字母登录成功)
#   row 60 → TestCase_AcuHMI_007_03_case01_3    (仅Numbers+Letters，无数字拒绝)
#   row 61 → TestCase_AcuHMI_007_03_case01_4    (仅Special Characters，含特殊字符登录成功)
#   row 62 → TestCase_AcuHMI_007_03_case01_5    (仅Special Characters，无特殊字符拒绝)
#   row 63 → TestCase_AcuHMI_007_03_case01_6    (Upper+Lower+Numbers+Letters，含大小写数字登录成功)
#   row 64 → TestCase_AcuHMI_007_03_case01_7    (Upper+Lower+Special，含大小写特殊字符登录成功)
#   row 65 → TestCase_AcuHMI_007_03_case01_8    (Numbers+Letters+Special，含数字字母特殊字符登录成功)
#   row 66 → TestCase_AcuHMI_007_03_case01_9    (Numbers+Letters+Special，无特殊字符拒绝)
#   row 67 → TestCase_AcuHMI_007_03_case01_10   (Numbers+Letters+Special，仅特殊字符拒绝)
#   row 68 → TestCase_AcuHMI_007_03_case01_11   (三重策略，含全部类型登录成功)
#   row 69 → TestCase_AcuHMI_007_03_case01_12   (三重策略，无数字拒绝)
#   row 70 → TestCase_AcuHMI_007_03_case01_13   (三重策略，无特殊字符拒绝)
#   row 71 → TestCase_AcuHMI_007_03_case02      (Password History=0 保存失败)
#   row 72 → TestCase_AcuHMI_007_03_case02_1    (Password History=1无限制，复用密码成功)
#   row 73 → TestCase_AcuHMI_007_03_case02_2    (Password History=2，复用失败，新密码成功)
#   row 74 → TestCase_AcuHMI_007_03_case02_3    (Password History=31，复用失败，新密码成功)
#   row 75 → TestCase_AcuHMI_007_03_case02_4    (Password History=32，复用失败，新密码成功)
#   row 76 → TestCase_AcuHMI_007_03_case02_5    (Password History=33 保存失败)
#   row 77 → TestCase_AcuHMI_007_03_case02_6    (Password History=小数/特殊/字母 保存失败)
#   row 78 → TestCase_AcuHMI_007_03_case03      (Min Password Age=-1 保存失败)
#   row 79 → TestCase_AcuHMI_007_03_case03_1    (Min Password Age=0，立即改密成功)
#   row 80 → TestCase_AcuHMI_007_03_case03_2    (Min Password Age=1，1天内改密失败)
#   row 83 → TestCase_AcuHMI_007_03_case03_5    (Min Password Age=91 保存失败)
#   row 84 → TestCase_AcuHMI_007_03_case03_6    (Min Password Age=小数/特殊/字母 保存失败)
#   row 85 → TestCase_AcuHMI_007_03_case04      (Password Expires=-1 保存失败)
#   row 86 → TestCase_AcuHMI_007_03_case04_1    (Password Expires=0 保存成功)
#   row 90 → TestCase_AcuHMI_007_03_case04_5    (Password Expires=91 保存失败)
#   row 91 → TestCase_AcuHMI_007_03_case04_6    (Password Expires=小数/特殊/字母 保存失败)
#   row 92 → TestCase_AcuHMI_007_03_case05      (Min Password Length=5 保存失败)
#   row 93 → TestCase_AcuHMI_007_03_case05_1    (Min Password Length=8 边界测试)
#   row 94 → TestCase_AcuHMI_007_03_case05_2    (Min Password Length=40)
#   row 95 → TestCase_AcuHMI_007_03_case05_3    (Min Password Length=64)
#   row 96 → TestCase_AcuHMI_007_03_case05_4    (Min Password Length=65 保存失败)
#   row 97 → TestCase_AcuHMI_007_03_case05_5    (Min Password Length=小数/字母/特殊 保存失败)
#   row 100 → TestCase_AcuHMI_007_03_case06_2   (Grace Period=30000 保存成功)
#   row 101 → TestCase_AcuHMI_007_03_case06_3   (Grace Period=65535 保存成功)
#   row 102 → TestCase_AcuHMI_007_03_case06_4   (Grace Period=-1/65536 保存失败)
#   row 103 → TestCase_AcuHMI_007_03_case06_5   (Grace Period=小数/字母/特殊 保存失败)
#   row 104 → TestCase_AcuHMI_007_03_case07     (Max Failed Attempts=0，无锁定)
#   row 105 → TestCase_AcuHMI_007_03_case07_1   (Max Failed Attempts=1，1次失败后锁定)
#   row 106 → TestCase_AcuHMI_007_03_case07_2   (Max Failed Attempts=29，29次失败后锁定)
#   row 107 → TestCase_AcuHMI_007_03_case07_3   (Max Failed Attempts=30，30次失败后锁定)
#   row 108 → TestCase_AcuHMI_007_03_case07_4   (Max Failed Attempts=-1/31 保存失败)
#   row 109 → TestCase_AcuHMI_007_03_case07_5   (Max Failed Attempts=小数/字母/特殊 保存失败)
#   row 110 → TestCase_AcuHMI_007_03_case08     (Failed Login Window=30s，窗口内3次失败后锁定/窗口重置)
#   row 111 → TestCase_AcuHMI_007_03_case08_02  (Failed Login Attempt Window=0，无限制尝试)
#   row 112 → TestCase_AcuHMI_007_03_case08_03  (Failed Login Attempt Window=120，保存成功)
#   row 113 → TestCase_AcuHMI_007_03_case08_04  (Failed Login Attempt Window=86400，保存成功)
#   row 114 → TestCase_AcuHMI_007_03_case08_05  (Failed Login Attempt Window=-1/86401 保存失败)
#   row 115 → TestCase_AcuHMI_007_03_case08_06  (Failed Login Attempt Window=小数/字母/特殊 保存失败)
#   row 116 → TestCase_AcuHMI_007_03_case09     (Failed Login Wait=0，无锁定)
#   row 117 → TestCase_AcuHMI_007_03_case09_01  (Failed Login Wait=30，5次失败后锁定30s)
#   row 118 → TestCase_AcuHMI_007_03_case09_02  (Failed Login Wait=40000，保存成功)
#   row 119 → TestCase_AcuHMI_007_03_case09_03  (Failed Login Wait=86400，保存成功)
#   row 120 → TestCase_AcuHMI_007_03_case09_04  (Failed Login Wait=-1/86401 保存失败)
#   row 121 → TestCase_AcuHMI_007_03_case09_05  (Failed Login Wait=小数/字母/特殊 保存失败)
debug_results = {
    2:  PASS_VAL,
    3:  PASS_VAL,
    4:  PASS_VAL,   # case1_03 — admin modifies own password (force=True fix)
    5:  PASS_VAL,
    6:  PASS_VAL,
    7:  PASS_VAL,
    11: PASS_VAL,
    17: PASS_VAL,
    # rows 20, 21 — case06_9/case06_10 are factory-reset related; no automation scripts (deleted)
    24: PASS_VAL,
    25: PASS_VAL,
    26: PASS_VAL,
    27: PASS_VAL,
    28: PASS_VAL,
    29: PASS_VAL,
    30: PASS_VAL,
    31: PASS_VAL,   # case01_7 — DataLog=view: login succeeds, "No Any Permissions" toast on About click
    32: PASS_VAL,
    33: PASS_VAL,
    34: PASS_VAL,   # case01_10 — SystemSettings=edit: confirmed passing
    35: PASS_VAL,
    36: PASS_VAL,
    37: PASS_VAL,
    38: PASS_VAL,
    39: PASS_VAL,
    40: PASS_VAL,
    41: PASS_VAL,
    42: PASS_VAL,
    45: PASS_VAL,   # case01_21 — FirmwareUpdate=edit: login succeeds, Browse visible, "No Any Permissions" on Devices click
    46: PASS_VAL,
    47: PASS_VAL,
    48: PASS_VAL,
    49: PASS_VAL,
    50: PASS_VAL,
    51: PASS_VAL,
    52: PASS_VAL,
    53: PASS_VAL,
    54: PASS_VAL,
    55: PASS_VAL,
    56: PASS_VAL,
    57: PASS_VAL,
    58: PASS_VAL,
    59: PASS_VAL,
    60: PASS_VAL,
    61: PASS_VAL,
    62: PASS_VAL,
    63: PASS_VAL,
    64: PASS_VAL,
    65: PASS_VAL,
    66: PASS_VAL,
    67: PASS_VAL,
    68: PASS_VAL,
    69: PASS_VAL,
    70: PASS_VAL,
    71: PASS_VAL,
    72: PASS_VAL,
    73: PASS_VAL,
    74: PASS_VAL,
    75: PASS_VAL,
    76: PASS_VAL,
    77: PASS_VAL,
    78: PASS_VAL,
    79: PASS_VAL,
    80: PASS_VAL,
    83: PASS_VAL,
    84: PASS_VAL,
    85: PASS_VAL,
    86: PASS_VAL,
    90: PASS_VAL,
    91: PASS_VAL,
    92: PASS_VAL,
    93: PASS_VAL,
    94: PASS_VAL,
    95: PASS_VAL,
    96: PASS_VAL,
    97: PASS_VAL,
    100: PASS_VAL,
    101: PASS_VAL,
    102: PASS_VAL,
    103: PASS_VAL,
    104: PASS_VAL,
    105: PASS_VAL,
    106: PASS_VAL,
    107: PASS_VAL,
    108: PASS_VAL,
    109: PASS_VAL,
    110: PASS_VAL,
    111: PASS_VAL,
    112: PASS_VAL,
    113: PASS_VAL,
    114: PASS_VAL,
    115: PASS_VAL,
    116: PASS_VAL,
    117: PASS_VAL,
    118: PASS_VAL,
    119: PASS_VAL,
    120: PASS_VAL,
    121: PASS_VAL,
    123: PASS_VAL,   # general case01_1 (Session Timeout=1)
    126: PASS_VAL,   # general case01_4
    127: PASS_VAL,   # general case01_5
    # === User Configuration (007_01) ===
    129: PASS_VAL,   # case02 — add 1/2/10 users
    130: PASS_VAL,   # case02_01
    # 131 skipped — case02_02 (skipped test)
    132: PASS_VAL,   # case02_03 — max 32 users (fixed: wait_for hidden)
    133: PASS_VAL,   # case02_04
    134: PASS_VAL,   # case03 — delete 1/2 users
    135: PASS_VAL,   # case03_1
    136: PASS_VAL,   # case03_2
    137: PASS_VAL,   # case04 — lock 2 users (fixed: confirmation dialog)
    138: PASS_VAL,   # case04_1 (fixed: confirmation dialog)
    139: PASS_VAL,   # case04_2
    140: PASS_VAL,   # case05 — role swap (fixed: no password in edit dialog)
    141: PASS_VAL,   # case05_01 — Multiple Login: 同一Context新开页签登录触发Unauthenticated弹框
    # 142 skipped — case05_02 (skipped test)
    143: PASS_VAL,   # case05_03 — none→edit role (fixed: accept login failure for none-permission)
    144: PASS_VAL,   # case05_1
    145: PASS_VAL,   # case05_2
    146: PASS_VAL,   # case06 — username len=1, Override Policy checkbox (fixed: .el-checkbox__inner)
    147: PASS_VAL,   # case06_1
    148: PASS_VAL,   # case06_2
    149: PASS_VAL,   # case06_3
    150: PASS_VAL,   # case06_4
}

wb = openpyxl.load_workbook(_EXCEL)
ws = wb['Sheet1']

# 写表头
hdr = ws.cell(1, 14)
hdr.value = "调试通过/不通过"
hdr.font  = Font(bold=True, color='FFFFFF')
hdr.fill  = HDR_FILL
hdr.alignment = Alignment(horizontal='center', wrap_text=True)

# 逐行填写
for row in ws.iter_rows(min_row=2):
    r = row[0].row
    result = debug_results.get(r, "")
    c = ws.cell(r, 14)
    c.value = result
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    if result == PASS_VAL:
        c.fill = GREEN_FILL
    elif result == FAIL_VAL:
        c.fill = RED_FILL
    else:
        c.fill = NO_FILL

ws.column_dimensions['N'].width = 18

buf = io.BytesIO()
wb.save(buf)
with open(_EXCEL, 'wb') as f:
    f.write(buf.getvalue())

print("完成：已在第 14 列新增'调试通过/不通过'")
print(f"  调试通过: {sum(1 for v in debug_results.values() if v == PASS_VAL)} 条")
print(f"  调试不通过: {sum(1 for v in debug_results.values() if v == FAIL_VAL)} 条")
