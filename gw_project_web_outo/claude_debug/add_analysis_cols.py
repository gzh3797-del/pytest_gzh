"""
分析全部 149 条用例，在 Excel 最后新增两列：
  列 12 —— 是否实现自动化
  列 13 —— 业务逻辑是否清楚
"""
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

_EXCEL = Path(__file__).parent.parent / 'Manual_testcase' / 'AcuHMI-1-7_用户管理用例.xlsx'

AUTO_YES      = "是"
AUTO_NO       = "否"
AUTO_CODE_ONLY = "是（生成代码但不执行）"

LOGIC_YES     = "是"
LOGIC_NO      = "否"
LOGIC_PARTIAL = "部分清楚"

# row_number -> (是否实现自动化, 业务逻辑是否清楚)
decisions = {
    # ── 密码配置 ──────────────────────────────────────────────
    2:  (AUTO_YES, LOGIC_YES),
    3:  (AUTO_YES, LOGIC_YES),

    # ── 密码修改 ──────────────────────────────────────────────
    4:  (AUTO_YES, LOGIC_PARTIAL),   # "修改后自动退出"行为待进一步确认
    5:  (AUTO_YES, LOGIC_YES),
    6:  (AUTO_YES, LOGIC_PARTIAL),   # Password Management 弹框是否需要输入当前密码待确认
    7:  (AUTO_YES, LOGIC_PARTIAL),   # 同 case1_05

    # ── 默认密码策略 ──────────────────────────────────────────
    8:  (AUTO_NO,  LOGIC_YES),       # 用户明确：不实现自动化
    9:  (AUTO_CODE_ONLY, LOGIC_PARTIAL),  # 涉及恢复出厂；默认密码 Admin@AABBCC 中的 SN 后 6 位未固化

    # ── 密码重置功能 ──────────────────────────────────────────
    10: (AUTO_NO, LOGIC_NO),         # 依赖外部工具生成临时密码
    11: (AUTO_YES, LOGIC_YES),       # 弹框文本已确认
    12: (AUTO_NO, LOGIC_NO),         # 外部工具 + 系统时间偏移
    13: (AUTO_NO, LOGIC_NO),
    14: (AUTO_NO, LOGIC_NO),
    15: (AUTO_NO, LOGIC_NO),

    # ── EULA ─────────────────────────────────────────────────
    16: (AUTO_CODE_ONLY, LOGIC_PARTIAL),  # 需出厂状态；"不接受"按钮文本未确认
    17: (AUTO_YES, LOGIC_YES),            # 新建用户首次登录触发；none 权限已确认
    18: (AUTO_CODE_ONLY, LOGIC_YES),      # 恢复出厂设置：生成代码但不执行
    19: (AUTO_CODE_ONLY, LOGIC_YES),      # 依赖固件升级：生成代码但不执行

    # ── 默认密码登录提醒 ──────────────────────────────────────
    20: (AUTO_CODE_ONLY, LOGIC_YES),      # 恢复出厂：生成代码但不执行
    21: (AUTO_CODE_ONLY, LOGIC_YES),
    22: (AUTO_CODE_ONLY, LOGIC_YES),
    23: (AUTO_CODE_ONLY, LOGIC_PARTIAL),  # "close" 按钮文本（与其他用例 Cancel 不一致）未最终确认

    # ── 用户角色配置 ──────────────────────────────────────────
    24: (AUTO_YES, LOGIC_YES),
    25: (AUTO_YES, LOGIC_YES),
    26: (AUTO_YES, LOGIC_YES),
    27: (AUTO_YES, LOGIC_YES),
    28: (AUTO_YES, LOGIC_YES),
    29: (AUTO_YES, LOGIC_YES),
    30: (AUTO_YES, LOGIC_YES),
    31: (AUTO_YES, LOGIC_YES),
    32: (AUTO_YES, LOGIC_YES),
    33: (AUTO_YES, LOGIC_YES),
    34: (AUTO_YES, LOGIC_PARTIAL),  # 步骤写"系统设置-视图"但预期写"系统设置-编辑"，不一致
    35: (AUTO_YES, LOGIC_YES),
    36: (AUTO_YES, LOGIC_YES),
    37: (AUTO_YES, LOGIC_YES),       # 用户确认：以"告警-视图"为准
    38: (AUTO_YES, LOGIC_YES),       # 用户确认：以"告警-编辑"为准
    39: (AUTO_YES, LOGIC_YES),
    40: (AUTO_YES, LOGIC_YES),
    41: (AUTO_YES, LOGIC_YES),
    42: (AUTO_YES, LOGIC_YES),
    43: (AUTO_NO,  LOGIC_NO),        # 用户确认：无效用例
    44: (AUTO_NO,  LOGIC_NO),        # 用户确认：无效用例
    45: (AUTO_YES, LOGIC_PARTIAL),   # 步骤写"固件更新-编辑"但预期写"固件更新-视图"，不一致
    46: (AUTO_YES, LOGIC_YES),
    47: (AUTO_YES, LOGIC_YES),
    48: (AUTO_YES, LOGIC_YES),
    49: (AUTO_YES, LOGIC_YES),
    50: (AUTO_YES, LOGIC_YES),
    51: (AUTO_YES, LOGIC_YES),
    52: (AUTO_YES, LOGIC_YES),
    53: (AUTO_YES, LOGIC_YES),
    54: (AUTO_YES, LOGIC_YES),
    55: (AUTO_YES, LOGIC_YES),
    56: (AUTO_YES, LOGIC_YES),

    # ── 用户密码策略 ──────────────────────────────────────────
    57: (AUTO_YES, LOGIC_PARTIAL),   # 步骤中 Password/Repeat Password 值不一致，疑似笔误
    58: (AUTO_YES, LOGIC_YES),
    59: (AUTO_YES, LOGIC_YES),
    60: (AUTO_YES, LOGIC_YES),
    61: (AUTO_YES, LOGIC_YES),
    62: (AUTO_YES, LOGIC_PARTIAL),   # 步骤中"密码不包含特殊字符"未给出具体密码值
    63: (AUTO_YES, LOGIC_YES),
    64: (AUTO_YES, LOGIC_PARTIAL),   # 步骤中密码值为描述性文字，未给出具体示例
    65: (AUTO_YES, LOGIC_PARTIAL),
    66: (AUTO_YES, LOGIC_PARTIAL),
    67: (AUTO_YES, LOGIC_PARTIAL),
    68: (AUTO_YES, LOGIC_YES),
    69: (AUTO_YES, LOGIC_PARTIAL),
    70: (AUTO_YES, LOGIC_PARTIAL),
    71: (AUTO_YES, LOGIC_YES),       # Password History=0 为无效值已确认
    72: (AUTO_YES, LOGIC_YES),
    73: (AUTO_YES, LOGIC_YES),
    74: (AUTO_YES, LOGIC_YES),
    75: (AUTO_YES, LOGIC_YES),
    76: (AUTO_YES, LOGIC_YES),
    77: (AUTO_YES, LOGIC_YES),
    78: (AUTO_YES, LOGIC_YES),
    79: (AUTO_YES, LOGIC_YES),
    80: (AUTO_YES, LOGIC_YES),       # 系统时间可修改已确认
    81: (AUTO_NO,  LOGIC_YES),       # 用户：不实现自动化
    82: (AUTO_NO,  LOGIC_YES),
    83: (AUTO_YES, LOGIC_YES),
    84: (AUTO_YES, LOGIC_YES),
    85: (AUTO_YES, LOGIC_YES),
    86: (AUTO_YES, LOGIC_YES),       # Password Expires=0：验证保存后永不过期即可
    87: (AUTO_NO,  LOGIC_YES),       # 用户：不实现自动化
    88: (AUTO_NO,  LOGIC_YES),
    89: (AUTO_NO,  LOGIC_YES),
    90: (AUTO_YES, LOGIC_YES),
    91: (AUTO_YES, LOGIC_YES),
    92: (AUTO_YES, LOGIC_YES),
    93: (AUTO_YES, LOGIC_YES),
    94: (AUTO_YES, LOGIC_YES),
    95: (AUTO_YES, LOGIC_YES),
    96: (AUTO_YES, LOGIC_YES),
    97: (AUTO_YES, LOGIC_YES),
    98: (AUTO_NO,  LOGIC_YES),       # 用户：不实现自动化
    99: (AUTO_NO,  LOGIC_YES),
   100: (AUTO_YES, LOGIC_YES),       # Grace Period=30000 边界值，验证保存成功
   101: (AUTO_YES, LOGIC_YES),       # Grace Period=65535 边界值
   102: (AUTO_YES, LOGIC_YES),
   103: (AUTO_YES, LOGIC_YES),
   104: (AUTO_YES, LOGIC_YES),       # Maximum Failed Attempts 字段名已确认
   105: (AUTO_YES, LOGIC_YES),
   106: (AUTO_YES, LOGIC_YES),
   107: (AUTO_YES, LOGIC_YES),
   108: (AUTO_YES, LOGIC_YES),
   109: (AUTO_YES, LOGIC_YES),
   110: (AUTO_YES, LOGIC_YES),
   111: (AUTO_YES, LOGIC_YES),
   112: (AUTO_YES, LOGIC_YES),
   113: (AUTO_YES, LOGIC_YES),
   114: (AUTO_YES, LOGIC_YES),
   115: (AUTO_YES, LOGIC_YES),
   116: (AUTO_YES, LOGIC_YES),
   117: (AUTO_YES, LOGIC_YES),
   118: (AUTO_YES, LOGIC_YES),
   119: (AUTO_YES, LOGIC_YES),
   120: (AUTO_YES, LOGIC_YES),
   121: (AUTO_YES, LOGIC_YES),

    # ── 用户配置 ──────────────────────────────────────────────
   122: (AUTO_YES, LOGIC_YES),
   123: (AUTO_YES, LOGIC_YES),       # 等待 1 分钟可接受
   124: (AUTO_NO,  LOGIC_YES),       # 用户：不实现（等待 31 分钟）
   125: (AUTO_NO,  LOGIC_YES),       # 用户：不实现（等待 61 分钟）
   126: (AUTO_YES, LOGIC_YES),
   127: (AUTO_YES, LOGIC_YES),
   128: (AUTO_NO,  LOGIC_YES),       # 用户：不实现（等待 60 分钟）
   129: (AUTO_YES, LOGIC_YES),
   130: (AUTO_YES, LOGIC_PARTIAL),   # 用例标题为空；部分字段（覆盖密码策略/过期）确切 UI 名称仍需确认
   131: (AUTO_YES, LOGIC_PARTIAL),   # Override Password Expire 已确认；步骤 3/7 需系统时间偏移
   132: (AUTO_YES, LOGIC_YES),
   133: (AUTO_YES, LOGIC_YES),
   134: (AUTO_YES, LOGIC_YES),
   135: (AUTO_YES, LOGIC_YES),
   136: (AUTO_YES, LOGIC_YES),
   137: (AUTO_YES, LOGIC_YES),
   138: (AUTO_YES, LOGIC_YES),
   139: (AUTO_YES, LOGIC_YES),
   140: (AUTO_YES, LOGIC_YES),
   141: (AUTO_NO,  LOGIC_YES),       # 用户：不实现自动化
   142: (AUTO_NO,  LOGIC_YES),       # 用户：不实现自动化（系统时间偏移+UI字段）
   143: (AUTO_YES, LOGIC_YES),       # none 权限已确认存在
   144: (AUTO_YES, LOGIC_YES),
   145: (AUTO_YES, LOGIC_YES),
   146: (AUTO_YES, LOGIC_YES),       # Override password policy 已确认存在
   147: (AUTO_YES, LOGIC_YES),
   148: (AUTO_YES, LOGIC_YES),
   149: (AUTO_YES, LOGIC_YES),
   150: (AUTO_YES, LOGIC_PARTIAL),   # 备注：用户名超 40 位可能被自动截断后仍保存成功，与预期矛盾
}

# 颜色
GREEN_FILL  = PatternFill('solid', fgColor='C6EFCE')
RED_FILL    = PatternFill('solid', fgColor='FFC7CE')
ORANGE_FILL = PatternFill('solid', fgColor='FFEB9C')
HDR_FILL    = PatternFill('solid', fgColor='4472C4')

wb = openpyxl.load_workbook(_EXCEL)
ws = wb['Sheet1']

# ── 写表头 ───────────────────────────────────────────────────
for col, title in [(12, '是否实现自动化'), (13, '业务逻辑是否清楚')]:
    cell = ws.cell(1, col)
    cell.value = title
    cell.font  = Font(bold=True, color='FFFFFF')
    cell.fill  = HDR_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ── 逐行填写 ─────────────────────────────────────────────────
for row in ws.iter_rows(min_row=2):
    r = row[0].row
    auto, logic = decisions.get(r, (AUTO_YES, LOGIC_YES))

    c12 = ws.cell(r, 12)
    c12.value = auto
    c12.alignment = Alignment(horizontal='center', wrap_text=True)
    if auto == AUTO_YES:
        c12.fill = GREEN_FILL
    elif auto == AUTO_NO:
        c12.fill = RED_FILL
    else:
        c12.fill = ORANGE_FILL

    c13 = ws.cell(r, 13)
    c13.value = logic
    c13.alignment = Alignment(horizontal='center', wrap_text=True)
    if logic == LOGIC_YES:
        c13.fill = GREEN_FILL
    elif logic == LOGIC_NO:
        c13.fill = RED_FILL
    else:
        c13.fill = ORANGE_FILL

ws.column_dimensions['L'].width = 22
ws.column_dimensions['M'].width = 20

buf = io.BytesIO()
wb.save(buf)
with open(_EXCEL, 'wb') as f:
    f.write(buf.getvalue())
print("完成：已新增列 12（是否实现自动化）和列 13（业务逻辑是否清楚）")

# ── 汇总统计 ─────────────────────────────────────────────────
auto_counts  = {AUTO_YES: 0, AUTO_NO: 0, AUTO_CODE_ONLY: 0}
logic_counts = {LOGIC_YES: 0, LOGIC_NO: 0, LOGIC_PARTIAL: 0}

for r, (a, l) in decisions.items():
    auto_counts[a]  = auto_counts.get(a, 0)  + 1
    logic_counts[l] = logic_counts.get(l, 0) + 1

print("\n=== 是否实现自动化 ===")
print(f"  是               : {auto_counts[AUTO_YES]}")
print(f"  否               : {auto_counts[AUTO_NO]}")
print(f"  生成代码但不执行  : {auto_counts[AUTO_CODE_ONLY]}")

print("\n=== 业务逻辑是否清楚 ===")
print(f"  是       : {logic_counts[LOGIC_YES]}")
print(f"  否       : {logic_counts[LOGIC_NO]}")
print(f"  部分清楚  : {logic_counts[LOGIC_PARTIAL]}")
