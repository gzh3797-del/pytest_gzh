"""
基于产品结构截图分析（2026-05-08），更新 Excel 第 13 列"业务逻辑是否清楚"。

升级 PARTIAL → YES 的行（共 11 行）：
  行 6/7   : Password Management Current Password 字段显示规则已在 struct 4.6/7.2 确认
  行 62    : Special Characters hint 文本已确认，可构造具体测试密码
  行 64~67 : 三条复杂度规则（大小写/数字字母/特殊字符）及 hint 完整确认
  行 69~70 : 同上
  行 130   : Multiple Login / Override Password Policy / Override Password Expire 字段名已在 struct 4.4 确认
  行 131   : Override Password Expire 已确认，时间偏移是执行手段非逻辑不清

保留 PARTIAL 的行（共 8 行）：
  行 4   : admin 修改自身密码后是否自动退出 — 未经截图验证
  行 9   : 默认密码依赖设备 SN 后 6 位 — 测试环境固定但通用性待确认
  行 16  : EULA "不接受"按钮文本 — 未见 EULA 弹框截图
  行 23  : 默认密码登录提醒弹框 "close" 按钮文本 — 与其他用例 Cancel 不一致，待确认
  行 34  : 用例步骤写"系统设置-视图"但预期写"系统设置-编辑"，文档自身矛盾
  行 45  : 用例步骤写"固件更新-编辑"但预期写"固件更新-视图"，文档自身矛盾
  行 57  : 步骤中 Password/Repeat Password 值不一致，疑似笔误
  行 150 : 用户名超 40 位行为（保存失败 vs 自动截断）仍不确定
"""
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Alignment

_EXCEL = Path(__file__).parent.parent / 'Manual_testcase' / 'AcuHMI-1-7_用户管理用例.xlsx'

LOGIC_YES     = "是"
LOGIC_PARTIAL = "部分清楚"

GREEN_FILL  = PatternFill('solid', fgColor='C6EFCE')
ORANGE_FILL = PatternFill('solid', fgColor='FFEB9C')

# 本次升级：PARTIAL → YES
upgrades = {
    6:   "admin 修改他人密码无 Current Password 字段；非 admin 有——struct §4.6/7.2 已确认",
    7:   "同行 6",
    62:  "Special Characters hint 已确认（e.g. '@#$%'），可直接构造无特殊字符测试密码",
    64:  "大小写/数字字母/特殊字符三条规则及 hint 文本已完整确认",
    65:  "同行 64",
    66:  "同行 64",
    67:  "同行 64",
    69:  "同行 64",
    70:  "同行 64",
    130: "Multiple Login / Override Password Policy / Override Password Expire 字段名已在 struct §4.4 确认",
    131: "Override Password Expire 字段名已确认；时间偏移是执行手段，不影响逻辑清晰度",
}

wb = openpyxl.load_workbook(_EXCEL)
ws = wb['Sheet1']

changed = []
for row_num, reason in upgrades.items():
    cell = ws.cell(row_num, 13)
    old_val = cell.value
    if old_val == LOGIC_PARTIAL:
        cell.value = LOGIC_YES
        cell.fill  = GREEN_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        changed.append((row_num, reason))

buf = io.BytesIO()
wb.save(buf)
with open(_EXCEL, 'wb') as f:
    f.write(buf.getvalue())

print("完成：共更新 %d 行 [业务逻辑是否清楚] 从 [部分清楚] -> [是]" % len(changed))
print()
for r, reason in changed:
    print(f"  行 {r:3d}: {reason}")
