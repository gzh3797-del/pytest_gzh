"""Check Excel column headers and a few data rows."""
import io, sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
excel = Path(__file__).parent.parent / 'Manual_testcase' / 'AcuHMI-1-7_用户管理用例.xlsx'
data = excel.read_bytes()
wb = openpyxl.load_workbook(io.BytesIO(data))
ws = wb['Sheet1']

# Print headers (row 1)
print("=== Headers (row 1) ===")
for j, cell in enumerate(ws[1], start=1):
    print(f"  col {j} ({chr(64+j)}): {cell.value!r}")

print()
print("=== Row 2 sample ===")
for j, cell in enumerate(ws[2], start=1):
    print(f"  col {j}: {cell.value!r}")

print()
print("=== Row 64 sample ===")
for j, cell in enumerate(ws[64], start=1):
    print(f"  col {j}: {cell.value!r}")
