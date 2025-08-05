#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @File     :AcuRev4100_fast_test.py
# @Author   :lcs
# @Time     :2025/8/5
# @Desc     :

import time
from AcuRev4100_modbus_get import *
import threading
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter

Log(str(__file__).split("\\")[-1])

# 设置全局异常钩子
sys.excepthook = global_exception_handler


yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_font = Font(color='FF0000')


# ModbusClient = ModbusRtuOrTcp(conn_mode=modbus_config['conn_mode'])

def e1_2w_fast_precision_measure(CT_type):
    set_service_configuration(0)
    input_list = []
    if CT_type == "mV":
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试')
    if CT_type == "mA":
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试 mA型CT')
    my_workbook = xlwt.Workbook()
    sheet = my_workbook.add_sheet('快速测试', cell_overwrite_ok=True)
    for i in range(len(input_list[0])):
        sheet.write(0, i, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        sheet.write(i + 1, 0, input_list[i + 1][0])
        sheet.write(i + 1, 1, input_list[i + 1][1])
        sheet.write(i + 1, 2, input_list[i + 1][2])
        sheet.write(i + 1, 3, input_list[i + 1][3])
        sheet.write(i + 1, 4, input_list[i + 1][4])
        sheet.write(i + 1, 5, input_list[i + 1][5])
        sheet.write(i + 1, 6, input_list[i + 1][6])
        if input_list[i + 1][0] != 'null':
            input_angle = e1_2w_pf_calculate_angle(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4], input_angle[5],
                         0, 0, input_list[i + 1][3], 0, 0, input_list[i + 1][5],
                         input_list[i + 1][2])
            Read_Phase_A_Vol = read_phase_a_voltage_new(input_list[i + 1][3], times=10)
            sheet.write(i + 1, 7, f'{Read_Phase_A_Vol[0]}')
            sheet.write(i + 1, 8, f'{Read_Phase_A_Vol[1]}')
            sheet.write(i + 1, 9, f'{Read_Phase_A_Vol[2]}')
            sheet.write(i + 1, 10, f'{input_list[i + 1][10]}')
            if Read_Phase_A_Vol[0][1] < input_list[i + 1][10] and Read_Phase_A_Vol[1][1] < input_list[i + 1][10] and \
                    Read_Phase_A_Vol[2][1] < input_list[i + 1][10]:
                sheet.write(i + 1, 11, "Passed")
            else:
                sheet.write(i + 1, 11, "Failed")
            Read_Input_Cur = read_input_current(input_list[i + 1][0], input_list[i + 1][5], times=10)
            sheet.write(i + 1, 12, f'{Read_Input_Cur[0]}')
            sheet.write(i + 1, 13, f'{Read_Input_Cur[1]}')
            sheet.write(i + 1, 14, f'{Read_Input_Cur[2]}')
            sheet.write(i + 1, 15, f'{input_list[i + 1][15]}')
            if Read_Input_Cur[0][1] < input_list[i + 1][15] and Read_Input_Cur[1][1] < input_list[i + 1][15] and \
                    Read_Input_Cur[2][1] < input_list[i + 1][15]:
                sheet.write(i + 1, 16, "Passed")
            else:
                sheet.write(i + 1, 16, "Failed")
            Phase_A_Voltage_Angle = read_phase_a_voltage_angle_new(input_angle[2], times=10)
            Input_Channel_Current_Angle = read_input_channel_current_angle_new(input_list[i + 1][0], input_angle[5],
                                                                               times=10)
            sheet.write(i + 1, 17, f'{Phase_A_Voltage_Angle}')
            sheet.write(i + 1, 18, f'{Input_Channel_Current_Angle}')
            sheet.write(i + 1, 19, f'{input_list[i + 1][19]}')
            if isinstance(input_list[i + 1][19], float):
                if input_angle[2] == 0:
                    if 359.9 <= Phase_A_Voltage_Angle or Phase_A_Voltage_Angle <= 0.1:
                        sheet.write(i + 1, 20, "Passed")
                        logging.info("Passed")
                    else:
                        sheet.write(i + 1, 20, "Failed")
                elif abs(Phase_A_Voltage_Angle - input_angle[2]) <= 0.1:
                    sheet.write(i + 1, 20, "Passed")
                    logging.info("Passed")
                else:
                    sheet.write(i + 1, 20, "Failed")
                if input_angle[5] == 0:
                    if 359.9 <= Input_Channel_Current_Angle or Input_Channel_Current_Angle <= 0.1:
                        sheet.write(i + 1, 21, "Passed")
                        logging.info("Passed")
                    else:
                        sheet.write(i + 1, 21, "Failed")
                        logging.info("failed")
                elif abs(Input_Channel_Current_Angle - input_angle[5]) <= 0.1:
                    sheet.write(i + 1, 21, "Passed")
                    logging.info("Passed")
                else:
                    sheet.write(i + 1, 21, "Failed")
                    logging.info("failed")
            Active_Power_standard = active_power_calculate(input_list[i + 1][3], input_list[i + 1][5], input_angle[2],
                                                           input_angle[5])
            Read_Input_Active_val = read_input_active_power(input_list[i + 1][0], Active_Power_standard, times=10)
            sheet.write(i + 1, 22, f'{Read_Input_Active_val[0]}')
            sheet.write(i + 1, 23, f'{Read_Input_Active_val[1]}')
            sheet.write(i + 1, 24, f'{Read_Input_Active_val[2]}')
            sheet.write(i + 1, 25, f'{input_list[i + 1][25]}')
            if isinstance(input_list[i + 1][25], float):
                if Read_Input_Active_val[0][1] < input_list[i + 1][25] and Read_Input_Active_val[1][1] < \
                        input_list[i + 1][25] and Read_Input_Active_val[2][1] < input_list[i + 1][25]:
                    sheet.write(i + 1, 26, "Passed")
                    logging.info("Passed")
                else:
                    sheet.write(i + 1, 26, "Failed")
                    logging.info("failed")
            Reactive_Power_standard = reactive_power_calculate(input_list[i + 1][3], input_list[i + 1][5],
                                                               input_angle[2], input_angle[5])
            Input_Channel_Reactive_Power = read_input_channel_reactive_power_new(input_list[i + 1][0],
                                                                                 Reactive_Power_standard, times=10)
            sheet.write(i + 1, 27, f'{Input_Channel_Reactive_Power[0]}')
            sheet.write(i + 1, 28, f'{Input_Channel_Reactive_Power[1]}')
            sheet.write(i + 1, 29, f'{Input_Channel_Reactive_Power[2]}')
            Apparent_Power = apparent_power_calculate(input_list[i + 1][3], input_list[i + 1][5])
            Read_Input_Apparent_Power = read_input_channel_apparent_power(input_list[i + 1][0], Apparent_Power,
                                                                          times=10)
            sheet.write(i + 1, 30, f'{Read_Input_Apparent_Power[0]}')
            sheet.write(i + 1, 31, f'{Read_Input_Apparent_Power[1]}')
            sheet.write(i + 1, 32, f'{Read_Input_Apparent_Power[2]}')
        my_workbook.save('Precision_Measure_{}.xls'.format(time.strftime('%Y%m%d')))
    my_workbook.save('Precision_Measure_{}.xls'.format(time.strftime('%Y%m%d%H%M%S')))


def e1_2w_fast_precision_measure_tp26_33(CT_type):
    set_service_configuration(0)
    input_list = []
    if CT_type == "mV":
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试mV型CT_new')
    if CT_type == "mA":
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试 mA型CT_new')
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        # if input_list[i + 1][0] == 1 and input_list[i + 1][1] in ('Test Point 20', 'Test Point 21',
        #                                                           'Test Point 22', 'Test Point 26', 'Test Point 27',
        #                                                           'Test Point 28',
        #                                                           'Test Point 29', 'Test Point 30', 'Test Point 31',
        #                                                           'Test Point 32', 'Test Point 33'):
        if input_list[i + 1][0] == 1:
            input_angle = e1_2w_pf_calculate_angle(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4], input_angle[5],
                         0, 0, input_list[i + 1][3], 0, 0, input_list[i + 1][5],
                         input_list[i + 1][2])
            # time.sleep(15)
            Read_Phase_A_Vol = read_phase_a_voltage_new(input_list[i + 1][3], times=10)
            ws.cell(i + 2, 8, f'{Read_Phase_A_Vol[0]}')
            ws.cell(i + 2, 9, f'{Read_Phase_A_Vol[1]}')
            ws.cell(i + 2, 10, f'{Read_Phase_A_Vol[2]}')
            ws.cell(i + 2, 11, f'{input_list[i + 1][10]}')
            if Read_Phase_A_Vol[0][1] < input_list[i + 1][10] and Read_Phase_A_Vol[1][1] < input_list[i + 1][10] and \
                    Read_Phase_A_Vol[2][1] < input_list[i + 1][10]:
                ws.cell(i + 2, 12, "Passed")
            else:
                ws.cell(i + 2, 12, "Failed")
            for j in range(24):
                k = j + 1
                j = j * 21
                Read_Input_Cur = read_input_current(k, input_list[i + 1][5], times=10)
                ws.cell(i + 2, 13 + j, f'{Read_Input_Cur[0]}')
                ws.cell(i + 2, 14 + j, f'{Read_Input_Cur[1]}')
                ws.cell(i + 2, 15 + j, f'{Read_Input_Cur[2]}')
                ws.cell(i + 2, 16 + j, f'{input_list[i + 1][15]}')
                if Read_Input_Cur[0][1] < input_list[i + 1][15] and Read_Input_Cur[1][1] < input_list[i + 1][15] and \
                        Read_Input_Cur[2][1] < input_list[i + 1][15]:
                    ws.cell(i + 2, 17 + j, "Passed")
                else:
                    ws.cell(i + 2, 17 + j, "Failed")
                Phase_A_Voltage_Angle = read_phase_a_voltage_angle_new(input_angle[2], times=10)
                Input_Channel_Current_Angle = read_input_channel_current_angle_new(k, input_angle[5],
                                                                                   times=10)
                ws.cell(i + 2, 18 + j, f'{Phase_A_Voltage_Angle}')
                ws.cell(i + 2, 19 + j, f'{Input_Channel_Current_Angle}')
                ws.cell(i + 2, 20 + j, f'{input_list[i + 1][19]}')
                if isinstance(input_list[i + 1][19], float):
                    if input_angle[2] == 0:
                        if 359.9 <= Phase_A_Voltage_Angle or Phase_A_Voltage_Angle <= 0.1:
                            ws.cell(i + 2, 21 + j, "Passed")
                            logging.info("Passed")
                        else:
                            ws.cell(i + 2, 21 + j, "Failed")
                    elif abs(Phase_A_Voltage_Angle - input_angle[2]) <= 0.1:
                        ws.cell(i + 2, 21 + j, "Passed")
                        logging.info("Passed")
                    else:
                        ws.cell(i + 2, 21 + j, "Failed")
                    if input_angle[5] == 0:
                        if 359.9 <= Input_Channel_Current_Angle or Input_Channel_Current_Angle <= 0.1:
                            ws.cell(i + 2, 22 + j, "Passed")
                            logging.info("Passed")
                        else:
                            ws.cell(i + 2, 22 + j, "Failed")
                            logging.info("failed")
                    elif abs(Input_Channel_Current_Angle - input_angle[5]) <= 0.1:
                        ws.cell(i + 2, 22 + j, "Passed")
                        logging.info("Passed")
                    else:
                        ws.cell(i + 2, 22 + j, "Failed")
                        logging.info("failed")
                Active_Power_standard = active_power_calculate(input_list[i + 1][3], input_list[i + 1][5],
                                                               input_angle[2],
                                                               input_angle[5])
                Read_Input_Active_val = read_input_active_power(k, Active_Power_standard, times=30)
                ws.cell(i + 2, 23 + j, f'{Read_Input_Active_val[0]}')
                ws.cell(i + 2, 24 + j, f'{Read_Input_Active_val[1]}')
                ws.cell(i + 2, 25 + j, f'{Read_Input_Active_val[2]}')
                ws.cell(i + 2, 26 + j, f'{input_list[i + 1][25]}')
                if isinstance(input_list[i + 1][25], float):
                    if Read_Input_Active_val[0][1] < input_list[i + 1][25] and Read_Input_Active_val[1][1] < \
                            input_list[i + 1][25] and Read_Input_Active_val[2][1] < input_list[i + 1][25]:
                        ws.cell(i + 2, 27 + j, "Passed")
                        logging.info("Passed")
                    else:
                        ws.cell(i + 2, 27 + j, "Failed")
                        logging.info("failed")
                Reactive_Power_standard = reactive_power_calculate(input_list[i + 1][3], input_list[i + 1][5],
                                                                   input_angle[2], input_angle[5])
                Input_Channel_Reactive_Power = read_input_channel_reactive_power_new(k,
                                                                                     Reactive_Power_standard, times=10)
                ws.cell(i + 2, 28 + j, f'{Input_Channel_Reactive_Power[0]}')
                ws.cell(i + 2, 29 + j, f'{Input_Channel_Reactive_Power[1]}')
                ws.cell(i + 2, 30 + j, f'{Input_Channel_Reactive_Power[2]}')
                Apparent_Power = apparent_power_calculate(input_list[i + 1][3], input_list[i + 1][5])
                Read_Input_Apparent_Power = read_input_channel_apparent_power(k, Apparent_Power, times=10)
                ws.cell(i + 2, 31 + j, f'{Read_Input_Apparent_Power[0]}')
                ws.cell(i + 2, 32 + j, f'{Read_Input_Apparent_Power[1]}')
                ws.cell(i + 2, 33 + j, f'{Read_Input_Apparent_Power[2]}')
        wb.save('E1_2W_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d')))
    wb.save('E1_2W_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


def e1_2w_fast_precision_measure_no_load():
    set_service_configuration(4)
    input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', 'No Load')
    # my_workbook = xlwt.workbook()
    # sheet = my_workbook.add_sheet('快速测试', cell_overwrite_ok=True)
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        # if input_list[i + 1][0] == 1 and input_list[i + 1][1] in (
        #         'Test Point 21', 'Test Point 26', 'Test Point 27', 'Test Point 28',
        #         'Test Point 29', 'Test Point 30', 'Test Point 31','Test Point 32','Test Point 33'):
        if input_list[i + 1][0] == 1 and input_list[i + 1][1] in ('Test Point 34', 'Test Point 35'):
            input_angle = e1_2w_pf_calculate_angle(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4], input_angle[5],
                         0, 0, input_list[i + 1][3], 0, 0, input_list[i + 1][5],
                         input_list[i + 1][2])
            time.sleep(15)
            Read_Phase_A_Vol = read_phase_a_voltage_new(input_list[i + 1][3], times=10)
            ws.cell(i + 2, 8, f'{Read_Phase_A_Vol[0]}')
            ws.cell(i + 2, 9, f'{Read_Phase_A_Vol[1]}')
            ws.cell(i + 2, 10, f'{Read_Phase_A_Vol[2]}')
            ws.cell(i + 2, 11, f'{input_list[i + 1][10]}')
            if Read_Phase_A_Vol[0][1] < input_list[i + 1][10] and Read_Phase_A_Vol[1][1] < input_list[i + 1][10] and \
                    Read_Phase_A_Vol[2][1] < input_list[i + 1][10]:
                ws.cell(i + 2, 12, "Passed")
            else:
                ws.cell(i + 2, 12, "Failed")
            for j in range(24):
                k = j + 1
                j = j + 1
                j = (j - 1) * 4
                Read_Input_Cur = read_input_current(k, input_list[i + 1][5], times=10)
                ws.cell(i + 2, 13 + j, f'{Read_Input_Cur[2][0]}')
                Active_Power_standard = active_power_calculate(input_list[i + 1][3], input_list[i + 1][5],
                                                               input_angle[2],
                                                               input_angle[5])
                Read_Input_Active_val = read_input_active_power(k, Active_Power_standard, times=10)
                ws.cell(i + 2, 14 + j, f'{Read_Input_Active_val[2][0]}')
                Reactive_Power_standard = reactive_power_calculate(input_list[i + 1][3], input_list[i + 1][5],
                                                                   input_angle[2], input_angle[5])
                Input_Channel_Reactive_Power = read_input_channel_reactive_power_new(k,
                                                                                     Reactive_Power_standard, times=10)
                ws.cell(i + 2, 15 + j, f'{Input_Channel_Reactive_Power[2][0]}')
                Apparent_Power = apparent_power_calculate(input_list[i + 1][3], input_list[i + 1][5])
                Read_Input_Apparent_Power = read_input_channel_apparent_power(k, Apparent_Power, times=10)
                ws.cell(i + 2, 16 + j, f'{Read_Input_Apparent_Power[2][0]}')
    wb.save('No_Load_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


def e3_4w_y_fast_precision_measure_new(Service, CT_type, input_ch_num: int):
    input_list = []
    if Service == "3E4WY":
        set_service_configuration(4)
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试_其他接线方式')
    if Service == "2E3W1P":
        set_service_configuration(1)
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试_其他接线方式')
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            logging.info('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        if input_list[i + 1][0] == CT_type:
            input_angle = []
            if Service == "3E4WY":
                input_angle = e3_4w_y_pf_calculate_angle_new(input_list[i + 1][6])
            if Service == "2E3W1P":
                input_angle = e2_3w_1p_pf_calculate_angle_new(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            input_current = input_list[i + 1][5]
            if CT_type == "RCT16_100" or CT_type == "RCT16_500" or CT_type == "RCT16_1000":
                input_current = input_list[i + 1][5] / 10
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            Phase_L_N_Voltage = []
            if Service == "3E4WY":
                ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                             input_angle[5],
                             input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3], input_current,
                             input_current, input_current,
                             input_list[i + 1][2])
                Phase_L_N_Voltage = read_phase_l_n_voltage(
                    [input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3]], times=10)
            if Service == "2E3W1P":
                ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                             input_angle[5],
                             input_list[i + 1][3], 0, input_list[i + 1][3], 0,
                             input_list[i + 1][5], input_list[i + 1][5],
                             input_list[i + 1][2])
                Phase_L_N_Voltage = read_phase_l_n_voltage(
                    [input_list[i + 1][3], 0, input_list[i + 1][3]], times=10)
            for j in range(9):
                ws.cell(i + 2, 8 + j, f'{Phase_L_N_Voltage[j]}')
            ws.cell(i + 2, 17, f'{input_list[i + 1][16]}')
            for j in range(9):
                if Phase_L_N_Voltage[j][1] <= input_list[i + 1][16]:
                    ws.cell(i + 2, 18, f'Passed')
                else:
                    cell = ws.cell(i + 2, 18, f'Failed')
                    cell.fill = yellow_fill
                    cell.font = red_font
                    break
            Phase_Angle_Voltage = read_phase_angle_voltage([input_angle[2], input_angle[1], input_angle[0]], times=10)
            for j in range(9):
                ws.cell(i + 2, 19 + j, f'{Phase_Angle_Voltage[j]}')
            ws.cell(i + 2, 28, f'{input_list[i + 1][27]}')
            for j in range(9):
                if Phase_Angle_Voltage[j][1] <= input_list[i + 1][27]:
                    ws.cell(i + 2, 29, f'Passed')
                else:
                    cell = ws.cell(i + 2, 29, f'Failed')
                    cell.fill = yellow_fill
                    cell.font = red_font
                    break
            input_Power_standard = []
            if Service == "3E4WY":
                input_Power_standard = e3_4w_y_input_power_standard_value(input_list[i + 1][3],
                                                                          input_list[i + 1][3],
                                                                          input_list[i + 1][3],
                                                                          input_list[i + 1][5],
                                                                          input_list[i + 1][5],
                                                                          input_list[i + 1][5],
                                                                          input_angle[2], input_angle[5],
                                                                          input_angle[1], input_angle[4],
                                                                          input_angle[0], input_angle[3], input_ch_num)
            if Service == "2E3W1P":
                input_Power_standard = e2_3w_1p_input_power_standard_value(input_list[i + 1][3], input_list[i + 1][3],
                                                                           input_list[i + 1][5],
                                                                           input_list[i + 1][5], input_angle[2],
                                                                           input_angle[5],
                                                                           input_angle[0],
                                                                           input_angle[4])
            System_Power = read_system_power(input_Power_standard[3], times=10)
            for j in range(3):
                ws.cell(i + 2, 30 + j, f'{System_Power[j]}')
            ws.cell(i + 2, 33, f'{input_list[i + 1][32]}')
            for j in range(3):
                if System_Power[j][1] <= input_list[i + 1][32]:
                    ws.cell(i + 2, 34, f'Passed')
                else:
                    cell = ws.cell(i + 2, 34, f'Failed')
                    cell.fill = yellow_fill
                    cell.font = red_font
                    break
            for j in range(3, 9):
                ws.cell(i + 2, 32 + j, f'{System_Power[j]}')

            for k in range(input_ch_num):
                input_ch = k + 1
                row = k * 21
                input1_Angle = []
                if k in (0, 3, 6, 9, 12, 15, 18, 21):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[5]], times=10)
                if k in (1, 4, 7, 10, 13, 16, 19, 22):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[4]], times=10)
                if k in (2, 5, 8, 11, 14, 17, 20, 23):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[3]], times=10)
                for j in range(3):
                    ws.cell(i + 2, 41 + j + row, f'{input1_Angle[j]}')
                ws.cell(i + 2, 44 + row, f'{input_list[i + 1][43]}')
                for j in range(3):
                    if input1_Angle[j][1] <= input_list[i + 1][43]:
                        ws.cell(i + 2, 45 + row, f'Passed')
                    else:
                        cell = ws.cell(i + 2, 45 + row, f'Failed')
                        cell.fill = yellow_fill
                        cell.font = red_font
                        break
                input1_Power = []
                if k in (0, 3, 6, 9, 12, 15, 18, 21):
                    input1_Power = read_input_power(input_ch, input_Power_standard[0], times=10)
                if k in (1, 4, 7, 10, 13, 16, 19, 22):
                    input1_Power = read_input_power(input_ch, input_Power_standard[1], times=10)
                if k in (2, 5, 8, 11, 14, 17, 20, 23):
                    input1_Power = read_input_power(input_ch, input_Power_standard[2], times=10)
                for j in range(3):
                    ws.cell(i + 2, 46 + j + row, f'{input1_Power[j]}')
                ws.cell(i + 2, 49 + row, f'{input_list[i + 1][48]}')
                for j in range(3):
                    if input1_Power[j][1] <= input_list[i + 1][48]:
                        ws.cell(i + 2, 50 + row, f'Passed')
                    else:
                        cell = ws.cell(i + 2, 50 + row, f'Failed')
                        cell.fill = yellow_fill
                        cell.font = red_font
                        break
                for j in range(3):
                    l = j + 3
                    ws.cell(i + 2, 51 + j + row, f'{input1_Power[l]}')
                ws.cell(i + 2, 54 + row, f'{input_list[i + 1][53]}')
                for j in range(3):
                    l = j + 3
                    if input1_Power[l][1] <= input_list[i + 1][53]:
                        ws.cell(i + 2, 55 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        cell = ws.cell(i + 2, 55 + row, f'Failed')
                        cell.fill = yellow_fill
                        cell.font = red_font
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 6
                    ws.cell(i + 2, 56 + j + row, f'{input1_Power[l]}')
            num = int(input_ch_num / 3)
            for k in range(num):
                input_ch = k + 1
                row = k * 11
                User_Power = read_user_power_nwe(input_ch, input_Power_standard[4], times=10)
                for j in range(3):
                    ws.cell(i + 2, 545 + j + row, f'{User_Power[j]}')
                ws.cell(i + 2, 548 + row, f'{input_list[i + 1][106]}')
                for j in range(3):
                    if User_Power[j][1] <= input_list[i + 1][106]:
                        ws.cell(i + 2, 549 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        cell = ws.cell(i + 2, 549 + row, f'Failed')
                        cell.fill = yellow_fill
                        cell.font = red_font
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 3
                    ws.cell(i + 2, 550 + j + row, f'{User_Power[l]}')
        wb.save('E3_4W_Y_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d')))
    wb.save('E3_4W_Y_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


def e2_3w1p_fast_precision_measure_new(Service, CT_type):
    input_list = []
    if Service == "3E4WY":
        set_service_configuration(4)
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试_其他接线方式')
    if Service == "2E3W1P":
        set_service_configuration(1)
        input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试_其他接线方式')
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        if input_list[i + 1][0] == CT_type:
            input_angle = []
            if Service == "3E4WY":
                input_angle = e3_4w_y_pf_calculate_angle_new(input_list[i + 1][6])
            if Service == "2E3W1P":
                input_angle = e2_3w_1p_pf_calculate_angle_new(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            Phase_L_N_Voltage = []
            if Service == "3E4WY":
                ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                             input_angle[5],
                             input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][5],
                             input_list[i + 1][5], input_list[i + 1][5],
                             input_list[i + 1][2])
                Phase_L_N_Voltage = read_phase_l_n_voltage(
                    [input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3]], times=10)
            if Service == "2E3W1P":
                ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                             input_angle[5],
                             input_list[i + 1][3], 0, input_list[i + 1][3], 0,
                             input_list[i + 1][5], input_list[i + 1][5],
                             input_list[i + 1][2])
                Phase_L_N_Voltage = read_phase_l_n_voltage(
                    [input_list[i + 1][3], 0, input_list[i + 1][3]], times=10)
            for j in range(9):
                ws.cell(i + 2, 8 + j, f'{Phase_L_N_Voltage[j]}')
            ws.cell(i + 2, 17, f'{input_list[i + 1][16]}')
            for j in range(9):
                if Phase_L_N_Voltage[j][1] <= input_list[i + 1][16]:
                    ws.cell(i + 2, 18, f'Passed')
                else:
                    ws.cell(i + 2, 18, f'Failed')
                    break
            Phase_Angle_Voltage = read_phase_angle_voltage([input_angle[2], input_angle[1], input_angle[0]], times=10)
            for j in range(9):
                ws.cell(i + 2, 19 + j, f'{Phase_Angle_Voltage[j]}')
            ws.cell(i + 2, 28, f'{input_list[i + 1][27]}')
            for j in range(9):
                if Phase_Angle_Voltage[j][1] <= input_list[i + 1][27]:
                    ws.cell(i + 2, 29, f'Passed')
                else:
                    ws.cell(i + 2, 29, f'Failed')
                    break
            input_Power_standard = []
            if Service == "3E4WY":
                input_Power_standard = e3_4w_y_input_power_standard_value(input_list[i + 1][3],
                                                                          input_list[i + 1][3],
                                                                          input_list[i + 1][3],
                                                                          input_list[i + 1][5],
                                                                          input_list[i + 1][5],
                                                                          input_list[i + 1][5],
                                                                          input_angle[2], input_angle[5],
                                                                          input_angle[1], input_angle[4],
                                                                          input_angle[0], input_angle[3], 24)
            if Service == "2E3W1P":
                input_Power_standard = e2_3w_1p_input_power_standard_value(input_list[i + 1][3], input_list[i + 1][3],
                                                                           input_list[i + 1][5],
                                                                           input_list[i + 1][5], input_angle[2],
                                                                           input_angle[5],
                                                                           input_angle[0],
                                                                           input_angle[4])
            System_Power = read_system_power(input_Power_standard[3], times=10)
            for j in range(3):
                ws.cell(i + 2, 30 + j, f'{System_Power[j]}')
            ws.cell(i + 2, 33, f'{input_list[i + 1][32]}')
            for j in range(3):
                if System_Power[j][1] <= input_list[i + 1][32]:
                    ws.cell(i + 2, 34, f'Passed')
                else:
                    ws.cell(i + 2, 34, f'Failed')
                    break
            for j in range(3, 9):
                ws.cell(i + 2, 32 + j, f'{System_Power[j]}')

            for k in range(24):
                input_ch = k + 1
                row = k * 21
                input1_Angle = []
                if k in (0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[5]], times=10)
                if k in (1, 3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[4]], times=10)
                for j in range(3):
                    ws.cell(i + 2, 41 + j + row, f'{input1_Angle[j]}')
                ws.cell(i + 2, 44 + row, f'{input_list[i + 1][43]}')
                for j in range(3):
                    if input1_Angle[j][1] <= input_list[i + 1][43]:
                        ws.cell(i + 2, 45 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 45 + row, f'Failed')
                        break
                input1_Power = []
                if k in (0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22):
                    input1_Power = read_input_power(input_ch, input_Power_standard[0], times=10)
                if k in (1, 3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23):
                    input1_Power = read_input_power(input_ch, input_Power_standard[1], times=10)
                for j in range(3):
                    ws.cell(i + 2, 46 + j + row, f'{input1_Power[j]}')
                ws.cell(i + 2, 49 + row, f'{input_list[i + 1][48]}')
                for j in range(3):
                    if input1_Power[j][1] <= input_list[i + 1][48]:
                        ws.cell(i + 2, 50 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 50 + row, f'Failed')
                        break
                for j in range(3):
                    l = j + 3
                    ws.cell(i + 2, 51 + j + row, f'{input1_Power[l]}')
                ws.cell(i + 2, 54 + row, f'{input_list[i + 1][53]}')
                for j in range(3):
                    l = j + 3
                    if input1_Power[l][1] <= input_list[i + 1][53]:
                        ws.cell(i + 2, 55 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 55 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 6
                    ws.cell(i + 2, 56 + j + row, f'{input1_Power[l]}')
            for k in range(12):
                input_ch = k + 1
                row = k * 11
                User_Power = read_user_power_nwe(input_ch, input_Power_standard[4], times=10)
                for j in range(3):
                    ws.cell(i + 2, 545 + j + row, f'{User_Power[j]}')
                ws.cell(i + 2, 548 + row, f'{input_list[i + 1][106]}')
                for j in range(3):
                    if User_Power[j][1] <= input_list[i + 1][106]:
                        ws.cell(i + 2, 549 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 549 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 3
                    ws.cell(i + 2, 550 + j + row, f'{User_Power[l]}')
    wb.save('E2_3W1P_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


def e2_3w_networkfast_precision_measure_new(CT_type):
    set_service_configuration(3)
    input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '快速测试_其他接线方式')
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        if input_list[i + 1][0] == CT_type:
            input_angle = []
            input_angle = e3_4w_y_pf_calculate_angle_new(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                         input_angle[5],
                         input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][5],
                         input_list[i + 1][5], input_list[i + 1][5],
                         input_list[i + 1][2])
            Phase_L_N_Voltage = read_phase_l_n_voltage(
                [input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3]], times=10)
            for j in range(9):
                ws.cell(i + 2, 8 + j, f'{Phase_L_N_Voltage[j]}')
            ws.cell(i + 2, 17, f'{input_list[i + 1][16]}')
            for j in range(9):
                if Phase_L_N_Voltage[j][1] <= input_list[i + 1][16]:
                    ws.cell(i + 2, 18, f'Passed')
                else:
                    ws.cell(i + 2, 18, f'Failed')
                    break
            Phase_Angle_Voltage = read_phase_angle_voltage([input_angle[2], input_angle[1], input_angle[0]], times=10)
            for j in range(9):
                ws.cell(i + 2, 19 + j, f'{Phase_Angle_Voltage[j]}')
            ws.cell(i + 2, 28, f'{input_list[i + 1][27]}')
            for j in range(9):
                if Phase_Angle_Voltage[j][1] <= input_list[i + 1][27]:
                    ws.cell(i + 2, 29, f'Passed')
                else:
                    ws.cell(i + 2, 29, f'Failed')
                    break
            input_Power_standard = []
            input_Power_standard = e2_3w_network_input_power_standard_value(input_list[i + 1][3],
                                                                            input_list[i + 1][3],
                                                                            input_list[i + 1][3],
                                                                            input_list[i + 1][5],
                                                                            input_list[i + 1][5],
                                                                            input_list[i + 1][5],
                                                                            input_angle[2], input_angle[5],
                                                                            input_angle[1], input_angle[4],
                                                                            input_angle[0], input_angle[3])
            System_Power = read_system_power(input_Power_standard[3], times=10)
            for j in range(3):
                ws.cell(i + 2, 30 + j, f'{System_Power[j]}')
            ws.cell(i + 2, 33, f'{input_list[i + 1][32]}')
            for j in range(3):
                if System_Power[j][1] <= input_list[i + 1][32]:
                    ws.cell(i + 2, 34, f'Passed')
                else:
                    ws.cell(i + 2, 34, f'Failed')
                    break
            for j in range(3, 9):
                ws.cell(i + 2, 32 + j, f'{System_Power[j]}')

            for k in range(24):
                input_ch = k + 1
                row = k * 21
                input1_Angle = []
                if k in (0, 3, 6, 9, 12, 15, 18, 21):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[5]], times=10)
                if k in (1, 4, 7, 10, 13, 16, 19, 22):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[4]], times=10)
                if k in (2, 5, 8, 11, 14, 17, 20, 23):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[3]], times=10)
                for j in range(3):
                    ws.cell(i + 2, 41 + j + row, f'{input1_Angle[j]}')
                ws.cell(i + 2, 44 + row, f'{input_list[i + 1][43]}')
                for j in range(3):
                    if input1_Angle[j][1] <= input_list[i + 1][43]:
                        ws.cell(i + 2, 45 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 45 + row, f'Failed')
                        break
                input1_Power = []
                if k in (0, 3, 6, 9, 12, 15, 18, 21):
                    input1_Power = read_input_power(input_ch, input_Power_standard[0], times=10)
                if k in (1, 4, 7, 10, 13, 16, 19, 22):
                    input1_Power = read_input_power(input_ch, input_Power_standard[1], times=10)
                if k in (2, 5, 8, 11, 14, 17, 20, 23):
                    input1_Power = read_input_power(input_ch, input_Power_standard[2], times=10)
                for j in range(3):
                    ws.cell(i + 2, 46 + j + row, f'{input1_Power[j]}')
                ws.cell(i + 2, 49 + row, f'{input_list[i + 1][48]}')
                for j in range(3):
                    if input1_Power[j][1] <= input_list[i + 1][48]:
                        ws.cell(i + 2, 50 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 50 + row, f'Failed')
                        break
                for j in range(3):
                    l = j + 3
                    ws.cell(i + 2, 51 + j + row, f'{input1_Power[l]}')
                ws.cell(i + 2, 54 + row, f'{input_list[i + 1][53]}')
                for j in range(3):
                    l = j + 3
                    if input1_Power[l][1] <= input_list[i + 1][53]:
                        ws.cell(i + 2, 55 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 55 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 6
                    ws.cell(i + 2, 56 + j + row, f'{input1_Power[l]}')
            for k in range(12):
                input_ch = k + 1
                row = k * 11
                User_Power = read_user_power_nwe(input_ch, input_Power_standard[4], times=10)
                for j in range(3):
                    ws.cell(i + 2, 545 + j + row, f'{User_Power[j]}')
                ws.cell(i + 2, 548 + row, f'{input_list[i + 1][106]}')
                for j in range(3):
                    if User_Power[j][1] <= input_list[i + 1][106]:
                        ws.cell(i + 2, 549 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 549 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 3
                    ws.cell(i + 2, 550 + j + row, f'{User_Power[l]}')
    wb.save('E2_3W_Network_Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


def e3_3w_delta_fast_precision_measure_new():
    set_service_configuration(5)
    input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '3E_3W_Delta接线 快速测试')
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        if input_list[i + 1][0] == "mA":
            input_angle = []
            input_angle = e3_3w_delta_pf_calculate_angle_new(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                         input_angle[5],
                         input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][5],
                         input_list[i + 1][5], input_list[i + 1][5],
                         input_list[i + 1][2])
            line_to_line_voltage = line_to_line_voltage_calculate(input_list[i + 1][3], input_list[i + 1][3],
                                                                  input_list[i + 1][3],
                                                                  input_angle[2], input_angle[1], input_angle[0])
            Phase_L_L_Voltage = read_phase_l_l_voltage(line_to_line_voltage, times=10)
            for j in range(9):
                ws.cell(i + 2, 8 + j, f'{Phase_L_L_Voltage[j]}')
            ws.cell(i + 2, 17, f'{input_list[i + 1][16]}')
            for j in range(9):
                if Phase_L_L_Voltage[j][1] <= input_list[i + 1][16]:
                    ws.cell(i + 2, 18, f'Passed')
                else:
                    ws.cell(i + 2, 18, f'Failed')
                    break
            Phase_Angle_Voltage = read_phase_angle_voltage([input_angle[2], input_angle[1], input_angle[0]], times=10)
            for j in range(9):
                ws.cell(i + 2, 19 + j, f'{Phase_Angle_Voltage[j]}')
            ws.cell(i + 2, 28, f'{input_list[i + 1][27]}')
            for j in range(9):
                if Phase_Angle_Voltage[j][1] <= input_list[i + 1][27]:
                    ws.cell(i + 2, 29, f'Passed')
                else:
                    ws.cell(i + 2, 29, f'Failed')
                    break
            input_Power_standard = []
            Vab = line_to_line_voltage[0]
            Vbc = line_to_line_voltage[1]
            Vca = line_to_line_voltage[2]
            input_Power_standard = e3_3w_delta_power_standard_value_new(Vab, Vbc, Vca,
                                                                        input_list[i + 1][5],
                                                                        input_list[i + 1][5],
                                                                        input_list[i + 1][5],
                                                                        input_angle[2], input_angle[5],
                                                                        input_angle[1], input_angle[4],
                                                                        input_angle[0], input_angle[3])
            System_Power = read_system_power(input_Power_standard[3], times=10)
            for j in range(3):
                ws.cell(i + 2, 30 + j, f'{System_Power[j]}')
            ws.cell(i + 2, 33, f'{input_list[i + 1][32]}')
            for j in range(3):
                if System_Power[j][1] <= input_list[i + 1][32]:
                    ws.cell(i + 2, 34, f'Passed')
                else:
                    ws.cell(i + 2, 34, f'Failed')
                    break
            for j in range(3, 9):
                ws.cell(i + 2, 32 + j, f'{System_Power[j]}')
            input_angle[5] = input_angle[5] - 30
            input_angle[4] = input_angle[4] - 30
            input_angle[3] = input_angle[3] - 30
            for k in range(24):
                input_ch = k + 1
                row = k * 21
                input1_Angle = []
                if k in (0, 3, 6, 9, 12, 15, 18, 21):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[5]], times=10)
                if k in (1, 4, 7, 10, 13, 16, 19, 22):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[4]], times=10)
                if k in (2, 5, 8, 11, 14, 17, 20, 23):
                    input1_Angle = read_input_angle_current(input_ch, [input_angle[3]], times=10)
                for j in range(3):
                    ws.cell(i + 2, 41 + j + row, f'{input1_Angle[j]}')
                ws.cell(i + 2, 44 + row, f'{input_list[i + 1][43]}')
                for j in range(3):
                    if input1_Angle[j][1] <= input_list[i + 1][43]:
                        ws.cell(i + 2, 45 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 45 + row, f'Failed')
                        break
                input1_Power = []
                if k in (0, 3, 6, 9, 12, 15, 18, 21):
                    input1_Power = read_input_power(input_ch, input_Power_standard[0], times=10)
                if k in (1, 4, 7, 10, 13, 16, 19, 22):
                    input1_Power = read_input_power(input_ch, input_Power_standard[1], times=10)
                if k in (2, 5, 8, 11, 14, 17, 20, 23):
                    input1_Power = read_input_power(input_ch, input_Power_standard[2], times=10)
                for j in range(3):
                    ws.cell(i + 2, 46 + j + row, f'{input1_Power[j]}')
                ws.cell(i + 2, 49 + row, f'{input_list[i + 1][48]}')
                for j in range(3):
                    if input1_Power[j][1] <= input_list[i + 1][48]:
                        ws.cell(i + 2, 50 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 50 + row, f'Failed')
                        break
                for j in range(3):
                    l = j + 3
                    ws.cell(i + 2, 51 + j + row, f'{input1_Power[l]}')
                ws.cell(i + 2, 54 + row, f'{input_list[i + 1][53]}')
                for j in range(3):
                    l = j + 3
                    if input1_Power[l][1] <= input_list[i + 1][53]:
                        ws.cell(i + 2, 55 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 55 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 6
                    ws.cell(i + 2, 56 + j + row, f'{input1_Power[l]}')
            for k in range(12):
                input_ch = k + 1
                row = k * 11
                User_Power = read_user_power_nwe(input_ch, input_Power_standard[4], times=10)
                for j in range(3):
                    ws.cell(i + 2, 545 + j + row, f'{User_Power[j]}')
                ws.cell(i + 2, 548 + row, f'{input_list[i + 1][106]}')
                for j in range(3):
                    if User_Power[j][1] <= input_list[i + 1][106]:
                        ws.cell(i + 2, 549 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 549 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 3
                    ws.cell(i + 2, 550 + j + row, f'{User_Power[l]}')
    wb.save('Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


def e2_3w_delta_fast_precision_measure_new(CT_type):
    set_service_configuration(2)
    input_list = data_read(r'./test_case/AcuRev4100/4100精度快速测试输入.xlsx', '2E_3W_Delta接线 快速测试')
    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(input_list[0])):
        j = i + 1
        ws.cell(1, j, f'{input_list[0][i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            break
        if i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        ws.cell(i + 2, 1, input_list[i + 1][0])
        ws.cell(i + 2, 2, input_list[i + 1][1])
        ws.cell(i + 2, 3, input_list[i + 1][2])
        ws.cell(i + 2, 4, input_list[i + 1][3])
        ws.cell(i + 2, 5, input_list[i + 1][4])
        ws.cell(i + 2, 6, input_list[i + 1][5])
        ws.cell(i + 2, 7, input_list[i + 1][6])
        if input_list[i + 1][0] == CT_type:
            input_angle = []
            input_angle = e2_3w_delta_pf_calculate_angle_new(input_list[i + 1][6])
            input_list[i + 1][5] = round(input_list[i + 1][5], 10)
            ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
            set_voltage_gear(input_list[i + 1][3])
            set_current_gear(input_list[i + 1][5])
            ret = set_ac(input_angle[0], input_angle[1], input_angle[2], input_angle[3], input_angle[4],
                         input_angle[5],
                         input_list[i + 1][3], input_list[i + 1][3], input_list[i + 1][3], 0,
                         input_list[i + 1][5], input_list[i + 1][5],
                         input_list[i + 1][2])
            line_to_line_voltage = line_to_line_voltage_calculate(input_list[i + 1][3], input_list[i + 1][3],
                                                                  input_list[i + 1][3],
                                                                  input_angle[2], input_angle[1], input_angle[0])
            Phase_L_L_Voltage = read_phase_l_l_voltage(line_to_line_voltage, times=10)
            for j in range(9):
                ws.cell(i + 2, 8 + j, f'{Phase_L_L_Voltage[j]}')
            ws.cell(i + 2, 17, f'{input_list[i + 1][16]}')
            for j in range(9):
                if Phase_L_L_Voltage[j][1] <= input_list[i + 1][16]:
                    ws.cell(i + 2, 18, f'Passed')
                else:
                    ws.cell(i + 2, 18, f'Failed')
                    break
            Phase_Angle_Voltage = read_phase_angle_voltage([input_angle[2], input_angle[1], input_angle[0]], times=10)
            for j in range(9):
                ws.cell(i + 2, 19 + j, f'{Phase_Angle_Voltage[j]}')
            ws.cell(i + 2, 28, f'{input_list[i + 1][27]}')
            for j in range(9):
                if Phase_Angle_Voltage[j][1] <= input_list[i + 1][27]:
                    ws.cell(i + 2, 29, f'Passed')
                else:
                    ws.cell(i + 2, 29, f'Failed')
                    break
            input_Power_standard = []
            Vab = line_to_line_voltage[0]
            Vbc = line_to_line_voltage[1]
            Vca = line_to_line_voltage[2]
            input_Power_standard = e2_3w_delta_power_standard_value_new(Vab, Vbc, Vca,
                                                                        input_list[i + 1][5],
                                                                        input_list[i + 1][5],
                                                                        input_list[i + 1][5],
                                                                        input_angle[2], input_angle[5],
                                                                        input_angle[1], input_angle[4],
                                                                        input_angle[0], input_angle[3])
            System_Power = read_system_power(input_Power_standard[3], times=10)
            for j in range(3):
                ws.cell(i + 2, 30 + j, f'{System_Power[j]}')
            ws.cell(i + 2, 33, f'{input_list[i + 1][32]}')
            for j in range(3):
                if System_Power[j][1] <= input_list[i + 1][32]:
                    ws.cell(i + 2, 34, f'Passed')
                else:
                    ws.cell(i + 2, 34, f'Failed')
                    break
            for j in range(3, 9):
                ws.cell(i + 2, 32 + j, f'{System_Power[j]}')
            Ia = input_angle[5] - 30
            if Ia < 0:
                Ia = Ia + 360
            Ib = input_angle[4] - 30
            if Ib < 0:
                Ib = Ib + 360
            Ic = input_angle[3] - 30
            if Ic < 0:
                Ic = Ic + 360
            for k in range(24):
                input_ch = k + 1
                row = k * 21
                input1_Angle = []
                if k in (0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22):
                    input1_Angle = read_input_angle_current(input_ch, [Ia], times=10)
                if k in (1, 3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23):
                    input1_Angle = read_input_angle_current(input_ch, [Ib], times=10)
                for j in range(3):
                    ws.cell(i + 2, 41 + j + row, f'{input1_Angle[j]}')
                ws.cell(i + 2, 44 + row, f'{input_list[i + 1][43]}')
                for j in range(3):
                    if input1_Angle[j][1] <= input_list[i + 1][43]:
                        ws.cell(i + 2, 45 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 45 + row, f'Failed')
                        break
                input1_Power = []
                if k in (0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22):
                    input1_Power = read_input_power(input_ch, input_Power_standard[0], times=10)
                if k in (1, 3, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23):
                    input1_Power = read_input_power(input_ch, input_Power_standard[1], times=10)
                for j in range(3):
                    ws.cell(i + 2, 46 + j + row, f'{input1_Power[j]}')
                ws.cell(i + 2, 49 + row, f'{input_list[i + 1][48]}')
                for j in range(3):
                    if input1_Power[j][1] <= input_list[i + 1][48]:
                        ws.cell(i + 2, 50 + row, f'Passed')
                    else:
                        ws.cell(i + 2, 50 + row, f'Failed')
                        break
                for j in range(3):
                    l = j + 3
                    ws.cell(i + 2, 51 + j + row, f'{input1_Power[l]}')
                ws.cell(i + 2, 54 + row, f'{input_list[i + 1][53]}')
                for j in range(3):
                    l = j + 3
                    if input1_Power[l][1] <= input_list[i + 1][53]:
                        ws.cell(i + 2, 55 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 55 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 6
                    ws.cell(i + 2, 56 + j + row, f'{input1_Power[l]}')
            for k in range(12):
                input_ch = k + 1
                row = k * 11
                User_Power = read_user_power_nwe(input_ch, input_Power_standard[4], times=10)
                for j in range(3):
                    ws.cell(i + 2, 545 + j + row, f'{User_Power[j]}')
                ws.cell(i + 2, 548 + row, f'{input_list[i + 1][106]}')
                for j in range(3):
                    if User_Power[j][1] <= input_list[i + 1][106]:
                        ws.cell(i + 2, 549 + row, f'Passed')
                        logging.info("有功功率Passed")
                    else:
                        ws.cell(i + 2, 549 + row, f'Failed')
                        logging.info("Failed")
                        break
                for j in range(6):
                    l = j + 3
                    ws.cell(i + 2, 550 + j + row, f'{User_Power[l]}')
    wb.save('Precision_Measure_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))


if __name__ == '__main__':
    print('====================Precision Measure Start====================')
    print('======================{}======================'.format(time.strftime('%Y_%m_%d %H:%M:%S')))
    start_time = time.time()
    switch_device_screen_interface(0x01)
    time.sleep(5)
    set_gear_switching_mode('00000000')
    time.sleep(5)
    CT_type = 'RCT16_500'
    # e1_2w_fast_precision_measure(CT_type)
    # e1_2w_fast_precision_measure_tp26_33(CT_type)
    # e1_2w_fast_precision_measure_no_load()
    e3_4w_y_fast_precision_measure_new("3E4WY", CT_type, 3)
    # e2_3w1p_fast_precision_measure_new("2E3W1P", CT_type)
    # e2_3w_networkfast_precision_measure_new(CT_type)
    # e2_3w_delta_fast_precision_measure_new(CT_type)
    ModbusClient.close()
    ret = set_ac(120, 240, 0, 120, 240, 0, 0, 0, 0, 0, 0, 0, 50)
    time.sleep(5)
    switch_device_screen_interface(0x00)
    print('====================测试总耗时:{}===================='.format(time.time() - start_time))
    print('====================={}====================='.format(time.strftime('%Y_%m_%d %H:%M:%S')))
    print('====================Precision Measure End====================')
