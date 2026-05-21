import json
import os.path
import time
from pathlib import Path

import openpyxl
import pandas as pd
import xlwt
from tools.log import root_path


def dcpara_addr_get(file_path, sheet_name):
    workbook = openpyxl.load_workbook(root_path + file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []
    data_dict = {}
    i = 0
    for row in sheet.rows:
        data.append([])
        for cell in row:
            data[i].append(cell.value)
        i += 1
    for i in range(1, len(data)):
        if sheet_name == 'System Infomation':
            if data[i][data[0].index('Descrption')] is None:
                continue
            data_dict[data[i][data[0].index('Descrption')]] = {}
            data_dict[data[i][data[0].index('Descrption')]]['Start(Dec)'] = data[i][data[0].index('Start(Dec)')]
            data_dict[data[i][data[0].index('Descrption')]]['Reg'] = data[i][data[0].index('Reg')]
        elif data[i][data[0].index('Descrption')] is None or data[i][data[0].index('Data type')] is None:
            continue
        else:
            data_dict[data[i][data[0].index('Descrption')] + ' ' + data[i][data[0].index('Data type')]] = {}
            data_dict[data[i][data[0].index('Descrption')] + ' ' + data[i][data[0].index('Data type')]]['Start(Dec)'] = \
                data[i][data[0].index('Start(Dec)')]
            data_dict[data[i][data[0].index('Descrption')] + ' ' + data[i][data[0].index('Data type')]]['Reg'] = \
                data[i][data[0].index('Reg')]
    # print(data_dict)
    return data_dict


def excel_write():
    my_workbook = xlwt.Workbook()
    sheet = my_workbook.add_sheet('AcuDC300 SYS TIME')
    sheet.write(0, 0, '0')
    my_workbook.save('TestCase_StabiLity_0809_case1_2.xlsx')


def data_read(file_path, sheet_name):
    """
    读取测试用例的sheet页数据
    :param file_path: 测试用例文件路径
    :param sheet_name: sheet页名
    :return: sheet页数据
    """
    workbook = openpyxl.load_workbook(root_path + file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []
    i = 0
    for row in sheet.rows:
        data.append([])
        for cell in row:
            data[i].append(cell.value)
        i += 1
    return data


# AcuDc260项目
# 读取excel寄存器地址表，转为字典格式，或者转为json类型存储

FILE_PATH = os.path.join(Path(__file__).parent.parent.parent, "test_case", "AcuDC_320",
                         "acudc320_modbus_addrs.xlsx")
OUTPUT_JSON_PATH = os.path.join(Path(__file__).parent.parent.parent, "test_case", "AcuDC_320",
                                "acudc320_modbus_addrs.json")
SHEET_NAMES = ["Readings"]


class HandleExcel:
    def __init__(self):
        self.file_path = FILE_PATH
        self.sheet_names = SHEET_NAMES
        self.output_json_path = OUTPUT_JSON_PATH
        self.data_dict = {}
        self.init_func()

    def init_func(self):
        self.handle_memory_addr()
        self.write_excel_to_json()

    def handle_memory_addr(self):
        """
        读取excel寄存器地址表，转为字典格式
        :return: 字典
        """
        for sheet_name in self.sheet_names:
            df = pd.read_excel(self.file_path, sheet_name)
            df.fillna(0, inplace=True)
            # acudc260
            if sheet_name == "Readings":
                self.data_dict["readings"] = dict(
                    zip(df["Descrption"], df["Start(Dec)"].astype(int)))
        return self.data_dict

    def write_excel_to_json(self):
        """
        将字典类型转为json类型存储
        :return:
        """
        with open(self.output_json_path, "w", encoding="utf-8") as f:
            json.dump(self.data_dict, f, ensure_ascii=False, indent=4)


def json_to_dict(json_path):
    """
    json文件数据转成字典类型
    :param json_path: json文件路径
    :return: 字典
    """
    with open(json_path, 'r', encoding='utf-8') as f:
        data_dict = json.load(f)
    return data_dict


if __name__ == '__main__':
    memory_data_dict = json_to_dict(json_path=OUTPUT_JSON_PATH)
    # memory_data_dict = HandleExcel().data_dict
    u_rms_addr = memory_data_dict["readings"].get("Current 1", 0)
    print(memory_data_dict,u_rms_addr)
