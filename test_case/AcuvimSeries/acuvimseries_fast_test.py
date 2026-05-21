#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
文件名称:acuvimseries_fast_test.py
功能描述:快速测量
创建日期:2025-08-05
作者:王洋
版本:v1.0
修改记录:
"""

import os
import statistics
import time
import math
import logging
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from acuvimseries_modbus_get import HandleMemory
from comm.source_control import (switch_device_screen_interface, set_gear_switching_mode, set_ac, set_voltage_gear,
                                 set_current_gear, up_source_ac)
from test_case.AcuvimSeries.memory_info import MemoryAddr
from test_case.AcuvimSeries.test_report_table_heading import TableTitle
from tools.excel_operate import data_read
from tools.log import Log
from power_calculate import CalculatePower

Log(str(__file__).split("\\")[-1])

test_case_path = r'./test_case/AcuvimSeries/acuvimseries_test_case.xlsx'
sheet_name_mV = "test_case_mV"
sheet_name_mA = "test_case_mA"
sheet_name_5A = "test_case_5A"

save_filedir = os.path.join(Path(__file__).parent, f"precision_measure_{time.strftime('%Y%m%d')}")
if not os.path.exists(save_filedir):
    os.makedirs(save_filedir, exist_ok=True)


class PrecisionMeasure:
    def __init__(self):
        self.handle_memory = HandleMemory(slave_id=1)

    def select_test_type(self, test_type, wire_type):
        """
        选择测试数据类型
        :param test_type: 0:mV, 1:mA, 2:5A
        :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
        :return:
        """
        if test_type == 0:
            self.select_wire_type(test_case_path, sheet_name_mV, wire_type)
        elif test_type == 1:
            self.select_wire_type(test_case_path, sheet_name_mA, wire_type)
        else:
            self.select_wire_type(test_case_path, sheet_name_5A, wire_type)

    def set_wire_type(self, voltage_wire_value, current_wire_value):
        """
        设置接线方式
        :param voltage_wire_value: 电压接线方式
        :param current_wire_value: 电流接线方式
        :return:
        """
        self.handle_memory.set_wire_mode_by_voltage(voltage_wire_mode=voltage_wire_value)
        self.handle_memory.set_wire_mode_by_current(current_wire_mode=current_wire_value)

    def select_test_case(self, test_type, wire_type):
        """
        选择测试标准，接线方式
        :param test_type: 0:mV, 1:mA, 2:5A
        :param wire_type:0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
        :return:
        """
        self.select_test_type(test_type, wire_type)

    @staticmethod
    def get_input_list_by_wire_type(source_input_list, wire_type):
        """
        获取接线方式的数据
        :param source_input_list: 测试用例数据
        :param wire_type: 接线类型
        :return: 接线方式的数据
        """
        input_list = [source_input_list[0]]
        if wire_type == 6:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "3E4wY":
                    input_list.append(source_input_list[i])
        elif wire_type == 5:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2.5E4wY":
                    input_list.append(source_input_list[i])
        elif wire_type == 4:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "3E3wD":
                    input_list.append(source_input_list[i])
        elif wire_type == 3:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2E3wD":
                    input_list.append(source_input_list[i])
        elif wire_type == 2:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2E3wN":
                    input_list.append(source_input_list[i])
        elif wire_type == 1:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2E3w1p":
                    input_list.append(source_input_list[i])
        elif wire_type == 0:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "1E2w1p":
                    input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_real_time_parameters_by_input_list_of_wire_type(source_input_list):
        """
        获取接线方式的快速测量实时数据
        :param source_input_list: 接线方式的数据
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            if not source_input_list[i][13]:
                input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def write_table_title_of_3e4wy_to_excel(file_path, wb, ws):
        """
        写入表头:3e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_3E4WY)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_3E4WY[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2p5e4wy_to_excel(file_path, wb, ws):
        """
        写入表头:2p5e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2p5E4WY)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2p5E4WY[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_1e2w1p_to_excel(file_path, wb, ws):
        """
        写入表头:1e2w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_1E2W1P)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_1E2W1P[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3w1p_to_excel(file_path, wb, ws):
        """
        写入表头:2e3w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2E3W1P)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2E3W1P[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_3e3wd_to_excel(file_path, wb, ws):
        """
        写入表头:3e3wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_3E3WD)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_3E3WD[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wd_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2E3WD)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2E3WD[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wn_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wn
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2E3WN)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_2E3WN[i]}')
        wb.save(file_path)

    @staticmethod
    def get_memory_addr():
        """
        获取寄存器值
        :return: 寄存器值
        """
        ua_rms_addr = MemoryAddr.ua_rms_addr
        ub_rms_addr = MemoryAddr.ub_rms_addr
        uc_rms_addr = MemoryAddr.uc_rms_addr
        u_avg_rms_addr = MemoryAddr.uv_avg_rms_addr
        ia_rms_addr = MemoryAddr.ia_rms_addr
        ib_rms_addr = MemoryAddr.ib_rms_addr
        ic_rms_addr = MemoryAddr.ic_rms_addr
        i_avg_rms_addr = MemoryAddr.iv_avg_rms_addr
        pa_rms_addr = MemoryAddr.pa_rms_addr
        pb_rms_addr = MemoryAddr.pb_rms_addr
        pc_rms_addr = MemoryAddr.pc_rms_addr
        p_total_rms_addr = MemoryAddr.p_total_rms_addr
        qa_rms_addr = MemoryAddr.qa_rms_addr
        qb_rms_addr = MemoryAddr.qb_rms_addr
        qc_rms_addr = MemoryAddr.qc_rms_addr
        q_total_rms_addr = MemoryAddr.q_total_rms_addr
        sa_rms_addr = MemoryAddr.sa_rms_addr
        sb_rms_addr = MemoryAddr.sb_rms_addr
        sc_rms_addr = MemoryAddr.sc_rms_addr
        s_total_rms_addr = MemoryAddr.s_total_rms_addr
        ua_p_rms_addr = MemoryAddr.ua_phase_angle_rms_addr
        ub_p_rms_addr = MemoryAddr.ub_phase_angle_rms_addr
        uc_p_rms_addr = MemoryAddr.uc_phase_angle_rms_addr
        ia_p_rms_addr = MemoryAddr.ia_phase_angle_rms_addr
        ib_p_rms_addr = MemoryAddr.ib_phase_angle_rms_addr
        ic_p_rms_addr = MemoryAddr.ic_phase_angle_rms_addr
        uab_rms_addr = MemoryAddr.uab_rms_addr
        ubc_rms_addr = MemoryAddr.ubc_rms_addr
        uca_rms_addr = MemoryAddr.uca_rms_addr
        ul_avg_rms_addr = MemoryAddr.ul_avg_rms_addr
        pf_a_rms_addr = MemoryAddr.pf_a_rms_addr
        pf_b_rms_addr = MemoryAddr.pf_b_rms_addr
        pf_c_rms_addr = MemoryAddr.pf_c_rms_addr
        pf_total_rms_addr = MemoryAddr.pf_total_rms_addr
        freq_rms_addr = MemoryAddr.freq_rms_addr
        return (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
                ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
                pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
                qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
                sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
                ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
                ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
                uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
                pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
                freq_rms_addr)

    @staticmethod
    def get_accuracy_by_datatype(accuracy_value):
        """
        处理N/A精度值
        :param accuracy_value: 精度值
        :return: 精度值
        """
        accuracy_res = accuracy_value
        if not isinstance(accuracy_value, float):
            accuracy_res = 999999
        return accuracy_res

    def get_test_case_info_of_accuracy(self, input_list, index_value):
        """
        获取测试精度信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 精度值
        """
        accuracy_of_voltage = input_list[index_value + 1][17]
        accuracy_of_current = input_list[index_value + 1][18]
        accuracy_of_phase_angle = input_list[index_value + 1][19]
        accuracy_of_active_power = input_list[index_value + 1][20]
        accuracy_of_reactive_power = input_list[index_value + 1][21]
        accuracy_of_apparent_power = input_list[index_value + 1][22]
        voltage_accuracy = self.get_accuracy_by_datatype(accuracy_of_voltage)
        current_accuracy = self.get_accuracy_by_datatype(accuracy_of_current)
        phase_angle_accuracy = self.get_accuracy_by_datatype(accuracy_of_phase_angle)
        # voltage_phase_angle_accuracy = self.get_accuracy_by_datatype(accuracy_of_phase_angle)
        # current_phase_angle_accuracy = self.get_accuracy_by_datatype(accuracy_of_phase_angle)
        active_power_accuracy = self.get_accuracy_by_datatype(accuracy_of_active_power)
        reactive_power_accuracy = self.get_accuracy_by_datatype(accuracy_of_reactive_power)
        apparent_power_accuracy = self.get_accuracy_by_datatype(accuracy_of_apparent_power)
        return (voltage_accuracy, current_accuracy, phase_angle_accuracy,
                active_power_accuracy, reactive_power_accuracy, apparent_power_accuracy)

    @staticmethod
    def get_test_case_info_of_input_value(input_list, index_value):
        """
        获取测试电压/电流、相位等信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 电压/电流、相位等信息
        """
        case_id = input_list[index_value + 1][0]
        ua = input_list[index_value + 1][1]
        ub = input_list[index_value + 1][2]
        uc = input_list[index_value + 1][3]
        ia = input_list[index_value + 1][4]
        ib = input_list[index_value + 1][5]
        ic = input_list[index_value + 1][6]
        ua_p = input_list[index_value + 1][7]
        ub_p = input_list[index_value + 1][8]
        uc_p = input_list[index_value + 1][9]
        ia_p = input_list[index_value + 1][10]
        ib_p = input_list[index_value + 1][11]
        ic_p = input_list[index_value + 1][12]
        freq = input_list[index_value + 1][16]
        sample_cnt = input_list[index_value + 1][23]
        sample_interval = input_list[index_value + 1][24]
        return case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, freq, sample_cnt, sample_interval

    @staticmethod
    def get_power_cal_res_of_phase(value_of_voltage, value_of_current,
                                   value_of_voltage_phase_angle, value_of_current_phase_angle):
        """
        通过电压计算功率
        :param value_of_voltage: 电压
        :param value_of_current: 电流
        :param value_of_voltage_phase_angle: 电压相位角度
        :param value_of_current_phase_angle: 电流相位角度
        :return: 功率
        """

        active_power_res = CalculatePower.calculate_active_power(value_of_voltage,
                                                                 value_of_current,
                                                                 value_of_voltage_phase_angle,
                                                                 value_of_current_phase_angle)
        reactive_power_res = CalculatePower.calculate_reactive_power(value_of_voltage,
                                                                     value_of_current,
                                                                     value_of_voltage_phase_angle,
                                                                     value_of_current_phase_angle)
        apparent_power_res = CalculatePower.calculate_apparent_power(value_of_voltage, value_of_current)

        return active_power_res, reactive_power_res, apparent_power_res

    # @staticmethod
    # def get_sys_power(power_of_phase_a, power_of_phase_b, power_of_phase_c):
    #     """
    #     计算系统功率
    #     :param power_of_phase_a: A相功率
    #     :param power_of_phase_b: B相功率
    #     :param power_of_phase_c: C相功率
    #     :return: 系统功率
    #     """
    #     sys_power = sum([power_of_phase_a, power_of_phase_b, power_of_phase_c])
    #     return sys_power

    @staticmethod
    def get_sys_power(*power_of_phase):
        """
        计算系统功率
        :param power_of_phase: A/B/C相功率
        :return: 系统功率
        """
        sys_power = sum(power_of_phase)
        return sys_power

    @staticmethod
    def write_common_values_to_excel(ws, index_value, common_values, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param common_values: case编号/电压/电流/电压相位角/电流相位角等输入值,精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(common_values)):
            ws.cell(index_value + 2, start_num + k, common_values[k])
        start_num += len(common_values)
        return start_num

    @staticmethod
    def write_accuracy_res_to_excel(file_path, wb, ws, index_value, accuracy_res, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param file_path: 待写入文件路径
        :param wb: 写入工作簿对象
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param accuracy_res: 精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(accuracy_res)):
            ws.cell(index_value + 2, start_num + k, accuracy_res[k])
        wb.save(file_path)
        start_num += len(accuracy_res)
        return start_num

    # def set_and_switch_source(self,uc_value, ub_value, ua_value,ic_value, ib_value, ia_value):
    #     """
    #     升源加设置档位
    #     :return:
    #     """
    #     # 设置电压/电流档位，当前是相同值，只配置了一种A相
    #     set_voltage_gear(uc_value, ub_value, ua_value)
    #     set_current_gear(ic_value, ib_value, ia_value)

    @staticmethod
    def get_line_current_angle(phase_current_angle):
        """
        delta类型接线方式，获取线电流角度
        :param phase_current_angle: 相电流角度
        :return: 线电流角度
        """
        phase_current_angle = phase_current_angle if ((phase_current_angle - 30) >= 0) else (phase_current_angle + 360)
        return phase_current_angle

    def get_phase_angle_of_line_current(self, ia_angle, ib_angle, ic_angle):
        """
        delta类型接线方式，获取线电流角度
        :param ia_angle: ia角度
        :param ib_angle: ib角度
        :param ic_angle: ic角度
        :return: 线电流角度
        """
        ia_angle = self.get_line_current_angle(ia_angle)
        ib_angle = self.get_line_current_angle(ib_angle)
        ic_angle = self.get_line_current_angle(ic_angle)
        return ia_angle, ib_angle, ic_angle

    def get_measure_values_of_phase_voltage(self):
        """
        获取相电压测量值
        :return: 相电压测量值
        """
        ua = self.handle_memory.read_ua_voltage()
        ub = self.handle_memory.read_ub_voltage()
        uc = self.handle_memory.read_uc_voltage()
        uv_avg = self.handle_memory.read_uv_avg_voltage()
        return ua, ub, uc, uv_avg

    def get_measure_values_of_line_voltage(self):
        """
        获取线电压测量值
        :return: 线电压测量值
        """
        uab = self.handle_memory.read_uab_voltage()
        ubc = self.handle_memory.read_ubc_voltage()
        uca = self.handle_memory.read_uca_voltage()
        ul_avg = self.handle_memory.read_ul_avg_voltage()
        return uab, ubc, uca, ul_avg

    def get_measure_values_of_phase_current(self):
        """
        获取电流测量值
        :return: 电流测量值
        """
        ia = self.handle_memory.read_ia_current()
        ib = self.handle_memory.read_ib_current()
        ic = self.handle_memory.read_ic_current()
        iv_avg = self.handle_memory.read_iv_avg_current()
        return ia, ib, ic, iv_avg

    def get_measure_values_of_in_current(self):
        """
        获取IN电流测量值
        :return: IN电流测量值
        """
        current_of_in = self.handle_memory.read_in_current()
        return current_of_in

    def get_measure_values_of_sys_power(self):
        """
        获取系统功率测量值
        :return: 系统功率测量值
        """
        p_total = self.handle_memory.read_p_total_power()
        q_total = self.handle_memory.read_q_total_power()
        s_total = self.handle_memory.read_s_total_power()
        return p_total, q_total, s_total

    def get_measure_values_of_active_power(self):
        """
        获取有功功率测量值
        :return: 有功功率测量值
        """
        pa = self.handle_memory.read_pa_power()
        pb = self.handle_memory.read_pb_power()
        pc = self.handle_memory.read_pc_power()
        p_total = self.handle_memory.read_p_total_power()
        return pa, pb, pc, p_total

    def get_measure_values_of_reactive_power(self):
        """
        获取无功功率测量值
        :return: 无功功率测量值
        """
        qa = self.handle_memory.read_qa_power()
        qb = self.handle_memory.read_qb_power()
        qc = self.handle_memory.read_qc_power()
        q_total = self.handle_memory.read_q_total_power()
        return qa, qb, qc, q_total

    def get_measure_values_of_apparent_power(self):
        """
        获取视在功率测量值
        :return: 视在功率测量值
        """
        sa = self.handle_memory.read_sa_power()
        sb = self.handle_memory.read_sb_power()
        sc = self.handle_memory.read_sc_power()
        s_total = self.handle_memory.read_s_total_power()
        return sa, sb, sc, s_total

    def get_measure_values_of_voltage_phase_angle(self):
        """
        获取电压相位角度测量值
        :return: 电压相位角度测量值
        """
        ua_phase_angle = self.handle_memory.read_ua_phase_angle()
        ub_phase_angle = self.handle_memory.read_ub_phase_angle()
        uc_phase_angle = self.handle_memory.read_uc_phase_angle()
        return ua_phase_angle, ub_phase_angle, uc_phase_angle

    def get_measure_values_of_current_phase_angle(self):
        """
        获取电流相位角度测量值
        :return: 获取电流相位角度测量值
        """
        ia_phase_angle = self.handle_memory.read_ia_phase_angle()
        ib_phase_angle = self.handle_memory.read_ib_phase_angle()
        ic_phase_angle = self.handle_memory.read_ic_phase_angle()
        return ia_phase_angle, ib_phase_angle, ic_phase_angle

    def get_measure_values_by_sample_cnt_of_3e4wy(self, sample_cnt, sample_interval):
        """
        获取测量值:3e4wy
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ua = []
        measure_values_of_ub = []
        measure_values_of_uc = []
        measure_values_of_ia = []
        measure_values_of_ib = []
        measure_values_of_ic = []
        measure_values_of_ua_angle = []
        measure_values_of_ub_angle = []
        measure_values_of_uc_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_ib_angle = []
        measure_values_of_ic_angle = []
        measure_values_of_pa = []
        measure_values_of_pb = []
        measure_values_of_pc = []
        measure_values_of_p_sys = []
        measure_values_of_qa = []
        measure_values_of_qb = []
        measure_values_of_qc = []
        measure_values_of_q_sys = []
        measure_values_of_sa = []
        measure_values_of_sb = []
        measure_values_of_sc = []
        measure_values_of_s_sys = []
        measure_values_of_uab = []
        measure_values_of_ubc = []
        measure_values_of_uca = []
        measure_values_of_in = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ua_measure, ub_measure, uc_measure, _ = self.get_measure_values_of_phase_voltage()
            ia_measure, ib_measure, ic_measure, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, ub_angle_measure, uc_angle_measure = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, ib_angle_measure, ic_angle_measure = self.get_measure_values_of_current_phase_angle()
            pa_measure, pb_measure, pc_measure, p_sys_measure = self.get_measure_values_of_active_power()
            qa_measure, qb_measure, qc_measure, q_sys_measure = self.get_measure_values_of_reactive_power()
            sa_measure, sb_measure, sc_measure, s_sys_measure = self.get_measure_values_of_apparent_power()
            uab_measure, ubc_measure, uca_measure, _ = self.get_measure_values_of_line_voltage()
            measure_of_in = self.get_measure_values_of_in_current()
            measure_values_of_ua.append(ua_measure)
            measure_values_of_ub.append(ub_measure)
            measure_values_of_uc.append(uc_measure)
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ib.append(ib_measure)
            measure_values_of_ic.append(ic_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ub_angle.append(ub_angle_measure)
            measure_values_of_uc_angle.append(uc_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_ib_angle.append(ib_angle_measure)
            measure_values_of_ic_angle.append(ic_angle_measure)
            measure_values_of_pa.append(pa_measure)
            measure_values_of_pb.append(pb_measure)
            measure_values_of_pc.append(pc_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_qa.append(qa_measure)
            measure_values_of_qb.append(qb_measure)
            measure_values_of_qc.append(qc_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_sa.append(sa_measure)
            measure_values_of_sb.append(sb_measure)
            measure_values_of_sc.append(sc_measure)
            measure_values_of_s_sys.append(s_sys_measure)
            measure_values_of_uab.append(uab_measure)
            measure_values_of_ubc.append(ubc_measure)
            measure_values_of_uca.append(uca_measure)
            measure_values_of_in.append(measure_of_in)
        return (
            measure_values_of_ua,
            measure_values_of_ub,
            measure_values_of_uc,
            measure_values_of_ia,
            measure_values_of_ib,
            measure_values_of_ic,
            measure_values_of_ua_angle,
            measure_values_of_ub_angle,
            measure_values_of_uc_angle,
            measure_values_of_ia_angle,
            measure_values_of_ib_angle,
            measure_values_of_ic_angle,
            measure_values_of_pa,
            measure_values_of_pb,
            measure_values_of_pc,
            measure_values_of_p_sys,
            measure_values_of_qa,
            measure_values_of_qb,
            measure_values_of_qc,
            measure_values_of_q_sys,
            measure_values_of_sa,
            measure_values_of_sb,
            measure_values_of_sc,
            measure_values_of_s_sys,
            measure_values_of_uab,
            measure_values_of_ubc,
            measure_values_of_uca,
            measure_values_of_in
        )

    def get_measure_values_by_sample_cnt_of_2e3w1p(self, sample_cnt, sample_interval):
        """
        获取测量值:2e3w1p
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ua = []
        measure_values_of_ub = []
        measure_values_of_ia = []
        measure_values_of_ib = []
        measure_values_of_ua_angle = []
        measure_values_of_ub_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_ib_angle = []
        measure_values_of_pa = []
        measure_values_of_pb = []
        measure_values_of_p_sys = []
        measure_values_of_qa = []
        measure_values_of_qb = []
        measure_values_of_q_sys = []
        measure_values_of_sa = []
        measure_values_of_sb = []
        measure_values_of_s_sys = []
        measure_values_of_uab = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ua_measure, ub_measure, _, _ = self.get_measure_values_of_phase_voltage()
            ia_measure, ib_measure, _, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, ub_angle_measure, _ = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, ib_angle_measure, _ = self.get_measure_values_of_current_phase_angle()
            pa_measure, pb_measure, _, p_sys_measure = self.get_measure_values_of_active_power()
            qa_measure, qb_measure, _, q_sys_measure = self.get_measure_values_of_reactive_power()
            sa_measure, sb_measure, _, s_sys_measure = self.get_measure_values_of_apparent_power()
            uab_measure, _, _, _ = self.get_measure_values_of_line_voltage()
            measure_values_of_ua.append(ua_measure)
            measure_values_of_ub.append(ub_measure)
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ib.append(ib_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ub_angle.append(ub_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_ib_angle.append(ib_angle_measure)
            measure_values_of_pa.append(pa_measure)
            measure_values_of_pb.append(pb_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_qa.append(qa_measure)
            measure_values_of_qb.append(qb_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_sa.append(sa_measure)
            measure_values_of_sb.append(sb_measure)
            measure_values_of_s_sys.append(s_sys_measure)
            measure_values_of_uab.append(uab_measure)
        return (
            measure_values_of_ua,
            measure_values_of_ub,
            measure_values_of_ia,
            measure_values_of_ib,
            measure_values_of_ua_angle,
            measure_values_of_ub_angle,
            measure_values_of_ia_angle,
            measure_values_of_ib_angle,
            measure_values_of_pa,
            measure_values_of_pb,
            measure_values_of_p_sys,
            measure_values_of_qa,
            measure_values_of_qb,
            measure_values_of_q_sys,
            measure_values_of_sa,
            measure_values_of_sb,
            measure_values_of_s_sys,
            measure_values_of_uab,
        )

    def get_measure_values_by_sample_cnt_of_1e2w1p(self, sample_cnt, sample_interval):
        """
        获取测量值:1e2w1p
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ua = []
        measure_values_of_ia = []
        measure_values_of_ua_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_pa = []
        measure_values_of_p_sys = []
        measure_values_of_qa = []
        measure_values_of_q_sys = []
        measure_values_of_sa = []
        measure_values_of_s_sys = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ua_measure, _, _, _ = self.get_measure_values_of_phase_voltage()
            ia_measure, _, _, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, _, _ = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, _, _ = self.get_measure_values_of_current_phase_angle()
            pa_measure, _, _, p_sys_measure = self.get_measure_values_of_active_power()
            qa_measure, _, _, q_sys_measure = self.get_measure_values_of_reactive_power()
            sa_measure, _, _, s_sys_measure = self.get_measure_values_of_apparent_power()
            measure_values_of_ua.append(ua_measure)
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_pa.append(pa_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_qa.append(qa_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_sa.append(sa_measure)
            measure_values_of_s_sys.append(s_sys_measure)
        return (
            measure_values_of_ua,
            measure_values_of_ia,
            measure_values_of_ua_angle,
            measure_values_of_ia_angle,
            measure_values_of_pa,
            measure_values_of_p_sys,
            measure_values_of_qa,
            measure_values_of_q_sys,
            measure_values_of_sa,
            measure_values_of_s_sys,
        )

    def get_measure_values_by_sample_cnt_of_2p5e4wy(self, sample_cnt, sample_interval):
        """
        获取测量值:2p5e4wy
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ua = []
        measure_values_of_ub = []
        measure_values_of_uc = []
        measure_values_of_ia = []
        measure_values_of_ib = []
        measure_values_of_ic = []
        measure_values_of_ua_angle = []
        measure_values_of_ub_angle = []
        measure_values_of_uc_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_ib_angle = []
        measure_values_of_ic_angle = []
        measure_values_of_pa = []
        measure_values_of_pb = []
        measure_values_of_pc = []
        measure_values_of_p_sys = []
        measure_values_of_qa = []
        measure_values_of_qb = []
        measure_values_of_qc = []
        measure_values_of_q_sys = []
        measure_values_of_sa = []
        measure_values_of_sb = []
        measure_values_of_sc = []
        measure_values_of_s_sys = []
        measure_values_of_uab = []
        measure_values_of_ubc = []
        measure_values_of_uca = []
        measure_values_of_in = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ua_measure, ub_measure, uc_measure, _ = self.get_measure_values_of_phase_voltage()
            ia_measure, ib_measure, ic_measure, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, ub_angle_measure, uc_angle_measure = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, ib_angle_measure, ic_angle_measure = self.get_measure_values_of_current_phase_angle()
            pa_measure, pb_measure, pc_measure, p_sys_measure = self.get_measure_values_of_active_power()
            qa_measure, qb_measure, qc_measure, q_sys_measure = self.get_measure_values_of_reactive_power()
            sa_measure, sb_measure, sc_measure, s_sys_measure = self.get_measure_values_of_apparent_power()
            uab_measure, ubc_measure, uca_measure, _ = self.get_measure_values_of_line_voltage()
            measure_of_in = self.get_measure_values_of_in_current()
            measure_values_of_ua.append(ua_measure)
            measure_values_of_ub.append(ub_measure)
            measure_values_of_uc.append(uc_measure)
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ib.append(ib_measure)
            measure_values_of_ic.append(ic_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ub_angle.append(ub_angle_measure)
            measure_values_of_uc_angle.append(uc_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_ib_angle.append(ib_angle_measure)
            measure_values_of_ic_angle.append(ic_angle_measure)
            measure_values_of_pa.append(pa_measure)
            measure_values_of_pb.append(pb_measure)
            measure_values_of_pc.append(pc_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_qa.append(qa_measure)
            measure_values_of_qb.append(qb_measure)
            measure_values_of_qc.append(qc_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_sa.append(sa_measure)
            measure_values_of_sb.append(sb_measure)
            measure_values_of_sc.append(sc_measure)
            measure_values_of_s_sys.append(s_sys_measure)
            measure_values_of_uab.append(uab_measure)
            measure_values_of_ubc.append(ubc_measure)
            measure_values_of_uca.append(uca_measure)
            measure_values_of_in.append(measure_of_in)
        return (
            measure_values_of_ua,
            measure_values_of_ub,
            measure_values_of_uc,
            measure_values_of_ia,
            measure_values_of_ib,
            measure_values_of_ic,
            measure_values_of_ua_angle,
            measure_values_of_ub_angle,
            measure_values_of_uc_angle,
            measure_values_of_ia_angle,
            measure_values_of_ib_angle,
            measure_values_of_ic_angle,
            measure_values_of_pa,
            measure_values_of_pb,
            measure_values_of_pc,
            measure_values_of_p_sys,
            measure_values_of_qa,
            measure_values_of_qb,
            measure_values_of_qc,
            measure_values_of_q_sys,
            measure_values_of_sa,
            measure_values_of_sb,
            measure_values_of_sc,
            measure_values_of_s_sys,
            measure_values_of_uab,
            measure_values_of_ubc,
            measure_values_of_uca,
        )

    def get_measure_values_by_sample_cnt_of_3e3wd(self, sample_cnt, sample_interval):
        """
        获取测量值:3e3wd
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ia = []
        measure_values_of_ib = []
        measure_values_of_ic = []
        measure_values_of_ua_angle = []
        measure_values_of_ub_angle = []
        measure_values_of_uc_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_ib_angle = []
        measure_values_of_ic_angle = []
        measure_values_of_p_sys = []
        measure_values_of_q_sys = []
        measure_values_of_s_sys = []
        measure_values_of_uab = []
        measure_values_of_ubc = []
        measure_values_of_uca = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ia_measure, ib_measure, ic_measure, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, ub_angle_measure, uc_angle_measure = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, ib_angle_measure, ic_angle_measure = self.get_measure_values_of_current_phase_angle()
            p_sys_measure, q_sys_measure, s_sys_measure = self.get_measure_values_of_sys_power()
            uab_measure, ubc_measure, uca_measure, _ = self.get_measure_values_of_line_voltage()
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ib.append(ib_measure)
            measure_values_of_ic.append(ic_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ub_angle.append(ub_angle_measure)
            measure_values_of_uc_angle.append(uc_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_ib_angle.append(ib_angle_measure)
            measure_values_of_ic_angle.append(ic_angle_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_s_sys.append(s_sys_measure)
            measure_values_of_uab.append(uab_measure)
            measure_values_of_ubc.append(ubc_measure)
            measure_values_of_uca.append(uca_measure)
        return (
            measure_values_of_ia,
            measure_values_of_ib,
            measure_values_of_ic,
            measure_values_of_ua_angle,
            measure_values_of_ub_angle,
            measure_values_of_uc_angle,
            measure_values_of_ia_angle,
            measure_values_of_ib_angle,
            measure_values_of_ic_angle,
            measure_values_of_p_sys,
            measure_values_of_q_sys,
            measure_values_of_s_sys,
            measure_values_of_uab,
            measure_values_of_ubc,
            measure_values_of_uca,
        )

    def get_measure_values_by_sample_cnt_of_2e3wd(self, sample_cnt, sample_interval):
        """
        获取测量值:2e3wd
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ia = []
        measure_values_of_ib = []
        measure_values_of_ic = []
        measure_values_of_ua_angle = []
        measure_values_of_ub_angle = []
        measure_values_of_uc_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_ib_angle = []
        measure_values_of_ic_angle = []
        measure_values_of_p_sys = []
        measure_values_of_q_sys = []
        measure_values_of_s_sys = []
        measure_values_of_uab = []
        measure_values_of_ubc = []
        measure_values_of_uca = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ia_measure, ib_measure, ic_measure, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, ub_angle_measure, uc_angle_measure = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, ib_angle_measure, ic_angle_measure = self.get_measure_values_of_current_phase_angle()
            p_sys_measure, q_sys_measure, s_sys_measure = self.get_measure_values_of_sys_power()
            uab_measure, ubc_measure, uca_measure, _ = self.get_measure_values_of_line_voltage()
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ib.append(ib_measure)
            measure_values_of_ic.append(ic_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ub_angle.append(ub_angle_measure)
            measure_values_of_uc_angle.append(uc_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_ib_angle.append(ib_angle_measure)
            measure_values_of_ic_angle.append(ic_angle_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_s_sys.append(s_sys_measure)
            measure_values_of_uab.append(uab_measure)
            measure_values_of_ubc.append(ubc_measure)
            measure_values_of_uca.append(uca_measure)
        return (
            measure_values_of_ia,
            measure_values_of_ib,
            measure_values_of_ic,
            measure_values_of_ua_angle,
            measure_values_of_ub_angle,
            measure_values_of_uc_angle,
            measure_values_of_ia_angle,
            measure_values_of_ib_angle,
            measure_values_of_ic_angle,
            measure_values_of_p_sys,
            measure_values_of_q_sys,
            measure_values_of_s_sys,
            measure_values_of_uab,
            measure_values_of_ubc,
            measure_values_of_uca,
        )

    def get_measure_values_by_sample_cnt_of_2e3wn(self, sample_cnt, sample_interval):
        """
        获取测量值:2e3wn
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_ia = []
        measure_values_of_ib = []
        measure_values_of_ic = []
        measure_values_of_ua_angle = []
        measure_values_of_ub_angle = []
        measure_values_of_uc_angle = []
        measure_values_of_ia_angle = []
        measure_values_of_ib_angle = []
        measure_values_of_ic_angle = []
        measure_values_of_p_sys = []
        measure_values_of_q_sys = []
        measure_values_of_s_sys = []
        measure_values_of_uab = []
        measure_values_of_ubc = []
        measure_values_of_uca = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            ia_measure, ib_measure, ic_measure, _ = self.get_measure_values_of_phase_current()
            ua_angle_measure, ub_angle_measure, uc_angle_measure = self.get_measure_values_of_voltage_phase_angle()
            ia_angle_measure, ib_angle_measure, ic_angle_measure = self.get_measure_values_of_current_phase_angle()
            p_sys_measure, q_sys_measure, s_sys_measure = self.get_measure_values_of_sys_power()
            uab_measure, ubc_measure, uca_measure, _ = self.get_measure_values_of_line_voltage()
            measure_values_of_ia.append(ia_measure)
            measure_values_of_ib.append(ib_measure)
            measure_values_of_ic.append(ic_measure)
            measure_values_of_ua_angle.append(ua_angle_measure)
            measure_values_of_ub_angle.append(ub_angle_measure)
            measure_values_of_uc_angle.append(uc_angle_measure)
            measure_values_of_ia_angle.append(ia_angle_measure)
            measure_values_of_ib_angle.append(ib_angle_measure)
            measure_values_of_ic_angle.append(ic_angle_measure)
            measure_values_of_p_sys.append(p_sys_measure)
            measure_values_of_q_sys.append(q_sys_measure)
            measure_values_of_s_sys.append(s_sys_measure)
            measure_values_of_uab.append(uab_measure)
            measure_values_of_ubc.append(ubc_measure)
            measure_values_of_uca.append(uca_measure)
        return (
            measure_values_of_ia,
            measure_values_of_ib,
            measure_values_of_ic,
            measure_values_of_ua_angle,
            measure_values_of_ub_angle,
            measure_values_of_uc_angle,
            measure_values_of_ia_angle,
            measure_values_of_ib_angle,
            measure_values_of_ic_angle,
            measure_values_of_p_sys,
            measure_values_of_q_sys,
            measure_values_of_s_sys,
            measure_values_of_uab,
            measure_values_of_ubc,
            measure_values_of_uca,
        )

    @staticmethod
    def get_cmp_accuracy_res(act_accuracy_min, act_accuracy_max, act_accuracy_avg, exp_accuracy):
        """
        获取精度比较结果
        :param act_accuracy_min: 精度最小值
        :param act_accuracy_max: 精度最大值
        :param act_accuracy_avg: 精度平均值
        :param exp_accuracy: 精度期望值
        :return: "Passed"/"Failed"
        """
        cmp_res = "Failed"
        if (act_accuracy_min <= exp_accuracy) and (act_accuracy_max <= exp_accuracy) and (
                act_accuracy_avg <= exp_accuracy):
            cmp_res = "Passed"
        return cmp_res

    @staticmethod
    def get_all_accuracy_values(all_accuracy_values):
        """
        获取最终精度比较结果
        :param all_accuracy_values: 全部精度值列表
        :return: 最终精度比较结果
        """
        all_cmp_res = "Passed"
        if "Failed" in all_accuracy_values:
            all_cmp_res = "Failed"
        return all_cmp_res

    def get_accuracy_res_by_exp_accuracy(self, standard_value, measure_values, exp_accuracy):
        """
        获取精度比较结果，带期望精度值
        :param standard_value: 输入值
        :param measure_values: 测量值
        :param exp_accuracy: 期望精度值
        :return: 测量值和精度比较结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_voltage_current_power(standard_value, measure_values)

        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        cmp_accuracy_res = self.get_cmp_accuracy_res(act_accuracy_min=min_accuracy, act_accuracy_max=max_accuracy,
                                                     act_accuracy_avg=avg_accuracy, exp_accuracy=exp_accuracy)
        accuracy_res = (min_value, min_accuracy, max_value, max_accuracy, avg_value, avg_accuracy, cmp_accuracy_res)
        return accuracy_res

    def get_accuracy_res_by_not_exp_accuracy(self, standard_value, measure_values):
        """
        获取精度比较结果，不带期望精度值
        :param standard_value: 输入值
        :param measure_values: 测量值
        :return: 测量值结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_voltage_current_power(standard_value, measure_values)

        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        accuracy_res = (min_value, min_accuracy, max_value, max_accuracy, avg_value, avg_accuracy)
        return accuracy_res

    def get_accuracy_res_by_phase_angle(self, standard_value, measure_values, exp_accuracy):
        """
        获取相位角度精度比较结果
        :param standard_value: 输入值
        :param measure_values: 测量值
        :param exp_accuracy: 期望精度值
        :return: 测量值和精度比较结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_phase_angle(standard_value, measure_values)
        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        cmp_accuracy_res = self.get_cmp_accuracy_res(act_accuracy_min=min_accuracy, act_accuracy_max=max_accuracy,
                                                     act_accuracy_avg=avg_accuracy, exp_accuracy=exp_accuracy)
        accuracy_res = (min_value, min_accuracy, max_value, max_accuracy, avg_value, avg_accuracy, cmp_accuracy_res)
        return accuracy_res

    @staticmethod
    def line_to_line_voltage_calculate(ua, ub, uc, va_angle, vb_angle, vc_angle):
        """
        计算线电压
        :param ua:ua
        :param ub:ub
        :param uc:uc
        :param va_angle:va_angle
        :param vb_angle:vb_angle
        :param vc_angle:vc_angle
        :return: vab, vbc, vca
        """
        ret = []
        va_complex = complex(math.cos(va_angle * math.pi / 180) * ua, math.sin(va_angle * math.pi / 180) * ua)
        vb_complex = complex(math.cos(vb_angle * math.pi / 180) * ub, math.sin(vb_angle * math.pi / 180) * ub)
        vc_complex = complex(math.cos(vc_angle * math.pi / 180) * uc, math.sin(vc_angle * math.pi / 180) * uc)
        vab = va_complex - vb_complex
        vbc = vb_complex - vc_complex
        vca = vc_complex - va_complex
        vab = abs(vab)
        vbc = abs(vbc)
        vca = abs(vca)
        ret.extend([vab, vbc, vca])
        return ret

    def fast_precision_measure_by_3e4wy(self, file_path, input_list):
        """
        接线方式:3e4wy
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            (pb, qb, sb) = self.get_power_cal_res_of_phase(ub, ib, ub_p, ib_p)
            (pc, qc, sc) = self.get_power_cal_res_of_phase(uc, ic, uc_p, ic_p)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            p_sys = self.get_sys_power(pa, pb, pc)
            q_sys = self.get_sys_power(qa, qb, qc)
            s_sys = self.get_sys_power(sa, sb, sc)
            value_of_in = ic
            start_num = 1
            common_values_of_3e4wy = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wy, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            (
                measure_values_of_ua,
                measure_values_of_ub,
                measure_values_of_uc,
                measure_values_of_ia,
                measure_values_of_ib,
                measure_values_of_ic,
                measure_values_of_ua_angle,
                measure_values_of_ub_angle,
                measure_values_of_uc_angle,
                measure_values_of_ia_angle,
                measure_values_of_ib_angle,
                measure_values_of_ic_angle,
                measure_values_of_pa,
                measure_values_of_pb,
                measure_values_of_pc,
                measure_values_of_p_sys,
                measure_values_of_qa,
                measure_values_of_qb,
                measure_values_of_qc,
                measure_values_of_q_sys,
                measure_values_of_sa,
                measure_values_of_sb,
                measure_values_of_sc,
                measure_values_of_s_sys,
                measure_values_of_uab,
                measure_values_of_ubc,
                measure_values_of_uca,
                measure_values_of_in
            ) = self.get_measure_values_by_sample_cnt_of_3e4wy(sample_cnt, sample_interval)
            ua_accuracy = self.get_accuracy_res_by_exp_accuracy(ua, measure_values_of_ua, voltage_accuracy)
            ub_accuracy = self.get_accuracy_res_by_exp_accuracy(ub, measure_values_of_ub, voltage_accuracy)
            uc_accuracy = self.get_accuracy_res_by_exp_accuracy(uc, measure_values_of_uc, voltage_accuracy)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ib_accuracy = self.get_accuracy_res_by_exp_accuracy(ib, measure_values_of_ib, current_accuracy)
            ic_accuracy = self.get_accuracy_res_by_exp_accuracy(ic, measure_values_of_ic, current_accuracy)

            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ub_angle_accuracy = self.get_accuracy_res_by_phase_angle(ub_p, measure_values_of_ub_angle,
                                                                     phase_angle_accuracy)
            uc_angle_accuracy = self.get_accuracy_res_by_phase_angle(uc_p, measure_values_of_uc_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            ib_angle_accuracy = self.get_accuracy_res_by_phase_angle(ib_p, measure_values_of_ib_angle,
                                                                     phase_angle_accuracy)
            ic_angle_accuracy = self.get_accuracy_res_by_phase_angle(ic_p, measure_values_of_ic_angle,
                                                                     phase_angle_accuracy)
            pa_accuracy = self.get_accuracy_res_by_exp_accuracy(pa, measure_values_of_pa, active_power_accuracy)
            pb_accuracy = self.get_accuracy_res_by_exp_accuracy(pb, measure_values_of_pb, active_power_accuracy)
            pc_accuracy = self.get_accuracy_res_by_exp_accuracy(pc, measure_values_of_pc, active_power_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            qa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qa, measure_values_of_qa)
            qb_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qb, measure_values_of_qb)
            qc_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qc, measure_values_of_qc)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            sa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sa, measure_values_of_sa)
            sb_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sb, measure_values_of_sb)
            sc_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sc, measure_values_of_sc)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)
            uab_accuracy = self.get_accuracy_res_by_exp_accuracy(uab, measure_values_of_uab, voltage_accuracy)
            ubc_accuracy = self.get_accuracy_res_by_exp_accuracy(ubc, measure_values_of_ubc, voltage_accuracy)
            uca_accuracy = self.get_accuracy_res_by_exp_accuracy(uca, measure_values_of_uca, voltage_accuracy)
            accuracy_of_in = self.get_accuracy_res_by_exp_accuracy(value_of_in, measure_values_of_in, current_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uab], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uab_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [ubc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ubc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uca], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uca_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [value_of_in], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, accuracy_of_in, start_num)

    def fast_precision_measure_by_2e3w1p(self, file_path, input_list):
        """
        接线方式:2e3w1p
        ①ua/ub/uab/ia/ib/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②uc/ubc/uca/ic/pc/qc/sc/,不处理，上位机显示N/A
        ③相位角度：ua_p和ub_p相差180°,ia_p和ib_p相差180°
        ④处理uc_p=0, ic_p=0
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2e3w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            # 处理uc_p/ic_p
            uc_p = 0
            ic_p = 0
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            (pb, qb, sb) = self.get_power_cal_res_of_phase(ub, ib, ub_p, ib_p)
            uab, _, _ = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            p_sys = self.get_sys_power(pa, pb)
            q_sys = self.get_sys_power(qa, qb)
            s_sys = self.get_sys_power(sa, sb)
            start_num = 1
            common_values_of_2e3w1p = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3w1p, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            (
                measure_values_of_ua,
                measure_values_of_ub,
                measure_values_of_ia,
                measure_values_of_ib,
                measure_values_of_ua_angle,
                measure_values_of_ub_angle,
                measure_values_of_ia_angle,
                measure_values_of_ib_angle,
                measure_values_of_pa,
                measure_values_of_pb,
                measure_values_of_p_sys,
                measure_values_of_qa,
                measure_values_of_qb,
                measure_values_of_q_sys,
                measure_values_of_sa,
                measure_values_of_sb,
                measure_values_of_s_sys,
                measure_values_of_uab
            ) = self.get_measure_values_by_sample_cnt_of_2e3w1p(sample_cnt, sample_interval)
            ua_accuracy = self.get_accuracy_res_by_exp_accuracy(ua, measure_values_of_ua, voltage_accuracy)
            ub_accuracy = self.get_accuracy_res_by_exp_accuracy(ub, measure_values_of_ub, voltage_accuracy)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ib_accuracy = self.get_accuracy_res_by_exp_accuracy(ib, measure_values_of_ib, current_accuracy)
            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ub_angle_accuracy = self.get_accuracy_res_by_phase_angle(ub_p, measure_values_of_ub_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            ib_angle_accuracy = self.get_accuracy_res_by_phase_angle(ib_p, measure_values_of_ib_angle,
                                                                     phase_angle_accuracy)
            pa_accuracy = self.get_accuracy_res_by_exp_accuracy(pa, measure_values_of_pa, active_power_accuracy)
            pb_accuracy = self.get_accuracy_res_by_exp_accuracy(pb, measure_values_of_pb, active_power_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            qa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qa, measure_values_of_qa)
            qb_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qb, measure_values_of_qb)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            sa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sa, measure_values_of_sa)
            sb_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sb, measure_values_of_sb)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)
            uab_accuracy = self.get_accuracy_res_by_exp_accuracy(uab, measure_values_of_uab, voltage_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uab], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, uab_accuracy, start_num)

    def fast_precision_measure_by_1e2w1p(self, file_path, input_list):
        """
        接线方式:1e2w1p
        ①ua/ia/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ub/uc/uab/ubc/uca/ib/ic/pb/pc/qb/qc/sb/sc,不处理，上位机显示N/A
        ③处理ub_p=0, ib_p =0, uc_p=0, ic_p=0
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_1e2w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            # 处理ub_p/ib_p/uc_p/ic_p
            ub_p = 0
            uc_p = 0
            ib_p = 0
            ic_p = 0
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            p_sys = self.get_sys_power(pa)
            q_sys = self.get_sys_power(qa)
            s_sys = self.get_sys_power(sa)
            start_num = 1
            common_values_of_1e2w1p = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_1e2w1p, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            (
                measure_values_of_ua,
                measure_values_of_ia,
                measure_values_of_ua_angle,
                measure_values_of_ia_angle,
                measure_values_of_pa,
                measure_values_of_p_sys,
                measure_values_of_qa,
                measure_values_of_q_sys,
                measure_values_of_sa,
                measure_values_of_s_sys,
            ) = self.get_measure_values_by_sample_cnt_of_1e2w1p(sample_cnt, sample_interval)
            ua_accuracy = self.get_accuracy_res_by_exp_accuracy(ua, measure_values_of_ua, voltage_accuracy)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            pa_accuracy = self.get_accuracy_res_by_exp_accuracy(pa, measure_values_of_pa, active_power_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            qa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qa, measure_values_of_qa)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            sa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sa, measure_values_of_sa)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

    def fast_precision_measure_by_3e3wd(self, file_path, input_list):
        """
        接线方式:3e3wd
        ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ua,ub,uc不处理，上位机显示N/A
        ③线电流滞后输入相电流30°
        ④功率计算：s_sys_power = √3￣ * ill_avg * ull_avg
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_3e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            # 关注点1，系统功率计算
            ull_avg = round(statistics.median([uab, ubc, uca]), 5)
            ill_avg = round(statistics.median([ia, ib, ic]), 5)
            s_sys = math.sqrt(3) * ull_avg * ill_avg
            p_sys = s_sys * (math.cos(math.radians(ua_p - ia_p)))
            q_sys = s_sys * (math.sin(math.radians(ua_p - ia_p)))

            start_num = 1
            common_values_of_3e3wd = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e3wd, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            (
                measure_values_of_ia,
                measure_values_of_ib,
                measure_values_of_ic,
                measure_values_of_ua_angle,
                measure_values_of_ub_angle,
                measure_values_of_uc_angle,
                measure_values_of_ia_angle,
                measure_values_of_ib_angle,
                measure_values_of_ic_angle,
                measure_values_of_p_sys,
                measure_values_of_q_sys,
                measure_values_of_s_sys,
                measure_values_of_uab,
                measure_values_of_ubc,
                measure_values_of_uca
            ) = self.get_measure_values_by_sample_cnt_of_3e3wd(sample_cnt, sample_interval)
            # 关注点2，线电流角度
            ia_p, ib_p, ic_p = self.get_phase_angle_of_line_current(ia_p, ib_p, ic_p)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ib_accuracy = self.get_accuracy_res_by_exp_accuracy(ib, measure_values_of_ib, current_accuracy)
            ic_accuracy = self.get_accuracy_res_by_exp_accuracy(ic, measure_values_of_ic, current_accuracy)
            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ub_angle_accuracy = self.get_accuracy_res_by_phase_angle(ub_p, measure_values_of_ub_angle,
                                                                     phase_angle_accuracy)
            uc_angle_accuracy = self.get_accuracy_res_by_phase_angle(uc_p, measure_values_of_uc_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            ib_angle_accuracy = self.get_accuracy_res_by_phase_angle(ib_p, measure_values_of_ib_angle,
                                                                     phase_angle_accuracy)
            ic_angle_accuracy = self.get_accuracy_res_by_phase_angle(ic_p, measure_values_of_ic_angle,
                                                                     phase_angle_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)
            uab_accuracy = self.get_accuracy_res_by_exp_accuracy(uab, measure_values_of_uab, voltage_accuracy)
            ubc_accuracy = self.get_accuracy_res_by_exp_accuracy(ubc, measure_values_of_ubc, voltage_accuracy)
            uca_accuracy = self.get_accuracy_res_by_exp_accuracy(uca, measure_values_of_uca, voltage_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uab], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uab_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [ubc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ubc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uca], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, uca_accuracy, start_num)

    def fast_precision_measure_by_2e3wd(self, file_path, input_list):
        """
        接线方式:2e3wd
        ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ua,ub,uc不处理，上位机显示N/A
        ③线电流滞后输入相电流30°
        ④功率计算：s_sys_power = (uab * ia + ubc * ib) * cos(30°)
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            # 关注点1, 系统功率计算
            p_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.cos(math.radians(ua_p - ia_p))
            q_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.sin(math.radians(ua_p - ia_p))
            s_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30))

            start_num = 1
            common_values_of_2e3wd = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wd, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            (
                measure_values_of_ia,
                measure_values_of_ib,
                measure_values_of_ic,
                measure_values_of_ua_angle,
                measure_values_of_ub_angle,
                measure_values_of_uc_angle,
                measure_values_of_ia_angle,
                measure_values_of_ib_angle,
                measure_values_of_ic_angle,
                measure_values_of_p_sys,
                measure_values_of_q_sys,
                measure_values_of_s_sys,
                measure_values_of_uab,
                measure_values_of_ubc,
                measure_values_of_uca
            ) = self.get_measure_values_by_sample_cnt_of_2e3wd(sample_cnt, sample_interval)
            # 关注点2, 线电流角度
            ia_p, ib_p, ic_p = self.get_phase_angle_of_line_current(ia_p, ib_p, ic_p)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ib_accuracy = self.get_accuracy_res_by_exp_accuracy(ib, measure_values_of_ib, current_accuracy)
            ic_accuracy = self.get_accuracy_res_by_exp_accuracy(ic, measure_values_of_ic, current_accuracy)
            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ub_angle_accuracy = self.get_accuracy_res_by_phase_angle(ub_p, measure_values_of_ub_angle,
                                                                     phase_angle_accuracy)
            uc_angle_accuracy = self.get_accuracy_res_by_phase_angle(uc_p, measure_values_of_uc_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            ib_angle_accuracy = self.get_accuracy_res_by_phase_angle(ib_p, measure_values_of_ib_angle,
                                                                     phase_angle_accuracy)
            ic_angle_accuracy = self.get_accuracy_res_by_phase_angle(ic_p, measure_values_of_ic_angle,
                                                                     phase_angle_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)
            uab_accuracy = self.get_accuracy_res_by_exp_accuracy(uab, measure_values_of_uab, voltage_accuracy)
            ubc_accuracy = self.get_accuracy_res_by_exp_accuracy(ubc, measure_values_of_ubc, voltage_accuracy)
            uca_accuracy = self.get_accuracy_res_by_exp_accuracy(uca, measure_values_of_uca, voltage_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uab], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uab_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [ubc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ubc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uca], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, uca_accuracy, start_num)

    def fast_precision_measure_by_2e3wn(self, file_path, input_list):
        """
        接线方式:2e3wn
        ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ua,ub,uc不处理，上位机显示N/A
        ③线电流滞后输入相电流30°
        ④功率计算：s_sys_power = (uab * ia + ubc * ib) * cos(30°)
        ⑤升源：输入ib = 0, 计算：ib = ic
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2e3wn_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            # 关注点1, ib输入值/升源值,ub计算值,档位值
            ib, ib_src = ic, ib
            # 关注点2, 系统功率计算
            p_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.cos(math.radians(ua_p - ia_p))
            q_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.sin(math.radians(ua_p - ia_p))
            s_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30))

            start_num = 1
            common_values_of_2e3wn = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub, uc, ia, ib_src, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wn, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib_src, ia, freq)
            (
                measure_values_of_ia,
                measure_values_of_ib,
                measure_values_of_ic,
                measure_values_of_ua_angle,
                measure_values_of_ub_angle,
                measure_values_of_uc_angle,
                measure_values_of_ia_angle,
                measure_values_of_ib_angle,
                measure_values_of_ic_angle,
                measure_values_of_p_sys,
                measure_values_of_q_sys,
                measure_values_of_s_sys,
                measure_values_of_uab,
                measure_values_of_ubc,
                measure_values_of_uca
            ) = self.get_measure_values_by_sample_cnt_of_2e3wn(sample_cnt, sample_interval)
            # 关注点2, 线电流角度
            ia_p, ib_p, ic_p = self.get_phase_angle_of_line_current(ia_p, ib_p, ic_p)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ib_accuracy = self.get_accuracy_res_by_exp_accuracy(ib, measure_values_of_ib, current_accuracy)
            ic_accuracy = self.get_accuracy_res_by_exp_accuracy(ic, measure_values_of_ic, current_accuracy)
            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ub_angle_accuracy = self.get_accuracy_res_by_phase_angle(ub_p, measure_values_of_ub_angle,
                                                                     phase_angle_accuracy)
            uc_angle_accuracy = self.get_accuracy_res_by_phase_angle(uc_p, measure_values_of_uc_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            ib_angle_accuracy = self.get_accuracy_res_by_phase_angle(ib_p, measure_values_of_ib_angle,
                                                                     phase_angle_accuracy)
            ic_angle_accuracy = self.get_accuracy_res_by_phase_angle(ic_p, measure_values_of_ic_angle,
                                                                     phase_angle_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)
            uab_accuracy = self.get_accuracy_res_by_exp_accuracy(uab, measure_values_of_uab, voltage_accuracy)
            ubc_accuracy = self.get_accuracy_res_by_exp_accuracy(ubc, measure_values_of_ubc, voltage_accuracy)
            uca_accuracy = self.get_accuracy_res_by_exp_accuracy(uca, measure_values_of_uca, voltage_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uab], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uab_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [ubc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ubc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uca], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, uca_accuracy, start_num)

    def fast_precision_measure_by_2p5e4wy(self, file_path, input_list):
        """
        接线方式:2p5e4wy
        关注点1：升源：输入ub = 0, 计算：ub = ua
        关注点2：不测In_value
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2p5e4wy_to_excel(file_path, wb, ws)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, phase_angle_accuracy,
             active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            # 关注ub输入值/升源值,ub计算值
            ub, ub_src = ua, ub
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            (pb, qb, sb) = self.get_power_cal_res_of_phase(ub, ib, ub_p, ib_p)
            (pc, qc, sc) = self.get_power_cal_res_of_phase(uc, ic, uc_p, ic_p)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            p_sys = self.get_sys_power(pa, pb, pc)
            q_sys = self.get_sys_power(qa, qb, qc)
            s_sys = self.get_sys_power(sa, sb, sc)

            start_num = 1
            common_values_of_2p5e4wy = (
                case_id, voltage_accuracy, current_accuracy, phase_angle_accuracy, active_power_accuracy,
                ua, ub_src, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2p5e4wy, start_num)
            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub_src, ua, ic, ib, ia, freq)
            (
                measure_values_of_ua,
                measure_values_of_ub,
                measure_values_of_uc,
                measure_values_of_ia,
                measure_values_of_ib,
                measure_values_of_ic,
                measure_values_of_ua_angle,
                measure_values_of_ub_angle,
                measure_values_of_uc_angle,
                measure_values_of_ia_angle,
                measure_values_of_ib_angle,
                measure_values_of_ic_angle,
                measure_values_of_pa,
                measure_values_of_pb,
                measure_values_of_pc,
                measure_values_of_p_sys,
                measure_values_of_qa,
                measure_values_of_qb,
                measure_values_of_qc,
                measure_values_of_q_sys,
                measure_values_of_sa,
                measure_values_of_sb,
                measure_values_of_sc,
                measure_values_of_s_sys,
                measure_values_of_uab,
                measure_values_of_ubc,
                measure_values_of_uca,
            ) = self.get_measure_values_by_sample_cnt_of_2p5e4wy(sample_cnt, sample_interval)
            ua_accuracy = self.get_accuracy_res_by_exp_accuracy(ua, measure_values_of_ua, voltage_accuracy)
            ub_accuracy = self.get_accuracy_res_by_exp_accuracy(ub, measure_values_of_ub, voltage_accuracy)
            uc_accuracy = self.get_accuracy_res_by_exp_accuracy(uc, measure_values_of_uc, voltage_accuracy)
            ia_accuracy = self.get_accuracy_res_by_exp_accuracy(ia, measure_values_of_ia, current_accuracy)
            ib_accuracy = self.get_accuracy_res_by_exp_accuracy(ib, measure_values_of_ib, current_accuracy)
            ic_accuracy = self.get_accuracy_res_by_exp_accuracy(ic, measure_values_of_ic, current_accuracy)

            ua_angle_accuracy = self.get_accuracy_res_by_phase_angle(ua_p, measure_values_of_ua_angle,
                                                                     phase_angle_accuracy)
            ub_angle_accuracy = self.get_accuracy_res_by_phase_angle(ub_p, measure_values_of_ub_angle,
                                                                     phase_angle_accuracy)
            uc_angle_accuracy = self.get_accuracy_res_by_phase_angle(uc_p, measure_values_of_uc_angle,
                                                                     phase_angle_accuracy)
            ia_angle_accuracy = self.get_accuracy_res_by_phase_angle(ia_p, measure_values_of_ia_angle,
                                                                     phase_angle_accuracy)
            ib_angle_accuracy = self.get_accuracy_res_by_phase_angle(ib_p, measure_values_of_ib_angle,
                                                                     phase_angle_accuracy)
            ic_angle_accuracy = self.get_accuracy_res_by_phase_angle(ic_p, measure_values_of_ic_angle,
                                                                     phase_angle_accuracy)
            pa_accuracy = self.get_accuracy_res_by_exp_accuracy(pa, measure_values_of_pa, active_power_accuracy)
            pb_accuracy = self.get_accuracy_res_by_exp_accuracy(pb, measure_values_of_pb, active_power_accuracy)
            pc_accuracy = self.get_accuracy_res_by_exp_accuracy(pc, measure_values_of_pc, active_power_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys,
                                                                   active_power_accuracy)
            qa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qa, measure_values_of_qa)
            qb_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qb, measure_values_of_qb)
            qc_accuracy = self.get_accuracy_res_by_not_exp_accuracy(qc, measure_values_of_qc)
            q_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(q_sys, measure_values_of_q_sys)
            sa_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sa, measure_values_of_sa)
            sb_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sb, measure_values_of_sb)
            sc_accuracy = self.get_accuracy_res_by_not_exp_accuracy(sc, measure_values_of_sc)
            s_sys_accuracy = self.get_accuracy_res_by_not_exp_accuracy(s_sys, measure_values_of_s_sys)
            uab_accuracy = self.get_accuracy_res_by_exp_accuracy(uab, measure_values_of_uab, voltage_accuracy)
            ubc_accuracy = self.get_accuracy_res_by_exp_accuracy(ubc, measure_values_of_ubc, voltage_accuracy)
            uca_accuracy = self.get_accuracy_res_by_exp_accuracy(uca, measure_values_of_uca, voltage_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ua_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ub_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uc_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ia_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ib_angle_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ic_angle_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uab], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, uab_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [ubc], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, ubc_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [uca], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, uca_accuracy, start_num)

    @staticmethod
    def get_save_filepath_of_3e4wy(filedir):
        """
        接线方式:3e4wy
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_3E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2p5e4wy(filedir):
        """
        接线方式:2p5e4wy
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_2p5E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_3e3wd(filedir):
        """
        接线方式:3e3wd
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_3E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wd(filedir):
        """
        接线方式:2e3wd
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_2E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wn(filedir):
        """
        接线方式:2e3wn
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_2E3WN_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3w1p(filedir):
        """
        接线方式:2e3w1p
        :param filedir:待写入文件目录
        :return:待写入文件路径
        """
        filename = f"Precision_Measure_2E3W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_1e2w1p(filedir):
        """
        接线方式:1e2w1p
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_1E2W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    def select_wire_type(self, case_path, sheet_name, wire_type):
        """
        选择接线方式
        :param case_path: 测试文件名
        :param sheet_name: sheet名
        :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
        :return:
        # 3 element 4 wire Wye       3LN     3CT    6   (0 0)
        # 2.5 element 4 wire Wye     3LN-2.5 3CT    5   (5 0)
        # 3 element 3 wire Delta     3LL     3CT    4   (3 0)
        # 2 element 3 wire Delta     2LL     3CT    3   (2 0)
        # 2 element 3 wire network   2LL     2CT    2   (2 2)
        # 2 element 3 wire 1 phase   1LL     2CT    1   (4 2)
        # 1 element 2 wire           1LN     1CT    0   (1 1)
        """
        if wire_type == 0:
            self.set_wire_type(voltage_wire_value=1, current_wire_value=1)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_1e2w1p(save_filedir)
            self.fast_precision_measure_by_1e2w1p(file_path, input_list)
        elif wire_type == 1:
            self.set_wire_type(voltage_wire_value=4, current_wire_value=2)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2e3w1p(save_filedir)
            self.fast_precision_measure_by_2e3w1p(file_path, input_list)
        elif wire_type == 2:
            self.set_wire_type(voltage_wire_value=2, current_wire_value=2)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2e3wn(save_filedir)
            self.fast_precision_measure_by_2e3wn(file_path, input_list)
        elif wire_type == 3:
            self.set_wire_type(voltage_wire_value=2, current_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2e3wd(save_filedir)
            self.fast_precision_measure_by_2e3wd(file_path, input_list)
        elif wire_type == 4:
            self.set_wire_type(voltage_wire_value=3, current_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_3e3wd(save_filedir)
            self.fast_precision_measure_by_3e3wd(file_path, input_list)
        elif wire_type == 5:
            self.set_wire_type(voltage_wire_value=5, current_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2p5e4wy(save_filedir)
            self.fast_precision_measure_by_2p5e4wy(file_path, input_list)
        else:
            self.set_wire_type(voltage_wire_value=0, current_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_3e4wy(save_filedir)
            self.fast_precision_measure_by_3e4wy(file_path, input_list)


def run_precision_measure_script(test_type, wire_type):
    """
    运行脚本入口
    :param test_type: 0:mV, 1:mA, 2:5A
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
    :return:
    """
    print(f"====================Precision Measure Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    switch_device_screen_interface(inter=0x01)  # 切换至交流界面
    set_gear_switching_mode('00000000')  # 档位切换归零
    precision_measure = PrecisionMeasure()

    precision_measure.select_test_case(test_type=test_type, wire_type=wire_type)

    # 关闭ModbusClient客户端连接
    precision_measure.handle_memory.modbus_client.close()
    up_source_ac()
    switch_device_screen_interface(inter=0x00)  # 切换至默认界面
    print(f"====================测试总耗时:{time.time() - start_time}====================")
    print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
    print(f"====================Precision Measure End====================")


if __name__ == '__main__':
    """
    :param test_type: 0:mV, 1:mA, 2:5A
    :param wire_type:0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
    """
    measure_mode = 2
    wire_mode = 6
    run_precision_measure_script(test_type=measure_mode, wire_type=wire_mode)
