#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
文件名称:acuvimseries_color_marked.py
功能描述:标记excel颜色
创建日期:2025-08-05
作者:王洋
版本:v1.0
修改记录:
"""

import os.path
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

current_time = datetime.now().strftime('%Y%m%d')
filedir = Path(__file__).parent

input_filename = "precision_measure_dc260_20250902115440.xlsx"
output_filename = f"precision_measure_color_marked_excel_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
INPUT_FILEPATH = os.path.join(filedir, f"precision_measure_{current_time}", input_filename)
OUTPUT_FILEPATH = os.path.join(filedir, f"color_marked_{current_time}", output_filename)
if not os.path.exists(OUTPUT_FILEPATH):
    os.makedirs(Path(OUTPUT_FILEPATH).parent, exist_ok=True)

input_filename_of_energy = "precision_measure_dc260_20250902115440.xlsx"
output_filename_of_energy = f"energy_measure_color_marked_excel_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
INPUT_FILEPATH_OF_ENERGY = os.path.join(filedir, f"energy_measure_{current_time}", input_filename_of_energy)
OUTPUT_FILEPATH_OF_ENERGY = os.path.join(filedir, f"color_marked_{current_time}",
                                         output_filename_of_energy)
if not os.path.exists(OUTPUT_FILEPATH_OF_ENERGY):
    os.makedirs(Path(OUTPUT_FILEPATH_OF_ENERGY).parent, exist_ok=True)

RED_FORMAT = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
YELLOW_FORMAT = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


class MarkColor:
    standard_columns = {
        'voltage_accuracy': [],
        'current_accuracy': [],
        'power_accuracy': [],
        'accuracy_res': [],
    }

    standard_columns_of_energy = {
        'power_accuracy': [],
        'accuracy_res': [],
    }

    def __init__(self):
        """
        初始化实例,默认文件路径
        """
        self.input_filepath = INPUT_FILEPATH
        self.output_filepath = OUTPUT_FILEPATH
        self.standard_columns = MarkColor.standard_columns

        self.input_filepath_of_energy = INPUT_FILEPATH_OF_ENERGY
        self.output_filepath_of_energy = OUTPUT_FILEPATH_OF_ENERGY
        self.energy_standard_columns = MarkColor.standard_columns_of_energy

    def read_excel_data(self):
        """
        读取和保存测试结果的表头
        :return: 保存表头的字典
        """
        df = pd.read_excel(self.input_filepath, sheet_name=0, nrows=0)
        columns = df.columns.tolist()
        for column in columns:
            if column.endswith("电压精度最小值") or column.endswith("电压精度最大值") or column.endswith(
                    "电压精度平均值"):
                self.standard_columns['voltage_accuracy'].append(column)
            elif column.endswith("电流1精度最小值") or column.endswith("电流1精度最大值") or column.endswith(
                    "电流1精度平均值"):
                self.standard_columns['current_accuracy'].append(column)
            elif column.endswith("电流2精度最小值") or column.endswith("电流2精度最大值") or column.endswith(
                    "电流2精度平均值"):
                self.standard_columns['current_accuracy'].append(column)
            elif column.endswith("功率1精度最小值") or column.endswith("功率1精度最大值") or column.endswith(
                    "功率1精度平均值"):
                self.standard_columns['power_accuracy'].append(column)
            elif column.endswith("功率2精度最小值") or column.endswith("功率2精度最大值") or column.endswith(
                    "功率2精度平均值"):
                self.standard_columns['power_accuracy'].append(column)
            elif column.endswith("功率总精度最小值") or column.endswith("功率总精度最大值") or column.endswith(
                    "功率总精度平均值"):
                self.standard_columns['power_accuracy'].append(column)
            elif column.endswith("精度测试结果") or column.endswith("精度测试结果") or column.endswith("精度测试结果"):
                self.standard_columns['accuracy_res'].append(column)
        return self.standard_columns

    def mark_color(self):
        """
        对测试结果进行颜色标记
        :return:
        """
        start_time = time.perf_counter()
        df = pd.read_excel(self.input_filepath, sheet_name=0)
        wb = openpyxl.load_workbook(self.input_filepath)
        ws = wb.active
        for row in df.itertuples(index=True, name="Pandas"):
            for refer_col, compare_cols in self.standard_columns.items():
                if refer_col == 'accuracy_res':
                    self.fill_color_by_measure_results(ws, df, row, compare_cols)
                else:
                    self.fill_color_by_measure_values(ws, df, row, refer_col, compare_cols)
        wb.save(self.output_filepath)
        print(f"output_filepath:{self.output_filepath}")
        print(f"real_time_color_marked cost_time:{time.perf_counter() - start_time}s")

    def fill_color_by_measure_values(self, ws, df, row_obj, refer_column, compare_columns):
        """
        对多个列的精度值进行标记
        :param ws:工作表对象
        :param df:读取的excel对象
        :param row_obj:行对象
        :param refer_column:基准列
        :param compare_columns:待标记颜色列
        :return:
        """
        for compare_column in compare_columns:
            self.fill_color_by_measure_value(ws, df, row_obj, refer_column, compare_column)

    @staticmethod
    def fill_color_by_measure_value(ws, df, row_obj, refer_column, compare_column):
        """
        对单个单元格的待对比数据进行标记
        :param ws:工作表对象
        :param df:读取的excel对象
        :param row_obj:行对象
        :param refer_column:基准列
        :param compare_column:待标记颜色列
        :return:
        """
        cell = ws.cell(row=row_obj.Index + 2, column=df.columns.get_loc(compare_column) + 1)
        if (getattr(row_obj, compare_column, 0) - getattr(row_obj, refer_column, 0)) > 0.002:
            cell.fill = RED_FORMAT
        elif (getattr(row_obj, compare_column, 0) - getattr(row_obj, refer_column, 0)) > 0:
            cell.fill = YELLOW_FORMAT
        return

    def fill_color_by_measure_results(self, ws, df, row_obj, compare_columns):
        """
        对单个行的"Failed"进行标记
        :param ws:工作表对象
        :param df:读取的excel对象
        :param row_obj:行对象
        :param compare_columns:待标记列
        :return:
        """
        for compare_column in compare_columns:
            self.fill_color_by_measure_result(ws, df, row_obj, compare_column)

    @staticmethod
    def fill_color_by_measure_result(ws, df, row_obj, compare_column):
        """
        对单个单元格的"Failed"进行标记
        :param ws:工作表对象
        :param df:读取的excel对象
        :param row_obj:行对象
        :param compare_column:待标记列
        :return:
        """
        cell = ws.cell(row=row_obj.Index + 2, column=df.columns.get_loc(compare_column) + 1)
        if getattr(row_obj, compare_column, 0) == "Failed":
            cell.fill = RED_FORMAT
            return

    def read_excel_data_of_energy_measure(self):
        """
        读取和保存测试结果的表头
        :return: 保存表头的字典
        """
        df = pd.read_excel(self.input_filepath_of_energy, sheet_name=0, nrows=0)
        columns = df.columns.tolist()
        for column in columns:
            if column.endswith("精度值"):
                self.energy_standard_columns['power_accuracy'].append(column)
            elif column.endswith("精度测试结果") or column.endswith("精度测试结果") or column.endswith("精度测试结果"):
                self.energy_standard_columns['accuracy_res'].append(column)
        return self.energy_standard_columns

    def mark_color_of_energy(self):
        """
        对测试结果进行颜色标记
        :return:
        """
        start_time = time.perf_counter()
        df = pd.read_excel(self.input_filepath_of_energy, sheet_name=0)
        wb = openpyxl.load_workbook(self.input_filepath_of_energy)
        ws = wb.active
        for row in df.itertuples(index=True, name="Pandas"):
            for refer_col, compare_cols in self.energy_standard_columns.items():
                if refer_col == 'accuracy_res':
                    self.fill_color_by_measure_results(ws, df, row, compare_cols)
                else:
                    self.fill_color_by_measure_values(ws, df, row, refer_col, compare_cols)
        wb.save(self.output_filepath_of_energy)
        print(f"output_filepath_of_energy:{self.output_filepath_of_energy}")
        print(f"energy_color_marked cost_time:{time.perf_counter() - start_time}s")

    def run_script(self, color_marked_flag=0):
        """
        运行脚本函数
        :return:
        """
        if not color_marked_flag:
            self.read_excel_data()
            self.mark_color()
        else:
            self.read_excel_data_of_energy_measure()
            self.mark_color_of_energy()


if __name__ == '__main__':
    marked_flag = 0
    MarkColor().run_script(color_marked_flag=marked_flag)
