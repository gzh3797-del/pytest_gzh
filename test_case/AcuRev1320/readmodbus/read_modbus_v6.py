"""
read_modbus.py
──────────────────────────────────────────────────────────────────────────────
从智能电表（Modbus TCP）读取寄存器值，写入 Excel。

目标 Sheet：HS_20ms | Energy_20ms | Real Time | 10S Freq | 200ms
设备 IP   ：192.168.3.129

【Bug修复记录】
  Fix-1: Energy_20ms ws.max_column=16379（Excel 隐藏格式污染）
         → 扫描表头行取真实末列号

  Fix-2: Start(Dec) 列存储 Excel 公式（如 =D142+I142），
         openpyxl 不计算公式，读到字符串后 int() 失败静默跳过，
         导致每个 Block 只有首行写入值。
         → 顺序遍历时维护 prev_start + prev_reg_num 自动推算地址，
           无需依赖公式计算结果，彻底解决。

  Fix-3: 200ms Harmonic 行 Reg Num=128 超过设备单次上限(120)
         → 自动按 BATCH_SIZE=64 分2批读取，合并后按每2个寄存器
           解析1个 float，输出 64 元素数组字符串写入 Decoded Value 列

依赖：pip install pymodbus openpyxl
"""

import struct
import shutil
import traceback
from pathlib import Path
from datetime import datetime, timedelta, timezone
from openpyxl.utils import get_column_letter

from pymodbus.client import ModbusTcpClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 配置 ──────────────────────────────────────────────────────────────────────
METER_IP   = "192.168.3.142"
METER_PORT = 502
UNIT_ID    = 1
TIMEOUT    = 5

SCRIPT_DIR  = Path(__file__).resolve().parent
SOURCE_FILE = SCRIPT_DIR / "Modbus_Address.xlsx"

TARGET_SHEETS = ["HS_20ms", "Energy_20ms", "Real Time", "10S Freq", "200ms"]

HEADER_ROW = 1

COL_START_DEC = 4
COL_DESC      = 6
COL_DTYPE     = 7
COL_REG_NUM   = 9

MAX_REGS_PER_READ = 120   # 设备单次最大寄存器读取数
BATCH_SIZE        = 64    # 分批大小（需 <= MAX_REGS_PER_READ）

TIMESTAMP_KEYWORDS = ("time stamp", "timestamp", "update time")
EPOCH_2006         = datetime(2006, 1, 1, tzinfo=timezone.utc)
CST                = timezone(timedelta(hours=8))   # 北京时间 UTC+8

# 电表时间戳格式说明:
#   格式A (200ms / HS_20ms 等): uint64 整体为毫秒级 Unix 时间戳
#                               直接 /1000 转 datetime，再转 UTC+8 输出
#                               判断: MS_EPOCH_MIN <= val <= MS_EPOCH_MAX
#   格式B (10S Freq 等):        uint64 值过大，无法作为毫秒时间戳
#                               hi32 = 从 2006-01-01(UTC) 起的秒数
#                               lo32 = 纳秒子秒 (取模 1e9)
#                               结果同样转 UTC+8 输出
MS_EPOCH_MIN = 946_684_800_000    # 2000-01-01 00:00:00 UTC 毫秒数
MS_EPOCH_MAX = 4_102_444_800_000  # 2100-01-01 00:00:00 UTC 毫秒数

# 背景色
STYLE_OK        = "E2EFDA"   # 浅绿
STYLE_ERR       = "FFCCCC"   # 浅红
STYLE_TIMESTAMP = "FFF2CC"   # 浅黄
STYLE_ARRAY     = "DDEEFF"   # 浅蓝（数组行）


# ── 样式工具 ──────────────────────────────────────────────────────────────────

def make_header_cell(cell, text, bg, fg="FFFFFF"):
    cell.value     = text
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)


def make_data_cell(cell, value, fill_color, wrap=False):
    cell.value     = value
    cell.font      = Font(name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=wrap)


# ── Fix-1: 真实末列 ───────────────────────────────────────────────────────────

def get_last_real_column(ws):
    """扫描表头行，取最后一个有值的列号，规避 max_column 被隐藏格式污染。"""
    last = 0
    for cell in ws[HEADER_ROW]:
        if cell.value is not None:
            last = cell.column
    return last if last > 0 else 1


# ── 时间戳 ────────────────────────────────────────────────────────────────────

def uint64_to_datetime(val):
    """
    电表存在两种 uint64 时间戳格式，通过值范围自动判断，结果统一转北京时间(UTC+8)：

    格式A — 毫秒级 Unix 时间戳 (200ms / HS_20ms 等)
      uint64 整体为从 1970-01-01 UTC 起的毫秒数
      判断条件: MS_EPOCH_MIN <= val <= MS_EPOCH_MAX (2000年~2100年范围内)
      示例: 0x0000019D_9C1C23E5 = 1776440583141 ms
            → 2026/04/17 23:43:03.141 (UTC+8) ✓

    格式B — 自定义 epoch 秒级时间戳 (10S Freq 等)
      uint64 值超出毫秒范围，采用拆分解析：
      hi32 = 从 2006-01-01 00:00:00 UTC 起的秒数
      lo32 = 纳秒子秒 (取模 1e9)
      示例: 0x2405C288_2405C278
            → hi32=604357256s from 2006-01-01 → 2025/02/25 05:00:56.604 (UTC+8)
    """
    if val == 0:
        return "0 (未同步)"
    try:
        if MS_EPOCH_MIN <= val <= MS_EPOCH_MAX:
            # 格式A: 毫秒级 Unix 时间戳
            ms_int = val % 1000
            dt_utc = datetime.fromtimestamp(val / 1000.0, tz=timezone.utc)
            dt_cst = dt_utc.astimezone(CST)
            return dt_cst.strftime(f"%Y/%m/%d %H:%M:%S.{ms_int:03d}")
        else:
            # 格式B: hi32=自定义epoch秒, lo32=纳秒子秒
            hi32   = (val >> 32) & 0xFFFFFFFF
            lo32   = val & 0xFFFFFFFF
            ns_sub = lo32 % 1_000_000_000
            ms_int = ns_sub // 1_000_000
            dt_utc = EPOCH_2006 + timedelta(seconds=hi32)
            dt_cst = dt_utc.astimezone(CST)
            return dt_cst.strftime(f"%Y/%m/%d %H:%M:%S.{ms_int:03d}")
    except Exception:
        return str(val)


# ── 解码 ──────────────────────────────────────────────────────────────────────

def regs_to_raw_hex(registers):
    return "0x" + "_".join(f"{r:04X}" for r in registers) if registers else ""


def decode_registers(registers, dtype, desc=""):
    if not registers:
        return "N/A"
    raw      = b"".join(r.to_bytes(2, "big") for r in registers)
    dl       = dtype.strip().lower()
    desc_low = desc.strip().lower()
    try:
        if dl == "float":
            return f"{struct.unpack('>f', raw[:4])[0]:.6g}"
        elif dl == "uint64_t":
            val = struct.unpack(">Q", raw[:8])[0]
            return uint64_to_datetime(val) if any(
                k in desc_low for k in TIMESTAMP_KEYWORDS) else str(val)
        elif dl == "uint32_t":
            return str(struct.unpack(">I", raw[:4])[0])
        elif dl == "int32_t":
            return str(struct.unpack(">i", raw[:4])[0])
        elif dl == "uint16_t":
            return str(struct.unpack(">H", raw[:2])[0])
        elif dl == "int16_t":
            return str(struct.unpack(">h", raw[:2])[0])
        else:
            return "0x" + raw.hex().upper()
    except struct.error as e:
        return f"DecodeErr({e})"


def decode_float_array(registers):
    """Fix-3: 每2个寄存器解析1个 float，128寄存器→64 float 列表。"""
    result = []
    for i in range(0, len(registers) - 1, 2):
        raw = registers[i].to_bytes(2, "big") + registers[i + 1].to_bytes(2, "big")
        try:
            result.append(round(struct.unpack(">f", raw)[0], 6))
        except struct.error:
            result.append(None)
    return result


# ── 读取 ──────────────────────────────────────────────────────────────────────

def read_registers_batched(client, addr, count):
    """Fix-3: 超限时分批读取，合并返回完整寄存器列表。"""
    all_regs = []
    offset   = 0
    while offset < count:
        batch = min(BATCH_SIZE, count - offset)
        try:
            resp = client.read_holding_registers(
                address=addr + offset, count=batch, slave=UNIT_ID)
        except Exception as e:
            return None, f"ConnErr batch@{addr+offset}({e})"
        if resp is None or resp.isError():
            return None, f"ModbusErr(addr=0x{addr+offset:04X}, count={batch})"
        all_regs.extend(resp.registers)
        offset += batch
    return all_regs, None


def read_row(client, addr, count, dtype, desc=""):
    """
    读取寄存器并解码。
    count > MAX_REGS_PER_READ → 分批读取 + float 数组输出。
    返回 (decoded_str, raw_hex_str)。
    """
    if count <= 0:
        return "", ""

    if count > MAX_REGS_PER_READ:
        all_regs, err = read_registers_batched(client, addr, count)
        if err:
            return err, ""
        raw_hex    = regs_to_raw_hex(all_regs)
        float_vals = decode_float_array(all_regs)
        arr_str    = "[" + ", ".join(
            "null" if v is None else f"{v:.4g}" for v in float_vals) + "]"
        return arr_str, raw_hex

    try:
        resp = client.read_holding_registers(
            address=addr, count=count, slave=UNIT_ID)
    except Exception as e:
        return f"ConnErr({e})", ""
    if resp is None or resp.isError():
        return f"ModbusErr(addr=0x{addr:04X}, count={count})", ""

    return decode_registers(resp.registers, dtype, desc), \
           regs_to_raw_hex(resp.registers)


# ── 处理单个 Sheet ────────────────────────────────────────────────────────────

def process_sheet(ws, client):
    """
    遍历每行读取寄存器，追加 Decoded Value 和 Raw Hex 两列。

    Fix-2 核心：
      Start(Dec) 可能是硬编码数值，也可能是 Excel 公式字符串（=D142+I142）。
      openpyxl 不计算公式，缓存值在未被 Excel 打开时也为 None。
      解决方案：顺序遍历时记录 prev_addr（上一有效行的起始地址）
      和 prev_count（上一有效行的寄存器数），当前行的地址 =
      prev_addr + prev_count，与公式 =D{n-1}+I{n-1} 完全等价。
    """
    max_col     = get_last_real_column(ws)
    col_decoded = max_col + 1
    col_hex     = max_col + 2

    print(f"  列布局: 末列={max_col}({get_column_letter(max_col)})  "
          f"Decoded={col_decoded}({get_column_letter(col_decoded)})  "
          f"RawHex={col_hex}({get_column_letter(col_hex)})")

    make_header_cell(ws.cell(HEADER_ROW, col_decoded), "Decoded Value", "2E75B6")
    make_header_cell(ws.cell(HEADER_ROW, col_hex),     "Raw Hex",       "375623")
    ws.column_dimensions[get_column_letter(col_decoded)].width = 28
    ws.column_dimensions[get_column_letter(col_hex)].width     = 22

    ok_count   = 0
    skip_count = 0
    prev_addr  = None   # Fix-2: 上一有效行起始地址
    prev_count = None   # Fix-2: 上一有效行寄存器数

    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        raw_start = ws.cell(row_idx, COL_START_DEC).value
        reg_num   = ws.cell(row_idx, COL_REG_NUM).value
        dtype     = ws.cell(row_idx, COL_DTYPE).value
        desc      = ws.cell(row_idx, COL_DESC).value or ""

        # 无类型 → 标题行/空行/合计行，跳过
        if dtype is None:
            skip_count += 1
            continue

        # ── Fix-2: 解析地址 ───────────────────────────────────────────────────
        # 规则：
        #   硬编码数值  → 直接使用，并重置推算基准（Block 首行必有硬编码）
        #   公式/None  → 用 prev_addr + prev_count 推算（等价于 =D{n-1}+I{n-1}）
        # Block 之间地址不连续，靠硬编码首行自动重置基准，确保跨 Block 正确。
        addr = None
        is_hardcoded = (raw_start is not None and not (
            isinstance(raw_start, str) and raw_start.startswith("=")))

        if is_hardcoded:
            try:
                addr = int(raw_start)
                # 硬编码行重置推算基准，避免跨 Block 地址污染
                prev_addr  = None
                prev_count = None
            except (ValueError, TypeError):
                pass

        if addr is None:
            # 公式行或 None → 用 prev_addr + prev_count 推算
            if prev_addr is not None and prev_count is not None:
                addr = prev_addr + prev_count
            else:
                print(f"  [WARN] row {row_idx}: 无法推算地址，跳过。({desc[:30]})")
                skip_count += 1
                continue

        # 寄存器数
        try:
            count = int(reg_num)
        except (TypeError, ValueError):
            skip_count += 1
            continue
        if count <= 0:
            skip_count += 1
            continue

        # ── 更新前驱（供下一行公式推算使用）────────────────────────────────
        prev_addr  = addr
        prev_count = count

        is_large = count > MAX_REGS_PER_READ
        print(f"  [{ws.title}] row {row_idx:>4}  addr={addr:<7}  "
              f"regs={count:<4}  dtype={str(dtype):<12}  "
              f"{'[BATCH]' if is_large else '       '}  {str(desc)[:35]}")

        decoded, raw_hex = read_row(client, addr, count, str(dtype), str(desc))

        is_ts  = (any(k in desc.lower() for k in TIMESTAMP_KEYWORDS)
                  and str(dtype).strip().lower() == "uint64_t")
        is_err = "Err" in decoded

        fill = (STYLE_ERR       if is_err  else
                STYLE_ARRAY     if is_large else
                STYLE_TIMESTAMP if is_ts   else
                STYLE_OK)

        make_data_cell(ws.cell(row_idx, col_decoded), decoded, fill,
                       wrap=is_large)
        make_data_cell(ws.cell(row_idx, col_hex), raw_hex, fill)

        if is_large:
            ws.row_dimensions[row_idx].height = 60

        ok_count += 1

    return ok_count, skip_count


# ── 主函数 ────────────────────────────────────────────────────────────────────

def main():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(
            f"\n找不到源文件！\n"
            f"  期望路径: {SOURCE_FILE}\n"
            f"  脚本目录: {SCRIPT_DIR}\n"
            f"请将 Modbus_Address.xlsx 与脚本放在同一目录。"
        )

    ts_suffix   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = SCRIPT_DIR / f"Modbus_Address_ReadResult_{ts_suffix}.xlsx"

    # 注意：使用 data_only=False 保留公式结构（Fix-2 已改用推算法，无需依赖缓存）
    shutil.copy2(SOURCE_FILE, output_file)
    print(f"[✓] 已复制原文件 → {output_file.name}")

    print(f"\n[→] 连接 {METER_IP}:{METER_PORT}  unit={UNIT_ID} …")
    client = ModbusTcpClient(host=METER_IP, port=METER_PORT, timeout=TIMEOUT)
    if not client.connect():
        raise ConnectionError(f"无法连接 {METER_IP}:{METER_PORT}")
    print("[✓] 连接成功\n")

    wb         = openpyxl.load_workbook(output_file, data_only=False)
    total_ok   = 0
    total_skip = 0
    run_start  = datetime.now()

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[!] Sheet '{sheet_name}' 不存在，跳过")
            continue
        print(f"\n{'='*62}")
        print(f"  处理 Sheet: {sheet_name}")
        print(f"{'='*62}")
        ok, skip    = process_sheet(wb[sheet_name], client)
        total_ok   += ok
        total_skip += skip
        print(f"  ✓ 完成：读取 {ok} 行，跳过 {skip} 行")

    client.close()
    wb.save(output_file)

    elapsed = (datetime.now() - run_start).total_seconds()
    print(f"\n{'='*62}")
    print(f"[✓] 完成！读取 {total_ok} 行，跳过 {total_skip} 行")
    print(f"    耗时 : {elapsed:.1f} 秒")
    print(f"    输出 : {output_file}")
    print(f"{'='*62}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[!] 用户中断")
    except Exception:
        traceback.print_exc()
