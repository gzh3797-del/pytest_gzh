#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
文件名称:acuvimserires_energy_measure.py
功能描述:能量测量
创建日期:2025-08-05
作者:王洋
版本:v1.0
修改记录:
"""

import os
import statistics
import threading
import time
import math
import cmath
import logging
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from test_case.AcuRev1320.fast_test.acuvimseries_modbus_get import HandleMemory
from comm.source_control import (switch_device_screen_interface, set_gear_switching_mode, set_ac, set_voltage_gear,
                                 set_current_gear, up_source_ac)
from test_case.AcuRev1320.fast_test.memory_addrs import MemoryAddr
from test_case.AcuvimSeries.test_report_table_heading import TableTitle
from tools.excel_operate import data_read
from tools.log import Log
from power_calculate import CalculatePower

Log(str(__file__).split("\\")[-1])

test_case_path = r'./test_case/AcuvimSeries/acuvimseries_test_case.xlsx'
sheet_name_mV = "test_case_mV"
sheet_name_mA = "test_case_mA"
sheet_name_rct = "test_case_rct"

save_filedir = os.path.join(Path(__file__).parent, f"energy_measure_{time.strftime('%Y%m%d')}")
if not os.path.exists(save_filedir):
    os.makedirs(save_filedir, exist_ok=True)


class EnergyMeasure:
    def __init__(self):
        self.handle_memory = HandleMemory(slave_id=1)

    def select_test_type(self, test_type, wire_type):
        """
        选择测试数据类型
        # :param test_type: 0:mV, 1:mA, 2:5A
        # :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY

        :param test_type: 0:mV, 1:mA, 2:rct
        :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 5:3e3wD  # AcuRev1320 项目
        :return:
        """
        #  目前西安实验室CT为333mV，脚本先适配333mV CT
        if test_type == 2:
            self.select_wire_type(test_case_path, sheet_name_mV, wire_type)
        #  目前西安实验室CT为100mA，脚本先适配100mA CT
        elif test_type == 0:
            self.select_wire_type(test_case_path, sheet_name_mA, wire_type)
        elif test_type == 3:
            self.select_wire_type(test_case_path, sheet_name_rct, wire_type)
        # 80mA 暂不适配
        elif test_type == 1:
            print("80mA CT暂未适配")

    def set_wire_type(self, voltage_wire_value):
        """
        设置接线方式
        :param voltage_wire_value: 电压接线方式
        # :param current_wire_value: 电流接线方式
        :return:
        """
        self.handle_memory.set_wire_mode_by_voltage(voltage_wire_mode=voltage_wire_value)
        # self.handle_memory.set_wire_mode_by_current(current_wire_mode=current_wire_value)

    def select_test_case(self, test_type, wire_type):
        """
        选择测试标准，接线方式
        :param test_type: 0:mV, 1:mA, 2:5A
        :param wire_type:0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
        :return:
        """
        self.select_test_type(test_type, wire_type)

    @staticmethod
    def get_input_list_by_wire_type(source_input_list, wire_type):
        """
        获取接线方式的数据
        :param source_input_list: 测试用例数据
        :param wire_type: 接线类型
        :return: 接线方式的数据
        """
        input_list = [source_input_list[0]]
        if wire_type == 6:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "3E4wY":
                    input_list.append(source_input_list[i])
        elif wire_type == 5:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2.5E4wY":
                    input_list.append(source_input_list[i])
        elif wire_type == 4:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "3E3wD":
                    input_list.append(source_input_list[i])
        elif wire_type == 3:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2E3wD":
                    input_list.append(source_input_list[i])
        elif wire_type == 2:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2E3wN":
                    input_list.append(source_input_list[i])
        elif wire_type == 1:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "2E3w1p":
                    input_list.append(source_input_list[i])
        elif wire_type == 0:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][15] == "1E2w1p":
                    input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_real_time_parameters_by_input_list_of_wire_type(source_input_list):
        """
        获取接线方式的快速测量实时数据
        :param source_input_list: 接线方式的数据
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            if not source_input_list[i][13]:
                input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_energy_parameters_by_input_list_of_wire_type(source_input_list):
        """
        获取接线方式的能量测量实时数据
        :param source_input_list: 接线方式的数据
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            if source_input_list[i][13]:
                input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def write_table_title_of_3e4wy_to_excel(file_path, wb, ws):
        """
        写入表头:3e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_3E4WY)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_3E4WY[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2p5e4wy_to_excel(file_path, wb, ws):
        """
        写入表头:2p5e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_2p5E4WY)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_2p5E4WY[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_1e2w1p_to_excel(file_path, wb, ws):
        """
        写入表头:1e2w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_1E2W1P)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_1E2W1P[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3w1p_to_excel(file_path, wb, ws):
        """
        写入表头:2e3w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_2E3W1P)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_2E3W1P[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_3e3wd_to_excel(file_path, wb, ws):
        """
        写入表头:3e3wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_3E3WD)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_3E3WD[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wd_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_2E3WD)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_2E3WD[i]}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wn_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wn
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.ENERGY_MEASURE_COLUMNS_OF_2E3WN)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.ENERGY_MEASURE_COLUMNS_OF_2E3WN[i]}')
        wb.save(file_path)

    @staticmethod
    def get_memory_addr():
        """
        获取寄存器值
        :return: 寄存器值
        """
        ua_rms_addr = MemoryAddr.ua_rms_addr
        ub_rms_addr = MemoryAddr.ub_rms_addr
        uc_rms_addr = MemoryAddr.uc_rms_addr
        u_avg_rms_addr = MemoryAddr.uv_avg_rms_addr
        ia_rms_addr = MemoryAddr.ia_rms_addr
        ib_rms_addr = MemoryAddr.ib_rms_addr
        ic_rms_addr = MemoryAddr.ic_rms_addr
        i_avg_rms_addr = MemoryAddr.iv_avg_rms_addr
        pa_rms_addr = MemoryAddr.pa_rms_addr
        pb_rms_addr = MemoryAddr.pb_rms_addr
        pc_rms_addr = MemoryAddr.pc_rms_addr
        p_total_rms_addr = MemoryAddr.p_total_rms_addr
        qa_rms_addr = MemoryAddr.qa_rms_addr
        qb_rms_addr = MemoryAddr.qb_rms_addr
        qc_rms_addr = MemoryAddr.qc_rms_addr
        q_total_rms_addr = MemoryAddr.q_total_rms_addr
        sa_rms_addr = MemoryAddr.sa_rms_addr
        sb_rms_addr = MemoryAddr.sb_rms_addr
        sc_rms_addr = MemoryAddr.sc_rms_addr
        s_total_rms_addr = MemoryAddr.s_total_rms_addr
        ua_p_rms_addr = MemoryAddr.ua_phase_angle_rms_addr
        ub_p_rms_addr = MemoryAddr.ub_phase_angle_rms_addr
        uc_p_rms_addr = MemoryAddr.uc_phase_angle_rms_addr
        ia_p_rms_addr = MemoryAddr.ia_phase_angle_rms_addr
        ib_p_rms_addr = MemoryAddr.ib_phase_angle_rms_addr
        ic_p_rms_addr = MemoryAddr.ic_phase_angle_rms_addr
        uab_rms_addr = MemoryAddr.uab_rms_addr
        ubc_rms_addr = MemoryAddr.ubc_rms_addr
        uca_rms_addr = MemoryAddr.uca_rms_addr
        ul_avg_rms_addr = MemoryAddr.ul_avg_rms_addr
        pf_a_rms_addr = MemoryAddr.pf_a_rms_addr
        pf_b_rms_addr = MemoryAddr.pf_b_rms_addr
        pf_c_rms_addr = MemoryAddr.pf_c_rms_addr
        pf_total_rms_addr = MemoryAddr.pf_total_rms_addr
        freq_rms_addr = MemoryAddr.freq_rms_addr
        return (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
                ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
                pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
                qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
                sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
                ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
                ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
                uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
                pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
                freq_rms_addr)

    @staticmethod
    def get_accuracy_by_datatype(accuracy_value):
        """
        处理N/A精度值
        :param accuracy_value: 精度值
        :return: 精度值
        """
        accuracy_res = accuracy_value
        if not isinstance(accuracy_value, float):
            accuracy_res = 999999
        return accuracy_res

    def get_test_case_info_of_accuracy(self, input_list, index_value):
        """
        获取测试精度信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 精度值
        """
        accuracy_of_voltage = input_list[index_value + 1][17]
        accuracy_of_current = input_list[index_value + 1][18]
        accuracy_of_phase_angle = input_list[index_value + 1][19]
        accuracy_of_active_power = input_list[index_value + 1][20]
        accuracy_of_reactive_power = input_list[index_value + 1][21]
        accuracy_of_apparent_power = input_list[index_value + 1][22]
        voltage_accuracy = self.get_accuracy_by_datatype(accuracy_of_voltage)
        current_accuracy = self.get_accuracy_by_datatype(accuracy_of_current)
        phase_angle_accuracy = self.get_accuracy_by_datatype(accuracy_of_phase_angle)
        # voltage_phase_angle_accuracy = self.get_accuracy_by_datatype(accuracy_of_phase_angle)
        # current_phase_angle_accuracy = self.get_accuracy_by_datatype(accuracy_of_phase_angle)
        active_power_accuracy = self.get_accuracy_by_datatype(accuracy_of_active_power)
        reactive_power_accuracy = self.get_accuracy_by_datatype(accuracy_of_reactive_power)
        apparent_power_accuracy = self.get_accuracy_by_datatype(accuracy_of_apparent_power)
        return (voltage_accuracy, current_accuracy, phase_angle_accuracy,
                active_power_accuracy, reactive_power_accuracy, apparent_power_accuracy)

    @staticmethod
    def get_test_case_info_of_input_value(input_list, index_value):
        """
        获取测试电压/电流、相位等信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 电压/电流、相位等信息
        """
        case_id = input_list[index_value + 1][0]
        ua = input_list[index_value + 1][1]
        ub = input_list[index_value + 1][2]
        uc = input_list[index_value + 1][3]
        ia = input_list[index_value + 1][4]
        ib = input_list[index_value + 1][5]
        ic = input_list[index_value + 1][6]
        ua_p = input_list[index_value + 1][7]
        ub_p = input_list[index_value + 1][8]
        uc_p = input_list[index_value + 1][9]
        ia_p = input_list[index_value + 1][10]
        ib_p = input_list[index_value + 1][11]
        ic_p = input_list[index_value + 1][12]
        freq = input_list[index_value + 1][16]
        sample_cnt = input_list[index_value + 1][23]
        sample_interval = input_list[index_value + 1][24]
        run_time = input_list[index_value + 1][13]
        return (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
                freq, sample_cnt, sample_interval, run_time)

    @staticmethod
    def get_power_cal_res_of_phase(value_of_voltage, value_of_current,
                                   value_of_voltage_phase_angle, value_of_current_phase_angle):
        """
        通过电压计算功率
        :param value_of_voltage: 电压
        :param value_of_current: 电流
        :param value_of_voltage_phase_angle: 电压相位角度
        :param value_of_current_phase_angle: 电流相位角度
        :return: 功率
        """

        active_power_res = CalculatePower.calculate_active_power(value_of_voltage,
                                                                 value_of_current,
                                                                 value_of_voltage_phase_angle,
                                                                 value_of_current_phase_angle)
        reactive_power_res = CalculatePower.calculate_reactive_power(value_of_voltage,
                                                                     value_of_current,
                                                                     value_of_voltage_phase_angle,
                                                                     value_of_current_phase_angle)
        apparent_power_res = CalculatePower.calculate_apparent_power(value_of_voltage, value_of_current)

        return active_power_res, reactive_power_res, apparent_power_res

    @staticmethod
    def get_sys_power(*power_of_phase):
        """
        计算系统功率
        :param power_of_phase: A/B/C相功率
        :return: 系统功率
        """
        sys_power = sum(power_of_phase)
        return sys_power

    @staticmethod
    def get_energy_of_imp_and_exp(test_time, phase_power):
        """
        计算能量进/出
        :param test_time: 时间
        :param phase_power: 功率
        :return: 进/出能量
        """
        energy_imp = 0
        energy_exp = 0
        # energy_imp,energy_exp均为正数
        if phase_power or phase_power == 0:
            energy_imp = phase_power * test_time
        else:
            energy_exp = abs(phase_power * test_time)
        return energy_imp, energy_exp

    def get_phase_energy_of_imp_and_exp(self, test_time, active_power, reactive_power, apparent_power):
        """
        计算有功/无功/视在的能量进/出
        :param self: 实例
        :param test_time:运行时间
        :param active_power:有功功率
        :param reactive_power:无功功率
        :param apparent_power:视在功率
        :return:有功/无功/视在的能量进/出
        """
        p_imp, p_exp = self.get_energy_of_imp_and_exp(test_time, active_power)
        q_imp, q_exp = self.get_energy_of_imp_and_exp(test_time, reactive_power)
        s_imp, s_exp = self.get_energy_of_imp_and_exp(test_time, apparent_power)
        return p_imp, p_exp, q_imp, q_exp, s_imp, s_exp

    @staticmethod
    def get_sys_energy_of_imp_or_exp(*energy_values):
        """
        计算系统能量进/出
        :param energy_values:
        :return: 系统能量进/出
        """
        sys_energy = sum(energy_values)
        return sys_energy

    @staticmethod
    def get_sys_energy_of_total_and_net(energy_imp, energy_exp):
        """
        计算系统能量进/出/净/总
        :param energy_imp: 进线能量
        :param energy_exp: 出线能量
        :return: 系统能量进/出/净总
        """
        sys_total_energy = energy_imp + energy_exp
        sys_net_energy = energy_imp - energy_exp
        return sys_total_energy, sys_net_energy

    @staticmethod
    def get_phase_energy_of_apparent(energy_imp, energy_exp):
        """
        计算系统能量进/出/净/总
        :param energy_imp: 进线能量
        :param energy_exp: 出线能量
        :return: 系统能量进/出/净总
        """
        phase_total_energy = energy_imp + energy_exp
        phase_net_energy = energy_imp - energy_exp
        return phase_total_energy, phase_net_energy

    @staticmethod
    def write_common_values_to_excel(ws, index_value, common_values, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param common_values: case编号/电压/电流/电压相位角/电流相位角等输入值,精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(common_values)):
            ws.cell(index_value + 2, start_num + k, common_values[k])
        start_num += len(common_values)
        return start_num

    @staticmethod
    def write_accuracy_res_to_excel(file_path, wb, ws, index_value, accuracy_res, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param file_path: 待写入文件路径
        :param wb: 写入工作簿对象
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param accuracy_res: 精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(accuracy_res)):
            ws.cell(index_value + 2, start_num + k, accuracy_res[k])
        wb.save(file_path)
        start_num += len(accuracy_res)
        return start_num

    # def set_and_switch_source(self,uc_value, ub_value, ua_value,ic_value, ib_value, ia_value):
    #     """
    #     升源加设置档位
    #     :return:
    #     """
    #     # 设置电压/电流档位，当前是相同值，只配置了一种A相
    #     set_voltage_gear(uc_value, ub_value, ua_value)
    #     set_current_gear(ic_value, ib_value, ia_value)

    @staticmethod
    def get_line_current_angle(phase_current_angle):
        """
        delta类型接线方式，获取线电流角度
        :param phase_current_angle: 相电流角度
        :return: 线电流角度
        """
        phase_current_angle = phase_current_angle if ((phase_current_angle - 30) >= 0) else (phase_current_angle + 360)
        return phase_current_angle

    def get_phase_angle_of_line_current(self, ia_angle, ib_angle, ic_angle):
        """
        delta类型接线方式，获取线电流角度
        :param ia_angle: ia角度
        :param ib_angle: ib角度
        :param ic_angle: ic角度
        :return: 线电流角度
        """
        ia_angle = self.get_line_current_angle(ia_angle)
        ib_angle = self.get_line_current_angle(ib_angle)
        ic_angle = self.get_line_current_angle(ic_angle)
        return ia_angle, ib_angle, ic_angle

    def get_measure_values_of_phase_voltage(self):
        """
        获取相电压测量值
        :return: 相电压测量值
        """
        ua = self.handle_memory.read_ua_voltage()
        ub = self.handle_memory.read_ub_voltage()
        uc = self.handle_memory.read_uc_voltage()
        uv_avg = self.handle_memory.read_uv_avg_voltage()
        return ua, ub, uc, uv_avg

    def get_measure_values_of_sys_power(self):
        """
        获取系统功率测量值
        :return: 系统功率测量值
        """
        p_total = self.handle_memory.read_p_total_power()
        q_total = self.handle_memory.read_q_total_power()
        s_total = self.handle_memory.read_s_total_power()
        return p_total, q_total, s_total

    def get_measure_values_of_active_energy_imp(self):
        """
        获取有功功率进线测量值
        :return: 有功功率进线测量值
        """
        pa_imp = self.handle_memory.read_pa_imp_energy()
        pb_imp = self.handle_memory.read_pb_imp_energy()
        pc_imp = self.handle_memory.read_pc_imp_energy()
        # p_sys_imp = self.handle_memory.read_p_sys_imp_energy()
        p_sys_imp = sum([pa_imp, pb_imp, pc_imp])
        return pa_imp, pb_imp, pc_imp, p_sys_imp

    def get_measure_values_of_active_energy_exp(self):
        """
        获取有功功率出线测量值
        :return: 有功功率出线测量值
        """
        pa_exp = self.handle_memory.read_pa_exp_energy()
        pb_exp = self.handle_memory.read_pb_exp_energy()
        pc_exp = self.handle_memory.read_pc_exp_energy()
        # p_sys_exp = self.handle_memory.read_p_sys_exp_energy()
        p_sys_exp = sum([pa_exp, pb_exp, pc_exp])
        return pa_exp, pb_exp, pc_exp, p_sys_exp

    def get_measure_values_of_apparent_energy(self):
        """
        获取视在功率进/出/总/净测量值
        :return: 视在功率进/出/总/净测量值
        """
        s_sys_imp = self.handle_memory.read_s_sys_imp_energy()
        s_sys_exp = self.handle_memory.read_s_sys_exp_energy()
        s_sys_total = self.handle_memory.read_q_sys_total_energy()
        return s_sys_imp, s_sys_exp, s_sys_total

    def get_measure_values_of_reactive_energy(self):
        """
        获取无功功率进/出/总/净测量值
        :return: 无功功率进/出/总/净测量值
        """
        q_sys_imp = self.handle_memory.read_q_sys_imp_energy()
        q_sys_exp = self.handle_memory.read_q_sys_exp_energy()
        q_sys_total = self.handle_memory.read_q_sys_total_energy()
        q_sys_net = self.handle_memory.read_q_sys_net_energy()
        return q_sys_imp, q_sys_exp, q_sys_total, q_sys_net

    def get_measure_values_of_active_energy(self):
        """
        获取有功功率进/出/总/净测量值
        :return: 有功功率进/出/总/净测量值
        """
        p_sys_imp = self.handle_memory.read_p_sys_imp_energy()
        p_sys_exp = self.handle_memory.read_p_sys_exp_energy()
        p_sys_total = self.handle_memory.read_p_sys_total_energy()
        p_sys_net = self.handle_memory.read_p_sys_net_energy()
        return p_sys_imp, p_sys_exp, p_sys_total, p_sys_net

    @staticmethod
    def get_measure_values_of_active_energy_total(p_sys_imp, p_sys_exp):
        """
        获取有功功率出线测量值
        :return: 有功功率出线测量值
        """
        p_sys_total = sum([p_sys_imp, p_sys_exp])
        return p_sys_total

    @staticmethod
    def get_measure_values_of_active_energy_net(p_sys_imp, p_sys_exp):
        """
        获取有功功率出线测量值
        :return: 有功功率出线测量值
        """
        p_sys_net = p_sys_imp - p_sys_exp
        return p_sys_net

    def get_measure_values_of_active_energy_total_bak(self):
        """
        获取有功功率出线测量值
        :return: 有功功率出线测量值
        """
        p_sys_total = self.handle_memory.read_p_sys_total_energy()
        return p_sys_total

    def get_measure_values_of_active_energy_net_bak(self):
        """
        获取有功功率出线测量值
        :return: 有功功率出线测量值
        """
        p_sys_net = self.handle_memory.read_p_sys_net_energy()
        return p_sys_net

    def get_measure_values_of_reactive_energy_imp(self):
        """
        获取无功功率进线测量值
        :return: 无功功率测量值
        """
        qa_imp = self.handle_memory.read_qa_imp_energy()
        qb_imp = self.handle_memory.read_qb_imp_energy()
        qc_imp = self.handle_memory.read_qc_imp_energy()
        # q_sys_imp = self.handle_memory.read_q_sys_imp_energy()
        q_sys_imp = sum([qa_imp, qb_imp, qc_imp])
        return qa_imp, qb_imp, qc_imp, q_sys_imp

    def get_measure_values_of_reactive_energy_exp(self):
        """
        获取无功功率出线测量值
        :return: 无功功率出线测量值
        """
        qa_exp = self.handle_memory.read_qa_exp_energy()
        qb_exp = self.handle_memory.read_qb_exp_energy()
        qc_exp = self.handle_memory.read_qc_exp_energy()
        # q_sys_exp = self.handle_memory.read_q_sys_exp_energy()
        q_sys_exp = sum([qa_exp, qb_exp, qc_exp])
        return qa_exp, qb_exp, qc_exp, q_sys_exp

    @staticmethod
    def get_measure_values_of_reactive_energy_total(q_sys_imp, q_sys_exp):
        """
        获取无功功率出线测量值
        :return: 获取无功功率出线测量值
        """
        q_sys_total = sum([q_sys_imp, q_sys_exp])
        return q_sys_total

    @staticmethod
    def get_measure_values_of_reactive_energy_net(q_sys_imp, q_sys_exp):
        """
        获取无功功率出线测量值
        :return: 无功功率出线测量值
        """
        q_sys_net = q_sys_imp - q_sys_exp
        return q_sys_net

    def get_measure_values_of_reactive_energy_total_bak(self):
        """
        获取无功功率出线测量值
        :return: 获取无功功率出线测量值
        """
        q_sys_total = self.handle_memory.read_q_sys_total_energy()
        return q_sys_total

    def get_measure_values_of_reactive_energy_net_bak(self):
        """
        获取无功功率出线测量值
        :return: 无功功率出线测量值
        """
        q_sys_net = self.handle_memory.read_q_sys_net_energy()
        return q_sys_net

    def get_measure_values_of_apparent_energy_imp(self):
        """
        获取视在功率测量值
        :return: 视在功率测量值
        """
        sa_imp = self.handle_memory.read_sa_imp_energy()
        sb_imp = self.handle_memory.read_sb_imp_energy()
        sc_imp = self.handle_memory.read_sc_imp_energy()
        # s_sys_imp = self.handle_memory.read_s_sys_imp_energy()
        s_sys_imp = sum([sa_imp, sb_imp, sc_imp])
        return sa_imp, sb_imp, sc_imp, s_sys_imp

    def get_measure_values_of_apparent_energy_exp(self):
        """
        获取视在功率测量值
        :return: 视在功率测量值
        """
        sa_exp = self.handle_memory.read_sa_exp_energy()
        sb_exp = self.handle_memory.read_sb_exp_energy()
        sc_exp = self.handle_memory.read_sc_exp_energy()
        # s_sys_exp = self.handle_memory.read_s_sys_exp_energy()
        s_sys_exp = sum([sa_exp, sb_exp, sc_exp])
        return sa_exp, sb_exp, sc_exp, s_sys_exp

    @staticmethod
    def get_measure_values_of_apparent_energy_total(s_sys_imp, s_sys_exp):
        """
        获取视在功率测量值
        :return: 视在功率测量值
        """
        s_sys_total = sum([s_sys_imp, s_sys_exp])
        return s_sys_total

    def get_measure_values_of_apparent_energy_total_bak(self):
        """
        获取视在功率测量值
        :return: 视在功率测量值
        """
        s_sys_total = self.handle_memory.read_s_sys_total_energy()
        return s_sys_total

    def get_measure_values_of_apparent_energy_app(self):
        """
        获取视在功率测量值
        :return: 视在功率测量值
        """
        sa_app = self.handle_memory.read_sa_app_energy()
        sb_app = self.handle_memory.read_sb_app_energy()
        sc_app = self.handle_memory.read_sc_app_energy()
        return sa_app, sb_app, sc_app

    def get_measure_values_by_sample_time_of_3e4wy(self):
        """
        获取测量值:3e4wy
        :return: 测量值
        """
        (
            pa_imp_measure,
            pb_imp_measure,
            pc_imp_measure,
            p_sys_imp_measure
        ) = self.get_measure_values_of_active_energy_imp()
        (
            pa_exp_measure,
            pb_exp_measure,
            pc_exp_measure,
            p_sys_exp_measure
        ) = self.get_measure_values_of_active_energy_exp()
        p_sys_total_measure = self.get_measure_values_of_active_energy_total(p_sys_imp_measure, p_sys_exp_measure)
        p_sys_net_measure = self.get_measure_values_of_active_energy_net(p_sys_imp_measure, p_sys_exp_measure)
        (
            qa_imp_measure,
            qb_imp_measure,
            qc_imp_measure,
            q_sys_imp_measure
        ) = self.get_measure_values_of_reactive_energy_imp()
        (
            qa_exp_measure,
            qb_exp_measure,
            qc_exp_measure,
            q_sys_exp_measure
        ) = self.get_measure_values_of_reactive_energy_exp()
        q_sys_total_measure = self.get_measure_values_of_reactive_energy_total(q_sys_imp_measure, q_sys_exp_measure)
        q_sys_net_measure = self.get_measure_values_of_reactive_energy_net(q_sys_imp_measure, q_sys_exp_measure)
        (
            sa_imp_measure,
            sb_imp_measure,
            sc_imp_measure,
            s_sys_imp_measure
        ) = self.get_measure_values_of_apparent_energy_imp()
        (
            sa_exp_measure,
            sb_exp_measure,
            sc_exp_measure,
            s_sys_exp_measure
        ) = self.get_measure_values_of_apparent_energy_exp()
        s_sys_total_measure = self.get_measure_values_of_apparent_energy_total(s_sys_imp_measure, s_sys_exp_measure)
        (
            sa_app_measure,
            sb_app_measure,
            sc_app_measure
        ) = self.get_measure_values_of_apparent_energy_app()
        return (
            pa_imp_measure,
            pb_imp_measure,
            pc_imp_measure,
            pa_exp_measure,
            pb_exp_measure,
            pc_exp_measure,
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            qa_imp_measure,
            qb_imp_measure,
            qc_imp_measure,
            qa_exp_measure,
            qb_exp_measure,
            qc_exp_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            sa_imp_measure,
            sb_imp_measure,
            sc_imp_measure,
            sa_exp_measure,
            sb_exp_measure,
            sc_exp_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
            sa_app_measure,
            sb_app_measure,
            sc_app_measure,
        )

    def get_measure_values_by_sample_time_of_2e3w1p(self):
        """
        获取测量值:2e3w1p
        :return: 测量值
        """
        (
            pa_imp_measure,
            pb_imp_measure,
            _,
            p_sys_imp_measure
        ) = self.get_measure_values_of_active_energy_imp()
        (
            pa_exp_measure,
            pb_exp_measure,
            _,
            p_sys_exp_measure
        ) = self.get_measure_values_of_active_energy_exp()
        p_sys_total_measure = self.get_measure_values_of_active_energy_total(p_sys_imp_measure, p_sys_exp_measure)
        p_sys_net_measure = self.get_measure_values_of_active_energy_net(p_sys_imp_measure, p_sys_exp_measure)
        (
            qa_imp_measure,
            qb_imp_measure,
            _,
            q_sys_imp_measure
        ) = self.get_measure_values_of_reactive_energy_imp()
        (
            qa_exp_measure,
            qb_exp_measure,
            _,
            q_sys_exp_measure
        ) = self.get_measure_values_of_reactive_energy_exp()
        q_sys_total_measure = self.get_measure_values_of_reactive_energy_total(q_sys_imp_measure, q_sys_exp_measure)
        q_sys_net_measure = self.get_measure_values_of_reactive_energy_net(q_sys_imp_measure, q_sys_exp_measure)
        (
            sa_imp_measure,
            sb_imp_measure,
            _,
            s_sys_imp_measure
        ) = self.get_measure_values_of_apparent_energy_imp()
        (
            sa_exp_measure,
            sb_exp_measure,
            _,
            s_sys_exp_measure
        ) = self.get_measure_values_of_apparent_energy_exp()
        s_sys_total_measure = self.get_measure_values_of_apparent_energy_total(s_sys_imp_measure, s_sys_exp_measure)
        (
            sa_app_measure,
            sb_app_measure,
            _
        ) = self.get_measure_values_of_apparent_energy_app()
        return (
            pa_imp_measure,
            pb_imp_measure,
            pa_exp_measure,
            pb_exp_measure,
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            qa_imp_measure,
            qb_imp_measure,
            qa_exp_measure,
            qb_exp_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            sa_imp_measure,
            sb_imp_measure,
            sa_exp_measure,
            sb_exp_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
            sa_app_measure,
            sb_app_measure,
        )

    def get_measure_values_by_sample_time_of_1e2w1p(self):
        """
        获取测量值:1e2w1p
        :return: 测量值
        """
        (
            pa_imp_measure,
            _,
            _,
            p_sys_imp_measure
        ) = self.get_measure_values_of_active_energy_imp()
        (
            pa_exp_measure,
            _,
            _,
            p_sys_exp_measure
        ) = self.get_measure_values_of_active_energy_exp()
        p_sys_total_measure = self.get_measure_values_of_active_energy_total(p_sys_imp_measure, p_sys_exp_measure)
        p_sys_net_measure = self.get_measure_values_of_active_energy_net(p_sys_imp_measure, p_sys_exp_measure)
        (
            qa_imp_measure,
            _,
            _,
            q_sys_imp_measure
        ) = self.get_measure_values_of_reactive_energy_imp()
        (
            qa_exp_measure,
            _,
            _,
            q_sys_exp_measure
        ) = self.get_measure_values_of_reactive_energy_exp()
        q_sys_total_measure = self.get_measure_values_of_reactive_energy_total(q_sys_imp_measure, q_sys_exp_measure)
        q_sys_net_measure = self.get_measure_values_of_reactive_energy_net(q_sys_imp_measure, q_sys_exp_measure)
        (
            sa_imp_measure,
            _,
            _,
            s_sys_imp_measure
        ) = self.get_measure_values_of_apparent_energy_imp()
        (
            sa_exp_measure,
            _,
            _,
            s_sys_exp_measure
        ) = self.get_measure_values_of_apparent_energy_exp()
        s_sys_total_measure = self.get_measure_values_of_apparent_energy_total(s_sys_imp_measure, s_sys_exp_measure)
        (
            sa_app_measure,
            _,
            _
        ) = self.get_measure_values_of_apparent_energy_app()
        return (
            pa_imp_measure,
            pa_exp_measure,
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            qa_imp_measure,
            qa_exp_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            sa_imp_measure,
            sa_exp_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
            sa_app_measure,
        )

    def get_measure_values_by_sample_time_of_2p5e4wy(self):
        """
        获取测量值:2p5e4wy
        :return: 测量值
        """
        (
            pa_imp_measure,
            pb_imp_measure,
            pc_imp_measure,
            p_sys_imp_measure
        ) = self.get_measure_values_of_active_energy_imp()
        (
            pa_exp_measure,
            pb_exp_measure,
            pc_exp_measure,
            p_sys_exp_measure
        ) = self.get_measure_values_of_active_energy_exp()
        p_sys_total_measure = self.get_measure_values_of_active_energy_total(p_sys_imp_measure, p_sys_exp_measure)
        p_sys_net_measure = self.get_measure_values_of_active_energy_net(p_sys_imp_measure, p_sys_exp_measure)
        (
            qa_imp_measure,
            qb_imp_measure,
            qc_imp_measure,
            q_sys_imp_measure
        ) = self.get_measure_values_of_reactive_energy_imp()
        (
            qa_exp_measure,
            qb_exp_measure,
            qc_exp_measure,
            q_sys_exp_measure
        ) = self.get_measure_values_of_reactive_energy_exp()
        q_sys_total_measure = self.get_measure_values_of_reactive_energy_total(q_sys_imp_measure, q_sys_exp_measure)
        q_sys_net_measure = self.get_measure_values_of_reactive_energy_net(q_sys_imp_measure, q_sys_exp_measure)
        (
            sa_imp_measure,
            sb_imp_measure,
            sc_imp_measure,
            s_sys_imp_measure
        ) = self.get_measure_values_of_apparent_energy_imp()
        (
            sa_exp_measure,
            sb_exp_measure,
            sc_exp_measure,
            s_sys_exp_measure
        ) = self.get_measure_values_of_apparent_energy_exp()
        s_sys_total_measure = self.get_measure_values_of_apparent_energy_total(s_sys_imp_measure, s_sys_exp_measure)
        (
            sa_app_measure,
            sb_app_measure,
            sc_app_measure
        ) = self.get_measure_values_of_apparent_energy_app()
        return (
            pa_imp_measure,
            pb_imp_measure,
            pc_imp_measure,
            pa_exp_measure,
            pb_exp_measure,
            pc_exp_measure,
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            qa_imp_measure,
            qb_imp_measure,
            qc_imp_measure,
            qa_exp_measure,
            qb_exp_measure,
            qc_exp_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            sa_imp_measure,
            sb_imp_measure,
            sc_imp_measure,
            sa_exp_measure,
            sb_exp_measure,
            sc_exp_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
            sa_app_measure,
            sb_app_measure,
            sc_app_measure,
        )

    def get_measure_values_by_sample_time_of_3e3wd(self):
        """
        获取测量值:3e3wd
        :return: 测量值
        """
        (
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure
        ) = self.get_measure_values_of_active_energy()
        (
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure
        ) = self.get_measure_values_of_reactive_energy()
        (
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
        ) = self.get_measure_values_of_apparent_energy()
        return (
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
        )

    def get_measure_values_by_sample_time_of_2e3wd(self):
        """
        获取测量值:2e3wd
        :return: 测量值
        """
        (
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure
        ) = self.get_measure_values_of_active_energy()
        (
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure
        ) = self.get_measure_values_of_reactive_energy()
        (
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
        ) = self.get_measure_values_of_apparent_energy()
        return (
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
        )

    def get_measure_values_by_sample_time_of_2e3wn(self):
        """
        获取测量值:2e3wn
        :return: 测量值
        """
        (
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure
        ) = self.get_measure_values_of_active_energy()
        (
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure
        ) = self.get_measure_values_of_reactive_energy()
        (
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
        ) = self.get_measure_values_of_apparent_energy()
        return (
            p_sys_imp_measure,
            p_sys_exp_measure,
            p_sys_total_measure,
            p_sys_net_measure,
            q_sys_imp_measure,
            q_sys_exp_measure,
            q_sys_total_measure,
            q_sys_net_measure,
            s_sys_imp_measure,
            s_sys_exp_measure,
            s_sys_total_measure,
        )

    @staticmethod
    def get_cmp_accuracy_res(act_accuracy, exp_accuracy):
        """
        获取精度比较结果
        :param act_accuracy: 精度测量值
        :param exp_accuracy: 精度期望值
        :return: "Passed"/"Failed"
        """
        cmp_res = "Failed"
        if act_accuracy <= exp_accuracy:
            cmp_res = "Passed"
        return cmp_res

    @staticmethod
    def get_all_accuracy_values(all_accuracy_values):
        """
        获取最终精度比较结果
        :param all_accuracy_values: 全部精度值列表
        :return: 最终精度比较结果
        """
        all_cmp_res = "Passed"
        if "Failed" in all_accuracy_values:
            all_cmp_res = "Failed"
        return all_cmp_res

    @staticmethod
    def get_measure_accuracy_by_energy(standard_value, measure_value):
        """
        获取能量精度计算结果
        :param standard_value: 输入能量值
        :param measure_value: 寄存器测量值
        :return: 能量精度计算结果
        """
        mea_val = measure_value
        if standard_value:
            mea_val_accuracy = round(abs((mea_val - standard_value) / standard_value), 5)
        else:
            if all(val == 0 for val in [mea_val, standard_value]) or mea_val < 0.001:
                mea_val_accuracy = 0
            else:
                mea_val_accuracy = 1
        return mea_val, mea_val_accuracy

    def get_accuracy_res_by_exp_accuracy(self, standard_value, measure_value, exp_accuracy):
        """
        获取精度比较结果，带期望精度值
        :param standard_value: 输入值
        :param measure_value: 测量值
        :param exp_accuracy: 期望精度值
        :return: 测量值和精度比较结果
        """
        (
            mea_value, mea_accuracy
        ) = self.get_measure_accuracy_by_energy(standard_value, measure_value)

        cmp_accuracy_res = self.get_cmp_accuracy_res(act_accuracy=mea_accuracy, exp_accuracy=exp_accuracy)
        accuracy_res = (mea_value, mea_accuracy, cmp_accuracy_res)
        return accuracy_res

    def get_accuracy_res_by_not_exp_accuracy(self, standard_value, measure_values):
        """
        获取精度比较结果，不带期望精度值
        :param standard_value: 输入值
        :param measure_values: 测量值
        :return: 测量值结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_voltage_current_power(standard_value, measure_values)

        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        accuracy_res = (min_value, min_accuracy, max_value, max_accuracy, avg_value, avg_accuracy)
        return accuracy_res

    @staticmethod
    def line_to_line_voltage_calculate(ua, ub, uc, va_angle, vb_angle, vc_angle):
        """
        计算线电压
        :param ua:ua
        :param ub:ub
        :param uc:uc
        :param va_angle:va_angle
        :param vb_angle:vb_angle
        :param vc_angle:vc_angle
        :return: vab, vbc, vca
        """
        ret = []
        va_complex = complex(math.cos(va_angle * math.pi / 180) * ua, math.sin(va_angle * math.pi / 180) * ua)
        vb_complex = complex(math.cos(vb_angle * math.pi / 180) * ub, math.sin(vb_angle * math.pi / 180) * ub)
        vc_complex = complex(math.cos(vc_angle * math.pi / 180) * uc, math.sin(vc_angle * math.pi / 180) * uc)
        vab = va_complex - vb_complex
        vbc = vb_complex - vc_complex
        vca = vc_complex - va_complex
        vab = abs(vab)
        vbc = abs(vbc)
        vca = abs(vca)
        ret.extend([vab, vbc, vca])
        return ret

    @staticmethod
    def calculate_2e3wn_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p):
        """
        2E3WN 系统功率计算（P_sys, Q_sys, S_sys）
        :param ua, ub, uc: 相电压幅值
        :param ua_p, ub_p, uc_p: 相电压相角，单位°
        :param ia, ib, ic: 相电流幅值
        :param ia_p, ib_p, ic_p: 相电流相角，单位°
        :return: P_sys, Q_sys, S_sys
        """
        # 构建复数相量
        va = cmath.rect(ua, math.radians(ua_p))
        vb = cmath.rect(ub, math.radians(ub_p))
        vc = cmath.rect(uc, math.radians(uc_p))

        ia_c = cmath.rect(ia, math.radians(ia_p))
        ic_c = cmath.rect(ic, math.radians(ic_p))

        # 线电压复数
        vab = va - vb
        vcb = vc - vb

        # 系统复功率（Aron 两元件法）
        s_total = vab * ia_c.conjugate() + vcb * ic_c.conjugate()

        p_sys = s_total.real
        q_sys = s_total.imag
        s_sys = abs(s_total)

        return p_sys, q_sys, s_sys

    @staticmethod
    def calculate_3e3wd_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p):
        """
        3E3WD 系统功率计算
        :param ua, ub, uc: 相电压幅值
        :param ua_p, ub_p, uc_p: 相电压相角，单位°
        :param ia, ib, ic: 相电流幅值
        :param ia_p, ib_p, ic_p: 相电流相角，单位°
        :return: P_sys, Q_sys, S_sys
        """
        # 构建电压复数
        va = cmath.rect(ua, math.radians(ua_p))
        vb = cmath.rect(ub, math.radians(ub_p))
        vc = cmath.rect(uc, math.radians(uc_p))

        # 线电压复数（3E3WD: AB, BC, CA）
        vab = va - vb
        vbc = vb - vc
        vca = vc - va

        # 构建电流复数
        ia_c = cmath.rect(ia, math.radians(ia_p))
        ib_c = cmath.rect(ib, math.radians(ib_p))
        ic_c = cmath.rect(ic, math.radians(ic_p))

        # 系统复功率
        s_total = vab * ia_c.conjugate() + vbc * ib_c.conjugate() + vca * ic_c.conjugate()

        p_sys = s_total.real
        q_sys = s_total.imag
        s_sys = abs(s_total)

        return p_sys, q_sys, s_sys

    @staticmethod
    def calculate_2e3wd_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p):
        """
        2e3wD sys功率计算方法
        Args:
            ua: 相电压A
            ub: 相电压B
            uc: 相电压C
            ua_p: 相电压A的相角
            ub_p: 相电压B的相角
            uc_p: 相电压C的相角
            ia: 电流A
            ib: 电流B
            ic: 电流C
            ia_p: 电流A的相角
            ib_p: 电流B的相角
            ic_p: 电流C的相角
        计算 U_AB/I_A 和 U_CB/I_C 相位差，并打印中间角度
        :return: p_sys, q_sys, s_sys
        """
        # 构建复数相量
        va = cmath.rect(ua, math.radians(ua_p))
        vb = cmath.rect(ub, math.radians(ub_p))
        vc = cmath.rect(uc, math.radians(uc_p))

        ia_c = cmath.rect(ia, math.radians(ia_p))
        ic_c = cmath.rect(ic, math.radians(ic_p))

        vab = va - vb
        vcb = vc - vb

        # 打印中间相角
        print(f"U_AB 相角: {math.degrees(cmath.phase(vab)):.2f}°")
        print(f"U_CB 相角: {math.degrees(cmath.phase(vcb)):.2f}°")
        print(f"I_A 相角: {math.degrees(cmath.phase(ia_c)):.2f}°")
        print(f"I_C 相角: {math.degrees(cmath.phase(ic_c)):.2f}°")

        # 相位差
        ua_ia_angle = math.degrees(cmath.phase(vab / ia_c))
        uc_ic_angle = math.degrees(cmath.phase(vcb / ic_c))

        # 归一化到 [-180, 180]
        ua_ia_angle = (ua_ia_angle + 180) % 360 - 180
        uc_ic_angle = (uc_ic_angle + 180) % 360 - 180
        # 打印中间相角
        print(f"U_AB与Ia夹角: {ua_ia_angle:.2f}°")
        print(f"U_CB与Ic夹角: {uc_ic_angle:.2f}°")

        # 有功、无功功率（复数法，考虑两条线）
        s_total = vab * ia_c.conjugate() + vcb * ic_c.conjugate()

        p_sys = s_total.real
        q_sys = s_total.imag
        s_sys = abs(s_total)

        return p_sys, q_sys, s_sys

    @staticmethod
    def wait_test_run_time(test_time):
        """
        等待指定时间后输出
        :param test_time: 测试时长
        :return:
        """
        time.sleep(test_time * 60 * 60)  # 转换成秒
        logging.info(f"等待 {test_time}h 结束!")
        print(f"等待 {test_time}h 结束！")

    def energy_precision_measure_by_3e4wy(self, file_path, input_list):
        """
        接线方式:3e4wy
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            (pb, qb, sb) = self.get_power_cal_res_of_phase(ub, ib, ub_p, ib_p)
            (pc, qc, sc) = self.get_power_cal_res_of_phase(uc, ic, uc_p, ic_p)
            # 计算功率进线、出现、视在等输入值
            pa_imp, pa_exp, qa_imp, qa_exp, sa_imp, sa_exp = self.get_phase_energy_of_imp_and_exp(run_time, pa, qa, sa)
            pb_imp, pb_exp, qb_imp, qb_exp, sb_imp, sb_exp = self.get_phase_energy_of_imp_and_exp(run_time, pb, qb, sb)
            pc_imp, pc_exp, qc_imp, qc_exp, sc_imp, sc_exp = self.get_phase_energy_of_imp_and_exp(run_time, pc, qc, sc)
            p_sys_imp = self.get_sys_energy_of_imp_or_exp(pa_imp, pb_imp, pc_imp)
            p_sys_exp = self.get_sys_energy_of_imp_or_exp(pa_exp, pb_exp, pc_exp)
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_imp = self.get_sys_energy_of_imp_or_exp(qa_imp, qb_imp, qc_imp)
            q_sys_exp = self.get_sys_energy_of_imp_or_exp(qa_exp, qb_exp, qc_exp)
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_imp = self.get_sys_energy_of_imp_or_exp(sa_imp, sb_imp, sc_imp)
            s_sys_exp = self.get_sys_energy_of_imp_or_exp(sa_exp, sb_exp, sc_exp)
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)
            sa_app = self.get_phase_energy_of_apparent(sa_imp, sa_exp)
            sb_app = self.get_phase_energy_of_apparent(sb_imp, sb_exp)
            sc_app = self.get_phase_energy_of_apparent(sc_imp, sc_exp)

            start_num = 1
            common_values_of_3e4wy = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wy, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                pa_imp_measure,
                pb_imp_measure,
                pc_imp_measure,
                pa_exp_measure,
                pb_exp_measure,
                pc_exp_measure,
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                qa_imp_measure,
                qb_imp_measure,
                qc_imp_measure,
                qa_exp_measure,
                qb_exp_measure,
                qc_exp_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                sa_imp_measure,
                sb_imp_measure,
                sc_imp_measure,
                sa_exp_measure,
                sb_exp_measure,
                sc_exp_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure,
                sa_app_measure,
                sb_app_measure,
                sc_app_measure,
            ) = self.get_measure_values_by_sample_time_of_3e4wy()

            pa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_imp, pa_imp_measure, active_power_accuracy)
            pa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_exp, pa_exp_measure, active_power_accuracy)
            qa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_imp, qa_imp_measure, active_power_accuracy)
            qa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_exp, qa_exp_measure, active_power_accuracy)
            sa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_imp, sa_imp_measure, active_power_accuracy)
            sa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_exp, sa_exp_measure, active_power_accuracy)
            sa_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_app, sa_app_measure, active_power_accuracy)

            pb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pb_imp, pb_imp_measure, active_power_accuracy)
            pb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pb_exp, pb_exp_measure, active_power_accuracy)
            qb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qb_imp, qb_imp_measure, active_power_accuracy)
            qb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qb_exp, qb_exp_measure, active_power_accuracy)
            sb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_imp, sb_imp_measure, active_power_accuracy)
            sb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_exp, sb_exp_measure, active_power_accuracy)
            sb_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_app, sb_app_measure, active_power_accuracy)

            pc_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pc_imp, pc_imp_measure, active_power_accuracy)
            pc_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pc_exp, pc_exp_measure, active_power_accuracy)
            qc_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qc_imp, qc_imp_measure, active_power_accuracy)
            qc_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qc_exp, qc_exp_measure, active_power_accuracy)
            sc_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sc_imp, sc_imp_measure, active_power_accuracy)
            sc_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sc_exp, sc_exp_measure, active_power_accuracy)
            sc_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sc_app, sc_app_measure, active_power_accuracy)

            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pc_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pc_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pc_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pc_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qc_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qc_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qc_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qc_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    def energy_precision_measure_by_2e3w1p(self, file_path, input_list):
        """
        接线方式:2e3w1p
        ①ua/ub/uab/ia/ib/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②uc/ubc/uca/ic/pc/qc/sc/,不处理，上位机显示N/A
        ③相位角度：ua_p和ub_p相差180°,ia_p和ib_p相差180°
        ④处理uc_p=0, ic_p=0
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2e3w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            # 处理uc_p/ic_p
            uc_p = 0
            ic_p = 0
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            (pb, qb, sb) = self.get_power_cal_res_of_phase(ub, ib, ub_p, ib_p)
            # 计算功率进线、出现、视在等输入值
            pa_imp, pa_exp, qa_imp, qa_exp, sa_imp, sa_exp = self.get_phase_energy_of_imp_and_exp(run_time, pa, qa, sa)
            pb_imp, pb_exp, qb_imp, qb_exp, sb_imp, sb_exp = self.get_phase_energy_of_imp_and_exp(run_time, pb, qb, sb)
            p_sys_imp = self.get_sys_energy_of_imp_or_exp(pa_imp, pb_imp)
            p_sys_exp = self.get_sys_energy_of_imp_or_exp(pa_exp, pb_exp)
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_imp = self.get_sys_energy_of_imp_or_exp(qa_imp, qb_imp)
            q_sys_exp = self.get_sys_energy_of_imp_or_exp(qa_exp, qb_exp)
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_imp = self.get_sys_energy_of_imp_or_exp(sa_imp, sb_imp)
            s_sys_exp = self.get_sys_energy_of_imp_or_exp(sa_exp, sb_exp)
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)
            sa_app = self.get_phase_energy_of_apparent(sa_imp, sa_exp)
            sb_app = self.get_phase_energy_of_apparent(sb_imp, sb_exp)

            start_num = 1
            common_values_of_2e3w1p = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3w1p, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                pa_imp_measure,
                pb_imp_measure,
                pa_exp_measure,
                pb_exp_measure,
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                qa_imp_measure,
                qb_imp_measure,
                qa_exp_measure,
                qb_exp_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                sa_imp_measure,
                sb_imp_measure,
                sa_exp_measure,
                sb_exp_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure,
                sa_app_measure,
                sb_app_measure
            ) = self.get_measure_values_by_sample_time_of_2e3w1p()
            pa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_imp, pa_imp_measure, active_power_accuracy)
            pa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_imp, pa_exp_measure, active_power_accuracy)
            qa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_imp, qa_imp_measure, active_power_accuracy)
            qa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_imp, qa_exp_measure, active_power_accuracy)
            sa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_imp, sa_imp_measure, active_power_accuracy)
            sa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_imp, sa_exp_measure, active_power_accuracy)
            sa_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_app, sa_app_measure, active_power_accuracy)

            pb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pb_imp, pb_imp_measure, active_power_accuracy)
            pb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pb_imp, pb_exp_measure, active_power_accuracy)
            qb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qb_imp, qb_imp_measure, active_power_accuracy)
            qb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qb_imp, qb_exp_measure, active_power_accuracy)
            sb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_imp, sb_imp_measure, active_power_accuracy)
            sb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_imp, sb_exp_measure, active_power_accuracy)
            sb_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_app, sb_app_measure, active_power_accuracy)

            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    def energy_precision_measure_by_1e2w1p(self, file_path, input_list):
        """
        接线方式:1e2w1p
        ①ua/ia/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ub/uc/uab/ubc/uca/ib/ic/pb/pc/qb/qc/sb/sc,不处理，上位机显示N/A
        ③处理ub_p=0, ib_p =0, uc_p=0, ic_p=0
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_1e2w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            # 处理ub_p/ib_p/uc_p/ic_p
            ub_p = 0
            uc_p = 0
            ib_p = 0
            ic_p = 0
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            # 计算功率进线、出现、视在等输入值，pa_imp, pa_exp, qa_imp, qa_exp, sa_imp, sa_exp 均为能量
            pa_imp, pa_exp, qa_imp, qa_exp, sa_imp, sa_exp = self.get_phase_energy_of_imp_and_exp(run_time, pa, qa, sa)
            p_sys_imp = self.get_sys_energy_of_imp_or_exp([pa_imp])
            p_sys_exp = self.get_sys_energy_of_imp_or_exp([pa_exp])
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_imp = self.get_sys_energy_of_imp_or_exp([qa_imp])
            q_sys_exp = self.get_sys_energy_of_imp_or_exp([qa_exp])
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_imp = self.get_sys_energy_of_imp_or_exp([sa_imp])
            s_sys_exp = self.get_sys_energy_of_imp_or_exp([sa_exp])
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)
            sa_app = self.get_phase_energy_of_apparent(sa_imp, sa_exp)

            start_num = 1
            common_values_of_1e2w1p = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_1e2w1p, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                pa_imp_measure,
                pa_exp_measure,
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                qa_imp_measure,
                qa_exp_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                sa_imp_measure,
                sa_exp_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure,
                sa_app_measure,
            ) = self.get_measure_values_by_sample_time_of_1e2w1p()
            pa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_imp, pa_imp_measure, active_power_accuracy)
            pa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_exp, pa_exp_measure, active_power_accuracy)
            qa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_imp, qa_imp_measure, active_power_accuracy)
            qa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_exp, qa_exp_measure, active_power_accuracy)
            sa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_imp, sa_imp_measure, active_power_accuracy)
            sa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_exp, sa_exp_measure, active_power_accuracy)
            sa_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_app, sa_app_measure, active_power_accuracy)

            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    def energy_precision_measure_by_3e3wd(self, file_path, input_list):
        """
        接线方式:3e3wd
        ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ua,ub,uc不处理，上位机显示N/A
        ③线电流滞后输入相电流30°
        ④功率计算：s_sys_power = √3￣ * ill_avg * ull_avg
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_3e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            # 关注点1，系统功率计算
            # ull_avg = round(statistics.median([uab, ubc, uca]), 5)
            # ill_avg = round(statistics.median([ia, ib, ic]), 5)
            # s_sys = math.sqrt(3) * ull_avg * ill_avg
            # p_sys = s_sys * (math.cos(math.radians(ua_p - ia_p)))
            # q_sys = s_sys * (math.sin(math.radians(ua_p - ia_p)))
            p_sys, q_sys, s_sys = self.calculate_3e3wd_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p)

            # 计算功率进线、出现、视在等输入值
            (
                p_sys_imp, p_sys_exp, q_sys_imp, q_sys_exp, s_sys_imp, s_sys_exp
            ) = self.get_phase_energy_of_imp_and_exp(run_time, p_sys, q_sys, s_sys)
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)

            start_num = 1
            common_values_of_3e3wd = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e3wd, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure
            ) = self.get_measure_values_by_sample_time_of_3e3wd()
            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    def energy_precision_measure_by_2e3wd(self, file_path, input_list):
        """
        接线方式:2e3wd
        ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ua,ub,uc不处理，上位机显示N/A
        ③线电流滞后输入相电流30°
        ④功率计算：s_sys_power = (uab * ia + ubc * ib) * cos(30°)
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            # 关注点1, 系统功率计算
            # p_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.cos(math.radians(ua_p - ia_p))
            # q_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.sin(math.radians(ua_p - ia_p))
            # s_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30))
            p_sys, q_sys, s_sys = self.calculate_3e3wd_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p)

            # 计算功率进线、出现、视在等输入值
            (
                p_sys_imp, p_sys_exp, q_sys_imp, q_sys_exp, s_sys_imp, s_sys_exp
            ) = self.get_phase_energy_of_imp_and_exp(run_time, p_sys, q_sys, s_sys)
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)

            start_num = 1
            common_values_of_2e3wd = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wd, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure
            ) = self.get_measure_values_by_sample_time_of_2e3wd()
            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    def energy_precision_measure_by_2e3wn(self, file_path, input_list):
        """
        接线方式:2e3wn
        ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
        ②ua,ub,uc不处理，上位机显示N/A
        ③线电流滞后输入相电流30°
        ④功率计算：s_sys_power = (uab * ia + ubc * ib) * cos(30°)
        ⑤升源：输入ib = 0, 计算：ib = ic
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2e3wn_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
            # 关注点1, ib输入值/升源值,ub计算值,档位值
            ib, ib_src = ic, ib
            # 关注点2, 系统功率计算
            # p_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.cos(math.radians(ua_p - ia_p))
            # q_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30)) * math.sin(math.radians(ua_p - ia_p))
            # s_sys = (uab * ia + ubc * ib) * math.cos(math.radians(30))
            p_sys, q_sys, s_sys = self.calculate_2e3wn_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p)

            # 计算功率进线、出现、视在等输入值
            (
                p_sys_imp, p_sys_exp, q_sys_imp, q_sys_exp, s_sys_imp, s_sys_exp
            ) = self.get_phase_energy_of_imp_and_exp(run_time, p_sys, q_sys, s_sys)
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)

            start_num = 1
            common_values_of_2e3wn = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wn, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib_src, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure
            ) = self.get_measure_values_by_sample_time_of_2e3wn()

            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    def energy_precision_measure_by_2p5e4wy(self, file_path, input_list):
        """
        接线方式:2p5e4wy
        关注点1：升源：输入ub = 0, 计算：ub = ua
        关注点2：不测In_value
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_2p5e4wy_to_excel(file_path, wb, ws)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (_, _, _, active_power_accuracy, _, _) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p,
             freq, _, _, run_time) = self.get_test_case_info_of_input_value(input_list, i)
            # 关注ub输入值/升源值,ub计算值
            ub, ub_src = ua, ub
            (pa, qa, sa) = self.get_power_cal_res_of_phase(ua, ia, ua_p, ia_p)
            (pb, qb, sb) = self.get_power_cal_res_of_phase(ub, ib, ub_p, ib_p)
            (pc, qc, sc) = self.get_power_cal_res_of_phase(uc, ic, uc_p, ic_p)
            # 计算功率进线、出现、视在等输入值
            pa_imp, pa_exp, qa_imp, qa_exp, sa_imp, sa_exp = self.get_phase_energy_of_imp_and_exp(run_time, pa, qa, sa)
            pb_imp, pb_exp, qb_imp, qb_exp, sb_imp, sb_exp = self.get_phase_energy_of_imp_and_exp(run_time, pb, qb, sb)
            pc_imp, pc_exp, qc_imp, qc_exp, sc_imp, sc_exp = self.get_phase_energy_of_imp_and_exp(run_time, pc, qc, sc)
            p_sys_imp = self.get_sys_energy_of_imp_or_exp(pa_imp, pb_imp, pc_imp)
            p_sys_exp = self.get_sys_energy_of_imp_or_exp(pa_exp, pb_exp, pc_exp)
            p_sys_total, p_sys_net = self.get_sys_energy_of_total_and_net(p_sys_imp, p_sys_exp)
            q_sys_imp = self.get_sys_energy_of_imp_or_exp(qa_imp, qb_imp, qc_imp)
            q_sys_exp = self.get_sys_energy_of_imp_or_exp(qa_exp, qb_exp, qc_exp)
            q_sys_total, q_sys_net = self.get_sys_energy_of_total_and_net(q_sys_imp, q_sys_exp)
            s_sys_imp = self.get_sys_energy_of_imp_or_exp(sa_imp, sb_imp, sc_imp)
            s_sys_exp = self.get_sys_energy_of_imp_or_exp(sa_exp, sb_exp, sc_exp)
            s_sys_total, _ = self.get_sys_energy_of_total_and_net(s_sys_imp, s_sys_exp)
            sa_app = self.get_phase_energy_of_apparent(sa_imp, sa_exp)
            sb_app = self.get_phase_energy_of_apparent(sb_imp, sb_exp)
            sc_app = self.get_phase_energy_of_apparent(sc_imp, sc_exp)

            start_num = 1
            common_values_of_2p5e4wy = (
                case_id, active_power_accuracy,
                ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, run_time
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_2p5e4wy, start_num)
            # 升源+设置电压/电流档位+关源
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub_src, ua, ic, ib, ia, freq)
            # 清除能量,保持485连接
            self.handle_memory.set_cleared_energy(clear_energy_flag=1)
            thread_of_wait_time = threading.Thread(target=self.wait_test_run_time, args=(run_time,))
            thread_of_hold_time = threading.Thread(target=self.handle_memory.hold_rs485_connect, args=(run_time,))
            thread_of_wait_time.start()
            thread_of_hold_time.start()
            thread_of_wait_time.join()
            thread_of_hold_time.join()
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            (
                pa_imp_measure,
                pb_imp_measure,
                pc_imp_measure,
                pa_exp_measure,
                pb_exp_measure,
                pc_exp_measure,
                p_sys_imp_measure,
                p_sys_exp_measure,
                p_sys_total_measure,
                p_sys_net_measure,
                qa_imp_measure,
                qb_imp_measure,
                qc_imp_measure,
                qa_exp_measure,
                qb_exp_measure,
                qc_exp_measure,
                q_sys_imp_measure,
                q_sys_exp_measure,
                q_sys_total_measure,
                q_sys_net_measure,
                sa_imp_measure,
                sb_imp_measure,
                sc_imp_measure,
                sa_exp_measure,
                sb_exp_measure,
                sc_exp_measure,
                s_sys_imp_measure,
                s_sys_exp_measure,
                s_sys_total_measure,
                sa_app_measure,
                sb_app_measure,
                sc_app_measure,
            ) = self.get_measure_values_by_sample_time_of_2p5e4wy()

            pa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_imp, pa_imp_measure, active_power_accuracy)
            pa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pa_exp, pa_exp_measure, active_power_accuracy)
            qa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_imp, qa_imp_measure, active_power_accuracy)
            qa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qa_exp, qa_exp_measure, active_power_accuracy)
            sa_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_imp, sa_imp_measure, active_power_accuracy)
            sa_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_exp, sa_exp_measure, active_power_accuracy)
            sa_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sa_app, sa_app_measure, active_power_accuracy)

            pb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pb_imp, pb_imp_measure, active_power_accuracy)
            pb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pb_exp, pb_exp_measure, active_power_accuracy)
            qb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qb_imp, qb_imp_measure, active_power_accuracy)
            qb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qb_exp, qb_exp_measure, active_power_accuracy)
            sb_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_imp, sb_imp_measure, active_power_accuracy)
            sb_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_exp, sb_exp_measure, active_power_accuracy)
            sb_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sb_app, sb_app_measure, active_power_accuracy)

            pc_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(pc_imp, pc_imp_measure, active_power_accuracy)
            pc_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(pc_exp, pc_exp_measure, active_power_accuracy)
            qc_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(qc_imp, qc_imp_measure, active_power_accuracy)
            qc_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(qc_exp, qc_exp_measure, active_power_accuracy)
            sc_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(sc_imp, sc_imp_measure, active_power_accuracy)
            sc_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(sc_exp, sc_exp_measure, active_power_accuracy)
            sc_app_accuracy = self.get_accuracy_res_by_exp_accuracy(sc_app, sc_app_measure, active_power_accuracy)

            p_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_imp, p_sys_imp_measure,
                                                                       active_power_accuracy)
            p_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_exp, p_sys_exp_measure,
                                                                       active_power_accuracy)
            p_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_total, p_sys_total_measure,
                                                                         active_power_accuracy)
            p_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys_net, p_sys_net_measure,
                                                                       active_power_accuracy)

            q_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_imp, q_sys_imp_measure,
                                                                       active_power_accuracy)
            q_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_exp, q_sys_exp_measure,
                                                                       active_power_accuracy)
            q_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_total, q_sys_total_measure,
                                                                         active_power_accuracy)
            q_sys_net_accuracy = self.get_accuracy_res_by_exp_accuracy(q_sys_net, q_sys_net_measure,
                                                                       active_power_accuracy)

            s_sys_imp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_imp, s_sys_imp_measure,
                                                                       active_power_accuracy)
            s_sys_exp_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_exp, s_sys_exp_measure,
                                                                       active_power_accuracy)
            s_sys_total_accuracy = self.get_accuracy_res_by_exp_accuracy(s_sys_total, s_sys_total_measure,
                                                                         active_power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sa_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sa_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sb_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sb_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pc_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pc_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [pc_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, pc_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qc_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qc_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [qc_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, qc_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [sc_app], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, sc_app_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_total], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_total_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [q_sys_net], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, q_sys_net_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_imp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_imp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_exp], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_exp_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [s_sys_total], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, s_sys_total_accuracy, start_num)

    @staticmethod
    def get_save_filepath_of_3e4wy(filedir):
        """
        接线方式:3e4wy
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_3E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2p5e4wy(filedir):
        """
        接线方式:2p5e4wy
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_2p5E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_3e3wd(filedir):
        """
        接线方式:3e3wd
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_3E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wd(filedir):
        """
        接线方式:2e3wd
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_2E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wn(filedir):
        """
        接线方式:2e3wn
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_2E3WN_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3w1p(filedir):
        """
        接线方式:2e3w1p
        :param filedir:待写入文件目录
        :return:待写入文件路径
        """
        filename = f"Precision_Measure_2E3W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_1e2w1p(filedir):
        """
        接线方式:1e2w1p
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"Precision_Measure_1E2W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    def select_wire_type(self, case_path, sheet_name, wire_type):
        """
        选择接线方式
        :param case_path: 测试文件名
        :param sheet_name: sheet名
        # :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
        :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 4:3e3wD  #AcuRev1320
        :return:
        # 3 element 4 wire Wye       3LN     3CT    6   (0 0)
        # 2.5 element 4 wire Wye     3LN-2.5 3CT    5   (5 0)
        # 3 element 3 wire Delta     3LL     3CT    4   (3 0)
        # 2 element 3 wire Delta     2LL     3CT    3   (2 0)
        # 2 element 3 wire network   2LL     2CT    2   (2 2)
        # 2 element 3 wire 1 phase   1LL     2CT    1   (4 2)
        # 1 element 2 wire           1LN     1CT    0   (1 1)
        """
        if wire_type == 0:
            # 1e2w1p AcuRev1320对应寄存器中值为0
            self.set_wire_type(voltage_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_1e2w1p(save_filedir)
            self.energy_precision_measure_by_1e2w1p(file_path, input_list)
        elif wire_type == 1:
            # 1e2w1p AcuRev1320对应寄存器中值为1
            self.set_wire_type(voltage_wire_value=1)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2e3w1p(save_filedir)
            self.energy_precision_measure_by_2e3w1p(file_path, input_list)
        elif wire_type == 3:
            # 2e3wN AcuRev1320对应寄存器中值为3
            self.set_wire_type(voltage_wire_value=3)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2e3wn(save_filedir)
            self.energy_precision_measure_by_2e3wn(file_path, input_list)
        elif wire_type == 2:
            # 2e3wd AcuRev1320对应寄存器中值为2
            self.set_wire_type(voltage_wire_value=2)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_2e3wd(save_filedir)
            self.energy_precision_measure_by_2e3wd(file_path, input_list)
        elif wire_type == 5:
            # 3e3wd AcuRev1320对应寄存器中值为5
            self.set_wire_type(voltage_wire_value=5)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_3e3wd(save_filedir)
            self.energy_precision_measure_by_3e3wd(file_path, input_list)
        # elif wire_type == 5:
        #     self.set_wire_type(voltage_wire_value=5, current_wire_value=0)
        #     source_input_list = data_read(case_path, sheet_name)
        #     input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
        #     input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
        #     file_path = self.get_save_filepath_of_2p5e4wy(save_filedir)
        #     self.energy_precision_measure_by_2p5e4wy(file_path, input_list)
        else:
            # 3e4wy 对应寄存器中值为4
            self.set_wire_type(voltage_wire_value=4)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_energy_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            file_path = self.get_save_filepath_of_3e4wy(save_filedir)
            self.energy_precision_measure_by_3e4wy(file_path, input_list)


def run_energy_measure_script(test_type, wire_type):
    """
    运行脚本入口
    :param test_type: 0:mV, 1:mA, 2:5A
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
    :return:
    """
    print(f"====================Energy Measure Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    switch_device_screen_interface(inter=0x01)  # 切换至交流界面
    set_gear_switching_mode('00000000')  # 档位切换归零
    energy_measure = EnergyMeasure()

    energy_measure.select_test_case(test_type=test_type, wire_type=wire_type)

    # 关闭ModbusClient客户端连接
    energy_measure.handle_memory.modbus_client.close()
    up_source_ac()
    switch_device_screen_interface(inter=0x00)  # 切换至默认界面
    print(f"====================测试总耗时:{time.time() - start_time}====================")
    print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
    print(f"====================Energy Measure End====================")


if __name__ == '__main__':
    """
    # :param test_type: 0:mV, 1:mA, 2:5A
    # :param wire_type:0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY
    
    :param measure_mode: 0:mV, 1:mA, 2:rct
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 5:3e3wD  # AcuRev1320 项目
    """
    measure_mode = 2
    wire_mode = 6
    run_energy_measure_script(test_type=measure_mode, wire_type=wire_mode)
