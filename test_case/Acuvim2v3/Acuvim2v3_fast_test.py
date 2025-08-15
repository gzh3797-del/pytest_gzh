#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
文件名称:Acuvim2v3_fast_test.py
功能描述:快速测量
创建日期:2025-08-05
作者:王洋
版本:v1.0
修改记录:
"""

import os.path
import time
import math

import xlwt

from Acuvim2v3_modbus_get import *
from openpyxl import Workbook
from test_case.Acuvim2v3.memory_addrs import MemoryAddr
from test_case.Acuvim2v3.test_report_table_heading import REAL_TIME_MEASURE_COLUMNS
from tools.excel_operate import data_read
from tools.log import Log

Log(str(__file__).split("\\")[-1])


def fast_precision_measure_by_2p5E4WY(input_list):
    """
    快速测量:2p5E4WY
    :param input_list: 接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = input_list[i + 1][8]
        uc_p = input_list[i + 1][9]
        ia_p = input_list[i + 1][10]
        ib_p = input_list[i + 1][11]
        ic_p = input_list[i + 1][12]
        ub_value = ub
        ub = ua
        pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        sa = Apparent_Power_calculate(ua, ia)
        sb = Apparent_Power_calculate(ub, ib)
        sc = Apparent_Power_calculate(uc, ia)
        uab, ubc, uca = line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
        # uab = ua
        # ubc = uc
        # uca = uc + ua
        # 需注意ub
        input_values = (
            case_id, ua, ub_value, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, pa, pb, pc, qa, qb, qc, sa, sb,
            sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源+设置电压/电流档位，当前是相同值，只配置了一种A相
        ret = set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        ret = set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub_value, ua, ic, ib, ia, freq)

        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压
        # uab, ubc, uca = line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
        # uab = ua
        # ubc = uc
        # uca = uc + ua
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = sum([pa, pb, pc])
        q_sys = sum([qa, qb, qc])
        s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)
    wb.save(f'{save_filedir}/Precision_Measure_2p5E4WY_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


def fast_precision_measure_by_2E3W1P(input_list):
    """
    快速测试:2E3W1P
    ua/ub/uab/ia/ib/pa/pb/qa/qb/sa/sb/pf_a/pf_b/p_sys 上位机有值
    :param input_list:接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = input_list[i + 1][8]
        uc_p = 0
        ia_p = input_list[i + 1][10]
        ib_p = input_list[i + 1][11]
        ic_p = 0
        pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        sa = Apparent_Power_calculate(ua, ia)
        sb = Apparent_Power_calculate(ub, ib)
        sc = Apparent_Power_calculate(uc, ia)
        input_values = (
            case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, pa, pb, pc, qa, qb, qc, sa, sb, sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源+设置电压/电流档位，当前是相同值，只配置了一种A相
        ret = set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        ret = set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)

        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        # qa_accuracy = get_cmp_accuracy_res(act_accuracy_min=qa_min_accuracy,
        #                                    act_accuracy_max=qa_max_accuracy,
        #                                    act_accuracy_avg=qa_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qa_accuracy_res = (qa_min_value, qa_min_accuracy,
        #                    qa_max_value, qa_max_accuracy,
        #                    qa_avg_value, qa_avg_accuracy,
        #                    qa_accuracy)
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        # qb_accuracy = get_cmp_accuracy_res(act_accuracy_min=qb_min_accuracy,
        #                                    act_accuracy_max=qb_max_accuracy,
        #                                    act_accuracy_avg=qb_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qb_accuracy_res = (qb_min_value, qb_min_accuracy,
        #                    qb_max_value, qb_max_accuracy,
        #                    qb_avg_value, qb_avg_accuracy,
        #                    qb_accuracy)
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        # qc_accuracy = get_cmp_accuracy_res(act_accuracy_min=qc_min_accuracy,
        #                                    act_accuracy_max=qc_max_accuracy,
        #                                    act_accuracy_avg=qc_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qc_accuracy_res = (qc_min_value, qc_min_accuracy,
        #                    qc_max_value, qc_max_accuracy,
        #                    qc_avg_value, qc_avg_accuracy,
        #                    qc_accuracy)
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        # sa_accuracy = get_cmp_accuracy_res(act_accuracy_min=sa_min_accuracy,
        #                                    act_accuracy_max=sa_max_accuracy,
        #                                    act_accuracy_avg=sa_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sa_accuracy_res = (sa_min_value, sa_min_accuracy,
        #                    sa_max_value, sa_max_accuracy,
        #                    sa_avg_value, sa_avg_accuracy,
        #                    sa_accuracy)
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        # sb_accuracy = get_cmp_accuracy_res(act_accuracy_min=sb_min_accuracy,
        #                                    act_accuracy_max=sb_max_accuracy,
        #                                    act_accuracy_avg=sb_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sb_accuracy_res = (sb_min_value, sb_min_accuracy,
        #                    sb_max_value, sb_max_accuracy,
        #                    sb_avg_value, sb_avg_accuracy,
        #                    sb_accuracy)
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        # sc_accuracy = get_cmp_accuracy_res(act_accuracy_min=sc_min_accuracy,
        #                                    act_accuracy_max=sc_max_accuracy,
        #                                    act_accuracy_avg=sc_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sc_accuracy_res = (sc_min_value, sc_min_accuracy,
        #                    sc_max_value, sc_max_accuracy,
        #                    sc_avg_value, sc_avg_accuracy,
        #                    sc_accuracy)
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压
        # uab, ubc, uca = line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
        uab = ua + ub
        ubc = ub
        uca = ua
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = sum([pa, pb, pc])
        q_sys = sum([qa, qb, qc])
        s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)
    wb.save(f'{save_filedir}/Precision_Measure_2E3W1P_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


def fast_precision_measure_by_2E3WN(input_list):
    """
    快速测试:2E3WN
    ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
    ②ua,ub,uc不处理，上位机显示N/A
    ③控源：实际+30，实际计算需-30
    :param input_list:接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        # 源输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = input_list[i + 1][8]
        uc_p = input_list[i + 1][9]
        ia_p = input_list[i + 1][10]
        ib_p = input_list[i + 1][11]
        ic_p = input_list[i + 1][12]

        uab, ubc, uca = line_to_line_voltage_calculate(
            ua=ua,
            ub=ub,
            uc=uc,
            va_angle=ua_p,
            vb_angle=ub_p,
            vc_angle=uc_p
        )
        # pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        # pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        # pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        # qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        # qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        # qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        # sa = Apparent_Power_calculate(ua, ia)
        # sb = Apparent_Power_calculate(ub, ib)
        # sc = Apparent_Power_calculate(uc, ia)
        pa = 0
        pb = 0
        pc = 0
        qa = 0
        qb = 0
        qc = 0
        sa = 0
        sb = 0
        sc = 0
        ib_value = ib
        ib = ic
        p_total = (uab * ia * (math.cos(math.radians(ua_p - ia_p)) * math.cos(math.radians(30)))
                   + ubc * ib * (math.cos(math.radians(uc_p - ic_p))) * math.cos(math.radians(30)))
        q_total = (uab * ia * (math.sin(math.radians(ua_p - ia_p)) * math.cos(math.radians(30)))
                   + ubc * ib * (math.sin(math.radians(uc_p - ic_p)) * math.cos(math.radians(30))))
        s_total = (uab * ia + ubc * ib) * math.cos(math.radians(30))
        print(f"case:{i}, p_total:{p_total}, q_total:{q_total}, s_total:{s_total}")
        logging.info(f"case:{i}, p_total:{p_total}, q_total:{q_total}, s_total:{s_total}")

        input_values = (case_id, ua, ub, uc, ia, ib_value, ic, ua_p, ub_p, uc_p,
                        ia_p, ib_p, ic_p,
                        pa, pb, pc, qa, qb, qc, sa, sb, sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源
        up_source_ac()
        # 设置电压/电流档位，当前是相同值，只配置了一种A相
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        set_ac(
            quc=uc_p,
            qub=ub_p,
            qua=ua_p,
            qic=ic_p,
            qib=ib_p,
            qia=ia_p,
            uc=uc,
            ub=ub,
            ua=ua,
            ic=ic,
            ib=ib_value,
            ia=ia,
            f=freq
        )
        ia_p = ia_p if ((ia_p - 30) >= 0) else (ia_p + 360)
        ib_p = ia_p if ((ib_p - 30) >= 0) else (ib_p + 360)
        ic_p = ic_p if ((ic_p - 30) >= 0) else (ic_p + 360)
        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr,
                                                                                     times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压

        uab = uab
        ubc = ubc
        uca = uca
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = p_total
        q_sys = q_total
        s_sys = s_total
        # p_sys = sum([pa, pb, pc])
        # q_sys = sum([qa, qb, qc])
        # s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)
    wb.save(f'{save_filedir}/Precision_Measure_2E3WN_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


def select_test_case(test_type, wire_type):
    """
    选择测试标准，接线方式
    :param test_type: 0:mV, 1:mA, 2:5A
    :param wire_type:0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
    :return:
    """
    select_test_type(test_type, wire_type)


filepath = r'./test_case/Acuvim2v3/acuvim2v3_test_case_0728.xlsx'

sheet_name_mV = "test_case_mV"
sheet_name_mV_1e2w1p = "test_case_mV-1e2w1p"
sheet_name_mA = "test_case_mA"
sheet_name_5A = "test_case_5A"

sheet_name_test_data = "test_data"

save_filedir = rf"./precision_measure_{time.strftime('%Y%m%d')}"
if not os.path.exists(save_filedir):
    os.makedirs(save_filedir, exist_ok=True)


def select_test_type(test_type, wire_type):
    """
    选择测试数据类型
    :param test_type: 0:mV, 1:mA, 2:5A
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
    :return:
    """
    if test_type == 0:
        select_wire_type(filepath, sheet_name_mV, wire_type)
    elif test_type == 1:
        select_wire_type(filepath, sheet_name_mA, wire_type)
    else:
        select_wire_type(filepath, sheet_name_5A, wire_type)


def set_wire_type(voltage_wire_value, current_wire_value):
    """
    设置接线方式
    :param voltage_wire_value: 电压接线方式
    :param current_wire_value: 电流接线方式
    :return:
    """
    set_service_configuration_by_voltage(addr=MemoryAddr.voltage_wire_addr, value=voltage_wire_value)
    set_service_configuration_by_current(addr=MemoryAddr.current_wire_addr, value=current_wire_value)


def get_input_list_by_wire_type(source_input_list, wire_type):
    """
    获取接线方式的数据
    :param source_input_list: 测试用例数据
    :param wire_type: 接线类型
    :return: 接线方式的数据
    """
    input_list = []
    input_list.append(source_input_list[0])
    if wire_type == 6:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "3E4wY":
                input_list.append(source_input_list[i])
    elif wire_type == 5:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "2.5E4wY":
                input_list.append(source_input_list[i])
    elif wire_type == 4:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "3E3wD":
                input_list.append(source_input_list[i])
    elif wire_type == 3:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "2E3wD":
                input_list.append(source_input_list[i])
    elif wire_type == 2:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "2E3wN":
                input_list.append(source_input_list[i])
    elif wire_type == 1:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "2E3w1p":
                input_list.append(source_input_list[i])
    elif wire_type == 0:
        for i in range(1, len(source_input_list)):
            if source_input_list[i][15] == "1E2w1p":
                input_list.append(source_input_list[i])
    return input_list


def select_wire_type(filepath, sheet_name, wire_type):
    """
    选择接线方式
    :param filepath: 测试文件路径
    :param sheet_name: sheet名
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,
    :return:
    # 3 element 4 wire Wye       3LN     3CT    6
    # 2.5 element 4 wire Wye     3LN-2.5 3CT    5
    # 3 element 3 wire Delta     3LL     3CT    4
    # 2 element 3 wire Delta     2LL     3CT    3
    # 2 element 3 wire network   2LL     2CT    2
    # 2 element 3 wire 1 phase   1LL     2CT    1
    # 1 element 2 wire           1LN     1CT    0
    """
    if wire_type == 0:
        set_wire_type(1, 1)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_1E2W1P(input_list)
    elif wire_type == 1:
        set_wire_type(4, 2)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_2E3W1P(input_list)
    elif wire_type == 2:
        set_wire_type(2, 2)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_2E3WN(input_list)
    elif wire_type == 3:
        set_wire_type(2, 0)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_2E3WD(input_list)
    elif wire_type == 4:
        set_wire_type(3, 0)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_3E3WD(input_list)
    elif wire_type == 5:
        set_wire_type(5, 0)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_2p5E4WY(input_list)
    else:
        set_wire_type(0, 0)
        source_input_list = data_read(filepath, sheet_name)
        input_list = get_input_list_by_wire_type(source_input_list, wire_type)
        fast_precision_measure_by_3E4WY(input_list)


def up_source_ac():
    """
    关源
    :return:
    """
    ret = set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
    return ret


def set_and_switch_source(voltage_gear=120, current_gear=0.5):
    """
    升源加设置档位
    设置电压/电流档位，当前是相同值，只配置了一种A相
    :param voltage_gear: 电压值
    :param current_gear: 电流值
    :return:
    """
    # 设置电压/电流档位，当前是相同值，只配置了一种A相
    set_voltage_gear(voltage_gear, voltage_gear, voltage_gear)
    set_current_gear(current_gear, current_gear, current_gear)


def fast_precision_measure_by_2E3WD(input_list):
    """
    快速测试:2E3WD
    ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
    ②ua,ub,uc不处理，上位机显示N/A
    ③控源：实际+30，实际计算需-30
    :param input_list:接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        # 源输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = input_list[i + 1][8]
        uc_p = input_list[i + 1][9]
        ia_p = input_list[i + 1][10]
        ib_p = input_list[i + 1][11]
        ic_p = input_list[i + 1][12]

        uab, ubc, uca = line_to_line_voltage_calculate(
            ua=ua,
            ub=ub,
            uc=uc,
            va_angle=ua_p,
            vb_angle=ub_p,
            vc_angle=uc_p
        )
        # pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        # pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        # pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        # qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        # qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        # qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        # sa = Apparent_Power_calculate(ua, ia)
        # sb = Apparent_Power_calculate(ub, ib)
        # sc = Apparent_Power_calculate(uc, ia)
        pa = 0
        pb = 0
        pc = 0
        qa = 0
        qb = 0
        qc = 0
        sa = 0
        sb = 0
        sc = 0
        # p_total = uab * ia * (math.cos(math.radians(ua_p - ia_p))) + ubc * ib * (math.cos(math.radians(ub_p - ib_p)))
        # q_total = uab * ia * (math.sin(math.radians(ua_p - ia_p))) + ubc * ib * (math.sin(math.radians(ub_p - ib_p)))
        # s_total = uab * ia + ubc * ib
        p_total = (uab * ia * (math.cos(math.radians(ua_p - ia_p)) * math.cos(math.radians(30)))
                   + ubc * ib * (math.cos(math.radians(uc_p - ic_p))) * math.cos(math.radians(30)))
        q_total = (uab * ia * (math.sin(math.radians(ua_p - ia_p)) * math.cos(math.radians(30)))
                   + ubc * ib * (math.sin(math.radians(uc_p - ic_p)) * math.cos(math.radians(30))))
        s_total = (uab * ia * math.cos(math.radians(30)) + ubc * ib * math.cos(math.radians(30)))
        print(f"case:{i}, p_total:{p_total}, q_total:{q_total}, s_total:{s_total}")
        logging.info(f"case:{i}, p_total:{p_total}, q_total:{q_total}, s_total:{s_total}")

        input_values = (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p,
                        ia_p, ib_p, ic_p,
                        pa, pb, pc, qa, qb, qc, sa, sb, sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源
        up_source_ac()
        # 设置电压/电流档位，当前是相同值，只配置了一种A相
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        set_ac(
            quc=uc_p,
            qub=ub_p,
            qua=ua_p,
            qic=ic_p,
            qib=ib_p,
            qia=ia_p,
            uc=uc,
            ub=ub,
            ua=ua,
            ic=ic,
            ib=ib,
            ia=ia,
            f=freq
        )
        ia_p = ia_p if ((ia_p - 30) >= 0) else (ia_p + 360)
        ib_p = ia_p if ((ib_p - 30) >= 0) else (ib_p + 360)
        ic_p = ic_p if ((ic_p - 30) >= 0) else (ic_p + 360)
        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr,
                                                                                     times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压

        uab = uab
        ubc = ubc
        uca = uca
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = p_total
        q_sys = q_total
        s_sys = s_total
        # p_sys = sum([pa, pb, pc])
        # q_sys = sum([qa, qb, qc])
        # s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)
    wb.save(f'{save_filedir}/Precision_Measure_2E3WD_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


def fast_precision_measure_by_3E3WD(input_list):
    """
    快速测试:3E3WD
    ①uab/ubc/uca/ia/ib/ic/p_sys/q_sys/s_sys/pf_sys 上位机有值
    ②ua,ub,uc不处理，上位机显示N/A
    ③控源：实际+30，实际计算需-30
    :param input_list:接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        # 源输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = input_list[i + 1][8]
        uc_p = input_list[i + 1][9]
        ia_p = input_list[i + 1][10]
        ib_p = input_list[i + 1][11]
        ic_p = input_list[i + 1][12]
        ia_p_input_value = ia_p
        ib_p_input_value = ib_p
        ic_p_input_value = ic_p
        uab, ubc, uca = line_to_line_voltage_calculate(
            ua=ua,
            ub=ub,
            uc=uc,
            va_angle=ua_p,
            vb_angle=ub_p,
            vc_angle=uc_p
        )
        # pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        # pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        # pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        # qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        # qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        # qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        # sa = Apparent_Power_calculate(ua, ia)
        # sb = Apparent_Power_calculate(ub, ib)
        # sc = Apparent_Power_calculate(uc, ia)
        pa = 0
        pb = 0
        pc = 0
        qa = 0
        qb = 0
        qc = 0
        sa = 0
        sb = 0
        sc = 0
        ull_avg = round(sum([uab, ubc, uca]) / len([uab, ubc, uca]), 5)
        ill_avg = round(sum([ia, ib, ic]) / len([ia, ib, ic]), 5)
        s_total = math.sqrt(3) * ull_avg * ill_avg
        p_total = s_total * (math.cos(math.radians(ua_p - ia_p)))
        q_total = s_total * (math.sin(math.radians(ua_p - ia_p)))
        print(f"case:{i}, p_total:{p_total}, q_total:{q_total}, s_total:{s_total}")
        logging.info(f"case:{i}, p_total:{p_total}, q_total:{q_total}, s_total:{s_total}")

        input_values = (case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p,
                        ia_p, ib_p, ic_p,
                        pa, pb, pc, qa, qb, qc, sa, sb, sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源
        up_source_ac()
        # 设置电压/电流档位，当前是相同值，只配置了一种A相
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        set_ac(
            quc=uc_p,
            qub=ub_p,
            qua=ua_p,
            qic=ic_p_input_value,
            qib=ib_p_input_value,
            qia=ia_p_input_value,
            uc=uc,
            ub=ub,
            ua=ua,
            ic=ic,
            ib=ib,
            ia=ia,
            f=freq
        )
        ia_p = ia_p if ((ia_p - 30) >= 0) else (ia_p + 360)
        ib_p = ia_p if ((ib_p - 30) >= 0) else (ib_p + 360)
        ic_p = ic_p if ((ic_p - 30) >= 0) else (ic_p + 360)
        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr,
                                                                                     times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压

        uab = uab
        ubc = ubc
        uca = uca
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = p_total
        q_sys = q_total
        s_sys = s_total
        # p_sys = sum([pa, pb, pc])
        # q_sys = sum([qa, qb, qc])
        # s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)
        # 总测试结果
    wb.save(f'{save_filedir}/Precision_Measure_3E3WD_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


def get_MemoryAddr():
    # 获取寄存器地址
    ua_rms_addr = MemoryAddr.ua_rms_addr
    ub_rms_addr = MemoryAddr.ub_rms_addr
    uc_rms_addr = MemoryAddr.uc_rms_addr
    u_avg_rms_addr = MemoryAddr.u_avg_rms_addr
    ia_rms_addr = MemoryAddr.ia_rms_addr
    ib_rms_addr = MemoryAddr.ib_rms_addr
    ic_rms_addr = MemoryAddr.ic_rms_addr
    i_avg_rms_addr = MemoryAddr.i_avg_rms_addr
    pa_rms_addr = MemoryAddr.pa_rms_addr
    pb_rms_addr = MemoryAddr.pb_rms_addr
    pc_rms_addr = MemoryAddr.pc_rms_addr
    p_total_rms_addr = MemoryAddr.p_total_rms_addr
    qa_rms_addr = MemoryAddr.qa_rms_addr
    qb_rms_addr = MemoryAddr.qb_rms_addr
    qc_rms_addr = MemoryAddr.qc_rms_addr
    q_total_rms_addr = MemoryAddr.q_total_rms_addr
    sa_rms_addr = MemoryAddr.sa_rms_addr
    sb_rms_addr = MemoryAddr.sb_rms_addr
    sc_rms_addr = MemoryAddr.sc_rms_addr
    s_total_rms_addr = MemoryAddr.s_total_rms_addr
    ua_p_rms_addr = MemoryAddr.ua_p_rms_addr
    ub_p_rms_addr = MemoryAddr.ub_p_rms_addr
    uc_p_rms_addr = MemoryAddr.uc_p_rms_addr
    ia_p_rms_addr = MemoryAddr.ia_p_rms_addr
    ib_p_rms_addr = MemoryAddr.ib_p_rms_addr
    ic_p_rms_addr = MemoryAddr.ic_p_rms_addr
    uab_rms_addr = MemoryAddr.uab_rms_addr
    ubc_rms_addr = MemoryAddr.ubc_rms_addr
    uca_rms_addr = MemoryAddr.uca_rms_addr
    ul_avg_rms_addr = MemoryAddr.ul_avg_rms_addr
    pf_a_rms_addr = MemoryAddr.pf_a_rms_addr
    pf_b_rms_addr = MemoryAddr.pf_b_rms_addr
    pf_c_rms_addr = MemoryAddr.pf_c_rms_addr
    pf_total_rms_addr = MemoryAddr.pf_total_rms_addr
    freq_rms_addr = MemoryAddr.freq_rms_addr
    return (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
            ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
            pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
            qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
            sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
            ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
            ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
            uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
            pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
            freq_rms_addr)


def get_cmp_accuracy_res(act_accuracy_min, act_accuracy_max, act_accuracy_avg, exp_accuracy):
    cmp_res = "Failed"
    if (act_accuracy_min <= exp_accuracy) and (act_accuracy_max <= exp_accuracy) and (
            act_accuracy_avg <= exp_accuracy):
        cmp_res = "Passed"
    return cmp_res


def get_all_accuracy_values(all_accuracy_values):
    all_cmp_res = "Passed"
    if "Failed" in all_accuracy_values:
        all_cmp_res = "Failed"
    return all_cmp_res


def fast_precision_measure_by_3E4WY(input_list):
    """
    快速测量:3E4WY
    :param input_list: 接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = input_list[i + 1][8]
        uc_p = input_list[i + 1][9]
        ia_p = input_list[i + 1][10]
        ib_p = input_list[i + 1][11]
        ic_p = input_list[i + 1][12]
        pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        sa = Apparent_Power_calculate(ua, ia)
        sb = Apparent_Power_calculate(ub, ib)
        sc = Apparent_Power_calculate(uc, ia)
        input_values = (
            case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, pa, pb, pc, qa, qb, qc, sa, sb, sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源+设置电压/电流档位，当前是相同值，只配置了一种A相
        ret = set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        ret = set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)

        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        # qa_accuracy = get_cmp_accuracy_res(act_accuracy_min=qa_min_accuracy,
        #                                    act_accuracy_max=qa_max_accuracy,
        #                                    act_accuracy_avg=qa_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qa_accuracy_res = (qa_min_value, qa_min_accuracy,
        #                    qa_max_value, qa_max_accuracy,
        #                    qa_avg_value, qa_avg_accuracy,
        #                    qa_accuracy)
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        # qb_accuracy = get_cmp_accuracy_res(act_accuracy_min=qb_min_accuracy,
        #                                    act_accuracy_max=qb_max_accuracy,
        #                                    act_accuracy_avg=qb_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qb_accuracy_res = (qb_min_value, qb_min_accuracy,
        #                    qb_max_value, qb_max_accuracy,
        #                    qb_avg_value, qb_avg_accuracy,
        #                    qb_accuracy)
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        # qc_accuracy = get_cmp_accuracy_res(act_accuracy_min=qc_min_accuracy,
        #                                    act_accuracy_max=qc_max_accuracy,
        #                                    act_accuracy_avg=qc_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qc_accuracy_res = (qc_min_value, qc_min_accuracy,
        #                    qc_max_value, qc_max_accuracy,
        #                    qc_avg_value, qc_avg_accuracy,
        #                    qc_accuracy)
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        # sa_accuracy = get_cmp_accuracy_res(act_accuracy_min=sa_min_accuracy,
        #                                    act_accuracy_max=sa_max_accuracy,
        #                                    act_accuracy_avg=sa_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sa_accuracy_res = (sa_min_value, sa_min_accuracy,
        #                    sa_max_value, sa_max_accuracy,
        #                    sa_avg_value, sa_avg_accuracy,
        #                    sa_accuracy)
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        # sb_accuracy = get_cmp_accuracy_res(act_accuracy_min=sb_min_accuracy,
        #                                    act_accuracy_max=sb_max_accuracy,
        #                                    act_accuracy_avg=sb_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sb_accuracy_res = (sb_min_value, sb_min_accuracy,
        #                    sb_max_value, sb_max_accuracy,
        #                    sb_avg_value, sb_avg_accuracy,
        #                    sb_accuracy)
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        # sc_accuracy = get_cmp_accuracy_res(act_accuracy_min=sc_min_accuracy,
        #                                    act_accuracy_max=sc_max_accuracy,
        #                                    act_accuracy_avg=sc_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sc_accuracy_res = (sc_min_value, sc_min_accuracy,
        #                    sc_max_value, sc_max_accuracy,
        #                    sc_avg_value, sc_avg_accuracy,
        #                    sc_accuracy)
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压
        uab, ubc, uca = line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
        uab = uab
        ubc = ubc
        uca = uca
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = sum([pa, pb, pc])
        q_sys = sum([qa, qb, qc])
        s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)

        # in_accuracy
        in_value = ic
        in_rms_addr = MemoryAddr.in_rms_addr
        read_current_in_min, read_current_in_max, read_current_in_avg = read_input_current(standard_value=in_value,
                                                                                           address=in_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        in_min_value, in_min_accuracy = read_current_in_min
        in_max_value, in_max_accuracy = read_current_in_max
        in_avg_value, in_avg_accuracy = read_current_in_avg
        in_accuracy = get_cmp_accuracy_res(act_accuracy_min=in_min_accuracy,
                                           act_accuracy_max=in_max_accuracy,
                                           act_accuracy_avg=in_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        in_accuracy_res = (in_value,
                           in_min_value, in_min_accuracy,
                           in_max_value, in_max_accuracy,
                           in_avg_value, in_avg_accuracy,
                           in_accuracy)
        for k in range(len(in_accuracy_res)):
            ws.cell(i + 2, index_value + k, in_accuracy_res[k])
        index_value += len(in_accuracy_res)
    wb.save(f'{save_filedir}/Precision_Measure_3E4WY_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


def fast_precision_measure_by_1E2W1P(input_list):
    """
    快速测试:1E2W1P
    ua/ia/pa/qa/sa/pf_a/p_sys上位机有值
    :param input_list:接线方式的测试数据
    :return:
    """
    # 寄存器全局设置
    pass_res = "Passed"
    fail_res = "Failed"
    (ua_rms_addr, ub_rms_addr, uc_rms_addr, u_avg_rms_addr,
     ia_rms_addr, ib_rms_addr, ic_rms_addr, i_avg_rms_addr,
     pa_rms_addr, pb_rms_addr, pc_rms_addr, p_total_rms_addr,
     qa_rms_addr, qb_rms_addr, qc_rms_addr, q_total_rms_addr,
     sa_rms_addr, sb_rms_addr, sc_rms_addr, s_total_rms_addr,
     ua_p_rms_addr, ub_p_rms_addr, uc_p_rms_addr,
     ia_p_rms_addr, ib_p_rms_addr, ic_p_rms_addr,
     uab_rms_addr, ubc_rms_addr, uca_rms_addr, ul_avg_rms_addr,
     pf_a_rms_addr, pf_b_rms_addr, pf_c_rms_addr, pf_total_rms_addr,
     freq_rms_addr) = get_MemoryAddr()

    # 创建新工作簿
    wb = Workbook()
    # 获取活动工作表
    ws = wb.active
    for i in range(len(REAL_TIME_MEASURE_COLUMNS)):
        j = i + 1
        ws.cell(1, j, f'{REAL_TIME_MEASURE_COLUMNS[i]}')
    for i in range(len(input_list)):
        if i == len(input_list) - 1:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            break
        elif i == 0:
            logging.info('测试进度:{}'.format(input_list[i]))
            print('测试进度:{}'.format(input_list[i]))
        else:
            logging.info('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
            print('测试进度:{},执行时间:{}'.format(input_list[i], time.strftime('%Y_%m_%d %H:%M:%S')))
        # 读取设备精度
        index_value = 1
        freq = input_list[i + 1][16]
        voltage_accuracy = input_list[i + 1][17]
        current_accuracy = input_list[i + 1][18]
        phase_accuracy = input_list[i + 1][19]
        if isinstance(phase_accuracy, float):
            voltage_phase_accuracy = phase_accuracy
            current_phase_accuracy = phase_accuracy
        else:
            voltage_phase_accuracy = 999999
            current_phase_accuracy = 999999

        power_accuracy = input_list[i + 1][20]
        if isinstance(phase_accuracy, float):
            active_power_accuracy = power_accuracy
        else:
            active_power_accuracy = 999999
        reactive_power_accuracy = input_list[i + 1][21]
        apparent_power_accuracy = input_list[i + 1][22]
        cyc_cnt = input_list[i + 1][23]
        interval_time = input_list[i + 1][24]

        # 计算case编号/电压/电流/电压相位角/电流相位角输入值
        case_id = input_list[i + 1][0]
        ua = input_list[i + 1][1]
        ub = input_list[i + 1][2]
        uc = input_list[i + 1][3]
        ia = input_list[i + 1][4]
        ib = input_list[i + 1][5]
        ic = input_list[i + 1][6]
        ua_p = input_list[i + 1][7]
        ub_p = 0
        uc_p = 0
        ia_p = input_list[i + 1][10]
        ib_p = 0
        ic_p = 0

        pa = Active_Power_calculate(ua, ia, ua_p, ia_p)
        pb = Active_Power_calculate(ub, ib, ub_p, ib_p)
        pc = Active_Power_calculate(uc, ia, uc_p, ic_p)
        qa = Reactive_Power_calculate(ua, ia, ua_p, ia_p)
        qb = Reactive_Power_calculate(ub, ib, ub_p, ib_p)
        qc = Reactive_Power_calculate(uc, ia, uc_p, ic_p)
        sa = Apparent_Power_calculate(ua, ia)
        sb = Apparent_Power_calculate(ub, ib)
        sc = Apparent_Power_calculate(uc, ia)
        input_values = (
            case_id, ua, ub, uc, ia, ib, ic, ua_p, ub_p, uc_p, ia_p, ib_p, ic_p, pa, pb, pc, qa, qb, qc, sa, sb, sc)
        # 写入case编号/电压/电流/电压相位角/电流相位角输入值
        for k in range(len(input_values)):
            ws.cell(i + 2, index_value + k, input_values[k])
        index_value += len(input_values)
        # 升源+设置电压/电流档位，当前是相同值，只配置了一种A相
        ret = set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        set_and_switch_source(voltage_gear=ua, current_gear=ia)
        ret = set_ac(uc_p, ub_p, ua_p, ic_p, ib_p, ia_p, uc, ub, ua, ic, ib, ia, freq)

        # 计算电压最小/最大/平均值，及精度是否符合要求
        # ua_accuracy
        read_phase_ua_min, read_phase_ua_max, read_phase_ua_avg = read_phase_voltage(standard_value=ua,
                                                                                     address=ua_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ua_min_value, ua_min_accuracy = read_phase_ua_min
        ua_max_value, ua_max_accuracy = read_phase_ua_max
        ua_avg_value, ua_avg_accuracy = read_phase_ua_avg
        ua_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_min_accuracy,
                                           act_accuracy_max=ua_max_accuracy,
                                           act_accuracy_avg=ua_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ua_accuracy_res = (ua_min_value, ua_min_accuracy,
                           ua_max_value, ua_max_accuracy,
                           ua_avg_value, ua_avg_accuracy,
                           ua_accuracy)
        for k in range(len(ua_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_accuracy_res[k])
        index_value += len(ua_accuracy_res)
        # ub_accuracy
        read_phase_ub_min, read_phase_ub_max, read_phase_ub_avg = read_phase_voltage(standard_value=ub,
                                                                                     address=ub_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        ub_min_value, ub_min_accuracy = read_phase_ub_min
        ub_max_value, ub_max_accuracy = read_phase_ub_max
        ub_avg_value, ub_avg_accuracy = read_phase_ub_avg
        ub_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_min_accuracy,
                                           act_accuracy_max=ub_max_accuracy,
                                           act_accuracy_avg=ub_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        ub_accuracy_res = (ub_min_value, ub_min_accuracy,
                           ub_max_value, ub_max_accuracy,
                           ub_avg_value, ub_avg_accuracy,
                           ub_accuracy)
        for k in range(len(ub_accuracy_res)):
            ws.cell(i + 2, index_value + k, ub_accuracy_res[k])
        index_value += len(ub_accuracy_res)
        # uc_accuracy
        read_phase_uc_min, read_phase_uc_max, read_phase_uc_avg = read_phase_voltage(standard_value=uc,
                                                                                     address=uc_rms_addr, times=cyc_cnt,
                                                                                     interval_time=interval_time)
        uc_min_value, uc_min_accuracy = read_phase_uc_min
        uc_max_value, uc_max_accuracy = read_phase_uc_max
        uc_avg_value, uc_avg_accuracy = read_phase_uc_avg
        uc_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_min_accuracy,
                                           act_accuracy_max=uc_max_accuracy,
                                           act_accuracy_avg=uc_avg_accuracy,
                                           exp_accuracy=voltage_accuracy)
        uc_accuracy_res = (uc_min_value, uc_min_accuracy,
                           uc_max_value, uc_max_accuracy,
                           uc_avg_value, uc_avg_accuracy,
                           uc_accuracy)
        for k in range(len(uc_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_accuracy_res[k])
        index_value += len(uc_accuracy_res)

        # 计算电流最小/最大/平均值，及精度是否符合要求
        # ia_accuracy
        read_current_ia_min, read_current_ia_max, read_current_ia_avg = read_input_current(standard_value=ia,
                                                                                           address=ia_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ia_min_value, ia_min_accuracy = read_current_ia_min
        ia_max_value, ia_max_accuracy = read_current_ia_max
        ia_avg_value, ia_avg_accuracy = read_current_ia_avg
        ia_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_min_accuracy,
                                           act_accuracy_max=ia_max_accuracy,
                                           act_accuracy_avg=ia_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ia_accuracy_res = (ia_min_value, ia_min_accuracy,
                           ia_max_value, ia_max_accuracy,
                           ia_avg_value, ia_avg_accuracy,
                           ia_accuracy)
        for k in range(len(ia_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_accuracy_res[k])
        index_value += len(ia_accuracy_res)
        # ib_accuracy
        read_current_ib_min, read_current_ib_max, read_current_ib_avg = read_input_current(standard_value=ib,
                                                                                           address=ib_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ib_min_value, ib_min_accuracy = read_current_ib_min
        ib_max_value, ib_max_accuracy = read_current_ib_max
        ib_avg_value, ib_avg_accuracy = read_current_ib_avg
        ib_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_min_accuracy,
                                           act_accuracy_max=ib_max_accuracy,
                                           act_accuracy_avg=ib_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ib_accuracy_res = (ib_min_value, ib_min_accuracy,
                           ib_max_value, ib_max_accuracy,
                           ib_avg_value, ib_avg_accuracy,
                           ib_accuracy)
        for k in range(len(ib_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_accuracy_res[k])
        index_value += len(ib_accuracy_res)
        # ic_accuracy
        read_current_ic_min, read_current_ic_max, read_current_ic_avg = read_input_current(standard_value=ic,
                                                                                           address=ic_rms_addr,
                                                                                           times=cyc_cnt,
                                                                                           interval_time=interval_time)
        ic_min_value, ic_min_accuracy = read_current_ic_min
        ic_max_value, ic_max_accuracy = read_current_ic_max
        ic_avg_value, ic_avg_accuracy = read_current_ic_avg
        ic_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_min_accuracy,
                                           act_accuracy_max=ic_max_accuracy,
                                           act_accuracy_avg=ic_avg_accuracy,
                                           exp_accuracy=current_accuracy)
        ic_accuracy_res = (ic_min_value, ic_min_accuracy,
                           ic_max_value, ic_max_accuracy,
                           ic_avg_value, ic_avg_accuracy,
                           ic_accuracy)
        for k in range(len(ic_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_accuracy_res[k])
        index_value += len(ic_accuracy_res)

        # 计算电压相位角最小/最大/平均值，及精度是否符合要求
        # ua_p
        read_voltage_angle_ua_p_min, read_voltage_angle_ua_p_max, read_voltage_angle_ua_p_avg = (
            read_phase_voltage_angle(standard_value=ua_p, address=ua_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ua_p_min_value, ua_p_min_accuracy = read_voltage_angle_ua_p_min
        ua_p_max_value, ua_p_max_accuracy = read_voltage_angle_ua_p_max
        ua_p_avg_value, ua_p_avg_accuracy = read_voltage_angle_ua_p_avg
        ua_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ua_p_min_accuracy,
                                             act_accuracy_max=ua_p_max_accuracy,
                                             act_accuracy_avg=ua_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ua_p_min_value, ua_p_min_accuracy,
                             ua_p_max_value, ua_p_max_accuracy,
                             ua_p_avg_value, ua_p_avg_accuracy,
                             ua_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # ub_p
        read_voltage_angle_ub_p_min, read_voltage_angle_ub_p_max, read_voltage_angle_ub_p_avg = (
            read_phase_voltage_angle(standard_value=ub_p, address=ub_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ub_p_min_value, ub_p_min_accuracy = read_voltage_angle_ub_p_min
        ub_p_max_value, ub_p_max_accuracy = read_voltage_angle_ub_p_max
        ub_p_avg_value, ub_p_avg_accuracy = read_voltage_angle_ub_p_avg
        ub_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ub_p_min_accuracy,
                                             act_accuracy_max=ub_p_max_accuracy,
                                             act_accuracy_avg=ub_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        ua_p_accuracy_res = (ub_p_min_value, ub_p_min_accuracy,
                             ub_p_max_value, ub_p_max_accuracy,
                             ub_p_avg_value, ub_p_avg_accuracy,
                             ub_p_accuracy)
        for k in range(len(ua_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ua_p_accuracy_res[k])
        index_value += len(ua_p_accuracy_res)
        # uc_p
        read_voltage_angle_uc_p_min, read_voltage_angle_uc_p_max, read_voltage_angle_uc_p_avg = (
            read_phase_voltage_angle(standard_value=uc_p, address=uc_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        uc_p_min_value, uc_p_min_accuracy = read_voltage_angle_uc_p_min
        uc_p_max_value, uc_p_max_accuracy = read_voltage_angle_uc_p_max
        uc_p_avg_value, uc_p_avg_accuracy = read_voltage_angle_uc_p_avg
        uc_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=uc_p_min_accuracy,
                                             act_accuracy_max=uc_p_max_accuracy,
                                             act_accuracy_avg=uc_p_avg_accuracy,
                                             exp_accuracy=voltage_phase_accuracy)
        uc_p_accuracy_res = (uc_p_min_value, uc_p_min_accuracy,
                             uc_p_max_value, uc_p_max_accuracy,
                             uc_p_avg_value, uc_p_avg_accuracy,
                             uc_p_accuracy)
        for k in range(len(uc_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, uc_p_accuracy_res[k])
        index_value += len(uc_p_accuracy_res)

        # 计算电流相位角最小/最大/平均值，及精度是否符合要求
        # ia_p_accuracy
        read_current_angle_ia_p_min, read_current_angle_ia_p_max, read_current_angle_ia_p_avg = (
            read_input_current_angle(standard_value=ia_p, address=ia_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ia_p_min_value, ia_p_min_accuracy = read_current_angle_ia_p_min
        ia_p_max_value, ia_p_max_accuracy = read_current_angle_ia_p_max
        ia_p_avg_value, ia_p_avg_accuracy = read_current_angle_ia_p_avg
        ia_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ia_p_min_accuracy,
                                             act_accuracy_max=ia_p_max_accuracy,
                                             act_accuracy_avg=ia_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ia_p_accuracy_res = (ia_p_min_value, ia_p_min_accuracy,
                             ia_p_max_value, ia_p_max_accuracy,
                             ia_p_avg_value, ia_p_avg_accuracy,
                             ia_p_accuracy)
        for k in range(len(ia_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ia_p_accuracy_res[k])
        index_value += len(ia_p_accuracy_res)
        # ib_p_accuracy
        read_current_angle_ib_p_min, read_current_angle_ib_p_max, read_current_angle_ib_p_avg = (
            read_input_current_angle(standard_value=ib_p, address=ib_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ib_p_min_value, ib_p_min_accuracy = read_current_angle_ib_p_min
        ib_p_max_value, ib_p_max_accuracy = read_current_angle_ib_p_max
        ib_p_avg_value, ib_p_avg_accuracy = read_current_angle_ib_p_avg
        ib_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ib_p_min_accuracy,
                                             act_accuracy_max=ib_p_max_accuracy,
                                             act_accuracy_avg=ib_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ib_p_accuracy_res = (ib_p_min_value, ib_p_min_accuracy,
                             ib_p_max_value, ib_p_max_accuracy,
                             ib_p_avg_value, ib_p_avg_accuracy,
                             ib_p_accuracy)
        for k in range(len(ib_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ib_p_accuracy_res[k])
        index_value += len(ib_p_accuracy_res)
        # ic_p_accuracy
        read_current_angle_ic_p_min, read_current_angle_ic_p_max, read_current_angle_ic_p_avg = (
            read_input_current_angle(standard_value=ic_p, address=ic_p_rms_addr, times=cyc_cnt,
                                     interval_time=interval_time))
        ic_p_min_value, ic_p_min_accuracy = read_current_angle_ic_p_min
        ic_p_max_value, ic_p_max_accuracy = read_current_angle_ic_p_max
        ic_p_avg_value, ic_p_avg_accuracy = read_current_angle_ic_p_avg
        ic_p_accuracy = get_cmp_accuracy_res(act_accuracy_min=ic_p_min_accuracy,
                                             act_accuracy_max=ic_p_max_accuracy,
                                             act_accuracy_avg=ic_p_avg_accuracy,
                                             exp_accuracy=current_phase_accuracy)
        ic_p_accuracy_res = (ic_p_min_value, ic_p_min_accuracy,
                             ic_p_max_value, ic_p_max_accuracy,
                             ic_p_avg_value, ic_p_avg_accuracy,
                             ic_p_accuracy)
        for k in range(len(ic_p_accuracy_res)):
            ws.cell(i + 2, index_value + k, ic_p_accuracy_res[k])
        index_value += len(ic_p_accuracy_res)

        # 计算有功功率相位角最小/最大/平均值，及精度是否符合要求
        # pa
        read_active_power_pa_min, read_active_power_pa_max, read_active_power_pa_avg = (
            read_input_active_power(standard_value=pa, address=pa_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pa_min_value, pa_min_accuracy = read_active_power_pa_min
        pa_max_value, pa_max_accuracy = read_active_power_pa_max
        pa_avg_value, pa_avg_accuracy = read_active_power_pa_avg
        pa_accuracy = get_cmp_accuracy_res(act_accuracy_min=pa_min_accuracy,
                                           act_accuracy_max=pa_max_accuracy,
                                           act_accuracy_avg=pa_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pa_accuracy_res = (pa_min_value, pa_min_accuracy,
                           pa_max_value, pa_max_accuracy,
                           pa_avg_value, pa_avg_accuracy,
                           pa_accuracy)
        for k in range(len(pa_accuracy_res)):
            ws.cell(i + 2, index_value + k, pa_accuracy_res[k])
        index_value += len(pa_accuracy_res)
        # pb
        read_active_power_pb_min, read_active_power_pb_max, read_active_power_pb_avg = (
            read_input_active_power(standard_value=pb, address=pb_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pb_min_value, pb_min_accuracy = read_active_power_pb_min
        pb_max_value, pb_max_accuracy = read_active_power_pb_max
        pb_avg_value, pb_avg_accuracy = read_active_power_pb_avg
        pb_accuracy = get_cmp_accuracy_res(act_accuracy_min=pb_min_accuracy,
                                           act_accuracy_max=pb_max_accuracy,
                                           act_accuracy_avg=pb_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pb_accuracy_res = (pb_min_value, pb_min_accuracy,
                           pb_max_value, pb_max_accuracy,
                           pb_avg_value, pb_avg_accuracy,
                           pb_accuracy)
        for k in range(len(pb_accuracy_res)):
            ws.cell(i + 2, index_value + k, pb_accuracy_res[k])
        index_value += len(pb_accuracy_res)
        # pc
        read_active_power_pc_min, read_active_power_pc_max, read_active_power_pc_avg = (
            read_input_active_power(standard_value=pc, address=pc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        pc_min_value, pc_min_accuracy = read_active_power_pc_min
        pc_max_value, pc_max_accuracy = read_active_power_pc_max
        pc_avg_value, pc_avg_accuracy = read_active_power_pc_avg
        pc_accuracy = get_cmp_accuracy_res(act_accuracy_min=pc_min_accuracy,
                                           act_accuracy_max=pc_max_accuracy,
                                           act_accuracy_avg=pc_avg_accuracy,
                                           exp_accuracy=active_power_accuracy)
        pc_accuracy_res = (pc_min_value, pc_min_accuracy,
                           pc_max_value, pc_max_accuracy,
                           pc_avg_value, pc_avg_accuracy,
                           pc_accuracy)
        for k in range(len(pc_accuracy_res)):
            ws.cell(i + 2, index_value + k, pc_accuracy_res[k])
        index_value += len(pc_accuracy_res)

        # 计算无功功率相位角最小/最大/平均值，及精度是否符合要求
        # qa
        read_reactive_power_qa_min, read_reactive_power_qa_max, read_reactive_power_qa_avg = (
            read_input_reactive_power(standard_value=qa, address=qa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qa_min_value, qa_min_accuracy = read_reactive_power_qa_min
        qa_max_value, qa_max_accuracy = read_reactive_power_qa_max
        qa_avg_value, qa_avg_accuracy = read_reactive_power_qa_avg
        qa_accuracy_res = (
            qa_min_value,
            qa_min_accuracy,
            qa_max_value,
            qa_max_accuracy,
            qa_avg_value,
            qa_avg_accuracy
        )
        # qa_accuracy = get_cmp_accuracy_res(act_accuracy_min=qa_min_accuracy,
        #                                    act_accuracy_max=qa_max_accuracy,
        #                                    act_accuracy_avg=qa_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qa_accuracy_res = (qa_min_value, qa_min_accuracy,
        #                    qa_max_value, qa_max_accuracy,
        #                    qa_avg_value, qa_avg_accuracy,
        #                    qa_accuracy)
        for k in range(len(qa_accuracy_res)):
            ws.cell(i + 2, index_value + k, qa_accuracy_res[k])
        index_value += len(qa_accuracy_res)
        # qb
        read_reactive_power_qb_min, read_reactive_power_qb_max, read_reactive_power_qb_avg = (
            read_input_reactive_power(standard_value=qb, address=qb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qb_min_value, qb_min_accuracy = read_reactive_power_qb_min
        qb_max_value, qb_max_accuracy = read_reactive_power_qb_max
        qb_avg_value, qb_avg_accuracy = read_reactive_power_qb_avg
        qb_accuracy_res = (
            qb_min_value,
            qb_min_accuracy,
            qb_max_value,
            qb_max_accuracy,
            qb_avg_value,
            qb_avg_accuracy
        )
        # qb_accuracy = get_cmp_accuracy_res(act_accuracy_min=qb_min_accuracy,
        #                                    act_accuracy_max=qb_max_accuracy,
        #                                    act_accuracy_avg=qb_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qb_accuracy_res = (qb_min_value, qb_min_accuracy,
        #                    qb_max_value, qb_max_accuracy,
        #                    qb_avg_value, qb_avg_accuracy,
        #                    qb_accuracy)
        for k in range(len(qb_accuracy_res)):
            ws.cell(i + 2, index_value + k, qb_accuracy_res[k])
        index_value += len(qb_accuracy_res)
        # qc
        read_reactive_power_qc_min, read_reactive_power_qc_max, read_reactive_power_qc_avg = (
            read_input_reactive_power(standard_value=qc, address=qc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        qc_min_value, qc_min_accuracy = read_reactive_power_qc_min
        qc_max_value, qc_max_accuracy = read_reactive_power_qc_max
        qc_avg_value, qc_avg_accuracy = read_reactive_power_qc_avg
        qc_accuracy_res = (
            qc_min_value,
            qc_min_accuracy,
            qc_max_value,
            qc_max_accuracy,
            qc_avg_value,
            qc_avg_accuracy
        )
        # qc_accuracy = get_cmp_accuracy_res(act_accuracy_min=qc_min_accuracy,
        #                                    act_accuracy_max=qc_max_accuracy,
        #                                    act_accuracy_avg=qc_avg_accuracy,
        #                                    exp_accuracy=reactive_power_accuracy)
        # qc_accuracy_res = (qc_min_value, qc_min_accuracy,
        #                    qc_max_value, qc_max_accuracy,
        #                    qc_avg_value, qc_avg_accuracy,
        #                    qc_accuracy)
        for k in range(len(qc_accuracy_res)):
            ws.cell(i + 2, index_value + k, qc_accuracy_res[k])
        index_value += len(qc_accuracy_res)
        # 计算视在功率相位角最小/最大/平均值，及精度是否符合要求
        # sa
        read_apparent_power_sa_min, read_apparent_power_sa_max, read_apparent_power_sa_avg = (
            read_input_apparent_power(standard_value=sa, address=sa_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sa_min_value, sa_min_accuracy = read_apparent_power_sa_min
        sa_max_value, sa_max_accuracy = read_apparent_power_sa_max
        sa_avg_value, sa_avg_accuracy = read_apparent_power_sa_avg
        sa_accuracy_res = (
            sa_min_value,
            sa_min_accuracy,
            sa_max_value,
            sa_max_accuracy,
            sa_avg_value,
            sa_avg_accuracy
        )
        # sa_accuracy = get_cmp_accuracy_res(act_accuracy_min=sa_min_accuracy,
        #                                    act_accuracy_max=sa_max_accuracy,
        #                                    act_accuracy_avg=sa_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sa_accuracy_res = (sa_min_value, sa_min_accuracy,
        #                    sa_max_value, sa_max_accuracy,
        #                    sa_avg_value, sa_avg_accuracy,
        #                    sa_accuracy)
        for k in range(len(sa_accuracy_res)):
            ws.cell(i + 2, index_value + k, sa_accuracy_res[k])
        index_value += len(sa_accuracy_res)
        # sb
        read_apparent_power_sb_min, read_apparent_power_sb_max, read_apparent_power_sb_avg = (
            read_input_apparent_power(standard_value=sb, address=sb_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sb_min_value, sb_min_accuracy = read_apparent_power_sb_min
        sb_max_value, sb_max_accuracy = read_apparent_power_sb_max
        sb_avg_value, sb_avg_accuracy = read_apparent_power_sb_avg
        sb_accuracy_res = (
            sb_min_value,
            sb_min_accuracy,
            sb_max_value,
            sb_max_accuracy,
            sb_avg_value,
            sb_avg_accuracy
        )
        # sb_accuracy = get_cmp_accuracy_res(act_accuracy_min=sb_min_accuracy,
        #                                    act_accuracy_max=sb_max_accuracy,
        #                                    act_accuracy_avg=sb_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sb_accuracy_res = (sb_min_value, sb_min_accuracy,
        #                    sb_max_value, sb_max_accuracy,
        #                    sb_avg_value, sb_avg_accuracy,
        #                    sb_accuracy)
        for k in range(len(sb_accuracy_res)):
            ws.cell(i + 2, index_value + k, sb_accuracy_res[k])
        index_value += len(sb_accuracy_res)
        # sc
        read_apparent_power_sc_min, read_apparent_power_sc_max, read_apparent_power_sc_avg = (
            read_input_apparent_power(standard_value=sc, address=sc_rms_addr, times=cyc_cnt,
                                      interval_time=interval_time))
        sc_min_value, sc_min_accuracy = read_apparent_power_sc_min
        sc_max_value, sc_max_accuracy = read_apparent_power_sc_max
        sc_avg_value, sc_avg_accuracy = read_apparent_power_sc_avg
        sc_accuracy_res = (
            sc_min_value,
            sc_min_accuracy,
            sc_max_value,
            sc_max_accuracy,
            sc_avg_value,
            sc_avg_accuracy
        )
        # sc_accuracy = get_cmp_accuracy_res(act_accuracy_min=sc_min_accuracy,
        #                                    act_accuracy_max=sc_max_accuracy,
        #                                    act_accuracy_avg=sc_avg_accuracy,
        #                                    exp_accuracy=apparent_power_accuracy)
        # sc_accuracy_res = (sc_min_value, sc_min_accuracy,
        #                    sc_max_value, sc_max_accuracy,
        #                    sc_avg_value, sc_avg_accuracy,
        #                    sc_accuracy)
        for k in range(len(sc_accuracy_res)):
            ws.cell(i + 2, index_value + k, sc_accuracy_res[k])
        index_value += len(sc_accuracy_res)

        # 计算线电压
        # uab, ubc, uca = line_to_line_voltage_calculate(ua, ub, uc, ua_p, ub_p, uc_p)
        uab = ua
        ubc = 0
        uca = ua
        # uab
        read_line_voltage_uab_min, read_line_voltage_uab_max, read_line_voltage_uab_avg = (
            read_line_voltage(standard_value=uab, address=uab_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uab_min_value, uab_min_accuracy = read_line_voltage_uab_min
        uab_max_value, uab_max_accuracy = read_line_voltage_uab_max
        uab_avg_value, uab_avg_accuracy = read_line_voltage_uab_avg
        uab_accuracy = get_cmp_accuracy_res(act_accuracy_min=uab_min_accuracy,
                                            act_accuracy_max=uab_max_accuracy,
                                            act_accuracy_avg=uab_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uab_accuracy_res = (uab,
                            uab_min_value, uab_min_accuracy,
                            uab_max_value, uab_max_accuracy,
                            uab_avg_value, uab_avg_accuracy,
                            uab_accuracy)
        for k in range(len(uab_accuracy_res)):
            ws.cell(i + 2, index_value + k, uab_accuracy_res[k])
        index_value += len(uab_accuracy_res)
        # ubc
        read_line_voltage_ubc_min, read_line_voltage_ubc_max, read_line_voltage_ubc_avg = (
            read_line_voltage(standard_value=ubc, address=ubc_rms_addr, times=cyc_cnt, interval_time=interval_time))
        ubc_min_value, ubc_min_accuracy = read_line_voltage_ubc_min
        ubc_max_value, ubc_max_accuracy = read_line_voltage_ubc_max
        ubc_avg_value, ubc_avg_accuracy = read_line_voltage_ubc_avg
        ubc_accuracy = get_cmp_accuracy_res(act_accuracy_min=ubc_min_accuracy,
                                            act_accuracy_max=ubc_max_accuracy,
                                            act_accuracy_avg=ubc_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        ubc_accuracy_res = (ubc,
                            ubc_min_value, ubc_min_accuracy,
                            ubc_max_value, ubc_max_accuracy,
                            ubc_avg_value, ubc_avg_accuracy,
                            ubc_accuracy)
        for k in range(len(ubc_accuracy_res)):
            ws.cell(i + 2, index_value + k, ubc_accuracy_res[k])
        index_value += len(ubc_accuracy_res)
        # uca
        read_line_voltage_uca_min, read_line_voltage_uca_max, read_line_voltage_uca_avg = (
            read_line_voltage(standard_value=uca, address=uca_rms_addr, times=cyc_cnt, interval_time=interval_time))
        uca_min_value, uca_min_accuracy = read_line_voltage_uca_min
        uca_max_value, uca_max_accuracy = read_line_voltage_uca_max
        uca_avg_value, uca_avg_accuracy = read_line_voltage_uca_avg
        uca_accuracy = get_cmp_accuracy_res(act_accuracy_min=uca_min_accuracy,
                                            act_accuracy_max=uca_max_accuracy,
                                            act_accuracy_avg=uca_avg_accuracy,
                                            exp_accuracy=voltage_phase_accuracy)
        uca_accuracy_res = (uca,
                            uca_min_value, uca_min_accuracy,
                            uca_max_value, uca_max_accuracy,
                            uca_avg_value, uca_avg_accuracy,
                            uca_accuracy)
        for k in range(len(uca_accuracy_res)):
            ws.cell(i + 2, index_value + k, uca_accuracy_res[k])
        index_value += len(uca_accuracy_res)

        # 计算系统有功,无功，视在功率
        p_sys = sum([pa, pb, pc])
        q_sys = sum([qa, qb, qc])
        s_sys = sum([sa, sb, sc])
        # p_abc
        read_sys_active_power_min, read_sys_active_power_max, read_sys_active_power_avg = (
            read_input_sys_power(standard_value=p_sys, address=p_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        p_sys_min_value, p_sys_min_accuracy = read_sys_active_power_min
        p_sys_max_value, p_sys_max_accuracy = read_sys_active_power_max
        p_sys_avg_value, p_sys_avg_accuracy = read_sys_active_power_avg
        p_sys_accuracy = get_cmp_accuracy_res(act_accuracy_min=p_sys_min_accuracy,
                                              act_accuracy_max=p_sys_max_accuracy,
                                              act_accuracy_avg=p_sys_avg_accuracy,
                                              exp_accuracy=active_power_accuracy)
        p_sys_accuracy_res = (p_sys,
                              p_sys_min_value, p_sys_min_accuracy,
                              p_sys_max_value, p_sys_max_accuracy,
                              p_sys_avg_value, p_sys_avg_accuracy,
                              p_sys_accuracy)
        for k in range(len(p_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, p_sys_accuracy_res[k])
        index_value += len(p_sys_accuracy_res)
        # q_abc
        read_sys_reactive_power_min, read_sys_reactive_power_max, read_sys_reactive_power_avg = (
            read_input_sys_power(standard_value=q_sys, address=q_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        q_sys_min_value, q_sys_min_accuracy = read_sys_reactive_power_min
        q_sys_max_value, q_sys_max_accuracy = read_sys_reactive_power_max
        q_sys_avg_value, q_sys_avg_accuracy = read_sys_reactive_power_avg
        q_sys_accuracy_res = (
            q_sys,
            q_sys_min_value,
            q_sys_min_accuracy,
            q_sys_max_value,
            q_sys_max_accuracy,
            q_sys_avg_value,
            q_sys_avg_accuracy
        )
        for k in range(len(q_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, q_sys_accuracy_res[k])
        index_value += len(q_sys_accuracy_res)
        # s_abc
        read_sys_apparent_power_min, read_sys_apparent_power_max, read_sys_apparent_power_avg = (
            read_input_sys_power(standard_value=s_sys, address=s_total_rms_addr, times=cyc_cnt,
                                 interval_time=interval_time))
        s_sys_min_value, s_sys_min_accuracy = read_sys_apparent_power_min
        s_sys_max_value, s_sys_max_accuracy = read_sys_apparent_power_max
        s_sys_avg_value, s_sys_avg_accuracy = read_sys_apparent_power_avg
        s_sys_accuracy_res = (
            s_sys,
            s_sys_min_value,
            s_sys_min_accuracy,
            s_sys_max_value,
            s_sys_max_accuracy,
            s_sys_avg_value,
            s_sys_avg_accuracy
        )
        for k in range(len(s_sys_accuracy_res)):
            ws.cell(i + 2, index_value + k, s_sys_accuracy_res[k])
        index_value += len(s_sys_accuracy_res)
    wb.save(f'{save_filedir}/Precision_Measure_1E2W1P_{time.strftime("%Y%m%d%H%M%S")}.xlsx')


if __name__ == '__main__':
    print(f"====================Precision Measure Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    switch_device_screen_interface(inter=0x01)  # 切换至交流界面
    set_gear_switching_mode('00000000')  # 档位切换归零
    time.sleep(5)

    # mV测量
    # select_test_case(test_type=0, wire_type=6)

    # mA测量
    # select_test_case(test_type=1, wire_type=6)

    # 5A测量
    select_test_case(test_type=2, wire_type=2)

    # 关闭ModbusClient客户端连接
    ModbusClient.close()
    up_source_ac()
    switch_device_screen_interface(inter=0x00)  # 切换至默认界面
    print(f"====================测试总耗时:{time.time() - start_time}====================")
    print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
    print(f"====================Precision Measure End====================")
