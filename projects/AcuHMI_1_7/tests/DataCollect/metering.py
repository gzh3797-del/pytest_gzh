# -*- coding: utf-8 -*-
"""
Metering 数据采集 + 寄存器匹配 + Modbus 比对（三合一）
支持设备：AcuRev-4100/4110、AcuRev-2100、Acuvim3、AcuvimIIW

运行：
  cd C:\\JrJ\\auto\\autotest\\Protocols
  python DataCollect/metering.py              # 完整三步
  python DataCollect/metering.py --collect    # 仅采集
  python DataCollect/metering.py --match      # 仅匹配
  python DataCollect/metering.py --compare    # 仅比对

输出：
  reports/metering_collect.json
  reports/metering_register_match.csv / .md
  reports/metering_compare.xlsx
"""
import argparse, csv as _csv, json, pathlib, re, struct, sys, zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from playwright.sync_api import sync_playwright, Page, TimeoutError as PWTimeout
from pymodbus.client import ModbusTcpClient

# 让本文件单独运行时也能 import 上级 config.py
_PROTO_DIR = pathlib.Path(__file__).resolve().parent.parent
if str(_PROTO_DIR) not in sys.path:
    sys.path.insert(0, str(_PROTO_DIR))
try:
    import config as _cfg
except Exception:
    _cfg = None

# ╔══════════════════════════════════════════════════════╗
# ║              运行前根据实际环境修改                  ║
# ╚══════════════════════════════════════════════════════╝
# 默认跟随 config.py（standalone 运行用），pytest 的 apply_config fixture 仍会覆盖
BASE_URL = getattr(_cfg, "HMI_URL", "https://192.168.2.9")    # AcuHMI 平台地址
USERNAME = getattr(_cfg, "HMI_USERNAME", "admin")             # 登录账号
PASSWORD = getattr(_cfg, "HMI_PASSWORD", "Admin@110001")      # 登录密码
# headless 模式下设备表格(Element-UI)渲染不出来 → 默认跟随 config.WEB_HEADLESS。
# config.WEB_HEADLESS=False 表示可见模式(headed)，对应 HEADED=True。
HEADED   = (not getattr(_cfg, "WEB_HEADLESS", True)) if _cfg else True
TOL_REL  = 0.01   # 相对容差 1%，超过则 FAIL
TOL_ABS  = 0.05   # 绝对容差，两者满足其一即 PASS
# ══════════════════════════════════════════════════════════

# 量类型分类（与 Device Mirror / Pass Through 的 quantity_class 保持一致）：
#   波动量（谐波/THD/相位角/K因子/电流/功率等）在网页采集与 Modbus 直读两次取值之间本身会变化，
#   差异单独报告、不计入 FAIL；稳定量（电压/频率/电能）与其他量要求一致，超容差即 FAIL。
_DYNAMIC_KW = ("power", "current", "thd", "crest", "harmonic", "angle",
               "demand", "flicker", "unbalance", "sequence", "predict", "factor")
_STABLE_KW = ("frequency", "voltage", "energy")


def quantity_class(text):
    t = (text or "").lower()
    if any(k in t for k in _DYNAMIC_KW):
        return "波动量"
    if any(k in t for k in _STABLE_KW):
        return "稳定量"
    return "其他"

TIMEOUT = 45000   # 网关网页有时较慢，给足超时（与 Pass Through/Device Mirror 一致）
VIEW_CANDIDATES = ["Realtime", "Demand", "Energy", "THD", "Sequence", "Harmonics"]

_BASE   = pathlib.Path(__file__).resolve().parent.parent   # tests/
_TMPL   = _BASE.parents[2] / "knowledge" / "shared" / "templates" / "raw"   # 仓库根 knowledge/：各设备 blockParams xlsx
REPORTS = pathlib.Path(__file__).resolve().parent          # 报告输出到本协议自身目录

# ── 设备名关键词 → 模板文件名前缀（按前缀 glob，兼容带版本号文件）────────────
DEVICE_TEMPLATES = [
    (["4110", "4100"],          "AcuRev-4100"),
    (["2100"],                  "AcuRev-2100"),
    (["acuvim3", "vim3"],       "Acuvim3"),
    (["acuvimiiw", "vimiiw", "iiiw", "iiw"], "AcuvimIIW"),
]

def _glob_latest(prefix: str):
    """在 _TMPL 下按前缀匹配 xlsx（如 AcuRev-2100 -> AcuRev-2100_v1.01_*.xlsx）。
    优先精确同名，其次取按文件名排序的最后一个（一般是最新版本）。"""
    exact = _TMPL / f"{prefix}.xlsx"
    if exact.exists():
        return exact
    cands = sorted(_TMPL.glob(f"{prefix}_*.xlsx")) + sorted(_TMPL.glob(f"{prefix}*.xlsx"))
    seen, uniq = set(), []
    for c in cands:
        if c not in seen:
            seen.add(c); uniq.append(c)
    return uniq[-1] if uniq else None

def _pick_template(device_name: str) -> pathlib.Path:
    """根据设备名关键词选择模板文件。"""
    dn = device_name.lower().replace("-", "").replace(" ", "")
    for keywords, prefix in DEVICE_TEMPLATES:
        if any(k.replace("-", "") in dn for k in keywords):
            p = _glob_latest(prefix)
            if p:
                return p
    # 默认回退到 4100
    fallback = _glob_latest("AcuRev-4100")
    log(f"  [警告] 未识别设备 '{device_name}'，使用默认模板 {fallback.name if fallback else 'AcuRev-4100'}")
    return fallback
JSON    = REPORTS / "metering_collect.json"        # 采集中间产物（流水线依赖）
CSV     = REPORTS / "metering_register_match.csv"  # 匹配中间产物（流水线依赖）
XLSX    = REPORTS / "数据对比结果.xlsx"             # 数据对比结果（交付物）

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def log(*a): print(*a, file=sys.stderr, flush=True)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — 采集
# ══════════════════════════════════════════════════════════════════════════════

def _login(page):
    page.goto(BASE_URL, wait_until="domcontentloaded")
    try: page.wait_for_url("**/login", timeout=4000)
    except PWTimeout: pass
    if "/login" not in page.url and "dashboard" in page.url.lower(): return
    page.locator("input[type=text]").first.fill(USERNAME)
    page.locator("input[type=password]").first.fill(PASSWORD)
    for sel in ["button:has-text('Sign in')", "button:has-text('Login')",
                "button:has-text('登录')", "button[type=submit]"]:
        b = page.locator(sel).first
        try: b.wait_for(state="visible", timeout=2500); b.click(); break
        except PWTimeout: continue
    for sel in [".el-message-box button:has-text('Cancel')",
                ".el-dialog button:has-text('取消')"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=2500)
            page.locator(sel).first.click(); break
        except PWTimeout: continue
    page.wait_for_url("**/dashboard", timeout=TIMEOUT)


def _list_modbus_tcp_devices(page):
    page.goto(BASE_URL + "/#/physicalDevices", wait_until="domcontentloaded")
    page.locator(".el-table__row, .el-table__body-row").first.wait_for(state="visible")
    page.wait_for_timeout(600)

    # 读取表头，找 Status 列索引
    headers = [h.strip().lower() for h in
               page.locator(".el-table__header th").all_inner_texts()]
    status_col = next((i for i, h in enumerate(headers) if "status" in h), None)

    rows = page.locator(".el-table__row, .el-table__body-row")
    out = []
    for i in range(rows.count()):
        cells = [c.strip() for c in rows.nth(i).locator("td").all_inner_texts()]
        if not cells: continue
        if "modbus tcp" not in " | ".join(cells).lower(): continue

        # 读取 Status 值
        status = ""
        if status_col is not None and status_col < len(cells):
            status = cells[status_col].strip().lower()
        else:
            # 找不到列时，也尝试全行扫描含 "off" 的单元格
            for c in cells:
                if c.strip().lower() in ("off", "offline"):
                    status = c.strip().lower(); break

        if status in ("off", "offline"):
            log(f"  [跳过] {cells[0]}  Status={status}")
            continue

        out.append({"index": i, "name": cells[0], "cells": cells, "status": status})

    log(f"Modbus TCP 设备 (Status=on): {[d['name'] for d in out]}")
    return out


def _expand_metering(page):
    m = page.get_by_text("Metering", exact=True).first
    try: m.wait_for(state="visible", timeout=8000)
    except PWTimeout: return
    rt = page.locator(".el-menu-item", has_text="Realtime").first
    for _ in range(2):
        if rt.count() and rt.is_visible(): return
        try: m.click()
        except Exception: pass
        page.wait_for_timeout(600)


def _open_device(page, name):
    page.goto(BASE_URL + "/#/physicalDevices", wait_until="domcontentloaded")
    page.locator(".el-table__row, .el-table__body-row").first.wait_for(state="visible")
    row = page.locator(".el-table__row, .el-table__body-row").filter(has_text=name).first
    row.wait_for(state="visible"); row.scroll_into_view_if_needed()
    cell0 = row.locator("td").first
    link = cell0.locator("a, .el-link, span, div").first
    (link if link.count() else cell0).click()
    page.get_by_text("Metering", exact=True).first.wait_for(state="visible")
    _expand_metering(page)


def _scrape_connection(page):
    st = page.locator(".el-sub-menu__title", has_text="Settings").first
    if st.count():
        try: st.click(); page.wait_for_timeout(400)
        except Exception: pass
    page.locator(".el-menu-item", has_text="Connection").first.click()
    page.wait_for_timeout(800)
    fields = {}
    items = page.locator(".el-form-item")
    for i in range(items.count()):
        it = items.nth(i)
        lbl = it.locator(".el-form-item__label").all_inner_texts()
        if not lbl: continue
        inp = it.locator("input")
        val = inp.first.input_value() if inp.count() else ""
        if not val:
            t = it.locator(".el-form-item__content").all_inner_texts()
            val = t[0].strip() if t else ""
        fields[lbl[0].strip().rstrip(':::').strip()] = val
    def g(k):
        for kk, v in fields.items():
            if k.lower() in kk.lower(): return v.strip()
        return ""
    return {"ip": g("IP Address"), "port": g("Port"), "modbus_id": g("Modbus ID")}


def _available_views(page):
    present = []
    for v in VIEW_CANDIDATES:
        loc = page.locator(".el-menu-item", has_text=v)
        for i in range(loc.count()):
            if loc.nth(i).inner_text().strip() == v:
                present.append(v); break
    return present


def _open_view(page, view):
    loc = page.locator(".el-menu-item", has_text=view)
    for i in range(loc.count()):
        if loc.nth(i).inner_text().strip() == view:
            if not loc.nth(i).is_visible(): _expand_metering(page)
            loc.nth(i).click(); page.wait_for_timeout(1000); return True
    return False


def _read_tables(page):
    rows_out = []; seen = set()
    for ti in range(page.locator(".el-table").count()):
        t = page.locator(".el-table").nth(ti)
        hdr = [h.strip() for h in t.locator(".el-table__header th").all_inner_texts()]
        if not hdr or hdr[0].lower() != "parameter": continue
        cols = hdr[1:]
        body = t.locator(".el-table__body-wrapper .el-table__row")
        for ri in range(body.count()):
            cells = [c.strip() for c in body.nth(ri).locator("td").all_inner_texts()]
            if not cells or not cells[0] or cells[0] in seen: continue
            seen.add(cells[0])
            values = {cols[ci - 1]: cells[ci]
                      for ci in range(1, min(len(cells), len(cols) + 1))}
            rows_out.append({"parameter": cells[0], "values": values})
    return rows_out


def _read_tables_sep(page):
    """按 .el-table 逐张读取，返回独立表格列表（**不跨表去重**，仅表内去重）：
        [{"param_header": 首列表头, "columns": [值列...],
          "rows": [{"parameter","values"}]}, ...]
    与 _read_tables 不同：
      1) 不要求首列表头为 "Parameter"——AcuRev-4100 的通道表首列是
         "Channel Parameter"，Harmonics 是 "Harmonic Order"，旧逻辑会整张丢弃；
      2) 不跨表去重——避免系统表与通道表同名参数互相覆盖。"""
    out = []
    for ti in range(page.locator(".el-table").count()):
        t = page.locator(".el-table").nth(ti)
        hdr = [h.strip() for h in t.locator(".el-table__header th").all_inner_texts()]
        if not hdr or len(hdr) < 2:
            continue
        cols = hdr[1:]
        rows = []; seen = set()
        body = t.locator(".el-table__body-wrapper .el-table__row")
        for ri in range(body.count()):
            cells = [c.strip() for c in body.nth(ri).locator("td").all_inner_texts()]
            if not cells or not cells[0] or cells[0] in seen:
                continue
            seen.add(cells[0])
            values = {cols[ci - 1]: cells[ci]
                      for ci in range(1, min(len(cells), len(cols) + 1))}
            rows.append({"parameter": cells[0], "values": values})
        if rows:
            out.append({"param_header": hdr[0], "columns": cols, "rows": rows})
    return out


def _is_channel_table(tbl):
    """判断是否为"随下拉切换的通道表"（区别于静态系统表）。
    依据 AcuRev-4100 实测表头特征：
      - 首列表头 == "Channel Parameter"（Realtime/Demand/Energy/Sequence 通道表）
      - 或末列(聚合列)为 "User Channel" / "Value"（THD/Harmonics 电流通道表）"""
    head = (tbl.get("param_header") or "").strip().lower()
    cols = [c.strip().lower() for c in tbl["columns"]]
    if head == "channel parameter":
        return True
    if cols and cols[-1] in ("user channel", "value"):
        return True
    return False


def _get_select_options(page, sel_locator):
    try:
        sel_locator.click(); page.wait_for_timeout(500)
        opts = page.locator(".el-select-dropdown__item:visible")
        items = [opts.nth(i).inner_text().strip() for i in range(opts.count())]
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
        return items
    except Exception:
        try: page.keyboard.press("Escape")
        except Exception: pass
        return []


def _select_option(page, sel_locator, opt_text):
    try:
        sel_locator.click(); page.wait_for_timeout(500)
        opts = page.locator(".el-select-dropdown__item:visible")
        for i in range(opts.count()):
            if opts.nth(i).inner_text().strip() == opt_text:
                opts.nth(i).click(); page.wait_for_timeout(800); return True
        page.keyboard.press("Escape"); return False
    except Exception:
        try: page.keyboard.press("Escape")
        except Exception: pass
        return False


def _has_no_data(page):
    """当前视图是否显示 'No Data'（通道未配置 / 无数据）。"""
    try:
        # Element Plus 空表格：.el-table__empty-text 或 .el-table__empty-block
        for sel in (".el-table__empty-text", ".el-table__empty-block",
                    ".el-empty__description", ".el-empty"):
            loc = page.locator(sel)
            for i in range(loc.count()):
                try:
                    txt = loc.nth(i).inner_text(timeout=500).strip().lower()
                    if "no data" in txt or "暂无数据" in txt:
                        return True
                except Exception:
                    pass
        # 兜底：页面上可见的 "No Data" 纯文本
        nd = page.locator("text=No Data")
        if nd.count() > 0:
            try:
                if nd.first.is_visible(timeout=500):
                    return True
            except Exception:
                pass
    except Exception:
        pass
    return False


def _read_view_all_dropdowns(page, view, on_table=None):
    """on_table(view, dropdown_opt, rows)：每抓完一张表立刻回调（用于即时读 Modbus，
    使网页值与 Modbus 值采样时刻贴近）。"""
    _open_view(page, view); page.wait_for_timeout(600)
    selects = page.locator(".el-select"); n_sel = selects.count()
    log(f"  [{view}] 下拉数={n_sel}")
    if n_sel == 0:
        if _has_no_data(page):
            log(f"  [{view}] No Data — 跳过")
            return {}
        tbl = _read_tables(page)
        if on_table: on_table(view, "(default)", tbl)
        return {"(default)": tbl}
    sel = selects.first
    options = _get_select_options(page, sel)
    log(f"    选项({len(options)}): {options[:4]}{'...' if len(options) > 4 else ''}")
    if not options:
        tbl = _read_tables(page)
        if on_table: on_table(view, "(default)", tbl)
        return {"(default)": tbl}
    result = {}
    prev_sig = None
    for opt in options:
        if not _select_option(page, sel, opt): continue
        if _has_no_data(page):
            log(f"      [{opt}] No Data — 跳过")
            continue
        rows = _read_tables(page)
        # 切换下拉后表格可能滞后刷新：若与上一选项内容完全相同，等待并重读，
        # 直到刷新或超时（避免把上一通道的残留数据当成本通道）
        waited = 0
        for _ in range(10):
            sig = json.dumps(rows, sort_keys=True, ensure_ascii=False)
            if prev_sig is None or sig != prev_sig:
                break
            page.wait_for_timeout(400); waited += 400
            rows = _read_tables(page)
        if waited:
            log(f"      [{opt}] 等待表格刷新 {waited}ms")
        if not rows:
            log(f"      [{opt}] 0 params — 跳过")
            continue
        prev_sig = json.dumps(rows, sort_keys=True, ensure_ascii=False)
        if on_table: on_table(view, opt, rows)   # 表格确认刷新后再读该表 Modbus
        result[opt] = rows
        log(f"      [{opt}] {len(rows)} params")
    return result


def _read_view(page, view, on_table=None):
    """通用视图采集（适配 AcuRev-4100「系统表 + 通道表」双表结构）。

    每个视图页面可能同时有：
      - 系统/静态表（首列 "Parameter"/"Harmonic Order"，含 Phase/Average/System 列），
        不随下拉变化 → 只采一次，key="(system)"
      - 通道表（_is_channel_table=True），随下拉(User Channel / Input Channel)切换 →
        每个选项各采一次，key=下拉选项原文（如 "User Channel 1:name1"、"Input Channel 5"），
        交由 _parse_dropdown 解析出通道号

    若存在下拉但识别不出通道表（其他型号的不同结构），回退旧版 _read_view_all_dropdowns，
    避免破坏 AcuvimIIW / Acuvim3 / AcuRev2100 等设备的既有行为。"""
    _open_view(page, view); page.wait_for_timeout(800)
    tables0 = _read_tables_sep(page)
    n_sel = page.locator(".el-select").count()
    chan0 = [t for t in tables0 if _is_channel_table(t)]
    sys0  = [t for t in tables0 if not _is_channel_table(t)]
    log(f"  [{view}] 下拉数={n_sel}  系统表={len(sys0)} 通道表={len(chan0)}")

    if n_sel > 0 and not chan0:
        log(f"  [{view}] 未识别到通道表，回退通用采集")
        return _read_view_all_dropdowns(page, view, on_table=on_table)

    result = {}

    # ── 1) 系统/静态表：合并所有系统表，只采一次 ─────────────────────────────
    sys_rows = [r for t in sys0 for r in t["rows"]]
    if sys_rows:
        result["(system)"] = sys_rows
        if on_table: on_table(view, "(system)", sys_rows)
        log(f"  [{view}] 系统表 {len(sys_rows)} 行")

    # ── 2) 通道表：随下拉逐通道采集 ──────────────────────────────────────────
    if n_sel > 0 and chan0:
        sel = page.locator(".el-select").first
        options = _get_select_options(page, sel)
        log(f"  [{view}] 通道选项({len(options)}): {options[:4]}{'...' if len(options) > 4 else ''}")

        def _chan_rows():
            return [r for t in _read_tables_sep(page) if _is_channel_table(t) for r in t["rows"]]

        for opt in options:
            if not _select_option(page, sel, opt):
                continue
            if _has_no_data(page):
                log(f"      [{opt}] No Data — 跳过")
                continue
            # 等表格稳定（连续两次读取一致即视为刷新完成），避免读到上一通道残留
            rows = _chan_rows()
            for _ in range(8):
                page.wait_for_timeout(400)
                cur = _chan_rows()
                if cur == rows:
                    break
                rows = cur
            if not rows:
                log(f"      [{opt}] 0 params — 跳过")
                continue
            result[opt] = rows
            if on_table: on_table(view, opt, rows)
            log(f"      [{opt}] {len(rows)} 行")
    return result


def _collect_with_page(page: Page) -> dict:
    """实际采集逻辑，接受已有 page 对象（已完成 set_default_timeout）。"""
    result = {}
    _login(page); log("登录成功")
    devices = _list_modbus_tcp_devices(page)
    for d in devices:
        name = d["name"]; log(f"\n===== 设备 {name} =====")
        try: _open_device(page, name)
        except Exception as e: log(f"  打开失败: {e}"); continue
        conn = {}
        try:
            conn = _scrape_connection(page)
            log(f"  Connection: {conn}"); _expand_metering(page)
        except Exception as e: log(f"  Connection 抓取失败: {e}")

        # 即时 Modbus：按设备加载模板 + 建连，抓完每张表立刻读对应寄存器
        mb_client = None; on_table = None
        ip = conn.get("ip", ""); unit = int(conn.get("modbus_id", 1) or 1)
        port = int(conn.get("port", 502) or 502)
        if ip:
            try:
                entries = _load_template(_pick_template(name))
                mb_client = ModbusTcpClient(ip, port=port, timeout=3)
                if mb_client.connect():
                    on_table = lambda view, opt, rows: _read_modbus_for_rows(mb_client, unit, entries, view, opt, rows)
                    log(f"  即时 Modbus 已连 {ip}:{port} unit={unit}")
                else:
                    log(f"  即时 Modbus 连接失败 {ip}:{port}"); mb_client.close(); mb_client = None
            except Exception as e:
                log(f"  即时 Modbus 初始化失败: {e}"); mb_client = None

            views = _available_views(page); log(f"  可用视图: {views}")
            dev_data = {}
            for v in views:
                try:
                    dev_data[v] = _read_view(page, v, on_table=on_table)
                except Exception as e: log(f"  [{v}] 读取失败: {e}"); dev_data[v] = {}
            if mb_client is not None:
                mb_client.close()
            result[name] = {"cells": d["cells"], "connection": conn, "metering": dev_data}
    return result


def collect(page: "Page | None" = None):
    """Step 1：登录 AcuHMI，采集 Metering 页面数据。

    page 参数：
      - 传入已有 Page 对象（pytest 模式）：复用共享 browser，不启动新 playwright 实例。
      - 不传（独立运行模式）：自行调用 sync_playwright() 启动浏览器，与独立执行兼容。
    """
    log("\n[Step 1] 采集页面数据 ...")
    if page is not None:
        page.set_default_timeout(TIMEOUT)
        result = _collect_with_page(page)
    else:
        with sync_playwright() as p:
            br = p.chromium.launch(headless=not HEADED, args=["--ignore-certificate-errors"])
            ctx = br.new_context(ignore_https_errors=True)
            _page = ctx.new_page(); _page.set_default_timeout(TIMEOUT)
            result = _collect_with_page(_page)
            ctx.close(); br.close()
    REPORTS.mkdir(parents=True, exist_ok=True)
    JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    total = sum(len(rows) for d in result.values()
                for vd in d["metering"].values() for rows in vd.values())
    log(f"\n已写入 {JSON}  参数条目合计={total}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — 匹配
# ══════════════════════════════════════════════════════════════════════════════

VOLTAGE_FAMILY = {
    "line to neutral voltage", "line to line voltage", "frequency",
    "voltage thd", "voltage thd odd", "voltage thd even",
    "voltage crest factor", "voltage thff factor",
    "voltage sequence magnitude", "voltage sequence angle", "voltage unbalance",
    "active energy", "reactive energy", "apparent energy",
}

PARAM_KW = {
    "line to neutral voltage phase angle": "line-to-neutral voltage phase angle",
    "line to neutral voltage":   "line-to-neutral voltage",
    "line to line voltage":      "line-to-line voltage",
    "frequency":                 "frequency",
    "current":                   "current",
    "active power":              "active power",
    "reactive power":            "reactive power",
    "apparent power":            "apparent power",
    "load nature":               "load nature",
    "power factor":              "power factor",
    "current demand":            "current demand",
    "import active power demand":"import active power demand",
    "export active power demand":"export active power demand",
    "import reactive power demand":"import reactive power demand",
    "export reactive power demand":"export reactive power demand",
    "apparent power demand":     "apparent power demand",
    "import active energy":      "active energy import",
    "export active energy":      "active energy export",
    "net active energy":         "active energy net",
    "total active energy":       "active energy total",
    "import reactive energy":    "reactive energy import",
    "export reactive energy":    "reactive energy export",
    "net reactive energy":       "reactive energy net",
    "total reactive energy":     "reactive energy total",
    "apparent energy":           "apparent energy",
    "voltage thd odd":           "voltage thd odd",
    "voltage thd even":          "voltage thd even",
    "voltage thd":               "voltage thd",
    "voltage crest factor":      "voltage crest factor",
    "voltage thff factor":       "voltage thff factor",
    "current thd odd":           "current thd odd",
    "current thd even":          "current thd even",
    "current thd":               "current thd",
    "current k factor":          "current k factor",
    "voltage sequence magnitude":"sequence magnitude",
    "voltage sequence angle":    "sequence angle",
    # 电流序量（User Channel 通道表）——必须区分 magnitude/angle，
    # 否则退化成关键词 "current"，magnitude 会被更短的 angle 寄存器误配
    "current sequence magnitude":"sequence magnitude",
    "current sequence angle":    "sequence angle",
    "voltage unbalance":         "voltage unbalance factor",
    # 中性线电流
    "neutral current":           "neutral current",
    # 电流相角（模板无此寄存器，保留 key 使其返回未匹配而非错误匹配）
    "current phase angle":       "current phase angle",
}

COL_MAP = {
    "phase a": "phase a", "phase b": "phase b", "phase c": "phase c",
    "average": "average", "system": "system", "value": "",
    "positive": "positive", "negative": "negative", "zero": "zero",
    "unbalance factor": "unbalance factor",
}


def _norm(s):
    s = re.sub(r"\(.*?\)", "", s or "")
    s = re.sub(r"[^a-zA-Z0-9 ]+", " ", s)
    return " ".join(s.lower().split())


def _is_voltage(param_norm):
    return any(kw in param_norm for kw in VOLTAGE_FAMILY)


def _col_kw(col, param_norm):
    c = col.strip().lower()
    if "line to line" in param_norm:
        return {"phase a": "phase a-b", "phase b": "phase b-c",
                "phase c": "phase c-a"}.get(c, c)
    return COL_MAP.get(c, c)


def _param_kw(param_norm):
    # 按键长降序匹配，确保长的（更具体的）关键词优先
    # 例："reactive power" 先于 "active power"，避免子串误匹配
    for k, v in sorted(PARAM_KW.items(), key=lambda x: -len(x[0])):
        if k in param_norm: return v
    return param_norm


def _parse_dropdown(opt, view=None):
    m = re.match(r"user channel\s+(\d+)", opt.lower())
    if m: return "user", int(m.group(1))
    m = re.match(r"input channel\s+(\d+)", opt.lower())
    if m: return "input", int(m.group(1))
    # 纯数字（如 AcuRev2100 下拉 "1","2",...）→ Input Channel N
    m = re.match(r"^\d+$", opt.strip())
    if m:
        # Realtime 视图的数字下拉只切相位角，主表(电压/电流/功率)是系统值，
        # 不能当 Input Channel（否则系统电流被错配到支路寄存器）。
        if "realtime" in (view or "").lower().replace(" ", ""):
            return "system", None
        return "input", int(opt.strip())
    return "system", None


def _blockparams_sheet_xml(z):
    """返回 blockParams 工作表对应的 xml 路径（按 sheet 名解析）；找不到回退 sheet2/sheet1。"""
    REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    names = z.namelist()
    try:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {r.get("Id"): r.get("Target") for r in rels}
        for sh in wb.iter(f"{{{NS}}}sheet"):
            if (sh.get("name") or "").strip().lower() == "blockparams":
                tgt = rid_to_target.get(sh.get(f"{{{REL_NS}}}id"), "").split("/")[-1]
                cand = f"xl/worksheets/{tgt}"
                if cand in names:
                    return cand
    except Exception:
        pass
    for fb in ("xl/worksheets/sheet2.xml", "xl/worksheets/sheet1.xml"):
        if fb in names:
            return fb
    return names[0]


def _load_template(path: pathlib.Path):
    """解析模板 blockParams sheet；按表头文字定位列（兼容新旧布局：
    旧 D/E/H/C/L，新 A/B/F/-/J）。paramId 在新模板无对应列时留空。"""
    with zipfile.ZipFile(str(path)) as z:
        ss = {}
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for i, si in enumerate(root.iter(f"{{{NS}}}si")):
                ss[i] = "".join(c.text or "" for c in si.iter()
                                if c.tag == f"{{{NS}}}t")
        root = ET.fromstring(z.read(_blockparams_sheet_xml(z)))
        raw = []
        for row in root.iter(f"{{{NS}}}row"):
            cells = {}
            for cell in row:
                ref = cell.get("r", ""); t = cell.get("t", "")
                v = cell.find(f"{{{NS}}}v")
                val = (ss.get(int(v.text), "") if t == "s" else v.text) \
                      if (v is not None and v.text) else ""
                cells["".join(c for c in ref if c.isalpha())] = val
            if cells: raw.append(cells)
    if not raw:
        log("模板条数: 0"); return []
    # 按表头定位列
    header = raw[0]
    c_addr = c_desc = c_dtype = c_pid = c_scale = None
    for col, txt in header.items():
        t = (txt or "").strip().lower()
        if c_addr is None and t.startswith("start"):       c_addr = col
        elif c_desc is None and t.startswith("descr"):     c_desc = col
        elif c_dtype is None and t in ("datatype", "data type"): c_dtype = col
        elif c_pid is None and t == "paramid":             c_pid = col
        elif c_scale is None and t == "scale":             c_scale = col
    c_addr = c_addr or "D"; c_desc = c_desc or "E"; c_dtype = c_dtype or "H"
    c_scale = c_scale or "L"   # c_pid 找不到则留空（新模板无 paramId 列，不回退避免取错）
    entries = []
    for r in raw[1:]:
        desc = r.get(c_desc, "").strip(); addr = r.get(c_addr, "").strip()
        if not desc or not addr: continue
        try: a = int(addr)
        except ValueError: continue
        scale_str = (r.get(c_scale, "") or "").strip() or "1"
        try: scale = float(scale_str)
        except ValueError: scale = 1.0
        entries.append({"desc": desc, "addr": a, "hex": f"0x{a:04X}",
                         "dtype": r.get(c_dtype, "").strip(),
                         "param_id": (r.get(c_pid, "").strip() if c_pid else ""),
                         "scale": scale})
    log(f"模板条数: {len(entries)}")
    return entries


def _match_entry(entries, ch_type, ch_num, param_kw, col_kw):
    def _find(kws):
        kws = [k for k in kws if k]
        if not kws: return None
        cands = [(len(e["desc"]), e["addr"], e) for e in entries
                 if all(k in e["desc"].lower() for k in kws)]
        if not cands:
            # 回退：按词元(token)匹配，兼容词序差异
            #   如网页 'voltage thd odd' ↔ 寄存器 'voltage odd thd'
            #      网页 'current phase angle'+'phase a' ↔ 'Phase Angle of Phase A Current to V1'
            # 用 \b 词边界，避免单字母 'a' 误配 'angle'，'v1' 误配等
            toks = [t for k in kws for t in k.split()]
            cands = [(len(e["desc"]), e["addr"], e) for e in entries
                     if all(re.search(r"\b" + re.escape(t) + r"\b", e["desc"].lower())
                            for t in toks)]
        if not cands: return None
        # 排除 real/imag 部寄存器（AcuvimIIW 序量用实部/虚部表示，网页显示的是幅值/角度，
        # 二者不可直接比对；网页从不出现 real/imag，故全局排除避免误配成 FAIL）
        _f = [(l, a, e) for l, a, e in cands
              if "real part" not in e["desc"].lower() and "imag" not in e["desc"].lower()]
        if _f: cands = _f
        # 当 param_kw 不含 "thd"/"demand" 时，优先排除含这些词的条目（避免误匹配）
        if "thd" not in param_kw:
            filtered = [(l, a, e) for l, a, e in cands if "thd" not in e["desc"].lower()]
            if filtered: cands = filtered
        if "demand" not in param_kw:
            filtered = [(l, a, e) for l, a, e in cands if "demand" not in e["desc"].lower()]
            if filtered: cands = filtered
        return sorted(cands, key=lambda x: (x[0], x[1]))[0][2]

    col_norm = col_kw.lower() if col_kw else ""

    if ch_type == "user" and ch_num:
        ch_kw = f"user channel {ch_num}"
        # 先带列名精确匹配（区分 Positive/Negative/Zero、Phase A/B/C 等）
        if col_norm and col_norm not in ("system", "average"):
            e = _find([ch_kw, col_norm, param_kw])
            if e: return e
        # 降级：不带列名
        return _find([ch_kw, param_kw])

    elif ch_type == "input" and ch_num:
        ch_kw = f"input channel {ch_num}"
        if col_norm and col_norm not in ("system", "average"):
            e = _find([ch_kw, col_norm, param_kw])
            if e: return e
        return _find([ch_kw, param_kw])

    else:
        # system / average：用列名作为通道区分词
        if col_norm and col_norm not in ("system", "average"):
            return _find([col_norm, param_kw])
        elif col_norm in ("system", "average"):
            e = _find([col_norm, param_kw])
            # 回退：去掉 col_kw（如 Neutral Current 无 System 前缀）
            if e is None:
                e = _find([param_kw])
            return e
        return _find([param_kw])


def _match_harmonic(entries, view, dropdown, param, col):
    """Harmonics 视图专用匹配（参数名是谐波次数 N，通用关键词匹配会乱配）。
      - 系统表(电压)：列 Phase A/B/C → 'Phase {X} Voltage individual Harmonics {N} percentage'
      - 通道表(电流)：dropdown=Input Channel M + 列 value
                     → 'Input Channel {M} Current individual Harmonics {N} percentage'
    按精确描述查找，命中即返回对应寄存器条目。"""
    order = str(param).strip()
    if not order.isdigit():
        return None
    col_l = col.strip().lower()
    if col_l.startswith("phase "):
        phase = col_l.split()[-1]                      # a/b/c
        target = f"phase {phase} voltage individual harmonics {order} percentage"
    else:
        # 电流谐波：从下拉解析 Input Channel 号
        ch_type, ch_num = _parse_dropdown(dropdown, view)
        if ch_type != "input" or not ch_num:
            return None
        target = f"input channel {ch_num} current individual harmonics {order} percentage"
    for e in entries:
        if e["desc"].strip().lower() == target:
            return e
    return None


def _match_label(desc, param_norm, col, ch_type, ch_num):
    """用「参数名 + 列名 + 通道」的词集合去抵消寄存器描述里的词，
    剩余词 ≤1 视为"精确"，否则"模糊"。比旧的"按关键词逐个 replace 后看剩余字符数"
    更稳健，且天然兼容词序差异（如网页 'Import Active Energy' ↔ 寄存器 'active energy import'）。"""
    used = set(param_norm.split()) | set(_norm(col).split())
    if ch_type == "user" and ch_num:
        used |= {"user", "channel", str(ch_num)}
    elif ch_type == "input" and ch_num:
        used |= {"input", "channel", str(ch_num)}
    dwords = re.sub(r"[^a-z0-9]+", " ", desc.lower()).split()
    leftover = [w for w in dwords if w not in used]
    return "精确" if len(leftover) <= 1 else "模糊"


def _is_no_register_param(param_norm, ch_type):
    """已知"页面有值、但 Modbus 模板无对应寄存器"的参数（非匹配失败，属设备固有）。
    这类条目标为 N/A(无寄存器) 而非未匹配——既不计入未匹配阈值，也被精确率统计排除。
      - User Channel 的逐相电流相角：模板只有 'Input Channel M Current Phase Angle'，
        无 'User Channel N Phase X' 版本（网页该值由所属 Input Channel 派生显示）。"""
    if "current phase angle" in param_norm and ch_type == "user":
        return True
    # 电压相角：部分型号（如 AcuvimIIW）以 V1 为基准，模板无电压相角寄存器。
    # 仅在匹配彻底失败时才会走到这里，对有该寄存器的型号（已匹配成功）无影响。
    if "voltage" in param_norm and "phase angle" in param_norm:
        return True
    return False


def match_registers(data=None):
    """Step 2：匹配寄存器地址（每台设备按名称自动选对应模板）。"""
    log("\n[Step 2] 匹配寄存器地址 ...")
    if data is None:
        data = json.loads(JSON.read_text(encoding="utf-8"))

    # 按设备预加载模板，避免重复读取
    _template_cache: dict[str, list] = {}
    def get_entries(dev_name):
        if dev_name not in _template_cache:
            tmpl = _pick_template(dev_name)
            log(f"  [{dev_name}] 使用模板: {tmpl.name}")
            _template_cache[dev_name] = _load_template(tmpl)
        return _template_cache[dev_name]

    rows_out = []
    for dev_name, dev_info in data.items():
        entries = get_entries(dev_name)
        for view, view_data in dev_info["metering"].items():
            is_harm = view.strip().lower() == "harmonics"
            for dropdown_opt, param_rows in view_data.items():
                ch_type_raw, ch_num_raw = _parse_dropdown(dropdown_opt, view)
                for r in param_rows:
                    param = r["parameter"]; pn = _norm(param); pkw = _param_kw(pn)
                    for col, val in r["values"].items():
                        if str(val).strip() in ("-", "--", "", "N/A"):
                            rows_out.append({
                                "device": dev_name, "view": view,
                                "dropdown": dropdown_opt, "parameter": param,
                                "column": col, "ui_value": val,
                                "reg_desc": "", "addr_dec": "", "addr_hex": "",
                                "dtype": "", "param_id": "", "scale": 1.0,
                                "match": "N/A(页面无值)"
                            }); continue
                        if is_harm:
                            # 谐波视图：参数名为次数，走专用精确匹配
                            e = _match_harmonic(entries, view, dropdown_opt, param, col)
                            if e:
                                rows_out.append({
                                    "device": dev_name, "view": view,
                                    "dropdown": dropdown_opt, "parameter": param,
                                    "column": col, "ui_value": val,
                                    "reg_desc": e["desc"], "addr_dec": e["addr"],
                                    "addr_hex": e["hex"], "dtype": e["dtype"],
                                    "param_id": e["param_id"],
                                    "scale": e.get("scale", 1.0), "match": "精确",
                                })
                            else:
                                rows_out.append({
                                    "device": dev_name, "view": view,
                                    "dropdown": dropdown_opt, "parameter": param,
                                    "column": col, "ui_value": val,
                                    "reg_desc": "", "addr_dec": "", "addr_hex": "",
                                    "dtype": "", "param_id": "", "scale": 1.0,
                                    "match": "未匹配"
                                })
                            continue
                        ckw = _col_kw(col, pn)
                        if _is_voltage(pn):
                            ch_type, ch_num = "system", None
                        else:
                            ch_type, ch_num = ch_type_raw, ch_num_raw
                        e = _match_entry(entries, ch_type, ch_num, pkw, ckw)
                        if e:
                            label = _match_label(e["desc"], pn, col, ch_type, ch_num)
                            rows_out.append({
                                "device": dev_name, "view": view,
                                "dropdown": dropdown_opt, "parameter": param,
                                "column": col, "ui_value": val,
                                "reg_desc": e["desc"], "addr_dec": e["addr"],
                                "addr_hex": e["hex"], "dtype": e["dtype"],
                                "param_id": e["param_id"],
                                "scale": e.get("scale", 1.0),
                                "match": label,
                            })
                        else:
                            no_reg = _is_no_register_param(pn, ch_type)
                            rows_out.append({
                                "device": dev_name, "view": view,
                                "dropdown": dropdown_opt, "parameter": param,
                                "column": col, "ui_value": val,
                                "reg_desc": "", "addr_dec": "", "addr_hex": "",
                                "dtype": "", "param_id": "", "scale": 1.0,
                                "match": "N/A(无寄存器)" if no_reg else "未匹配"
                            })

    # ── 塌缩检测 ──────────────────────────────────────────────────────────────
    # 同一参数行的多个列（Phase A/B/C 等）匹配到【同一寄存器地址】，说明设备只有
    # 聚合/单通道寄存器、无逐相寄存器（如 AcuRev-4100 'User Channel N Current'、
    # AcuRev-2100 'Input Channel N Current THD'）。拿同一聚合值与不同的逐相网页值比对
    # 语义错误（4100 因各相值接近碰巧 PASS，2100 相差略大即 FAIL）。
    # 处理：保留聚合列(system/average/user channel/value/空)，其余逐相列标 N/A(无逐相寄存器)。
    _AGG_COLS = ("system", "average", "user channel", "value", "")
    _grp: dict = {}
    for r in rows_out:
        if not r["addr_dec"]:
            continue
        key = (r["device"], r["view"], r["dropdown"], r["parameter"])
        _grp.setdefault(key, {}).setdefault(str(r["addr_dec"]), []).append(r)
    for _k, _addrmap in _grp.items():
        for _addr, _rs in _addrmap.items():
            if len(_rs) < 2:
                continue
            _agg = [r for r in _rs if r["column"].strip().lower() in _AGG_COLS]
            _keep = _agg[0] if _agg else None
            for r in _rs:
                if r is _keep:
                    continue
                r.update({"reg_desc": "", "addr_dec": "", "addr_hex": "",
                          "dtype": "", "param_id": "", "scale": 1.0,
                          "match": "N/A(无逐相寄存器)"})

    REPORTS.mkdir(parents=True, exist_ok=True)
    if rows_out:
        with CSV.open("w", encoding="utf-8-sig", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
            w.writeheader(); w.writerows(rows_out)
    total = len(rows_out)
    exact = sum(1 for r in rows_out if r["match"] == "精确")
    fuzzy = sum(1 for r in rows_out if r["match"] == "模糊")
    nomatch = sum(1 for r in rows_out if r["match"] == "未匹配")
    na = sum(1 for r in rows_out if r["match"].startswith("N/A"))
    log(f"总={total}  精确={exact}  模糊={fuzzy}  未匹配={nomatch}  N/A={na}")
    return rows_out


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — 比对
# ══════════════════════════════════════════════════════════════════════════════

RED    = PatternFill("solid", fgColor="FFCCCC")
GREEN  = PatternFill("solid", fgColor="CCFFCC")
YELLOW = PatternFill("solid", fgColor="FFFACC")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _modbus_read(client, addr, dtype, slave):
    count = {"float32": 2, "double": 4, "uint32": 2, "uint16": 1,
             "int32": 2, "int16": 1}.get(dtype.lower(), 2)
    try:
        rr = client.read_holding_registers(addr, count=count, slave=slave)
        if rr.isError(): return None
        regs = rr.registers
        if dtype.lower() == "float32":
            return struct.unpack(">f", struct.pack(">HH", regs[0], regs[1]))[0]
        elif dtype.lower() == "double":
            return struct.unpack(">d", struct.pack(">HHHH", *regs[:4]))[0]
        elif dtype.lower() in ("uint32", "int32"):
            return struct.unpack(">I", struct.pack(">HH", regs[0], regs[1]))[0]
        elif dtype.lower() in ("uint16", "int16"):
            return float(regs[0])
        else:
            return struct.unpack(">f", struct.pack(">HH", regs[0], regs[1]))[0]
    except Exception:
        return None


def _parse_ui(val_str):
    s = re.sub(r"[^\d.\-eE]", "", str(val_str).strip())
    try: return float(s)
    except ValueError: return None


def _in_tol(ui, mb):
    if ui is None or mb is None: return None
    diff = abs(ui - mb)
    return diff <= TOL_ABS or diff / (abs(ui) + 1e-9) <= TOL_REL


def _read_modbus_for_rows(client, unit, entries, view, dropdown_opt, rows):
    """抓完一张表后立刻读该表所有参数的 Modbus 寄存器，把 (值/地址/类型) 标注到 row['mb'][col]。
    与 match_registers 用同一套匹配逻辑（含 voltage 特例 + view 感知），保证网页值与 Modbus 值采样时刻贴近。"""
    if client is None or not entries:
        return
    is_harm = view.strip().lower() == "harmonics"
    ch_type_raw, ch_num_raw = _parse_dropdown(dropdown_opt, view)
    for r in rows:
        param = r["parameter"]; pn = _norm(param); pkw = _param_kw(pn)
        mb = r.setdefault("mb", {})
        for col, val in r["values"].items():
            if str(val).strip() in ("-", "--", "", "N/A"):
                continue
            if is_harm:
                e = _match_harmonic(entries, view, dropdown_opt, param, col)
            else:
                ckw = _col_kw(col, pn)
                if _is_voltage(pn):
                    ch_type, ch_num = "system", None
                else:
                    ch_type, ch_num = ch_type_raw, ch_num_raw
                e = _match_entry(entries, ch_type, ch_num, pkw, ckw)
            if not e:
                continue
            raw = _modbus_read(client, e["addr"], e["dtype"], unit)
            if raw is not None:
                raw = raw * e.get("scale", 1.0)
            mb[col] = {"value": round(raw, 4) if raw is not None else None,
                       "addr": e["addr"], "hex": e["hex"], "dtype": e["dtype"],
                       "reg_desc": e["desc"], "param_id": e.get("param_id", ""),
                       "scale": e.get("scale", 1.0)}


def compare(collect_data=None, match_rows=None):
    """Step 3：Modbus 直读比对，生成 Excel 报告。"""
    log("\n[Step 3] Modbus 比对 ...")
    if collect_data is None:
        collect_data = json.loads(JSON.read_text(encoding="utf-8"))
    if match_rows is None:
        with CSV.open(encoding="utf-8-sig") as f:
            match_rows = list(_csv.DictReader(f))

    match_idx = {}
    no_reg = set()   # 匹配阶段判定为"无寄存器/无逐相寄存器"的列 → 即使采集时存了 mb 也不比对
    for r in match_rows:
        key = (r["device"], r["view"], r["dropdown"], r["parameter"], r["column"])
        if str(r.get("match", "")).startswith("N/A(无"):
            no_reg.add(key)
        if r["match"] in ("N/A(页面无值)", "") or not r["addr_dec"]: continue
        try:
            match_idx[key] = {"addr": int(r["addr_dec"]), "dtype": r["dtype"],
                               "hex": r["addr_hex"], "param_id": r["param_id"],
                               "reg_desc": r["reg_desc"],
                               "scale": float(r.get("scale", 1) or 1)}
        except ValueError:
            pass

    all_results = []
    for dev_name, dev_info in collect_data.items():
        conn = dev_info.get("connection", {})
        ip   = conn.get("ip", "")
        port = int(conn.get("port", 502) or 502)
        unit = int(conn.get("modbus_id", 1) or 1)
        # 优先用采集时即时读到的 mb（与网页值同表抓取，时刻贴近）；缺失才回退实时读
        client = None
        if ip:
            client = ModbusTcpClient(ip, port=port, timeout=3)
            if not client.connect():
                log(f"[{dev_name}] 实时回退连接失败 {ip}:{port}（将仅用采集时 Modbus 值）")
                client.close(); client = None
        for view, view_data in dev_info["metering"].items():
            for dropdown, param_rows in view_data.items():
                for r in param_rows:
                    param = r["parameter"]
                    mb_stored = r.get("mb") or {}
                    for col, ui_str in r["values"].items():
                        key = (dev_name, view, dropdown, param, col)
                        ui_num = _parse_ui(ui_str)
                        # 塌缩/无寄存器列：不比对，直接记"无寄存器"
                        if key in no_reg:
                            all_results.append({
                                "device": dev_name, "view": view, "dropdown": dropdown,
                                "parameter": param, "column": col, "ui_value": ui_str,
                                "modbus_value": "", "addr_dec": "", "addr_hex": "",
                                "dtype": "", "reg_desc": "", "param_id": "",
                                "qclass": quantity_class(f"{param} {col}"),
                                "diff": "", "diff_pct": "", "result": "无寄存器"
                            }); continue
                        st = mb_stored.get(col)
                        if st and st.get("addr") not in (None, ""):
                            # 采集时即时读到的值（已 scale/round）
                            addr=st["addr"]; hexs=st["hex"]; dtype=st["dtype"]
                            reg_desc=st.get("reg_desc",""); pid=st.get("param_id","")
                            mb_val = st.get("value")
                        else:
                            info = match_idx.get(key)
                            if info is None:
                                all_results.append({
                                    "device": dev_name, "view": view, "dropdown": dropdown,
                                    "parameter": param, "column": col, "ui_value": ui_str,
                                    "modbus_value": "", "addr_dec": "", "addr_hex": "",
                                    "dtype": "", "reg_desc": "", "param_id": "",
                                    "qclass": quantity_class(f"{param} {col}"),
                                    "diff": "", "diff_pct": "", "result": "无寄存器"
                                }); continue
                            addr=info["addr"]; hexs=info["hex"]; dtype=info["dtype"]
                            reg_desc=info["reg_desc"]; pid=info["param_id"]
                            if client is None:
                                mb_val = None
                            else:
                                mb_raw = _modbus_read(client, addr, dtype, unit)
                                if mb_raw is not None:
                                    mb_raw = mb_raw * info.get("scale", 1.0)
                                mb_val = round(mb_raw, 4) if mb_raw is not None else None
                        # Power Factor: device may report signed (−1…+1), UI shows absolute
                        _pf_param = "power factor" in param.lower()
                        _ui_cmp  = abs(ui_num)  if (_pf_param and ui_num  is not None) else ui_num
                        _mb_cmp  = abs(mb_val)  if (_pf_param and mb_val  is not None) else mb_val
                        _angle_param = "angle" in param.lower()
                        diff = diff_pct = ""; result = "N/A"
                        if ui_num is not None and mb_val is not None:
                            if _angle_param:
                                # 角度按环形差比较：0°≈360°（如 0.008 vs 359.99 实为同一角度）
                                cd = abs(_ui_cmp - _mb_cmp) % 360.0
                                cd = min(cd, 360.0 - cd)
                                diff = round(cd, 4)
                                diff_pct = round(cd / 360.0 * 100, 2)
                                result = "PASS" if (cd <= TOL_ABS or cd <= TOL_REL * 360.0) else "FAIL"
                            else:
                                d = _ui_cmp - _mb_cmp
                                diff = round(d, 4)
                                diff_pct = round(abs(d) / (abs(_ui_cmp) + 1e-9) * 100, 2)
                                result = "PASS" if _in_tol(_ui_cmp, _mb_cmp) else "FAIL"
                        elif mb_val is None:
                            result = "Modbus读取失败"
                        all_results.append({
                            "device": dev_name, "view": view, "dropdown": dropdown,
                            "parameter": param, "column": col, "ui_value": ui_str,
                            "modbus_value": mb_val if mb_val is not None else "ERR",
                            "addr_dec": addr, "addr_hex": hexs,
                            "dtype": dtype, "reg_desc": reg_desc,
                            "param_id": pid,
                            "qclass": quantity_class(f"{param} {col} {reg_desc}"),
                            "diff": diff, "diff_pct": diff_pct, "result": result
                        })
        if client is not None:
            client.close()

    # 写 xlsx
    wb = openpyxl.Workbook()
    ws_sum = wb.active; ws_sum.title = "汇总"
    total  = len(all_results)
    passed = sum(1 for r in all_results if r["result"] == "PASS")
    failed = sum(1 for r in all_results if r["result"] == "FAIL")
    # 波动量 FAIL 单独统计（不计入严格失败）；稳定量/其他 FAIL 才是 failed_stable
    failed_dynamic = sum(1 for r in all_results
                         if r["result"] == "FAIL" and r.get("qclass") == "波动量")
    failed_stable = failed - failed_dynamic
    noread = sum(1 for r in all_results if r["result"] == "Modbus读取失败")
    noregs = sum(1 for r in all_results if r["result"] == "无寄存器")
    comparable = total - noread - noregs
    pass_rate = f"{passed/comparable*100:.1f}%" if comparable > 0 else "N/A"
    for row in [["项目","数量"],["总参数条目",total],["PASS",passed],
                ["FAIL(稳定量/其他)",failed_stable],["FAIL(波动量,仅报告)",failed_dynamic],
                ["Modbus 读取失败",noread],["无寄存器映射",noregs],["通过率",pass_rate]]:
        ws_sum.append(row)
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 12

    ws = wb.create_sheet("明细")
    headers = ["设备","视图","通道/选项","参数","列","量类型","网页值","Modbus值","差值","差值%",
               "寄存器地址(Dec)","地址(Hex)","类型","寄存器描述","paramId","结果"]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    col_map = ["device","view","dropdown","parameter","column","qclass","ui_value","modbus_value",
               "diff","diff_pct","addr_dec","addr_hex","dtype","reg_desc","param_id","result"]
    all_results.sort(key=lambda r: (-float(r["diff_pct"]) if str(r["diff_pct"]).replace('.','',1).isdigit() else 0,))
    for r in all_results:
        ws.append([r.get(k, "") for k in col_map])
        ri = ws.max_row; res = r["result"]
        if res == "FAIL":
            for c in range(1, len(headers)+1): ws.cell(ri, c).fill = RED
        elif res == "PASS":
            ws.cell(ri, len(headers)).fill = GREEN
        elif res in ("Modbus读取失败", "无寄存器"):
            ws.cell(ri, len(headers)).fill = YELLOW
    for i, w in enumerate([12,10,22,35,10,8,12,12,8,8,14,10,8,35,10,8], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    # ── 用例 Mapping sheet ─────────────────────────────────────────────────
    DC_MAPPING = [
        ("DC-001", "TestCollect::test_json_generated",          "采集", "Step1 采集结果文件 metering_collect.json 正常生成"),
        ("DC-002", "TestCollect::test_has_devices",             "采集", "至少采集到 1 台设备数据"),
        ("DC-003", "TestCollect::test_connection_info",         "采集", "每台设备均抓取到 Connection IP 地址"),
        ("DC-004", "TestCollect::test_metering_views",          "采集", "每台设备至少包含 1 个 Metering 视图数据"),
        ("DC-005", "TestCollect::test_param_count",             "采集", "总采集参数条目数量 > 0"),
        ("DC-006", "TestMatch::test_csv_generated",             "匹配", "Step2 匹配结果文件 metering_register_match.csv 正常生成"),
        ("DC-007", "TestMatch::test_match_count",               "匹配", "匹配结果总条目数量 > 0"),
        ("DC-008", "TestMatch::test_exact_match_ratio",         "匹配", "精确匹配率 ≥ 80%（排除 N/A 条目）"),
        ("DC-009", "TestMatch::test_no_unmatched_excess",       "匹配", "未匹配条目数量 ≤ 20"),
        ("DC-010", "TestCompare::test_xlsx_generated",          "比对", "Step3 比对报告 metering_compare.xlsx 正常生成"),
        ("DC-011", "TestCompare::test_fail_count_zero",         "比对", "稳定量/其他 FAIL = 0（波动量差异单独报告，不计入）"),
        ("DC-012", "TestCompare::test_pass_rate",               "比对", "有效比对通过率 ≥ 95%（排除无寄存器映射条目）"),
        ("DC-013", "TestCompare::test_modbus_read_failure_zero","比对", "Modbus 读取失败条目 = 0"),
    ]
    wm = wb.create_sheet("用例Mapping")
    HF_M = PatternFill("solid", fgColor="305496")
    HFONT_M = Font(color="FFFFFF", bold=True)
    MH = ["用例ID", "用例名称（Class::method）", "测试步骤", "验证点"]
    for ci, h in enumerate(MH, 1):
        c = wm.cell(1, ci, h); c.fill = HF_M; c.font = HFONT_M
        c.alignment = Alignment(horizontal="center")
    wm.freeze_panes = "A2"
    for row in DC_MAPPING:
        wm.append(list(row))
    for i, w in enumerate([8, 42, 8, 52], 1):
        wm.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    REPORTS.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(XLSX)
        log(f"写入 {XLSX}")
    except PermissionError:
        alt = REPORTS / "数据对比结果_new.xlsx"
        wb.save(alt)
        log(f"[警告] {XLSX.name} 被占用，已改写到 {alt.name}（请关闭 Excel 后重跑覆盖）")
    log(f"总={total} PASS={passed} FAIL(稳定/其他)={failed_stable} "
        f"FAIL(波动量)={failed_dynamic} 读取失败={noread} 无寄存器={noregs}")
    return {"total": total, "passed": passed, "failed": failed,
            "failed_stable": failed_stable, "failed_dynamic": failed_dynamic,
            "noread": noread, "noregs": noregs}


# ══════════════════════════════════════════════════════════════════════════════
# 入口
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Metering 数据采集 + 比对")
    parser.add_argument("--collect", action="store_true", help="仅执行采集")
    parser.add_argument("--match",   action="store_true", help="仅执行匹配")
    parser.add_argument("--compare", action="store_true", help="仅执行比对")
    args = parser.parse_args()

    run_all = not (args.collect or args.match or args.compare)

    collected = None
    matched   = None

    if args.collect or run_all:
        collected = collect()
    if args.match or run_all:
        matched = match_registers(collected)
    if args.compare or run_all:
        compare(collected, matched)


if __name__ == "__main__":
    main()
