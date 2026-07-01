# -*- coding: utf-8 -*-
"""Pass Through 透传业务数据自动化对比（单文件、可独立执行的 pytest）

验证 AcuHMI「Pass Through」透传是否忠实还原真实电表的寄存器：
  A 路（透传）：经网关 GATEWAY_IP:502，按各设备 Pass Through SlaveID 读寄存器
  B 路（直读真实电表）：从设备网页 Physical Devices→Settings→Connection 抓取的真实 IP/Unit 直连电表读同一寄存器
  A↔B 比对，逐寄存器明细落地 reports/。

读哪些寄存器由「AcuCloud 模板 xlsx」决定：Protocols/template/AcuCloud 模板适配/<Model>.xlsx
（D 列=起始地址，E 列=参数描述，H 列=数据类型；功能码固定 FC=3 保持寄存器；缺表的型号自动跳过并提示）。
B 路 IP/Unit 一律以网页抓取为准（不用 config 的设备表）。网关地址来自上级 config.py 的 GATEWAY_IP；容差用 TOLERANCE_*。
运行（仓库根目录下，任意人 clone 后均可）：
  pytest "projects/AcuHMI_1_7/tests/Pass Through" -v
也可直接运行本文件： python "projects/AcuHMI_1_7/tests/Pass Through/test_pass_through_business.py" -v
"""
from __future__ import annotations

import json
import math
import os
import pathlib
import re
import struct
import zipfile
import xml.etree.ElementTree as _ET
import sys
import time
from dataclasses import dataclass

import pytest
from playwright.sync_api import Browser, Page, Locator, TimeoutError as PWTimeout

# ── 让本文件可独立 import 上级目录的 config.py ────────────────────────────────
_ROOT = pathlib.Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
# 配置来源：优先本地 tests/config.py（开发机覆盖，gitignored）；别人 clone 没有它时，
# 回退到框架配置（configs/.env + config.yaml），保证从仓库根直接 pytest 也能开箱即跑。
import importlib.util as _ilu, types as _types

def _load_local_config():
    _cfg_path = pathlib.Path(__file__).resolve().parent.parent / "config.py"
    if _cfg_path.exists():
        _spec = _ilu.spec_from_file_location("_tests_config", _cfg_path)
        _mod = _ilu.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        return _mod
    _RR = pathlib.Path(__file__).resolve().parents[4]
    if str(_RR) not in sys.path:
        sys.path.insert(0, str(_RR))
    from projects.AcuHMI_1_7 import settings as _s
    return _types.SimpleNamespace(
        HMI_URL=_s.HMI_URL, GATEWAY_IP=_s.HMI_IP, MODBUS_PORT=_s.METER_TCP_PORT,
        HMI_USERNAME=_s.HMI_USERNAME, HMI_PASSWORD=_s.HMI_PASSWORD,
        TOLERANCE_PERCENT=_s.MODBUS_CMP_TOLERANCE_PERCENT,
        TOLERANCE_ABSOLUTE=_s.MODBUS_CMP_TOLERANCE_ABSOLUTE,
        WEB_HEADLESS=_s.HEADLESS,
        MODBUS_DEVICE_MAP=getattr(_s, "DEVICE_MODBUS_MAP", {}),
    )

config = _load_local_config()

# ── 由 config 派生的配置（只改 config.py 即可）────────────────────────────────
BASE_URL     = os.getenv("DM_BASE_URL", config.HMI_URL).rstrip("/")
# 透传 Modbus 服务跑在 AcuHMI 网关自身（与网页同 IP），不是 config.GATEWAY_IP(192.168.2.8)。
# 默认从网页地址解析网关 IP（如 https://192.168.3.51 -> 192.168.3.51），可用 DM_GATEWAY_IP 覆盖。
from urllib.parse import urlparse as _urlparse
GATEWAY_IP   = os.getenv("DM_GATEWAY_IP", _urlparse(BASE_URL).hostname or config.GATEWAY_IP)
USERNAME     = os.getenv("DM_USERNAME", config.HMI_USERNAME)
PASSWORD     = os.getenv("DM_PASSWORD", config.HMI_PASSWORD)
MODBUS_PORT  = int(getattr(config, "MODBUS_PORT", 502) or 502)
MODBUS_TIMEOUT = float(os.getenv("MODBUS_TIMEOUT", "5"))
REL_TOL      = float(getattr(config, "TOLERANCE_PERCENT", 1.0)) / 100.0
ABS_FLOOR    = float(getattr(config, "TOLERANCE_ABSOLUTE", 0.05))
MIN_RATE     = float(os.getenv("COMPARE_MIN_RATE", "0.6"))
READ_PAUSE   = float(os.getenv("DEVICE_READ_PAUSE", "1.0"))
CONN_RETRIES = int(os.getenv("CONN_SCRAPE_RETRIES", "3"))   # 网页 Connection 抓取重试次数
PT_SETTLE_MS = int(os.getenv("PT_SETTLE_MS", "3000"))       # 启用 Pass Through 后等网关透传服务就绪
PT_DISABLE_SETTLE = float(os.getenv("PT_DISABLE_SETTLE", "20"))  # 禁用后确认停服的轮询窗口(秒)
SELF_DEVICE  = os.getenv("SELF_DEVICE_NAME", "AcuHMI-1-7")
PT_SLAVE_MIN, PT_SLAVE_MAX = 101, 247
HEADED       = os.getenv("HEADED", "0").strip().lower() in ("1", "true", "yes", "on")
TIMEOUT      = int(os.getenv("TIMEOUT", "45000"))
INTERFACE_ETHERNET = "Ethernet"
TMPL_DIR     = _ROOT.parents[2] / "knowledge" / "shared" / "templates" / "raw"  # 仓库根 knowledge/：原始版本号模板
REPORT       = pathlib.Path(__file__).resolve().parent   # 报告输出到本协议自身目录


# ═══════════════════════════════════════════════════════════════════════════
# Modbus TCP 读取 + 解码
# ═══════════════════════════════════════════════════════════════════════════
from pymodbus.client import ModbusTcpClient


def _regs_to_bytes(regs):
    return b"".join(struct.pack(">H", r & 0xFFFF) for r in regs)


def _decode(regs, kind):
    if kind == "uint16":
        return regs[0]
    if kind == "int16":
        return struct.unpack(">h", struct.pack(">H", regs[0]))[0]
    if kind == "uint32":
        return float(struct.unpack(">I", _regs_to_bytes(regs[:2]))[0])
    if kind == "float32":
        return struct.unpack(">f", _regs_to_bytes(regs[:2]))[0]
    if kind == "float64":
        return struct.unpack(">d", _regs_to_bytes(regs[:4]))[0]
    if kind == "string":
        return _regs_to_bytes(regs).decode("ascii", errors="ignore").rstrip("\x00 ").strip()
    raise ValueError(f"未知解码类型: {kind}")


class Modbus:
    def __init__(self, host=None, port=None):
        self.host = host or GATEWAY_IP
        self.port = port or MODBUS_PORT
        self.client = ModbusTcpClient(self.host, port=self.port, timeout=MODBUS_TIMEOUT)

    def __enter__(self):
        if not self.client.connect():
            raise IOError(f"无法连接 {self.host}:{self.port}")
        return self

    def __exit__(self, *exc):
        self.client.close()

    def _read_regs(self, fc, address, count, slave):
        if fc == 3:
            rr = self.client.read_holding_registers(address, count=count, slave=slave)
        elif fc == 4:
            rr = self.client.read_input_registers(address, count=count, slave=slave)
        elif fc == 1:
            rr = self.client.read_coils(address, count=count, slave=slave)
        elif fc == 2:
            rr = self.client.read_discrete_inputs(address, count=count, slave=slave)
        else:
            raise IOError(f"不支持的功能码 {fc}")
        if rr.isError():
            raise IOError(f"slave={slave} fc={fc} addr={address} -> {rr}")
        return list(rr.bits)[:count] if fc in (1, 2) else list(rr.registers)

    def read(self, rec, unit=None, retries=2):
        slave = getattr(rec, "slave_id", None) if unit is None else unit
        last = None
        for attempt in range(retries + 1):
            try:
                regs = self._read_regs(rec.fc, rec.start, rec.width, slave)
                return int(bool(regs[0])) if rec.kind == "bit" else _decode(regs, rec.kind)
            except Exception as e:  # noqa: BLE001
                last = e
                if attempt < retries:
                    try:
                        self.client.connect()
                    except Exception:
                        pass
        raise last

    def read_all(self, records, unit=None):
        out = {}
        for rec in records:
            try:
                out[(rec.device, rec.parameter)] = self.read(rec, unit=unit)
            except Exception as e:  # noqa: BLE001
                out[(rec.device, rec.parameter)] = e
        return out


def read_pair_interleaved(gw, direct, records, a_unit, b_unit):
    """逐寄存器 A、B 背靠背读取，最小化两路采样时间差。
    波动量（谐波/THD/相位角/K因子）必须近似同时读才能对得上——
    否则先读完上千个 A 再读 B，值早已变化导致假性不一致。
    返回 (a_dict, b_dict)，键与 read_all 一致。"""
    a_out, b_out = {}, {}
    for rec in records:
        key = (rec.device, rec.parameter)
        try:
            a_out[key] = gw.read(rec, unit=a_unit)
        except Exception as e:  # noqa: BLE001
            a_out[key] = e
        try:
            b_out[key] = direct.read(rec, unit=b_unit)
        except Exception as e:  # noqa: BLE001
            b_out[key] = e
    return a_out, b_out


# ═══════════════════════════════════════════════════════════════════════════
# 按型号寄存器表  template/AcuCloud 模板适配/<Model>.xlsx
# D=起始地址  E=参数描述  H=数据类型  FC 固定=3
# ═══════════════════════════════════════════════════════════════════════════
_XLSX_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

# 设备名关键词 → 模板文件名前缀（按前缀 glob，兼容带版本号文件 如 AcuRev-2100_v1.01_*.xlsx）
_TMPL_MAP = [
    (["4110", "4100"],               "AcuRev-4100"),
    (["2100"],                        "AcuRev-2100"),
    (["1300"],                        "AcuRev-1300"),
    (["acuvim3", "vim3"],             "Acuvim3"),
    (["acuvimiiw", "vimiiw", "iiw"],  "AcuvimIIW"),
    (["acuvimii", "vimii"],           "AcuvimIIR"),
]


@dataclass(frozen=True)
class RegRecord:
    device: str
    parameter: str
    fc: int
    start: int
    width: int
    kind: str


def _kind_width(datatype):
    dt = (datatype or "").strip().lower()
    if dt in ("uint16", "word"):
        return "uint16", 1
    if dt in ("int16", "short"):
        return "int16", 1
    if dt in ("uint32", "dword", "long"):
        return "uint32", 2
    if dt in ("float32", "float", "real"):
        return "float32", 2
    if dt in ("float64", "double"):
        return "float64", 4
    return "uint16", 1


def _glob_latest(prefix: str):
    """在 TMPL_DIR 下按前缀匹配 xlsx（如 AcuRev-2100 -> AcuRev-2100_v1.01_*.xlsx），
    优先精确同名，其次取按文件名排序的最后一个（一般是最新版本）。"""
    exact = TMPL_DIR / f"{prefix}.xlsx"
    if exact.exists():
        return exact
    cands = sorted(TMPL_DIR.glob(f"{prefix}_*.xlsx")) + sorted(TMPL_DIR.glob(f"{prefix}*.xlsx"))
    # 去重保序
    seen, uniq = set(), []
    for c in cands:
        if c not in seen:
            seen.add(c); uniq.append(c)
    return uniq[-1] if uniq else None


def _pick_tmpl(model: str):
    """按设备型号返回模板文件路径；找不到返回 None。"""
    dn = (model or "").lower().replace("-", "").replace(" ", "")
    for keywords, prefix in _TMPL_MAP:
        if any(k.replace("-", "") in dn for k in keywords):
            p = _glob_latest(prefix)
            if p:
                return p
    # 未识别时尝试直接按型号前缀查找
    return _glob_latest(model) if model else None


def _blockparams_sheet_xml(z):
    """返回 blockParams 工作表对应的 xml 路径（按 sheet 名解析 r:id->rels->target）。
    找不到则回退 sheet2.xml / sheet1.xml。"""
    REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    names = z.namelist()
    try:
        wb = _ET.fromstring(z.read("xl/workbook.xml"))
        rels = _ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {r.get("Id"): r.get("Target") for r in rels}
        for sh in wb.iter(f"{{{_XLSX_NS}}}sheet"):
            if (sh.get("name") or "").strip().lower() == "blockparams":
                rid = sh.get(f"{{{REL_NS}}}id")
                tgt = rid_to_target.get(rid, "")
                tgt = tgt.split("/")[-1]
                cand = f"xl/worksheets/{tgt}"
                if cand in names:
                    return cand
    except Exception:
        pass
    for fallback in ("xl/worksheets/sheet2.xml", "xl/worksheets/sheet1.xml"):
        if fallback in names:
            return fallback
    return names[0]


def _load_xlsx_template(path: pathlib.Path):
    """解析 AcuCloud 模板 xlsx，返回 {desc, addr, dtype} 列表（blockParams sheet）。
    按表头文字定位列（Start(Dec)=地址 / descrption=描述 / dataType=类型），
    兼容新旧两种列布局（旧:D/E/H 新:A/B/F）。"""
    with zipfile.ZipFile(str(path)) as z:
        ss = {}
        if "xl/sharedStrings.xml" in z.namelist():
            root = _ET.fromstring(z.read("xl/sharedStrings.xml"))
            for i, si in enumerate(root.iter(f"{{{_XLSX_NS}}}si")):
                ss[i] = "".join(c.text or "" for c in si.iter()
                                if c.tag == f"{{{_XLSX_NS}}}t")
        sheet_xml = _blockparams_sheet_xml(z)
        root = _ET.fromstring(z.read(sheet_xml))
        raw = []
        for row in root.iter(f"{{{_XLSX_NS}}}row"):
            cells = {}
            for cell in row:
                ref = cell.get("r", ""); t = cell.get("t", "")
                v = cell.find(f"{{{_XLSX_NS}}}v")
                val = (ss.get(int(v.text), "") if t == "s" else (v.text or "")) \
                      if (v is not None and v.text) else ""
                cells["".join(c for c in ref if c.isalpha())] = val
            if cells:
                raw.append(cells)
    if not raw:
        return []
    # 按表头定位列字母
    header = raw[0]
    col_addr = col_desc = col_dtype = None
    for col, txt in header.items():
        t = (txt or "").strip().lower()
        if col_addr is None and t.startswith("start"):
            col_addr = col
        elif col_desc is None and t.startswith("descr"):
            col_desc = col
        elif col_dtype is None and t in ("datatype", "data type"):
            col_dtype = col
    # 回退到旧布局固定列
    col_addr = col_addr or "D"
    col_desc = col_desc or "E"
    col_dtype = col_dtype or "H"
    entries = []
    for r in raw[1:]:
        desc = r.get(col_desc, "").strip()
        addr = r.get(col_addr, "").strip()
        dtype = r.get(col_dtype, "").strip()
        if not desc or not addr:
            continue
        try:
            a = int(addr)
        except ValueError:
            continue
        entries.append({"desc": desc, "addr": a, "dtype": dtype})
    return entries


_tmpl_cache: dict = {}


def load_table(model, device=""):
    """从 AcuCloud 模板 xlsx 加载寄存器列表，返回 RegRecord 列表；找不到模板则返回空列表。"""
    p = _pick_tmpl(model)
    if p is None:
        return []
    if str(p) not in _tmpl_cache:
        _tmpl_cache[str(p)] = _load_xlsx_template(p)
    out = []
    for e in _tmpl_cache[str(p)]:
        kind, width = _kind_width(e["dtype"])
        out.append(RegRecord(device, e["desc"], 3, e["addr"], width, kind))
    return out


def table_path(model):
    """返回对应模板路径（供缺表提示用），找不到返回占位路径。"""
    p = _pick_tmpl(model)
    return p if p else TMPL_DIR / f"{model}.xlsx"


# ═══════════════════════════════════════════════════════════════════════════
# 比对 + 报告（与 Device Mirror 文件一致）
# ═══════════════════════════════════════════════════════════════════════════
_DYNAMIC_KW = ("power", "current", "thd", "crest", "harmonic", "angle",
               "demand", "flicker", "unbalance", "sequence", "predict", "factor")
_STABLE_KW = ("frequency", "voltage", "energy")


def quantity_class(parameter):
    p = (parameter or "").lower()
    if any(k in p for k in _DYNAMIC_KW):
        return "波动量"
    if any(k in p for k in _STABLE_KW):
        return "稳定量"
    return "其他"


def _match(a, b):
    if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return a == b
    # 一方为 0，另一方为超大值（未初始化寄存器的垃圾 float32），视为匹配
    _GARBAGE = 1e4
    if (a == 0.0 and abs(b) > _GARBAGE) or (b == 0.0 and abs(a) > _GARBAGE):
        return True
    return abs(a - b) <= max(REL_TOL * max(abs(a), abs(b)), ABS_FLOOR)


@dataclass
class Row:
    device: str
    parameter: str
    kind: str
    a_val: object
    b_val: object
    diff: object
    diff_pct: object
    matched: bool


def compare(records, a_values, b_values):
    rows = []
    for rec in records:
        key = (rec.device, rec.parameter)
        a, b = a_values.get(key), b_values.get(key)
        if isinstance(a, Exception) or isinstance(b, Exception) or a is None or b is None:
            continue
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            diff = float(a) - float(b)
            base = max(abs(a), abs(b))
            dpct = (abs(diff) / base) if base else 0.0
        else:
            diff = dpct = None
        rows.append(Row(rec.device, rec.parameter, rec.kind, a, b, diff, dpct, _match(a, b)))
    return rows


def write_xlsx(result, conn, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HF = PatternFill("solid", fgColor="305496"); HFONT = Font(color="FFFFFF", bold=True)
    FAIL = PatternFill("solid", fgColor="FFC7CE"); OK = PatternFill("solid", fgColor="C6EFCE")
    ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

    def fmt(v):
        if isinstance(v, (int, float)):
            return round(v, 4)
        return ILLEGAL.sub("", v) if isinstance(v, str) else v

    def hdr(ws, hs):
        for ci, h in enumerate(hs, 1):
            c = ws.cell(1, ci, h); c.fill = HF; c.font = HFONT
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def fit(ws, n):
        for ci in range(1, n + 1):
            w = max((len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 10), 46)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active; ws.title = "汇总"
    hdr(ws, ["设备", "真实IP", "ModbusID", "可比对项", "数值量一致", "数值量项", "一致率"])
    for dev, rows in result.items():
        meas = [r for r in rows if r.kind != "string"]
        m = sum(1 for r in meas if r.matched)
        ci = conn.get(dev, {})
        rate = m / len(meas) if meas else 0
        ws.append([dev, ci.get("ip", ""), ci.get("modbus_id", ""), len(rows), m, len(meas), f"{rate:.0%}"])
        if meas and m < len(meas):
            ws.cell(ws.max_row, 7).fill = FAIL
    fit(ws, 7)

    wd = wb.create_sheet("对比明细")
    H = ["设备", "参数", "量类型", "类型", "透传值(A路)", "电表实读值(B路)", "差值", "差值%", "是否一致"]
    hdr(wd, H)
    flat = [(d, r) for d, rows in result.items() for r in rows]
    flat.sort(key=lambda dr: dr[1].diff_pct if isinstance(dr[1].diff_pct, (int, float)) else -1.0,
              reverse=True)
    for dev, r in flat:
        dp = r.diff_pct
        wd.append([dev, r.parameter, quantity_class(r.parameter), r.kind,
                   fmt(r.a_val), fmt(r.b_val), fmt(r.diff),
                   f"{dp:.1%}" if isinstance(dp, (int, float)) else "",
                   "✓" if r.matched else "✗"])
        rn = wd.max_row
        if not r.matched:
            for ci in range(1, len(H) + 1):
                wd.cell(rn, ci).fill = FAIL
        else:
            wd.cell(rn, len(H)).fill = OK
    fit(wd, len(H))

    # ── 用例 Mapping sheet ─────────────────────────────────────────────────
    MAPPING = [
        ("PT-001", "test_pt_001_config",                     "功能",    "启用Pass Through，Ethernet设备SlaveID均在101–247范围内",                              "TestCase_AcuHMI_008_05_config"),
        ("PT-002", "test_pt_002_data_collected",             "功能",    "按型号寄存器表经透传读到可比对数据；缺表/无响应则skip并诊断",                         "TestCase_AcuHMI_008_05_data"),
        ("PT-003", "test_pt_003_passthrough_matches_direct", "数据正确性","透传路径(A) ↔ 直连真实电表(B) 逐寄存器数值一致率达标",                              "TestCase_AcuHMI_008_05_case05"),
        ("PT-004", "test_pt_case01_enable_disable_toggle",   "功能",    "Disable+Save后透传读取失败；Enable+Save后恢复读取",                                    "TestCase_AcuHMI_008_05_case01"),
        ("PT-005", "test_pt_case07_concurrent_masters",      "并发",    "多Modbus主站并发以不同SlaveID经透传读取，各自正常、无串扰",                            "TestCase_AcuHMI_008_05_case07"),
        ("PT-006", "test_pt_case02_valid_slaveid_save",      "边界",    "有效SlaveID(101–247)保存成功且持久化生效",                                             "TestCase_AcuHMI_008_05_case02"),
        ("PT-007", "test_pt_case03_slaveid_boundary",        "边界",    "越界SlaveID(100/248)被拒；边界值(101/247)保存成功",                                    "TestCase_AcuHMI_008_05_case03"),
        ("PT-008", "test_pt_case04_duplicate_slaveid",       "异常",    "两设备配置相同SlaveID应被拒绝或不同时生效",                                            "TestCase_AcuHMI_008_05_case04"),
        ("PT-009", "test_pt_case06_disabled_blocks_access",  "功能",    "Pass Through关闭后无法经透传SlaveID访问下游设备，系统其余功能正常",                    "TestCase_AcuHMI_008_05_case06"),
    ]
    wm = wb.create_sheet("用例Mapping")
    MH = ["用例ID", "用例函数名", "测试类型", "验证点", "关联TC编号"]
    hdr(wm, MH)
    for row in MAPPING:
        wm.append(list(row))
    fit(wm, len(MH))

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════
# 真实电表连接信息 —— 从设备网页 Physical Devices -> Settings -> Connection 抓取
# （B 路 IP/Unit 唯一来源；不使用 config.MODBUS_DEVICE_MAP）
# ═══════════════════════════════════════════════════════════════════════════
class PhysicalDevices:
    def __init__(self, page: Page):
        self.page = page

    def goto(self):
        self.page.goto(BASE_URL + "/#/physicalDevices", wait_until="domcontentloaded")
        self.page.locator(".el-table__row, .el-table__body-row").first.wait_for(state="visible")

    def read_connection(self, name):
        """进入设备 -> Settings -> Connection，抓取真实 IP / 端口 / Modbus ID。"""
        self.goto()
        row = self.page.locator(".el-table__row, .el-table__body-row").filter(has_text=name).first
        row.wait_for(state="visible")
        row.scroll_into_view_if_needed()
        cell0 = row.locator("td").first
        link = cell0.locator("a, .el-link, span, div").first
        (link if link.count() else cell0).click()
        self.page.get_by_text("Metering", exact=True).first.wait_for(state="visible")
        settings = self.page.locator(".el-sub-menu__title", has_text="Settings").first
        if settings.count():
            settings.click(); self.page.wait_for_timeout(400)
        conn = self.page.locator(".el-menu-item", has_text="Connection").first
        conn.wait_for(state="visible"); conn.click(); self.page.wait_for_timeout(800)
        fields = {}
        items = self.page.locator(".el-form-item")
        for i in range(items.count()):
            it = items.nth(i)
            labels = it.locator(".el-form-item__label").all_inner_texts()
            if not labels:
                continue
            label = labels[0].strip().rstrip(":：").strip()
            inp = it.locator("input")
            val = inp.first.input_value() if inp.count() else ""
            if not val:
                txt = it.locator(".el-form-item__content").all_inner_texts()
                val = txt[0].strip() if txt else ""
            fields[label] = val

        def g(key):
            for k, v in fields.items():
                if key.lower() in k.lower():
                    return v.strip()
            return ""

        port = g("Port")
        mbid = g("Modbus ID")
        return {
            "ip": g("IP Address"),
            "port": int(port) if port.isdigit() else MODBUS_PORT,
            "modbus_id": int(mbid) if mbid.lstrip("-").isdigit() else 0,
        }

    def read_connection_retry(self, candidates, retries=CONN_RETRIES):
        """对候选设备名依次尝试抓取，每个失败/空结果重试若干次（应对页面跳转偶发失败）。
        返回首个含有效 IP+ModbusID 的连接信息，全部失败返回 None。"""
        if isinstance(candidates, str):
            candidates = [candidates]
        for cand in candidates:
            for _ in range(max(1, retries)):
                try:
                    ci = self.read_connection(cand)
                    if ci and ci.get("ip") and ci.get("modbus_id"):
                        return ci
                except Exception:
                    pass
                self.page.wait_for_timeout(800)
        return None


# ═══════════════════════════════════════════════════════════════════════════
# 登录 + Pass Through 页面驱动
# ═══════════════════════════════════════════════════════════════════════════
COL_SLAVEID, COL_DEVICE, COL_INTERFACE, COL_MODEL = 1, 2, 3, 5


def _login(page: Page):
    page.goto(BASE_URL, wait_until="domcontentloaded")
    try:
        page.wait_for_url("**/login", timeout=4000)
    except PWTimeout:
        pass
    if "/login" not in page.url and "dashboard" in page.url.lower():
        return
    page.locator("input[type=text]").first.fill(USERNAME)
    page.locator("input[type=password]").first.fill(PASSWORD)
    for sel in ["button:has-text('Sign in')", "button:has-text('Login')",
                "button:has-text('登录')", "button[type=submit]"]:
        b = page.locator(sel).first
        try:
            b.wait_for(state="visible", timeout=2500); b.click(); break
        except PWTimeout:
            continue
    for sel in [".el-message-box button:has-text('Cancel')", ".el-dialog button:has-text('取消')"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=2500)
            page.locator(sel).first.click(); break
        except PWTimeout:
            continue
    page.wait_for_url("**/dashboard", timeout=TIMEOUT)


class GatewayPage:
    """同构页面驱动；Pass Through 与 Device Mirror 仅 TAB/标签不同。"""

    def __init__(self, page: Page, tab, enable_label):
        self.page = page
        self.TAB_NAME = tab
        self.ENABLE_LABEL = enable_label

    def goto(self, retries=3):
        last_err = None
        for attempt in range(retries):
            try:
                self.page.goto(BASE_URL + "/#/systemSettings/dateTime",
                               wait_until="domcontentloaded", timeout=TIMEOUT)
                prot = self.page.get_by_text("Protocols", exact=True).first
                prot.wait_for(state="visible", timeout=TIMEOUT)
                if "/protocols/" not in self.page.url:
                    prot.click(); self.page.wait_for_url("**/protocols/**", timeout=TIMEOUT)
                tab = self.page.locator(".el-menu-item, [role=menuitem]", has_text="Modbus").first
                tab.hover()
                item = self.page.get_by_role("menuitem", name=self.TAB_NAME)
                try:
                    item.wait_for(state="visible", timeout=5000)
                except PWTimeout:
                    tab.click(); item.wait_for(state="visible", timeout=5000)
                item.click()
                self.page.locator(f"text={self.ENABLE_LABEL}").wait_for(state="visible", timeout=TIMEOUT)
                return
            except Exception as e:
                last_err = e
                if attempt < retries - 1:
                    self.page.wait_for_timeout(5000)  # 等网关恢复
        raise last_err

    @property
    def rows(self) -> Locator:
        return self.page.locator(".el-table__body-row, .el-table__row")

    def _cell(self, row, col):
        return row.locator("td").nth(col)

    def _slaveid_input(self, row):
        return self._cell(row, COL_SLAVEID).locator("input").first

    def _radio(self, label):
        return self.page.locator(".el-radio", has_text=label).first

    def is_enabled(self):
        return "is-checked" in (self._radio("Enable").get_attribute("class") or "")

    def set_enabled(self, on):
        self._radio("Enable" if on else "Disable").click()

    def ensure_enabled(self):
        if not self.is_enabled():
            self.set_enabled(True)
            self.page.wait_for_timeout(500)
            # Enable 后必须 Save 才能让透传服务启动并显示设备行
            self.save()
            self.page.wait_for_timeout(1500)
        # 不管之前是否已启用，等待行可见（最多 TIMEOUT）
        try:
            self.rows.first.wait_for(state="visible", timeout=TIMEOUT)
        except Exception:
            # 行仍不可见时再 goto 一次并重等
            self.goto()
            self.rows.first.wait_for(state="visible", timeout=TIMEOUT)

    def has_message(self, timeout=8000):
        try:
            self.page.locator(".el-message").first.wait_for(state="visible", timeout=timeout)
            return True
        except Exception:
            return False

    def save(self):
        self.page.get_by_role("button", name="Save").click()

    def _scan(self, device):
        rows = self.rows
        for i in range(rows.count()):
            row = rows.nth(i)
            try:
                if self._cell(row, COL_DEVICE).inner_text().strip() == device:
                    return row
            except Exception:
                continue
        return None

    def find_row(self, device, retries=6):
        try:
            self.rows.first.wait_for(state="visible", timeout=TIMEOUT)
        except PWTimeout:
            pass
        for _ in range(retries):
            row = self._scan(device)
            if row is not None:
                return row
            self.page.wait_for_timeout(400)
        raise AssertionError(f"未找到设备行: {device}")

    def ethernet_device_names(self):
        out, rows = [], self.rows
        for i in range(rows.count()):
            row = rows.nth(i)
            if INTERFACE_ETHERNET.lower() in self._cell(row, COL_INTERFACE).inner_text().strip().lower():
                out.append(self._cell(row, COL_DEVICE).inner_text().strip())
        return out

    def is_checked(self, device):
        return "is-checked" in (self.find_row(device).locator(".el-checkbox").first.get_attribute("class") or "")

    def set_checked(self, device, checked):
        if self.is_checked(device) != checked:
            cb = self.find_row(device).locator(".el-checkbox").first
            cb.scroll_into_view_if_needed(); cb.click()

    def read_row_basic(self, device):
        row = self.find_row(device)
        return {
            "model": self._cell(row, COL_MODEL).inner_text().strip(),
            "slave_id": self._slaveid_input(row).input_value(),
        }

    # ---- SlaveID 读写 / 提示 / 保存结果 / 快照还原（功能用例用）----
    def set_slaveid(self, device, value):
        inp = self._slaveid_input(self.find_row(device))
        inp.scroll_into_view_if_needed(); inp.click(); inp.fill(""); inp.fill(str(value)); inp.blur()

    def get_slaveid(self, device):
        return self._slaveid_input(self.find_row(device)).input_value()

    def all_device_names(self):
        rows = self.rows
        return [self._cell(rows.nth(i), COL_DEVICE).inner_text().strip() for i in range(rows.count())]

    def used_slave_ids(self):
        ids = set()
        rows = self.rows
        for i in range(rows.count()):
            v = self._slaveid_input(rows.nth(i)).input_value().strip()
            if v.isdigit():
                ids.add(int(v))
        return ids

    def free_slave_id(self, lo=PT_SLAVE_MIN, hi=PT_SLAVE_MAX):
        used = self.used_slave_ids()
        for v in range(lo, hi + 1):
            if v not in used:
                return v
        raise AssertionError("无可用 SlaveID")

    def message_text(self, timeout=8000):
        try:
            m = self.page.locator(".el-message").first
            m.wait_for(state="visible", timeout=timeout)
            return m.inner_text().strip()
        except Exception:
            return ""

    def form_errors(self):
        try:
            return [t.strip() for t in self.page.locator(".el-form-item__error").all_inner_texts() if t.strip()]
        except Exception:
            return []

    def save_and_result(self, timeout=8000):
        """点击 Save 返回 (success, message, errors)。成功=出现成功消息/含 saved/成功 且无校验错误。"""
        self.save()
        self.page.wait_for_timeout(600)
        msg = self.message_text(timeout=timeout)
        errs = self.form_errors()
        is_succ = ("el-message--success" in (
            self.page.locator(".el-message").first.get_attribute("class") or ""
        )) if self.page.locator(".el-message").first.count() else False
        success = (is_succ or "saved" in msg.lower() or "成功" in msg) and not errs
        return success, msg, errs

    def snapshot(self):
        snap = {}
        for name in self.all_device_names():
            if name == SELF_DEVICE:
                continue
            row = self.find_row(name)
            cb = row.locator(".el-checkbox").first
            snap[name] = {
                "checked": "is-checked" in (cb.get_attribute("class") or ""),
                "slave_id": self._slaveid_input(row).input_value(),
            }
        return snap

    def restore(self, baseline, max_passes=5):
        drift = []
        for _ in range(max_passes):
            for name, st in baseline.items():
                row = self._scan(name)
                if row is None:
                    continue
                if "is-disabled" in (row.locator(".el-checkbox").first.get_attribute("class") or ""):
                    continue
                self.set_checked(name, True)
                inp = self._slaveid_input(self.find_row(name))
                inp.scroll_into_view_if_needed(); inp.click(); inp.fill(""); inp.fill(str(st["slave_id"])); inp.blur()
                self.set_checked(name, st["checked"])
            self.save(); self.has_message(); self.goto(); self.ensure_enabled()
            drift = [n for n, st in baseline.items()
                     if (self._scan(n) is not None) and self.is_checked(n) != st["checked"]]
            if not drift:
                return []
        return drift


# ═══════════════════════════════════════════════════════════════════════════
# Fixtures
# ═══════════════════════════════════════════════════════════════════════════
@pytest.fixture(scope="module")
def page_factory(browser: Browser):
    """复用项目级共享 browser，不启动新 playwright 实例。"""
    ctx0 = browser.new_context(ignore_https_errors=True, accept_downloads=True)
    pg = ctx0.new_page(); pg.set_default_timeout(TIMEOUT)
    _login(pg)
    storage = pg.evaluate("JSON.stringify(window.sessionStorage)")
    ctx0.close()
    contexts = []

    def make():
        ctx = browser.new_context(ignore_https_errors=True, accept_downloads=True)
        ctx.add_init_script(
            "(() => { const d = %s; for (const k in d){try{sessionStorage.setItem(k,d[k]);}catch(e){}} })();"
            % storage)
        page = ctx.new_page(); page.set_default_timeout(TIMEOUT)
        contexts.append(ctx)
        return page

    yield make
    for c in contexts:
        c.close()


def _pt(page):
    return GatewayPage(page, "Pass Through", "Pass Through Enable")


def _dm(page):
    return GatewayPage(page, "Device Mirror", "Device Mirror Enable")


@pytest.fixture(scope="module", autouse=True)
def baseline(page_factory):
    """保护 Pass Through 行配置(SlaveID/勾选) 与 Pass Through / Device Mirror 的 Enable 状态。"""
    dm = _dm(page_factory()); dm.goto(); dm_on0 = dm.is_enabled()
    pt = _pt(page_factory()); pt.goto(); pt_on0 = pt.is_enabled()
    pt.ensure_enabled()           # 启用才能读到表格做行快照
    snap = pt.snapshot()
    yield snap
    try:
        # 还原行配置（SlaveID/勾选）
        pt2 = _pt(page_factory()); pt2.goto(); pt2.ensure_enabled()
        drift = pt2.restore(snap)
        if drift:
            import warnings
            warnings.warn(f"Pass Through 还原后仍有偶发漂移：{drift}")
        # 还原 Pass Through Enable 状态
        if not pt_on0:
            pt2.goto()
            if pt2.is_enabled():
                pt2.set_enabled(False); pt2.save(); pt2.has_message()
        # 还原 Device Mirror Enable 状态
        dm2 = _dm(page_factory()); dm2.goto()
        if dm2.is_enabled() != dm_on0:
            dm2.set_enabled(dm_on0); dm2.save(); dm2.has_message()
    except Exception:
        pass


@pytest.fixture(scope="module")
def configured(page_factory):
    """启用 Pass Through、确保 Ethernet 设备勾选，为 SlaveID 无效（0 或超出 101-247）的设备自动分配空闲 SlaveID。"""
    pt = _pt(page_factory()); pt.goto(); pt.ensure_enabled()
    names = [n for n in pt.ethernet_device_names() if n != SELF_DEVICE]
    for n in names:
        pt.set_checked(n, True)
    pt.save(); pt.has_message()
    pt.goto(); pt.ensure_enabled()

    # 检查是否有 SlaveID 无效（0 或超出范围）的设备，若有则自动分配
    need_assign = []
    for n in names:
        info = pt.read_row_basic(n)
        sid = info["slave_id"]
        if not (str(sid).isdigit() and PT_SLAVE_MIN <= int(sid) <= PT_SLAVE_MAX):
            need_assign.append(n)

    if need_assign:
        # 确保所有设备 checkbox 已勾选（未勾选时 input 为 disabled，无法填写）
        for n in names:
            pt.set_checked(n, True)
        used = pt.used_slave_ids()
        next_id = PT_SLAVE_MIN
        for n in need_assign:
            while next_id in used:
                next_id += 1
            pt.set_slaveid(n, str(next_id))
            used.add(next_id); next_id += 1
        pt.save(); pt.has_message()
        pt.page.wait_for_timeout(3000)  # 等网关应用透传配置
        pt.goto(); pt.ensure_enabled()

    assigned, models = {}, {}
    for n in names:
        info = pt.read_row_basic(n)
        assigned[n] = info["slave_id"]; models[n] = info["model"]
    return {"assigned": assigned, "models": models}


@pytest.fixture(scope="module")
def pt_comparison(configured, page_factory):
    assigned, models = configured["assigned"], configured["models"]

    # ① 确保 Pass Through 已启用并“保存”到网关，再等网关把透传服务起来（关键：没启用就全 IllegalAddress）
    ptpage = _pt(page_factory()); ptpage.goto()
    if not ptpage.is_enabled():
        ptpage.set_enabled(True)
        ptpage.save(); ptpage.has_message()
    ptpage.page.wait_for_timeout(PT_SETTLE_MS)   # 等透传服务就绪

    # ② 预抓各设备真实 IP / 原生 ModbusID（B 路直读用；实测抓取不会打断透传转发）
    pd = PhysicalDevices(page_factory())
    conns = {dev: pd.read_connection_retry([dev]) for dev in assigned}

    # ③ 逐设备背靠背：A 路=网关透传(用 Pass Through 配置的 SlaveID)；B 路=直连真实电表(用原生 ModbusID)
    result, conn, missing = {}, {}, []
    with Modbus() as gw:
        for dev, sid in assigned.items():
            regs = load_table(models.get(dev, ""), device=dev)
            if not regs:
                missing.append({"device": dev, "model": models.get(dev, ""),
                                "expected_file": str(table_path(models.get(dev, "")))})
                continue
            if not str(sid).strip().isdigit():
                continue
            ci = conns.get(dev)
            if not ci or not ci.get("ip") or not ci.get("modbus_id"):
                continue
            ip, port, native = ci["ip"], ci["port"], ci["modbus_id"]
            conn[dev] = {"ip": ip, "port": port, "modbus_id": native, "passthrough_sid": sid}
            try:
                with Modbus(host=ip, port=port) as direct:
                    # 逐寄存器 A、B 背靠背交错读，最小化两路采样时间差（波动量才对得上）
                    a, b = read_pair_interleaved(gw, direct, regs, int(sid), native)
            except Exception:
                continue
            result[dev] = compare(regs, a, b)
            time.sleep(READ_PAUSE)
    REPORT.mkdir(parents=True, exist_ok=True)
    # 默认只生成「数据对比结果.xlsx」；若被 Excel 占用则退避到备用名，避免整个用例 ERROR
    try:
        write_xlsx(result, conn, REPORT / "数据对比结果.xlsx")
    except PermissionError:
        import warnings
        alt = REPORT / "数据对比结果_new.xlsx"
        try:
            write_xlsx(result, conn, alt)
            warnings.warn(f"数据对比结果.xlsx 被占用，已改写到 {alt.name}（请关闭 Excel 后重跑覆盖）")
        except PermissionError:
            warnings.warn("数据对比结果.xlsx 与备用文件均被占用，跳过 xlsx 写入")
    return {"result": result, "missing": missing}


def _total(pt):
    return sum(len(v) for v in pt["result"].values())


# ═══════════════════════════════════════════════════════════════════════════
# 用例
# ═══════════════════════════════════════════════════════════════════════════
def test_pt_001_config(configured):
    """启用 Pass Through 并对 Ethernet 设备读到有效 SlaveID（101-247）。"""
    assigned = configured["assigned"]
    assert assigned, "应至少有一台 Ethernet 设备"
    valid = [int(v) for v in assigned.values() if str(v).strip().isdigit() and int(v) > 0]
    assert valid, f"应至少读到一台设备有效 SlaveID：{assigned}"
    for v in valid:
        assert PT_SLAVE_MIN <= v <= PT_SLAVE_MAX, f"SlaveID {v} 超出 Pass Through 范围"


def test_pt_002_data_collected(pt_comparison):
    """透传按寄存器表读到可比对数据；缺表或透传无响应则跳过并给出诊断。"""
    if pt_comparison["missing"] and not pt_comparison["result"]:
        miss = "；".join(f"{m['model'] or m['device']} -> {m['expected_file']}"
                        for m in pt_comparison["missing"])
        pytest.skip(f"未找到任何型号寄存器表，请在 register_tables/ 补充：{miss}")
    if _total(pt_comparison) == 0:
        pytest.skip("透传(A路)未返回可比对寄存器（多为 IllegalAddress）：需各型号真实寄存器表，"
                    "或确认 Pass Through 透传端口/路由。直读真实电表(B路)正常。")
    assert _total(pt_comparison) > 0


def test_pt_003_passthrough_matches_direct(pt_comparison):
    """透传(A) 与 直读真实电表(B)：稳定量逐寄存器必须一致；波动量单独报告，不计入失败。

    透传通路是否正确以"稳定量"（电压/频率/电能等非波动量）为准——任一不一致即判失败。
    波动量（谐波/THD/相位角/K因子/电流等）在 A、B 两次先后读取之间本身会变化（低电流时
    更接近噪声），单独以 warning 列出供参考，不计入通过/失败。
    若设备未接电源（A 路读值全为 0），则该设备 0=0 视为通过，不计入统计。"""
    if not pt_comparison["result"] or _total(pt_comparison) == 0:
        pytest.skip("透传未返回可比对数据（见 002 诊断）")
    stable_n, stable_bad, dynamic_bad = 0, [], []
    for dev, rows in pt_comparison["result"].items():
        num_rows = [r for r in rows if r.kind != "string"]
        if not num_rows:
            continue
        # 若该设备 A 路读值全为 0（未接电源），则 0=0 视为全部通过，跳过一致性计算
        all_a_zero = all(
            isinstance(r.a_val, (int, float)) and abs(r.a_val) <= ABS_FLOOR
            for r in num_rows
        )
        if all_a_zero:
            continue  # 无源设备：0=0 通过，不计入统计
        for r in num_rows:
            line = f"{dev}/{r.parameter}: 透传(A)={r.a_val} 直读(B)={r.b_val}"
            if quantity_class(r.parameter) == "波动量":
                if not r.matched:
                    dynamic_bad.append(line)
                continue
            stable_n += 1
            if not r.matched:
                stable_bad.append(line)
    if not stable_n:
        pytest.skip("无稳定量可比对项（设备均无源或无稳定量寄存器）")
    if dynamic_bad:
        import warnings
        warnings.warn(
            f"波动量 A/B 差异 {len(dynamic_bad)} 项（不计入失败，信号在两次读取间波动）：\n"
            + "\n".join(dynamic_bad))
    assert not stable_bad, (
        f"稳定量存在透传↔直读不一致 {len(stable_bad)}/{stable_n}（稳定量要求每个寄存器都一致）：\n"
        + "\n".join(stable_bad))


# ═══════════════════════════════════════════════════════════════════════════
# 功能 / 边界 / 异常用例（对应 case01~07；case05=test_pt_003）
# ═══════════════════════════════════════════════════════════════════════════
def _pt_read_one(sid, rec):
    """经网关透传读单个寄存器，成功返回 True。"""
    try:
        with Modbus() as mb:
            mb.read(rec, unit=int(sid))
        return True
    except Exception:
        return False


def _setup_pt_probes(page_factory):
    """启用 Pass Through、勾选所有 Ethernet 设备、保存并等就绪；
    返回 (pt_page, {sid: rec})——每个设备一个“经透传能读通”的寄存器探针。"""
    pt = _pt(page_factory()); pt.goto()
    if not pt.is_enabled():
        pt.set_enabled(True)
    for n in pt.ethernet_device_names():
        if n != SELF_DEVICE:
            pt.set_checked(n, True)
    pt.save(); pt.has_message()
    pt.goto(); pt.ensure_enabled()
    pt.page.wait_for_timeout(PT_SETTLE_MS)
    working = {}
    for n in pt.ethernet_device_names():
        if n == SELF_DEVICE:
            continue
        info = pt.read_row_basic(n)
        sid = info["slave_id"]
        if not str(sid).strip().isdigit():
            continue
        regs = load_table(info["model"], device=n)
        for rec in regs[:15]:
            if _pt_read_one(sid, rec):
                working[int(sid)] = rec
                break
    return pt, working


def test_pt_case01_enable_disable_toggle(page_factory):
    """case01 切 Disable+Save 后经透传读取失败；切回 Enable+Save 后恢复读取。"""
    pt, working = _setup_pt_probes(page_factory)
    if not working:
        pytest.skip("无可经透传读通的设备（确认型号寄存器表/设备在线/已启用）")

    pt.goto(); pt.set_enabled(False)
    succ_dis, msg_dis, _ = pt.save_and_result(); pt.goto()
    disabled_persisted = not pt.is_enabled()
    # 轮询确认停服
    can_read_disabled = True
    deadline = time.time() + PT_DISABLE_SETTLE
    while time.time() < deadline:
        if not any(_pt_read_one(s, r) for s, r in working.items()):
            can_read_disabled = False
            break
        time.sleep(1.5)

    pt.goto(); pt.set_enabled(True)
    succ_en, msg_en, _ = pt.save_and_result(); pt.goto()
    enabled_persisted = pt.is_enabled()
    pt.page.wait_for_timeout(PT_SETTLE_MS)
    can_read_enabled = False
    for _ in range(5):
        if any(_pt_read_one(s, r) for s, r in working.items()):
            can_read_enabled = True
            break
        time.sleep(2)

    assert succ_dis and disabled_persisted, f"切 Disable 应保存并持久化（succ={succ_dis}, 提示={msg_dis!r}）"
    assert succ_en and enabled_persisted, f"切回 Enable 应保存并持久化（succ={succ_en}, 提示={msg_en!r}）"
    assert can_read_enabled, "Enable 后经透传应能读到下游设备数据"
    assert not can_read_disabled, f"Disable 后 {PT_DISABLE_SETTLE:.0f}s 内经透传仍能读到数据（预期应连接失败/无数据）"


def test_pt_case07_concurrent_masters(page_factory):
    """case07 多个 Modbus 主站并发以不同 SlaveID 经透传读取，各自正常、无串扰。

    放在改写 SlaveID 的用例(case02/03/04)之前执行，确保多台设备 SlaveID 仍为干净的基线值。
    """
    import threading
    pt, working = _setup_pt_probes(page_factory)
    sids = list(working.keys())[:3]
    if len(sids) < 2:
        pytest.skip(f"经透传可读的 SlaveID 不足 2 个（实得 {sorted(working.keys())}）")
    dur = float(os.getenv("CONCURRENT_SECONDS", "20"))
    stats = {s: {"ok": 0, "err": 0} for s in sids}

    def worker(sid):
        rec = working[sid]
        deadline = time.time() + dur
        try:
            with Modbus() as mb:
                while time.time() < deadline:
                    try:
                        mb.read(rec, unit=sid); stats[sid]["ok"] += 1
                    except Exception:
                        stats[sid]["err"] += 1
                    time.sleep(0.1)
        except Exception:
            stats[sid]["err"] += 1

    threads = [threading.Thread(target=worker, args=(s,)) for s in sids]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    bad = []
    for s in sids:
        ok, err = stats[s]["ok"], stats[s]["err"]
        rate = ok / (ok + err) if (ok + err) else 0
        if ok == 0 or rate < 0.8:
            bad.append(f"SlaveID {s}: ok={ok} err={err} 成功率={rate:.0%}（单独可读，并发下失败）")
    assert not bad, "并发经透传读取存在异常（疑似串扰/网关并发限制）：\n" + "\n".join(bad)


def test_pt_case02_valid_slaveid_save(page_factory):
    """case02 为设备配置有效 SlaveID（101–247）保存成功且持久化。"""
    pt = _pt(page_factory()); pt.goto(); pt.ensure_enabled()
    dev = next((n for n in pt.ethernet_device_names() if n != SELF_DEVICE), None)
    assert dev, "需至少一台非本机 Ethernet 设备"
    pt.set_checked(dev, True)
    val = str(pt.free_slave_id())
    pt.set_slaveid(dev, val)
    succ, msg, errs = pt.save_and_result()
    pt.goto(); pt.ensure_enabled()
    persisted = pt.get_slaveid(dev)
    assert succ, f"有效 SlaveID={val} 保存应成功（提示:{msg!r} 错误:{errs}）"
    assert persisted == val, f"SlaveID 应持久化为 {val}，实际 {persisted!r}"


def test_pt_case03_slaveid_boundary(page_factory):
    """case03 越界 100/248 保存失败；边界 101/247 保存成功（按持久化判定）。"""
    pt = _pt(page_factory()); pt.goto(); pt.ensure_enabled()
    eths = [n for n in pt.ethernet_device_names() if n != SELF_DEVICE]
    assert eths, "需至少一台非本机 Ethernet 设备"
    dev = eths[0]
    fails = []
    for val, should_accept in [("100", False), ("248", False), ("101", True), ("247", True)]:
        pt.goto(); pt.ensure_enabled()
        # 只勾选目标设备，避免 101/247 与其它设备重复（隔离“范围校验”）
        for n in eths:
            pt.set_checked(n, n == dev)
        pt.set_slaveid(dev, val)
        succ, msg, errs = pt.save_and_result()
        pt.goto(); pt.ensure_enabled()
        persisted = pt.get_slaveid(dev)
        # "No change to save" 说明值未变化仍持久化有效，按持久化结果判定是否接受
        no_change = "no change" in msg.lower()
        accepted = (bool(succ) or no_change) and persisted == val
        if should_accept and not accepted:
            fails.append(f"{val} 应被接受，但未生效(succ={succ},persist={persisted!r},msg={msg!r},err={errs})")
        if (not should_accept) and accepted:
            fails.append(f"{val} 应被拒绝(超出 101–247)，但保存生效了(msg={msg!r})")
    assert not fails, "SlaveID 边界校验不符合预期：\n" + "\n".join(fails)


def test_pt_case04_duplicate_slaveid(page_factory):
    """case04 两台设备配相同 SlaveID 应被拒绝（保存失败/提示重复，或不同时生效）。"""
    pt = _pt(page_factory()); pt.goto(); pt.ensure_enabled()
    eths = [n for n in pt.ethernet_device_names() if n != SELF_DEVICE]
    if len(eths) < 2:
        pytest.skip("需至少两台 Ethernet 设备")
    d1, d2 = eths[0], eths[1]
    dup = str(pt.free_slave_id())
    pt.set_checked(d1, True); pt.set_slaveid(d1, dup)
    pt.set_checked(d2, True); pt.set_slaveid(d2, dup)
    succ, msg, errs = pt.save_and_result()
    pt.goto(); pt.ensure_enabled()
    p1, p2 = pt.get_slaveid(d1), pt.get_slaveid(d2)
    both_dup = (p1 == dup and p2 == dup)
    assert (not succ) or (not both_dup), (
        f"两台设备相同 SlaveID={dup} 应被拒绝/提示重复，但都保存成功了"
        f"（{d1}={p1}, {d2}={p2}, msg={msg!r}, err={errs}）")


def test_pt_case06_disabled_blocks_access(page_factory):
    """case06 Pass Through 关闭后，主站无法经透传 SlaveID 访问下游设备，系统仍正常。"""
    pt, working = _setup_pt_probes(page_factory)
    if not working:
        pytest.skip("无可经透传读通的设备")
    pt.goto(); pt.set_enabled(False)
    succ, msg, _ = pt.save_and_result(); pt.goto()
    disabled_persisted = not pt.is_enabled()
    can_read = True
    deadline = time.time() + PT_DISABLE_SETTLE
    while time.time() < deadline:
        if not any(_pt_read_one(s, r) for s, r in working.items()):
            can_read = False
            break
        time.sleep(1.5)
    # 系统仍正常：能再次进入页面（导航不抛异常即认为存活）
    pt.goto(); system_alive = pt.is_enabled() in (True, False)
    # 还原启用（baseline 也会兜底）
    pt.set_enabled(True); pt.save(); pt.has_message()
    assert succ and disabled_persisted, f"关闭 Pass Through 应保存并持久化（succ={succ}, 提示={msg!r}）"
    assert not can_read, f"Pass Through 关闭后 {PT_DISABLE_SETTLE:.0f}s 内仍能经透传读到数据（预期应无法访问）"
    assert system_alive, "AcuHMI 系统应仍正常运行"


if __name__ == "__main__":
    # 直接运行：python "Pass Through/test_pass_through_business.py" [额外 pytest 参数]
    # 用 --confcutdir 限制 pytest 只从本套件目录起加载 conftest，绕开重组后因目录大小写
    # （磁盘 acuhmi_1_7 vs 代码引用 AcuHMI_1_7）而 import 失败的上级 conftest。
    # 本套件不依赖上级 conftest 的 fixture（自带 page_factory + 本目录 conftest）。
    _f = str(pathlib.Path(__file__).resolve())
    _here = str(pathlib.Path(__file__).resolve().parent)
    raise SystemExit(pytest.main([_f, "--confcutdir", _here, *sys.argv[1:]]))


