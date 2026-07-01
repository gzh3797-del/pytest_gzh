# -*- coding: utf-8 -*-
"""Device Mirror 业务数据自动化对比（单文件、可独立执行的 pytest）

验证 AcuHMI「Device Mirror」镜像透传是否忠实还原真实电表数据：
  A 路（镜像）：经网关 GATEWAY_IP:502，按 Download All 导出的镜像 SlaveID 读寄存器
  B 路（直读真实电表）：从设备网页 Physical Devices→Settings→Connection 抓取的真实 IP/Unit 直连电表读同一寄存器
  A↔B 比对（数值测量量整体一致率），逐寄存器明细落地 reports/。

仅比对 Interface=Ethernet（Modbus TCP）设备。B 路 IP/Unit 一律以网页抓取为准（不用 config 的设备表）。
网关地址来自上级 config.py 的 GATEWAY_IP；容差用 TOLERANCE_*。
运行（仓库根目录下，任意人 clone 后均可）：
  pytest "projects/AcuHMI_1_7/tests/Device Mirror" -v
也可直接运行本文件： python "projects/AcuHMI_1_7/tests/Device Mirror/test_device_mirror_business.py" -v
"""
from __future__ import annotations

import json
import os
import pathlib
import re
import struct
import sys
import time
from dataclasses import dataclass

import pytest
from playwright.sync_api import Browser, Page, Locator, TimeoutError as PWTimeout

# ── 让本文件可独立 import 上级目录的 config.py ────────────────────────────────
_ROOT = pathlib.Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
# 配置来源：优先本地 tests/config.py（开发机覆盖，gitignored）；别人 clone 没有它时，
# 回退到框架配置（configs/.env + config.yaml），保证从仓库根直接 pytest 也能开箱即跑。
import importlib.util as _ilu, types as _types

def _load_local_config():
    _cfg_path = pathlib.Path(__file__).resolve().parent.parent / "config.py"
    if _cfg_path.exists():
        _spec = _ilu.spec_from_file_location("_tests_config", _cfg_path)
        _mod = _ilu.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        return _mod
    _RR = pathlib.Path(__file__).resolve().parents[4]
    if str(_RR) not in sys.path:
        sys.path.insert(0, str(_RR))
    from projects.AcuHMI_1_7 import settings as _s
    return _types.SimpleNamespace(
        HMI_URL=_s.HMI_URL, GATEWAY_IP=_s.HMI_IP, MODBUS_PORT=_s.METER_TCP_PORT,
        HMI_USERNAME=_s.HMI_USERNAME, HMI_PASSWORD=_s.HMI_PASSWORD,
        TOLERANCE_PERCENT=_s.MODBUS_CMP_TOLERANCE_PERCENT,
        TOLERANCE_ABSOLUTE=_s.MODBUS_CMP_TOLERANCE_ABSOLUTE,
        WEB_HEADLESS=_s.HEADLESS,
        MODBUS_DEVICE_MAP=getattr(_s, "DEVICE_MODBUS_MAP", {}),
    )

config = _load_local_config()

# ── 由 config 派生的配置（只改 config.py 即可）────────────────────────────────
BASE_URL     = os.getenv("DM_BASE_URL", config.HMI_URL).rstrip("/")
# 镜像 Modbus 服务跑在 AcuHMI 网关自身（与网页同 IP），不是 config.GATEWAY_IP(192.168.2.8)。
# 默认从网页地址解析网关 IP（如 https://192.168.3.51 -> 192.168.3.51），可用 DM_GATEWAY_IP 覆盖。
from urllib.parse import urlparse as _urlparse
GATEWAY_IP   = os.getenv("DM_GATEWAY_IP", _urlparse(BASE_URL).hostname or config.GATEWAY_IP)
USERNAME     = os.getenv("DM_USERNAME", config.HMI_USERNAME)
PASSWORD     = os.getenv("DM_PASSWORD", config.HMI_PASSWORD)
MODBUS_PORT  = int(getattr(config, "MODBUS_PORT", 502) or 502)
MODBUS_TIMEOUT = float(os.getenv("MODBUS_TIMEOUT", "5"))
REL_TOL      = float(getattr(config, "TOLERANCE_PERCENT", 1.0)) / 100.0
ABS_FLOOR    = float(getattr(config, "TOLERANCE_ABSOLUTE", 0.05))
MIN_RATE     = float(os.getenv("COMPARE_MIN_RATE", "0.6"))
READ_PAUSE   = float(os.getenv("DEVICE_READ_PAUSE", "1.0"))
CONN_RETRIES = int(os.getenv("CONN_SCRAPE_RETRIES", "3"))   # 网页 Connection 抓取重试次数
DISABLE_SETTLE = float(os.getenv("DISABLE_SETTLE", "30"))   # 禁用后等网关停止供数的最长秒数(轮询)
ENABLE_SETTLE  = float(os.getenv("ENABLE_SETTLE", "10"))    # 启用+保存后等网关镜像服务起来的最长秒数(轮询)
SETTLE_POLL    = float(os.getenv("SETTLE_POLL", "2.0"))     # 就绪轮询间隔(秒)
SELF_DEVICE  = os.getenv("SELF_DEVICE_NAME", "AcuHMI-1-7")
SLAVE_MIN, SLAVE_MAX = 2, 99
HEADED       = os.getenv("HEADED", "0").strip().lower() in ("1", "true", "yes", "on")
TIMEOUT      = int(os.getenv("TIMEOUT", "15000"))
INTERFACE_ETHERNET = "Ethernet"
REPORT = pathlib.Path(__file__).resolve().parent   # 报告输出到本协议自身目录
DOWNLOAD_DIR = REPORT / "downloads"                 # 镜像配置 CSV 下载到本协议目录


# ═══════════════════════════════════════════════════════════════════════════
# Modbus TCP 读取 + 解码（pymodbus 3.x，slave= 参数）
# ═══════════════════════════════════════════════════════════════════════════
from pymodbus.client import ModbusTcpClient


def _regs_to_bytes(regs):
    return b"".join(struct.pack(">H", r & 0xFFFF) for r in regs)


def _decode(regs, kind):
    if kind == "uint16":
        return regs[0]
    if kind == "float32":
        return struct.unpack(">f", _regs_to_bytes(regs[:2]))[0]
    if kind == "float64":
        return struct.unpack(">d", _regs_to_bytes(regs[:4]))[0]
    if kind == "string":
        return _regs_to_bytes(regs).decode("ascii", errors="ignore").rstrip("\x00 ").strip()
    raise ValueError(f"未知解码类型: {kind}")


class Modbus:
    """Modbus TCP 客户端（context manager）。read_all 支持指定 unit；读取自动重连重试。"""

    def __init__(self, host=None, port=None):
        self.host = host or GATEWAY_IP
        self.port = port or MODBUS_PORT
        self.client = ModbusTcpClient(self.host, port=self.port, timeout=MODBUS_TIMEOUT)

    def __enter__(self):
        if not self.client.connect():
            raise IOError(f"无法连接 {self.host}:{self.port}")
        return self

    def __exit__(self, *exc):
        self.client.close()

    def _read_regs(self, fc, address, count, slave):
        if fc == 3:
            rr = self.client.read_holding_registers(address, count=count, slave=slave)
        elif fc == 4:
            rr = self.client.read_input_registers(address, count=count, slave=slave)
        elif fc == 1:
            rr = self.client.read_coils(address, count=count, slave=slave)
        elif fc == 2:
            rr = self.client.read_discrete_inputs(address, count=count, slave=slave)
        else:
            raise IOError(f"不支持的功能码 {fc}")
        if rr.isError():
            raise IOError(f"slave={slave} fc={fc} addr={address} -> {rr}")
        return list(rr.bits)[:count] if fc in (1, 2) else list(rr.registers)

    def read(self, rec, unit=None, retries=2):
        slave = rec.slave_id if unit is None else unit
        last = None
        for attempt in range(retries + 1):
            try:
                regs = self._read_regs(rec.fc, rec.start, rec.width, slave)
                return int(bool(regs[0])) if rec.kind == "bit" else _decode(regs, rec.kind)
            except Exception as e:  # noqa: BLE001
                last = e
                if attempt < retries:
                    try:
                        self.client.connect()
                    except Exception:
                        pass
        raise last

    def read_all(self, records, unit=None):
        out = {}
        for rec in records:
            try:
                out[(rec.device, rec.parameter)] = self.read(rec, unit=unit)
            except Exception as e:  # noqa: BLE001
                out[(rec.device, rec.parameter)] = e
        return out


def read_pair_interleaved(gw, direct, records, a_unit, b_unit):
    """逐寄存器 A、B 背靠背读取，最小化两路采样时间差。
    波动量（谐波/THD/相位角/K因子）必须近似同时读才能对得上——
    否则先读完全部 A 再读 B，值已变化导致假性不一致。
    返回 (a_dict, b_dict)，键与 read_all 一致。a_unit=None 时 A 路用各记录自身 slave_id。"""
    a_out, b_out = {}, {}
    for rec in records:
        key = (rec.device, rec.parameter)
        try:
            a_out[key] = gw.read(rec, unit=a_unit)
        except Exception as e:  # noqa: BLE001
            a_out[key] = e
        try:
            b_out[key] = direct.read(rec, unit=b_unit)
        except Exception as e:  # noqa: BLE001
            b_out[key] = e
    return a_out, b_out


# ═══════════════════════════════════════════════════════════════════════════
# MirrorConfALL.csv 解析（Download All 导出）
# ═══════════════════════════════════════════════════════════════════════════
import csv as _csv


@dataclass(frozen=True)
class ParamRecord:
    slave_id: int
    device: str
    parameter: str
    fc: int
    start: int
    end: int

    @property
    def width(self):
        return self.end - self.start + 1

    @property
    def kind(self):
        if self.fc in (1, 2):
            return "bit"
        w = self.width
        return {1: "uint16", 2: "float32", 4: "float64"}.get(w, "string")


def parse_mirror_csv(path):
    out = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in _csv.DictReader(f):
            out.append(ParamRecord(
                slave_id=int(str(row["Slave Id"]).strip()),
                device=row["Device"].strip(),
                parameter=row["Parameter"].strip(),
                fc=int(str(row["Type"]).strip()),
                start=int(str(row["Start Address(Dec.)"]).strip()),
                end=int(str(row["End Address(Dec.)"]).strip()),
            ))
    return out


# ═══════════════════════════════════════════════════════════════════════════
# 比对
# ═══════════════════════════════════════════════════════════════════════════
_DYNAMIC_KW = ("power", "current", "thd", "crest", "harmonic", "angle",
               "demand", "flicker", "unbalance", "sequence", "predict", "factor")
_STABLE_KW = ("frequency", "voltage", "energy")


def quantity_class(parameter):
    p = (parameter or "").lower()
    if any(k in p for k in _DYNAMIC_KW):
        return "波动量"
    if any(k in p for k in _STABLE_KW):
        return "稳定量"
    return "其他"


def _match(a, b):
    if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return a == b
    return abs(a - b) <= max(REL_TOL * max(abs(a), abs(b)), ABS_FLOOR)


@dataclass
class Row:
    device: str
    parameter: str
    kind: str
    a_val: object
    b_val: object
    diff: object
    diff_pct: object
    matched: bool


def compare(records, a_values, b_values):
    rows = []
    for rec in records:
        key = (rec.device, rec.parameter)
        a, b = a_values.get(key), b_values.get(key)
        if isinstance(a, Exception) or isinstance(b, Exception) or a is None or b is None:
            continue
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            diff = float(a) - float(b)
            base = max(abs(a), abs(b))
            dpct = (abs(diff) / base) if base else 0.0
        else:
            diff = dpct = None
        rows.append(Row(rec.device, rec.parameter, rec.kind, a, b, diff, dpct, _match(a, b)))
    return rows


# ═══════════════════════════════════════════════════════════════════════════
# 报告 xlsx（汇总 + 逐寄存器明细：量类型 + 按差值%降序 + 不一致标红）
# ═══════════════════════════════════════════════════════════════════════════
def write_xlsx(result, conn, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HF = PatternFill("solid", fgColor="305496"); HFONT = Font(color="FFFFFF", bold=True)
    FAIL = PatternFill("solid", fgColor="FFC7CE"); OK = PatternFill("solid", fgColor="C6EFCE")
    ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

    def fmt(v):
        if isinstance(v, (int, float)):
            return round(v, 4)
        return ILLEGAL.sub("", v) if isinstance(v, str) else v

    def hdr(ws, hs):
        for ci, h in enumerate(hs, 1):
            c = ws.cell(1, ci, h); c.fill = HF; c.font = HFONT
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def fit(ws, n):
        for ci in range(1, n + 1):
            w = max((len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 10), 46)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active; ws.title = "汇总"
    hdr(ws, ["设备", "真实IP", "ModbusID", "可比对项", "数值量一致", "数值量项", "一致率"])
    for dev, rows in result.items():
        meas = [r for r in rows if r.kind != "string"]
        m = sum(1 for r in meas if r.matched)
        ci = conn.get(dev, {})
        rate = m / len(meas) if meas else 0
        ws.append([dev, ci.get("ip", ""), ci.get("modbus_id", ""),
                   len(rows), m, len(meas), f"{rate:.0%}"])
        if meas and m < len(meas):
            ws.cell(ws.max_row, 7).fill = FAIL
    fit(ws, 7)

    wd = wb.create_sheet("对比明细")
    H = ["设备", "参数", "量类型", "类型", "网关值(A路)", "电表实读值(B路)", "差值", "差值%", "是否一致"]
    hdr(wd, H)
    flat = [(d, r) for d, rows in result.items() for r in rows]
    flat.sort(key=lambda dr: dr[1].diff_pct if isinstance(dr[1].diff_pct, (int, float)) else -1.0,
              reverse=True)
    for dev, r in flat:
        dp = r.diff_pct
        wd.append([dev, r.parameter, quantity_class(r.parameter), r.kind,
                   fmt(r.a_val), fmt(r.b_val), fmt(r.diff),
                   f"{dp:.1%}" if isinstance(dp, (int, float)) else "",
                   "✓" if r.matched else "✗"])
        rn = wd.max_row
        if not r.matched:
            for ci in range(1, len(H) + 1):
                wd.cell(rn, ci).fill = FAIL
        else:
            wd.cell(rn, len(H)).fill = OK
    fit(wd, len(H))

    # ── 用例 Mapping sheet ─────────────────────────────────────────────────
    MAPPING = [
        ("DM-001", "test_dm_001_config_and_export",          "功能",    "启用Device Mirror + 选全Ethernet设备 + 唯一SlaveID + Download All + 映射表可解析",  "TestCase_AcuHMI_008_04_case06"),
        ("DM-002", "test_dm_002_modbus_read",                "功能",    "按导出CSV经网关逐参数Modbus读取，读取成功率达标",                                   "TestCase_AcuHMI_008_04_case05(子)"),
        ("DM-003", "test_dm_003_mirror_matches_direct",      "数据正确性","镜像路径(A) ↔ 直连真实电表(B) 逐寄存器数值一致率达标",                           "TestCase_AcuHMI_008_04_case05"),
        ("DM-004", "test_dm_case01_enable_disable_toggle",   "功能",    "Disable+Save后经Mirror读取停止；Enable+Save后恢复读取",                             "TestCase_AcuHMI_008_04_case01"),
        ("DM-005", "test_dm_case02_valid_slaveid_save",      "边界",    "有效SlaveID(2–99)保存成功且持久化生效",                                             "TestCase_AcuHMI_008_04_case02"),
        ("DM-006", "test_dm_case03_slaveid_boundary",        "边界",    "越界SlaveID(1/100)被拒；边界值(2/99)保存成功",                                      "TestCase_AcuHMI_008_04_case03"),
        ("DM-007", "test_dm_case04_duplicate_slaveid",       "异常",    "两设备配置相同SlaveID应被拒绝或不同时生效",                                          "TestCase_AcuHMI_008_04_case04"),
        ("DM-008", "test_dm_case07_offline_stability",       "稳定性",  "设备离线后系统稳定、其他设备不受影响（需人工执行，自动化skip）",                     "TestCase_AcuHMI_008_04_case07"),
        ("DM-009", "test_dm_case08_concurrent_masters",      "并发",    "多Modbus主站并发读不同SlaveID，各自正常、无串扰",                                    "TestCase_AcuHMI_008_04_case08"),
    ]
    wm = wb.create_sheet("用例Mapping")
    MH = ["用例ID", "用例函数名", "测试类型", "验证点", "关联TC编号"]
    hdr(wm, MH)
    for row in MAPPING:
        wm.append(list(row))
    fit(wm, len(MH))

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════
# 真实电表连接信息 —— 从设备网页 Physical Devices -> Settings -> Connection 抓取
# （B 路 IP/Unit 唯一来源；不使用 config.MODBUS_DEVICE_MAP）
# ═══════════════════════════════════════════════════════════════════════════
class PhysicalDevices:
    def __init__(self, page: Page):
        self.page = page

    def goto(self):
        self.page.goto(BASE_URL + "/#/physicalDevices", wait_until="domcontentloaded")
        self.page.locator(".el-table__row, .el-table__body-row").first.wait_for(state="visible")

    def read_connection(self, name):
        """进入设备 -> Settings -> Connection，抓取真实 IP / 端口 / Modbus ID。"""
        self.goto()
        row = self.page.locator(".el-table__row, .el-table__body-row").filter(has_text=name).first
        row.wait_for(state="visible")
        row.scroll_into_view_if_needed()
        cell0 = row.locator("td").first
        link = cell0.locator("a, .el-link, span, div").first
        (link if link.count() else cell0).click()
        self.page.get_by_text("Metering", exact=True).first.wait_for(state="visible")
        settings = self.page.locator(".el-sub-menu__title", has_text="Settings").first
        if settings.count():
            settings.click(); self.page.wait_for_timeout(400)
        conn = self.page.locator(".el-menu-item", has_text="Connection").first
        conn.wait_for(state="visible"); conn.click(); self.page.wait_for_timeout(800)
        fields = {}
        items = self.page.locator(".el-form-item")
        for i in range(items.count()):
            it = items.nth(i)
            labels = it.locator(".el-form-item__label").all_inner_texts()
            if not labels:
                continue
            label = labels[0].strip().rstrip(":：").strip()
            inp = it.locator("input")
            val = inp.first.input_value() if inp.count() else ""
            if not val:
                txt = it.locator(".el-form-item__content").all_inner_texts()
                val = txt[0].strip() if txt else ""
            fields[label] = val

        def g(key):
            for k, v in fields.items():
                if key.lower() in k.lower():
                    return v.strip()
            return ""

        port = g("Port")
        mbid = g("Modbus ID")
        return {
            "ip": g("IP Address"),
            "port": int(port) if port.isdigit() else MODBUS_PORT,
            "modbus_id": int(mbid) if mbid.lstrip("-").isdigit() else 0,
            "_fields": fields,   # 诊断用：原始 Connection 表单字段
        }

    def read_connection_retry(self, candidates, retries=CONN_RETRIES):
        """对候选设备名依次尝试抓取，每个失败/空结果重试若干次（应对页面跳转偶发失败）。
        返回首个含有效 IP+ModbusID 的连接信息，全部失败返回 None。"""
        if isinstance(candidates, str):
            candidates = [candidates]
        for cand in candidates:
            for _ in range(max(1, retries)):
                try:
                    ci = self.read_connection(cand)
                    if ci and ci.get("ip") and ci.get("modbus_id"):
                        return ci
                except Exception:
                    pass
                self.page.wait_for_timeout(800)
        return None


# ═══════════════════════════════════════════════════════════════════════════
# 登录 + Device Mirror 页面驱动（内联）
# ═══════════════════════════════════════════════════════════════════════════
COL_SLAVEID, COL_DEVICE, COL_INTERFACE, COL_MODEL = 1, 2, 3, 5


def _login(page: Page):
    page.goto(BASE_URL, wait_until="domcontentloaded")
    try:
        page.wait_for_url("**/login", timeout=4000)
    except PWTimeout:
        pass
    if "/login" not in page.url and "dashboard" in page.url.lower():
        return
    page.locator("input[type=text]").first.fill(USERNAME)
    page.locator("input[type=password]").first.fill(PASSWORD)
    for sel in ["button:has-text('Sign in')", "button:has-text('Login')",
                "button:has-text('登录')", "button[type=submit]"]:
        b = page.locator(sel).first
        try:
            b.wait_for(state="visible", timeout=2500); b.click(); break
        except PWTimeout:
            continue
    for sel in [".el-message-box button:has-text('Cancel')", ".el-dialog button:has-text('取消')"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=2500)
            page.locator(sel).first.click(); break
        except PWTimeout:
            continue
    page.wait_for_url("**/dashboard", timeout=TIMEOUT)


class MirrorPage:
    TAB_NAME = "Device Mirror"
    ENABLE_LABEL = "Device Mirror Enable"

    def __init__(self, page: Page):
        self.page = page

    def goto(self):
        self.page.goto(BASE_URL + "/#/systemSettings/dateTime", wait_until="domcontentloaded")
        prot = self.page.get_by_text("Protocols", exact=True).first
        prot.wait_for(state="visible")
        if "/protocols/" not in self.page.url:
            prot.click(); self.page.wait_for_url("**/protocols/**", timeout=TIMEOUT)
        tab = self.page.locator(".el-menu-item, [role=menuitem]", has_text="Modbus").first
        tab.hover()
        item = self.page.get_by_role("menuitem", name=self.TAB_NAME)
        try:
            item.wait_for(state="visible", timeout=3000)
        except PWTimeout:
            tab.click(); item.wait_for(state="visible", timeout=3000)
        item.click()
        self.page.locator(f"text={self.ENABLE_LABEL}").wait_for(state="visible")

    @property
    def rows(self) -> Locator:
        return self.page.locator(".el-table__body-row, .el-table__row")

    def _cell(self, row, col):
        return row.locator("td").nth(col)

    def _slaveid_input(self, row):
        return self._cell(row, COL_SLAVEID).locator("input").first

    def _radio(self, label):
        return self.page.locator(".el-radio", has_text=label).first

    def is_enabled(self):
        return "is-checked" in (self._radio("Enable").get_attribute("class") or "")

    def set_enabled(self, on):
        self._radio("Enable" if on else "Disable").click()

    def ensure_enabled(self):
        if not self.is_enabled():
            self.set_enabled(True); self.page.wait_for_timeout(800)
            self.rows.first.wait_for(state="visible", timeout=TIMEOUT)

    def has_message(self, timeout=8000):
        try:
            self.page.locator(".el-message").first.wait_for(state="visible", timeout=timeout)
            return True
        except Exception:
            return False

    def message_text(self, timeout=8000):
        """返回顶部提示(.el-message)文本，无则空串。"""
        try:
            m = self.page.locator(".el-message").first
            m.wait_for(state="visible", timeout=timeout)
            return m.inner_text().strip()
        except Exception:
            return ""

    def form_errors(self):
        """当前可见的表单校验错误文案。"""
        try:
            return [t.strip() for t in self.page.locator(".el-form-item__error").all_inner_texts() if t.strip()]
        except Exception:
            return []

    def save(self):
        self.page.get_by_role("button", name="Save").click()

    def save_and_result(self, timeout=8000):
        """点击 Save 并返回 (success: bool, message: str, errors: list)。
        success 判定：出现成功类消息或文案含 'saved'/'成功'，且无表单校验错误。"""
        errs_before = self.form_errors()
        self.save()
        self.page.wait_for_timeout(600)
        msg = self.message_text(timeout=timeout)
        errs = self.form_errors() or errs_before
        is_succ_msg = ("el-message--success" in (
            self.page.locator(".el-message").first.get_attribute("class") or ""
        )) if self.page.locator(".el-message").first.count() else False
        success = (is_succ_msg or "saved" in msg.lower() or "成功" in msg) and not errs
        return success, msg, errs

    def download_all(self):
        with self.page.expect_download(timeout=TIMEOUT) as dl:
            self.page.get_by_role("button", name="Download All").click()
        return dl.value

    def _scan(self, device):
        rows = self.rows
        for i in range(rows.count()):
            row = rows.nth(i)
            try:
                if self._cell(row, COL_DEVICE).inner_text().strip() == device:
                    return row
            except Exception:
                continue
        return None

    def find_row(self, device, retries=6):
        try:
            self.rows.first.wait_for(state="visible", timeout=TIMEOUT)
        except PWTimeout:
            pass
        for _ in range(retries):
            row = self._scan(device)
            if row is not None:
                return row
            self.page.wait_for_timeout(400)
        raise AssertionError(f"未找到设备行: {device}")

    def all_device_names(self):
        rows = self.rows
        return [self._cell(rows.nth(i), COL_DEVICE).inner_text().strip() for i in range(rows.count())]

    def ethernet_device_names(self):
        out, rows = [], self.rows
        for i in range(rows.count()):
            row = rows.nth(i)
            if INTERFACE_ETHERNET.lower() in self._cell(row, COL_INTERFACE).inner_text().strip().lower():
                out.append(self._cell(row, COL_DEVICE).inner_text().strip())
        return out

    def is_checked(self, device):
        return "is-checked" in (self.find_row(device).locator(".el-checkbox").first.get_attribute("class") or "")

    def set_checked(self, device, checked):
        if self.is_checked(device) != checked:
            cb = self.find_row(device).locator(".el-checkbox").first
            cb.scroll_into_view_if_needed(); cb.click()

    def set_slaveid(self, device, value):
        inp = self._slaveid_input(self.find_row(device))
        inp.scroll_into_view_if_needed(); inp.click(); inp.fill(""); inp.fill(str(value)); inp.blur()

    def get_slaveid(self, device):
        return self._slaveid_input(self.find_row(device)).input_value()

    def used_slave_ids(self):
        ids = set()
        rows = self.rows
        for i in range(rows.count()):
            v = self._slaveid_input(rows.nth(i)).input_value().strip()
            if v.isdigit():
                ids.add(int(v))
        return ids

    def free_slave_id(self, lo=2, hi=99):
        used = self.used_slave_ids()
        for v in range(lo, hi + 1):
            if v not in used:
                return v
        raise AssertionError("无可用 SlaveID")

    def read_row_basic(self, device):
        row = self.find_row(device)
        cb = row.locator(".el-checkbox").first
        return {
            "model": self._cell(row, COL_MODEL).inner_text().strip(),
            "slave_id": self._slaveid_input(row).input_value(),
            "checked": "is-checked" in (cb.get_attribute("class") or ""),
            "disabled": "is-disabled" in (cb.get_attribute("class") or ""),
        }

    def enable_all_distinct(self, slave_min=SLAVE_MIN):
        self.ensure_enabled()
        assigned, nid = {}, slave_min
        for name in self.ethernet_device_names():
            if name == SELF_DEVICE:
                continue
            self.set_checked(name, True)
            self.set_slaveid(name, str(nid))
            assigned[name] = str(nid); nid += 1
        self.save(); self.has_message()
        return assigned

    def snapshot(self):
        snap = {}
        for name in self.all_device_names():
            if name == SELF_DEVICE:
                continue
            info = self.read_row_basic(name)
            snap[name] = {"checked": info["checked"], "slave_id": info["slave_id"]}
        return snap

    def restore(self, baseline, max_passes=5):
        drift = []
        for _ in range(max_passes):
            for name, st in baseline.items():
                row = self._scan(name)
                if row is None:
                    continue
                if "is-disabled" in (row.locator(".el-checkbox").first.get_attribute("class") or ""):
                    continue
                self.set_checked(name, True)
                inp = self._slaveid_input(self.find_row(name))
                inp.scroll_into_view_if_needed(); inp.click(); inp.fill(""); inp.fill(str(st["slave_id"])); inp.blur()
                self.set_checked(name, st["checked"])
            self.save(); self.has_message(); self.goto()
            drift = [n for n, st in baseline.items()
                     if (self._scan(n) is not None) and self.is_checked(n) != st["checked"]]
            if not drift:
                return []
        return drift


# ═══════════════════════════════════════════════════════════════════════════
# Fixtures
# ═══════════════════════════════════════════════════════════════════════════
@pytest.fixture(scope="module")
def page_factory(browser: Browser):
    """复用项目级共享 browser，不启动新 playwright 实例。"""
    ctx0 = browser.new_context(ignore_https_errors=True, accept_downloads=True)
    pg = ctx0.new_page(); pg.set_default_timeout(TIMEOUT)
    _login(pg)
    storage = pg.evaluate("JSON.stringify(window.sessionStorage)")
    ctx0.close()
    contexts = []

    def make():
        ctx = browser.new_context(ignore_https_errors=True, accept_downloads=True)
        ctx.add_init_script(
            "(() => { const d = %s; for (const k in d){try{sessionStorage.setItem(k,d[k]);}catch(e){}} })();"
            % storage)
        page = ctx.new_page(); page.set_default_timeout(TIMEOUT)
        contexts.append(ctx)
        return page

    yield make
    for c in contexts:
        c.close()


@pytest.fixture(scope="module", autouse=True)
def baseline(page_factory):
    dm = MirrorPage(page_factory()); dm.goto()
    enabled0 = dm.is_enabled()
    dm.ensure_enabled()        # 启用才能读到表格做快照
    snap = dm.snapshot()
    yield snap
    dm2 = MirrorPage(page_factory()); dm2.goto()
    dm2.ensure_enabled()
    drift = dm2.restore(snap)
    # 还原 Enable/Disable 开关到原始状态
    dm2.goto()
    if dm2.is_enabled() != enabled0:
        dm2.set_enabled(enabled0)
        dm2.save(); dm2.has_message()
    if drift:
        import warnings
        warnings.warn(f"Device Mirror 还原后仍有偶发漂移：{drift}")


def _wait_gateway_ready(records, timeout=ENABLE_SETTLE, poll=SETTLE_POLL):
    """启用+保存后轮询，直到网关镜像服务开始供数（任一探针读成功即就绪）。
    返回 (ready: bool, 最终成功率)。用每台设备少量探针降低开销。"""
    probes = _sample_records(records, per_device=2) or records[:10]
    deadline = time.time() + timeout
    last_rate = 0.0
    while time.time() < deadline:
        try:
            with Modbus() as mb:
                vals = mb.read_all(probes)
            ok = sum(1 for v in vals.values() if not isinstance(v, Exception))
            last_rate = ok / len(vals) if vals else 0.0
            if last_rate > 0:
                return True, last_rate
        except Exception:
            pass
        time.sleep(poll)
    return False, last_rate


@pytest.fixture(scope="module")
def exported_csv(page_factory):
    dm = MirrorPage(page_factory()); dm.goto()
    # 确保 Device Mirror 已启用：未启用则置 Enable 并保存
    if not dm.is_enabled():
        dm.set_enabled(True); dm.save(); dm.has_message()
        dm.goto()
    assigned = dm.enable_all_distinct()   # 勾选全部 Ethernet 设备 + 唯一 SlaveID + Save
    dm.goto()
    dl = dm.download_all()
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    target = DOWNLOAD_DIR / dl.suggested_filename
    dl.save_as(str(target))
    records = parse_mirror_csv(str(target))
    # 关键：等网关把镜像服务真正起来（否则全 IllegalAddress）
    ready, rate = _wait_gateway_ready(records)
    print(f"[exported_csv] 网关镜像就绪={ready} 探针成功率={rate:.0%}")
    return {"assigned": assigned, "records": records, "ready": ready}


@pytest.fixture(scope="module")
def comparison(exported_csv, page_factory):
    by_dev = {}
    for r in exported_csv["records"]:
        by_dev.setdefault(r.device, []).append(r)
    # 镜像 SlaveID -> Device Name（enable_all 返回的是页面 Device Name，用于网页定位设备）
    name_by_sid = {str(v): k for k, v in exported_csv["assigned"].items()}

    # ① 先从网页 Connection 抓各设备真实 IP/Unit（B 路唯一来源）
    pd = PhysicalDevices(page_factory())
    scraped: dict[str, dict] = {}
    for device, recs in by_dev.items():
        if device == SELF_DEVICE:
            continue
        dev_name = name_by_sid.get(str(recs[0].slave_id), device)
        # 先按 Device Name 再按 CSV 名定位行，每个候选失败/空结果自动重试
        ci = pd.read_connection_retry([dev_name, device])
        scraped[device] = {"dev_name": dev_name, "ci": ci}

    # ②' 读前再确认网关镜像服务在供数（中途若被其它用例 disable/enable 扰动则重等）
    if not _gateway_can_read(_sample_records(exported_csv["records"], per_device=2)):
        dm = MirrorPage(page_factory()); dm.goto()
        if not dm.is_enabled():
            dm.set_enabled(True); dm.save(); dm.has_message()
        _wait_gateway_ready(exported_csv["records"])

    # ② 逐设备 A 镜像读完紧接 B 直读（B 用网页抓到的 IP/Unit）
    result, conn, diag = {}, {}, {}
    with Modbus() as gw:
        for device, recs in by_dev.items():
            if device == SELF_DEVICE:
                continue
            info = scraped.get(device) or {}
            ci = info.get("ci")
            d = {"dev_name": info.get("dev_name"), "csv_records": len(recs)}
            if not ci or not ci.get("ip") or not ci.get("modbus_id"):
                d["stage"] = "网页未抓到有效 IP/ModbusID（行未找到或字段为空）"
                d["scraped_fields"] = (ci or {}).get("_fields")
                diag[device] = d
                continue
            ip, port, unit = ci["ip"], ci["port"], ci["modbus_id"]
            d.update({"scraped_ip": ip, "scraped_port": port, "scraped_modbus_id": unit,
                      "scraped_fields": ci.get("_fields")})
            conn[device] = {"ip": ip, "port": port, "modbus_id": unit,
                            "device_name": info.get("dev_name")}
            try:
                with Modbus(host=ip, port=port) as direct:
                    # 逐寄存器 A、B 背靠背交错读，最小化两路采样时间差（波动量才对得上）
                    # A 路用各记录自身镜像 slave_id（a_unit=None），B 路用网页抓到的原生 unit
                    a, b = read_pair_interleaved(gw, direct, recs, None, unit)
            except Exception as e:
                d["stage"] = f"B 路连接异常: {type(e).__name__}: {e}"
                diag[device] = d
                continue
            a_err = sum(1 for v in a.values() if isinstance(v, Exception))
            b_err = sum(1 for v in b.values() if isinstance(v, Exception))
            d["A路错误"] = f"{a_err}/{len(a)}"
            d["B路错误"] = f"{b_err}/{len(b)}"
            d["A样本错误"] = next((str(v) for v in a.values() if isinstance(v, Exception)), "")
            d["B样本错误"] = next((str(v) for v in b.values() if isinstance(v, Exception)), "")
            rows = compare(recs, a, b)
            d["可比对行"] = len(rows)
            d["stage"] = "ok" if rows else "比对为空（A 路或 B 路读取全部失败，见样本错误）"
            diag[device] = d
            result[device] = rows
            time.sleep(READ_PAUSE)
    REPORT.mkdir(parents=True, exist_ok=True)
    # 默认只生成「数据对比结果.xlsx」；若被 Excel 占用则退避到备用名，避免整个用例 ERROR
    try:
        write_xlsx(result, conn, REPORT / "数据对比结果.xlsx")
    except PermissionError:
        import warnings
        alt = REPORT / "数据对比结果_new.xlsx"
        try:
            write_xlsx(result, conn, alt)
            warnings.warn(f"数据对比结果.xlsx 被占用，已改写到 {alt.name}（请关闭 Excel 后重跑覆盖）")
        except PermissionError:
            warnings.warn("数据对比结果.xlsx 与备用文件均被占用，跳过 xlsx 写入")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# 用例
# ═══════════════════════════════════════════════════════════════════════════
def test_dm_001_config_and_export(exported_csv):
    """选中全部 Ethernet 设备 + 唯一 SlaveID + Download All，导出可解析且 SlaveID 唯一。"""
    assigned, records = exported_csv["assigned"], exported_csv["records"]
    assert assigned, "应至少启用一个 Ethernet 设备"
    assert len(set(assigned.values())) == len(assigned), f"SlaveID 应互不相同：{assigned}"
    assert records, "CSV 应解析出参数记录"


def test_dm_002_modbus_read(exported_csv):
    """按 CSV 经网关 Modbus 读取，成功率达标。"""
    with Modbus() as mb:
        results = mb.read_all(exported_csv["records"])
    ok = sum(1 for v in results.values() if not isinstance(v, Exception))
    total = len(results)
    rate = ok / total if total else 0
    assert rate >= 0.6, f"网关 Modbus 读取成功率过低 {rate:.0%}（{ok}/{total}）"


def test_dm_003_mirror_matches_direct(comparison):
    """镜像(A) 与 直读真实电表(B)：稳定量逐寄存器必须一致；波动量单独报告，不计入失败。

    镜像通路是否正确以"稳定量"（电压/频率/电能等非波动量）为准——这些量两次读取应
    完全相同（容差内），任一不一致即判失败。波动量（谐波/THD/相位角/K因子/电流等）在
    A、B 两次先后读取之间本身会变化（低电流时更接近噪声），单独以 warning 列出供参考，
    不计入通过/失败。逐寄存器明细见本目录「数据对比结果.xlsx」。
    """
    if not comparison:
        pytest.skip("无可比对设备（网页 Connection 未抓到真实 IP/ModbusID，或设备不在线）")
    stable_n, stable_bad, dynamic_bad = 0, [], []
    for dev, rows in comparison.items():
        for r in rows:
            if r.kind == "string":
                continue
            line = f"{dev}/{r.parameter}: 镜像(A)={r.a_val} 直读(B)={r.b_val}"
            if quantity_class(r.parameter) == "波动量":
                if not r.matched:
                    dynamic_bad.append(line)
                continue
            stable_n += 1
            if not r.matched:
                stable_bad.append(line)
    assert stable_n, "无稳定量可比对项"
    if dynamic_bad:
        import warnings
        warnings.warn(
            f"波动量 A/B 差异 {len(dynamic_bad)} 项（不计入失败，信号在两次读取间波动）：\n"
            + "\n".join(dynamic_bad))
    assert not stable_bad, (
        f"稳定量存在镜像↔直读不一致 {len(stable_bad)}/{stable_n}（稳定量要求每个寄存器都一致）：\n"
        + "\n".join(stable_bad))


# ═══════════════════════════════════════════════════════════════════════════
# 功能 / 边界 / 异常用例（对应 TestCase_AcuHMI_008_04_case01~08）
#   说明：case05=test_dm_003(更强：镜像↔真实电表)，case06=test_dm_001(下载+解析)
# ═══════════════════════════════════════════════════════════════════════════
def _gateway_can_read(records):
    """经网关读这些记录，任一成功即返回 True；连不上/全错返回 False。"""
    try:
        with Modbus() as mb:
            vals = mb.read_all(records)
        return any(not isinstance(v, Exception) for v in vals.values())
    except Exception:
        return False


def _sample_records(records, per_device=1):
    """每台设备取若干条非字符串(测量)记录，作为读取探针。"""
    out, cnt = [], {}
    for r in records:
        if r.kind == "string":
            continue
        if cnt.get(r.device, 0) < per_device:
            out.append(r); cnt[r.device] = cnt.get(r.device, 0) + 1
    return out


def test_dm_case01_enable_disable_toggle(exported_csv, page_factory):
    """case01 Disable+Save 后经 Mirror 读取停止；Enable+Save 后恢复读取。

    稳健化：先校验开关确实持久化；禁用后在 DISABLE_SETTLE 秒内轮询，读取停止即视为生效；
    启用后多次重试确认能读到。


    """
    probe = _sample_records(exported_csv["records"]) or exported_csv["records"][:3]
    dm = MirrorPage(page_factory()); dm.goto()

    # —— 切 Disable 并确认持久化 ——
    dm.set_enabled(False)
    succ_dis, msg_dis, _ = dm.save_and_result()
    dm.goto()
    disabled_persisted = not dm.is_enabled()
    # 轮询：禁用后等网关停止供数（最长 DISABLE_SETTLE 秒）
    still_serving = True
    deadline = time.time() + DISABLE_SETTLE
    while time.time() < deadline:
        if not _gateway_can_read(probe):
            still_serving = False
            break
        time.sleep(2)

    # —— 切回 Enable 并确认持久化 + 能读 ——
    dm.goto(); dm.set_enabled(True)
    succ_en, msg_en, _ = dm.save_and_result()
    dm.goto()
    enabled_persisted = dm.is_enabled()
    can_read_enabled = False
    for _ in range(6):
        if _gateway_can_read(probe):
            can_read_enabled = True
            break
        time.sleep(2)

    assert succ_dis and disabled_persisted, \
        f"切到 Disable 应保存并持久化（succ={succ_dis}, 仍enabled={not disabled_persisted}, 提示={msg_dis!r}）"
    assert succ_en and enabled_persisted, \
        f"切回 Enable 应保存并持久化（succ={succ_en}, 提示={msg_en!r}）"
    assert can_read_enabled, "Enable 后经 Device Mirror 应能读到数据"
    assert not still_serving, \
        f"Disable 后 {DISABLE_SETTLE:.0f}s 内经 Device Mirror 仍持续能读到数据（预期应停止供数）"


def test_dm_case02_valid_slaveid_save(page_factory):
    """case02 为设备配置有效 SlaveID（2–99）保存成功且持久化生效。"""
    dm = MirrorPage(page_factory()); dm.goto(); dm.ensure_enabled()
    dev = next((n for n in dm.ethernet_device_names() if n != SELF_DEVICE), None)
    assert dev, "需至少一台非本机 Ethernet 设备"
    dm.set_checked(dev, True)
    val = str(dm.free_slave_id())
    dm.set_slaveid(dev, val)
    succ, msg, errs = dm.save_and_result()
    dm.goto(); dm.ensure_enabled()
    persisted = dm.get_slaveid(dev)
    assert succ, f"有效 SlaveID={val} 保存应成功（提示:{msg!r} 错误:{errs}）"
    assert persisted == val, f"SlaveID 应持久化为 {val}，实际 {persisted!r}"


def test_dm_case03_slaveid_boundary(page_factory):
    """case03 越界 1/100 保存失败；边界 2/99 保存成功（按持久化结果判定，兼容客户端钳制）。"""
    dm = MirrorPage(page_factory()); dm.goto(); dm.ensure_enabled()
    dev = next((n for n in dm.ethernet_device_names() if n != SELF_DEVICE), None)
    assert dev, "需至少一台非本机 Ethernet 设备"
    fails = []
    for val, should_accept in [("1", False), ("100", False), ("2", True), ("99", True)]:
        dm.goto(); dm.ensure_enabled()
        dm.set_checked(dev, True)
        dm.set_slaveid(dev, val)
        succ, msg, errs = dm.save_and_result()
        dm.goto(); dm.ensure_enabled()
        persisted = dm.get_slaveid(dev)
        accepted = bool(succ) and persisted == val
        if should_accept and not accepted:
            fails.append(f"{val} 应被接受，但未生效(succ={succ},persist={persisted!r},msg={msg!r},err={errs})")
        if (not should_accept) and accepted:
            fails.append(f"{val} 应被拒绝(超出 2–99)，但保存生效了(msg={msg!r})")
    assert not fails, "SlaveID 边界校验不符合预期：\n" + "\n".join(fails)


def test_dm_case04_duplicate_slaveid(page_factory):
    """case04 两台设备配相同 SlaveID 应被拒绝（保存失败/提示重复，或不同时生效）。"""
    dm = MirrorPage(page_factory()); dm.goto(); dm.ensure_enabled()
    eths = [n for n in dm.ethernet_device_names() if n != SELF_DEVICE]
    if len(eths) < 2:
        pytest.skip("需至少两台 Ethernet 设备")
    d1, d2 = eths[0], eths[1]
    dup = str(dm.free_slave_id())
    dm.set_checked(d1, True); dm.set_slaveid(d1, dup)
    dm.set_checked(d2, True); dm.set_slaveid(d2, dup)
    succ, msg, errs = dm.save_and_result()
    dm.goto(); dm.ensure_enabled()
    p1, p2 = dm.get_slaveid(d1), dm.get_slaveid(d2)
    both_dup = (p1 == dup and p2 == dup)
    assert (not succ) or (not both_dup), (
        f"两台设备相同 SlaveID={dup} 应被拒绝/提示重复，但都保存成功了"
        f"（{d1}={p1}, {d2}={p2}, msg={msg!r}, err={errs}）")


@pytest.mark.skip(reason="case07 需物理断开下游设备网络模拟离线，自动化无法控制设备上下线；请人工执行")
def test_dm_case07_offline_stability():
    """case07 下游设备离线后 Mirror 数据稳定、系统不崩溃、其他设备不受影响（人工）。"""


def test_dm_case08_concurrent_masters(exported_csv):
    """case08 多个 Modbus 主站并发读不同 SlaveID，各自正常、无串扰、系统稳定。

    稳健化：先为每个 SlaveID 找一个“单独读能成功”的探针寄存器，只对这些响应正常的
    SlaveID 做并发；若某 SlaveID 单独可读但并发失败，则坐实是并发串扰/限制（断言失败）。
    """
    import threading
    recs = exported_csv["records"]
    by_sid_recs = {}
    for r in recs:
        if r.kind != "string":
            by_sid_recs.setdefault(r.slave_id, []).append(r)

    # 单独验证：每个 SlaveID 找一个能读通的寄存器
    working = {}
    for sid, rs in by_sid_recs.items():
        for rec in rs[:12]:
            try:
                with Modbus() as mb:
                    mb.read(rec, unit=sid)
                working[sid] = rec
                break
            except Exception:
                continue
    sids = list(working.keys())[:3]
    if len(sids) < 2:
        pytest.skip(f"单独读取可成功的 SlaveID 不足 2 个（实得 {sorted(working.keys())}），无法做并发对比")

    dur = float(os.getenv("CONCURRENT_SECONDS", "20"))
    stats = {s: {"ok": 0, "err": 0} for s in sids}

    def worker(sid):
        rec = working[sid]
        deadline = time.time() + dur
        try:
            with Modbus() as mb:
                while time.time() < deadline:
                    try:
                        mb.read(rec, unit=sid)
                        stats[sid]["ok"] += 1
                    except Exception:
                        stats[sid]["err"] += 1
                    time.sleep(0.1)
        except Exception:
            stats[sid]["err"] += 1

    threads = [threading.Thread(target=worker, args=(s,)) for s in sids]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    bad = []
    for s in sids:
        ok, err = stats[s]["ok"], stats[s]["err"]
        rate = ok / (ok + err) if (ok + err) else 0
        if ok == 0 or rate < 0.8:
            bad.append(f"SlaveID {s}: ok={ok} err={err} 成功率={rate:.0%}（单独可读，并发下失败）")
    assert not bad, "并发读取存在异常（疑似串扰/网关并发限制）：\n" + "\n".join(bad)


if __name__ == "__main__":
    # 直接运行：python "Device Mirror/test_device_mirror_business.py" [额外 pytest 参数]
    # 用 --confcutdir 限制 pytest 只从本套件目录起加载 conftest，绕开重组后因目录大小写
    # （磁盘 acuhmi_1_7 vs 代码引用 AcuHMI_1_7）而 import 失败的上级 conftest。
    # 本套件不依赖上级 conftest 的 fixture（自带 page_factory + 本目录 conftest）。
    _f = str(pathlib.Path(__file__).resolve())
    _here = str(pathlib.Path(__file__).resolve().parent)
    raise SystemExit(pytest.main([_f, "--confcutdir", _here, *sys.argv[1:]]))
