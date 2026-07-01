"""
快速精度测试工具 - 主界面
完全独立运行，不依赖 autotest 工程
依赖：pymodbus, pyserial, openpyxl（pip install pymodbus pyserial openpyxl）
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import queue
import logging
import os
import sys
import time
import winsound

from core.meter_comm import MeterComm
from core.source_comm import SourceComm
from core.addr_loader import load_from_excel, DEFAULT_ADDRS, WIRE_MODE_MAP
from core.test_engine import TestEngine

_LOG_LEVEL = logging.DEBUG if "--debug" in sys.argv else logging.INFO
logging.basicConfig(level=_LOG_LEVEL,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    handlers=[logging.StreamHandler(sys.stdout)])

RESULTS_DIR   = os.path.join(os.path.dirname(__file__), "results")
TEST_POINT_DIR = os.path.join(os.path.dirname(__file__), "test_point")
ADDR_DIR      = os.path.join(os.path.dirname(__file__), "modbus_addr")


def _default_addr_excel() -> str:
    """扫描 modbus_addr/ 目录，唯一 xlsx 自动选中，多个或零个返回空串"""
    if not os.path.isdir(ADDR_DIR):
        return ""
    files = [f for f in os.listdir(ADDR_DIR) if f.lower().endswith(".xlsx")]
    if len(files) == 1:
        return os.path.join(ADDR_DIR, files[0])
    return ""

WIRE_OPTIONS = ["1E2W1P", "2E3W1P", "2E3WN", "3E3WD", "3E4WY", "2E3WD", "3E4WY-P"]
CT_OPTIONS = ["100mA", "333mV", "RCT"]


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("快速精度测试工具 v1.0")
        self.resizable(True, True)
        self.minsize(1200, 700)

        # 状态对象
        self._meter: MeterComm | None = None
        self._source: SourceComm | None = None
        self._engine: TestEngine | None = None
        self._addrs: dict = dict(DEFAULT_ADDRS)
        self._rewire_event: threading.Event | None = None
        self._mon_threshold: dict[str, float] = {}   # key → accuracy ratio (0.001 = 0.1%)
        self._mon_expected: dict[str, float] = {}    # key → 当前测试点期望值（来自 test_engine）
        self._mon_queue: queue.Queue = queue.Queue(maxsize=1)   # worker → 主线程，最多积压 1 帧
        self._src_disconnect_notified = False         # 防止重复推送断开通知

        self._build_ui()
        self._load_tp_thresholds()   # 启动时按默认测点文件预加载门限
        self._start_monitor_loop()

    # ═══════════════════════════════════════════════════════
    # UI 构建
    # ═══════════════════════════════════════════════════════

    def _build_ui(self):
        pad = dict(padx=6, pady=4)

        # ── 顶部：通信设置 ────────────────────────────────
        comm_frame = ttk.LabelFrame(self, text="通信设置")
        comm_frame.pack(fill="x", **pad)

        # 电表 RTU
        ttk.Label(comm_frame, text="电表 RTU 串口:").grid(row=0, column=0, sticky="e", **pad)
        self.sv_meter_port = tk.StringVar(value="COM3")
        ttk.Entry(comm_frame, textvariable=self.sv_meter_port, width=8).grid(row=0, column=1, **pad)
        ttk.Label(comm_frame, text="波特率:").grid(row=0, column=2, sticky="e", **pad)
        self.sv_baud = tk.StringVar(value="19200")
        ttk.Combobox(comm_frame, textvariable=self.sv_baud, width=8,
                     values=["9600", "19200", "38400", "115200"]).grid(row=0, column=3, **pad)
        ttk.Label(comm_frame, text="Slave ID:").grid(row=0, column=4, sticky="e", **pad)
        self.sv_slave = tk.StringVar(value="1")
        ttk.Entry(comm_frame, textvariable=self.sv_slave, width=5).grid(row=0, column=5, **pad)
        self.btn_meter_conn = ttk.Button(comm_frame, text="连接电表", command=self._connect_meter)
        self.btn_meter_conn.grid(row=0, column=6, **pad)
        self.lbl_meter_status = ttk.Label(comm_frame, text="● 未连接", foreground="gray")
        self.lbl_meter_status.grid(row=0, column=7, **pad)

        # 源通信
        ttk.Label(comm_frame, text="源通信:").grid(row=1, column=0, sticky="e", **pad)
        self.sv_src_mode = tk.StringVar(value="TCP")
        ttk.Radiobutton(comm_frame, text="TCP/UDP", variable=self.sv_src_mode,
                        value="TCP", command=self._toggle_src_mode).grid(row=1, column=1, **pad)
        ttk.Radiobutton(comm_frame, text="串口", variable=self.sv_src_mode,
                        value="serial", command=self._toggle_src_mode).grid(row=1, column=2, **pad)
        # TCP 参数（源 IP/Port 默认与 autotest config.json 一致）
        self.frm_tcp = ttk.Frame(comm_frame)
        self.frm_tcp.grid(row=1, column=3, columnspan=2, sticky="w")
        ttk.Label(self.frm_tcp, text="IP:").pack(side="left")
        self.sv_src_ip = tk.StringVar(value="192.168.0.50")
        ttk.Entry(self.frm_tcp, textvariable=self.sv_src_ip, width=14).pack(side="left", padx=2)
        ttk.Label(self.frm_tcp, text="Port:").pack(side="left")
        self.sv_src_port = tk.StringVar(value="10003")
        ttk.Entry(self.frm_tcp, textvariable=self.sv_src_port, width=6).pack(side="left", padx=2)
        ttk.Label(self.frm_tcp, text="本地Port:").pack(side="left")
        self.sv_src_local_port = tk.StringVar(value="10005")
        ttk.Entry(self.frm_tcp, textvariable=self.sv_src_local_port, width=6).pack(side="left", padx=2)
        # 串口参数（默认显示，与 sv_src_mode="serial" 一致）
        self.frm_serial = ttk.Frame(comm_frame)
        self.frm_serial.grid(row=1, column=3, columnspan=2, sticky="w")
        ttk.Label(self.frm_serial, text="串口:").pack(side="left")
        self.sv_src_serial = tk.StringVar(value="COM19")
        ttk.Entry(self.frm_serial, textvariable=self.sv_src_serial, width=8).pack(side="left", padx=2)
        ttk.Label(self.frm_serial, text="波特率:").pack(side="left")
        self.sv_src_baud = tk.StringVar(value="9600")
        ttk.Combobox(self.frm_serial, textvariable=self.sv_src_baud, width=8,
                     values=["9600", "19200", "38400", "115200"]).pack(side="left", padx=2)
        # 默认 TCP/UDP 模式：串口面板隐藏
        self.frm_serial.grid_remove()

        self.btn_src_conn = ttk.Button(comm_frame, text="连接源", command=self._connect_source)
        self.btn_src_conn.grid(row=1, column=5, **pad)
        self.lbl_src_status = ttk.Label(comm_frame, text="● 未连接", foreground="gray")
        self.lbl_src_status.grid(row=1, column=6, **pad)

        # 地址表
        ttk.Label(comm_frame, text="Modbus 地址表:").grid(row=2, column=0, sticky="e", **pad)
        _default_addr = _default_addr_excel()
        self.sv_addr_file = tk.StringVar(value=_default_addr if _default_addr else "（使用内置默认）")
        ttk.Entry(comm_frame, textvariable=self.sv_addr_file, width=48).grid(row=2, column=1, columnspan=5, **pad)
        ttk.Button(comm_frame, text="浏览", command=self._browse_addr_file).grid(row=2, column=6, **pad)
        ttk.Button(comm_frame, text="加载", command=self._load_addr).grid(row=2, column=7, **pad)

        # ── 中部：左侧配置 + 右侧实时监控 ───────────────────
        mid = ttk.Frame(self)
        mid.pack(fill="both", expand=True, **pad)

        # 左：配置
        cfg = ttk.LabelFrame(mid, text="测试配置")
        cfg.pack(side="left", fill="y", **pad)

        ttk.Label(cfg, text="CT 类型:").grid(row=0, column=0, sticky="e", **pad)
        self.sv_ct = tk.StringVar(value="100mA")
        ttk.Combobox(cfg, textvariable=self.sv_ct, values=CT_OPTIONS, width=10, state="readonly").grid(
            row=0, column=1, **pad)

        ttk.Label(cfg, text="测点文件:").grid(row=1, column=0, sticky="e", **pad)
        self.sv_tp_file = tk.StringVar(value=self._default_tp())
        ttk.Entry(cfg, textvariable=self.sv_tp_file, width=26).grid(row=1, column=1, **pad)
        ttk.Button(cfg, text="浏览", command=self._browse_tp).grid(row=1, column=2, **pad)

        ttk.Label(cfg, text="接线方式（多选）:").grid(row=2, column=0, sticky="ne", **pad)
        self.wire_vars: dict[str, tk.BooleanVar] = {}
        wire_frame = ttk.Frame(cfg)
        wire_frame.grid(row=2, column=1, sticky="w", **pad)
        for i, wt in enumerate(WIRE_OPTIONS):
            v = tk.BooleanVar(value=True)
            self.wire_vars[wt] = v
            ttk.Checkbutton(wire_frame, text=wt, variable=v).grid(row=i // 2, column=i % 2, sticky="w")

        ttk.Label(cfg, text="源稳定等待(s):").grid(row=3, column=0, sticky="e", **pad)
        self.sv_settle = tk.StringVar(value="5")
        ttk.Entry(cfg, textvariable=self.sv_settle, width=6).grid(row=3, column=1, sticky="w", **pad)

        ttk.Label(cfg, text="采样间隔(ms):").grid(row=4, column=0, sticky="e", **pad)
        self.sv_sample_int = tk.StringVar(value="200")
        samp_int_frame = ttk.Frame(cfg)
        samp_int_frame.grid(row=4, column=1, sticky="w", **pad)
        ttk.Entry(samp_int_frame, textvariable=self.sv_sample_int, width=6).pack(side="left")
        ttk.Label(samp_int_frame, text="(20–500)", foreground="gray").pack(side="left", padx=2)

        ttk.Label(cfg, text="采样次数:").grid(row=5, column=0, sticky="e", **pad)
        self.sv_sample_cnt = tk.StringVar(value="20")
        ttk.Entry(cfg, textvariable=self.sv_sample_cnt, width=6).grid(row=5, column=1, sticky="w", **pad)

        # 操作按钮
        btn_frame = ttk.Frame(cfg)
        btn_frame.grid(row=6, column=0, columnspan=3, pady=8)
        self.btn_start   = ttk.Button(btn_frame, text="▶ 开始测试",  command=self._start_test)
        self.btn_stop    = ttk.Button(btn_frame, text="■ 停止",      command=self._stop_test,    state="disabled")
        self.btn_restart = ttk.Button(btn_frame, text="↺ 重新开始",  command=self._restart_test, state="disabled")
        self.btn_start.pack(side="left", padx=4)
        self.btn_stop.pack(side="left", padx=4)
        self.btn_restart.pack(side="left", padx=4)

        # 进度
        self.lbl_progress = ttk.Label(cfg, text="就绪")
        self.lbl_progress.grid(row=7, column=0, columnspan=3, **pad)
        self.progress = ttk.Progressbar(cfg, length=280, mode="determinate")
        self.progress.grid(row=8, column=0, columnspan=3, **pad)
        self.lbl_status = ttk.Label(cfg, text="", wraplength=280, foreground="gray")
        self.lbl_status.grid(row=9, column=0, columnspan=3, **pad)

        # 右：实时监控（九列表格）
        mon = ttk.LabelFrame(mid, text="实时监控（200ms 刷新）")
        mon.pack(side="left", fill="both", expand=True, **pad)

        self._mon_history: dict[str, list] = {}

        # 表头
        hpad = dict(padx=3, pady=2)
        _BOLD = ("", 9, "bold")
        ttk.Label(mon, text="参数",     font=_BOLD, width=13, anchor="w"     ).grid(row=0, column=0, **hpad)
        ttk.Label(mon, text="当前值",   font=_BOLD, width=10, anchor="center").grid(row=0, column=1, **hpad)
        ttk.Label(mon, text="最小值",   font=_BOLD, width=10, anchor="center").grid(row=0, column=2, **hpad)
        ttk.Label(mon, text="最大值",   font=_BOLD, width=10, anchor="center").grid(row=0, column=3, **hpad)
        ttk.Label(mon, text="平均值",   font=_BOLD, width=10, anchor="center").grid(row=0, column=4, **hpad)
        ttk.Label(mon, text="精度门限", font=_BOLD, width=9,  anchor="center").grid(row=0, column=5, **hpad)
        ttk.Label(mon, text="最小误差%",font=_BOLD, width=10, anchor="center").grid(row=0, column=6, **hpad)
        ttk.Label(mon, text="最大误差%",font=_BOLD, width=10, anchor="center").grid(row=0, column=7, **hpad)
        ttk.Label(mon, text="平均误差%",font=_BOLD, width=10, anchor="center").grid(row=0, column=8, **hpad)
        ttk.Button(mon, text="清零", width=5, command=self._reset_monitor).grid(row=0, column=9, padx=4)
        ttk.Separator(mon, orient="horizontal").grid(row=1, column=0, columnspan=10, sticky="ew", pady=1)

        fields = [
            ("freq",  "频率 (Hz)"),
            ("ua",    "Ua (V)"),     ("ub",    "Ub (V)"),     ("uc",    "Uc (V)"),
            ("ia",    "Ia (A)"),     ("ib",    "Ib (A)"),     ("ic",    "Ic (A)"),
            ("in",    "In (A)"),
            ("pa",    "Pa (kW)"),    ("qa",    "Qa (kvar)"),  ("pf_a",  "PF_a"),
            ("pb",    "Pb (kW)"),    ("qb",    "Qb (kvar)"),  ("pf_b",  "PF_b"),
            ("pc",    "Pc (kW)"),    ("qc",    "Qc (kvar)"),  ("pf_c",  "PF_c"),
            ("p_sys", "P_sys (kW)"), ("q_sys", "Q_sys (kvar)"),
        ]

        self.mon_cur:     dict[str, tk.Label] = {}
        self.mon_min:     dict[str, tk.Label] = {}
        self.mon_max:     dict[str, tk.Label] = {}
        self.mon_avg:     dict[str, tk.Label] = {}
        self.mon_thr:     dict[str, tk.Label] = {}
        self.mon_min_err: dict[str, tk.Label] = {}
        self.mon_max_err: dict[str, tk.Label] = {}
        self.mon_avg_err: dict[str, tk.Label] = {}

        def _vlbl(parent, col, row, fg="#1a5276"):
            l = tk.Label(parent, text="—", width=10, anchor="center",
                         fg=fg, bg="white", font=("Consolas", 9))
            l.grid(row=row, column=col, **hpad)
            return l

        for i, (key, label) in enumerate(fields):
            r = i + 2
            self._mon_history[key] = []
            ttk.Label(mon, text=label, width=13, anchor="w").grid(row=r, column=0, **hpad)
            self.mon_cur[key]     = _vlbl(mon, 1, r, "#1a5276")
            self.mon_min[key]     = _vlbl(mon, 2, r, "#1a6634")
            self.mon_max[key]     = _vlbl(mon, 3, r, "#7b1a1a")
            self.mon_avg[key]     = _vlbl(mon, 4, r, "#4a4a4a")
            self.mon_thr[key]     = _vlbl(mon, 5, r, "#5a5a5a")
            self.mon_min_err[key] = _vlbl(mon, 6, r, "#4a4a4a")
            self.mon_max_err[key] = _vlbl(mon, 7, r, "#4a4a4a")
            self.mon_avg_err[key] = _vlbl(mon, 8, r, "#4a4a4a")

        # keep backward-compat alias
        self.mon_labels = self.mon_cur

        # ── 底部：结果表格 ────────────────────────────────
        res_frame = ttk.LabelFrame(self, text="测试结果")
        res_frame.pack(fill="both", expand=True, **pad)

        cols = ("Case ID", "Wire", "CT",
                "Ua avg", "Ua err%", "Ua✓",
                "Ia avg", "Ia err%", "Ia✓",
                "Pa avg(kW)", "Pa err%", "Pa✓",
                "Psys avg", "Psys err%", "Psys✓",
                "Qsys avg", "Ssys avg")
        self.tree = ttk.Treeview(res_frame, columns=cols, show="headings", height=10)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=80, anchor="center")
        self.tree.column("Case ID", width=120)
        vsb = ttk.Scrollbar(res_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.tag_configure("pass", background="#C6EFCE")
        self.tree.tag_configure("fail", background="#FFC7CE")

        # 换线弹窗（初始隐藏）
        self._rewire_banner: tk.Toplevel | None = None

    # ═══════════════════════════════════════════════════════
    # 通信操作
    # ═══════════════════════════════════════════════════════

    def _connect_meter(self):
        if self._meter and self._meter.connected:
            self._meter.disconnect()
            self.lbl_meter_status.config(text="● 未连接", foreground="gray")
            self.btn_meter_conn.config(text="连接电表")
            return
        self.btn_meter_conn.config(state="disabled", text="连接中…")

        def _do():
            try:
                port  = self.sv_meter_port.get().strip()
                baud  = int(self.sv_baud.get())
                slave = int(self.sv_slave.get())
                meter = MeterComm(port=port, baudrate=baud, slave_id=slave)
                ok    = meter.connect()
                self.after(0, lambda: _upd(meter, ok, None))
            except Exception as e:
                self.after(0, lambda: _upd(None, False, e))

        def _upd(meter, ok, err):
            if err:
                self.btn_meter_conn.config(state="normal", text="连接电表")
                messagebox.showerror("错误", f"电表连接失败：{err}")
            elif ok:
                self._meter = meter
                self.lbl_meter_status.config(text="● 已连接", foreground="green")
                self.btn_meter_conn.config(state="normal", text="断开电表")
            else:
                self.lbl_meter_status.config(text="● 连接失败", foreground="red")
                self.btn_meter_conn.config(state="normal", text="连接电表")

        threading.Thread(target=_do, daemon=True).start()

    def _connect_source(self):
        if self._source and self._source.connected:
            self._source.disconnect()
            self._source = None
            self.lbl_src_status.config(text="● 未连接", foreground="gray")
            self.btn_src_conn.config(text="连接源")
            return
        self.btn_src_conn.config(state="disabled", text="连接中…")

        def _do():
            import threading as _th
            result = [None, None, None]   # [src, ok, err]
            done   = _th.Event()

            def _work():
                try:
                    mode = self.sv_src_mode.get().lower()
                    if mode == "tcp":
                        host       = self.sv_src_ip.get().strip()
                        port       = int(self.sv_src_port.get())
                        local_port = int(self.sv_src_local_port.get()) if self.sv_src_local_port.get().strip() else 10005
                        src  = SourceComm(mode="tcp", host=host, port=port, local_port=local_port)
                    else:
                        sp  = self.sv_src_serial.get().strip()
                        sb  = int(self.sv_src_baud.get())
                        src = SourceComm(mode="serial", serial_port=sp, baudrate=sb)
                    ok = src.connect()
                    result[:] = [src, ok, None]
                except Exception as e:
                    result[:] = [None, False, e]
                finally:
                    done.set()

            _th.Thread(target=_work, daemon=True).start()
            # 最多等待 5 秒，防止网络问题导致按钮永久卡死
            if not done.wait(timeout=5):
                result[:] = [None, False, TimeoutError("连接超时（5秒），请检查源的 IP/串口配置")]
            self.after(0, lambda: _upd(result[0], result[1], result[2]))

        def _upd(src, ok, err):
            if err:
                self.btn_src_conn.config(state="normal", text="连接源")
                messagebox.showerror("错误", f"源连接失败：{err}")
            elif ok:
                self._source = src
                self.lbl_src_status.config(text="● 已连接", foreground="green")
                self.btn_src_conn.config(state="normal", text="断开源")
            else:
                self.lbl_src_status.config(text="● 连接失败", foreground="red")
                self.btn_src_conn.config(state="normal", text="连接源")

        threading.Thread(target=_do, daemon=True).start()

    def _toggle_src_mode(self):
        if self.sv_src_mode.get() == "TCP":
            self.frm_serial.grid_remove()
            self.frm_tcp.grid()
        else:
            self.frm_tcp.grid_remove()
            self.frm_serial.grid()

    def _browse_addr_file(self):
        f = filedialog.askopenfilename(
            initialdir=ADDR_DIR if os.path.isdir(ADDR_DIR) else os.path.dirname(__file__),
            filetypes=[("Excel", "*.xlsx")])
        if f:
            self.sv_addr_file.set(f)

    def _load_addr(self):
        f = self.sv_addr_file.get()
        if os.path.exists(f):
            self._addrs = load_from_excel(f)
            messagebox.showinfo("成功", f"已加载 {len(self._addrs)} 个地址")
        else:
            self._addrs = dict(DEFAULT_ADDRS)
            messagebox.showinfo("提示", "文件不存在，使用内置默认地址")

    # ═══════════════════════════════════════════════════════
    # 测试控制
    # ═══════════════════════════════════════════════════════

    def _start_test(self):
        if not self._meter or not self._meter.connected:
            messagebox.showwarning("提示", "请先连接电表")
            return
        if not self._source or not self._source.connected:
            messagebox.showwarning("提示", "请先连接源")
            return

        selected_wires = [wt for wt, v in self.wire_vars.items() if v.get()]
        if not selected_wires:
            messagebox.showwarning("提示", "请至少选择一种接线方式")
            return

        tp_file = self.sv_tp_file.get()
        if not os.path.exists(tp_file):
            messagebox.showwarning("提示", f"测点文件不存在：{tp_file}")
            return

        ct_type = self.sv_ct.get()
        settle_s = float(self.sv_settle.get() or "5")

        # 验证采样参数
        try:
            sample_int = int(self.sv_sample_int.get())
            if not (20 <= sample_int <= 500):
                messagebox.showwarning("提示", "采样间隔须在 20–500 ms 之间")
                return
        except ValueError:
            messagebox.showwarning("提示", "采样间隔须为整数")
            return
        try:
            sample_cnt = int(self.sv_sample_cnt.get())
            if sample_cnt < 1:
                messagebox.showwarning("提示", "采样次数须 ≥ 1")
                return
        except ValueError:
            messagebox.showwarning("提示", "采样次数须为整数")
            return

        # 清空结果表
        for row in self.tree.get_children():
            self.tree.delete(row)

        self._engine = TestEngine(
            meter=self._meter,
            source=self._source,
            addrs=self._addrs,
            testpoint_file=tp_file,
            ct_type=ct_type,
            wire_types=selected_wires,
            results_dir=RESULTS_DIR,
            on_progress=self._on_progress,
            on_result=self._on_result,
            on_status=self._on_status,
            on_rewire=self._on_rewire,
            on_expected=self._on_expected,
            settle_s=settle_s,
            sample_cnt=sample_cnt,
            sample_int_ms=sample_int,
        )
        self._engine.start()
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_restart.config(state="normal")

    def _stop_test(self):
        if self._engine:
            self._engine.stop()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        # 保留 restart 可用（已有结果，用户可能想重跑）

    def _restart_test(self):
        """停止当前测试、清空结果、重置 UI 至就绪状态。"""
        if self._engine:
            self._engine.stop()
            self._engine = None
        # 清空结果表
        for row in self.tree.get_children():
            self.tree.delete(row)
        # 重置进度与状态
        self.progress.config(value=0)
        self.lbl_progress.config(text="就绪")
        self.lbl_status.config(text="")
        # 按钮回到初始状态
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_restart.config(state="disabled")

    # ═══════════════════════════════════════════════════════
    # 引擎回调（在主线程更新 UI）
    # ═══════════════════════════════════════════════════════

    def _on_progress(self, current: int, total: int):
        self.after(0, lambda: (
            self.progress.config(value=current / total * 100),
            self.lbl_progress.config(text=f"进度 {current}/{total}")
        ))

    def _on_status(self, msg: str):
        logging.info(msg)
        def _upd():
            self.lbl_status.config(text=msg)
            # 测试自然结束时恢复按钮状态
            if msg.startswith("✅"):
                self.btn_start.config(state="normal")
                self.btn_stop.config(state="disabled")
                self.btn_restart.config(state="normal")
                self._beep_once([(523, 200), (659, 200), (784, 300)], repeat=7)
            elif msg.startswith("⚠"):
                self.btn_start.config(state="normal")
                self.btn_stop.config(state="disabled")
                self.btn_restart.config(state="normal")
                self._beep_once([(600, 300), (400, 400)], repeat=7)
        self.after(0, _upd)

    def _on_result(self, wire: str, case_id: str, row_data: dict):
        def _upd():
            def _av(d, key="avg"):
                return f"{d[key]:.4f}" if d and d.get(key) is not None else "—"
            def _er(d):
                return f"{d['avg_err']*100:.4f}%" if d and d.get("avg_err") is not None else "—"
            def _ps(d):
                if d is None or d.get("pass") is None: return "—"
                return "✓" if d["pass"] else "✗"

            row = (
                case_id, wire, self.sv_ct.get(),
                _av(row_data.get("ua")),  _er(row_data.get("ua")),  _ps(row_data.get("ua")),
                _av(row_data.get("ia")),  _er(row_data.get("ia")),  _ps(row_data.get("ia")),
                _av(row_data.get("pa")),  _er(row_data.get("pa")),  _ps(row_data.get("pa")),
                _av(row_data.get("p_sys")), _er(row_data.get("p_sys")), _ps(row_data.get("p_sys")),
                _av(row_data.get("q_sys")), _av(row_data.get("s_sys")),
            )
            iid = self.tree.insert("", "end", values=row)
            self.tree.see(iid)
            # 着色
            all_pass = all(
                row_data.get(k, {}) and row_data[k].get("pass") is True
                for k in ("ua", "ia", "pa", "p_sys")
                if row_data.get(k) is not None
            )
            tag = "pass" if all_pass else "fail"
            self.tree.item(iid, tags=(tag,))
        self.after(0, _upd)

    def _on_expected(self, pt: dict, exp: dict):
        """test_engine 每个测试点开始采样时调用，更新监控面板的期望值"""
        mapping = {
            "ua": pt["ua"], "ub": pt["ub"], "uc": pt["uc"],
            "ia": pt["ia"], "ib": pt["ib"], "ic": pt["ic"],
            "pa": exp.get("pa", 0), "pb": exp.get("pb", 0), "pc": exp.get("pc", 0),
            "p_sys": exp.get("p_sys", 0),
            "qa": exp.get("qa", 0), "qb": exp.get("qb", 0),
            "qc": exp.get("qc", 0), "q_sys": exp.get("q_sys", 0),
        }
        self.after(0, lambda m=mapping: self._mon_expected.update(m))

    def _beep_loop_until(self, stop_event: threading.Event):
        def _play():
            while not stop_event.is_set():
                winsound.Beep(880, 250)
                stop_event.wait(0.1)
        threading.Thread(target=_play, daemon=True).start()

    def _beep_once(self, pattern: list, repeat: int = 1):
        def _play():
            for _ in range(repeat):
                for freq, ms in pattern:
                    winsound.Beep(freq, ms)
        threading.Thread(target=_play, daemon=True).start()

    def _on_rewire(self, wire_type: str, event: threading.Event):
        self._rewire_event = event
        self.after(0, lambda: self._show_rewire_dialog(wire_type, event))

    def _show_rewire_dialog(self, wire_type: str, event: threading.Event):
        self._beep_loop_until(event)
        dlg = tk.Toplevel(self)
        dlg.title("换线提示")
        dlg.grab_set()
        dlg.resizable(False, False)
        msg = (f"即将测试接线方式：{wire_type}\n\n"
               f"请完成换线后点击【确认继续】")
        ttk.Label(dlg, text=msg, wraplength=300, padding=20).pack()
        def confirm():
            event.set()
            dlg.destroy()
        ttk.Button(dlg, text="确认继续", command=confirm).pack(pady=10)
        dlg.protocol("WM_DELETE_WINDOW", confirm)

    # ═══════════════════════════════════════════════════════
    # 实时监控
    # ═══════════════════════════════════════════════════════

    def _start_monitor_loop(self):
        self._monitor_active = True
        threading.Thread(target=self._monitor_worker, daemon=True).start()
        self._poll_monitor()   # 主线程自调度消费

    def _monitor_worker(self):
        """后台线程：读 Modbus，结果放队列；满了就丢弃（主线程来不及消费时不积压）"""
        mon_keys = ["freq", "ua", "ub", "uc", "ia", "ib", "ic", "in",
                    "pa", "qa", "pf_a", "pb", "qb", "pf_b", "pc", "qc", "pf_c",
                    "p_sys", "q_sys"]
        while self._monitor_active:
            # 巡检源断开（_send_raw 失败后 _connected 已置 False）
            if self._source and not self._source.connected:
                if not self._src_disconnect_notified:
                    self._src_disconnect_notified = True
                    # 用队列特殊值通知主线程，避免直接 after() 积压
                    try:
                        self._mon_queue.put_nowait("__src_disconnected__")
                    except queue.Full:
                        pass
            else:
                self._src_disconnect_notified = False

            if self._meter and self._meter.connected:
                try:
                    addr_map = {k: self._addrs[k] for k in mon_keys if k in self._addrs}
                    vals = self._meter.read_measure_batch(addr_map)
                    try:
                        self._mon_queue.put_nowait(vals)   # 队列满则丢弃本帧，不阻塞
                    except queue.Full:
                        pass
                except Exception:
                    pass
            time.sleep(0.2)

    def _poll_monitor(self):
        """主线程每 200ms 轮询一次队列，保证 UI 更新始终在主线程执行"""
        if not self._monitor_active:
            return
        try:
            item = self._mon_queue.get_nowait()
            if item == "__src_disconnected__":
                self._on_source_disconnected()
            elif isinstance(item, dict):
                self._update_monitor(item)
        except queue.Empty:
            pass
        self.after(200, self._poll_monitor)   # 自调度，固定 200ms 间隔

    def _on_source_disconnected(self):
        """源在运行中断开时更新 UI 状态（主线程调用）"""
        self.lbl_src_status.config(text="● 已断开", foreground="red")
        self.btn_src_conn.config(text="连接源")
        if self._source:
            self._source.disconnect()
            self._source = None
        self._src_disconnect_notified = False

    def _update_monitor(self, vals: dict):
        YELLOW    = "#FFEB9C"
        LIGHT_RED = "#FFC7CE"
        NORMAL    = "white"

        def _fv(x):  return f"{x:.5f}"  if x is not None else "—"
        def _fp(x):  return f"{x*100:.4f}%" if x is not None else "—"

        for key in self._mon_history:
            v = vals.get(key)
            if v is not None:
                self._mon_history[key].append(v)
            self.mon_cur[key].config(text=_fv(v))

            hist = self._mon_history[key]
            if not hist:
                for d in (self.mon_min, self.mon_max, self.mon_avg,
                          self.mon_min_err, self.mon_max_err, self.mon_avg_err):
                    d[key].config(text="—", bg=NORMAL)
                continue

            mn  = min(hist)
            mx  = max(hist)
            avg = sum(hist) / len(hist)
            self.mon_min[key].config(text=_fv(mn))
            self.mon_max[key].config(text=_fv(mx))
            self.mon_avg[key].config(text=_fv(avg))

            thr = self._mon_threshold.get(key)

            # 误差基准：测试运行时用期望值，否则用均值近似
            ref = self._mon_expected.get(key)
            if ref is None or ref == 0:
                ref = avg if avg != 0 else None

            if ref and len(hist) >= 2:
                min_err = (mn  - ref) / abs(ref)
                max_err = (mx  - ref) / abs(ref)
                avg_err = (avg - ref) / abs(ref)

                self.mon_min_err[key].config(text=_fp(min_err))
                self.mon_max_err[key].config(text=_fp(max_err))
                self.mon_avg_err[key].config(text=_fp(avg_err))

                if thr:
                    bg_min = YELLOW    if abs(min_err) > thr else NORMAL
                    bg_max = YELLOW    if abs(max_err) > thr else NORMAL
                    bg_avg = LIGHT_RED if avg_err > thr      else NORMAL
                    self.mon_min[key].config(bg=bg_min)
                    self.mon_min_err[key].config(bg=bg_min)
                    self.mon_max[key].config(bg=bg_max)
                    self.mon_max_err[key].config(bg=bg_max)
                    self.mon_avg[key].config(bg=bg_avg)
                    self.mon_avg_err[key].config(bg=bg_avg)
            else:
                for d in (self.mon_min_err, self.mon_max_err, self.mon_avg_err):
                    d[key].config(text="—", bg=NORMAL)

    def _reset_monitor(self):
        NORMAL = "white"
        for key in self._mon_history:
            self._mon_history[key].clear()
        for d in (self.mon_min, self.mon_max, self.mon_avg,
                  self.mon_min_err, self.mon_max_err, self.mon_avg_err):
            for lbl in d.values():
                lbl.config(text="—", bg=NORMAL)

    # ═══════════════════════════════════════════════════════
    # 报告导出
    # ═══════════════════════════════════════════════════════


    # ═══════════════════════════════════════════════════════
    # 工具方法
    # ═══════════════════════════════════════════════════════

    def _default_tp(self) -> str:
        candidates = [
            os.path.join(TEST_POINT_DIR, "acuvimseries_test_case.xlsx"),
            os.path.join(TEST_POINT_DIR, "rev1320_test_case.xlsx"),
        ]
        for p in candidates:
            if os.path.exists(p):
                return p
        return ""

    def _browse_tp(self):
        f = filedialog.askopenfilename(
            initialdir=TEST_POINT_DIR,
            filetypes=[("Excel", "*.xlsx")])
        if f:
            self.sv_tp_file.set(f)
            self._load_tp_thresholds()

    def _load_tp_thresholds(self):
        """从测点 Excel 读取精度门限，更新监控表格中的门限列"""
        tp_file = self.sv_tp_file.get()
        if not tp_file or not os.path.exists(tp_file):
            return
        ct_type   = self.sv_ct.get()
        sheet_map = {"100mA": "test_case_mA", "333mV": "test_case_mV", "RCT": "test_case_rct"}
        sheet     = sheet_map.get(ct_type, "test_case_mA")
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(tp_file, data_only=True)
            if sheet not in wb.sheetnames:
                return
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                v_acc = float(row[17] or 0.001)
                i_acc = float(row[18] or 0.001)
                p_acc = float(row[20] or 0.002)
                q_acc = float(row[21] or 0.002) if len(row) > 21 and row[21] else p_acc
                s_acc = float(row[22] or 0.002) if len(row) > 22 and row[22] else p_acc
                # 电压
                for k in ("ua", "ub", "uc"):
                    self._mon_threshold[k] = v_acc
                # 电流
                for k in ("ia", "ib", "ic", "in"):
                    self._mon_threshold[k] = i_acc
                # 有功功率（各相 + 系统）
                for k in ("pa", "pb", "pc", "p_sys"):
                    self._mon_threshold[k] = p_acc
                # 无功功率（各相 + 系统）
                for k in ("qa", "qb", "qc", "q_sys"):
                    self._mon_threshold[k] = q_acc
                # 视在功率
                for k in ("s_sys",):
                    self._mon_threshold[k] = s_acc
                # PF 无专项列，沿用 p_acc 作为参考
                for k in ("pf_a", "pf_b", "pf_c", "pf_sys"):
                    self._mon_threshold[k] = p_acc
                break  # 只取第一行的门限值
        except Exception as e:
            logging.warning(f"Load thresholds: {e}")
            return

        # 刷新门限列显示
        for key, lbl in self.mon_thr.items():
            thr = self._mon_threshold.get(key)
            lbl.config(text=f"{thr*100:.3f}%" if thr is not None else "—")

    def on_closing(self):
        self._monitor_active = False          # 停止 poll_monitor 重调度
        if self._engine:
            self._engine.stop()
        if self._source:
            try: self._source.disconnect()
            except Exception: pass
        if self._meter:
            try: self._meter.disconnect()
            except Exception: pass
        self.quit()       # 退出 mainloop（必须先于 destroy，否则 after 回调可能卡住）
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
