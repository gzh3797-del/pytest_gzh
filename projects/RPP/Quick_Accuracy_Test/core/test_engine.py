"""
精度测试执行引擎（独立，不依赖 autotest 工程）
在独立线程中运行，通过回调向 GUI 汇报进度和结果
"""
import time
import logging
import threading
import os
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from core.meter_comm import MeterComm
from core.source_comm import SourceComm
from core.addr_loader import WIRE_MODE_MAP, CT_TYPE_MAP, get_measure_addr_map
from core.testpoint_reader import load_testpoints
from core import power_calc as pc


# ── 需要换线才能继续的接线方式组 ───────────────────────────
# 同一套物理接线可连续测试：1E2W1P, 2E3W1P, 2E3WN, 3E3WD
# 需要单独换线：3E4WY, 2E3WD, 3E4WY-P
REWIRE_REQUIRED = {"3E4WY", "2E3WD", "3E4WY-P"}

# 各接线方式在 Excel 报告中需要置空的列组
# Key 命名规则：
#   "{量}_all"  → 整组置空（含输入值列）
#   "{量}_meas" → 仅置空测量列（保留输入值列）
# 列组对应 _write_row 中的写入单元：
#   write_7      : ua/ub/uc/ua_angle/ub_angle/uc_angle/ia_angle/ib_angle/ic_angle
#   write_8      : pa/pb/pc/p_sys/uab/ubc/uca/in_
#   write_7_nopass: qa/qb/qc/q_sys/sa/sb/sc/s_sys
BLANK_SPEC: dict[str, set[str]] = {
    "1E2W1P": {
        "uab_all", "ubc_all", "uca_all", "in_all",
    },
    "2E3W1P": {
        "uab_meas",   # 保留 Uab 输入值，测量列置空
        "ubc_all",
    },
    "2E3WD": {
        "ua_meas", "ub_meas", "uc_meas",
        "ub_angle_meas", "uc_angle_meas",
        "ia_angle_meas", "ib_angle_meas", "ic_angle_meas",
        "pa_all", "pb_all", "pc_all", "p_sys_all",
        "qa_all", "qb_all", "qc_all", "q_sys_all",
        "sa_all", "sb_all", "sc_all",
        # s_sys 不置空（系统视在功率有效）
    },
    "2E3WN": {
        "ub_angle_meas", "ib_angle_meas",
        "uab_meas",   # 保留 Uab 输入值，测量列置空
        "ubc_all",
    },
}

SHEET_MAP = {
    "100mA": "test_case_mA",
    "333mV": "test_case_mV",
    "RCT":   "test_case_rct",
}


class TestEngine:
    def __init__(self,
                 meter: MeterComm,
                 source: SourceComm,
                 addrs: dict,
                 testpoint_file: str,
                 ct_type: str,
                 wire_types: list[str],
                 results_dir: str,
                 on_progress=None,
                 on_result=None,
                 on_status=None,
                 on_rewire=None,
                 on_expected=None,
                 settle_s: float = 5.0,
                 sample_cnt: int | None = None,
                 sample_int_ms: int | None = None):
        self.meter = meter
        self.source = source
        self.addrs = addrs
        self.testpoint_file = testpoint_file
        self.ct_type = ct_type
        self.wire_types = wire_types
        self.results_dir = results_dir
        self.on_progress = on_progress or (lambda c, t: None)
        self.on_result   = on_result   or (lambda w, cid, r: None)
        self.on_status   = on_status   or (lambda m: None)
        self.on_rewire   = on_rewire   or (lambda w, e: e.set())
        self.on_expected = on_expected or (lambda pt, exp: None)
        self.settle_s = settle_s
        # None 表示使用测点 Excel 中的值
        self._sample_cnt = sample_cnt
        self._sample_int_ms = sample_int_ms

        self._stop_event = threading.Event()
        self._thread: threading.Thread | None = None
        self.all_results: list[dict] = []

    # ── 控制 ─────────────────────────────────────────────────

    def start(self):
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop_event.set()

    def is_running(self) -> bool:
        return self._thread is not None and self._thread.is_alive()

    # ── 主流程 ────────────────────────────────────────────────

    def _run(self):
        self.on_status("测试开始")
        sheet_name = SHEET_MAP.get(self.ct_type, "test_case_mA")

        # 统计总测试点数（用于进度条）
        total_pts = 0
        pts_by_wire: dict[str, list[dict]] = {}
        for wt in self.wire_types:
            wire_int = WIRE_MODE_MAP.get(wt, 0)
            pts = load_testpoints(self.testpoint_file, sheet_name, wire_int)
            pts_by_wire[wt] = pts
            total_pts += len(pts)
        done = 0

        # 设置电表 CT 类型
        ct_val = CT_TYPE_MAP.get(self.ct_type, 0)
        self.on_status(f"写入 CT 类型寄存器：{self.ct_type}（值={ct_val}）")
        self.meter.set_ct_type(
            self.addrs["ct_ch1"],
            self.addrs["ct_ch2"],
            self.addrs["ct_ch3"],
            ct_val,
        )
        time.sleep(1)

        prev_freq: int | None = None

        for wt in self.wire_types:
            if self._stop_event.is_set():
                break

            pts = pts_by_wire[wt]
            if not pts:
                self.on_status(f"{wt}：无测试点，跳过")
                continue

            # 换线提示
            if wt in REWIRE_REQUIRED:
                ev = threading.Event()
                self.on_rewire(wt, ev)
                self.on_status(f"等待用户完成 {wt} 换线并确认…")
                ev.wait()

            # 设置接线方式
            wire_val = WIRE_MODE_MAP.get(wt, 0)
            self.on_status(f"设置接线方式：{wt}（寄存器值={wire_val}）")
            self.meter.set_wire_mode(self.addrs["wire_mode"], wire_val)
            time.sleep(1)

            # 切换源至交流界面，参照参考脚本 run_precision_measure_script 的初始化序列
            self.source.switch_screen(0x01)
            time.sleep(5)                          # 等待屏幕切换完成（参考脚本有5秒等待）
            self.source.set_gear_mode('00000000')
            # 发送初始化帧，使源进入"就绪输出"状态（参考脚本必须步骤）
            self.source.set_ac(0, 0, 0, 0, 0, 0, 0, 0, 220, 0, 0, 0, 0, settle_s=5)

            # 结果 Excel
            result_file = self._make_result_file(wt)
            wb = Workbook()
            ws = wb.active
            ws.title = wt
            self._write_header(ws, wt)

            for idx, pt in enumerate(pts):
                if self._stop_event.is_set():
                    break

                self.on_status(f"{wt} [{idx+1}/{len(pts)}] {pt['case_id']}")

                # 频率切换 → 软重启
                freq_hz = int(pt["freq"])
                freq_reg_val = 0 if freq_hz == 50 else 1
                if prev_freq is not None and prev_freq != freq_reg_val:
                    self.on_status(f"频率变更 {50 if prev_freq==0 else 60}Hz → {freq_hz}Hz，软重启中…")
                    self.meter.set_freq_mode(self.addrs["freq_set"], freq_hz)
                    self.meter.soft_reboot(self.addrs["reboot"])
                    time.sleep(8)
                prev_freq = freq_reg_val

                # 计算预期功率
                expected = self._calc_expected(pt, wt)

                # 驱动源输出
                try:
                    # 按接线方式修正源输出相角（与参考脚本各 fast_precision_measure_by_xxx 一致）
                    uc_p, ub_p, ua_p = pt["uc_p"], pt["ub_p"], pt["ua_p"]
                    ic_p, ib_p, ia_p = pt["ic_p"], pt["ib_p"], pt["ia_p"]
                    if wt == "1E2W1P":
                        # 只有 A 相接线，B/C 相角度强制清零，避免源输出无关信号
                        ub_p = uc_p = ib_p = ic_p = 0.0
                    elif wt in ("2E3W1P", "2E3WN"):
                        # B 相仅作参考，B 相电压/电流相角强制为 0
                        ub_p = ib_p = 0.0

                    self.source.set_voltage_gear(pt["uc"], pt["ub"], pt["ua"])
                    self.source.set_current_gear(pt["ic"], pt["ib"], pt["ia"])
                    time.sleep(0.5)   # 档位切换后留时间让源稳定再输出
                    self.source.set_ac(
                        uc_p, ub_p, ua_p,
                        ic_p, ib_p, ia_p,
                        pt["uc"],   pt["ub"],   pt["ua"],
                        pt["ic"],   pt["ib"],   pt["ia"],
                        float(freq_hz),
                        settle_s=self.settle_s,
                    )
                except Exception as src_err:
                    logging.error(f"源输出失败 [{pt['case_id']}]: {src_err}")
                    self.on_status(f"⚠ 源命令失败，跳过 {pt['case_id']}: {src_err}")
                    done += 1
                    self.on_progress(done, total_pts)
                    continue

                # 采样（优先使用界面输入值，否则用测点 Excel 中的值）
                n   = self._sample_cnt    if self._sample_cnt    is not None else pt["sample_cnt"]
                ims = self._sample_int_ms if self._sample_int_ms is not None else pt["sample_int_ms"]
                addr_map = get_measure_addr_map(self.addrs, wt)
                samples = self.meter.sample_n_times(addr_map, n, ims)

                # 精度计算
                row_data = self._calc_row(pt, expected, samples, wt)
                self.all_results.append({"wire": wt, "case_id": pt["case_id"], **row_data})
                # 把当前测试点期望值推送给 UI 监控面板，用于正确计算误差
                self.on_expected(pt, expected)
                self.on_result(wt, pt["case_id"], row_data)

                # 写入 Excel
                self._write_row(ws, idx + 2, pt, expected, row_data, wt)
                wb.save(result_file)

                done += 1
                self.on_progress(done, total_pts)

            # 关源
            self.source.set_zero(settle_s=2)
            self.source.switch_screen(0x00)

        self.on_status("✅ 测试完成" if not self._stop_event.is_set() else "⚠ 测试已停止")

    # ── 计算预期功率 ──────────────────────────────────────────

    def _calc_expected(self, pt: dict, wt: str) -> dict:
        ua, ub, uc = pt["ua"], pt["ub"], pt["uc"]
        ia, ib, ic = pt["ia"], pt["ib"], pt["ic"]
        ua_p, ub_p, uc_p = pt["ua_p"], pt["ub_p"], pt["uc_p"]
        ia_p, ib_p, ic_p = pt["ia_p"], pt["ib_p"], pt["ic_p"]

        uab, ubc, uca = pc.line_to_line_voltage(ua, ub, uc, ua_p, ub_p, uc_p)
        pa, qa, sa = pc.phase_power(ua, ia, ua_p, ia_p)
        pb, qb, sb = pc.phase_power(ub, ib, ub_p, ib_p)
        pcc, qc, sc = pc.phase_power(uc, ic, uc_p, ic_p)

        if wt in ("3E4WY", "3E4WY-P"):
            p_sys = pc.sys_power(pa, pb, pcc)
            q_sys = pc.sys_power(qa, qb, qc)
            s_sys = pc.sys_power(sa, sb, sc)
        elif wt == "2E3W1P":
            p_sys = pc.sys_power(pa, pcc)
            q_sys = pc.sys_power(qa, qc)
            s_sys = pc.sys_power(sa, sc)
        elif wt == "1E2W1P":
            p_sys = pa; q_sys = qa; s_sys = sa
        elif wt == "2E3WD":
            p_sys, q_sys, s_sys = pc.calc_2e3wd_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p)
        elif wt == "3E3WD":
            p_sys, q_sys, s_sys = pc.calc_3e4wd_power(ua, ub, uc, ua_p, ub_p, uc_p, ia, ib, ic, ia_p, ib_p, ic_p)
        elif wt == "2E3WN":
            # 两元件接线：仅 A+C 两相计入系统功率（与参考脚本 fast_precision_measure_by_2e3wn 一致）
            p_sys = pc.sys_power(pa, pcc)
            q_sys = pc.sys_power(qa, qc)
            s_sys = pc.sys_power(sa, sc)
        else:
            p_sys = pc.sys_power(pa, pb, pcc)
            q_sys = pc.sys_power(qa, qb, qc)
            s_sys = pc.sys_power(sa, sb, sc)

        return {
            "pa": pa, "pb": pb, "pc": pcc, "p_sys": p_sys,
            "qa": qa, "qb": qb, "qc": qc, "q_sys": q_sys,
            "sa": sa, "sb": sb, "sc": sc, "s_sys": s_sys,
            "uab": uab, "ubc": ubc, "uca": uca,
        }

    # ── 精度汇总 ──────────────────────────────────────────────

    def _calc_row(self, pt: dict, exp: dict, samples: dict[str, list], wt: str) -> dict:
        M = self.meter  # 精度计算静态方法
        v_acc = pt["v_acc"]
        i_acc = pt["i_acc"]
        ang_acc = pt["angle_acc"]
        p_acc = pt["p_acc"]

        row = {}
        # 电压
        for name in ("ua", "ub", "uc"):
            std = pt[name]
            row[name] = M.calc_accuracy(std, samples[name], v_acc) if std > 0 else None
        # 电流
        for name in ("ia", "ib", "ic"):
            std = pt[name]
            row[name] = M.calc_accuracy(std, samples[name], i_acc) if std > 0 else None
        # 相角
        for name in ("ua_angle", "ub_angle", "uc_angle", "ia_angle", "ib_angle", "ic_angle"):
            base = name.replace("_angle", "")
            std_ang = pt[f"{base}_p"]
            row[name] = M.calc_phase_angle_accuracy(std_ang, samples[name], ang_acc)
        # 线电压
        for name, std_key in [("uab", "uab"), ("ubc", "ubc"), ("uca", "uca")]:
            std = exp[std_key]
            row[name] = M.calc_accuracy(std, samples[name], v_acc) if std > 0 else None
        # 有功
        for name in ("pa", "pb", "pc", "p_sys"):
            std = exp.get(name, 0)
            row[name] = M.calc_accuracy(std, samples[name], p_acc) if std != 0 else None
        # 无功（无阈值）
        for name in ("qa", "qb", "qc", "q_sys"):
            row[name] = M.calc_accuracy_no_threshold(exp.get(name, 0), samples[name])
        # 视在（无阈值）
        for name in ("sa", "sb", "sc", "s_sys"):
            row[name] = M.calc_accuracy_no_threshold(exp.get(name, 0), samples[name])
        return row

    # ── Excel 输出 ────────────────────────────────────────────

    def _make_result_file(self, wire_type: str) -> str:
        os.makedirs(self.results_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = f"Precision_Measure_{wire_type}_{ts}.xlsx"
        return os.path.join(self.results_dir, name)

    def _write_header(self, ws, wire_type: str):
        # ── 输入参数列（固定 17 列）──
        headers = [
            "用例编号", "voltage_accuracy", "current_accuracy",
            "phase_angle_accuracy", "active_power_accuracy",
            "A相电压输入值", "B相电压输入值", "C相电压输入值",
            "A相电流输入值", "B相电流输入值", "C相电流输入值",
            "A相电压相位输入值", "B相电压相位输入值", "C相电压相位输入值",
            "A相电流相位输入值", "B相电流相位输入值", "C相电流相位输入值",
        ]
        # ── 每个物理量的 7 列模板（无输入值）──
        def _cols7(name):
            return [f"{name}最小值", f"{name}最小误差",
                    f"{name}最大值", f"{name}最大误差",
                    f"{name}平均值", f"{name}平均误差",
                    f"{name}精度测试结果"]
        # ── 每个物理量的 8 列模板（含输入值）──
        def _cols8(name):
            return [f"{name}输入值"] + _cols7(name)
        # ── 无 pass 的 7 列（Q/S）──
        def _cols7_nopass(name):
            return [f"{name}输入值",
                    f"{name}最小值", f"{name}最小误差",
                    f"{name}最大值", f"{name}最大误差",
                    f"{name}平均值", f"{name}平均误差"]

        for ph in ("A相电压", "B相电压", "C相电压"):
            headers += _cols7(ph)
        for ph in ("A相电流", "B相电流", "C相电流"):
            headers += _cols7(ph)
        for ph in ("A相电压相位", "B相电压相位", "C相电压相位",
                   "A相电流相位", "B相电流相位", "C相电流相位"):
            headers += _cols7(ph)
        for ph in ("A相有功功率", "B相有功功率", "C相有功功率", "系统有功功率"):
            headers += _cols8(ph)
        for ph in ("A相无功功率", "B相无功功率", "C相无功功率", "系统无功功率"):
            headers += _cols7_nopass(ph)
        for ph in ("A相视在功率", "B相视在功率", "C相视在功率", "系统视在功率"):
            headers += _cols7_nopass(ph)
        for ph in ("AB线电压", "BC线电压", "CA线电压"):
            headers += _cols8(ph)
        headers += _cols8("N相电流")

        hdr_fill = PatternFill("solid", fgColor="4F81BD")
        hdr_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    def _write_row(self, ws, row_num: int, pt: dict, exp: dict, row_data: dict, wire_type: str = ""):
        """写一行数据，超门限着色；按接线方式对无效列置空"""
        RED    = PatternFill("solid", fgColor="FFC7CE")
        YELLOW = PatternFill("solid", fgColor="FFEB9C")
        GREEN  = PatternFill("solid", fgColor="C6EFCE")

        col = 1
        blanks = BLANK_SPEC.get(wire_type, set())

        def wc(value, fill=None):
            nonlocal col
            cell = ws.cell(row=row_num, column=col, value=value)
            if fill:
                cell.fill = fill
            col += 1

        def blank(n):
            for _ in range(n): wc("")

        # ── 输入参数（17列，始终写入）──
        wc(pt["case_id"])
        wc(pt["v_acc"]); wc(pt["i_acc"]); wc(pt["angle_acc"]); wc(pt["p_acc"])
        wc(pt["ua"]); wc(pt["ub"]); wc(pt["uc"])
        wc(pt["ia"]); wc(pt["ib"]); wc(pt["ic"])
        wc(pt["ua_p"]); wc(pt["ub_p"]); wc(pt["uc_p"])
        wc(pt["ia_p"]); wc(pt["ib_p"]); wc(pt["ic_p"])

        v_thr = pt["v_acc"]
        i_thr = pt["i_acc"]
        a_thr = pt["angle_acc"]
        p_thr = pt["p_acc"]

        # ── 内部写入函数（不含置空判断）──

        def _write7(d, thr):
            """7列有阈值：min/err max/err avg/err result"""
            if d is None:
                blank(7); return
            mn_e, mx_e, av_e = d["min_err"], d["max_err"], d["avg_err"]
            min_over = thr is not None and abs(mn_e) > thr
            max_over = thr is not None and abs(mx_e) > thr
            avg_over = thr is not None and abs(av_e) > thr
            wc(d["min"],   YELLOW if min_over else None)
            wc(mn_e,       YELLOW if min_over else None)
            wc(d["max"],   YELLOW if max_over else None)
            wc(mx_e,       YELLOW if max_over else None)
            wc(d["avg"],   RED    if avg_over else None)
            wc(av_e,       RED    if avg_over else None)
            wc("Passed" if d["pass"] else "Failed",
               GREEN if d["pass"] else RED)

        def _write7np(d):
            """6列无阈值：min/err max/err avg/err（无 result 列）"""
            if d is None:
                blank(6); return
            wc(d["min"]); wc(d["min_err"])
            wc(d["max"]); wc(d["max_err"])
            wc(d["avg"]); wc(d["avg_err"])

        # ── 带置空的写入代理 ──

        def w7(key, d, thr):
            """write_7 代理：置空时写 7 个空"""
            if f"{key}_all" in blanks or f"{key}_meas" in blanks:
                blank(7)
            else:
                _write7(d, thr)

        def w8(key, exp_val, d, thr):
            """write_8 代理（输入值+7列）：
            _all  → 全部 8 个空
            _meas → 保留输入值，测量 7 列置空"""
            if f"{key}_all" in blanks:
                blank(8)
            elif f"{key}_meas" in blanks:
                wc(exp_val); blank(7)
            else:
                wc(exp_val); _write7(d, thr)

        def w7np(key, exp_val, d):
            """write_7_nopass 代理（输入值+6列）：
            _all  → 全部 7 个空
            _meas → 保留输入值，测量 6 列置空"""
            if f"{key}_all" in blanks:
                blank(7)
            elif f"{key}_meas" in blanks:
                wc(exp_val); blank(6)
            else:
                wc(exp_val); _write7np(d)

        # ── 按列顺序写入（与表头对应）──

        # 相电压（各 7 列）
        w7("ua", row_data.get("ua"), v_thr)
        w7("ub", row_data.get("ub"), v_thr)
        w7("uc", row_data.get("uc"), v_thr)
        # 相电流（各 7 列）
        w7("ia", row_data.get("ia"), i_thr)
        w7("ib", row_data.get("ib"), i_thr)
        w7("ic", row_data.get("ic"), i_thr)
        # 电压相角（各 7 列）
        w7("ua_angle", row_data.get("ua_angle"), a_thr)
        w7("ub_angle", row_data.get("ub_angle"), a_thr)
        w7("uc_angle", row_data.get("uc_angle"), a_thr)
        # 电流相角（各 7 列）
        w7("ia_angle", row_data.get("ia_angle"), a_thr)
        w7("ib_angle", row_data.get("ib_angle"), a_thr)
        w7("ic_angle", row_data.get("ic_angle"), a_thr)
        # 有功（各 8 列）
        w8("pa",    exp.get("pa",    0), row_data.get("pa"),    p_thr)
        w8("pb",    exp.get("pb",    0), row_data.get("pb"),    p_thr)
        w8("pc",    exp.get("pc",    0), row_data.get("pc"),    p_thr)
        w8("p_sys", exp.get("p_sys", 0), row_data.get("p_sys"), p_thr)
        # 无功（各 7 列无 pass）
        w7np("qa",    exp.get("qa",    0), row_data.get("qa"))
        w7np("qb",    exp.get("qb",    0), row_data.get("qb"))
        w7np("qc",    exp.get("qc",    0), row_data.get("qc"))
        w7np("q_sys", exp.get("q_sys", 0), row_data.get("q_sys"))
        # 视在（各 7 列无 pass）
        w7np("sa",    exp.get("sa",    0), row_data.get("sa"))
        w7np("sb",    exp.get("sb",    0), row_data.get("sb"))
        w7np("sc",    exp.get("sc",    0), row_data.get("sc"))
        w7np("s_sys", exp.get("s_sys", 0), row_data.get("s_sys"))
        # 线电压（各 8 列）
        w8("uab", exp.get("uab", 0), row_data.get("uab"), v_thr)
        w8("ubc", exp.get("ubc", 0), row_data.get("ubc"), v_thr)
        w8("uca", exp.get("uca", 0), row_data.get("uca"), v_thr)
        # N 相电流（8 列）
        w8("in_", pt.get("in_expected", pt["ia"]), row_data.get("in"), i_thr)
