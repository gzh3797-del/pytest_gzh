"""xlsx test-point I/O for the energy-meter auto-test.

Public API
----------
find_columns(headers)                          -- pure: map logical names -> col index
list_sheets(path)                              -- return all worksheet names
read_test_points(path, sheet, ct_type)         -- read voltage/current/pf/pc rows
write_results(path, results, sheet)            -- write min_s/max_s/avg_s back
"""
from __future__ import annotations

import re
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalize(value) -> str:
    """Stringify, strip, lower, collapse any whitespace (incl. newlines) to single space."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"\s+", " ", text)


# ---------------------------------------------------------------------------
# Public: list_sheets
# ---------------------------------------------------------------------------

def list_testpoint_sheets(path: str) -> list[str]:
    """Return only the runnable test-point worksheet names from an xlsx file.

    A sheet qualifies when BOTH conditions hold:
    1. Name filter: the sheet name contains "测点" and does NOT contain "分析".
    2. Structural filter: ``_find_header_row`` succeeds on the sheet (i.e. a row
       with voltage, current, and power_factor columns exists).

    If no sheet passes both tests the function falls back to ``list_sheets(path)``
    so the GUI dropdown is never empty for an unrecognised template.

    Parameters
    ----------
    path : str
        Path to the xlsx file.

    Returns
    -------
    list[str]
        Filtered (or fallback) worksheet names.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_names: list[str] = list(wb.sheetnames)

    qualified: list[str] = []
    for name in all_names:
        # --- Name filter ---
        if "测点" not in name or "分析" in name:
            continue
        # --- Structural filter ---
        ws = wb[name]
        try:
            _find_header_row(ws)
            qualified.append(name)
        except ValueError:
            pass

    wb.close()
    return qualified if qualified else all_names


def list_sheets(path: str) -> list[str]:
    """Return all worksheet names in the workbook, in order.

    Parameters
    ----------
    path : str
        Path to the xlsx file.

    Returns
    -------
    list[str]
        Worksheet names suitable for a GUI dropdown.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ---------------------------------------------------------------------------
# Public: find_columns
# ---------------------------------------------------------------------------

def find_columns(headers: list) -> dict:
    """Map logical column names to 0-based indices for all present columns.

    Parameters
    ----------
    headers : list
        Cell values (str or None) from one worksheet row.

    Returns
    -------
    dict
        Subset of keys that were found:
        "voltage", "current", "power_factor", "pulse_constant",
        "min_s", "max_s", "avg_s".
    """
    result: dict = {}
    for idx, raw in enumerate(headers):
        norm = _normalize(raw)
        if not norm:
            continue

        # Exact-match result columns first (prevent accidental substring hits).
        if norm in ("mim(s)", "min(s)"):
            result["min_s"] = idx
            continue
        if norm == "max(s)":
            result["max_s"] = idx
            continue
        if norm == "avg(s)":
            result["avg_s"] = idx
            continue

        # Substring matches for the remaining read columns.
        if "ct类型" in norm or "ct type" in norm:
            result["ct_type"] = idx
        elif "voltage" in norm:
            result["voltage"] = idx
        elif "current" in norm:
            result["current"] = idx
        elif "power factor" in norm:
            result["power_factor"] = idx
        elif "pulse constant" in norm:
            result["pulse_constant"] = idx

    return result


# ---------------------------------------------------------------------------
# Internal: locate header row in a worksheet
# ---------------------------------------------------------------------------

def _find_header_row(ws) -> tuple[int, dict]:
    """Return (1-based row number, column_map) for the first row that has
    at least voltage, current, and power_factor columns.

    Raises ValueError if no such row is found.
    """
    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        col_map = find_columns(values)
        if {"voltage", "current", "power_factor"}.issubset(col_map):
            return row[0].row, col_map
    raise ValueError(
        "No header row found: could not locate columns for voltage, current,"
        " and power_factor in any row of the worksheet."
    )


# ---------------------------------------------------------------------------
# Public: read_test_points
# ---------------------------------------------------------------------------

def read_test_points(
    path: str,
    sheet: Optional[str] = None,
    ct_type: Optional[str] = None,
) -> list[dict]:
    """Read test-point rows from an xlsx workbook.

    The header row is auto-detected as the first row whose find_columns()
    result contains at least "voltage", "current", and "power_factor".
    Data rows follow the header; reading stops at the first row whose
    voltage cell is empty/None.

    Parameters
    ----------
    path : str
        Path to the xlsx file.
    sheet : str, optional
        Worksheet name.  If None, the active sheet is used.
    ct_type : str, optional
        If given and not in (None, "", "all", "全部"), only include rows
        whose CT类型 cell normalized-equals this value (case-insensitive).
        If the sheet has no CT类型 column, this filter is silently ignored.

    Returns
    -------
    list[dict]
        Each dict: {"row": int, "voltage": float, "current": float,
                    "power_factor": float, "pulse_constant": int/float/None}.

    Raises
    ------
    ValueError
        If no header row is found.
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet] if sheet else wb.active

    header_row_num, col_map = _find_header_row(ws)

    # Column indices are 0-based from find_columns; openpyxl cells are 1-based.
    v_col = col_map["voltage"] + 1
    # Current column: 0-based index of the "current" header cell.
    a_col_0 = col_map["current"]          # 0-based
    a_col_1 = a_col_0 + 1                 # 1-based (the header column itself)
    pf_col_0 = col_map["power_factor"]    # 0-based
    pf_col = pf_col_0 + 1                 # 1-based
    pc_col = col_map.get("pulse_constant")
    pc_col_1 = (pc_col + 1) if pc_col is not None else None
    ct_col = col_map.get("ct_type")
    ct_col_1 = (ct_col + 1) if ct_col is not None else None

    # Determine whether we need to apply a ct_type filter.
    _ct_filter: Optional[str]
    if ct_type is None or ct_type in ("", "all", "全部"):
        _ct_filter = None
    else:
        _ct_filter = _normalize(ct_type)

    points: list[dict] = []
    current_ct: Optional[str] = None   # forward-filled CT类型 value
    for row_num in range(header_row_num + 1, ws.max_row + 1):
        v_val = ws.cell(row=row_num, column=v_col).value
        if v_val is None or v_val == "":
            break   # stop at the first empty voltage cell

        # --- CT类型 forward-fill ---
        # Merged cells in openpyxl: only the top cell holds the value; the rest
        # read as None.  We keep current_ct and update it only when a non-empty
        # cell is encountered, effectively forward-filling down each merged block.
        if ct_col_1 is not None:
            raw_ct = _normalize(ws.cell(row=row_num, column=ct_col_1).value)
            if raw_ct:  # non-empty → new block starts
                current_ct = raw_ct
            # if empty (None/blank), current_ct stays as the previous value

        # --- CT类型 filter ---
        if _ct_filter is not None and ct_col_1 is not None:
            if current_ct != _ct_filter:
                continue

        # --- Numeric current: scan from current-header column up to (not including)
        #     power_factor column.  This handles both:
        #       (a) simple layout: numeric directly in current column
        #       (b) real-template: text label in current column, numeric in next column
        current_val: Optional[float] = None
        for scan_col in range(a_col_1, pf_col):   # 1-based, exclusive of pf_col
            raw = ws.cell(row=row_num, column=scan_col).value
            if isinstance(raw, (int, float)):
                current_val = float(raw)
                break

        if current_val is None:
            continue   # skip rows with no numeric current

        points.append({
            "row": row_num,
            "voltage": float(v_val),
            "current": current_val,
            "power_factor": float(ws.cell(row=row_num, column=pf_col).value),
            "pulse_constant": (
                ws.cell(row=row_num, column=pc_col_1).value
                if pc_col_1 is not None else None
            ),
        })

    return points


# ---------------------------------------------------------------------------
# Public: write_results
# ---------------------------------------------------------------------------

def write_results(
    path: str,
    results: list[dict],
    sheet: Optional[str] = None,
) -> None:
    """Write frequency-counter results back into the xlsx workbook.

    Only the min_s, max_s, and avg_s cells are touched; all other cells
    (including formula cells) are preserved.

    Parameters
    ----------
    path : str
        Path to the xlsx file (overwritten in place).
    results : list[dict]
        Each dict: {"row": int, "min_s": float, "max_s": float, "avg_s": float}.
    sheet : str, optional
        Worksheet name.  If None, the active sheet is used.

    Raises
    ------
    ValueError
        If the min_s, max_s, or avg_s columns cannot be located.
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet] if sheet else wb.active

    _header_row_num, col_map = _find_header_row(ws)

    if not {"min_s", "max_s", "avg_s"}.issubset(col_map):
        missing = {"min_s", "max_s", "avg_s"} - col_map.keys()
        raise ValueError(
            f"Cannot write results: result column(s) not found: {missing}. "
            "Expected headers 'mim(s)', 'max(s)', 'avg(s)'."
        )

    min_col = col_map["min_s"] + 1
    max_col = col_map["max_s"] + 1
    avg_col = col_map["avg_s"] + 1

    for entry in results:
        row_num = entry["row"]
        ws.cell(row=row_num, column=min_col).value = entry["min_s"]
        ws.cell(row=row_num, column=max_col).value = entry["max_s"]
        ws.cell(row=row_num, column=avg_col).value = entry["avg_s"]

    wb.save(path)
