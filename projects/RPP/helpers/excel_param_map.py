"""
设备参数模板加载器 —— 从 data/map 目录的 Excel blockParams 表读取参数映射。

职责：
  paramType → {"addr": int, "type": str, "scale": float, "unit": str}

用途：
  作为 MIB OBJECT-TYPE 名称 → Modbus 地址/类型/系数 的权威来源，
  取代各模块中分散的手写地址表。

路径约定：
  Excel 文件默认位于 <project_root>/data/map/。
  若环境变量 ACUHMI_DATA_MAP_DIR 存在，则优先使用该路径，
  便于在路径变更或 CI 环境中覆盖默认位置。

数据类型映射（Excel dataType → 内部类型字符串）：
  float32  → "float"       (03H, 2 regs, IEEE754 大端序)
  double   → "double"      (03H, 4 regs, IEEE754 大端序)
  uint32   → "uint32"      (03H, 2 regs, 无符号大端序)
  int32    → "int32"       (03H, 2 regs, 有符号大端序)
  uint16   → "word"        (03H, 1 reg,  无符号)
  word     → "word"        (03H, 1 reg,  无符号)
  int16    → "word_signed" (03H, 1 reg,  有符号，需补码转换)
  其他类型（uint64/string/isOnline/enum）→ 跳过，不参与 Modbus 读取对比
"""

import os
import logging
from pathlib import Path

log = logging.getLogger(__name__)

# ── 路径配置 ─────────────────────────────────────────────────────────────────
# 默认：本文件在 helpers/，向上 3 层是仓库根，再进入 knowledge/shared/templates/raw/
_DEFAULT_DATA_MAP_DIR = Path(__file__).resolve().parents[3] / "knowledge" / "shared" / "templates" / "raw"

def _get_data_map_dir() -> Path:
    """优先读取环境变量 ACUHMI_DATA_MAP_DIR，否则使用相对于本文件的默认路径。"""
    env = os.environ.get("ACUHMI_DATA_MAP_DIR")
    if env:
        return Path(env)
    return _DEFAULT_DATA_MAP_DIR


# ── 设备 → Excel 文件名 ───────────────────────────────────────────────────────
_EXCEL_FILES: dict[str, str] = {
    "AcuRev4100": "AcuRev-4100_v1.01_20260427.xlsx",
    "AcuRev2100": "AcuRev-2100_v1.01_20260416.xlsx",
    "AcuvimIIW":  "AcuvimIIW_v1.01_20260509.xlsx",
    "Acuvim3":    "Acuvim3_v1.01_20260416.xlsx",
    "AcuRev1300": "AcuRev-1300_v1.01_20260416.xlsx",
    "AcuvimIIR":  "AcuvimIIR_v1.01_20260509.xlsx",
}

# ── Excel dataType → 内部类型字符串 ───────────────────────────────────────────
_DTYPE_MAP: dict[str, str | None] = {
    "float32":  "float",
    "double":   "double",
    "uint32":   "uint32",
    "int32":    "int32",
    "uint16":   "word",
    "word":     "word",
    "int16":    "word_signed",
    # 以下不参与保持寄存器读取对比，跳过
    "uint64":   None,
    "string":   None,
    "isOnline": None,
    "enum":     None,
}

# ── 加载缓存 ──────────────────────────────────────────────────────────────────
_cache: dict = {}


def load_device_params(device_name: str, snmp_only: bool = True) -> dict:
    """
    加载设备参数模板，返回 {paramType: {"addr", "type", "scale", "unit"}}。

    参数：
        device_name: 设备名称，对应 _EXCEL_FILES 中的键
        snmp_only:   True 时只返回 SNMP 列非空的行（即通过 SNMP 暴露的参数）

    返回值说明：
        "addr"  — Modbus 寄存器起始地址（十进制）
        "type"  — 内部类型字符串（见模块注释）
        "scale" — Modbus 原始值 × scale = 物理量
        "unit"  — 单位字符串（仅供显示）

    注意：scale=None 的行（Excel 中该格为空）会被跳过，由调用方的 legacy 逻辑处理。
    """
    key = (device_name, snmp_only)
    if key in _cache:
        return _cache[key]

    fname = _EXCEL_FILES.get(device_name)
    if fname is None:
        log.debug("[ExcelMap] 设备 %s 无对应 Excel 文件", device_name)
        return {}

    excel_path = _get_data_map_dir() / fname
    if not excel_path.exists():
        log.warning("[ExcelMap] 文件不存在: %s", excel_path)
        return {}

    try:
        import openpyxl
    except ImportError:
        log.error("[ExcelMap] 需要安装 openpyxl: pip install openpyxl")
        return {}

    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        ws = wb["blockParams"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        log.error("[ExcelMap] 读取 %s 失败: %s", excel_path, e)
        return {}

    if not rows:
        return {}

    # 找到包含 "paramType" 的 header 行（通常是第一行，但做容错）
    header_idx = next(
        (i for i, row in enumerate(rows)
         if any(str(c).strip().lower() == "paramtype" for c in row if c is not None)),
        None,
    )
    if header_idx is None:
        log.error("[ExcelMap] %s 中未找到 blockParams header 行", fname)
        return {}

    header = rows[header_idx]

    def ci(name: str) -> int | None:
        nl = name.lower()
        for i, h in enumerate(header):
            if h is not None and str(h).strip().lower() == nl:
                return i
        return None

    col_addr  = ci("start(dec)")
    col_ptype = ci("paramtype")
    col_dtype = ci("datatype")
    col_scale = ci("scale")
    col_unit  = ci("unit")
    col_snmp  = ci("snmp")
    col_desc  = ci("descrption")  # Excel typo: missing 'i'

    if any(c is None for c in [col_addr, col_ptype, col_dtype, col_scale]):
        log.error("[ExcelMap] %s 列缺失: addr=%s ptype=%s dtype=%s scale=%s",
                  fname, col_addr, col_ptype, col_dtype, col_scale)
        return {}

    result: dict = {}
    skipped_no_scale = 0

    for row in rows[header_idx + 1:]:
        if row is None or all(c is None for c in row):
            continue

        ptype = row[col_ptype]
        if ptype is None:
            continue
        ptype = str(ptype).strip()
        if not ptype:
            continue

        # SNMP 过滤
        if snmp_only and col_snmp is not None:
            snmp_val = row[col_snmp]
            if not snmp_val or str(snmp_val).strip() in ("", "None"):
                continue

        addr_raw = row[col_addr]
        if addr_raw is None:
            continue
        try:
            addr = int(addr_raw)
        except (ValueError, TypeError):
            continue

        dtype_raw = row[col_dtype]
        dtype_str = str(dtype_raw).strip() if dtype_raw is not None else "float32"
        internal_type = _DTYPE_MAP.get(dtype_str)
        if internal_type is None:
            continue  # 不支持的类型（isOnline/uint64/string 等）

        scale_raw = row[col_scale]
        if scale_raw is None:
            # scale 缺失的行由 legacy 逻辑处理，跳过
            skipped_no_scale += 1
            continue
        try:
            scale = float(scale_raw)
        except (ValueError, TypeError):
            skipped_no_scale += 1
            continue

        unit = ""
        if col_unit is not None and row[col_unit] is not None:
            unit = str(row[col_unit]).strip()

        desc = ""
        if col_desc is not None and row[col_desc] is not None:
            desc = str(row[col_desc]).strip()

        result[ptype] = {
            "addr":  addr,
            "type":  internal_type,
            "scale": scale,
            "unit":  unit,
            "desc":  desc,
        }

    log.info("[ExcelMap] %s(%s): 加载 %d 个参数，跳过 scale 缺失 %d 个",
             device_name, "SNMP" if snmp_only else "全量", len(result), skipped_no_scale)
    _cache[key] = result
    return result


def get_param(device_name: str, param_type: str, snmp_only: bool = True) -> dict | None:
    """
    获取单个参数的映射信息。
    返回 {"addr", "type", "scale", "unit"} 或 None（未找到）。
    """
    return load_device_params(device_name, snmp_only).get(param_type)


def list_snmp_params(device_name: str) -> list[str]:
    """返回设备所有 SNMP 暴露参数的 paramType 列表（按加载顺序）。"""
    return list(load_device_params(device_name, snmp_only=True).keys())


def load_bacnet_params(device_name: str) -> dict:
    """
    加载设备参数模板中 BACnet/IP 支持的参数。

    过滤规则：BACnetIP 列非空（任意非空字符串视为支持）。
    返回值结构同 load_device_params：
        {paramType: {"addr", "type", "scale", "unit", "desc"}}
    """
    cache_key = (device_name, "bacnet")
    if cache_key in _cache:
        return _cache[cache_key]

    fname = _EXCEL_FILES.get(device_name)
    if fname is None:
        log.debug("[ExcelMap] 设备 %s 无对应 Excel 文件", device_name)
        return {}

    excel_path = _get_data_map_dir() / fname
    if not excel_path.exists():
        log.warning("[ExcelMap] 文件不存在: %s", excel_path)
        return {}

    try:
        import openpyxl
    except ImportError:
        log.error("[ExcelMap] 需要安装 openpyxl: pip install openpyxl")
        return {}

    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        ws = wb["blockParams"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        log.error("[ExcelMap] 读取 %s 失败: %s", excel_path, e)
        return {}

    if not rows:
        return {}

    header_idx = next(
        (i for i, row in enumerate(rows)
         if any(str(c).strip().lower() == "paramtype" for c in row if c is not None)),
        None,
    )
    if header_idx is None:
        log.error("[ExcelMap] %s 中未找到 blockParams header 行", fname)
        return {}

    header = rows[header_idx]

    def _ci(name: str) -> int | None:
        nl = name.lower()
        for idx, h in enumerate(header):
            if h is not None and str(h).strip().lower() == nl:
                return idx
        return None

    col_addr   = _ci("start(dec)")
    col_ptype  = _ci("paramtype")
    col_dtype  = _ci("datatype")
    col_scale  = _ci("scale")
    col_unit   = _ci("unit")
    col_bacnet = _ci("bacnetip")
    col_desc   = _ci("descrption")  # Excel typo: missing 'i'

    if any(c is None for c in [col_addr, col_ptype, col_dtype, col_scale]):
        log.error("[ExcelMap] %s 必要列缺失: addr=%s ptype=%s dtype=%s scale=%s",
                  fname, col_addr, col_ptype, col_dtype, col_scale)
        return {}

    if col_bacnet is None:
        log.warning("[ExcelMap] %s 无 BACnetIP 列，返回空", fname)
        return {}

    result: dict = {}
    skipped_no_scale = 0

    for row in rows[header_idx + 1:]:
        if row is None or all(c is None for c in row):
            continue

        ptype = row[col_ptype]
        if ptype is None:
            continue
        ptype = str(ptype).strip()
        if not ptype:
            continue

        # BACnetIP 列过滤
        bacnet_val = row[col_bacnet]
        if not bacnet_val or str(bacnet_val).strip() in ("", "None"):
            continue

        addr_raw = row[col_addr]
        if addr_raw is None:
            continue
        try:
            addr = int(addr_raw)
        except (ValueError, TypeError):
            continue

        dtype_raw = row[col_dtype]
        dtype_str = str(dtype_raw).strip() if dtype_raw is not None else "float32"
        internal_type = _DTYPE_MAP.get(dtype_str)
        if internal_type is None:
            continue

        scale_raw = row[col_scale]
        if scale_raw is None:
            skipped_no_scale += 1
            continue
        try:
            scale = float(scale_raw)
        except (ValueError, TypeError):
            skipped_no_scale += 1
            continue

        unit = ""
        if col_unit is not None and row[col_unit] is not None:
            unit = str(row[col_unit]).strip()

        desc = ""
        if col_desc is not None and row[col_desc] is not None:
            desc = str(row[col_desc]).strip()

        result[ptype] = {
            "addr":  addr,
            "type":  internal_type,
            "scale": scale,
            "unit":  unit,
            "desc":  desc,
        }

    log.info("[ExcelMap] %s(BACnetIP): 加载 %d 个参数，跳过 scale 缺失 %d 个",
             device_name, len(result), skipped_no_scale)
    _cache[cache_key] = result
    return result


def list_bacnet_params(device_name: str) -> list[str]:
    """返回设备所有 BACnet/IP 暴露参数的 paramType 列表（按加载顺序）。"""
    return list(load_bacnet_params(device_name).keys())
