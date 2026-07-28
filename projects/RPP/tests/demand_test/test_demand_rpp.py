import cmath
import itertools
import logging
import math
import sys
from datetime import datetime

import openpyxl
import pytest
from openpyxl import Workbook

from comm.source_control import *

from tools.excel_operate import data_read
from projects.RPP.tests.demand_test.acuvimseries_modbus_get import HandleMemory
from projects.RPP.tests.demand_test.memory_addrs import MemoryAddr
from projects.RPP.tests.demand_test.demand_table_heading import TableTitle
from projects.RPP.tests.demand_test.demand_addr_reader import load_demand_addr

Log(str(__file__).split("\\")[-1])
CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
# data_read() 内部以 root_path(仓库根) + file_path 拼路径, 须传仓库根相对路径(以 / 开头),
# 与 comm/test_data 等其他调用方一致; 不能传绝对路径(否则 root_path + 绝对路径 非法)。
test_case_path = "/projects/RPP/tests/demand_test/demand_test_case.xlsx"

# 注: demand_test_case.xlsx 中 test_case_mA / test_case_mV 均为需量(14 列)格式可用。
# 当前数据分布(截至本次核对):
#   - test_case_mA: 17 行, 全为 3E4WY(wire_type=4); fixed 10 / sliding 7。
#   - test_case_mV: 17 行, 2E3WN(wire_type=3) 15 行(fixed 8 / sliding 7) + 3E4WY(wire_type=4)
#                   2 行(fixed 2)。
#   - test_case_rct: sheet 不存在 -> 选 rct 模式时该组合会因无数据被 pytest.skip。
# 其余接线方式需按同样 14 列格式补齐对应数据行。文件内的 *-1e2w1p / test_data 是 26 列
# 精度测试旧格式, 与本模块不兼容。
sheet_name_mV = "test_case_mV"
sheet_name_mA = "test_case_mA"
sheet_name_rct = "test_case_rct"

# 调试/冒烟开关(环境变量, 默认不设 -> 不影响正式运行):
#   DEMAND_CASE_ID=xxx : 只跑 test_case 列等于该值的那一行(如 AcuRev1320_case1)
#   DEMAND_MAX_CASES=N : 每个组合只跑前 N 条用例(如 =1 只跑第一行)
#   DEMAND_FAST=1      : 把需量等待 sleep 压到极短, 几秒跑完整条链路(控源/Modbus 读/
#                        写表/断言), 仅验证"代码能跑"; 需量不会真正累积 -> 精度断言通常
#                        会失败(预期)。可叠加: DEMAND_FAST=1 DEMAND_CASE_ID=AcuRev1320_case1。
_CASE_ID = os.environ.get("DEMAND_CASE_ID") or None
_MAX_CASES = int(os.environ["DEMAND_MAX_CASES"]) if os.environ.get("DEMAND_MAX_CASES") else None
_FAST = os.environ.get("DEMAND_FAST") == "1"

# 长等待进度的实时输出: 直接写真实终端(Windows CONOUT$ / Unix /dev/tty), 绕过 pytest 对
# stdout 的捕获(否则 print 会被憋到用例结束才一次性输出); 无可用终端(如 CI/重定向)时回退 print。
_LIVE_TTY = None
_LIVE_TTY_TRIED = False


def _live_print(msg):
    """把一行进度实时写到真实终端(绕过 pytest 捕获); 无终端时回退普通 print。"""
    global _LIVE_TTY, _LIVE_TTY_TRIED
    if not _LIVE_TTY_TRIED:
        _LIVE_TTY_TRIED = True
        try:
            _LIVE_TTY = open(
                "CONOUT$" if os.name == "nt" else "/dev/tty", "w",
                encoding=(getattr(sys.stdout, "encoding", None) or "utf-8"), errors="replace"
            )
        except OSError:
            _LIVE_TTY = None
    if _LIVE_TTY is not None:
        try:
            _LIVE_TTY.write(msg + "\n")
            _LIVE_TTY.flush()
            return
        except OSError:
            pass
    print(msg, flush=True)

ALL_SAVE_DIRS_BY_WIRE_MODE = {
    0: "save_filedir_1e2w1p",
    1: "save_filedir_2e3w1p",
    2: "save_filedir_2e3wd",
    3: "save_filedir_2e3wn",
    4: "save_filedir_3e4wy",
    5: "save_filedir_3e4wd",
}

ALL_SAVE_DIRS = ["1E2W1P", "2E3W1P", "2E3WD", "2E3WN", "3E4WY", "3E4WD"]


# save_filedir = os.path.join(
#         CURRENT_PATH, f"{time.strftime('%Y%m%d')}", "Demand_3E4WY",
#         f"demand_test_res_{time.strftime('%Y%m%d%H%M%S')}"
#     )
#     if not os.path.exists(save_filedir):
#         os.makedirs(save_filedir, exist_ok=True)

def init_filepath():
    for i in range(len(ALL_SAVE_DIRS)):
        ALL_SAVE_DIRS_BY_WIRE_MODE[i] = os.path.join(
            CURRENT_PATH,
            f"{time.strftime('%Y%m%d')}",
            f"Demand_{ALL_SAVE_DIRS[i]}",
            f"demand_test_res_{time.strftime('%Y%m%d%H%M%S')}"
        )
        if not os.path.exists(ALL_SAVE_DIRS_BY_WIRE_MODE[i]):
            os.makedirs(ALL_SAVE_DIRS_BY_WIRE_MODE[i], exist_ok=True)
    return ALL_SAVE_DIRS_BY_WIRE_MODE


class DemandTest:
    def __init__(self):
        # 方案B: 需量寄存器地址从知识库官方地址表读取并覆盖 memory_addrs 中的旧硬编码值
        # (原 0xC4xx 为 IIV3 残留, 在 AcuRev1320 上读不到 -> 返回 None -> 崩)。
        MemoryAddr.demand_addr = load_demand_addr()
        self.handle_memory = HandleMemory(slave_id=1)
        # 逐条需量精度判定结果(True/False)累积, 供 pytest 层做真实断言。
        self.case_results = []
        # 结构化结果记录, 供生成"概览 + 明细"双 sheet 报告(见 write_demand_report)。
        # 每条: {meta, cycle, kind, std[7], meas[7], passed}; kind: "measure"/"clear"。
        self.demand_records = []
        self._case_meta = None   # 当前用例的公共信息(下发参数), 由 _begin_case 设置
        self._cycle_no = 0       # 当前用例内的周期计数(1st/2nd/3rd), 每用例重置
        init_filepath()

    # 结果记录列顺序(与 check_demand_power_current_is_pass 的 std/meas 一致)
    _DEMAND_QTY_LABELS = ["P(kW)", "Q(kvar)", "S(kVA)", "Ia(A)", "Ib(A)", "Ic(A)", "In(A)"]

    def _begin_case(self, meta):
        """在每条用例开始时登记其公共信息并重置周期计数。

        meta 顺序: (case_id, accuracy, voltage, angle, current, freq, wire_type,
                    demand_method, interval, update_rate, trigger)
        """
        self._case_meta = meta
        self._cycle_no = 0

    def _record_clear_case(self, cleared):
        """登记清零分支用例(电压<9.5 或 电流=0)的结果, 供报告显示(否则该类用例不入表)。"""
        if self._case_meta is not None:
            self.demand_records.append({
                "meta": self._case_meta, "cycle": None, "kind": "clear",
                "std": None, "meas": None, "passed": bool(cleared),
            })
        return cleared

    def select_test_case(self, test_type, wire_type, demand_type):
        test_type_to_sheet = {
            0: sheet_name_mA,
            1: sheet_name_mA,
            2: sheet_name_mV,
            3: sheet_name_rct,
        }
        for key in test_type_to_sheet.keys():
            if test_type == key:
                self.select_wire_type(test_case_path, test_type_to_sheet[key], wire_type, demand_type)
                return

    def select_wire_type(self, case_path, sheet_name, wire_type, demand_type):
        """
        选择接线方式
        :param case_path: 测试文件名
        :param sheet_name: sheet名
        :param wire_type: 接线方式
        :param demand_type: 触发方式
        """
        if wire_type == 0:
            self.set_wire_type(voltage_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_1e2w1p(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_1e2w1p(file_path, input_list, demand_type)
        elif wire_type == 1:
            self.set_wire_type(voltage_wire_value=1)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_2e3w1p(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_2e3w1p(file_path, input_list, demand_type)
        elif wire_type == 2:
            self.set_wire_type(voltage_wire_value=2)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_2e3wd(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_2e3wd(file_path, input_list, demand_type)
        elif wire_type == 3:
            self.set_wire_type(voltage_wire_value=3)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_2e3wn(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_2e3wn(file_path, input_list, demand_type)
        elif wire_type == 4:
            self.set_wire_type(voltage_wire_value=4)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_3e4wy(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_3e4wy(file_path, input_list, demand_type)
        elif wire_type == 5:
            self.set_wire_type(voltage_wire_value=5)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_3e4wd(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_3e4wd(file_path, input_list, demand_type)

    def set_wire_type(self, voltage_wire_value):
        """
        设置接线方式
        :param voltage_wire_value: 电压接线方式
        :return:
        """
        self.handle_memory.set_wire_mode_by_voltage(voltage_wire_mode=voltage_wire_value)

    @staticmethod
    def get_input_list_by_wire_type(source_input_list, wire_type):
        """
        获取接线方式的数据
        :param source_input_list: 测试用例数据
        :param wire_type: 接线类型
        :return: 接线方式的数据
        “”“ 函数基于IIV3代码， 修改了接线方式对应关系”“”
        """
        input_list = [source_input_list[0]]
        # Excel 第 4 列(col4)存放数字型接线方式(0~5), 与 wire_type 取值同义,
        # 直接按数值比对筛选(0:1E2W1P 1:2E3W1P 2:2E3WD 3:2E3WN 4:3E4WY 5:3E4WD)。
        for i in range(1, len(source_input_list)):
            if source_input_list[i][4] == wire_type:
                input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_demand_para_by_input_list_of_wire_type(source_input_list, demand_type):
        """
        获取接线方式的快速测量实时数据
        :param source_input_list: 接线方式的数据
        :param demand_type
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            if source_input_list[i][6] == demand_type:
                input_list.append(source_input_list[i])
        # 调试: 按 test_case 列只保留指定的那一行(表头 + 命中行); 默认 None 不过滤。
        if _CASE_ID is not None:
            input_list = [input_list[0]] + [r for r in input_list[1:] if str(r[0]) == _CASE_ID]
        # 调试: 只保留前 _MAX_CASES 条用例(表头 + N 行); 默认 None 全跑。
        if _MAX_CASES is not None:
            input_list = input_list[:_MAX_CASES + 1]
        return input_list

    @staticmethod
    def get_save_filepath_of_3e4wd(filedir, demand_type):
        """
        接线方式:3e4wd
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_3E4WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_3E4WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wn(filedir, demand_type):
        """
        接线方式:2e3wn
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_2E3WN_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_2E3WN_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wd(filedir, demand_type):
        """
        接线方式:2e3wd
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_2E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_2E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3w1p(filedir, demand_type):
        """
        接线方式:2e3w1p
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_2E3W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_2E3W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_1e2w1p(filedir, demand_type):
        """
        接线方式:1e2w1p
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_1E2W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_1E2W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_3e4wy(filedir, demand_type):
        """
        接线方式:3e4wy
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_3E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_3E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    def demand_test_by_3e4wd(self, file_path, input_list, demand_type):
        """
        接线方式:3e4wd
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_3e4wd(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_3e4wd(file_path, input_list)

    def demand_test_by_2e3wn(self, file_path, input_list, demand_type):
        """
        接线方式:2e3wn
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_2e3wn(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_2e3wn(file_path, input_list)

    def demand_test_by_2e3wd(self, file_path, input_list, demand_type):
        """
        接线方式:2e3wd
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_2e3wd(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_2e3wd(file_path, input_list)

    def demand_test_by_2e3w1p(self, file_path, input_list, demand_type):
        """
        接线方式:2e3w1p
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_2e3w1p(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_2e3w1p(file_path, input_list)

    def demand_test_by_1e2w1p(self, file_path, input_list, demand_type):
        """
        接线方式:1e2w1p
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_1e2w1p(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_1e2w1p(file_path, input_list)

    def demand_test_by_3e4wy(self, file_path, input_list, demand_type):
        """
        接线方式:3e4wy
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_3e4wy(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_3e4wy(file_path, input_list)

    @staticmethod
    def write_table_title_of_3e4wy_to_excel(file_path, wb, ws):
        """
        写入表头:3e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_3E4WY)):
            j = i + 1
            ws.cell(1, j, TableTitle.DEMAND_COLUMNS_OF_3E4WY[i])
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_1e2w1p_to_excel(file_path, wb, ws):
        """
        写入表头:1e2w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_1E2W1P)):
            j = i + 1
            ws.cell(1, j, TableTitle.DEMAND_COLUMNS_OF_1E2W1P[i])
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3w1p_to_excel(file_path, wb, ws):
        """
        写入表头:2e3w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_2E3W1P)):
            j = i + 1
            ws.cell(1, j, TableTitle.DEMAND_COLUMNS_OF_2E3W1P[i])
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wn_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wn
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_2E3WN)):
            j = i + 1
            ws.cell(1, j, TableTitle.DEMAND_COLUMNS_OF_2E3WN[i])
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wd_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_2E3WD)):
            j = i + 1
            ws.cell(1, j, TableTitle.DEMAND_COLUMNS_OF_2E3WD[i])
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_3e4wd_to_excel(file_path, wb, ws):
        """
        写入表头:3e4wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_3E4WD)):
            j = i + 1
            ws.cell(1, j, TableTitle.DEMAND_COLUMNS_OF_3E4WD[i])
        wb.save(file_path)

    @staticmethod
    def get_test_case_info_of_accuracy(input_list, index_value):
        """
        获取测试精度信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 精度值
        """
        accuracy_of_demand = input_list[index_value + 1][10]
        return accuracy_of_demand

    @staticmethod
    def get_test_case_info_of_input_value(input_list, index_value):
        """
        获取测试电压/电流、相位等信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 电压/电流、相位等信息
        """
        # Excel 列契约(0 基): 0 test_case 1 voltage 2 angle 3 current 4 wire_type(数字)
        # 5 freq 6 demand_method 7 interval 8 update_rate 9 demand_trigger
        # 10 voltage_accuracy 11 freq(冗余) 12 抽样次数 13 抽样间隔
        case_id = input_list[index_value + 1][0]
        u = input_list[index_value + 1][1]
        ui_angle = input_list[index_value + 1][2]
        i = input_list[index_value + 1][3]
        wire_type = input_list[index_value + 1][4]
        freq = input_list[index_value + 1][5]
        method = input_list[index_value + 1][6]
        interval = input_list[index_value + 1][7]
        update_rate = input_list[index_value + 1][8]
        trigger = input_list[index_value + 1][9]
        sample_cnt = input_list[index_value + 1][12]
        sample_interval = input_list[index_value + 1][13]
        return case_id, u, ui_angle, i, freq, method, interval, update_rate, trigger, sample_cnt, sample_interval, wire_type

    @staticmethod
    def calculate_active_power(voltage, current, voltage_current_angle):
        """
        计算有功功率(单位 kW)。

        除以 1000 与表需量功率寄存器的原生单位(kW)对齐: 读取端
        read_demand_sys_active_power 直接返回寄存器原生 float(kW), 期望端按 V×I 算得 W,
        故此处 /1000 统一为 kW, 否则功率类需量判定恒差 1000 倍。

        :param voltage:电压
        :param current:电流
        :param voltage_current_angle:电压电流相位角度
        :return: 有功功率, 单位 kW
        """
        active_power = (voltage * current * math.cos(math.radians(voltage_current_angle))) / 1000
        return active_power

    @staticmethod
    def calculate_reactive_power_by_sin(voltage, current, voltage_current_angle):
        """
        计算有功功率
        :param voltage:电压
        :param current:电流
        :param voltage_current_angle:电压电流相位角度
        :return:
        """
        reactive_power = (voltage * current * math.sin(math.radians(voltage_current_angle)))
        return reactive_power

    @staticmethod
    def calculate_reactive_power(active_power, apparent_power):
        """
        计算有功功率
        :param active_power:有功功率
        :param apparent_power:视在功率
        :return:
        """
        reactive_power_pow2 = apparent_power ** 2 - active_power ** 2
        reactive_power = math.sqrt(reactive_power_pow2) if reactive_power_pow2 else 0
        return reactive_power

    @staticmethod
    def calculate_apparent_power(voltage, current):
        """
        计算视在功率(单位 kVA)。

        /1000 与表需量视在功率寄存器原生单位(kVA)对齐, 理由同 calculate_active_power。

        :param voltage:电压
        :param current:电流
        :return: 视在功率, 单位 kVA
        """
        apparent_power = (voltage * current) / 1000
        return apparent_power

    def set_demand_para(self, demand_method, demand_interval, demand_update_rate):
        """
        Args:
            demand_method: Fixed Window: 0  Sliding Window: 1
            demand_interval: 1~30 minute
            demand_update_rate: 1~30 minute
        Returns:
        """
        # 设置demand method, sliding or fixed
        self.handle_memory.set_demand_method(demand_method)
        self.handle_memory.set_demand_interval(demand_interval)
        self.handle_memory.set_demand_update_rate(demand_update_rate)

    def set_time_trigger(self, sys_millisecond):
        """
        sys_millisecond,0-999
        """
        self.handle_memory.set_sys_millisecond(sys_millisecond)

    def set_demand_trigger(self, clear_max_demand):
        """
        clear_max_demand
        0, None;
        1: clear
        """
        self.handle_memory.set_clear_max_demand(clear_max_demand)

    @staticmethod
    def check_demand_is_clear(measure_value, tolerance=0.01):
        """
        用于判断demand清除是否成功
        """
        if abs(measure_value) <= tolerance:
            return True, measure_value,
        return False, measure_value

    @staticmethod
    def check_demand_is_pass(standard_value, measure_value, tolerance=0.01):
        """
        用于判断demand是否符合预期。
        期望值为 0 时(如纯阻性负载的无功需量)无法算相对误差, 改用绝对误差与容差比较,
        避免除零(与 check_demand_is_clear 一致)。
        """
        if standard_value == 0:
            return abs(measure_value) <= tolerance, measure_value
        relative_error = round((measure_value - standard_value) / standard_value, 5)
        return abs(relative_error) <= tolerance, measure_value

    def check_demand_power_is_pass(self, standard_power_values, tolerance):
        std_sys_active_power, std_sys_reactive_power, std_sys_apparent_power = standard_power_values
        sys_active_power = self.handle_memory.read_demand_sys_active_power()
        p_res, p_val = self.check_demand_is_pass(std_sys_active_power, sys_active_power, tolerance)
        sys_reactive_power = self.handle_memory.read_demand_sys_reactive_power()
        q_res, q_val = self.check_demand_is_pass(std_sys_reactive_power, sys_reactive_power, tolerance)
        sys_apparent_power = self.handle_memory.read_demand_sys_apparent_power()
        s_res, s_val = self.check_demand_is_pass(std_sys_apparent_power, sys_apparent_power, tolerance)
        return p_res, q_res, s_res, p_val, q_val, s_val

    def check_demand_current_is_pass(self, standard_current_values, tolerance):
        std_demand_ia, std_demand_ib, std_demand_ic, std_demand_in = standard_current_values
        demand_ia = self.handle_memory.read_demand_ia()
        ia_res, ia_val = self.check_demand_is_pass(std_demand_ia, demand_ia, tolerance)
        demand_ib = self.handle_memory.read_demand_ib()
        ib_res, ib_val = self.check_demand_is_pass(std_demand_ib, demand_ib, tolerance)
        demand_ic = self.handle_memory.read_demand_ic()
        ic_res, ic_val = self.check_demand_is_pass(std_demand_ic, demand_ic, tolerance)
        demand_in = self.handle_memory.read_demand_in()
        if demand_in is None:  # RPP 无中线电流需量寄存器, 跳过 In 比对(记 N/A)
            in_res, in_val = True, None
        else:
            in_res, in_val = self.check_demand_is_pass(std_demand_in, demand_in, tolerance)
        return ia_res, ib_res, ic_res, in_res, ia_val, ib_val, ic_val, in_val

    def check_demand_power_current_is_pass(self, standard_power_values, standard_current_values, tolerance):
        p_res, q_res, s_res, p_val, q_val, s_val = self.check_demand_power_is_pass(standard_power_values, tolerance)
        ia_res, ib_res, ic_res, in_res, ia_val, ib_val, ic_val, in_val = self.check_demand_current_is_pass(
            standard_current_values, tolerance)
        power_current_res = [p_res, q_res, s_res, ia_res, ib_res, ic_res, in_res]
        power_current_vals = [p_val, q_val, s_val, ia_val, ib_val, ic_val, in_val]
        is_pass = bool(len(power_current_res) and all(power_current_res))
        # 记录该次需量精度判定结果, 供 pytest 层聚合断言(option 2)。
        self.case_results.append(is_pass)
        # 结构化记录本周期(供双 sheet 报告); std/meas 顺序均为 [P,Q,S,Ia,Ib,Ic,In]。
        self._cycle_no += 1
        if self._case_meta is not None:
            self.demand_records.append({
                "meta": self._case_meta, "cycle": self._cycle_no, "kind": "measure",
                "std": list(standard_power_values) + list(standard_current_values),
                "meas": list(power_current_vals), "passed": is_pass,
            })
        logging.info("demand_power_current_is_pass %s", "Succeed" if is_pass else "Failed")
        return is_pass, power_current_vals

    def check_demand_power_is_clear(self, tolerance):
        sys_active_power = self.handle_memory.read_demand_sys_active_power()
        p_res, _ = self.check_demand_is_clear(sys_active_power, tolerance)
        sys_reactive_power = self.handle_memory.read_demand_sys_reactive_power()
        q_res, _ = self.check_demand_is_clear(sys_reactive_power, tolerance)
        sys_apparent_power = self.handle_memory.read_demand_sys_apparent_power()
        s_res, _ = self.check_demand_is_clear(sys_apparent_power, tolerance)
        return p_res, q_res, s_res

    def check_demand_current_is_clear(self, tolerance):
        demand_ia = self.handle_memory.read_demand_ia()
        ia_res, _ = self.check_demand_is_clear(demand_ia, tolerance)
        demand_ib = self.handle_memory.read_demand_ib()
        ib_res, _ = self.check_demand_is_clear(demand_ib, tolerance)
        demand_ic = self.handle_memory.read_demand_ic()
        ic_res, _ = self.check_demand_is_clear(demand_ic, tolerance)
        demand_in = self.handle_memory.read_demand_in()
        if demand_in is None:  # RPP 无中线电流需量寄存器, 跳过 In 清零校验
            in_res = True
        else:
            in_res, _ = self.check_demand_is_clear(demand_in, tolerance)
        return ia_res, ib_res, ic_res, in_res

    def check_demand_power_current_is_clear(self, tolerance):
        p_res, q_res, s_res = self.check_demand_power_is_clear(tolerance)
        ia_res, ib_res, ic_res, in_res = self.check_demand_current_is_clear(tolerance)
        power_current_res = [p_res, q_res, s_res, ia_res, ib_res, ic_res, in_res]
        if len(power_current_res) and all(power_current_res):
            logging.info("demand_power_current_is_clear Succeed")
            return True
        else:
            logging.info("demand_power_current_is_clear Failed")
            return False

    def assert_source_output_alive(self, exp_voltages, exp_currents, min_ratio=0.5):
        """
        源在线探针: 加源后、进入需量长等待前回读表实时 RMS 电压/电流。

        控源走 UDP 单向下发(set_ac 只发不收), 源软件没开/没输出时下发不会报错,
        用例会白等整个需量周期才在精度断言处失败。此处主动回读, 命令值非零而实测
        明显偏低即判定源无输出, 立即抛 SourceControlError 快速失败。

        表为 1:1 直采(需量精度断言以 ~0.001 容差直接比对下发值可通过 -> 无 PT/CT 变比),
        故实测与下发同量纲, 用比例判据即可; min_ratio 取 0.5 留足源稳定/量化裕度。

        :param exp_voltages: 期望的各相电压 [ua, ub, uc], 该相值为 0 时跳过校验
        :param exp_currents: 期望的各相电流 [ia, ib, ic], 该相值为 0 时跳过校验
        :param min_ratio: 实测/期望 的最小比例, 低于则判定源无输出
        """
        volt_readers = [
            ("Ua", self.handle_memory.read_ua_voltage),
            ("Ub", self.handle_memory.read_ub_voltage),
            ("Uc", self.handle_memory.read_uc_voltage),
        ]
        curr_readers = [
            ("Ia", self.handle_memory.read_ia_current),
            ("Ib", self.handle_memory.read_ib_current),
            ("Ic", self.handle_memory.read_ic_current),
        ]
        dead = []
        for exp, (label, reader) in zip(exp_voltages, volt_readers):
            if exp:
                meas = reader()
                if meas < exp * min_ratio:
                    dead.append(f"{label} exp={exp:.4f} meas={meas:.4f}")
        for exp, (label, reader) in zip(exp_currents, curr_readers):
            if exp:
                meas = reader()
                if meas < exp * min_ratio:
                    dead.append(f"{label} exp={exp:.4f} meas={meas:.4f}")
        if dead:
            raise SourceControlError(
                "Source output not detected on meter (source off or not outputting). "
                "Check whether the control-source software is on and outputting: "
                + "; ".join(dead)
            )
        logging.info("source output alive check passed")

    @staticmethod
    def _fmt_hms(seconds):
        """把秒数格式化为 'HhMMmSSs'(去掉为 0 的高位), 用于进度提示。"""
        seconds = int(round(seconds))
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        if h:
            return f"{h}h{m:02d}m{s:02d}s"
        if m:
            return f"{m}m{s:02d}s"
        return f"{s}s"

    def sleep_with_progress(self, seconds, step_desc, prefix="", tick=60):
        """分段睡眠, 期间周期性打印"当前步骤 + 预计剩余时间", 避免长等待阶段无任何输出。

        - 剩余时间按"请求时长"递减(非墙钟), 以兼容 DEMAND_FAST(time.sleep 被压缩)。
        - DEMAND_FAST 下不分段、只打印一行, 保持冒烟模式的极速。

        :param seconds: 需等待的秒数(可为浮点)
        :param step_desc: 当前步骤描述
        :param prefix: 行首前缀(如 "case1[3/5]"), 便于区分用例/步骤
        :param tick: 正常模式下每隔多少秒打印一次剩余时间
        """
        seconds = int(round(seconds))
        head = f"[需量]{(' ' + prefix) if prefix else ''} {step_desc}"
        if seconds <= 0:
            return
        if _FAST:
            _live_print(f"{head} (FAST 跳过, 名义等待 {self._fmt_hms(seconds)})")
            time.sleep(seconds)
            return
        start_msg = f"{head} 开始, 预计等待 {self._fmt_hms(seconds)}"
        _live_print(start_msg)
        logging.info(start_msg)
        remaining = seconds
        while remaining > 0:
            nap = min(tick, remaining)
            time.sleep(nap)
            remaining -= nap
            if remaining > 0:
                msg = f"{head} 进行中, 剩余约 {self._fmt_hms(remaining)}"
                _live_print(msg)
                logging.info(msg)
        done_msg = f"{head} 完成"
        _live_print(done_msg)
        logging.info(done_msg)

    @staticmethod
    def get_wait_seconds(demand_interval):
        """
        Args:
            demand_interval: 需量窗口间隔时间min

        Returns: 当前时刻基于demand_interval等待时间，s

        """
        if not demand_interval or demand_interval > 30:
            logging.warning("demand_interval is set to zero or more than 30, skip waiting")
            return 0
        init_start_min = 60
        now_time = datetime.now()
        current_minute = now_time.minute
        current_seconds = now_time.second + 1 if now_time.microsecond else now_time.second
        demand_intervals = []
        for i in range(1, 60 // demand_interval + 1):
            if (i * demand_interval in range(current_minute, 60)) and (i * demand_interval > current_minute):
                demand_intervals.append(i * demand_interval)
                break
        if len(demand_intervals):
            predict_start_min = demand_intervals[0]
            if (init_start_min - predict_start_min) < demand_interval:
                exp_wait_min = init_start_min - current_minute
            else:
                exp_wait_min = predict_start_min - current_minute
        else:
            exp_wait_min = init_start_min - current_minute
        return exp_wait_min * 60 - current_seconds

    @staticmethod
    def write_common_values_to_excel(ws, index_value, common_values, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param common_values: case编号/电压/电流/电压相位角/电流相位角等输入值,精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(common_values)):
            ws.cell(index_value + 2, start_num + k, common_values[k])
        start_num += len(common_values)
        return start_num

    @staticmethod
    def write_accuracy_res_to_excel(file_path, wb, ws, index_value, accuracy_res, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param file_path: 待写入文件路径
        :param wb: 写入工作簿对象
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param accuracy_res: 精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(accuracy_res)):
            ws.cell(index_value + 2, start_num + k, accuracy_res[k])
        wb.save(file_path)
        start_num += len(accuracy_res)
        return start_num

    def write_demand_report(self, file_path):
        """把 self.demand_records 生成"概览(Summary) + 明细(Detail)"双 sheet 报告, 覆盖写 file_path。

        - Summary: 每条用例一行(下发参数 + 总判定 PASS/FAIL/PASS(清零));
        - Detail : 每条用例每周期一行, 每量给出 std/实测/误差%, 行尾判定;
                   清零分支用例(电压<9.5 或 电流=0)记为 1 行(kind=clear)。
        meta 顺序: (case_id, accuracy, voltage, angle, current, freq, wire_type,
                    method, interval, update_rate, trigger)
        """
        wb = openpyxl.Workbook()
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_det = wb.create_sheet("Detail")

        # ---- 概览: 按 case 聚合, 总判定 = 该 case 全部记录 passed 之与 ----
        ws_sum.append(["case", "accuracy", "voltage", "angle", "current", "freq",
                       "wire_type", "method", "interval", "update_rate", "trigger", "结果"])
        agg = {}
        order = []
        for rec in self.demand_records:
            cid = rec["meta"][0]
            if cid not in agg:
                agg[cid] = {"meta": rec["meta"], "passed": True, "clear": False}
                order.append(cid)
            agg[cid]["passed"] = agg[cid]["passed"] and bool(rec["passed"])
            if rec["kind"] == "clear":
                agg[cid]["clear"] = True
        for cid in order:
            e = agg[cid]
            if not e["passed"]:
                result = "FAIL"
            else:
                result = "PASS(清零)" if e["clear"] else "PASS"
            ws_sum.append(list(e["meta"]) + [result])

        # ---- 明细: 每记录一行 ----
        det_hdr = ["case", "周期", "voltage", "current", "angle", "interval", "trigger"]
        for q in self._DEMAND_QTY_LABELS:
            det_hdr += [f"{q}_std", f"{q}_实测", f"{q}_误差%"]
        det_hdr += ["判定"]
        ws_det.append(det_hdr)
        for rec in self.demand_records:
            m = rec["meta"]
            row = [m[0], (str(rec["cycle"]) if rec["cycle"] else "清零"),
                   m[2], m[4], m[3], m[8], m[10]]
            if rec["kind"] == "clear":
                for _ in self._DEMAND_QTY_LABELS:
                    row += ["", "", ""]
            else:
                for std, meas in zip(rec["std"], rec["meas"]):
                    err = "" if std in (0, None) else round((meas - std) / std * 100, 2)
                    row += [round(std, 5), round(meas, 5), err]
            row += ["PASS" if rec["passed"] else "FAIL"]
            ws_det.append(row)

        ws_sum.freeze_panes = "A2"
        ws_det.freeze_panes = "A2"
        wb.save(file_path)
        wb.close()

    @staticmethod
    def line_to_line_voltage_calculate(ua, ub, uc, va_angle, vb_angle, vc_angle):
        """
        计算线电压
        :param ua:ua
        :param ub:ub
        :param uc:uc
        :param va_angle:va_angle
        :param vb_angle:vb_angle
        :param vc_angle:vc_angle
        :return: vab, vbc, vca
        """
        ret = []
        va_complex = ua * cmath.exp(1j * math.radians(va_angle))
        vb_complex = ub * cmath.exp(1j * math.radians(vb_angle))
        vc_complex = uc * cmath.exp(1j * math.radians(vc_angle))
        vab = va_complex - vb_complex
        vbc = vb_complex - vc_complex
        vca = vc_complex - va_complex
        vab = abs(vab)
        vbc = abs(vbc)
        vca = abs(vca)
        ret.extend([vab, vbc, vca])
        return ret

    @staticmethod
    def get_line_to_line_voltage(u1, u2, u1_angle, u2_angle):
        """
        计算线电压
        :param u1:ua
        :param u2:ub
        :param u1_angle:u1_angle
        :param u2_angle:u2_angle
        :return: vab, vbc, vca
        """
        ret = []
        va_complex = u1 * cmath.exp(1j * math.radians(u1_angle))
        vb_complex = u2 * cmath.exp(1j * math.radians(u2_angle))
        vab = va_complex - vb_complex
        vba = vb_complex - va_complex
        vab = abs(vab)
        vba = abs(vba)
        ret.extend([vab, vba])
        return ret

    @staticmethod
    def get_sys_p_power_by_2e3wd(uab, ia, ucb, ic, vc_angle):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :param vc_angle:vc_angle
        :return: sys_p_power_by_delta, 单位 kW
        """
        # /1000 与表需量功率寄存器原生单位(kW)对齐, 理由同 calculate_active_power
        ret = (uab * ia + ucb * ic) * math.cos(math.radians(30)) * math.cos(math.radians(vc_angle)) / 1000
        return ret

    @staticmethod
    def get_sys_s_power_by_2e3wd(uab, ia, ucb, ic):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :return: sys_s_power_by_delta, 单位 kVA
        """
        # /1000 与表需量视在功率寄存器原生单位(kVA)对齐
        ret = (uab * ia + ucb * ic) * math.cos(math.radians(30)) / 1000
        return ret

    @staticmethod
    def get_sys_q_power_by_2e3wd(sys_p_power_by_2e3wd, sys_s_power_by_2e3wd):
        """
        计算线电压
        :param sys_p_power_by_2e3wd:sys_p_power_by_2e3wd
        :param sys_s_power_by_2e3wd:sys_s_power_by_2e3wd
        :return: sys_q_power_by_delta
        """
        sys_q_power_by_delta = sys_s_power_by_2e3wd ** 2 - sys_p_power_by_2e3wd ** 2
        sys_q_power_by_delta = math.sqrt(sys_q_power_by_delta) if sys_q_power_by_delta > 0 else 0
        return sys_q_power_by_delta

    @staticmethod
    def get_sys_p_power_by_3e4wd(uab, ia, ucb, ic, ubc, in_val, vc_angle):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :param ubc:ubc
        :param in_val:in_val
        :param vc_angle:vc_angle
        :return: sys_p_power_by_delta, 单位 kW
        """
        # /1000 与表需量功率寄存器原生单位(kW)对齐, 理由同 calculate_active_power
        ret = (uab * ia + ucb * ic + ubc * in_val * 0.5) * math.cos(math.radians(30)) * math.cos(
            math.radians(vc_angle)) / 1000
        return ret

    @staticmethod
    def get_sys_s_power_by_3e4wd(uab, ia, ucb, ic, ubc, in_val):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :param ubc:ubc
        :param in_val:in_val
        :return: sys_s_power_by_delta, 单位 kVA
        """
        # /1000 与表需量视在功率寄存器原生单位(kVA)对齐
        ret = (uab * ia + ucb * ic + ubc * in_val * 0.5) * math.cos(math.radians(30)) / 1000
        return ret

    @staticmethod
    def get_sys_q_power_by_3e4wd(sys_p_power_by_3e4wd, sys_s_power_by_3e4wd):
        """
        计算线电压
        :param sys_p_power_by_3e4wd:sys_p_power_by_3e4wd
        :param sys_s_power_by_3e4wd:sys_s_power_by_3e4wd
        :return: sys_q_power_by_delta
        """
        sys_q_power_by_delta = sys_s_power_by_3e4wd ** 2 - sys_p_power_by_3e4wd ** 2
        sys_q_power_by_delta = math.sqrt(sys_q_power_by_delta) if sys_q_power_by_delta > 0 else 0
        return sys_q_power_by_delta

    def fixed_demand_test_by_3e4wd(self, file_path, input_list):
        """
        IN−rms = IA−rms + IB−rms + IC−rms
        Psys=VAB∗IA−VBC∗IC+12VBC∗IN
        Psys=VAB∗IA+VCB∗IC+12VBC∗IN
        Ssys=(VAB−rms∗IA−rms+VBC−rms∗IC−rms+12VBC−rms∗IN−rms)∗Cos30deg
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = sum([ia, ib, ic])

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_angle, ub_angle, uc_angle)
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)
            exp_demand_p_sys_1st = self.get_sys_p_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_3e4wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )

            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·前半窗", prefix=f"{case_id} [4/5]")
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = sum([ia_2nd, ib_2nd, ic_2nd])
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                # 第2周期电流需量取窗口平均(前半窗满值 + 后半窗半值, 各占 0.5), 与功率 2nd 口径一致
                exp_demand_ia_2nd = ia * 0.5 + ia_2nd * 0.5
                exp_demand_ib_2nd = ib * 0.5 + ib_2nd * 0.5
                exp_demand_ic_2nd = ic * 0.5 + ic_2nd * 0.5
                exp_demand_in_2nd = in_val * 0.5 + in_2nd * 0.5

                # 等待 后1/2 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·后半窗", prefix=f"{case_id} [5/5]")
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
        self.write_demand_report(file_path)

    def sliding_demand_test_by_3e4wd(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = sum([ia, ib, ic])

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(
                ua, ub, uc,
                ua_angle, ub_angle, uc_angle
            )
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)

            exp_demand_p_sys_1st = self.get_sys_p_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_3e4wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = sum([ia_2nd, ib_2nd, ic_2nd])
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第2周期·滑动更新", prefix=f"{case_id} [4/5]")
                # 查看需量是否正确，计算电压、电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_3e4wd(
                    uab, ia, ucb, ic, ubc, in_val, vc_angle
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_3e4wd(
                    uab, ia, ucb, ic, ubc, in_val,
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = round(
                    ia * (demand_interval - demand_update_rate) / demand_interval
                    + ia_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ib_2nd = round(
                    ib * (demand_interval - demand_update_rate) / demand_interval
                    + ib_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ic_2nd = round(
                    ic * (demand_interval - demand_update_rate) / demand_interval
                    + ic_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_in_2nd = round(
                    in_val * (demand_interval - demand_update_rate) / demand_interval
                    + in_2nd * demand_update_rate / demand_interval, 5)

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第3周期·滑动更新", prefix=f"{case_id} [5/5]")
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = sum([ia_3rd, ib_3rd, ic_3rd])

                uab_3rd, ubc_3rd, uca_3rd = self.line_to_line_voltage_calculate(
                    ua_3rd, ub_3rd, uc_3rd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_3rd, = self.get_line_to_line_voltage(ub_3rd, uc_3rd, ub_angle, uc_angle)

                exp_demand_p_sys_3rd_by_down_half = self.get_sys_p_power_by_3e4wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd, ubc_3rd, in_3rd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = self.get_sys_s_power_by_3e4wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd, ubc_3rd, in_3rd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                    ) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                    ) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_3e4wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = self.get_sys_p_power_by_3e4wd(
                        uab, ia, ucb, ic, ubc, in_val, vc_angle
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = self.get_sys_s_power_by_3e4wd(
                        uab, ia, ucb, ic, ubc, in_val
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.get_sys_q_power_by_3e4wd(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )

                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_3e4wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                if demand_interval - demand_update_rate >= demand_update_rate:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * demand_update_rate / demand_interval
                        + ia * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * demand_update_rate / demand_interval
                        + ib * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * demand_update_rate / demand_interval
                        + ic * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * demand_update_rate / demand_interval
                        + in_val * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                else:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)
        self.write_demand_report(file_path)

    def fixed_demand_test_by_2e3wd(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_angle, ub_angle, uc_angle)
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)
            exp_demand_p_sys_1st = self.get_sys_p_power_by_2e3wd(
                uab, ia, ucb, ic, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_2e3wd(
                uab, ia, ucb, ic
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_2e3wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )

            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            # 2E3W Delta(2e3wd)下 1320/4100 用二表法且 N 接 B 相: 仅 Ia/Ic 两个 CT(B 相不接电流),
            # 且 B 相为公共参考端 -> 表侧 Ub/Ib 实时值恒为 0; 故 B 相电压/电流均传 0 让探针跳过。
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, 0, uc], [ia, 0, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·前半窗", prefix=f"{case_id} [4/5]")
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                # 第2周期电流需量取窗口平均(前半窗满值 + 后半窗半值, 各占 0.5), 与功率 2nd 口径一致
                exp_demand_ia_2nd = ia * 0.5 + ia_2nd * 0.5
                exp_demand_ib_2nd = ib * 0.5 + ib_2nd * 0.5
                exp_demand_ic_2nd = ic * 0.5 + ic_2nd * 0.5
                exp_demand_in_2nd = in_val * 0.5 + in_2nd * 0.5

                # 等待 后1/2 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·后半窗", prefix=f"{case_id} [5/5]")
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
        self.write_demand_report(file_path)

    def sliding_demand_test_by_2e3wd(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(
                ua, ub, uc,
                ua_angle, ub_angle, uc_angle
            )
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)

            exp_demand_p_sys_1st = self.get_sys_p_power_by_2e3wd(
                uab, ia, ucb, ic, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_2e3wd(
                uab, ia, ucb, ic
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_2e3wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            # 2E3W Delta(2e3wd)下 1320/4100 用二表法且 N 接 B 相: 仅 Ia/Ic 两个 CT(B 相不接电流),
            # 且 B 相为公共参考端 -> 表侧 Ub/Ib 实时值恒为 0; 故 B 相电压/电流均传 0 让探针跳过。
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, 0, uc], [ia, 0, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第2周期·滑动更新", prefix=f"{case_id} [4/5]")
                # 查看需量是否正确，计算电压、电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_2e3wd(
                    uab, ia, ucb, ic, vc_angle
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_2e3wd(
                    uab, ia, ucb, ic
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = round(
                    ia * (demand_interval - demand_update_rate) / demand_interval
                    + ia_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ib_2nd = round(
                    ib * (demand_interval - demand_update_rate) / demand_interval
                    + ib_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ic_2nd = round(
                    ic * (demand_interval - demand_update_rate) / demand_interval
                    + ic_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_in_2nd = round(
                    in_val * (demand_interval - demand_update_rate) / demand_interval
                    + in_2nd * demand_update_rate / demand_interval, 5)

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第3周期·滑动更新", prefix=f"{case_id} [5/5]")
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = 0

                uab_3rd, ubc_3rd, uca_3rd = self.line_to_line_voltage_calculate(
                    ua_3rd, ub_3rd, uc_3rd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_3rd, = self.get_line_to_line_voltage(ub_3rd, uc_3rd, ub_angle, uc_angle)

                exp_demand_p_sys_3rd_by_down_half = self.get_sys_p_power_by_2e3wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = self.get_sys_s_power_by_2e3wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                    ) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                    ) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_2e3wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = self.get_sys_p_power_by_2e3wd(
                        uab, ia, ucb, ic, vc_angle
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = self.get_sys_s_power_by_2e3wd(
                        uab, ia, ucb, ic
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.get_sys_q_power_by_2e3wd(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )

                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_2e3wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                if demand_interval - demand_update_rate >= demand_update_rate:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * demand_update_rate / demand_interval
                        + ia * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * demand_update_rate / demand_interval
                        + ib * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * demand_update_rate / demand_interval
                        + ic * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * demand_update_rate / demand_interval
                        + in_val * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                else:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)
        self.write_demand_report(file_path)

    def fixed_demand_test_by_2e3wn(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wn_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            # 2E3W Network(2e3wn)为两元件三线, 表只测 A/C 相(电压 2Phase + 电流 2CT), B 相不接入 ->
            # 表侧 Ub/Ib 实时值恒为 0; 此处 B 相传 0 让探针跳过, 避免把"接线本就不测 B"误判成源无输出。
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, 0, uc], [ia, 0, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·前半窗", prefix=f"{case_id} [4/5]")
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                # 第2周期电流需量取窗口平均(前半窗满值 + 后半窗半值, 各占 0.5), 与功率 2nd 口径一致
                exp_demand_ia_2nd = ia * 0.5 + ia_2nd * 0.5
                exp_demand_ib_2nd = ib * 0.5 + ib_2nd * 0.5
                exp_demand_ic_2nd = ic * 0.5 + ic_2nd * 0.5
                exp_demand_in_2nd = in_val * 0.5 + in_2nd * 0.5

                # 等待 后1/2 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·后半窗", prefix=f"{case_id} [5/5]")
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wn = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wn, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
        self.write_demand_report(file_path)

    def sliding_demand_test_by_2e3wn(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wn_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            # 2E3W Network(2e3wn)为两元件三线, 表只测 A/C 相(电压 2Phase + 电流 2CT), B 相不接入 ->
            # 表侧 Ub/Ib 实时值恒为 0; 此处 B 相传 0 让探针跳过, 避免把"接线本就不测 B"误判成源无输出。
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, 0, uc], [ia, 0, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第2周期·滑动更新", prefix=f"{case_id} [4/5]")
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = round(
                    ia * (demand_interval - demand_update_rate) / demand_interval
                    + ia_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ib_2nd = round(
                    ib * (demand_interval - demand_update_rate) / demand_interval
                    + ib_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ic_2nd = round(
                    ic * (demand_interval - demand_update_rate) / demand_interval
                    + ic_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_in_2nd = round(
                    in_val * (demand_interval - demand_update_rate) / demand_interval
                    + in_2nd * demand_update_rate / demand_interval, 5)

                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第3周期·滑动更新", prefix=f"{case_id} [5/5]")
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle),
                    self.calculate_active_power(ub_3rd, ib_3rd, vc_angle),
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd),
                    self.calculate_apparent_power(ub_3rd, ib_3rd),
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    ]) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                    ]) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                        self.calculate_active_power(ub, ib, vc_angle),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                        self.calculate_apparent_power(ub, ib),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                if demand_interval - demand_update_rate >= demand_update_rate:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * demand_update_rate / demand_interval
                        + ia * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * demand_update_rate / demand_interval
                        + ib * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * demand_update_rate / demand_interval
                        + ic * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * demand_update_rate / demand_interval
                        + in_val * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                else:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wn = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wn, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)
        self.write_demand_report(file_path)

    def fixed_demand_test_by_2e3w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))
            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 180
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            ua = voltage
            ub = 0
            uc = voltage
            ia = current
            ib = 0
            ic = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(exp_demand_p_sys_1st, exp_demand_s_sys_1st)
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st, exp_demand_ib_1st, exp_demand_ic_1st, exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·前半窗", prefix=f"{case_id} [4/5]")
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = voltage_2nd
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd),
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle),
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                # 第2周期电流需量取窗口平均(前半窗满值 + 后半窗半值, 各占 0.5), 与功率 2nd 口径一致
                exp_demand_ia_2nd = ia * 0.5 + ia_2nd * 0.5
                exp_demand_ib_2nd = ib * 0.5 + ib_2nd * 0.5
                exp_demand_ic_2nd = ic * 0.5 + ic_2nd * 0.5
                exp_demand_in_2nd = in_val * 0.5 + in_2nd * 0.5

                # 等待 后1/2 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·后半窗", prefix=f"{case_id} [5/5]")
                # 第二个周期开始计算
                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd, exp_demand_ib_2nd, exp_demand_ic_2nd, exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
        self.write_demand_report(file_path)

    def sliding_demand_test_by_2e3w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 180
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            ua = voltage
            ub = 0
            uc = voltage
            ia = current
            ib = 0
            ic = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = voltage_2nd
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第2周期·滑动更新", prefix=f"{case_id} [4/5]")
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_update_rate / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd)
                ]) * demand_update_rate / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = round(
                    ia * (demand_interval - demand_update_rate) / demand_interval
                    + ia_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ib_2nd = round(
                    ib * (demand_interval - demand_update_rate) / demand_interval
                    + ib_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ic_2nd = round(
                    ic * (demand_interval - demand_update_rate) / demand_interval
                    + ic_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_in_2nd = round(
                    in_val * (demand_interval - demand_update_rate) / demand_interval
                    + in_2nd * demand_update_rate / demand_interval, 5)

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第3周期·滑动更新", prefix=f"{case_id} [5/5]")
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                ua_3rd = voltage_3rd
                ub_3rd = 0
                uc_3rd = voltage_3rd
                ia_3rd = current_3rd
                ib_3rd = 0
                ic_3rd = current_3rd
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle),
                    self.calculate_active_power(uc_3rd, ic_3rd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd),
                    self.calculate_apparent_power(uc_3rd, ic_3rd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * demand_update_rate / demand_interval
                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * demand_update_rate / demand_interval
                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                        self.calculate_active_power(uc, ic, vc_angle)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval
                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                        self.calculate_apparent_power(uc, ic)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval
                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * (demand_interval - demand_update_rate) / demand_interval
                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * (demand_interval - demand_update_rate) / demand_interval
                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                if demand_interval - demand_update_rate >= demand_update_rate:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * demand_update_rate / demand_interval
                        + ia * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * demand_update_rate / demand_interval
                        + ib * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * demand_update_rate / demand_interval
                        + ic * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * demand_update_rate / demand_interval
                        + in_val * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                else:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)
        self.write_demand_report(file_path)

    def fixed_demand_test_by_1e2w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_1e2w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))
            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 0
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = 0
            ua = voltage
            ub = 0
            uc = 0
            ia = current
            ib = 0
            ic = 0
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st, exp_demand_ib_1st, exp_demand_ic_1st, exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·前半窗", prefix=f"{case_id} [4/5]")
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = 0
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = 0
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd)
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval
                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval
                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_p_sys_2nd_by_down_half
                ])
                exp_demand_q_sys_2nd = sum([
                    exp_demand_q_sys_2nd_by_upper_half,
                    exp_demand_q_sys_2nd_by_down_half
                ])
                exp_demand_s_sys_2nd = sum([
                    exp_demand_s_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_down_half
                ])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                # 第2周期电流需量取窗口平均(前半窗满值 + 后半窗半值, 各占 0.5), 与功率 2nd 口径一致
                exp_demand_ia_2nd = ia * 0.5 + ia_2nd * 0.5
                exp_demand_ib_2nd = ib * 0.5 + ib_2nd * 0.5
                exp_demand_ic_2nd = ic * 0.5 + ic_2nd * 0.5
                exp_demand_in_2nd = in_val * 0.5 + in_2nd * 0.5

                # 等待 后1/2 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·后半窗", prefix=f"{case_id} [5/5]")
                # 第二个周期开始计算
                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd, exp_demand_ib_2nd, exp_demand_ic_2nd, exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_1e2w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_1e2w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
        self.write_demand_report(file_path)

    def sliding_demand_test_by_1e2w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_1e2w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 0
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = 0
            ua = voltage
            ub = 0
            uc = 0
            ia = current
            ib = 0
            ic = 0
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = 0
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = 0
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第2周期·滑动更新", prefix=f"{case_id} [4/5]")
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                ]) * demand_update_rate / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd)
                ]) * demand_update_rate / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = round(
                    ia * (demand_interval - demand_update_rate) / demand_interval
                    + ia_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ib_2nd = round(
                    ib * (demand_interval - demand_update_rate) / demand_interval
                    + ib_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ic_2nd = round(
                    ic * (demand_interval - demand_update_rate) / demand_interval
                    + ic_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_in_2nd = round(
                    in_val * (demand_interval - demand_update_rate) / demand_interval
                    + in_2nd * demand_update_rate / demand_interval, 5)

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第3周期·滑动更新", prefix=f"{case_id} [5/5]")
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                ua_3rd = voltage_3rd
                ub_3rd = 0
                uc_3rd = 0
                ia_3rd = current_3rd
                ib_3rd = 0
                ic_3rd = 0
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                    ]) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                    ]) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                if demand_interval - demand_update_rate >= demand_update_rate:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * demand_update_rate / demand_interval
                        + ia * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * demand_update_rate / demand_interval
                        + ib * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * demand_update_rate / demand_interval
                        + ic * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * demand_update_rate / demand_interval
                        + in_val * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                else:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_1e2w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_1e2w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)
        self.write_demand_report(file_path)

    def fixed_demand_test_by_3e4wy(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 仅对期望真实累积需量的用例(有压有流)校验, 与下方 pass 判定分支同条件;
            # 源没开/无输出则此处立即失败, 避免白等整个需量周期。
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·前半窗", prefix=f"{case_id} [4/5]")
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd)
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                # 第2周期电流需量取窗口平均(前半窗满值 + 后半窗半值, 各占 0.5), 与功率 2nd 口径一致
                exp_demand_ia_2nd = ia * 0.5 + ia_2nd * 0.5
                exp_demand_ib_2nd = ib * 0.5 + ib_2nd * 0.5
                exp_demand_ic_2nd = ic * 0.5 + ic_2nd * 0.5
                exp_demand_in_2nd = in_val * 0.5 + in_2nd * 0.5

                # 等待 后1/2 * demand_interval后
                self.sleep_with_progress(0.5 * demand_interval * 60, "第2周期·后半窗", prefix=f"{case_id} [5/5]")
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wy = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wy, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
        self.write_demand_report(file_path)

    def sliding_demand_test_by_3e4wy(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval, wire_type) = self.get_test_case_info_of_input_value(input_list, i)
            self._begin_case((case_id, demand_accuracy, voltage, vc_angle, current, freq, wire_type, demand_method, demand_interval, demand_update_rate, demand_trigger))

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 源在线探针: 有压有流的用例, 加源后回读表实时值, 源没开/无输出即快速失败
            if voltage >= 9.5 and current != 0:
                self.assert_source_output_alive([ua, ub, uc], [ia, ib, ic])
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            self.sleep_with_progress(300, "清零沉降(5min)", prefix=f"{case_id} [1/5]")
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            self.sleep_with_progress(wait_time, "对齐需量窗口起点", prefix=f"{case_id} [2/5]")
            # 需量第一次上报时间
            self.sleep_with_progress(demand_interval * 60, "第1个需量窗口累积", prefix=f"{case_id} [3/5]")
            if voltage < 9.5 or current == 0:
                self._record_clear_case(self.check_demand_power_current_is_clear(tolerance=demand_accuracy))
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第2周期·滑动更新", prefix=f"{case_id} [4/5]")
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = round(
                    ia * (demand_interval - demand_update_rate) / demand_interval
                    + ia_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ib_2nd = round(
                    ib * (demand_interval - demand_update_rate) / demand_interval
                    + ib_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_ic_2nd = round(
                    ic * (demand_interval - demand_update_rate) / demand_interval
                    + ic_2nd * demand_update_rate / demand_interval, 5)
                exp_demand_in_2nd = round(
                    in_val * (demand_interval - demand_update_rate) / demand_interval
                    + in_2nd * demand_update_rate / demand_interval, 5)

                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                self.sleep_with_progress(demand_update_rate * 60, "第3周期·滑动更新", prefix=f"{case_id} [5/5]")
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle),
                    self.calculate_active_power(ub_3rd, ib_3rd, vc_angle),
                    self.calculate_active_power(uc_3rd, ic_3rd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd),
                    self.calculate_apparent_power(ub_3rd, ib_3rd),
                    self.calculate_apparent_power(uc_3rd, ic_3rd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                        self.calculate_active_power(ub, ib, vc_angle),
                        self.calculate_active_power(uc, ic, vc_angle)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                        self.calculate_apparent_power(ub, ib),
                        self.calculate_apparent_power(uc, ic)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                if demand_interval - demand_update_rate >= demand_update_rate:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * demand_update_rate / demand_interval
                        + ia * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * demand_update_rate / demand_interval
                        + ib * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * demand_update_rate / demand_interval
                        + ic * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * demand_update_rate / demand_interval
                        + in_val * (demand_interval - 2 * demand_update_rate) / demand_interval, 5)
                else:
                    exp_demand_ia_3rd = round(
                        ia_3rd * demand_update_rate / demand_interval
                        + ia_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ib_3rd = round(
                        ib_3rd * demand_update_rate / demand_interval
                        + ib_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_ic_3rd = round(
                        ic_3rd * demand_update_rate / demand_interval
                        + ic_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)
                    exp_demand_in_3rd = round(
                        in_3rd * demand_update_rate / demand_interval
                        + in_2nd * (demand_interval - demand_update_rate) / demand_interval, 5)

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wy = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq, wire_type,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wy, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)
        self.write_demand_report(file_path)


_ORIG_SLEEP = None


def _patch_fast_sleep(cap=0.5):
    """DEMAND_FAST: 把 time.sleep 压到 <=cap 秒, 仅用于快速验证代码链路(需量不累积)。"""
    global _ORIG_SLEEP
    if _ORIG_SLEEP is not None:
        return
    _ORIG_SLEEP = time.sleep

    def _fast(seconds):
        _ORIG_SLEEP(min(seconds, cap))

    time.sleep = _fast


def _restore_sleep():
    """还原被 _patch_fast_sleep 替换的 time.sleep。"""
    global _ORIG_SLEEP
    if _ORIG_SLEEP is not None:
        time.sleep = _ORIG_SLEEP
        _ORIG_SLEEP = None


def _sync_project_conn_config():
    """把分层配置(global.yaml ← projects/RPP/config.yaml)的连接参数同步进 comm 层
    共享的 modbus_config, 使本(老版)demand 测试也尊重项目级 config.yaml, 而非只吃 global。
    RPP 暂无 config.yaml 时, load_config 自动回退为纯 global 配置。

    必须在建立 Modbus 连接(DemandTest())之前调用: ModbusRtuOrTcp 在 __init__ 时读取
    modbus_config['conn_mode']/['rtu']/['tcp'], 此处就地更新同一 dict 对象即可生效。
    注: 同目录 acuvimseries_modbus_get 仍直接读 global, 未走本同步; 若项目 config 与
    global 的 rtu 分叉, 该模块不会跟随(需另行改造), 故当前二者的 rtu 仍应在 global 中保持一致。
    """
    from framework.config.loader import load_config
    from modbus_config import modbus_config
    merged = load_config("RPP")
    if merged.get("conn_mode"):
        modbus_config["conn_mode"] = merged["conn_mode"]
    for key in ("rtu", "tcp"):
        if isinstance(merged.get(key), dict):
            modbus_config.setdefault(key, {}).update(merged[key])
    logging.info("demand 连接参数已按项目 config 同步: conn_mode=%s rtu=%s",
                 modbus_config.get("conn_mode"), modbus_config.get("rtu"))


def run_demand_test_script(test_type, wire_type, demand_type):
    """
    运行脚本入口
    :param test_type: 0/1:mA, 2:mV, 3:rct (以 select_test_case 映射为准)
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 5:3e4wD  #AcuRev1320
    :param demand_type: 0:fixed, 1:slid
    :return: 逐条需量精度判定结果列表(供 pytest 层断言)
    """
    _sync_project_conn_config()  # 先按项目 config 同步连接参数, 再建 Modbus 连接
    if _FAST:
        _patch_fast_sleep()
    print("====================Demand Test Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    try:
        switch_device_screen_interface(inter=0x01)  # 切换至交流界面
        set_gear_switching_mode('00000000')  # 档位切换归零,自动
        demand_test = DemandTest()

        demand_test.select_test_case(test_type=test_type, wire_type=wire_type, demand_type=demand_type)

        # 关闭ModbusClient客户端连接
        demand_test.handle_memory.modbus_client.close()
        up_source_ac()
        switch_device_screen_interface(inter=0x00)  # 切换至默认界面
        set_gear_switching_mode(mode='01000000')  # 档位切换归零,手动
        print(f"====================测试总耗时:{time.time() - start_time}====================")
        print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
        print("====================Demand Test End====================")
        # 返回逐条需量精度判定结果, 供 pytest 层做真实断言。
        return demand_test.case_results
    finally:
        if _FAST:
            _restore_sleep()


# ---------------------------------------------------------------------------
# pytest 入口
# ---------------------------------------------------------------------------
# 说明:
# - 需量用例需真实功率源 + 电表在环, 单条接线方式动辄数十分钟(含多次 5 分钟级
#   sleep), 且会持续改变设备状态; 统一打 slow + destructive 标记, 需显式选择运行。
# - 按团队约定的"接线方式 x 触发方式"两个维度参数化: 每个组合为一个独立用例项,
#   通过/失败独立报告; 测量模式(mV/mA/rct)由命令行选项 --measure 选定(见 conftest.py),
#   三项配置同一条 pytest 命令内搞定, 不再用环境变量。
# - option 1(去假绿): Excel 中无该组合数据时直接 pytest.skip(标黄), 不再空跑判 PASS。
# - option 2(真断言): 逐条需量精度判定结果由 DemandTest.case_results 累积, 本层据此
#   断言"有判定且全部通过"; 任一条精度不达标即判该组合 FAILED(不再只写 xlsx)。

# 测量模式: select_test_case 实际映射 0/1 -> mA, 2 -> mV, 3 -> rct
# 测量模式 -> sheet 名(与 select_test_case 内的映射保持一致)
_MEASURE_MODE_TO_SHEET = {
    0: sheet_name_mA,
    1: sheet_name_mA,
    2: sheet_name_mV,
    3: sheet_name_rct,
}

# --measure 选项名(mv/ma/rct) -> 测量模式取值; 也接受直接传 0/1/2/3
_MEASURE_NAME_TO_MODE = {"ma": 0, "mv": 2, "rct": 3}
# 缺省测量模式(未传 --measure 时): 2 = mV
_DEFAULT_MEASURE_MODE = 2


def _resolve_measure_mode(config):
    """解析本次运行的测量模式(mV/mA/rct)。

    单一入口: pytest 命令行 --measure(mv|ma|rct 或 0/1/2/3), 缺省 mV。让接线/触发(-k)
    与测量模式(--measure)在同一条 pytest 命令内选定, 无需环境变量, 无需改代码。
    """
    raw = config.getoption("--measure")
    if raw is None:
        return _DEFAULT_MEASURE_MODE
    key = str(raw).strip().lower()
    if key in _MEASURE_NAME_TO_MODE:
        return _MEASURE_NAME_TO_MODE[key]
    if key.isdigit() and int(key) in _MEASURE_MODE_TO_SHEET:
        return int(key)
    raise pytest.UsageError(
        f"--measure 取值非法: {raw!r}; 应为 mv|ma|rct 或 0/1/2/3"
    )

# wire_type 取值 -> 接线方式标识(与 ALL_SAVE_DIRS 一致)
_WIRE_TYPES = [
    (0, "1e2w1p"),
    (1, "2e3w1p"),
    (2, "2e3wd"),
    (3, "2e3wn"),
    (4, "3e4wy"),
    (5, "3e4wd"),
]
# demand_type 取值 -> 触发(计算)方式
_DEMAND_TYPES = [(0, "fixed"), (1, "sliding")]

_DEMAND_PARAMS = [
    pytest.param(wire_type, demand_type, id=f"{wire_name}-{demand_name}")
    for wire_type, wire_name in _WIRE_TYPES
    for demand_type, demand_name in _DEMAND_TYPES
]


def _count_demand_cases(test_type, wire_type, demand_type):
    """统计 Excel 中某(测量模式, 接线方式, 触发方式)组合实际命中的用例条数。

    与 select_test_case/select_wire_type 的筛选口径一致, 仅读 Excel、不连设备,
    供 option 1 在运行前判断是否需要 skip。
    """
    sheet_name = _MEASURE_MODE_TO_SHEET.get(test_type)
    if not sheet_name:
        return 0
    try:
        source_input_list = data_read(test_case_path, sheet_name)
    except KeyError:
        # sheet 不存在(如 test_case_rct 尚未建) -> 视为 0 条 -> 交由上层 pytest.skip,
        # 不因缺 sheet 崩溃报错。
        return 0
    by_wire = DemandTest.get_input_list_by_wire_type(source_input_list, wire_type)
    final = DemandTest.get_demand_para_by_input_list_of_wire_type(by_wire, demand_type)
    return max(len(final) - 1, 0)  # 减去表头行


@pytest.mark.slow
@pytest.mark.destructive
@pytest.mark.parametrize("wire_type, demand_type", _DEMAND_PARAMS)
def test_demand(wire_type, demand_type, pytestconfig):
    """需量测试: 按 接线方式 x 触发方式 组合驱动。

    三项配置均在一条 pytest 命令内选定, 无需环境变量:
      - 接线方式 / 触发方式: pytest 原生 -k 选参数化 id(如 -k "3e4wy-fixed");
      - 测量模式(mV/mA/rct): 命令行 --measure(mv|ma|rct, 缺省 mv)。

    option 1: 无该组合数据时 skip(不假绿);
    option 2: 对逐条需量精度判定结果做真实断言(有判定且全部通过)。
    """
    measure_mode = _resolve_measure_mode(pytestconfig)
    # option 1: Excel 无该组合用例 -> 跳过, 避免"空跑即 PASS"的假绿
    case_count = _count_demand_cases(measure_mode, wire_type, demand_type)
    if case_count == 0:
        pytest.skip(
            f"{_MEASURE_MODE_TO_SHEET.get(measure_mode)} 中无 "
            f"wire_type={wire_type} demand_type={demand_type} 的用例数据"
        )

    results = run_demand_test_script(
        test_type=measure_mode, wire_type=wire_type, demand_type=demand_type
    )

    # option 2: 真实断言 —— 必须产生了精度判定, 且全部通过
    failed = [r for r in results if not r]
    assert results, (
        f"该组合有 {case_count} 条用例但未产生任何需量精度判定"
        f"(可能全部因电压<9.5/电流=0 走了清零分支), 请检查测试数据"
    )
    assert not failed, (
        f"需量精度判定失败 {len(failed)}/{len(results)} 条, 详见结果 xlsx"
    )
