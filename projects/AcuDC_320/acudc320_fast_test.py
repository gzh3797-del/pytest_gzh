#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
文件名称:acudc320_fast_test.py
功能描述:快速测量
创建日期:2025-08-28
作者:王洋
版本:v1.0
修改记录:
"""

import os
import statistics
import time
import math
import logging
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from acudc320_modbus_get import HandleMemory
from comm.source_control import (switch_device_screen_interface, set_gear_switching_mode, set_ac, set_voltage_gear,
                                 set_current_gear, up_source_ac, sour_output, sour_stop, mv_sour_output)
from projects.AcuDC_320.acudc320_memory_info import MemoryAddr
from projects.AcuDC_320.acudc320_test_report_table_heading import TableTitle
from tools.excel_operate import data_read
from tools.log import Log
from power_calculate import CalculatePower

Log(str(__file__).split("\\")[-1])

test_case_path = r'./test_case/AcuDC_320/acudc320_test_case.xlsx'
sheet_name_dc260 = "Sheet2"

save_filedir = os.path.join(Path(__file__).parent, f"precision_measure_{time.strftime('%Y%m%d')}")
if not os.path.exists(save_filedir):
    os.makedirs(save_filedir, exist_ok=True)


class PrecisionMeasure:
    def __init__(self):
        self.handle_memory = HandleMemory(slave_id=1)

    def select_test_case(self, measure_mode):
        """
        选择测试标准
        :param measure_mode:测试类型
        :return:
        """
        case_path = test_case_path
        sheet_name = sheet_name_dc260
        self.select_wire_type(case_path, sheet_name, measure_mode)

    @staticmethod
    def get_real_time_parameters_by_input_list_of_wire_type(source_input_list):
        """
        获取接线方式的快速测量实时数据
        :param source_input_list: 接线方式的数据
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            if not source_input_list[i][4]:
                input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def write_table_title_of_dc260_to_excel(file_path, wb, ws):
        """
        写入表头:dc260
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_DC320)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.REAL_TIME_MEASURE_COLUMNS_OF_DC320[i]}')
        wb.save(file_path)

    @staticmethod
    def get_accuracy_by_datatype(accuracy_value):
        """
        处理N/A精度值
        :param accuracy_value: 精度值
        :return: 精度值
        """
        accuracy_res = accuracy_value
        if not isinstance(accuracy_value, float):
            accuracy_res = 999999
        return accuracy_res

    def get_test_case_info_of_accuracy(self, input_list, index_value):
        """
        获取测试精度信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 精度值
        """
        accuracy_of_voltage = input_list[index_value + 1][5]
        accuracy_of_current = input_list[index_value + 1][6]
        accuracy_of_power = input_list[index_value + 1][7]
        voltage_accuracy = accuracy_of_voltage
        current_accuracy = accuracy_of_current
        power_accuracy = accuracy_of_power
        return voltage_accuracy, current_accuracy, power_accuracy

    @staticmethod
    def get_test_case_info_of_input_value(input_list, index_value):
        """
        获取测试电压/电流、相位等信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 电压/电流、相位等信息
        """
        case_id = input_list[index_value + 1][0]
        u = input_list[index_value + 1][1]
        i1 = input_list[index_value + 1][2]
        i2 = input_list[index_value + 1][3]
        sample_cnt = input_list[index_value + 1][8]
        sample_interval = input_list[index_value + 1][9]
        return case_id, u, i1, i2, sample_cnt, sample_interval

    @staticmethod
    def get_power_cal_res_of_phase(value_of_voltage, value_of_current):
        """
        通过电压计算功率
        :param value_of_voltage: 电压
        :param value_of_current: 电流
        :return: 功率
        """
        apparent_power_res = CalculatePower.calculate_apparent_power(value_of_voltage, value_of_current)
        return round(apparent_power_res / 1000, 5)

    # @staticmethod
    # def get_sys_power(power_of_phase_a, power_of_phase_b, power_of_phase_c):
    #     """
    #     计算系统功率
    #     :param power_of_phase_a: A相功率
    #     :param power_of_phase_b: B相功率
    #     :param power_of_phase_c: C相功率
    #     :return: 系统功率
    #     """
    #     sys_power = sum([power_of_phase_a, power_of_phase_b, power_of_phase_c])
    #     return sys_power

    @staticmethod
    def get_sys_power(*power_of_phase):
        """
        计算系统功率
        :param power_of_phase: A/B/C相功率
        :return: 系统功率
        """
        sys_power = sum(power_of_phase)
        return sys_power

    @staticmethod
    def write_common_values_to_excel(ws, index_value, common_values, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param common_values: case编号/电压/电流/电压相位角/电流相位角等输入值,精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(common_values)):
            ws.cell(index_value + 2, start_num + k, common_values[k])
        start_num += len(common_values)
        return start_num

    @staticmethod
    def write_accuracy_res_to_excel(file_path, wb, ws, index_value, accuracy_res, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param file_path: 待写入文件路径
        :param wb: 写入工作簿对象
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param accuracy_res: 精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(accuracy_res)):
            ws.cell(index_value + 2, start_num + k, accuracy_res[k])
        wb.save(file_path)
        start_num += len(accuracy_res)
        return start_num

    # def set_and_switch_source(self,uc_value, ub_value, ua_value,ic_value, ib_value, ia_value):
    #     """
    #     升源加设置档位
    #     :return:
    #     """
    #     # 设置电压/电流档位，当前是相同值，只配置了一种A相
    #     set_voltage_gear(uc_value, ub_value, ua_value)
    #     set_current_gear(ic_value, ib_value, ia_value)

    @staticmethod
    def get_line_current_angle(phase_current_angle):
        """
        delta类型接线方式，获取线电流角度
        :param phase_current_angle: 相电流角度
        :return: 线电流角度
        """
        phase_current_angle = phase_current_angle if ((phase_current_angle - 30) >= 0) else (phase_current_angle + 360)
        return phase_current_angle

    def get_phase_angle_of_line_current(self, ia_angle, ib_angle, ic_angle):
        """
        delta类型接线方式，获取线电流角度
        :param ia_angle: ia角度
        :param ib_angle: ib角度
        :param ic_angle: ic角度
        :return: 线电流角度
        """
        ia_angle = self.get_line_current_angle(ia_angle)
        ib_angle = self.get_line_current_angle(ib_angle)
        ic_angle = self.get_line_current_angle(ic_angle)
        return ia_angle, ib_angle, ic_angle

    def get_measure_values_of_phase_voltage(self):
        """
        获取相电压测量值
        :return: 相电压测量值
        """
        u = self.handle_memory.read_u_voltage()
        return u

    def get_measure_values_of_phase_current(self):
        """
        获取电流测量值
        :return: 电流测量值
        """
        i1 = self.handle_memory.read_i1_current()
        # i2 = self.handle_memory.read_i2_current()
        return i1

    def get_measure_values_of_power(self):
        """
        获取有功功率测量值
        :return: 有功功率测量值
        """
        p1 = self.handle_memory.read_p1_power()
        return p1

    def get_measure_values_by_sample_cnt_of_dc260(self, sample_cnt, sample_interval):
        """
        获取测量值:dc260
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_u = []
        measure_values_of_i1 = []
        measure_values_of_p1 = []
        for j in range(sample_cnt):
            time.sleep(sample_interval)
            u_measure = self.get_measure_values_of_phase_voltage()
            i1_measure = self.get_measure_values_of_phase_current()
            p1_measure = self.get_measure_values_of_power()
            measure_values_of_u.append(u_measure)
            measure_values_of_i1.append(i1_measure)
            measure_values_of_p1.append(p1_measure)
        return (
            measure_values_of_u,
            measure_values_of_i1,
            measure_values_of_p1,
        )

    @staticmethod
    def get_cmp_accuracy_res(act_accuracy_min, act_accuracy_max, act_accuracy_avg, exp_accuracy):
        """
        获取精度比较结果
        :param act_accuracy_min: 精度最小值
        :param act_accuracy_max: 精度最大值
        :param act_accuracy_avg: 精度平均值
        :param exp_accuracy: 精度期望值
        :return: "Passed"/"Failed"
        """
        cmp_res = "Failed"
        if (act_accuracy_min <= exp_accuracy) and (act_accuracy_max <= exp_accuracy) and (
                act_accuracy_avg <= exp_accuracy):
            cmp_res = "Passed"
        return cmp_res

    @staticmethod
    def get_all_accuracy_values(all_accuracy_values):
        """
        获取最终精度比较结果
        :param all_accuracy_values: 全部精度值列表
        :return: 最终精度比较结果
        """
        all_cmp_res = "Passed"
        if "Failed" in all_accuracy_values:
            all_cmp_res = "Failed"
        return all_cmp_res

    def get_accuracy_res_by_exp_accuracy(self, standard_value, measure_values, exp_accuracy):
        """
        获取精度比较结果，带期望精度值
        :param standard_value: 输入值
        :param measure_values: 测量值
        :param exp_accuracy: 期望精度值
        :return: 测量值和精度比较结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_voltage_current_power(standard_value, measure_values)

        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        cmp_accuracy_res = self.get_cmp_accuracy_res(act_accuracy_min=min_accuracy, act_accuracy_max=max_accuracy,
                                                     act_accuracy_avg=avg_accuracy, exp_accuracy=exp_accuracy)
        accuracy_res = (min_value, min_accuracy, max_value, max_accuracy, avg_value, avg_accuracy, cmp_accuracy_res)
        return accuracy_res

    def get_accuracy_res_by_not_exp_accuracy(self, standard_value, measure_values):
        """
        获取精度比较结果，不带期望精度值
        :param standard_value: 输入值
        :param measure_values: 测量值
        :return: 测量值结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_voltage_current_power(standard_value, measure_values)

        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        accuracy_res = (min_value, min_accuracy, max_value, max_accuracy, avg_value, avg_accuracy)
        return accuracy_res

    def fast_precision_measure_by_dc260_ui(self, file_path, input_list):
        """
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_dc260_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, power_accuracy) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, u, i1, _, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            p1 = self.get_power_cal_res_of_phase(u, i1)

            start_num = 1
            common_values_of_dc260 = (
                case_id, voltage_accuracy, current_accuracy, power_accuracy, u, i1,
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_dc260, start_num)
            # 升源+设置电压/电流档位
            sour_output(voltage=u, current=i1)

            (
                measure_values_of_u,
                measure_values_of_i1,
                measure_values_of_p1,
            ) = self.get_measure_values_by_sample_cnt_of_dc260(sample_cnt, sample_interval)
            u_accuracy = self.get_accuracy_res_by_exp_accuracy(u, measure_values_of_u, voltage_accuracy)
            i1_accuracy = self.get_accuracy_res_by_exp_accuracy(i1, measure_values_of_i1, current_accuracy)
            p1_accuracy = self.get_accuracy_res_by_exp_accuracy(p1, measure_values_of_p1, power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, u_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, i1_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p1], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p1_accuracy, start_num)
            # start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p2], start_num)
            # start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p2_accuracy, start_num)
            # start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            # self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

    def fast_precision_measure_by_dc260_u(self, file_path, input_list):
        """
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_dc260_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, power_accuracy) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, u, i1, i2, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            i1 = 0
            i2 = 0
            p1 = self.get_power_cal_res_of_phase(u, i1)
            p2 = self.get_power_cal_res_of_phase(u, i2)
            p_sys = self.get_sys_power(p1, p2)

            start_num = 1
            common_values_of_dc260 = (
                case_id, voltage_accuracy, current_accuracy, power_accuracy, u, i1, i2
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_dc260, start_num)
            # 升源+设置电压/电流档位
            sour_output(voltage=u, current=i1)
            (
                measure_values_of_u,
                measure_values_of_i1,
                measure_values_of_i2,
                measure_values_of_p1,
                measure_values_of_p2,
                measure_values_of_p_sys,
            ) = self.get_measure_values_by_sample_cnt_of_dc260(sample_cnt, sample_interval)
            u_accuracy = self.get_accuracy_res_by_exp_accuracy(u, measure_values_of_u, voltage_accuracy)
            i1_accuracy = self.get_accuracy_res_by_exp_accuracy(i1, measure_values_of_i1, current_accuracy)
            i2_accuracy = self.get_accuracy_res_by_exp_accuracy(i1, measure_values_of_i2, current_accuracy)
            p1_accuracy = self.get_accuracy_res_by_exp_accuracy(p1, measure_values_of_p1, power_accuracy)
            p2_accuracy = self.get_accuracy_res_by_exp_accuracy(p2, measure_values_of_p2, power_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys, power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, u_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, i1_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, i2_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p1], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p1_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p2], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p2_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

    def fast_precision_measure_by_dc260_i(self, file_path, input_list):
        """
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :return:
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.write_table_title_of_dc260_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            (voltage_accuracy, current_accuracy, power_accuracy) = self.get_test_case_info_of_accuracy(input_list, i)
            (case_id, u, i1, i2, sample_cnt, sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            u = 0
            p1 = self.get_power_cal_res_of_phase(u, i1)
            p2 = self.get_power_cal_res_of_phase(u, i2)
            p_sys = self.get_sys_power(p1, p2)

            start_num = 1
            common_values_of_dc260 = (
                case_id, voltage_accuracy, current_accuracy, power_accuracy, u, i1, i2
            )
            start_num = self.write_common_values_to_excel(ws, i, common_values_of_dc260, start_num)
            # 升源+设置电压/电流档位
            sour_output(voltage=u, current=i1)
            (
                measure_values_of_u,
                measure_values_of_i1,
                measure_values_of_i2,
                measure_values_of_p1,
                measure_values_of_p2,
                measure_values_of_p_sys,
            ) = self.get_measure_values_by_sample_cnt_of_dc260(sample_cnt, sample_interval)
            u_accuracy = self.get_accuracy_res_by_exp_accuracy(u, measure_values_of_u, voltage_accuracy)
            i1_accuracy = self.get_accuracy_res_by_exp_accuracy(i1, measure_values_of_i1, current_accuracy)
            i2_accuracy = self.get_accuracy_res_by_exp_accuracy(i1, measure_values_of_i2, current_accuracy)
            p1_accuracy = self.get_accuracy_res_by_exp_accuracy(p1, measure_values_of_p1, power_accuracy)
            p2_accuracy = self.get_accuracy_res_by_exp_accuracy(p2, measure_values_of_p2, power_accuracy)
            p_sys_accuracy = self.get_accuracy_res_by_exp_accuracy(p_sys, measure_values_of_p_sys, power_accuracy)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, u_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, i1_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, i2_accuracy, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p1], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p1_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p2], start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, p2_accuracy, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [p_sys], start_num)
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, p_sys_accuracy, start_num)

    @staticmethod
    def get_save_filepath_of_dc260(filedir):
        """
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"precision_measure_dc320_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    def select_wire_type(self, case_path, sheet_name, measure_mode):
        """
        选择接线方式
        :param case_path: 测试文件名
        :param sheet_name: sheet名
        :param measure_mode: 测试类型
        :return:
        """
        source_input_list = data_read(case_path, sheet_name)
        input_list = self.get_real_time_parameters_by_input_list_of_wire_type(source_input_list)
        file_path = self.get_save_filepath_of_dc260(save_filedir)
        if measure_mode == 0:
            self.fast_precision_measure_by_dc260_ui(file_path, input_list)
        elif measure_mode == 1:
            self.fast_precision_measure_by_dc260_u(file_path, input_list)
        elif measure_mode == 2:
            self.fast_precision_measure_by_dc260_i(file_path, input_list)


def run_precision_measure_script(measure_mode):
    """
    运行脚本入口
    """
    print(f"====================Precision Measure Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    precision_measure = PrecisionMeasure()
    precision_measure.select_test_case(measure_mode)

    # 关闭ModbusClient客户端连接
    precision_measure.handle_memory.modbus_client.close()

    # sour_output(0, 0)
    # sour_stop()

    print(f"====================测试总耗时:{time.time() - start_time}====================")
    print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
    print(f"====================Precision Measure End====================")


if __name__ == '__main__':
    """
    程序主入口
    :param measure_mode: 0:ui, 1:u, 2,i
    """
    measure_type = 0
    run_precision_measure_script(measure_mode=measure_type)
