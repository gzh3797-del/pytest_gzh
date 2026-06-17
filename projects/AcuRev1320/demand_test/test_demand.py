import cmath
import itertools
import logging

import openpyxl
from openpyxl import Workbook
from datetime import datetime, time
from comm.source_control import *
import math

from tools.excel_operate import data_read
from projects.AcuRev1320.fast_test.acuvimseries_modbus_get import HandleMemory
from projects.AcuRev1320.demand_test.demand_table_heading import TableTitle

Log(str(__file__).split("\\")[-1])
CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
test_case_path = os.path.join(CURRENT_PATH, "demand_test_case.xlsx")

sheet_name_mV = "test_case_mV"
sheet_name_mA = "test_case_mA"
sheet_name_rct = "test_case_rct"

ALL_SAVE_DIRS_BY_WIRE_MODE = {
    0: "save_filedir_1e2w1p",
    1: "save_filedir_2e3w1p",
    2: "save_filedir_2e3wd",
    3: "save_filedir_2e3wn",
    4: "save_filedir_3e4wy",
    5: "save_filedir_3e4wd",
}

ALL_SAVE_DIRS = ["1E2W1P", "2E3W1P", "2E3WD", "2E3WN", "3E4WY", "3E4WD"]


# save_filedir = os.path.join(
#         CURRENT_PATH, f"{time.strftime('%Y%m%d')}", "Demand_3E4WY",
#         f"demand_test_res_{time.strftime('%Y%m%d%H%M%S')}"
#     )
#     if not os.path.exists(save_filedir):
#         os.makedirs(save_filedir, exist_ok=True)

def init_filepath():
    for i in range(len(ALL_SAVE_DIRS)):
        ALL_SAVE_DIRS_BY_WIRE_MODE[i] = os.path.join(
            CURRENT_PATH,
            f"{time.strftime('%Y%m%d')}",
            f"Demand_{ALL_SAVE_DIRS[i]}",
            f"demand_test_res_{time.strftime('%Y%m%d%H%M%S')}"
        )
        if not os.path.exists(ALL_SAVE_DIRS_BY_WIRE_MODE[i]):
            os.makedirs(ALL_SAVE_DIRS_BY_WIRE_MODE[i], exist_ok=True)
    return ALL_SAVE_DIRS_BY_WIRE_MODE


class DemandTest:
    def __init__(self):
        self.handle_memory = HandleMemory(slave_id=1)
        init_filepath()

    def select_test_case(self, test_type, wire_type, demand_type):
        test_type_to_sheet = {
            0: sheet_name_mA,
            1: sheet_name_mA,
            2: sheet_name_mV,
            3: sheet_name_rct,
        }
        for key in test_type_to_sheet.keys():
            if test_type == key:
                self.select_wire_type(test_case_path, test_type_to_sheet[key], wire_type, demand_type)
                return

    def select_wire_type(self, case_path, sheet_name, wire_type, demand_type):
        """
        选择接线方式
        :param case_path: 测试文件名
        :param sheet_name: sheet名
        :param wire_type: 接线方式
        :param demand_type: 触发方式
        """
        if wire_type == 0:
            self.set_wire_type(voltage_wire_value=0)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_1e2w1p(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_1e2w1p(file_path, input_list, demand_type)
        elif wire_type == 1:
            self.set_wire_type(voltage_wire_value=1)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_2e3w1p(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_2e3w1p(file_path, input_list, demand_type)
        elif wire_type == 2:
            self.set_wire_type(voltage_wire_value=2)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_2e3wd(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_2e3wd(file_path, input_list, demand_type)
        elif wire_type == 3:
            self.set_wire_type(voltage_wire_value=3)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_2e3wn(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_2e3wn(file_path, input_list, demand_type)
        elif wire_type == 4:
            self.set_wire_type(voltage_wire_value=4)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_3e4wy(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_3e4wy(file_path, input_list, demand_type)
        elif wire_type == 5:
            self.set_wire_type(voltage_wire_value=5)
            source_input_list = data_read(case_path, sheet_name)
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, wire_type)
            input_list = self.get_demand_para_by_input_list_of_wire_type(input_list_of_wire_type, demand_type)
            file_path = self.get_save_filepath_of_3e4wd(ALL_SAVE_DIRS_BY_WIRE_MODE[wire_type], demand_type)
            self.demand_test_by_3e4wd(file_path, input_list, demand_type)

    def set_wire_type(self, voltage_wire_value):
        """
        设置接线方式
        :param voltage_wire_value: 电压接线方式
        :return:
        """
        self.handle_memory.set_wire_mode_by_voltage(voltage_wire_mode=voltage_wire_value)

    @staticmethod
    def get_input_list_by_wire_type(source_input_list, wire_type):
        """
        获取接线方式的数据
        :param source_input_list: 测试用例数据
        :param wire_type: 接线类型
        :return: 接线方式的数据
        “”“ 函数基于IIV3代码， 修改了接线方式对应关系”“”
        """
        input_list = [source_input_list[0]]
        # AcuRev1320中 接线方式3E4wY对应值是4
        if wire_type == 0:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][5] == "1E2w1p":
                    input_list.append(source_input_list[i])
        elif wire_type == 1:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][5] == "2E3w1p":
                    input_list.append(source_input_list[i])
        elif wire_type == 2:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][5] == "2E3wD":
                    input_list.append(source_input_list[i])
        elif wire_type == 3:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][5] == "2E3wN":
                    input_list.append(source_input_list[i])
        elif wire_type == 4:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][5] == "3E4wY":
                    input_list.append(source_input_list[i])
        elif wire_type == 5:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][5] == "3E4wD":
                    input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_demand_para_by_input_list_of_wire_type(source_input_list, demand_type):
        """
        获取接线方式的快速测量实时数据
        :param source_input_list: 接线方式的数据
        :param demand_type
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            if source_input_list[i][6] == demand_type:
                input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_save_filepath_of_3e4wd(filedir, demand_type):
        """
        接线方式:3e4wd
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_3E4WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_3E4WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wn(filedir, demand_type):
        """
        接线方式:2e3wn
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_2E3WN_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_2E3WN_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3wd(filedir, demand_type):
        """
        接线方式:2e3wd
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_2E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_2E3WD_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_2e3w1p(filedir, demand_type):
        """
        接线方式:2e3w1p
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_2E3W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_2E3W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_1e2w1p(filedir, demand_type):
        """
        接线方式:1e2w1p
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_1E2W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_1E2W1P_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def get_save_filepath_of_3e4wy(filedir, demand_type):
        """
        接线方式:3e4wy
        :param filedir: 待写入文件目录
        :param demand_type
        :return: 待写入文件路径
        """
        filename = None
        if demand_type == 0:
            filename = f"Fixed_Demand_Test_3E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        if demand_type == 1:
            filename = f"Sliding_Demand_Test_3E4WY_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    def demand_test_by_3e4wd(self, file_path, input_list, demand_type):
        """
        接线方式:3e4wd
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_3e4wd(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_3e4wd(file_path, input_list)

    def demand_test_by_2e3wn(self, file_path, input_list, demand_type):
        """
        接线方式:2e3wn
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_2e3wn(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_2e3wn(file_path, input_list)

    def demand_test_by_2e3wd(self, file_path, input_list, demand_type):
        """
        接线方式:2e3wd
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_2e3wd(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_2e3wd(file_path, input_list)

    def demand_test_by_2e3w1p(self, file_path, input_list, demand_type):
        """
        接线方式:2e3w1p
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_2e3w1p(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_2e3w1p(file_path, input_list)

    def demand_test_by_1e2w1p(self, file_path, input_list, demand_type):
        """
        接线方式:1e2w1p
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_1e2w1p(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_1e2w1p(file_path, input_list)

    def demand_test_by_3e4wy(self, file_path, input_list, demand_type):
        """
        接线方式:3e4wy
        :param file_path: 待写入文件路径
        :param input_list: 接线方式的数据
        :param demand_type: 需量的计算方法
        :return:
        """
        if demand_type == 0:
            self.fixed_demand_test_by_3e4wy(file_path, input_list)
        elif demand_type == 1:
            self.sliding_demand_test_by_3e4wy(file_path, input_list)

    @staticmethod
    def write_table_title_of_3e4wy_to_excel(file_path, wb, ws):
        """
        写入表头:3e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_3E4WY)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.DEMAND_COLUMNS_OF_3E4WY}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_1e2w1p_to_excel(file_path, wb, ws):
        """
        写入表头:1e2w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_1E2W1P)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.DEMAND_COLUMNS_OF_1E2W1P}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3w1p_to_excel(file_path, wb, ws):
        """
        写入表头:2e3w1p
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_2E3W1P)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.DEMAND_COLUMNS_OF_2E3W1P}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wn_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wn
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_2E3WN)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.DEMAND_COLUMNS_OF_2E3WN}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_2e3wd_to_excel(file_path, wb, ws):
        """
        写入表头:2e3wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_2E3WD)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.DEMAND_COLUMNS_OF_2E3WD}')
        wb.save(file_path)

    @staticmethod
    def write_table_title_of_3e4wd_to_excel(file_path, wb, ws):
        """
        写入表头:3e4wd
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :return:
        """
        for i in range(len(TableTitle.DEMAND_COLUMNS_OF_3E4WD)):
            j = i + 1
            ws.cell(1, j, f'{TableTitle.DEMAND_COLUMNS_OF_3E4WD}')
        wb.save(file_path)

    @staticmethod
    def get_test_case_info_of_accuracy(input_list, index_value):
        """
        获取测试精度信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 精度值
        """
        accuracy_of_demand = input_list[index_value + 1][10]
        return accuracy_of_demand

    @staticmethod
    def get_test_case_info_of_input_value(input_list, index_value):
        """
        获取测试电压/电流、相位等信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: 电压/电流、相位等信息
        """
        case_id = input_list[index_value + 1][0]
        u = input_list[index_value + 1][1]
        ui_angle = input_list[index_value + 1][2]
        i = input_list[index_value + 1][3]
        freq = input_list[index_value + 1][4]
        method = input_list[index_value + 1][6]
        interval = input_list[index_value + 1][7]
        update_rate = input_list[index_value + 1][8]
        trigger = input_list[index_value + 1][9]
        sample_cnt = input_list[index_value + 1][11]
        sample_interval = input_list[index_value + 1][12]
        return case_id, u, ui_angle, i, freq, method, interval, update_rate, trigger, sample_cnt, sample_interval

    @staticmethod
    def calculate_active_power(voltage, current, voltage_current_angle):
        """
        计算有功功率
        :param voltage:电压
        :param current:电流
        :param voltage_current_angle:电压电流相位角度
        :return:
        """
        active_power = (voltage * current * math.cos(math.radians(voltage_current_angle)))
        return active_power

    @staticmethod
    def calculate_reactive_power_by_sin(voltage, current, voltage_current_angle):
        """
        计算有功功率
        :param voltage:电压
        :param current:电流
        :param voltage_current_angle:电压电流相位角度
        :return:
        """
        reactive_power = (voltage * current * math.sin(math.radians(voltage_current_angle)))
        return reactive_power

    @staticmethod
    def calculate_reactive_power(active_power, apparent_power):
        """
        计算有功功率
        :param active_power:有功功率
        :param apparent_power:视在功率
        :return:
        """
        reactive_power_pow2 = apparent_power ** 2 - active_power ** 2
        reactive_power = math.sqrt(reactive_power_pow2) if reactive_power_pow2 else 0
        return reactive_power

    @staticmethod
    def calculate_apparent_power(voltage, current):
        """
        计算有功功率
        :param voltage:电压
        :param current:电流
        :return:
        """
        apparent_power = (voltage * current)
        return apparent_power

    def set_demand_para(self, demand_method, demand_interval, demand_update_rate):
        """
        Args:
            demand_method: Fixed Window: 0  Sliding Window: 1
            demand_interval: 1~30 minute
            demand_update_rate: 1~30 minute
        Returns:
        """
        # 设置demand method, sliding or fixed
        self.handle_memory.set_demand_method(demand_method)
        self.handle_memory.set_demand_interval(demand_interval)
        self.handle_memory.set_demand_update_rate(demand_update_rate)

    def set_time_trigger(self, sys_millisecond):
        """
        sys_millisecond,0-999
        """
        self.handle_memory.set_sys_millisecond(sys_millisecond)

    def set_demand_trigger(self, clear_max_demand):
        """
        clear_max_demand
        0, None;
        1: clear
        """
        self.handle_memory.set_clear_max_demand(clear_max_demand)

    @staticmethod
    def check_demand_is_clear(measure_value, tolerance=0.01):
        """
        用于判断demand清除是否成功
        """
        if abs(measure_value) <= tolerance:
            return True, measure_value,
        return False, measure_value

    @staticmethod
    def check_demand_is_pass(standard_value, measure_value, tolerance=0.01):
        """
        用于判断demand是否符合预期
        """
        relative_error = round((measure_value - standard_value) / standard_value, 5)
        if abs(relative_error) <= tolerance:
            return True, measure_value,
        return False, measure_value

    def check_demand_power_is_pass(self, standard_power_values, tolerance):
        std_sys_active_power, std_sys_reactive_power, std_sys_apparent_power = standard_power_values
        sys_active_power = self.handle_memory.read_demand_sys_active_power()
        p_res, p_val = self.check_demand_is_pass(std_sys_active_power, sys_active_power, tolerance)
        sys_reactive_power = self.handle_memory.read_demand_sys_reactive_power()
        q_res, q_val = self.check_demand_is_pass(std_sys_reactive_power, sys_reactive_power, tolerance)
        sys_apparent_power = self.handle_memory.read_demand_sys_apparent_power()
        s_res, s_val = self.check_demand_is_pass(std_sys_apparent_power, sys_apparent_power, tolerance)
        return p_res, q_res, s_res, p_val, q_val, s_val

    def check_demand_current_is_pass(self, standard_current_values, tolerance):
        std_demand_ia, std_demand_ib, std_demand_ic, std_demand_in = standard_current_values
        demand_ia = self.handle_memory.read_demand_ia()
        ia_res, ia_val = self.check_demand_is_pass(std_demand_ia, demand_ia, tolerance)
        demand_ib = self.handle_memory.read_demand_ib()
        ib_res, ib_val = self.check_demand_is_pass(std_demand_ib, demand_ib, tolerance)
        demand_ic = self.handle_memory.read_demand_ic()
        ic_res, ic_val = self.check_demand_is_pass(std_demand_ic, demand_ic, tolerance)
        demand_in = self.handle_memory.read_demand_in()
        in_res, in_val = self.check_demand_is_pass(std_demand_in, demand_in, tolerance)
        return ia_res, ib_res, ic_res, in_res, ia_val, ib_val, ic_val, in_val

    def check_demand_power_current_is_pass(self, standard_power_values, standard_current_values, tolerance):
        p_res, q_res, s_res, p_val, q_val, s_val = self.check_demand_power_is_pass(standard_power_values, tolerance)
        ia_res, ib_res, ic_res, in_res, ia_val, ib_val, ic_val, in_val = self.check_demand_current_is_pass(
            standard_current_values, tolerance)
        power_current_res = [p_res, q_res, s_res, ia_res, ib_res, ic_res, in_res]
        power_current_vals = [p_val, q_val, s_val, ia_val, ib_val, ic_val, in_val]
        if len(power_current_res) and all(power_current_res):
            logging.info("demand_power_current_is_pass Succeed")
            return True, power_current_vals
        else:
            logging.info("demand_power_current_is_pass Failed")
            return False, power_current_vals

    def check_demand_power_is_clear(self, tolerance):
        sys_active_power = self.handle_memory.read_demand_sys_active_power()
        p_res, _ = self.check_demand_is_clear(sys_active_power, tolerance)
        sys_reactive_power = self.handle_memory.read_demand_sys_reactive_power()
        q_res, _ = self.check_demand_is_clear(sys_reactive_power, tolerance)
        sys_apparent_power = self.handle_memory.read_demand_sys_apparent_power()
        s_res, _ = self.check_demand_is_clear(sys_apparent_power, tolerance)
        return p_res, q_res, s_res

    def check_demand_current_is_clear(self, tolerance):
        demand_ia = self.handle_memory.read_demand_ia()
        ia_res, _ = self.check_demand_is_clear(demand_ia, tolerance)
        demand_ib = self.handle_memory.read_demand_ib()
        ib_res, _ = self.check_demand_is_clear(demand_ib, tolerance)
        demand_ic = self.handle_memory.read_demand_ic()
        ic_res, _ = self.check_demand_is_clear(demand_ic, tolerance)
        demand_in = self.handle_memory.read_demand_in()
        in_res, _ = self.check_demand_is_clear(demand_in, tolerance)
        return ia_res, ib_res, ic_res, in_res

    def check_demand_power_current_is_clear(self, tolerance):
        p_res, q_res, s_res = self.check_demand_power_is_clear(tolerance)
        ia_res, ib_res, ic_res, in_res = self.check_demand_current_is_clear(tolerance)
        power_current_res = [p_res, q_res, s_res, ia_res, ib_res, ic_res, in_res]
        if len(power_current_res) and all(power_current_res):
            logging.info("demand_power_current_is_clear Succeed")
            return True
        else:
            logging.info("demand_power_current_is_clear Failed")
            return False

    @staticmethod
    def calc_hms_by_seconds(seconds):
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        return h % 24, m, s

    @staticmethod
    def get_wait_seconds(demand_interval):
        """
        Args:
            demand_interval: 需量窗口间隔时间min

        Returns: 当前时刻基于demand_interval等待时间，s

        """
        if not demand_interval or demand_interval > 30:
            return f"demand_interval is set to zero or more than 30"
        init_start_min = 60
        now_time = datetime.now()
        current_minute = now_time.minute
        current_seconds = now_time.second + 1 if now_time.microsecond else now_time.second
        demand_intervals = []
        for i in range(1, 60 // demand_interval + 1):
            if (i * demand_interval in range(current_minute, 60)) and (i * demand_interval > current_minute):
                demand_intervals.append(i * demand_interval)
                break
        if len(demand_intervals):
            predict_start_min = demand_intervals[0]
            if (init_start_min - predict_start_min) < demand_interval:
                exp_wait_min = init_start_min - current_minute
            else:
                exp_wait_min = predict_start_min - current_minute
        else:
            exp_wait_min = init_start_min - current_minute
        return exp_wait_min * 60 - current_seconds

    @staticmethod
    def write_common_values_to_excel(ws, index_value, common_values, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param common_values: case编号/电压/电流/电压相位角/电流相位角等输入值,精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(common_values)):
            ws.cell(index_value + 2, start_num + k, common_values[k])
        start_num += len(common_values)
        return start_num

    @staticmethod
    def write_accuracy_res_to_excel(file_path, wb, ws, index_value, accuracy_res, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param file_path: 待写入文件路径
        :param wb: 写入工作簿对象
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param accuracy_res: 精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(accuracy_res)):
            ws.cell(index_value + 2, start_num + k, accuracy_res[k])
        wb.save(file_path)
        start_num += len(accuracy_res)
        return start_num

    @staticmethod
    def line_to_line_voltage_calculate(ua, ub, uc, va_angle, vb_angle, vc_angle):
        """
        计算线电压
        :param ua:ua
        :param ub:ub
        :param uc:uc
        :param va_angle:va_angle
        :param vb_angle:vb_angle
        :param vc_angle:vc_angle
        :return: vab, vbc, vca
        """
        ret = []
        va_complex = ua * cmath.exp(1j * math.radians(va_angle))
        vb_complex = ub * cmath.exp(1j * math.radians(vb_angle))
        vc_complex = uc * cmath.exp(1j * math.radians(vc_angle))
        vab = va_complex - vb_complex
        vbc = vb_complex - vc_complex
        vca = vc_complex - va_complex
        vab = abs(vab)
        vbc = abs(vbc)
        vca = abs(vca)
        ret.extend([vab, vbc, vca])
        return ret

    @staticmethod
    def get_line_to_line_voltage(u1, u2, u1_angle, u2_angle):
        """
        计算线电压
        :param u1:ua
        :param u2:ub
        :param u1_angle:u1_angle
        :param u2_angle:u2_angle
        :return: vab, vbc, vca
        """
        ret = []
        va_complex = u1 * cmath.exp(1j * math.radians(u1_angle))
        vb_complex = u2 * cmath.exp(1j * math.radians(u2_angle))
        vab = va_complex - vb_complex
        vba = vb_complex - va_complex
        vab = abs(vab)
        vba = abs(vba)
        ret.extend([vab, vba])
        return ret

    @staticmethod
    def get_sys_p_power_by_2e3wd(uab, ia, ucb, ic, vc_angle):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :param vc_angle:vc_angle
        :return: sys_p_power_by_delta
        """
        ret = (uab * ia + ucb * ic) * math.cos(math.radians(30)) * math.cos(math.radians(vc_angle))
        return ret

    @staticmethod
    def get_sys_s_power_by_2e3wd(uab, ia, ucb, ic):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :return: sys_p_power_by_delta
        """
        ret = (uab * ia + ucb * ic) * math.cos(math.radians(30))
        return ret

    @staticmethod
    def get_sys_q_power_by_2e3wd(sys_p_power_by_2e3wd, sys_s_power_by_2e3wd):
        """
        计算线电压
        :param sys_p_power_by_2e3wd:sys_p_power_by_2e3wd
        :param sys_s_power_by_2e3wd:sys_s_power_by_2e3wd
        :return: sys_q_power_by_delta
        """
        sys_q_power_by_delta = sys_p_power_by_2e3wd ** 2 - sys_s_power_by_2e3wd ** 2
        sys_q_power_by_delta = math.sqrt(sys_q_power_by_delta) if sys_q_power_by_delta else 0
        return sys_q_power_by_delta

    @staticmethod
    def get_sys_p_power_by_3e4wd(uab, ia, ucb, ic, ubc, in_val, vc_angle):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :param ubc:ubc
        :param in_val:in_val
        :param vc_angle:vc_angle
        :return: sys_p_power_by_delta
        """
        ret = (uab * ia + ucb * ic + ubc * in_val * 0.5) * math.cos(math.radians(30)) * math.cos(math.radians(vc_angle))
        return ret

    @staticmethod
    def get_sys_s_power_by_3e4wd(uab, ia, ucb, ic, ubc, in_val):
        """
        计算线电压
        :param uab:uab
        :param ia:ub
        :param ucb:ucb
        :param ic:ic
        :param ubc:ubc
        :param in_val:in_val
        :return: sys_p_power_by_delta
        """
        ret = (uab * ia + ucb * ic + ubc * in_val * 0.5) * math.cos(math.radians(30))
        return ret

    @staticmethod
    def get_sys_q_power_by_3e4wd(sys_p_power_by_3e4wd, sys_s_power_by_3e4wd):
        """
        计算线电压
        :param sys_p_power_by_3e4wd:sys_p_power_by_3e4wd
        :param sys_s_power_by_3e4wd:sys_s_power_by_3e4wd
        :return: sys_q_power_by_delta
        """
        sys_q_power_by_delta = sys_p_power_by_3e4wd ** 2 - sys_s_power_by_3e4wd ** 2
        sys_q_power_by_delta = math.sqrt(sys_q_power_by_delta) if sys_q_power_by_delta else 0
        return sys_q_power_by_delta

    def fixed_demand_test_by_3e4wd(self, file_path, input_list):
        """
        IN−rms = IA−rms + IB−rms + IC−rms
        Psys=VAB∗IA−VBC∗IC+12VBC∗IN
        Psys=VAB∗IA+VCB∗IC+12VBC∗IN
        Ssys=(VAB−rms∗IA−rms+VBC−rms∗IC−rms+12VBC−rms∗IN−rms)∗Cos30deg
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = sum([ia, ib, ic])

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_angle, ub_angle, uc_angle)
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)
            exp_demand_p_sys_1st = self.get_sys_p_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_3e4wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )

            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = sum([ia_2nd, ib_2nd, ic_2nd])
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                # 等待 后1/2 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)

    def sliding_demand_test_by_3e4wd(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = sum([ia, ib, ic])

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(
                ua, ub, uc,
                ua_angle, ub_angle, uc_angle
            )
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)

            exp_demand_p_sys_1st = self.get_sys_p_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_3e4wd(
                uab, ia, ucb, ic, ubc, in_val
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_3e4wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = sum([ia_2nd, ib_2nd, ic_2nd])
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_3e4wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd,
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = sum([ia_3rd, ib_3rd, ic_3rd])

                uab_3rd, ubc_3rd, uca_3rd = self.line_to_line_voltage_calculate(
                    ua_3rd, ub_3rd, uc_3rd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_3rd, = self.get_line_to_line_voltage(ub_3rd, uc_3rd, ub_angle, uc_angle)

                exp_demand_p_sys_3rd_by_down_half = self.get_sys_p_power_by_3e4wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd, ubc_3rd, in_3rd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = self.get_sys_s_power_by_3e4wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd, ubc_3rd, in_3rd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.get_sys_q_power_by_3e4wd(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                    ) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                    ) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_3e4wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = self.get_sys_p_power_by_3e4wd(
                        uab, ia, ucb, ic, ubc, in_val, vc_angle
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = self.get_sys_s_power_by_3e4wd(
                        uab, ia, ucb, ic, ubc, in_val
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.get_sys_q_power_by_3e4wd(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )

                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd, vc_angle
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_3e4wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, ubc_2nd, in_2nd
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_3e4wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                exp_demand_ia_3rd = ia_3rd
                exp_demand_ib_3rd = ib_3rd
                exp_demand_ic_3rd = ic_3rd
                exp_demand_in_3rd = in_3rd

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)

    def fixed_demand_test_by_2e3wd(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(ua, ub, uc, ua_angle, ub_angle, uc_angle)
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)
            exp_demand_p_sys_1st = self.get_sys_p_power_by_2e3wd(
                uab, ia, ucb, ic, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_2e3wd(
                uab, ia, ucb, ic
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_2e3wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )

            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                # 等待 后1/2 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)

    def sliding_demand_test_by_2e3wd(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wd_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            uab, ubc, uca = self.line_to_line_voltage_calculate(
                ua, ub, uc,
                ua_angle, ub_angle, uc_angle
            )
            _, ucb = self.get_line_to_line_voltage(ub, uc, ub_angle, uc_angle)

            exp_demand_p_sys_1st = self.get_sys_p_power_by_2e3wd(
                uab, ia, ucb, ic, vc_angle
            )
            exp_demand_s_sys_1st = self.get_sys_s_power_by_2e3wd(
                uab, ia, ucb, ic
            )
            exp_demand_q_sys_1st = self.get_sys_q_power_by_2e3wd(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                uab_2nd, ubc_2nd, uca_2nd = self.line_to_line_voltage_calculate(
                    ua_2nd, ub_2nd, uc_2nd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_2nd = self.get_line_to_line_voltage(ub_2nd, uc_2nd, ub_angle, uc_angle)

                exp_demand_p_sys_2nd_by_down_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = self.get_sys_p_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = self.get_sys_s_power_by_2e3wd(
                    uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                ) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = 0

                uab_3rd, ubc_3rd, uca_3rd = self.line_to_line_voltage_calculate(
                    ua_3rd, ub_3rd, uc_3rd,
                    ua_angle, ub_angle, uc_angle
                )
                _, ucb_3rd, = self.get_line_to_line_voltage(ub_3rd, uc_3rd, ub_angle, uc_angle)

                exp_demand_p_sys_3rd_by_down_half = self.get_sys_p_power_by_2e3wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd, vc_angle
                ) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = self.get_sys_s_power_by_2e3wd(
                    uab_3rd, ia_3rd, ucb_3rd, ic_3rd
                ) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.get_sys_q_power_by_2e3wd(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                    ) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                    ) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_2e3wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = self.get_sys_p_power_by_2e3wd(
                        uab, ia, ucb, ic, vc_angle
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = self.get_sys_s_power_by_2e3wd(
                        uab, ia, ucb, ic
                    ) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.get_sys_q_power_by_2e3wd(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )

                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = self.get_sys_p_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd, vc_angle
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = self.get_sys_s_power_by_2e3wd(
                        uab_2nd, ia_2nd, ucb_2nd, ic_2nd
                    ) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.get_sys_q_power_by_2e3wd(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                exp_demand_ia_3rd = ia_3rd
                exp_demand_ib_3rd = ib_3rd
                exp_demand_ic_3rd = ic_3rd
                exp_demand_in_3rd = in_3rd

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wd = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)

    def fixed_demand_test_by_2e3wn(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wn_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                # 等待 后1/2 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wn = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wn, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)

    def sliding_demand_test_by_2e3wn(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3wn_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle),
                    self.calculate_active_power(ub_3rd, ib_3rd, vc_angle),
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd),
                    self.calculate_apparent_power(ub_3rd, ib_3rd),
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    ]) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                    ]) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                        self.calculate_active_power(ub, ib, vc_angle),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                        self.calculate_apparent_power(ub, ib),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                exp_demand_ia_3rd = ia_3rd
                exp_demand_ib_3rd = ib_3rd
                exp_demand_ic_3rd = ic_3rd
                exp_demand_in_3rd = in_3rd

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3wn = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3wn, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)

    def fixed_demand_test_by_2e3w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 180
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            ua = voltage
            ub = 0
            uc = voltage
            ia = current
            ib = 0
            ic = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(exp_demand_p_sys_1st, exp_demand_s_sys_1st)
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st, exp_demand_ib_1st, exp_demand_ic_1st, exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = voltage_2nd
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd),
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle),
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                # 等待 后1/2 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 第二个周期开始计算
                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd, exp_demand_ib_2nd, exp_demand_ic_2nd, exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)

    def sliding_demand_test_by_2e3w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_2e3w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 180
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            ua = voltage
            ub = 0
            uc = voltage
            ia = current
            ib = 0
            ic = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = voltage_2nd
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_update_rate / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd)
                ]) * demand_update_rate / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                ua_3rd = voltage_3rd
                ub_3rd = 0
                uc_3rd = voltage_3rd
                ia_3rd = current_3rd
                ib_3rd = 0
                ic_3rd = current_3rd
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle),
                    self.calculate_active_power(uc_3rd, ic_3rd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd),
                    self.calculate_apparent_power(uc_3rd, ic_3rd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * demand_update_rate / demand_interval
                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * demand_update_rate / demand_interval
                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                        self.calculate_active_power(uc, ic, vc_angle)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval
                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                        self.calculate_apparent_power(uc, ic)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval
                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * (demand_interval - demand_update_rate) / demand_interval
                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * (demand_interval - demand_update_rate) / demand_interval
                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                exp_demand_ia_3rd = ia_3rd
                exp_demand_ib_3rd = ib_3rd
                exp_demand_ic_3rd = ic_3rd
                exp_demand_in_3rd = in_3rd

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_2e3w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_2e3w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)

    def fixed_demand_test_by_1e2w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_1e2w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)
            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 0
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = 0
            ua = voltage
            ub = 0
            uc = 0
            ia = current
            ib = 0
            ic = 0
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st, exp_demand_ib_1st, exp_demand_ic_1st, exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = 0
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = 0
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd)
                ]) * demand_interval * 0.5 / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval
                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval
                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_p_sys_2nd_by_down_half
                ])
                exp_demand_q_sys_2nd = sum([
                    exp_demand_q_sys_2nd_by_upper_half,
                    exp_demand_q_sys_2nd_by_down_half
                ])
                exp_demand_s_sys_2nd = sum([
                    exp_demand_s_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_down_half
                ])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                # 等待 后1/2 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 第二个周期开始计算
                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd, exp_demand_ib_2nd, exp_demand_ic_2nd, exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_1e2w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_1e2w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)

    def sliding_demand_test_by_1e2w1p(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_1e2w1p_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            # 接线方式不通,电压、电流、压流角都不相同,需区分
            ua_angle = 0
            ub_angle = 0
            uc_angle = 0
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = 0
            ic_angle = 0
            ua = voltage
            ub = 0
            uc = 0
            ia = current
            ib = 0
            ic = 0
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [
                    exp_demand_p_sys_1st,
                    exp_demand_q_sys_1st,
                    exp_demand_s_sys_1st
                ]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st,
                    standard_current_values_1st,
                    tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                ua_2nd = voltage_2nd
                ub_2nd = 0
                uc_2nd = 0
                ia_2nd = current_2nd
                ib_2nd = 0
                ic_2nd = 0
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                ]) * demand_update_rate / demand_interval
                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd)
                ]) * demand_update_rate / demand_interval
                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                standard_power_values_2nd = [
                    exp_demand_p_sys_2nd,
                    exp_demand_q_sys_2nd,
                    exp_demand_s_sys_2nd
                ]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd,
                    standard_current_values_2nd,
                    tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                ua_3rd = voltage_3rd
                ub_3rd = 0
                uc_3rd = 0
                ia_3rd = current_3rd
                ib_3rd = 0
                ic_3rd = 0
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                    ]) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                    ]) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                exp_demand_ia_3rd = ia_3rd
                exp_demand_ib_3rd = ib_3rd
                exp_demand_ic_3rd = ic_3rd
                exp_demand_in_3rd = in_3rd

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_1e2w1p = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_1e2w1p, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)

    def fixed_demand_test_by_3e4wy(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )

                # 第二个周期开始计算
                # 等待前0.5 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 查看需量是否正确，voltage*5/8用于计算功率需量，current_value * 1/2用于计算电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd)
                ]) * demand_interval * 0.5 / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_interval * 0.5) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )

                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])

                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                # 等待 后1/2 * demand_interval后
                time.sleep(0.5 * demand_interval * 60)
                # 第二个周期开始计算
                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wy = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wy, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)

    def sliding_demand_test_by_3e4wy(self, file_path, input_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws)
        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            # 从输入excel表格中获取电压、电流、角度、有功要求的精度, 例如0.001
            demand_accuracy = self.get_test_case_info_of_accuracy(input_list, i)
            # 从输入excel表格中获取电压、电流、角度， 抽样次数、抽样间隔
            (case_id, voltage, vc_angle, current, freq, demand_method, demand_interval, demand_update_rate,
             demand_trigger, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i)

            ua_angle = 0
            ub_angle = 240
            uc_angle = 120
            ia_angle = (ua_angle - vc_angle) if (ua_angle - vc_angle) else (ua_angle - vc_angle + 360)
            ib_angle = (ub_angle - vc_angle) if (ub_angle - vc_angle) else (ub_angle - vc_angle + 360)
            ic_angle = (uc_angle - vc_angle) if (uc_angle - vc_angle) else (uc_angle - vc_angle + 360)
            uc = voltage
            ub = voltage
            ua = voltage
            ic = current
            ib = current
            ia = current
            in_val = 0

            # 初始需量值
            exp_demand_p_sys_1st = sum([
                self.calculate_active_power(ua, ia, vc_angle),
                self.calculate_active_power(ub, ib, vc_angle),
                self.calculate_active_power(uc, ic, vc_angle)
            ])
            exp_demand_s_sys_1st = sum([
                self.calculate_apparent_power(ua, ia),
                self.calculate_apparent_power(ub, ib),
                self.calculate_apparent_power(uc, ic)
            ])
            exp_demand_q_sys_1st = self.calculate_reactive_power(
                exp_demand_p_sys_1st,
                exp_demand_s_sys_1st
            )
            exp_demand_p_sys_1st = round(exp_demand_p_sys_1st, 5)
            exp_demand_s_sys_1st = round(exp_demand_s_sys_1st, 5)
            exp_demand_q_sys_1st = round(exp_demand_q_sys_1st, 5)
            exp_demand_ia_1st = ia
            exp_demand_ib_1st = ib
            exp_demand_ic_1st = ic
            exp_demand_in_1st = in_val
            # 初始需量测量需量设置
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(uc, ub, ua)
            set_current_gear(ic, ib, ia)
            set_ac(
                quc=uc_angle,
                qub=ub_angle,
                qua=ua_angle,
                qic=ic_angle,
                qib=ib_angle,
                qia=ia_angle,
                uc=uc,
                ub=ub,
                ua=ua,
                ic=ic,
                ib=ib,
                ia=ia,
                f=freq
            )
            # 设置需量参数
            self.set_demand_para(demand_method, demand_interval, demand_update_rate)
            time.sleep(300)
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            # 触发需量重置
            if demand_trigger == 0:
                self.set_demand_para(demand_method, demand_interval, demand_update_rate)  # 设置需量参数
            elif demand_trigger == 1:  # 设置时间重置需量
                self.set_time_trigger(sys_millisecond=0)
            else:
                self.set_demand_trigger(clear_max_demand=1)  # 设置重置需量最大值来重置需量
            # 检查需量值是否为0
            self.check_demand_power_current_is_clear(tolerance=demand_accuracy)  # "设置需量参数后，需量没有置为0"
            # 计算基于当前时间最近的demand_interval整数倍的时刻，需等待时间(s)
            wait_time = self.get_wait_seconds(demand_interval)
            logging.info(f"等待开始时间,{wait_time}")
            # 等待需量开始时间
            time.sleep(wait_time)  # 距离最近的demand_interval整数倍时刻期间的时间，不计算需量
            # 需量第一次上报时间
            time.sleep(demand_interval * 60)
            if voltage < 9.5 or current == 0:
                self.check_demand_power_current_is_clear(tolerance=demand_accuracy)
            else:
                # 第一个周期开始计算
                standard_power_values_1st = [exp_demand_p_sys_1st, exp_demand_q_sys_1st, exp_demand_s_sys_1st]
                standard_current_values_1st = [
                    exp_demand_ia_1st,
                    exp_demand_ib_1st,
                    exp_demand_ic_1st,
                    exp_demand_in_1st
                ]
                check_res_1st, measure_vals_1st = self.check_demand_power_current_is_pass(
                    standard_power_values_1st, standard_current_values_1st, tolerance=demand_accuracy
                )
                # 第二个周期开始计算
                # 降低电压到voltage * 0.5, 电流到current_value * 0.5
                voltage_2nd = voltage * 0.5
                current_2nd = current * 0.5
                uc_2nd = voltage_2nd
                ub_2nd = voltage_2nd
                ua_2nd = voltage_2nd
                ic_2nd = current_2nd
                ib_2nd = current_2nd
                ia_2nd = current_2nd
                in_2nd = 0
                # 控源操作
                set_voltage_gear(uc_2nd, ub_2nd, ua_2nd)
                set_current_gear(ic_2nd, ib_2nd, ia_2nd)
                set_ac(
                    quc=uc_angle,
                    qub=ub_angle,
                    qua=ua_angle,
                    qic=ic_angle,
                    qib=ib_angle,
                    qia=ia_angle,
                    uc=uc_2nd,
                    ub=ub_2nd,
                    ua=ua_2nd,
                    ic=ic_2nd,
                    ib=ib_2nd,
                    ia=ia_2nd,
                    f=freq
                )
                # 第二个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                exp_demand_p_sys_2nd_by_down_half = sum([
                    self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                    self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                    self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_2nd_by_down_half = sum([
                    self.calculate_apparent_power(ua_2nd, ia_2nd),
                    self.calculate_apparent_power(ub_2nd, ib_2nd),
                    self.calculate_apparent_power(uc_2nd, ic_2nd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_2nd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_down_half,
                    exp_demand_s_sys_2nd_by_down_half
                )

                exp_demand_p_sys_2nd_by_upper_half = sum([
                    self.calculate_active_power(ua, ia, vc_angle),
                    self.calculate_active_power(ub, ib, vc_angle),
                    self.calculate_active_power(uc, ic, vc_angle)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_s_sys_2nd_by_upper_half = sum([
                    self.calculate_apparent_power(ua, ia),
                    self.calculate_apparent_power(ub, ib),
                    self.calculate_apparent_power(uc, ic)
                ]) * (demand_interval - demand_update_rate) / demand_interval

                exp_demand_q_sys_2nd_by_upper_half = self.calculate_reactive_power(
                    exp_demand_p_sys_2nd_by_upper_half,
                    exp_demand_s_sys_2nd_by_upper_half
                )
                exp_demand_p_sys_2nd = sum([exp_demand_p_sys_2nd_by_upper_half, exp_demand_p_sys_2nd_by_down_half])
                exp_demand_q_sys_2nd = sum([exp_demand_q_sys_2nd_by_upper_half, exp_demand_q_sys_2nd_by_down_half])
                exp_demand_s_sys_2nd = sum([exp_demand_s_sys_2nd_by_upper_half, exp_demand_s_sys_2nd_by_down_half])
                exp_demand_p_sys_2nd = round(exp_demand_p_sys_2nd, 5)
                exp_demand_q_sys_2nd = round(exp_demand_q_sys_2nd, 5)
                exp_demand_s_sys_2nd = round(exp_demand_s_sys_2nd, 5)
                exp_demand_ia_2nd = ia_2nd
                exp_demand_ib_2nd = ib_2nd
                exp_demand_ic_2nd = ic_2nd
                exp_demand_in_2nd = in_2nd

                standard_power_values_2nd = [exp_demand_p_sys_2nd, exp_demand_q_sys_2nd, exp_demand_s_sys_2nd]
                standard_current_values_2nd = [
                    exp_demand_ia_2nd,
                    exp_demand_ib_2nd,
                    exp_demand_ic_2nd,
                    exp_demand_in_2nd
                ]
                check_res_2nd, measure_vals_2nd = self.check_demand_power_current_is_pass(
                    standard_power_values_2nd, standard_current_values_2nd, tolerance=demand_accuracy
                )

                # 第三个周期开始计算
                # 等待demand_update_rate后
                time.sleep(demand_update_rate * 60)
                # 查看需量是否正确，计算电压、电流需量
                voltage_3rd = voltage_2nd
                current_3rd = current_2nd
                uc_3rd = voltage_3rd
                ub_3rd = voltage_3rd
                ua_3rd = voltage_3rd
                ic_3rd = current_3rd
                ib_3rd = current_3rd
                ia_3rd = current_3rd
                in_3rd = 0
                exp_demand_p_sys_3rd_by_down_half = sum([
                    self.calculate_active_power(ua_3rd, ia_3rd, vc_angle),
                    self.calculate_active_power(ub_3rd, ib_3rd, vc_angle),
                    self.calculate_active_power(uc_3rd, ic_3rd, vc_angle)
                ]) * demand_update_rate / demand_interval

                exp_demand_s_sys_3rd_by_down_half = sum([
                    self.calculate_apparent_power(ua_3rd, ia_3rd),
                    self.calculate_apparent_power(ub_3rd, ib_3rd),
                    self.calculate_apparent_power(uc_3rd, ic_3rd)
                ]) * demand_update_rate / demand_interval

                exp_demand_q_sys_3rd_by_down_half = self.calculate_reactive_power(
                    exp_demand_p_sys_3rd_by_down_half,
                    exp_demand_s_sys_3rd_by_down_half
                )

                if demand_interval - demand_update_rate >= demand_update_rate:
                    # 场景1: demand_interval - demand_update_rate > demand_update_rate
                    # 场景2: demand_interval - demand_update_rate == demand_update_rate
                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * demand_update_rate / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * demand_update_rate / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                    exp_demand_p_sys_3rd_by_upper_half = sum([
                        self.calculate_active_power(ua, ia, vc_angle),
                        self.calculate_active_power(ub, ib, vc_angle),
                        self.calculate_active_power(uc, ic, vc_angle)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_upper_half = sum([
                        self.calculate_apparent_power(ua, ia),
                        self.calculate_apparent_power(ub, ib),
                        self.calculate_apparent_power(uc, ic)
                    ]) * (demand_interval - 2 * demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_upper_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_upper_half,
                        exp_demand_s_sys_3rd_by_upper_half
                    )
                else:
                    # 场景1: demand_interval - demand_update_rate < demand_update_rate
                    # 场景2: demand_interval == demand_update_rate
                    exp_demand_p_sys_3rd_by_upper_half = 0
                    exp_demand_s_sys_3rd_by_upper_half = 0
                    exp_demand_q_sys_3rd_by_upper_half = 0

                    exp_demand_p_sys_3rd_by_middle_half = sum([
                        self.calculate_active_power(ua_2nd, ia_2nd, vc_angle),
                        self.calculate_active_power(ub_2nd, ib_2nd, vc_angle),
                        self.calculate_active_power(uc_2nd, ic_2nd, vc_angle)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_s_sys_3rd_by_middle_half = sum([
                        self.calculate_apparent_power(ua_2nd, ia_2nd),
                        self.calculate_apparent_power(ub_2nd, ib_2nd),
                        self.calculate_apparent_power(uc_2nd, ic_2nd)
                    ]) * (demand_interval - demand_update_rate) / demand_interval

                    exp_demand_q_sys_3rd_by_middle_half = self.calculate_reactive_power(
                        exp_demand_p_sys_3rd_by_middle_half,
                        exp_demand_s_sys_3rd_by_middle_half
                    )

                exp_demand_p_sys_3rd = sum([
                    exp_demand_p_sys_3rd_by_upper_half,
                    exp_demand_p_sys_3rd_by_middle_half,
                    exp_demand_p_sys_3rd_by_down_half
                ])
                exp_demand_q_sys_3rd = sum([
                    exp_demand_q_sys_3rd_by_upper_half,
                    exp_demand_q_sys_3rd_by_middle_half,
                    exp_demand_q_sys_3rd_by_down_half
                ])
                exp_demand_s_sys_3rd = sum([
                    exp_demand_s_sys_3rd_by_upper_half,
                    exp_demand_s_sys_3rd_by_middle_half,
                    exp_demand_s_sys_3rd_by_down_half
                ])
                exp_demand_p_sys_3rd = round(exp_demand_p_sys_3rd, 5)
                exp_demand_q_sys_3rd = round(exp_demand_q_sys_3rd, 5)
                exp_demand_s_sys_3rd = round(exp_demand_s_sys_3rd, 5)
                exp_demand_ia_3rd = ia_3rd
                exp_demand_ib_3rd = ib_3rd
                exp_demand_ic_3rd = ic_3rd
                exp_demand_in_3rd = in_3rd

                standard_power_values_3rd = [
                    exp_demand_p_sys_3rd,
                    exp_demand_q_sys_3rd,
                    exp_demand_s_sys_3rd
                ]
                standard_current_values_3rd = [
                    exp_demand_ia_3rd,
                    exp_demand_ib_3rd,
                    exp_demand_ic_3rd,
                    exp_demand_in_3rd
                ]
                check_res_3rd, measure_vals_3rd = self.check_demand_power_current_is_pass(
                    standard_power_values_3rd,
                    standard_current_values_3rd,
                    tolerance=demand_accuracy
                )

                # 处理数据,保持和表列相同
                standard_power_current_values_1st = list(
                    itertools.chain(standard_power_values_1st, standard_current_values_1st)
                )
                standard_power_current_values_2nd = list(
                    itertools.chain(standard_power_values_2nd, standard_current_values_2nd)
                )
                standard_power_current_values_3rd = list(
                    itertools.chain(standard_power_values_3rd, standard_current_values_3rd)
                )
                std_measure_vals_1st = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_1st, measure_vals_1st))
                )
                std_measure_vals_2nd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_2nd, measure_vals_2nd))
                )
                std_measure_vals_3rd = list(
                    itertools.chain.from_iterable(zip(standard_power_current_values_3rd, measure_vals_3rd))
                )
                # 写入数据
                start_num = 1  # 输出excel表格的列编号
                common_values_of_3e4wy = (
                    case_id, demand_accuracy,
                    voltage, vc_angle, current, freq,
                    demand_method, demand_interval, demand_update_rate, demand_trigger
                )
                # 从第二行开始，逐行抄写输入的测试数据到结果excel的前半列，后半列用于存储测试到的数据，返回下次写的列标
                start_num = self.write_common_values_to_excel(ws, i, common_values_of_3e4wy, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_1st, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_1st], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_2nd, start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_2nd], start_num)
                start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, std_measure_vals_3rd, start_num)
                self.write_accuracy_res_to_excel(file_path, wb, ws, i, [check_res_3rd], start_num)


def run_demand_test_script(test_type, wire_type, demand_type):
    """
    运行脚本入口
    :param test_type: 0:mV, 1:mA, 2:rct
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 5:3e3wD  #AcuRev1320
    :param demand_type: 0:fixed, 1:slid
    :return:
    """
    print(f"====================Demand Test Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    switch_device_screen_interface(inter=0x01)  # 切换至交流界面
    set_gear_switching_mode('00000000')  # 档位切换归零,自动
    demand_test = DemandTest()

    demand_test.select_test_case(test_type=test_type, wire_type=wire_type, demand_type=demand_type)

    # 关闭ModbusClient客户端连接
    demand_test.handle_memory.modbus_client.close()
    up_source_ac()
    switch_device_screen_interface(inter=0x00)  # 切换至默认界面
    set_gear_switching_mode(mode='01000000')  # 档位切换归零,手动
    print(f"====================测试总耗时:{time.time() - start_time}====================")
    print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
    print(f"====================Demand Test End====================")


if __name__ == '__main__':
    """
    :param measure_mode: 0:mV, 1:mA, 2:rct
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 5:3e3wD  # AcuRev1320 项目
    :param demand_type: 0:fixed, 1:sliding
    """
    measure_mode = 2
    wire_mode = 6
    demand_mode = 0
    run_demand_test_script(test_type=measure_mode, wire_type=wire_mode, demand_type=demand_mode)
