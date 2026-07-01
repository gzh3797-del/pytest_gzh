#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
"""从知识库官方 Modbus 地址表读取 AcuRev1320 需量寄存器地址(方案B)。

不再硬编码需量数值寄存器地址(原 memory_addrs.py 中的 0xC4xx 为 IIV3 残留, 在
AcuRev1320 上读不到), 改为运行时从知识库权威地址表的 Demand sheet 解析:

    knowledge/shared/modbus_tables/raw/AcuRev-1320 Modbus Address Table v1.01 20260330.xlsx

返回的键与 HandleMemory.read_demand_* 使用的键一致, 可直接覆盖 MemoryAddr.demand_addr。
"""
import os
import re

import openpyxl

from tools.log import root_path

# 知识库官方地址表(权威来源)
DEMAND_TABLE_PATH = os.path.join(
    root_path, "knowledge", "shared", "modbus_tables", "raw",
    "AcuRev-1320 Modbus Address Table v1.01 20260330.xlsx",
)
DEMAND_SHEET = "Demand"

# 地址表内的参数描述 -> 代码使用的键(与 acuvimseries_modbus_get.read_demand_* 对应)
_DESC_TO_KEY = {
    "System Active Power Demand": "system_active_power",
    "System Reactive Power Demand": "system_reactive_power",
    "System Apparent Power Demand": "system_apparent_power",
    "Phase A Current Demand": "phase_a_current",
    "Phase B Current Demand": "phase_b_current",
    "Phase C Current Demand": "phase_c_current",
    "Neutral Current Demand": "phase_n_current",
}
_HEX_RE = re.compile(r"^0x[0-9A-Fa-f]+$")

# 解析结果缓存(地址表不变, 进程内只读一次)
_cache = None


def load_demand_addr(table_path=DEMAND_TABLE_PATH, sheet_name=DEMAND_SHEET):
    """解析地址表 Demand sheet, 返回 {key: int 地址}。结果缓存。

    :raises FileNotFoundError: 地址表文件不存在
    :raises ValueError: 地址表缺少某个需量寄存器或 sheet 不存在
    """
    global _cache
    if _cache is not None:
        return _cache

    if not os.path.isfile(table_path):
        raise FileNotFoundError(f"AcuRev1320 Modbus 地址表不存在: {table_path}")

    workbook = openpyxl.load_workbook(table_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"地址表无 '{sheet_name}' sheet: {table_path}")
        sheet = workbook[sheet_name]
        result = {}
        for row in sheet.iter_rows(values_only=True):
            descriptions = [str(cell).strip() for cell in row if isinstance(cell, str)]
            key = next((_DESC_TO_KEY[d] for d in descriptions if d in _DESC_TO_KEY), None)
            if key is None or key in result:
                continue
            # 取该行第一个形如 0xXXXX 的单元格作为起始地址
            addr_hex = next(
                (cell.strip() for cell in row
                 if isinstance(cell, str) and _HEX_RE.match(cell.strip())),
                None,
            )
            if addr_hex is not None:
                result[key] = int(addr_hex, 16)
    finally:
        workbook.close()

    missing = set(_DESC_TO_KEY.values()) - set(result)
    if missing:
        raise ValueError(f"地址表 '{sheet_name}' 缺少需量寄存器: {sorted(missing)}")

    _cache = result
    return result


if __name__ == '__main__':
    for name, addr in load_demand_addr().items():
        print(f"{name:24s} = 0x{addr:04X}")
