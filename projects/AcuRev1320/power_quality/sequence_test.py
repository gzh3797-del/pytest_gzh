import cmath
import math
import statistics

import openpyxl
from openpyxl import Workbook
from datetime import time
from comm.source_control import *
from pathlib import Path

from tools.excel_operate import data_read
from projects.AcuRev1320.power_quality.sequence_modbus_get import HandleMemory
from projects.AcuRev1320.power_quality.sequence_table_heading import TableTitle

Log(str(__file__).split("\\")[-1])
CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
test_case_path = r'./test_case/AcuRev1320/power_quality/sequence_test_case.xlsx'

sheet_name_3E4WY = "Sheet4"
sheet_name_2E3W1P = "Sheet2"

save_filedir = os.path.join(Path(__file__).parent, f"precision_measure_{time.strftime('%Y%m%d')}")
if not os.path.exists(save_filedir):
    os.makedirs(save_filedir, exist_ok=True)

ALL_SAVE_DIRS = ["1E2W1P", "2E3W1P", "2E3WD", "2E3WN", "3E4WY", "3E4WD"]


class SequenceTest:
    def __init__(self):
        self.handle_memory = HandleMemory(slave_id=1)

    def set_wire_type(self, voltage_wire_value):
        """
        设置接线方式
        :param voltage_wire_value: 电压接线方式
        # :param current_wire_value: 电流接线方式
        :return:
        """
        self.handle_memory.set_wire_mode_by_voltage(voltage_wire_mode=voltage_wire_value)
        # self.handle_memory.set_wire_mode_by_current(current_wire_mode=current_wire_value)

    def set_phase_order(self, phase_order):
        """
        设置接线方式
        :param phase_order: 电压接线方式
        # :param current_wire_value: 电流接线方式
        :return:
        """
        self.handle_memory.set_phase_order_by_voltage(phase_order=phase_order)

    def select_test_case(self, test_type, wire_type):
        """
        选择测试标准，接线方式
        # :param test_type: 0:mV, 1:mA, 2:rct
        # :param wire_type:0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,

        :param test_type: 0:mV, 1:mA, 2:rct
        :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 4:3e3wD  #AcuRev1320
        :return:
        """
        if wire_type == 1:
            sheet_name = sheet_name_2E3W1P
        else:
            sheet_name = sheet_name_3E4WY
        self.select_wire_type(test_case_path, sheet_name, wire_type, test_type)

    @staticmethod
    def get_input_list_by_wire_type(source_input_list, test_type):
        """
        获取接线方式的数据
        :param source_input_list: 测试用例数据
        :param test_type: 接线类型
        :return: 接线方式的数据
        “”“ 函数基于IIV3代码， 修改了接线方式对应关系”“”
        """
        input_list = [source_input_list[0]]
        # AcuRev1320中 接线方式3E4wY对应值是6
        if test_type == 0:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][11] == "VUF":
                    input_list.append(source_input_list[i])
        elif test_type == 1:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][11] == "CUF":
                    input_list.append(source_input_list[i])
        elif test_type == 2:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][11] == "相电压序分量":
                    input_list.append(source_input_list[i])
        else:
            for i in range(1, len(source_input_list)):
                if source_input_list[i][11] == "用户序分量":
                    input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_real_time_parameters_by_input_list_of_wire_type(source_input_list):
        """
        获取接线方式的快速测量实时数据
        :param source_input_list: 接线方式的数据
        :return: 接线方式的实时数据
        """
        input_list = [source_input_list[0]]
        for i in range(1, len(source_input_list)):
            input_list.append(source_input_list[i])
        return input_list

    @staticmethod
    def get_save_filepath_of_3e4wy(filedir):
        """
        接线方式:3e3wd
        :param filedir: 待写入文件目录
        :return: 待写入文件路径
        """
        filename = f"sequence_component_precision_measure_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(filedir, filename)
        wb = Workbook()
        wb.save(file_path)
        wb.close()
        return file_path

    @staticmethod
    def write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type):
        """
        写入表头:3e4wy
        :param file_path: 文件路径
        :param wb: 工作簿对象
        :param ws: 工作sheet对象
        :param test_type:
        :return:
        """
        if test_type == 0:
            for i in range(len(TableTitle.SEQUENCE_COLUMNS_OF_3E4WY_VUF)):
                j = i + 1
                ws.cell(1, j, f'{TableTitle.SEQUENCE_COLUMNS_OF_3E4WY_VUF[i]}')
        elif test_type == 1:
            for i in range(len(TableTitle.SEQUENCE_COLUMNS_OF_3E4WY_CUF)):
                j = i + 1
                ws.cell(1, j, f'{TableTitle.SEQUENCE_COLUMNS_OF_3E4WY_CUF[i]}')
        else:
            for i in range(len(TableTitle.SEQUENCE_COLUMNS_OF_3E4WY)):
                j = i + 1
                ws.cell(1, j, f'{TableTitle.SEQUENCE_COLUMNS_OF_3E4WY[i]}')
        wb.save(file_path)

    @staticmethod
    def get_test_case_info_of_accuracy_wire_type(input_list, index_value):
        """
        获取测试精度信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :return: accuracy_of_demand, wire_type_number, wire_order_number
        """
        accuracy_of_demand = input_list[index_value + 1][10]
        wire_type = input_list[index_value + 1][12]
        phase_order = input_list[index_value + 1][13]
        if wire_type == '3e4wY':
            wire_type_number = 4
        elif wire_type == '2e3wN':
            wire_type_number = 3
        elif wire_type == '2e3w1p':
            wire_type_number = 1
        else:
            wire_type_number = 4

        if phase_order == 'ACB':
            wire_order_number = 1
        else:
            wire_order_number = 0
        return accuracy_of_demand, wire_type_number, wire_order_number

    @staticmethod
    def get_test_case_info_of_input_value(input_list, index_value, wire_order_number):
        """
        获取测试电压/电流、相位等信息
        :param input_list: 接线方式输入数据
        :param index_value: 接线方式数据行索引
        :param wire_order_number：wire_order：1 ACB、0 ABC
        :return: 电压/电流、相位等信息
        """
        case_id = input_list[index_value + 1][0]
        a_amplitude = input_list[index_value + 1][1]
        b_amplitude = input_list[index_value + 1][2]
        c_amplitude = input_list[index_value + 1][3]
        a_angle = input_list[index_value + 1][4]
        b_angle = input_list[index_value + 1][5]
        c_angle = input_list[index_value + 1][6]
        freq = input_list[index_value + 1][7]
        sample_cnt = input_list[index_value + 1][8]
        sample_interval = input_list[index_value + 1][9]
        # if wire_order_number == 1:
        #     return (case_id, a_amplitude, c_amplitude, b_amplitude, a_angle, c_angle, b_angle, freq, sample_cnt,
        #             sample_interval)
        # else:
        return (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
                sample_interval)

    @staticmethod
    def write_common_values_to_excel(ws, index_value, common_values, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param common_values: case编号/电压/电流/电压相位角/电流相位角等输入值,精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(common_values)):
            ws.cell(index_value + 2, start_num + k, common_values[k])
        start_num += len(common_values)
        return start_num

    @staticmethod
    def calculate_angle(sequence_component_complex, sequence_component):
        if sequence_component > 0.00001:
            phase = cmath.phase(sequence_component_complex)
            angle = math.degrees(phase)

            if abs(angle) < 0.000001:
                angle = 0

            # 如果角度为负数，调整为 0 到 360 度之间
            if angle < 0:
                angle += 360

            return angle
        else:
            return 0

    @staticmethod
    def get_accuracy_res_by_sequence_rea_imag(measure_values):
        """
        获取精度比较结果，带期望精度值
        :param measure_values: 测量值
        :return: 测量值
        """
        min_val = min(measure_values)
        max_val = max(measure_values)
        avg_val = statistics.mean(measure_values)
        accuracy_res = (min_val, max_val, avg_val)
        return accuracy_res

    def sequence_component_calculation(self, a: float, b: float, c: float, a_angle: float, b_angle: float,
                                       c_angle: float):
        va_complex = complex(math.cos(a_angle * math.pi / 180) * a, math.sin(a_angle * math.pi / 180) * a)
        vb_complex = complex(math.cos(b_angle * math.pi / 180) * b, math.sin(b_angle * math.pi / 180) * b)
        vc_complex = complex(math.cos(c_angle * math.pi / 180) * c, math.sin(c_angle * math.pi / 180) * c)
        rotation_factor = complex(-1 / 2, math.sqrt(3) / 2)
        rotation_factor_square = rotation_factor ** 2
        zero_sequence_component_complex = (va_complex + vb_complex + vc_complex) / 3
        # # 实部
        # z_real = zero_sequence_component_complex.real
        # # 虚部
        # z_imag = zero_sequence_component_complex.imag
        # print(z_real,z_imag)
        zero_sequence_component = abs(zero_sequence_component_complex)
        zero_seq_calculate_angle = round(self.calculate_angle(zero_sequence_component_complex, zero_sequence_component),
                                         3)

        positive_sequence_component_complex = (va_complex / 3) + ((vb_complex * rotation_factor) / 3) + (
                (rotation_factor_square * vc_complex) / 3)
        # # 实部
        # p_real = positive_sequence_component_complex.real
        # # 虚部
        # p_imag = positive_sequence_component_complex.imag
        # print(positive_sequence_component_complex)
        # print(p_real,p_imag)
        positive_sequence_component = abs(positive_sequence_component_complex)
        positive_seq_calculate_angle = round(
            self.calculate_angle(positive_sequence_component_complex, positive_sequence_component), 3)

        negative_sequence_component_complex = (va_complex / 3) + ((vb_complex * rotation_factor_square) / 3) + (
                (rotation_factor * vc_complex) / 3)
        negative_sequence_component = abs(negative_sequence_component_complex)
        negative_seq_calculate_angle = round(
            self.calculate_angle(negative_sequence_component_complex, negative_sequence_component), 3)
        try:
            vuf_cuf = (round((negative_sequence_component / positive_sequence_component), 10)) * 100
        except:
            vuf_cuf = 0
        if vuf_cuf > 150:
            vuf_cuf = 150

        return (round(zero_sequence_component, 3), zero_seq_calculate_angle,
                round(positive_sequence_component, 3), positive_seq_calculate_angle,
                round(negative_sequence_component, 3), negative_seq_calculate_angle, vuf_cuf)



    def sequence_2e3w1p(self, a_amplitude, b_amplitude):
        """
        计算 a_amplitude 和 b_amplitude 相对于均值的最大相对偏差
        等价于 Excel:
        =MAX(ABS(A4-(A4+D4)/2)/((A4+D4)/2), ABS(D4-(A4+D4)/2)/((A4+D4)/2))
        @param a_amplitude: a_幅值
        @param b_amplitude: b_幅值
        @return:
        """
        if a_amplitude + b_amplitude == 0:
            return 0.0  # 或者 raise ValueError("a_amplitude + b_amplitude 不能为 0")

        return abs(a_amplitude - b_amplitude) / (a_amplitude + b_amplitude)

    def get_sequence_component_of_test_type(self, test_type, sequence_component_values):
        if test_type == 0:
            sequence_component_values = [sequence_component_values[6]]
            return sequence_component_values
        elif test_type == 1:
            sequence_component_values = [sequence_component_values[6]]
            return sequence_component_values
        else:
            return list(sequence_component_values)

    def get_measure_values_of_voltage_sequence(self):
        """
        获取相电压测量值
        :return:
        """
        vuf = self.handle_memory.read_voltage_unbalance_negative()
        (z_real, z_imag, z_magnitude, z_angle) = self.handle_memory.read_voltage_zero_sequence()
        (p_real, p_imag, p_magnitude, p_angle) = self.handle_memory.read_voltage_positive_sequence()
        (n_real, n_imag, n_magnitude, n_angle) = self.handle_memory.read_voltage_negative_sequence()
        return (vuf,
                (z_real, z_imag, z_magnitude, z_angle),
                (p_real, p_imag, p_magnitude, p_angle),
                (n_real, n_imag, n_magnitude, n_angle))

    def get_measure_values_of_current_sequence(self):
        """
        获取相电压测量值
        :return:
        """
        cuf = self.handle_memory.read_current_unbalance_negative()
        (z_real, z_imag, z_magnitude, z_angle) = self.handle_memory.read_current_zero_sequence()
        (p_real, p_imag, p_magnitude, p_angle) = self.handle_memory.read_current_positive_sequence()
        (n_real, n_imag, n_magnitude, n_angle) = self.handle_memory.read_current_negative_sequence()
        return (cuf,
                (z_real, z_imag, z_magnitude, z_angle),
                (p_real, p_imag, p_magnitude, p_angle),
                (n_real, n_imag, n_magnitude, n_angle))

    def get_measure_values_by_sample_cnt_voltage_sequence(self, sample_cnt, sample_interval):
        """
        获取测量值:
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_vuf = []
        measure_values_of_voltage_zero_sequence_real = []
        measure_values_of_voltage_zero_sequence_imaginary = []
        measure_values_of_voltage_zero_sequence_magnitude = []
        measure_values_of_voltage_zero_sequence_angle = []

        measure_values_of_voltage_positive_sequence_real = []
        measure_values_of_voltage_positive_sequence_imaginary = []
        measure_values_of_voltage_positive_sequence_magnitude = []
        measure_values_of_voltage_positive_sequence_angle = []

        measure_values_of_voltage_negative_sequence_real = []
        measure_values_of_voltage_negative_sequence_imaginary = []
        measure_values_of_voltage_negative_sequence_magnitude = []
        measure_values_of_voltage_negative_sequence_angle = []

        for j in range(sample_cnt):
            time.sleep(sample_interval)
            (vuf,
             (z_real, z_imag, z_magnitude, z_angle),
             (p_real, p_imag, p_magnitude, p_angle),
             (n_real, n_imag, n_magnitude, n_angle)) = self.get_measure_values_of_voltage_sequence()
            measure_values_of_vuf.append(vuf)
            measure_values_of_voltage_zero_sequence_real.append(z_real)
            measure_values_of_voltage_zero_sequence_imaginary.append(z_imag)
            measure_values_of_voltage_zero_sequence_magnitude.append(z_magnitude)
            measure_values_of_voltage_zero_sequence_angle.append(z_angle)

            measure_values_of_voltage_positive_sequence_real.append(p_real)
            measure_values_of_voltage_positive_sequence_imaginary.append(p_imag)
            measure_values_of_voltage_positive_sequence_magnitude.append(p_magnitude)
            measure_values_of_voltage_positive_sequence_angle.append(p_angle)

            measure_values_of_voltage_negative_sequence_real.append(n_real)
            measure_values_of_voltage_negative_sequence_imaginary.append(n_imag)
            measure_values_of_voltage_negative_sequence_magnitude.append(n_magnitude)
            measure_values_of_voltage_negative_sequence_angle.append(n_angle)
        return (measure_values_of_vuf,
                measure_values_of_voltage_zero_sequence_real,
                measure_values_of_voltage_zero_sequence_imaginary,
                measure_values_of_voltage_zero_sequence_magnitude,
                measure_values_of_voltage_zero_sequence_angle,
                measure_values_of_voltage_positive_sequence_real,
                measure_values_of_voltage_positive_sequence_imaginary,
                measure_values_of_voltage_positive_sequence_magnitude,
                measure_values_of_voltage_positive_sequence_angle,
                measure_values_of_voltage_negative_sequence_real,
                measure_values_of_voltage_negative_sequence_imaginary,
                measure_values_of_voltage_negative_sequence_magnitude,
                measure_values_of_voltage_negative_sequence_angle)

    def get_measure_values_by_sample_cnt_current_sequence(self, sample_cnt, sample_interval):
        """
        获取测量值:
        :param sample_cnt: 采样次数
        :param sample_interval: 采样时间间隔
        :return: 测量值
        """
        measure_values_of_cuf = []
        measure_values_of_current_zero_sequence_real = []
        measure_values_of_current_zero_sequence_imaginary = []
        measure_values_of_current_zero_sequence_magnitude = []
        measure_values_of_current_zero_sequence_angle = []

        measure_values_of_current_positive_sequence_real = []
        measure_values_of_current_positive_sequence_imaginary = []
        measure_values_of_current_positive_sequence_magnitude = []
        measure_values_of_current_positive_sequence_angle = []

        measure_values_of_current_negative_sequence_real = []
        measure_values_of_current_negative_sequence_imaginary = []
        measure_values_of_current_negative_sequence_magnitude = []
        measure_values_of_current_negative_sequence_angle = []

        for j in range(sample_cnt):
            time.sleep(sample_interval)
            (cuf,
             (z_real, z_imag, z_magnitude, z_angle),
             (p_real, p_imag, p_magnitude, p_angle),
             (n_real, n_imag, n_magnitude, n_angle)) = self.get_measure_values_of_current_sequence()
            measure_values_of_cuf.append(cuf)
            measure_values_of_current_zero_sequence_real.append(z_real)
            measure_values_of_current_zero_sequence_imaginary.append(z_imag)
            measure_values_of_current_zero_sequence_magnitude.append(z_magnitude)
            measure_values_of_current_zero_sequence_angle.append(z_angle)

            measure_values_of_current_positive_sequence_real.append(p_real)
            measure_values_of_current_positive_sequence_imaginary.append(p_imag)
            measure_values_of_current_positive_sequence_magnitude.append(p_magnitude)
            measure_values_of_current_positive_sequence_angle.append(p_angle)

            measure_values_of_current_negative_sequence_real.append(n_real)
            measure_values_of_current_negative_sequence_imaginary.append(n_imag)
            measure_values_of_current_negative_sequence_magnitude.append(n_magnitude)
            measure_values_of_current_negative_sequence_angle.append(n_angle)
        return (measure_values_of_cuf,
                measure_values_of_current_zero_sequence_real,
                measure_values_of_current_zero_sequence_imaginary,
                measure_values_of_current_zero_sequence_magnitude,
                measure_values_of_current_zero_sequence_angle,
                measure_values_of_current_positive_sequence_real,
                measure_values_of_current_positive_sequence_imaginary,
                measure_values_of_current_positive_sequence_magnitude,
                measure_values_of_current_positive_sequence_angle,
                measure_values_of_current_negative_sequence_real,
                measure_values_of_current_negative_sequence_imaginary,
                measure_values_of_current_negative_sequence_magnitude,
                measure_values_of_current_negative_sequence_angle)

    @staticmethod
    def get_cmp_accuracy_res(accuracy_min, accuracy_max, accuracy_avg, standard_value, exp_accuracy):
        """
        获取精度比较结果
        :param accuracy_min: 精度最小值
        :param accuracy_max: 精度最大值
        :param accuracy_avg: 精度平均值
        :param standard_value: 计算得到的标准值
        :param exp_accuracy: 精度期望值
        :return: "Passed"/"Failed"
        """
        act_accuracy_min = accuracy_min
        act_accuracy_max = accuracy_max
        act_accuracy_avg = accuracy_avg
        act_standard_value = standard_value
        cmp_res = "Failed"
        if ((act_accuracy_min - exp_accuracy <= act_standard_value <= act_accuracy_min + exp_accuracy)
                and (act_accuracy_max - exp_accuracy <= act_standard_value <= act_accuracy_max + exp_accuracy)
                and (act_accuracy_avg - exp_accuracy <= act_standard_value <= act_accuracy_avg + exp_accuracy)):
            cmp_res = "Passed"
        return cmp_res

    def get_accuracy_res_by_sequence_accuracy(self, standard_value, measure_values, exp_accuracy):
        """
        获取精度比较结果，带期望精度值
        :param standard_value: 输入值
        :param measure_values: 测量值
        :param exp_accuracy: 期望精度值
        :return: 测量值和精度比较结果
        """
        (
            min_measure, max_measure, avg_measure
        ) = self.handle_memory.get_measure_accuracy_by_vuf(standard_value, measure_values)

        min_value, min_accuracy = min_measure
        max_value, max_accuracy = max_measure
        avg_value, avg_accuracy = avg_measure
        cmp_accuracy_res = self.get_cmp_accuracy_res(accuracy_min=min_value, accuracy_max=max_value,
                                                     accuracy_avg=avg_value, standard_value=standard_value,
                                                     exp_accuracy=exp_accuracy)
        accuracy_res = (min_value, max_value, avg_value, cmp_accuracy_res)
        return accuracy_res

    @staticmethod
    def write_accuracy_res_to_excel(file_path, wb, ws, index_value, accuracy_res, start_num):
        """
        写入case编号/电压/电流/电压相位角/电流相位角输入值,精度值
        :param file_path: 待写入文件路径
        :param wb: 写入工作簿对象
        :param ws: 写入工作sheet对象
        :param index_value: 接线方式输入数据的行索引
        :param accuracy_res: 精度值
        :param start_num: openpyxl写入列索引
        :return: openpyxl写入列索引
        """
        for k in range(len(accuracy_res)):
            ws.cell(index_value + 2, start_num + k, accuracy_res[k])
        wb.save(file_path)
        start_num += len(accuracy_res)
        return start_num

    def sequence_component_vuf_2e3w1p_precision_measure(self, file_path, input_list, test_type):
        """

        :param file_path: 待写入文件路径
        :param input_list: testcase的数据
        :param test_type: 0:VUF, 1:CUF, 2:相电压序分量, 3:用户序分量
        :return:
        """

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            print(input_list)

            # 从输入excel表格中获取要求的精度, 例如0.001
            (sequence_accuracy, wire_type, wire_order_number) = self.get_test_case_info_of_accuracy_wire_type(
                input_list, i)
            print(sequence_accuracy)
            self.set_wire_type(voltage_wire_value=wire_type)
            self.set_phase_order(phase_order=wire_order_number)

            # 从输入excel表格中获取test case输入值
            (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i, wire_order_number)
            print(case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
                  sample_interval)

            if wire_order_number == 1:
                ub = c_amplitude
                uc = b_amplitude
            else:
                ub = b_amplitude
                uc = c_amplitude

            # 通过excel提供的电压、电流幅值，计算对应的sequence_component
            (VUF_CUF) = self.sequence_2e3w1p(a_amplitude, uc)
            sequence_component_values = VUF_CUF
            print(sequence_component_values)
            sequence_component_values = self.get_sequence_component_of_test_type(test_type,
                                                                                 sequence_component_values)
            write_common_values = [case_id, sequence_accuracy, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle,
                                   c_angle] + sequence_component_values
            start_num = 1
            start_num = self.write_common_values_to_excel(ws, i, write_common_values, start_num)

            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(c_amplitude, b_amplitude, a_amplitude)
            set_current_gear(0, 0, 0)
            set_ac(c_angle, b_angle, a_angle, 120, 240, 0, c_amplitude, b_amplitude, a_amplitude, 1, 1, 1, freq)

            # 获取寄存器测量到的vuf值
            (measure_values_of_vuf, _, _, _, _, _, _, _, _, _, _, _,
             _) = self.get_measure_values_by_sample_cnt_voltage_sequence(
                sample_cnt,
                sample_interval)
            vuf_accuracy = self.get_accuracy_res_by_sequence_accuracy(VUF_CUF, measure_values_of_vuf, sequence_accuracy)

            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, vuf_accuracy, start_num)

    def sequence_component_cuf_2e3w1p_precision_measure(self, file_path, input_list, test_type):
        """

        :param file_path: 待写入文件路径
        :param input_list: testcase的数据
        :param test_type:
        :return:
        """

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            print(input_list)

            # 从输入excel表格中获取要求的精度, 例如0.001
            (sequence_accuracy, wire_type, wire_order_number) = self.get_test_case_info_of_accuracy_wire_type(
                input_list, i)
            print(sequence_accuracy)
            self.set_wire_type(voltage_wire_value=wire_type)
            self.set_phase_order(phase_order=wire_order_number)
            # 从输入excel表格中获取test case输入值
            (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i, wire_order_number)
            print(case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
                  sample_interval)

            # 通过excel提供的电压、电流幅值，计算对应的sequence_component
            (VUF_CUF) = self.sequence_2e3w1p(a_amplitude, b_amplitude)
            sequence_component_values = VUF_CUF
            print(sequence_component_values)
            sequence_component_values = self.get_sequence_component_of_test_type(test_type,
                                                                                 sequence_component_values)
            write_common_values = [case_id, sequence_accuracy, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle,
                                   c_angle] + sequence_component_values
            start_num = 1
            start_num = self.write_common_values_to_excel(ws, i, write_common_values, start_num)

            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(c_amplitude, b_amplitude, a_amplitude)
            set_current_gear(0, 0, 0)
            set_ac(120, 240, 0, c_angle, b_angle, a_angle, 50, 50, 50, c_amplitude, b_amplitude, a_amplitude, freq)

            # 获取寄存器测量到的vuf值
            (measure_values_of_cuf, _, _, _, _, _, _, _, _, _, _, _,
             _) = self.get_measure_values_by_sample_cnt_current_sequence(
                sample_cnt,
                sample_interval)
            cuf_accuracy = self.get_accuracy_res_by_sequence_accuracy(VUF_CUF, measure_values_of_cuf, sequence_accuracy)

            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, cuf_accuracy, start_num)

    def sequence_component_vuf_precision_measure(self, file_path, input_list, test_type):
        """

        :param file_path: 待写入文件路径
        :param input_list: testcase的数据
        :param test_type: 0:VUF, 1:CUF, 2:相电压序分量, 3:用户序分量
        :return:
        """

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            print(input_list)

            # 从输入excel表格中获取要求的精度, 例如0.001
            (sequence_accuracy, wire_type, wire_order_number) = self.get_test_case_info_of_accuracy_wire_type(
                input_list, i)
            print(sequence_accuracy)
            self.set_wire_type(voltage_wire_value=wire_type)
            self.set_phase_order(phase_order=wire_order_number)

            # 从输入excel表格中获取test case输入值
            (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i, wire_order_number)
            print(case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
                  sample_interval)

            if wire_order_number == 1:
                ub = c_amplitude
                uc = b_amplitude
                b = c_angle
                c = b_angle
            else:
                ub = b_amplitude
                uc = c_amplitude
                b = b_angle
                c = c_angle

            # 通过excel提供的电压、电流幅值，计算对应的sequence_component
            (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
             positive_seq_calculate_angle,
             negative_sequence_component, negative_seq_calculate_angle, VUF_CUF) = self.sequence_component_calculation(
                a_amplitude, ub, uc, a_angle, b, c)
            sequence_component_values = (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
                                         positive_seq_calculate_angle, negative_sequence_component,
                                         negative_seq_calculate_angle, VUF_CUF)
            print(sequence_component_values)
            sequence_component_values = self.get_sequence_component_of_test_type(test_type,
                                                                                 sequence_component_values)
            write_common_values = [case_id, sequence_accuracy, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle,
                                   c_angle] + sequence_component_values
            start_num = 1
            start_num = self.write_common_values_to_excel(ws, i, write_common_values, start_num)

            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(c_amplitude, b_amplitude, a_amplitude)
            set_current_gear(0, 0, 0)
            set_ac(c_angle, b_angle, a_angle, 120, 240, 0, c_amplitude, b_amplitude, a_amplitude, 1, 1, 1, freq)

            # 获取寄存器测量到的vuf值
            (measure_values_of_vuf, _, _, _, _, _, _, _, _, _, _, _,
             _) = self.get_measure_values_by_sample_cnt_voltage_sequence(
                sample_cnt,
                sample_interval)
            vuf_accuracy = self.get_accuracy_res_by_sequence_accuracy(VUF_CUF, measure_values_of_vuf, sequence_accuracy)

            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, vuf_accuracy, start_num)

    def sequence_component_cuf_precision_measure(self, file_path, input_list, test_type):
        """

        :param file_path: 待写入文件路径
        :param input_list: testcase的数据
        :param test_type:
        :return:
        """

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
            print(input_list)

            # 从输入excel表格中获取要求的精度, 例如0.001
            (sequence_accuracy, wire_type, wire_order_number) = self.get_test_case_info_of_accuracy_wire_type(
                input_list, i)
            print(sequence_accuracy)
            self.set_wire_type(voltage_wire_value=wire_type)
            self.set_phase_order(phase_order=wire_order_number)
            # 从输入excel表格中获取test case输入值
            (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i, wire_order_number)
            print(case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
                  sample_interval)

            if wire_order_number == 1:
                ub = c_amplitude
                uc = b_amplitude
                b = c_angle
                c = b_angle
            else:
                ub = b_amplitude
                uc = c_amplitude
                b = b_angle
                c = c_angle

            # 通过excel提供的电压、电流幅值，计算对应的sequence_component
            (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
             positive_seq_calculate_angle,
             negative_sequence_component, negative_seq_calculate_angle, VUF_CUF) = self.sequence_component_calculation(
                a_amplitude, ub, uc, a_angle, b, c)
            sequence_component_values = (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
                                         positive_seq_calculate_angle, negative_sequence_component,
                                         negative_seq_calculate_angle, VUF_CUF)
            print(sequence_component_values)
            sequence_component_values = self.get_sequence_component_of_test_type(test_type,
                                                                                 sequence_component_values)
            write_common_values = [case_id, sequence_accuracy, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle,
                                   c_angle] + sequence_component_values
            start_num = 1
            start_num = self.write_common_values_to_excel(ws, i, write_common_values, start_num)

            # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(c_amplitude, b_amplitude, a_amplitude)
            set_current_gear(0, 0, 0)
            set_ac(120, 240, 0, c_angle, b_angle, a_angle, 50, 50, 50, c_amplitude, b_amplitude, a_amplitude, freq)

            # 获取寄存器测量到的vuf值
            (measure_values_of_cuf, _, _, _, _, _, _, _, _, _, _, _,
             _) = self.get_measure_values_by_sample_cnt_current_sequence(
                sample_cnt,
                sample_interval)
            cuf_accuracy = self.get_accuracy_res_by_sequence_accuracy(VUF_CUF, measure_values_of_cuf, sequence_accuracy)

            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, cuf_accuracy, start_num)

    def sequence_component_phase_voltage_precision_measure(self, file_path, input_list, test_type):
        """

        :param file_path: 待写入文件路径
        :param input_list: testcase的数据
        :param test_type:
        :return:
        """

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")

            # 从输入excel表格中获取要求的精度, 例如0.001
            (sequence_accuracy, wire_type, wire_order_number) = self.get_test_case_info_of_accuracy_wire_type(
                input_list, i)
            print(sequence_accuracy)
            self.set_wire_type(voltage_wire_value=wire_type)
            self.set_phase_order(phase_order=wire_order_number)
            # 从输入excel表格中获取test case输入值
            (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i, wire_order_number)

            if wire_order_number == 1:
                ub = c_amplitude
                uc = b_amplitude
                b = c_angle
                c = b_angle
            else:
                ub = b_amplitude
                uc = c_amplitude
                b = b_angle
                c = c_angle

            # 通过excel提供的电压、电流幅值，计算对应的sequence_component
            (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
             positive_seq_calculate_angle,
             negative_sequence_component, negative_seq_calculate_angle, VUF_CUF) = self.sequence_component_calculation(
                a_amplitude, ub, uc, a_angle, b, c)
            sequence_component_values = (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
                                         positive_seq_calculate_angle, negative_sequence_component,
                                         negative_seq_calculate_angle, VUF_CUF)
            sequence_component_values = self.get_sequence_component_of_test_type(test_type,
                                                                                 sequence_component_values)
            write_common_values = [case_id, sequence_accuracy, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle,
                                   c_angle] + sequence_component_values
            print(write_common_values)
            start_num = 1
            start_num = self.write_common_values_to_excel(ws, i, write_common_values, start_num)

            # # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(c_amplitude, b_amplitude, a_amplitude)
            set_current_gear(0, 0, 0)
            set_ac(c_angle, b_angle, a_angle, 120, 240, 0, c_amplitude, b_amplitude, a_amplitude, 1, 1, 1, freq)
            # 获取寄存器测量到的sequence值
            (measure_values_of_vuf,
             measure_values_of_voltage_zero_sequence_real,
             measure_values_of_voltage_zero_sequence_imaginary,
             measure_values_of_voltage_zero_sequence_magnitude,
             measure_values_of_voltage_zero_sequence_angle,
             measure_values_of_voltage_positive_sequence_real,
             measure_values_of_voltage_positive_sequence_imaginary,
             measure_values_of_voltage_positive_sequence_magnitude,
             measure_values_of_voltage_positive_sequence_angle,
             measure_values_of_voltage_negative_sequence_real,
             measure_values_of_voltage_negative_sequence_imaginary,
             measure_values_of_voltage_negative_sequence_magnitude,
             measure_values_of_voltage_negative_sequence_angle) = (
                self.get_measure_values_by_sample_cnt_voltage_sequence(sample_cnt, sample_interval))
            zero_sequence_real = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_zero_sequence_real)
            zero_sequence_imag = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_zero_sequence_imaginary)
            zero_sequence_magnitude = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_zero_sequence_magnitude)
            zero_sequence_angle = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_zero_sequence_angle)

            positive_sequence_real = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_positive_sequence_real)
            positive_sequence_imaginary = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_positive_sequence_imaginary)
            positive_sequence_magnitude = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_positive_sequence_magnitude)
            positive_sequence_angle = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_positive_sequence_angle)

            negative_sequence_real = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_negative_sequence_real)
            negative_sequence_imaginary = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_negative_sequence_imaginary)
            negative_sequence_magnitude = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_negative_sequence_magnitude)
            negative_sequence_angle = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_voltage_negative_sequence_angle)
            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_real, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_imag, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_magnitude, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_angle, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_real, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_imaginary, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_magnitude, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_angle, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_real, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_imaginary, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_magnitude, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_angle, start_num)

            vuf_accuracy = self.get_accuracy_res_by_sequence_accuracy(VUF_CUF, measure_values_of_vuf, sequence_accuracy)

            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, vuf_accuracy, start_num)

    def sequence_component_current_precision_measure(self, file_path, input_list, test_type):
        """

        :param file_path: 待写入文件路径
        :param input_list: testcase的数据
        :param test_type:
        :return:
        """

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 写结果excel的第一行(Title),title内容在report_table_heading文件中定义
        self.write_table_title_of_3e4wy_to_excel(file_path, wb, ws, test_type)

        for i in range(len(input_list)):
            if i == len(input_list) - 1:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                break
            else:
                logging.info(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")
                print(f"测试进度:{input_list[i]},执行时间:{time.strftime('%Y_%m_%d %H:%M:%S')}")

            # 从输入excel表格中获取要求的精度, 例如0.001
            (sequence_accuracy, wire_type, wire_order_number) = self.get_test_case_info_of_accuracy_wire_type(
                input_list, i)
            print(sequence_accuracy)
            self.set_wire_type(voltage_wire_value=wire_type)
            self.set_phase_order(phase_order=wire_order_number)
            # 从输入excel表格中获取test case输入值
            (case_id, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle, c_angle, freq, sample_cnt,
             sample_interval) = self.get_test_case_info_of_input_value(input_list, i, wire_order_number)

            if wire_order_number == 1:
                ub = c_amplitude
                uc = b_amplitude
                b = c_angle
                c = b_angle
            else:
                ub = b_amplitude
                uc = c_amplitude
                b = b_angle
                c = c_angle

            # 通过excel提供的电压、电流幅值，计算对应的sequence_component
            (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
             positive_seq_calculate_angle,
             negative_sequence_component, negative_seq_calculate_angle, VUF_CUF) = self.sequence_component_calculation(
                a_amplitude, ub, uc, a_angle, b, c)
            sequence_component_values = (zero_sequence_component, zero_seq_calculate_angle, positive_sequence_component,
                                         positive_seq_calculate_angle, negative_sequence_component,
                                         negative_seq_calculate_angle, VUF_CUF)
            sequence_component_values = self.get_sequence_component_of_test_type(test_type,
                                                                                 sequence_component_values)
            write_common_values = [case_id, sequence_accuracy, a_amplitude, b_amplitude, c_amplitude, a_angle, b_angle,
                                   c_angle] + sequence_component_values
            print(write_common_values)
            start_num = 1
            start_num = self.write_common_values_to_excel(ws, i, write_common_values, start_num)

            # # 升源+设置电压/电流档位
            set_ac(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            set_voltage_gear(c_amplitude, b_amplitude, a_amplitude)
            set_current_gear(0, 0, 0)
            set_ac(120, 240, 0, c_angle, b_angle, a_angle, 50, 50, 50, c_amplitude, b_amplitude, a_amplitude, freq)

            # 获取寄存器测量到的sequence值
            (measure_values_of_cuf,
             measure_values_of_current_zero_sequence_real,
             measure_values_of_current_zero_sequence_imaginary,
             measure_values_of_current_zero_sequence_magnitude,
             measure_values_of_current_zero_sequence_angle,
             measure_values_of_current_positive_sequence_real,
             measure_values_of_current_positive_sequence_imaginary,
             measure_values_of_current_positive_sequence_magnitude,
             measure_values_of_current_positive_sequence_angle,
             measure_values_of_current_negative_sequence_real,
             measure_values_of_current_negative_sequence_imaginary,
             measure_values_of_current_negative_sequence_magnitude,
             measure_values_of_current_negative_sequence_angle) = (
                self.get_measure_values_by_sample_cnt_current_sequence(sample_cnt, sample_interval))
            zero_sequence_real = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_zero_sequence_real)
            zero_sequence_imag = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_zero_sequence_imaginary)
            zero_sequence_magnitude = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_zero_sequence_magnitude)
            zero_sequence_angle = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_zero_sequence_angle)

            positive_sequence_real = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_positive_sequence_real)
            positive_sequence_imaginary = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_positive_sequence_imaginary)
            positive_sequence_magnitude = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_positive_sequence_magnitude)
            positive_sequence_angle = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_positive_sequence_angle)

            negative_sequence_real = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_negative_sequence_real)
            negative_sequence_imaginary = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_negative_sequence_imaginary)
            negative_sequence_magnitude = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_negative_sequence_magnitude)
            negative_sequence_angle = self.get_accuracy_res_by_sequence_rea_imag(
                measure_values_of_current_negative_sequence_angle)
            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_real, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_imag, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_magnitude, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, zero_sequence_angle, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_real, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_imaginary, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_magnitude, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, positive_sequence_angle, start_num)

            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_real, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_imaginary, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_magnitude, start_num)
            start_num = self.write_accuracy_res_to_excel(file_path, wb, ws, i, negative_sequence_angle, start_num)

            vuf_accuracy = self.get_accuracy_res_by_sequence_accuracy(VUF_CUF, measure_values_of_cuf, sequence_accuracy)

            # 第i个case，继续在每一列中写值，write_accuracy_res_to_excel中返回列数自加后的值start_numm供下一个数据写入使用。
            # vuf_accuracy写入元组数据， 包含max、min、avg，compare
            self.write_accuracy_res_to_excel(file_path, wb, ws, i, vuf_accuracy, start_num)

    def select_wire_type(self, case_path, sheet_name, wire_type, test_type):
        """
        选择接线方式
        :param case_path: 测试文件名
        :param sheet_name: sheet名
        :param test_type:
        # :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wN, 3:2e3wD, 4:3e3wD, 5:2.5e4wY, 6:3e4wY,

        :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 4:3e3wD  #AcuRev1320
        :return:
        # 3 element 4 wire Wye       3LN     3CT    6   (0 0)
        # 2.5 element 4 wire Wye     3LN-2.5 3CT    5   (5 0)
        # 3 element 3 wire Delta     3LL     3CT    4   (3 0)
        # 2 element 3 wire Delta     2LL     3CT    3   (2 0)
        # 2 element 3 wire network   2LL     2CT    2   (2 2)
        # 2 element 3 wire 1 phase   1LL     2CT    1   (4 2)
        # 1 element 2 wire           1LN     1CT    0   (1 1)
        """
        if wire_type == 1:
            # 2e3w1p AcuRev1320对应寄存器中值为0
            # self.set_wire_type(voltage_wire_value=2)

            # 读取test case excel文件中的数据
            source_input_list = data_read(case_path, sheet_name)

            # 按照wire_type选择需要执行的test case
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, test_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            # 当前目录生成结果excel空文件
            file_path = self.get_save_filepath_of_3e4wy(save_filedir)
            # # 快速精度测试主函数
            if test_type == 0:
                self.sequence_component_vuf_2e3w1p_precision_measure(file_path, input_list, test_type)
            elif test_type == 1:
                self.sequence_component_cuf_2e3w1p_precision_measure(file_path, input_list, test_type)
        if wire_type == 6:
            # 3e4wy AcuRev1320对应寄存器中值为0
            # self.set_wire_type(voltage_wire_value=4)

            # 读取test case excel文件中的数据
            source_input_list = data_read(case_path, sheet_name)

            # 按照wire_type选择需要执行的test case
            input_list_of_wire_type = self.get_input_list_by_wire_type(source_input_list, test_type)
            input_list = self.get_real_time_parameters_by_input_list_of_wire_type(input_list_of_wire_type)
            # 当前目录生成结果excel空文件
            file_path = self.get_save_filepath_of_3e4wy(save_filedir)
            # # 快速精度测试主函数
            if test_type == 0:
                self.sequence_component_vuf_precision_measure(file_path, input_list, test_type)
            elif test_type == 1:
                self.sequence_component_cuf_precision_measure(file_path, input_list, test_type)
            elif test_type == 2:
                self.sequence_component_phase_voltage_precision_measure(file_path, input_list, test_type)
            else:
                self.sequence_component_current_precision_measure(file_path, input_list, test_type)


def run_sequence_test_script(test_type, wire_type):
    """
    运行脚本入口
    :param test_type: 0:mV, 1:mA, 2:rct
    :param wire_type: 0:1e2w1p, 1:2e3w1p, 2:2e3wD, 3:2e3wN, 4:3e4wY, 5:3e3wD  #AcuRev1320
    :return:
    """
    print(f"====================Demand Test Start====================")
    print(f"======================{time.strftime('%Y_%m_%d %H:%M:%S')}======================")
    start_time = time.time()
    switch_device_screen_interface(inter=0x01)  # 切换至交流界面
    set_gear_switching_mode('00000000')  # 档位切换归零,自动

    precision_measure = SequenceTest()

    precision_measure.select_test_case(test_type=test_type, wire_type=wire_type)

    # 关闭ModbusClient客户端连接
    precision_measure.handle_memory.modbus_client.close()
    up_source_ac()
    switch_device_screen_interface(inter=0x00)  # 切换至默认界面
    set_gear_switching_mode(mode='01000000')  # 档位切换归零,手动
    print(f"====================测试总耗时:{time.time() - start_time}====================")
    print(f"====================={time.strftime('%Y_%m_%d %H:%M:%S')}=====================")
    print(f"====================Demand Test End====================")


if __name__ == '__main__':
    """
    :param load_type: 0:Three-phase load 1:Two-phase load
    :param test_mode_list: 0:VUF, 1:CUF, 2:相电压序分量, 3:用户序分量
    :param wire_mode: 
    """

    test_mode_list = [2, 3]
    wire_mode = 6
    for test_mode in test_mode_list:
        run_sequence_test_script(test_type=test_mode, wire_type=wire_mode)

    # precision_measure = SequenceTest()
    # r = precision_measure.sequence_2e3w1phase(230, 250)
    # print(r)
    # source_input_list = data_read(test_case_path, sheet_name_mV)
    # print(source_input_list)
