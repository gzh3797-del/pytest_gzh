import logging
from comm.Source_control_3021dc import *
from tools.excel_operate import dcpara_addr_get, data_read
from comm.modbus_rtu_tcp import *
import openpyxl

base_dir = os.path.dirname(os.path.abspath(__file__))
base_dir = base_dir.replace("\\", "/")
iom_project_configfile_path = '/test_case/IOM/test_data/'
iom_project_outputfile_path = '/test_case/IOM/output_data/'
iom_project_configfile = 'IOM_input_parameters.xlsx'

if not os.path.exists(base_dir + '/output_data/'):
    os.makedirs(base_dir + '/output_data/')

modbus_address_setting = dcpara_addr_get(iom_project_configfile_path + iom_project_configfile, 'modbus_address_aiao_setting')
modbus_address_reading = dcpara_addr_get(iom_project_configfile_path + iom_project_configfile, 'modbus_address_aiao_reading')

# 默认设置为Rtu通信方式
ModbusClient = ModbusRtuOrTcp()


def read_iom_ai_voltage_current(standard_value: float = 0.000, ai_num: int = 1, times: int = 1):
    """
    功能说明:
       输入电压或电流获取寄存器地址值
    参数：
        standard_value: 输入标准值
        ai_num        ：端子序号
        time          : 获取寄存器数据次数
    返回：
        返回误差最大的那个值
    """
    address = modbus_address_reading['AI{} input original data float32'.format(ai_num)]['Start(Dec)']
    count = modbus_address_reading['AI{} input original data float32'.format(ai_num)]['Reg']
    val_list = []
    for v in range(times):
        time.sleep(0.2)
        value = ModbusClient.read_measurement(address, count, slave=1)
        logging.info('AI{} input original data float32 ret is: {}'.format(ai_num,value,))
        reg = hex(value[0]).replace('0x', '').zfill(4) + hex(value[1]).replace('0x', '').zfill(4)
        hex_num = reg.replace('0x', '')
        integer_num = int(hex_num, 16)
        value_measu = struct.unpack('!f', struct.pack('!I', integer_num))[0]
        val_list.append(value_measu)
    val_list.sort()
    if abs(val_list[-1] - standard_value) > abs(val_list[0] - standard_value):
        return val_list[-1]
    return val_list[0]


def set_iom_ai_ao_configuration(ai_num: int, line_num: int, type_value: int, bop_limit: int, top_limit: int, point_x: list, point_y: list):
    """
    功能说明:
        设置单个AI或AO配置信息。
    参数:
        ai_num    : 用于标识那个AI端子
        line_num  : 折线数量
        type_value：输入信号的类型 0: voltage：0-10 V
                                1: voltage：2-10 V
                                2:current：0-20mA
                                3:current：4-20mA
        bop_limit: 电压/电流底部限制
        top_limit： 电压/电流顶部限制
        point_x  ： X轴三个值
        point_y  ： y轴三个值
    返回：
        无
    """
    address_ai_type = modbus_address_setting['AI{} Type uint16_t'.format(ai_num)]['Start(Dec)']
    logging.info('AI{} Type: {}'.format(ai_num, address_ai_type))

    address_line_num = modbus_address_setting['AI{} Line Num uint16_t'.format(ai_num)]['Start(Dec)']
    logging.info('AI{} Line Num: {}'.format(ai_num, address_line_num))

    address_bot_limit = modbus_address_setting['AI{} Bot Limit uint32_t'.format(ai_num)]['Start(Dec)']
    logging.info('AI{} bot Limit: {}'.format(ai_num, address_bot_limit))

    address_top_limit = modbus_address_setting['AI{} Top Limit uint32_t'.format(ai_num)]['Start(Dec)']
    logging.info('AI{} Top Limit: {}'.format(ai_num, address_top_limit))

    address_line_point_x = modbus_address_setting['AI{} Line Point\n X1 X2 X3 X4 uint32_t'.format(ai_num)]['Start(Dec)']
    logging.info('AI{} Line Point_X: {}'.format(ai_num, address_line_point_x))

    address_line_point_y = modbus_address_setting['AI{} Line Point \nY1 Y2 Y3 Y4 int32_t'.format(ai_num)]['Start(Dec)']
    logging.info('AI{} Line Point_Y: {}'.format(ai_num, address_line_point_y))

    ModbusClient.write_registers(address=address_ai_type, values=type_value, slave=1)
    ModbusClient.write_registers(address=address_line_num, values=line_num, slave=1)
    ModbusClient.write_registers(address=address_bot_limit, values=bop_limit, slave=1)
    ModbusClient.write_registers(address=address_top_limit, values=top_limit, slave=1)
    num_x = 0
    num_y = 0
    for i in range(4):
        ModbusClient.write_registers(address=address_line_point_x + num_x, values=point_x[i], slave=1)
        num_x = num_x + 2
    for i in range(4):
        ModbusClient.write_registers(address=address_line_point_y + num_y, values=point_y[i], slave=1)
        num_y = num_y + 2


def get_ai_y_from_x(x: float,
                    points: list,
                    limit_bot_top_x: list,
                    type_value: int = 0,
                    line_num: int = 1):
    """
    功能说明:
        根据输入x值，在由4个点构成的折线上线性插值计算y值。
        若x超出范围，则返回边界点对应的y值。
    参数：
        x             : 用户输入的x值
        points        : 4个点的列表，每个点是(x, y)，例如 [(x1,y1), (x2,y2), (x3,y3), (x4,y4)]
        limit_bot_top : 用于限制电压和电流最小值和最大值
        type_value    ：输入信号的类型 0: voltage：0-10 V
                                           1: voltage：2-10 V
                                           2:current：0-20mA
                                           3:current：4-20mA
        line_num      : 折线段数

    返回：
        ret_standard    ： 对应的y值（float）
        ret_lower       ：下限
        ret_upper       ：上限
        precision_value ：精度
    """
    #  精度变量获取
    if (type_value == 0) or (type_value == 1):
        precision_value = 0.02
    elif (type_value == 3) or (type_value == 4):
        precision_value = 0.04
    else:
        print("输入信号类型错误，退出运算")
        return False
    # 标识电压电流范围数据
    voltage_current_range_data = {1: [0, 10], 2: [2, 10], 3: [0, 20], 4: [4, 20]}.get(type_value)
    # 定义bot和top
    bot_x = limit_bot_top_x[0]
    top_x = limit_bot_top_x[1]

    # 数据合法性校验
    if (
        (bot_x >= voltage_current_range_data[0]) and
        (top_x <= voltage_current_range_data[1]) and
        (points[0][0] < points[1][0]) and
        (points[1][0] < points[2][0]) and
        (points[0][1] < points[1][1]) and
        (points[1][1] < points[2][1])):
        print("数据校验成功，进行计算")
    else:
        print("数据校验失败，中断计算")
        return None
    # 线性插值公式：y = y0 + (x - x0) * (y1 - y0) / (x1 - x0)
    if line_num == 1:
        x1, x2 = [points[0][0], points[1][0]]
        y1, y2 = [points[0][1], points[1][1]]
        k1 = (y2 - y1) / (x2 - x1)
        #  bot处在第一段，x <= bot_x, bot_x <= x <= x1
        ret_standard = y1 + k1 * (x - x1)
        ret_lower = y1 + k1 * (x - precision_value - x1)
        ret_upper = y1 + k1 * (x + precision_value - x1)
        if x <= bot_x:
            ret_standard = y1 + k1 * (bot_x - x1)
            ret_lower = y1 + k1 * (bot_x - precision_value - x1)
            ret_upper = y1 + k1 * (bot_x + precision_value - x1)
        if x >= top_x:
            ret_standard = y1 + k1 * (top_x - x1)
            ret_lower = y1 + k1 * (top_x - precision_value - x1)
            ret_upper = y1 + k1 * (top_x + precision_value - x1)
        return ret_standard, ret_lower, ret_upper, precision_value
    elif line_num == 2:
        x1, x2, x3 = [points[0][0], points[1][0], points[2][0]]
        y1, y2, y3 = [points[0][1], points[1][1], points[2][1]]
        k1 = (y2 - y1) / (x2 - x1)
        k2 = (y3 - y2) / (x3 - x2)
        #  bot处在第一段，x <= bot_x, bot_x <= x <= x1
        if x <= x2:
            ret_standard = y1 + k1 * (x - x1)
            ret_lower = y1 + k1 * (x - precision_value - x1)
            ret_upper = y1 + k1 * (x + precision_value - x1)
            if (x <= bot_x) and (x2 >= bot_x):
                ret_standard = y1 + k1 * (bot_x - x1)
                ret_lower = y1 + k1 * (bot_x - precision_value - x1)
                ret_upper = y1 + k1 * (bot_x + precision_value -  x1)
            if (x <= bot_x) and (x2 <= bot_x):
                ret_standard = y2 + k2 * (bot_x - x2)
                ret_lower = y2 + k2 * (bot_x - precision_value - x2)
                ret_upper = y2 + k2 * (bot_x + precision_value - x2)
            elif x >= top_x:
                ret_standard = y1 + k1 * (top_x - x1)
                ret_lower = y1 + k1 * (top_x - precision_value - x1)
                ret_upper = y1 + k1 * (top_x + precision_value - x1)
            else:
                pass
            # print("testing： ",ret_standard, ret_lower, ret_upper, precision_value)
            return ret_standard, ret_lower, ret_upper, precision_value
        elif x >= x2:
            ret_standard = y2 + k2 * (x - x2)
            ret_lower = y2 + k2 * (x - precision_value - x2)
            ret_upper = y2 + k2 * (x + precision_value - x2)
            if x <= bot_x:
                ret_standard = y2 + k2 * (bot_x - x2)
                ret_lower = y2 + k2 * (bot_x - precision_value - x2)
                ret_upper = y2 + k2 * (bot_x + precision_value - x2)
            elif (x >= top_x) and (x2 <= top_x):
                ret_standard = y2 + k2 * (top_x - x2)
                ret_lower = y2 + k2 * (top_x - precision_value - x2)
                ret_upper = y2 + k2 * (top_x + precision_value - x2)
            elif (x >= top_x) and (x2 >= top_x):
                ret_standard = y1 + k1 * (top_x - x1)
                ret_lower = y1 + k1 * (top_x - precision_value - x1)
                ret_upper = y1 + k1 * (top_x + precision_value - x1)
            else:
                pass
            # print("testing： ", ret_standard, ret_lower, ret_upper, precision_value)
            return ret_standard, ret_lower, ret_upper, precision_value
        else:
            pass
    elif line_num == 3:
        x1, x2, x3, x4 = [points[0][0], points[1][0], points[2][0],points[3][0]]
        y1, y2, y3, y4 = [points[0][1], points[1][1], points[2][1],points[3][1]]
        k1 = (y2 - y1) / (x2 - x1)
        k2 = (y3 - y2) / (x3 - x2)
        k3 = (y4 - y3) / (x4 - x3)
        #  bot处在第一段，top处在第一段,x <= bot_x, bot_x <= x <= top_x, top_x <= x
        if x <= x2:
            ret_standard = y1 + k1 * (x - x1)
            ret_lower = y1 + k1 * (x - precision_value - x1)
            ret_upper = y1 + k1 * (x + precision_value - x1)
            if (x <= bot_x) and (bot_x <= x2):
                ret_standard = y1 + k1 * (bot_x - x1)
                ret_lower = y1 + k1 * (bot_x - precision_value - x1)
                ret_upper = y1 + k1 * (bot_x + precision_value - x1)
            elif (x <= bot_x) and (bot_x >= x2) and (bot_x <= x3):
                ret_standard = y2 + k2 * (bot_x - x2)
                ret_lower = y2 + k2 * (bot_x - precision_value - x2)
                ret_upper = y2 + k2 * (bot_x + precision_value - x2)
            elif (x <= bot_x) and (bot_x >= x3):
                ret_standard = y3 + k3 * (bot_x - x3)
                ret_lower = y3 + k3 * (bot_x - precision_value - x3)
                ret_upper = y3 + k3 * (bot_x + precision_value - x3)
            elif x >= top_x:
                ret_standard = y1 + k1 * (top_x - x1)
                ret_lower = y1 + k1 * (top_x - precision_value - x1)
                ret_upper = y1 + k1 * (top_x + precision_value - x1)
            else:
                pass
            # print("testing： ", ret_standard, ret_lower, ret_upper, precision_value)
            return ret_standard, ret_lower, ret_upper, precision_value
        elif (x >= x2) and (x <= x3):
            ret_standard = y2 + k2 * (x - x2)
            ret_lower = y2 + k2 * (x - precision_value - x2)
            ret_upper = y2 + k2 * (x + precision_value - x2)
            if (x <= bot_x) and (bot_x <= x3):
                ret_standard = y2 + k2 * (bot_x - x2)
                ret_lower = y2 + k2 * (bot_x - precision_value - x2)
                ret_upper = y2 + k2 * (bot_x + precision_value - x2)
            elif (x <= bot_x) and (bot_x >= x3):
                ret_standard = y3 + k3 * (bot_x - x3)
                ret_lower = y3 + k3 * (bot_x - precision_value - x3)
                ret_upper = y3 + k3 * (bot_x + precision_value - x3)
            elif (x >= top_x) and (x2 <= top_x):
                ret_standard = y2 + k2 * (top_x - x2)
                ret_lower = y2 + k2 * (top_x - precision_value - x2)
                ret_upper = y2 + k2 * (top_x + precision_value - x2)
            elif (x >= top_x) and (x2 >= top_x):
                ret_standard = y1 + k1 * (top_x - x1)
                ret_lower = y1 + k1 * (top_x - precision_value - x1)
                ret_upper = y1 + k1 * (top_x + precision_value - x1)
            else:
                pass
            # print("testing： ", ret_standard, ret_lower, ret_upper, precision_value)
            return ret_standard, ret_lower, ret_upper, precision_value
        elif x >= x3:
            ret_standard = y3 + k3 * (x - x3)
            ret_lower = y3 + k3 * (x - precision_value - x3)
            ret_upper = y3 + k3 * (x + precision_value - x3)
            if x <= bot_x:
                ret_standard = y3 + k3 * (bot_x - x3)
                ret_lower = y3 + k3 * (bot_x - precision_value - x3)
                ret_upper = y3 + k3 * (bot_x + precision_value - x3)
            elif (x >= top_x) and (x3 <= top_x):
                ret_standard = y3 + k3 * (top_x - x3)
                ret_lower = y3 + k3 * (top_x - precision_value - x3)
                ret_upper = y3 + k3 * (top_x + precision_value - x3)
            elif (x >= top_x) and (x2 <= top_x) and (x3 >= top_x):
                ret_standard = y2 + k2 * (top_x - x2)
                ret_lower = y2 + k2 * (top_x - precision_value - x2)
                ret_upper = y2 + k2 * (top_x + precision_value - x2)
            elif (x >= top_x) and (x2 >= top_x):
                ret_standard = y1 + k1 * (top_x - x1)
                ret_lower = y1 + k1 * (top_x - precision_value - x1)
                ret_upper = y1 + k1 * (top_x + precision_value - x1)
            else:
                pass
            # print("testing： ", ret_standard, ret_lower, ret_upper, precision_value)
            return ret_standard, ret_lower, ret_upper, precision_value
        else:
            pass


def get_ao_y_from_x(x: float,
                    points: list,
                    limit_y: list,
                    type_value: int = 1,
                    line_num: int = 1):
    """
    功能说明:
        根据输入x值，在由4个点构成的折线上线性插值计算y值。
        若x超出范围，则返回边界点对应的y值。
    参数：
        x                    : 用户输入的x值
        points               : 4个点的列表，每个点是(x, y)，例如 [(x1,y1), (x2,y2), (x3,y3), (x4,y4)]
        limit_bot_top        : 用于限制电压和电流最小值和最大值
        type_value           ：输入信号的类型 0: voltage：0-10 V
                                           1: voltage：2-10 V
                                           2:current：0-20mA
                                           3:current：4-20mA
        line_num             : 折线段数
    返回：
        对应的y值（float）
    """
    #  精度定义
    if (type_value == 0) or (type_value == 1):
        precision_value_y = 0.05
    elif (type_value == 3) or (type_value == 4):
        precision_value_y = 0.10
    else:
        print("输入信号类型错误，退出运算")
        return False

    # 标识电压电流范围数据
    voltage_current_range_data = {1: [0, 10], 2: [2, 10], 3: [0, 20], 4: [4, 20]}.get(type_value)
    # 定义bot和top
    bot_y = limit_y[0]
    top_y = limit_y[1]

    # 数据合法性校验
    if (
            (bot_y >= voltage_current_range_data[0]) and
            (top_y <= voltage_current_range_data[1]) and
            (points[0][0] < points[1][0]) and
            (points[1][0] < points[2][0]) and
            (points[0][1] < points[1][1]) and
            (points[1][1] < points[2][1])):
        print("数据校验成功，进行计算")
    else:
        print("数据校验失败，中断计算")
        return None
    # 线性插值公式：y = y0 + (x - x0) * (y1 - y0) / (x1 - x0)
    if line_num == 1:
        x1, x2 = [points[0][0], points[1][0]]
        y1, y2 = [points[0][1], points[1][1]]
        k1 = (y2 - y1) / (x2 - x1)
        #  换算bot对应X轴值
        #  b1 = y1-k1*x1
        #  x = （y-b）/k
        bot_y_to_x = (bot_y - y1 + k1*x1) / k1
        top_y_to_x = (top_y - y1 + k1*x1) / k1
        precision_value_x_k1 = (precision_value_y - y1 + k1*x1) / k1

        #  bot处在第一段，x <= bot_x, bot_x <= x <= x1
        ret_standard = y1 + k1 * (x - x1)
        ret_lower = 0
        ret_upper = 0
        if x <= bot_y_to_x:
            ret_standard = y1 + k1 * (bot_y_to_x - x1)
            ret_lower = y1 + k1 * (bot_y_to_x - precision_value_x_k1 - x1)
            ret_upper = y1 + k1 * (bot_y_to_x + precision_value_x_k1 - x1)
        elif x >= top_y_to_x:
            ret_standard = y1 + k1 * (top_y_to_x - x1)
            ret_lower = y1 + k1 * (top_y_to_x - precision_value_x_k1 - x1)
            ret_upper = y1 + k1 * (top_y_to_x + precision_value_x_k1 - x1)
        else:
            pass
        return ret_standard, ret_lower, ret_upper, precision_value_y
    elif line_num == 2:
        x1, x2, x3 = [points[0][0], points[1][0], points[2][0]]
        y1, y2, y3 = [points[0][1], points[1][1], points[2][1]]
        k1 = (y2 - y1) / (x2 - x1)
        k2 = (y3 - y2) / (x3 - x2)
        #  换算bot对应X轴值
        #  b = y1-k1*x1
        #  x = （y-b）/k
        bot_y_to_x = (bot_y - y1 + k1 * x1) / k1
        top_y_to_x = (top_y - y1 + k1 * x1) / k1

        precision_value_x_k1 = (precision_value_y - y1 + k1*x1) / k1
        precision_value_x_k2 = (precision_value_y - y2 + k2*x2) / k2

        if bot_y >= y2:
            bot_y_to_x = (bot_y - y2 + k2 * x2) / k2
        if top_y >= y2:
            top_y_to_x = (top_y - y2 + k2 * x2) / k2

        # bot处在第一段，x <= bot_x, bot_x <= x <= x1
        if x <= x2:
            ret_standard = y1 + k1 * (x - x1)
            ret_lower = y1 + k1 * (x - precision_value_x_k1 - x1)
            ret_upper = y1 + k1 * (x + precision_value_x_k1 - x1)
            if (x <= bot_y_to_x) and (x2 >= bot_y_to_x):
                ret_standard = y1 + k1 * (bot_y_to_x - x1)
                ret_lower = y1 + k1 * (bot_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (bot_y_to_x + precision_value_x_k1 - x1)
            if (x <= bot_y_to_x) and (x2 <= bot_y_to_x):
                ret_standard = y2 + k2 * (bot_y_to_x - x2)
                ret_lower = y2 + k2 * (bot_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (bot_y_to_x + precision_value_x_k2 - x2)
            elif x >= top_y_to_x:
                ret_standard = y1 + k1 * (top_y_to_x - x1)
                ret_lower = y1 + k1 * (top_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (top_y_to_x + precision_value_x_k1 - x1)
            else:
                pass
            return ret_standard, ret_lower, ret_upper, precision_value_y
        elif x >= x2:
            ret_standard = y2 + k2 * (x - x2)
            ret_lower = y2 + k2 * (x - precision_value_x_k2 - x2)
            ret_upper = y2 + k2 * (x + precision_value_x_k2 - x2)
            if x <= bot_y_to_x:
                ret_standard = y2 + k2 * (bot_y_to_x - x2)
                ret_lower = y2 + k2 * (bot_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (bot_y_to_x + precision_value_x_k2 - x2)
            elif (x >= top_y_to_x) and (x2 <= top_y_to_x):
                ret_standard = y2 + k2 * (top_y_to_x - x2)
                ret_lower = y2 + k2 * (top_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (top_y_to_x + precision_value_x_k2 - x2)
            elif (x >= top_y_to_x) and (x2 >= top_y_to_x):
                ret_standard = y1 + k1 * (top_y_to_x - x1)
                ret_lower = y1 + k1 * (top_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (top_y_to_x + precision_value_x_k1 - x1)
            else:
                pass
            return ret_standard, ret_lower, ret_upper, precision_value_y
        else:
            pass
    elif line_num == 3:
        x1, x2, x3, x4 = [points[0][0], points[1][0], points[2][0], points[3][0]]
        y1, y2, y3, y4 = [points[0][1], points[1][1], points[2][1], points[3][1]]
        k1 = (y2 - y1) / (x2 - x1)
        k2 = (y3 - y2) / (x3 - x2)
        k3 = (y4 - y3) / (x4 - x3)
        #  换算bot对应X轴值
        #  b = y1-k1*x1
        #  x = （y-b）/k
        bot_y_to_x = (bot_y - y1 + k1 * x1) / k1
        top_y_to_x = (top_y - y1 + k1 * x1) / k1

        precision_value_x_k1 = (precision_value_y - y1 + k1*x1) / k1
        precision_value_x_k2 = (precision_value_y - y2 + k2*x2) / k2
        precision_value_x_k3 = (precision_value_y - y3 + k3*x3) / k3

        if (bot_y >= y2) and (bot_y <= y3):
            bot_y_to_x = (bot_y + y2 + k2 * x2) / k2
        if bot_y >= y3:
            bot_y_to_x = (bot_y - y3 + k3 * x3) / k3

        if (top_y >= y2) and (top_y <= y3):
            top_y_to_x = (top_y - y2 + k2 * x2) / k2
        if top_y >= y3:
            top_y_to_x = (top_y - y3 + k3 * x3) / k3

        #  bot处在第一段，top处在第一段,x <= bot_x, bot_x <= x <= top_x, top_x <= x
        if x <= x2:
            ret_standard = y1 + k1 * (x - x1)
            ret_lower = y1 + k1 * (x - precision_value_x_k1 - x1)
            ret_upper = y1 + k1 * (x + precision_value_x_k1 - x1)
            if (x <= bot_y_to_x) and (bot_y_to_x <= x2):
                ret_standard = y1 + k1 * (bot_y_to_x - x1)
                ret_lower = y1 + k1 * (bot_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (bot_y_to_x + precision_value_x_k1 - x1)
            elif (x <= bot_y_to_x) and (bot_y_to_x >= x2) and (bot_y_to_x <= x3):
                ret_standard = y2 + k2 * (bot_y_to_x - x2)
                ret_lower = y2 + k2 * (bot_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (bot_y_to_x + precision_value_x_k2 - x2)
            elif (x <= bot_y_to_x) and (bot_y_to_x >= x3):
                ret_standard = y3 + k3 * (bot_y_to_x - x3)
                ret_lower = y3 + k3 * (bot_y_to_x - precision_value_x_k3 - x3)
                ret_upper = y3 + k3 * (bot_y_to_x + precision_value_x_k3 - x3)
            elif x >= top_y_to_x:
                ret_standard = y1 + k1 * (top_y_to_x - x1)
                ret_lower = y1 + k1 * (top_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (top_y_to_x + precision_value_x_k1 - x1)
            else:
                pass
            return ret_standard, ret_lower, ret_upper, precision_value_y
        elif (x >= x2) and (x <= x3):
            ret_standard = y2 + k2 * (x - x2)
            ret_lower = y2 + k2 * (x - precision_value_x_k2 - x2)
            ret_upper = y2 + k2 * (x + precision_value_x_k2 - x2)
            if (x <= bot_y_to_x) and (bot_y_to_x <= x3):
                ret_standard = y2 + k2 * (bot_y_to_x - x2)
                ret_lower = y2 + k2 * (bot_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (bot_y_to_x + precision_value_x_k2 - x2)
            elif (x <= bot_y_to_x) and (bot_y_to_x >= x3):
                ret_standard = y3 + k3 * (bot_y_to_x - x3)
                ret_lower = y3 + k3 * (bot_y_to_x - precision_value_x_k3 - x3)
                ret_upper = y3 + k3 * (bot_y_to_x + precision_value_x_k3 - x3)
            elif (x >= top_y_to_x) and (x2 <= top_y_to_x):
                ret_standard = y2 + k2 * (top_y_to_x - x2)
                ret_lower = y2 + k2 * (top_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (top_y_to_x + precision_value_x_k2 - x2)
            elif (x >= top_y_to_x) and (x2 >= top_y_to_x):
                ret_standard = y1 + k1 * (top_y_to_x - x1)
                ret_lower = y1 + k1 * (top_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (top_y_to_x + precision_value_x_k1 - x1)
            else:
                pass
            return ret_standard, ret_lower, ret_upper, precision_value_y
        elif x >= x3:
            ret_standard = y3 + k3 * (x - x3)
            ret_lower = y3 + k3 * (x - precision_value_x_k3 - x3)
            ret_upper = y3 + k3 * (x + precision_value_x_k3 - x3)
            if x <= bot_y_to_x:
                ret_standard = y3 + k3 * (bot_y_to_x - x3)
                ret_lower = y3 + k3 * (bot_y_to_x - precision_value_x_k3 - x3)
                ret_upper = y3 + k3 * (bot_y_to_x + precision_value_x_k3 - x3)
            elif (x >= top_y_to_x) and (x3 <= top_y_to_x):
                ret_standard = y3 + k3 * (top_y_to_x - x3)
                ret_lower = y3 + k3 * (top_y_to_x - precision_value_x_k3 - x3)
                ret_upper = y3 + k3 * (top_y_to_x + precision_value_x_k3 - x3)
            elif (x >= top_y_to_x) and (x3 >= top_y_to_x) and (x2 <= top_y_to_x):
                ret_standard = y2 + k2 * (top_y_to_x - x2)
                ret_lower = y2 + k2 * (top_y_to_x - precision_value_x_k2 - x2)
                ret_upper = y2 + k2 * (top_y_to_x + precision_value_x_k2 - x2)
            elif (x >= top_y_to_x) and (x2 >= top_y_to_x):
                ret_standard = y1 + k1 * (top_y_to_x - x1)
                ret_lower = y1 + k1 * (top_y_to_x - precision_value_x_k1 - x1)
                ret_upper = y1 + k1 * (top_y_to_x + precision_value_x_k1 - x1)
            else:
                pass
            return ret_standard, ret_lower, ret_upper, precision_value_y
        else:
            pass


def calculator_ai():
    """
    通过AI算法计算出y轴值， 并回填到到excel表中
    参数:
    Voltage_or_current:
                      1: 仅计算电压
                      2：仅计算电流
                      3：电压电流均计算
    """
    comm_setting_data = data_read(iom_project_configfile_path + iom_project_configfile, 'comm_setting')
    input_list = data_read(iom_project_configfile_path + iom_project_configfile, 'set_ai_voltage_current_dc')
    input_list0 = input_list[0]
    input_list0.extend(['预期电压上限', '预期电压下限', '电压精度', '预期电流上限', '预期电流下限', '电流精度'])
    # 创建新工作簿
    wb = openpyxl.Workbook()
    # 删除默认创建的 Sheet（可选）
    default_sheet = wb.active
    wb.remove(default_sheet)
    # 添加一个新 Sheet，命名为“电流测量”
    wb.create_sheet("AI测量计算")
    # 获取活动工作表
    ws = wb.active

    for i in range(1, len(input_list0)+1):
        ws.cell(row=1, column=i, value=input_list0[i-1])
    # 获取AI寄存器中的值
    points = [(comm_setting_data[0][1], comm_setting_data[4][1]),
              (comm_setting_data[1][1], comm_setting_data[5][1]),
              (comm_setting_data[2][1], comm_setting_data[6][1]),
              (comm_setting_data[3][1], comm_setting_data[7][1])]
    limit_list = [comm_setting_data[8][1], comm_setting_data[9][1]]
    type_input = comm_setting_data[10][1]
    line_number = comm_setting_data[11][1]

    # 计算AI电压或电流
    if (type_input == 1) or (type_input == 2):
        for j in range(1, len(input_list)):
            voltage_standard = input_list[j][1]
            # print("数据： ", voltage_standard, points, limit_list, type_input, line_number)
            ret = get_ai_y_from_x(voltage_standard, points, limit_list, type_input, line_number)
            lower_value = ret[1]
            upper_value = ret[2]
            precision_value = ret[3]
            ws.cell(row=j+1, column=1, value=input_list[j][0])
            ws.cell(row=j+1, column=2, value=input_list[j][1])
            ws.cell(row=j+1, column=3, value=input_list[j][2])
            ws.cell(row=j+1, column=4, value=lower_value)
            ws.cell(row=j+1, column=5, value=upper_value)
            ws.cell(row=j+1, column=6, value=precision_value)
    elif (type_input == 3) or (type_input == 4):
        #  计算电流的AI值
        for k in range(1, len(input_list)):
            voltage_standard = input_list[k][2]
            ret = get_ai_y_from_x(voltage_standard, points, limit_list, type_input, line_number)
            lower_value = ret[1]
            upper_value = ret[2]
            precision_value = ret[3]
            ws.cell(row=k+1, column=1, value=input_list[k][0])
            ws.cell(row=k+1, column=2, value=input_list[k][1])
            ws.cell(row=k+1, column=3, value=input_list[k][2])
            ws.cell(row=k+1, column=7, value=lower_value)
            ws.cell(row=k+1, column=8, value=upper_value)
            ws.cell(row=k+1, column=9, value=precision_value)
    wb.save(base_dir + '/output_data/' + f"AI_voltage_{time.strftime('%Y%m%d%H%M%S')}.xlsx")


def calculator_ao():
    """
    通过AO算法计算出y轴值， 并回填到到excel表中
    参数:
    Voltage_or_current:
                      1: 仅计算电压
                      2：仅计算电流
                      3：电压电流均计算
    """
    comm_setting_data = data_read(iom_project_configfile_path + iom_project_configfile, 'comm_setting')
    input_list = data_read(iom_project_configfile_path + iom_project_configfile, 'set_ao_voltage_current_dc')
    input_list0 = input_list[0]
    input_list0.extend(['预期电压上限', '预期电压下限', '电压精度', '预期电流上限', '预期电流下限', '电流精度'])
    # 创建新工作簿
    wb = openpyxl.Workbook()
    # 删除默认创建的 Sheet（可选）
    default_sheet = wb.active
    wb.remove(default_sheet)
    # 添加一个新 Sheet，命名为“电流测量”
    wb.create_sheet("AO测量计算")
    # 获取活动工作表
    ws = wb.active
    for i in range(1, len(input_list0)+1):
        ws.cell(row=1, column=i, value=input_list0[i-1])

    #  获取comm_setting_data中值
    points = [(comm_setting_data[0][1], comm_setting_data[4][1]),
              (comm_setting_data[1][1], comm_setting_data[5][1]),
              (comm_setting_data[2][1], comm_setting_data[6][1]),
              (comm_setting_data[3][1], comm_setting_data[7][1])]
    limit_list = [comm_setting_data[8][1], comm_setting_data[9][1]]
    type_input = comm_setting_data[10][1]
    line_number = comm_setting_data[11][1]
    #  计算电压的AO值
    if (type_input == 1) or (type_input == 2):
        for i in range(1, len(input_list)):
            voltage_standard = input_list[i][1]
            ret = get_ao_y_from_x(voltage_standard, points, limit_list, type_input, line_number)
            lower_value = ret[1]
            upper_value = ret[2]
            precision_value = ret[3]
            ws.cell(row=i+1, column=1, value=input_list[i][0])
            ws.cell(row=i+1, column=2, value=input_list[i][1])
            ws.cell(row=i+1, column=3, value=input_list[i][2])
            ws.cell(row=i+1, column=4, value=lower_value)
            ws.cell(row=i+1, column=5, value=upper_value)
            ws.cell(row=i+1, column=6, value=precision_value)
    #  计算电流的AO值
    elif (type_input == 3) or (type_input == 4):
        for i in range(1, len(input_list)):
            voltage_standard = input_list[i][2]
            ret = get_ai_y_from_x(voltage_standard, points, limit_list, type_input, line_number)
            lower_value = ret[1]
            upper_value = ret[2]
            precision_value = ret[3]
            ws.cell(row=i+1, column=1, value=input_list[i][0])
            ws.cell(row=i+1, column=2, value=input_list[i][1])
            ws.cell(row=i+1, column=3, value=input_list[i][2])
            ws.cell(row=i+1, column=7, value=lower_value)
            ws.cell(row=i+1, column=8, value=upper_value)
            ws.cell(row=i+1, column=9, value=precision_value)
    wb.save(base_dir + '/output_data/' + f"AO_voltage_{time.strftime('%Y%m%d%H%M%S')}.xlsx")


def measuring_ai():
    comm_setting_data = data_read(r'/comm/test_data/IOM_input_parameters.xlsx', 'comm_setting')
    input_list = data_read(r'/comm/test_data/IOM_input_parameters.xlsx', 'set_ai_voltage_current_dc')
    input_list0 = input_list[0]
    input_list0.extend(['预期电压上限', '预期电压下限', '精度', '实际电压', '实际测试结果'])
    # 创建新工作簿
    wb = openpyxl.Workbook()
    wb.create_sheet("AI电压测量")
    # 获取活动工作表
    ws = wb.active

    for i in range(len(input_list0)):
        j = i + 1
        ws.cell(row=1, column=j, value=input_list[0][i])

    wb.save('AI_voltage_{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S')))

    """"AI 测量"""
    #  通过DC源输入电压
    current_input = 0.00
    for i in range(16):
        voltage_input = input_list0[1]
        set_output_dc(1, voltage_input, current_input)

    # 获取AI寄存器中的值
    for i in range(len(input_list)):
        voltage_standard = input_list[i][1]
        ret_register = read_iom_ai_voltage_current(voltage_standard, 1, 10)
        points = [(comm_setting_data[0][1], comm_setting_data[4][1]),
                  (comm_setting_data[1][1], comm_setting_data[5][1]),
                  (comm_setting_data[2][1], comm_setting_data[6][1]),
                  (comm_setting_data[3][1], comm_setting_data[7][1])]
        limit_list = [comm_setting_data[8][1], comm_setting_data[9][1]]
        ret = get_ai_y_from_x(voltage_standard, points, limit_list, 0, 1)
        lower_value = ret[1]
        upper_value = ret[2]
        precision_value = ret[3]
        ws.cell(row=i, column=4, value=lower_value)
        ws.cell(row=i, column=5, value=upper_value)
        ws.cell(row=i, column=6, value=precision_value)
        ws.cell(row=i, column=7, value=ret_register)


if __name__ == '__main__':
    calculator_ao()
    calculator_ai()
    print('====================AI Measure Start====================')
    print('======================{}======================'.format(time.strftime('%Y_%m_%d %H:%M:%S')))
    start_time = time.time()

    #  通过DC源输入电流

    # 获取AI寄存器中的值

    # 把获取的值和实际计算的值进行比较，把比较的结果写到寄存器中
    """"AO 测量"""
    #  通过DC源输入电压

    # 获取AI寄存器中的值

    # 把获取的值和实际计算的值进行比较，把比较的结果写到寄存器中

    #  通过DC源输入电流

    # 获取AI寄存器中的值

    #  把获取的值和实际计算的值进行比较，把比较的结果写到寄存器中

