"""
测点文件读取模块
从 Excel 测点文件读取测试参数列表
"""
import openpyxl
import logging

# 测点 Excel 列索引（0-based），与原 acuvimseries_test_case.xlsx 格式一致
COL_CASE_ID      = 0
COL_UA           = 1
COL_UB           = 2
COL_UC           = 3
COL_IA           = 4
COL_IB           = 5
COL_IC           = 6
COL_UA_P         = 7
COL_UB_P         = 8
COL_UC_P         = 9
COL_IA_P         = 10
COL_IB_P         = 11
COL_IC_P         = 12
COL_FREQ         = 16
COL_V_ACC        = 17
COL_I_ACC        = 18
COL_ANGLE_ACC    = 19
COL_P_ACC        = 20
COL_Q_ACC        = 21
COL_S_ACC        = 22
COL_SAMPLE_CNT   = 23
COL_SAMPLE_INT   = 24
COL_WIRE_TYPE    = 15   # 接线方式字符串，如 "3E4wY"

# 接线方式字符串 → 内部 wire_type 整数
WIRE_STR_MAP = {
    "1E2w1p":  0,
    "2E3w1p":  1,
    "2E3wD":   2,
    "2E3wN":   3,
    "3E4wY":   4,
    "3E4wD":   5,
    "3E4wY-P": 6,
}


def load_testpoints(excel_path: str, sheet_name: str, wire_type_int: int) -> list[dict]:
    """
    读取测点 Excel，筛选指定接线方式的测试行
    返回 list[dict]，每个 dict 是一个测试点的参数
    """
    # 找到 wire_type_int 对应的字符串标签
    target_labels = [k for k, v in WIRE_STR_MAP.items() if v == wire_type_int]

    points = []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            logging.error(f"Sheet '{sheet_name}' not found in {excel_path}")
            return []
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[COL_CASE_ID] is None:
                continue
            wire_label = str(row[COL_WIRE_TYPE] or '').strip()
            if wire_label not in target_labels:
                continue
            try:
                pt = {
                    "case_id":      str(row[COL_CASE_ID]),
                    "ua":           float(row[COL_UA] or 0),
                    "ub":           float(row[COL_UB] or 0),
                    "uc":           float(row[COL_UC] or 0),
                    "ia":           float(row[COL_IA] or 0),
                    "ib":           float(row[COL_IB] or 0),
                    "ic":           float(row[COL_IC] or 0),
                    "ua_p":         float(row[COL_UA_P] or 0),
                    "ub_p":         float(row[COL_UB_P] or 0),
                    "uc_p":         float(row[COL_UC_P] or 0),
                    "ia_p":         float(row[COL_IA_P] or 0),
                    "ib_p":         float(row[COL_IB_P] or 0),
                    "ic_p":         float(row[COL_IC_P] or 0),
                    "freq":         float(row[COL_FREQ] or 50),
                    "v_acc":        float(row[COL_V_ACC] or 0.001),
                    "i_acc":        float(row[COL_I_ACC] or 0.001),
                    "angle_acc":    float(row[COL_ANGLE_ACC] or 0.1),
                    "p_acc":        float(row[COL_P_ACC] or 0.002),
                    "sample_cnt":   int(row[COL_SAMPLE_CNT] or 20),
                    # Excel 中 sample_interval 列存的是秒（如0.2s），需转换为毫秒
                    "sample_int_ms":int(float(row[COL_SAMPLE_INT] or 0.2) * 1000),
                    "wire_label":   wire_label,
                }
                points.append(pt)
            except (TypeError, ValueError) as e:
                logging.warning(f"Skip row {row[COL_CASE_ID]}: {e}")
    except Exception as e:
        logging.error(f"Failed to load testpoints: {e}")

    return points
