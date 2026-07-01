"""
Modbus 地址表加载器
从标准 Modbus Address Table Excel 中解析所需地址
也提供内置默认地址（AcuRev1320），无需 Excel 即可运行
"""
import openpyxl
import logging

# ── 内置默认地址（AcuRev1320）────────────────────────────────
DEFAULT_ADDRS = {
    # 测量量（200ms 刷新，float32，每量占 2 个寄存器）
    "freq":       0x9307,
    "ua":         0x9309,
    "ub":         0x930B,
    "uc":         0x930D,
    "ua_angle":   0x930F,
    "ub_angle":   0x9311,
    "uc_angle":   0x9313,
    "uab":        0x9317,
    "ubc":        0x9319,
    "uca":        0x931B,
    "uab_angle":  0x931D,
    "ubc_angle":  0x931F,
    "uca_angle":  0x9321,
    "ia":         0x9325,
    "ib":         0x9327,
    "ic":         0x9329,
    "in":         0x932B,
    "ia_angle":   0x932D,
    "ib_angle":   0x932F,
    "ic_angle":   0x9331,
    "pa":         0x9337,
    "pb":         0x9339,
    "pc":         0x933B,
    "p_sys":      0x933D,
    "qa":         0x933F,
    "qb":         0x9341,
    "qc":         0x9343,
    "q_sys":      0x9345,
    "sa":         0x9347,
    "sb":         0x9349,
    "sc":         0x934B,
    "s_sys":      0x934D,
    "pf_a":       0x934F,
    "pf_b":       0x9351,
    "pf_c":       0x9353,
    "pf_sys":     0x9355,
    # 配置寄存器（uint16，单寄存器）
    "freq_set":   0x1041,   # 频率设置：0=50Hz，1=60Hz
    "wire_mode":  0x1042,   # 接线方式
    "ct_ch1":     0x1049,   # 通道1 CT类型
    "ct_ch2":     0x104D,   # 通道2 CT类型
    "ct_ch3":     0x1051,   # 通道3 CT类型
    "reboot":     0x1147,   # 软重启：写 0x0001
}

# 接线方式名称 → 寄存器值
WIRE_MODE_MAP = {
    "1E2W1P":  0,
    "2E3W1P":  1,
    "2E3WD":   2,
    "2E3WN":   3,
    "3E4WY":   4,
    "3E3WD":   5,
    "3E4WY-P": 4,   # 与 3E4WY 共用同一寄存器值
}

# CT 类型名称 → 寄存器值
CT_TYPE_MAP = {
    "100mA": 0,
    "333mV": 2,
    "RCT":   3,
}


def load_from_excel(excel_path: str) -> dict:
    """
    从标准 Modbus Address Table Excel 解析地址
    规则：
      - 遍历 '200ms' 和 'Basic Setting' 两个 Sheet
      - 找到 Description 列匹配关键词的行，取 Start(Hex) 列的值
    如果解析失败，返回内置默认地址
    """
    # 关键词 → 内部名称
    keyword_map = {
        "System Frequency":                    "freq",
        "Phase A Line-to-Neutral Voltage":     "ua",
        "Phase B Line-to-Neutral Voltage":     "ub",
        "Phase C Line-to-Neutral Voltage":     "uc",
        "Phase A Line-to-Neutral Voltage Phase Angle": "ua_angle",
        "Phase B Line-to-Neutral Voltage Phase Angle": "ub_angle",
        "Phase C Line-to-Neutral Voltage Phase Angle": "uc_angle",
        "Phase AB Line-to-Line Voltage":       "uab",
        "Phase BC Line-to-Line Voltage":       "ubc",
        "Phase CA Line-to-Line Voltage":       "uca",
        "Phase AB Line-to-Line Voltage Phase Angle": "uab_angle",
        "Phase BC Line-to-Line Voltage Phase Angle": "ubc_angle",
        "Phase CA Line-to-Line Voltage Phase Angle": "uca_angle",
        "Phase A Current":                     "ia",
        "Phase B Current":                     "ib",
        "Phase C Current":                     "ic",
        "Neutral Current":                     "in",
        "Phase A Current Phase Angle":         "ia_angle",
        "Phase B Current Phase Angle":         "ib_angle",
        "Phase C Current Phase Angle":         "ic_angle",
        "Phase A Active Power":                "pa",
        "Phase B Active Power":                "pb",
        "Phase C Active Power":                "pc",
        "System Active Power":                 "p_sys",
        "Phase A Reactive Power":              "qa",
        "Phase B Reactive Power":              "qb",
        "Phase C Reactive Power":              "qc",
        "System Reactive Power":               "q_sys",
        "Phase A Apparent Power":              "sa",
        "Phase B Apparent Power":              "sb",
        "Phase C Apparent Power":              "sc",
        "System Apparent Power":               "s_sys",
        "Phase A Power Factor":                "pf_a",
        "Phase B Power Factor":                "pf_b",
        "Phase C Power Factor":                "pf_c",
        "System Power Factor":                 "pf_sys",
        "Frequency Selection":                 "freq_set",
        "Service Configuration":               "wire_mode",
        "Device Reboot":                       "reboot",
    }
    # CT 通道关键词单独处理（多条相似描述）
    ct_keywords = [
        ("channel 1 input CT Type", "ct_ch1"),
        ("channel 2 input CT Type", "ct_ch2"),
        ("channel 3 input CT Type", "ct_ch3"),
    ]

    result = dict(DEFAULT_ADDRS)  # 以默认值为基础，覆盖找到的

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        target_sheets = [s for s in wb.sheetnames if s in ('200ms', 'Basic Setting')]

        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            # 找列索引
            header = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            try:
                desc_col = header.index('Description')
                addr_col = header.index('Start(Hex)')
            except ValueError:
                logging.warning(f"Sheet '{sheet_name}' missing expected columns")
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                desc = str(row[desc_col] or '').strip()
                addr_raw = row[addr_col]
                if not desc or addr_raw is None:
                    continue
                try:
                    addr = int(str(addr_raw), 16) if isinstance(addr_raw, str) else int(addr_raw)
                except (ValueError, TypeError):
                    continue

                # 精确匹配
                for keyword, name in keyword_map.items():
                    if keyword.lower() == desc.lower():
                        result[name] = addr
                        break

                # CT 通道（前缀匹配）
                for keyword, name in ct_keywords:
                    if keyword.lower() in desc.lower():
                        result[name] = addr
                        break

        logging.info(f"Loaded {len(result)} addresses from {excel_path}")
    except Exception as e:
        logging.error(f"Failed to load address table from Excel: {e}，使用默认地址")

    return result


def get_measure_addr_map(addrs: dict, wire_type: str) -> dict:
    """
    根据接线方式返回需要采样的寄存器字典
    只返回该接线方式有效的测量量
    """
    all_meas = {
        "freq":      addrs["freq"],
        "ua":        addrs["ua"],
        "ub":        addrs["ub"],
        "uc":        addrs["uc"],
        "ua_angle":  addrs["ua_angle"],
        "ub_angle":  addrs["ub_angle"],
        "uc_angle":  addrs["uc_angle"],
        "uab":       addrs["uab"],
        "ubc":       addrs["ubc"],
        "uca":       addrs["uca"],
        "uab_angle": addrs["uab_angle"],
        "ubc_angle": addrs["ubc_angle"],
        "uca_angle": addrs["uca_angle"],
        "ia":        addrs["ia"],
        "ib":        addrs["ib"],
        "ic":        addrs["ic"],
        "in":        addrs["in"],
        "ia_angle":  addrs["ia_angle"],
        "ib_angle":  addrs["ib_angle"],
        "ic_angle":  addrs["ic_angle"],
        "pa":        addrs["pa"],
        "pb":        addrs["pb"],
        "pc":        addrs["pc"],
        "p_sys":     addrs["p_sys"],
        "qa":        addrs["qa"],
        "qb":        addrs["qb"],
        "qc":        addrs["qc"],
        "q_sys":     addrs["q_sys"],
        "sa":        addrs["sa"],
        "sb":        addrs["sb"],
        "sc":        addrs["sc"],
        "s_sys":     addrs["s_sys"],
        "pf_sys":    addrs["pf_sys"],
    }

    # 实时监控始终读全部；采样时按接线方式过滤（test_engine 中处理）
    return all_meas
