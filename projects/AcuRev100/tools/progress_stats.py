# -*- coding: utf-8 -*-
"""AcuRev-100 手工用例→自动化 转化进度统计。

从手工用例 xlsx（全量分母）与 tests/<模块目录>/ 已生成 pytest 文件统计进度，
生成 projects/AcuRev100/PROGRESS.md。统计口径见生成文件头部说明。

用法（仓库根目录执行）: python projects/AcuRev100/tools/progress_stats.py
"""
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_DIR = Path(__file__).resolve().parents[1]
XLSX_PATH = PROJECT_DIR / "tests" / "manual_testcase" / "AcuRev-100测试用例_已回填自动化覆盖情况.xlsx"
# 自动化用例目录：tests/ 全树递归（各模块目录）
AUTO_DIRS = (PROJECT_DIR / "tests",)
OUT_PATH = PROJECT_DIR / "PROGRESS.md"

# 模块号 → 模块名（xlsx 模块列存在合并单元格，前向填充不可靠，以基线文档为准）
MODULE_NAMES = {
    "001": "交流频率", "002": "相电压", "003": "线电压", "004": "电流测量",
    "005": "PF·功率", "006": "相角", "007": "电能累计", "008": "脉冲",
    "009": "接线验证", "010": "接线检查", "011": "计时功能",
    "012": "LCD显示及LED指示灯", "013": "铅封", "014": "Firmware升级",
    "015": "数据重置及恢复出厂设置", "016": "通讯模块", "017": "密码",
}
# 本期自动化范围（VERIFIER_BASELINE.md 第三节；011 于 2026-07-14 按用户口头指令补入并已生成 7 条）
IN_SCOPE = {"001", "002", "003", "004", "005", "006", "007", "008", "009", "011", "014", "015", "016", "017"}

CASE_ID_RE = re.compile(r"_(\d{3})_(\d{2})_case(\d+(?:_\d+)*?)(?:\.py)?$")


def norm_case_key(text: str) -> str | None:
    """从用例编号或文件名提取归一化键 <模块>_<子模块>_case<N>（保留拆分后缀，如 case7_01）。"""
    m = CASE_ID_RE.search(text)
    return f"{m.group(1)}_{m.group(2)}_case{m.group(3)}" if m else None


def load_manual_cases() -> tuple[list[dict], list[str]]:
    """返回 (用例列表, 编号不合规范的原始编号列表)。编号异常的用例仍计入所属模块分母。"""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    # 数据页按名取（首页现为「进展统计」汇总页）
    ws = wb["AcuRev-100测试用例"] if "AcuRev-100测试用例" in wb.sheetnames else wb[wb.sheetnames[0]]
    cases, bad_ids = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        case_id = row[2]
        if not case_id:
            continue
        raw = str(case_id).strip()
        key = norm_case_key(raw)
        if key is None:
            bad_ids.append(raw)
            loose = re.search(r"_(\d{3})_", raw)
            key, module = raw, (loose.group(1) if loose else "???")
        else:
            module = key[:3]
        cases.append({
            "key": key,
            "module": module,
            "automatable": str(row[8]).strip() if len(row) > 8 and row[8] else "",
            "debug": str(row[9]).strip() if len(row) > 9 and row[9] else "",
            "result": str(row[10]).strip() if len(row) > 10 and row[10] else "",
            "remark": str(row[11]).strip() if len(row) > 11 and row[11] else "",
        })
    wb.close()
    return cases, bad_ids


def load_generated_keys() -> set[str]:
    keys = set()
    for auto_dir in AUTO_DIRS:
        for py in auto_dir.rglob("test_acuview_*.py"):
            key = norm_case_key(py.name)
            if key:
                keys.add(key)
    return keys


def debug_status(text: str) -> str:
    """调试结果列非空时归类。

    回填约定（2026-07-14 用户裁决）：首行去掉 [日期/标签] 前缀后以「通过」或「已调试」开头
    且不含失败字样 → passed；其余非空内容一律计 pending（待调试/待重测/转手工/疑似缺陷等），
    避免正文中出现的"验证通过后"等字样造成误判。
    """
    if not text:
        return "none"
    first = re.sub(r"^(?:\s*\[[^\]]*\])*\s*", "", text.splitlines()[0])
    if "不通过" in first or "失败" in first:
        return "pending"
    return "passed" if first.startswith(("通过", "已调试")) else "pending"


# 状态/原因归类（依据 xlsx『调试结果』『测试结果』『备注』三列的回填约定, 2026-07-14）。
# 顺序即优先级；未生成子原因按备注关键词匹配（关键词有序, 先匹配先归类）。
CAT_PASS_REAL = "已跑通-实测通过"
CAT_PASS_ADC = "已跑通-仅剩读数为0（已知ADC损坏, 换板复跑出有效判定）"
_UNGEN_RULES = [
    ("脉冲采集台架", "未生成-需脉冲采集台架（008, 规划accuracy_pulse引擎）"),
    ("本期转换范围外", "未生成-本期范围外（010, 已另备判据级用例集59条）"),
    ("铅封", "未生成-铅封需人工拨码配合（待人工配合场次）"),
    ("升级", "未生成-Acuview升级GUI驱动未实现（014）"),
    ("boot界面", "未生成-Acuview升级GUI驱动未实现（014）"),
    ("评审", "未生成-2026-07-14新增/重写, 待评审转换"),
    ("GUI", "未生成-GUI write_verify引擎未就绪（015/016批, 含密码/通讯/复位）"),
    ("015/016", "未生成-GUI write_verify引擎未就绪（015/016批, 含密码/通讯/复位）"),
]


def categorize(c: dict, generated: set[str]) -> str:
    """单条用例 → 状态/原因类别（供模块备注与全量分布两处复用）。"""
    remark, debug, result = c["remark"], c["debug"], c["result"]
    if debug_status(debug) == "passed":
        return CAT_PASS_ADC if "未通过" in result else CAT_PASS_REAL
    if remark.startswith("待重测") or result.startswith("待重测"):
        return "待重测-07-08报告证据链断裂（复跑补证据即可）"
    if remark.startswith("转手工") or (debug and "转手工" in debug.splitlines()[0]):
        return "转手工"
    if c["automatable"] == "否":
        return "不可自动化（标否, 判否原因见xlsx备注列）"
    if remark.startswith("待调试") or "待调试" in debug:
        if "CT回路" in remark or "333mV" in remark:
            return "待调试-未接5A/333mV CT回路"
        if "RCT" in remark or "80mA" in remark:
            return "待调试-RCT/80mA无CT枚举, 待固件确认"
        return "待调试-其他"
    if c["key"] not in generated:
        for kw, name in _UNGEN_RULES:
            if kw in remark:
                return name
        return "未生成-其他"
    return "已生成-未回填"


# 模块备注里用的短标签
_SHORT = {
    CAT_PASS_ADC: "仅剩读数为0(ADC)",
    "待重测-07-08报告证据链断裂（复跑补证据即可）": "待重测(证据链)",
    "待调试-未接5A/333mV CT回路": "待调试:未接mV CT回路",
    "待调试-RCT/80mA无CT枚举, 待固件确认": "待调试:RCT/80mA待固件",
    "待调试-其他": "待调试:其他",
    "转手工": "转手工",
    "不可自动化（标否, 判否原因见xlsx备注列）": "标否不可自动化",
    "未生成-需脉冲采集台架（008, 规划accuracy_pulse引擎）": "未生成:待脉冲台架",
    "未生成-本期范围外（010, 已另备判据级用例集59条）": "未生成:本期范围外",
    "未生成-铅封需人工拨码配合（待人工配合场次）": "未生成:待铅封拨码配合",
    "未生成-Acuview升级GUI驱动未实现（014）": "未生成:待升级GUI驱动",
    "未生成-2026-07-14新增/重写, 待评审转换": "未生成:待评审转换",
    "未生成-GUI write_verify引擎未就绪（015/016批, 含密码/通讯/复位）": "未生成:待GUI引擎",
    "未生成-其他": "未生成:其他",
    "已生成-未回填": "已生成未回填",
}


def build_report() -> str:
    cases, bad_ids = load_manual_cases()
    generated = load_generated_keys()

    stats = defaultdict(lambda: defaultdict(int))
    mod_cats: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    cat_mods: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for c in cases:
        s = stats[c["module"]]
        s["total"] += 1
        if c["key"] in generated:
            s["generated"] += 1
        cat = categorize(c, generated)
        if cat == CAT_PASS_REAL or cat == CAT_PASS_ADC:
            s["debug_passed"] += 1
        if cat == CAT_PASS_REAL:
            s["pass_real"] += 1
        elif cat == CAT_PASS_ADC:
            s["pass_adc"] += 1
        mod_cats[c["module"]][cat] += 1
        cat_mods[cat][c["module"]] += 1

    orphan = sorted(k for k in generated if not any(c["key"] == k for c in cases))

    def pct(n: int, d: int) -> str:
        return f"{n / d * 100:.0f}%" if d else "—"

    lines = [
        "# AcuRev-100 自动化用例转化进度",
        "",
        f"> 更新时间：{datetime.now():%Y-%m-%d %H:%M}　·　由 `tools/progress_stats.py` 自动生成，勿手改",
        ">",
        "> **口径**：分母=手工用例全量；「已生成」=`tests/<模块目录>/`存在与用例编号逐字符对应的 `test_*.py`；",
        "> 「调试通过」=**自动化流程已实跑跑通**（『调试结果』列回填「已调试/通过」）——其中多数当前实测仍被",
        "> 已知板卡 ADC 损坏（判据读数为 0）阻塞，属预期问题，换板复跑即可出有效判定（2026-07-14 口径澄清）；",
        "> 其余未调试/未生成用例的原因见下方「状态与原因分布」及 xlsx『备注』列。",
        "",
    ]

    total = {"total": 0, "generated": 0, "debug_passed": 0, "pass_real": 0, "pass_adc": 0}
    rows = []
    for mod in sorted(stats):
        s = stats[mod]
        for k in total:
            total[k] += s[k]
        remark = []
        if mod not in IN_SCOPE:
            remark.append("不在本期自动化范围")
        remark += [f"{_SHORT.get(cat, cat)}×{n}"
                   for cat, n in sorted(mod_cats[mod].items(), key=lambda kv: -kv[1])
                   if cat != CAT_PASS_REAL]
        rows.append(
            f"| {mod} | {MODULE_NAMES.get(mod, '?')} | {s['total']} | {s['generated']} | "
            f"{pct(s['generated'], s['total'])} | {s['debug_passed']} | "
            f"{pct(s['debug_passed'], s['total'])} | {'；'.join(remark)} |"
        )

    lines += [
        f"**总览**：手工用例全量 **{total['total']}** 条｜已生成 **{total['generated']} 条"
        f"（{pct(total['generated'], total['total'])}）**｜调试通过（已跑通）**{total['debug_passed']} 条"
        f"（{pct(total['debug_passed'], total['total'])}）**——其中实测通过 {total['pass_real']} 条、"
        f"仅剩读数为0待换板复跑 {total['pass_adc']} 条",
        "",
        "| 模块 | 模块名 | 手工用例数 | 已生成 | 生成进度 | 调试通过 | 调试进度 | 备注（原因×条数） |",
        "|:---:|---|:---:|:---:|:---:|:---:|:---:|---|",
        *rows,
        f"| — | **合计** | **{total['total']}** | **{total['generated']}** | "
        f"**{pct(total['generated'], total['total'])}** | **{total['debug_passed']}** | "
        f"**{pct(total['debug_passed'], total['total'])}** | |",
        "",
        "## 状态与原因分布（全量）",
        "",
        "| 状态-原因 | 条数 | 模块分布 |",
        "|---|:---:|---|",
    ]
    for cat, mods in sorted(cat_mods.items(), key=lambda kv: -sum(kv[1].values())):
        dist = "、".join(f"{m}×{n}" for m, n in sorted(mods.items()))
        lines.append(f"| {cat} | {sum(mods.values())} | {dist} |")
    lines += [
        "",
        "> 逐条原因明细见手工用例 xlsx 的『备注』列（2026-07-14 已按 未生成/待调试/未通过/转手工/不可自动化 逐行标注）。",
    ]

    if orphan or bad_ids:
        lines += ["", "**⚠ 编号异常**（按全局约定 12 需整理清单交负责人修正）："]
        lines += [f"- 已生成文件在用例表中无对应编号：`{k}`" for k in orphan]
        lines += [f"- 用例表编号不合命名规范：`{k}`" for k in bad_ids]

    lines.append("")
    return "\n".join(lines)


def main() -> None:
    report = build_report()
    OUT_PATH.write_text(report, encoding="utf-8")
    print(f"written: {OUT_PATH}")


if __name__ == "__main__":
    main()
