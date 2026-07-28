"""测点文件读取模块（AcuRev-100 / ACmeter 口径）。

测点 Excel 列格式见 test_point/gen_acurev100_testpoints.py 与 README「测试点文件格式」。
两点与 1320 的老格式不同，改测点表时务必注意：
  1. 电流列 4~6 是**源侧输出电流(A)**，电表期望读数 = 源电流 × CT Primary ÷ 台体CT额定
     （台体CT额定见 config.yaml source.bench_ct_a）→ 新增列 25 = CT Primary(A)
  2. 阈值列 17~22 必须与 config.yaml accuracy 段一致（load 时用 check_thresholds 复核，
     确需放宽的点在列 26「备注」写"宽口径"标记，复核时跳过）
"""
import openpyxl
import logging

# 测点 Excel 列索引（0-based）
COL_CASE_ID      = 0
COL_UA           = 1
COL_UB           = 2
COL_UC           = 3
COL_IA           = 4     # 源侧输出电流（非电表读数）
COL_IB           = 5
COL_IC           = 6
COL_UA_P         = 7
COL_UB_P         = 8
COL_UC_P         = 9
COL_IA_P         = 10
COL_IB_P         = 11
COL_IC_P         = 12
COL_WAIT_H       = 13    # 保留列（能量类用例的等待时长，精度测试恒 0）
COL_PF           = 14    # 参考列（不参与判定，仅便于人读）
COL_WIRE_TYPE    = 15    # 接线方式字符串："1E2W" / "2E3W1P" / "3E4WY"
COL_FREQ         = 16
COL_V_ACC        = 17
COL_I_ACC        = 18
COL_ANGLE_ACC    = 19    # 绝对量（°）
COL_P_ACC        = 20
COL_Q_ACC        = 21
COL_S_ACC        = 22
COL_SAMPLE_CNT   = 23
COL_SAMPLE_INT   = 24    # 秒
COL_CT_PRIMARY   = 25    # CT 一次侧额定电流（A），期望值换算用
COL_REMARK       = 26    # 备注；含"宽口径"= 该行阈值有意偏离 config，不参与一致性复核

# 接线方式标签别名 → 规范名（兼容 1320 老表的 1E2w1p / 2E3w1p / 3E4wY 写法）
WIRE_LABEL_ALIAS = {
    "1e2w":    "1E2W",
    "1e2w1p":  "1E2W",
    "2e3w1p":  "2E3W1P",
    "3e4wy":   "3E4WY",
}
# AcuRev-100 支持的三种接线（与 addr_loader.WIRE_MODE_MAP 一致）
WIRE_TYPES = ("1E2W", "2E3W1P", "3E4WY")


def normalize_wire_label(label: str) -> str:
    """测点表接线标签 → 规范名；无法识别时原样返回（调用方按"不匹配"处理）。"""
    return WIRE_LABEL_ALIAS.get(str(label or "").strip().lower(), str(label or "").strip())


def load_testpoints(excel_path: str, sheet_name: str, wire_type: str) -> list[dict]:
    """读测点 Excel，筛出指定接线方式（规范名，如 "3E4WY"）的测试行。

    ⚠️ 历史坑：老版本这里收的是"寄存器值 int"，而寄存器值 2(3E4WY) 会被老映射表解成
    1320 的 "2E3wD" 标签 → 3E4WY 恒选不到测点、整组被静默跳过。现改为按接线名直筛。
    """
    target = normalize_wire_label(wire_type)
    points: list[dict] = []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        logging.error(f"测点文件打不开: {e}")
        return []
    try:
        if sheet_name not in wb.sheetnames:
            logging.error(f"Sheet '{sheet_name}' not found in {excel_path}")
            return []
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= COL_FREQ or row[COL_CASE_ID] is None:
                continue
            if normalize_wire_label(row[COL_WIRE_TYPE]) != target:
                continue

            def _cell(idx: int, default=None):
                return row[idx] if len(row) > idx and row[idx] is not None else default

            try:
                pt = {
                    "case_id":      str(row[COL_CASE_ID]),
                    "ua":           float(_cell(COL_UA, 0) or 0),
                    "ub":           float(_cell(COL_UB, 0) or 0),
                    "uc":           float(_cell(COL_UC, 0) or 0),
                    # 源侧输出电流；电表期望电流由 test_engine 按 ct_primary 换算
                    "ia":           float(_cell(COL_IA, 0) or 0),
                    "ib":           float(_cell(COL_IB, 0) or 0),
                    "ic":           float(_cell(COL_IC, 0) or 0),
                    "ua_p":         float(_cell(COL_UA_P, 0) or 0),
                    "ub_p":         float(_cell(COL_UB_P, 0) or 0),
                    "uc_p":         float(_cell(COL_UC_P, 0) or 0),
                    "ia_p":         float(_cell(COL_IA_P, 0) or 0),
                    "ib_p":         float(_cell(COL_IB_P, 0) or 0),
                    "ic_p":         float(_cell(COL_IC_P, 0) or 0),
                    "pf":           _cell(COL_PF),
                    "freq":         float(_cell(COL_FREQ, 50)),
                    "v_acc":        float(_cell(COL_V_ACC, 0.002)),
                    "i_acc":        float(_cell(COL_I_ACC, 0.002)),
                    "angle_acc":    float(_cell(COL_ANGLE_ACC, 0.5)),
                    "p_acc":        float(_cell(COL_P_ACC, 0.002)),
                    "q_acc":        float(_cell(COL_Q_ACC, 0.005)),
                    "s_acc":        float(_cell(COL_S_ACC, 0.005)),
                    "sample_cnt":   int(_cell(COL_SAMPLE_CNT, 10)),
                    # Excel 中 sample_interval 列存的是秒（如 0.2s），转换为毫秒
                    "sample_int_ms": int(float(_cell(COL_SAMPLE_INT, 0.2)) * 1000),
                    "ct_primary":   float(_cell(COL_CT_PRIMARY, 200)),
                    "remark":       str(_cell(COL_REMARK, "") or ""),
                    "wire_label":   target,
                }
                points.append(pt)
            except (TypeError, ValueError) as e:
                logging.warning(f"Skip row {row[COL_CASE_ID]}: {e}")
    except Exception as e:
        logging.error(f"Failed to load testpoints: {e}")
    finally:
        wb.close()

    return points


def check_thresholds(points: list[dict], cfg_thr: dict) -> list[str]:
    """复核测点表阈值列与 config.yaml accuracy 段是否一致。

    返回不一致描述列表（空=一致）。备注含"宽口径"的行有意偏离，跳过复核。
    """
    # 列名 → config 键名（同名，见 project_config.accuracy_thresholds）
    fields = ("v_acc", "i_acc", "angle_acc", "p_acc", "q_acc", "s_acc")
    diffs: list[str] = []
    for pt in points:
        if "宽口径" in pt.get("remark", ""):
            continue
        for f in fields:
            table_val = pt.get(f)
            cfg_val = cfg_thr.get(f)
            if cfg_val is None or table_val is None:
                continue
            if abs(float(table_val) - float(cfg_val)) > 1e-12:
                diffs.append(f"{pt['case_id']} {f}: 测点表={table_val} ≠ config={cfg_val}")
    return diffs
