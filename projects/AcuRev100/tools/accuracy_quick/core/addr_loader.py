"""
Modbus 地址表加载器
从标准 Modbus Address Table Excel 中解析所需地址
也提供内置默认地址（AcuRev-100 / RACG），无需 Excel 即可运行

⚠️ AcuRev-100 无中性线电流（RACG 地址表 0x9078 标 `Reserved(1320 Neutral Current)`），
   本模块不提供 "in" 键，测量与报告均不含 N 相电流。
"""
import openpyxl
import logging

# ── 内置默认地址（AcuRev-100 / RACG，Real Time 表 0x9054~0x90A3）──────────
DEFAULT_ADDRS = {
    # 测量量（100ms 刷新，float32，每量占 2 个寄存器）
    "freq":       0x9054,
    "ua":         0x9056,
    "ub":         0x9058,
    "uc":         0x905A,
    "ua_angle":   0x905C,
    "ub_angle":   0x905E,
    "uc_angle":   0x9060,
    "uab":        0x9064,
    "ubc":        0x9066,
    "uca":        0x9068,
    "uab_angle":  0x906A,
    "ubc_angle":  0x906C,
    "uca_angle":  0x906E,
    "ia":         0x9072,
    "ib":         0x9074,
    "ic":         0x9076,
    "ia_angle":   0x907A,
    "ib_angle":   0x907C,
    "ic_angle":   0x907E,
    # 功率单位：W / var / VA（不是 kW，2026-07 台面实证 220V×50A → 11000）
    "pa":         0x9084,
    "pb":         0x9086,
    "pc":         0x9088,
    "p_sys":      0x908A,
    "qa":         0x908C,
    "qb":         0x908E,
    "qc":         0x9090,
    "q_sys":      0x9092,
    "sa":         0x9094,
    "sb":         0x9096,
    "sc":         0x9098,
    "s_sys":      0x909A,
    "pf_a":       0x909C,
    "pf_b":       0x909E,
    "pf_c":       0x90A0,
    "pf_sys":     0x90A2,
    # 配置寄存器（uint16，单寄存器）
    "seal":        0x1040,  # 铅封状态：0x0A=锁定(测量配置全局拒写) / 0=解锁
    "freq_set":    0x1041,  # 频率设置：0=50Hz，1=60Hz（写后电表延迟 30~60s 重启）
    "wire_mode":   0x1042,  # 接线方式（Service Configuration）
    "ct_ch1":      0x1049,  # 通道1 CT 类型
    "ct_pri_ch1":  0x104A,  # 通道1 CT 一次侧额定电流（字面值，5~2000A）
    "ct_ch2":      0x104D,
    "ct_pri_ch2":  0x104E,
    "ct_ch3":      0x1051,
    "ct_pri_ch3":  0x1052,
    "reboot":      0x1147,  # 软重启：写 0x0001（Device Reboot）
}

# 接线方式名称 → 寄存器值（AcuRev-100 仅三种，Service Configuration 0x1042）
#   0: ELEMENT_1_WIRE_2 (1E2W) / 1: ELEMENT_2_WIRE_3_PHASE_1 (2E3W1P, A+C) / 2: ELEMENT_3_WIRE_4_Y (3E4WY)
WIRE_MODE_MAP = {
    "1E2W":    0,
    "2E3W1P":  1,
    "3E4WY":   2,
}

# CT 类型名称 → 寄存器值（与 tests/helpers_accuracy.CT_TYPE_MAP 保持一致）
#   80mA(mA 型) 与 RCT(mV 型, SRS v1.03 新增) 的寄存器值固件侧未确认，确认后再补
CT_TYPE_MAP = {
    "100mA": 0,
    "333mV": 2,
}

SEAL_UNLOCKED = 0        # 0=解锁可写；0x0A=铅封锁定，测量类配置写入被拒(exception 1)


def load_from_excel(excel_path: str) -> dict:
    """
    从标准 Modbus Address Table Excel 解析地址
    规则：
      - 遍历 'Real Time' 和 'Basic Setting' 两个 Sheet（RACG 地址表 sheet 名）
      - 找到 Description 列匹配关键词的行，取 Start(Hex) 列的值
    如果解析失败，返回内置默认地址
    """
    # 关键词 → 内部名称（须与 RACG 地址表 Description 列**逐字一致**，
    # 大小写不敏感；1320 的 "Phase A Line-to-Neutral Voltage Phase Angle" 类写法
    # 在 RACG 表里是 "Phase A Voltage Phase Angle"，写错即静默回退默认地址）
    keyword_map = {
        "System Frequency":                    "freq",
        "Phase A Line-to-Neutral Voltage":     "ua",
        "Phase B Line-to-Neutral Voltage":     "ub",
        "Phase C Line-to-Neutral Voltage":     "uc",
        "Phase A Voltage Phase Angle":         "ua_angle",
        "Phase B Voltage Phase Angle":         "ub_angle",
        "Phase C Voltage Phase Angle":         "uc_angle",
        "Phase AB Line-to-Line Voltage":       "uab",
        "Phase BC Line-to-Line Voltage":       "ubc",
        "Phase CA Line-to-Line Voltage":       "uca",
        "Phase AB Line-to-Line Voltage Phase Angle": "uab_angle",
        "Phase BC Line-to-Line Voltage Phase Angle": "ubc_angle",
        "Phase CA Line-to-Line Voltage Phase Angle": "uca_angle",
        "Phase A Current":                     "ia",
        "Phase B Current":                     "ib",
        "Phase C Current":                     "ic",
        "Phase A Current Phase Angle":         "ia_angle",
        "Phase B Current Phase Angle":         "ib_angle",
        "Phase C Current Phase Angle":         "ic_angle",
        "Phase A Active Power":                "pa",
        "Phase B Active Power":                "pb",
        "Phase C Active Power":                "pc",
        "System Active Power":                 "p_sys",
        "Phase A Reactive Power":              "qa",
        "Phase B Reactive Power":              "qb",
        "Phase C Reactive Power":              "qc",
        "System Reactive Power":               "q_sys",
        "Phase A Apparent Power":              "sa",
        "Phase B Apparent Power":              "sb",
        "Phase C Apparent Power":              "sc",
        "System Apparent Power":               "s_sys",
        "Phase A Power Factor":                "pf_a",
        "Phase B Power Factor":                "pf_b",
        "Phase C Power Factor":                "pf_c",
        "System Power Factor":                 "pf_sys",
        "Sealing Status":                      "seal",
        "Frequency Selection":                 "freq_set",
        "Service Configuration":               "wire_mode",
        "Device Reboot":                       "reboot",
    }
    # CT 通道关键词单独处理（多条相似描述，前缀/包含匹配）
    ct_keywords = [
        ("channel 1 input CT Type",    "ct_ch1"),
        ("channel 2 input CT Type",    "ct_ch2"),
        ("channel 3 input CT Type",    "ct_ch3"),
        ("channel 1 input CT Primary", "ct_pri_ch1"),
        ("channel 2 input CT Primary", "ct_pri_ch2"),
        ("channel 3 input CT Primary", "ct_pri_ch3"),
    ]

    result = dict(DEFAULT_ADDRS)  # 以默认值为基础，覆盖找到的
    hit = 0

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        target_sheets = [s for s in wb.sheetnames if s in ('Real Time', 'Basic Setting')]

        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            # 找列索引
            header = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            try:
                desc_col = header.index('Description')
                addr_col = header.index('Start(Hex)')
            except ValueError:
                logging.warning(f"Sheet '{sheet_name}' missing expected columns")
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                desc = str(row[desc_col] or '').strip()
                addr_raw = row[addr_col]
                if not desc or addr_raw is None:
                    continue
                try:
                    addr = int(str(addr_raw), 16) if isinstance(addr_raw, str) else int(addr_raw)
                except (ValueError, TypeError):
                    continue

                # 精确匹配
                for keyword, name in keyword_map.items():
                    if keyword.lower() == desc.lower():
                        result[name] = addr
                        hit += 1
                        break

                # CT 通道（包含匹配）
                for keyword, name in ct_keywords:
                    if keyword.lower() in desc.lower():
                        result[name] = addr
                        hit += 1
                        break

        missing = sorted(set(keyword_map.values()) | {n for _, n in ct_keywords})
        logging.info("从 %s 命中 %d 条地址（共需 %d 条），未命中项沿用内置默认",
                     excel_path, hit, len(missing))
    except Exception as e:
        logging.error(f"Failed to load address table from Excel: {e}，使用默认地址")

    return result


# 采样/监控读取的测量量键（顺序即报告列顺序的基础）
MEASURE_KEYS = (
    "freq",
    "ua", "ub", "uc", "ua_angle", "ub_angle", "uc_angle",
    "uab", "ubc", "uca",
    "ia", "ib", "ic", "ia_angle", "ib_angle", "ic_angle",
    "pa", "pb", "pc", "p_sys",
    "qa", "qb", "qc", "q_sys",
    "sa", "sb", "sc", "s_sys",
    "pf_a", "pf_b", "pf_c", "pf_sys",
)


def get_measure_addr_map(addrs: dict) -> dict:
    """需要采样的测量量地址字典（实时监控与采样都读全量，按接线方式的判据过滤在 test_engine）。"""
    return {k: addrs[k] for k in MEASURE_KEYS if k in addrs}
