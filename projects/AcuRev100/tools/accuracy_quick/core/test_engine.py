"""精度测试执行引擎（AcuRev-100 / ACmeter 专用）。

在独立线程中运行，通过回调向 GUI 汇报进度和结果。

AcuRev-100 台面的三条硬前提（与 tests/helpers_accuracy.py 同源，均为台面实证结论）：
  1. **电表自供电**：电源取自 CL3021 的 Va/Vn。关源 = 给电表断电 → 全程用保活点
     （A 相 ≥ config source.supply_guard.phase_a_keepalive_v），不用 set_zero/退AC版面。
  2. **测量类配置受铅封门禁**：Sealing Status(0x1040)≠0 时接线/CT/频率写入被拒
     （exception 1）→ 写前先查铅封；写入采幂等口径（读回比对只写差异），
     写频率选择后电表延迟 30~60s 重启，需吃满窗口并期间救源。
  3. **电流经台体 CT 降流进表**：电表期望读数 = 源侧输出电流 × CT Primary ÷ 台体CT额定
     （config source.bench_ct_a），测点表存的是源侧电流。
"""
import time
import logging
import threading
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from core.meter_comm import MeterComm
from core.addr_loader import WIRE_MODE_MAP, CT_TYPE_MAP, SEAL_UNLOCKED, get_measure_addr_map
from core.testpoint_reader import load_testpoints, check_thresholds
from core import power_calc as pc
from core import project_config as cfg


# ── 需要换线才能继续的接线方式（AcuRev-100 仅三种接线）──────────
# 同一套物理接线可连续测试：1E2W（仅 A 相）、2E3W1P（A+C，B 相不接线）
# 需要单独换线：3E4WY（三相星形，需接入 B 相）
REWIRE_REQUIRED = {"3E4WY"}

# CT 类型 → 测点 sheet（mA 型与 mV 型两套源侧电流量级）
SHEET_MAP = {
    "100mA": "test_case_mA",
    "80mA":  "test_case_mA",
    "333mV": "test_case_mV",
    "RCT":   "test_case_mV",
}

# ── 时间常数（沿用 tests/helpers_accuracy 的台面实证值）──────────
# ⏱ 恢复类等待的量级依据（2026-07-27 用户澄清）：**只要源把电压拉起来，电表 <2s 就恢复工作**。
#    因此"等电表"一律取秒级，不再沿用当初"冷启动 3.5min"那套分钟级预算——读不通就是源没出力，
#    该做的是立刻救源，而不是傻等（memory: ≥30s 空窗 = 源输出关死）。
BOOT_TIMEOUT_S       = 30    # 电表复活的最大容忍（超过即认定源没出力，不是电表慢）
DEAD_DETECT_S        = 3     # 设点后无应答即判源掉输出，立刻救源
REARM_BOOT_S         = 10    # 救源后等电表复活窗口（<2s 就够，留 5 倍裕量）
MEASURE_READY_S      = 4     # 设源后测量就绪等待（频率读值 > 10 才允许读判据）
MEASURE_RETRY_S      = 20    # 首轮没就绪时的放宽窗口（覆盖测量块重启后爬坡）
CFG_STABILIZE_S      = 90    # 配置写入后恢复总预算（电表恢复只需秒级；超时即另有问题）
CFG_STABLE_HOLD_S    = 3     # 判定恢复：频率读值连续有效满该秒数（电压稳定 + <2s 恢复 ⇒ 3s 足够）
# 写频率选择后"等延迟重启出现"的窗口。2026-07-13 记录写 4161 后 30~60s 才重启，故原为 90s；
# 2026-07-27 两轮实测（17:01、17:37）全程未观察到重启、纯等 91s，且电表恢复只需 <2s ⇒ 默认 10s。
# 安全网 = _sample_guarded：重启若砸在采样期，本点自动重采而非判 FAIL（代价约 2s + 一次重采）。
# config run.cfg_restart_window_s 可调：0=完全不等（全靠重采网）/ 90=回老保守口径。
CFG_RESTART_WINDOW_S = 10
CFG_PROGRESS_S       = 5     # 配置等待期间的进度播报间隔
SAMPLE_TRIES         = 3     # 采样期测量中断时的重采次数
SRC_REARM_OFFLINE_S  = 12    # 失联持续该时长即认定源掉输出，补拉保活重臂
RESCUE_TRIES         = 3     # 单个测点最多救源轮数（重发本点 → 冷重臂）；超过即停批留证
FREQ_SWITCH_TRIES    = 3     # 频率切换（唯一会掉源输出的动作）的重试轮数，每轮含冷重臂

FREQ_SEL_MAP = {50: 0, 60: 1}

# 各接线方式**有效判据**的量（其余量电表不测，报告置空、不判定）
_ALL_QTY = (
    "ua", "ub", "uc", "ia", "ib", "ic",
    "ua_angle", "ub_angle", "uc_angle", "ia_angle", "ib_angle", "ic_angle",
    "pa", "pb", "pc", "p_sys", "qa", "qb", "qc", "q_sys",
    "sa", "sb", "sc", "s_sys", "uab", "ubc", "uca",
)
PROBE_V = 30.0               # 接线自检电压（源最小档 ≤30V，即使短路能量也远小于 120V+）

# 各接线方式**需要加压**的相（其余相加压=近似短路，源会报过载）
VOLT_PHASES = {
    "1E2W":   ("a",),
    "2E3W1P": ("a", "c"),
    "3E4WY":  ("a", "b", "c"),
}

VALID_QTY: dict[str, set[str]] = {
    "1E2W": {"ua", "ia", "ua_angle", "ia_angle",
             "pa", "qa", "sa", "p_sys", "q_sys", "s_sys"},
    # A+C 两相，B 相端子不接线；线电压只有 Uca 有效
    "2E3W1P": {"ua", "uc", "ia", "ic", "ua_angle", "uc_angle", "ia_angle", "ic_angle",
               "pa", "pc", "qa", "qc", "sa", "sc", "p_sys", "q_sys", "s_sys", "uca"},
    "3E4WY": set(_ALL_QTY),
}

# 报告的物理量块（每块 8 列：期望值 + 最小/最小误差/最大/最大误差/平均/平均误差/结果）
REPORT_BLOCKS: list[tuple[str, str]] = [
    ("ua", "A相电压(V)"), ("ub", "B相电压(V)"), ("uc", "C相电压(V)"),
    ("ia", "A相电流(A)"), ("ib", "B相电流(A)"), ("ic", "C相电流(A)"),
    ("ua_angle", "A相电压相角(°)"), ("ub_angle", "B相电压相角(°)"), ("uc_angle", "C相电压相角(°)"),
    ("ia_angle", "A相电流相角(°)"), ("ib_angle", "B相电流相角(°)"), ("ic_angle", "C相电流相角(°)"),
    ("pa", "A相有功(W)"), ("pb", "B相有功(W)"), ("pc", "C相有功(W)"), ("p_sys", "系统有功(W)"),
    ("qa", "A相无功(var)"), ("qb", "B相无功(var)"), ("qc", "C相无功(var)"), ("q_sys", "系统无功(var)"),
    ("sa", "A相视在(VA)"), ("sb", "B相视在(VA)"), ("sc", "C相视在(VA)"), ("s_sys", "系统视在(VA)"),
    ("uab", "AB线电压(V)"), ("ubc", "BC线电压(V)"), ("uca", "CA线电压(V)"),
]


class TestEngine:
    def __init__(self,
                 meter: MeterComm,
                 source,
                 addrs: dict,
                 testpoint_file: str,
                 ct_type: str,
                 wire_types: list[str],
                 results_dir: str,
                 on_progress=None,
                 on_result=None,
                 on_status=None,
                 on_rewire=None,
                 on_expected=None,
                 settle_s: float = 5.0,
                 sample_cnt: int | None = None,
                 sample_int_ms: int | None = None):
        self.meter = meter
        self.source = source                 # SourceUdp（UDP 台面）或 SourceComm（串口）
        self.addrs = addrs
        self.testpoint_file = testpoint_file
        self.ct_type = ct_type
        self.wire_types = wire_types
        self.results_dir = results_dir
        self.on_progress = on_progress or (lambda c, t: None)
        self.on_result   = on_result   or (lambda w, cid, r: None)
        self.on_status   = on_status   or (lambda m: None)
        self.on_rewire   = on_rewire   or (lambda w, e: e.set())
        self.on_expected = on_expected or (lambda pt, exp: None)
        self.settle_s = settle_s
        # None 表示使用测点 Excel 中的值
        self._sample_cnt = sample_cnt
        self._sample_int_ms = sample_int_ms

        self._stop_event = threading.Event()
        self._thread: threading.Thread | None = None
        self.all_results: list[dict] = []
        # 最近一次已生效的源频率；频率切换是唯一会掉源输出的动作，见 _settle_at_zero_current
        self._last_freq: float | None = None
        self._wire_current: str | None = None    # 当前接线（保活点据此决定哪几相可加压）

    # ── 控制 ─────────────────────────────────────────────────

    def start(self):
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop_event.set()

    def is_running(self) -> bool:
        return self._thread is not None and self._thread.is_alive()

    # ── 保活 / 供电护栏 ───────────────────────────────────────

    def _keepalive_point(self, wire_type: str | None = None) -> dict:
        """保活点：A 相带电（电表电源）、零电流、标准三相角度。

        🔴 未接线相一律不出压：1E2W 的 B/C、2E3W1P 的 B 若被加压，台面上等于近似短路，
        源会报 "Ub/Uc 过载" 并可能锁存保护（2026-07-27 实机）。wire_type 缺省时取当前接线。
        频率沿用当前已生效值——频率切换是唯一会掉源输出的动作，保活点没有理由再切一次。
        """
        keep_v = cfg.supply_guard()[1]
        wt = wire_type or self._wire_current
        wired = VOLT_PHASES.get(wt or "", ("a", "b", "c"))
        return dict(qua=0.0, qub=240.0, quc=120.0, qia=0.0, qib=240.0, qic=120.0,
                    ua=keep_v,
                    ub=keep_v if "b" in wired else 0.0,
                    uc=keep_v if "c" in wired else 0.0,
                    ia=0.0, ib=0.0, ic=0.0,
                    freq=self._last_freq if self._last_freq else 50.0)

    def _probe_wiring(self, wt: str, pt: dict, src_pt: dict):
        """换线后的接线自检：先在 **30V 低压档 + 0A** 上确认该接线要带电的相真的接到了电表。

        🔴 2026-07-27 实机：3E4WY 一开始就 "Ub+Uc 过载" —— 台面仍是 1E2W/2E3W1P 接法
        （B/C 电压端子并接/接 N），直接上 120V 就是近似短路。低压档先探一次：
          - 电表能读到 ≈30V ⇒ 该相确实接到电表，继续正常流程；
          - 读到 ≈0V 或源直接掉输出 ⇒ 立刻降压并报错，不在高压下持续过载。
        1E2W 只有 A 相带电，无需自检。
        """
        expect = [ph for ph in VOLT_PHASES.get(wt, ()) if ph != "a"]
        if not expect:
            return
        probe = self._keepalive_point(wt)
        for ph in expect:
            probe[f"u{ph}"] = PROBE_V
        probe["freq"] = src_pt["freq"]

        # 电表要按目标接线方式配置才会测 B/C（幂等写，此时 30V/0A，安全）
        self._sync_meter_cfg(wt, pt, probe)
        self.on_status(f"{wt} 接线自检：{PROBE_V:g}V/0A 低压探 "
                       f"{'/'.join(p.upper() for p in expect)} 相是否接到电表")
        try:
            self.source.set_point(probe, settle_s=2)
        except Exception as e:
            logging.warning("接线自检下发失败: %s", e)
        if self._ensure_meter_ready(probe) != "ok":
            raise RuntimeError(
                f"{wt} 接线自检：{PROBE_V:g}V 低压探测后源未出力/电表未复活。"
                f"最可能是 {'/'.join(p.upper() for p in expect)} 相电压端子并接或接 N（近似短路）"
                f"——请确认已按 {wt} 换线；若源面板显示过载/保护，需手动复位。")

        vals = self.meter.read_measure_batch({k: self.addrs[k] for k in (f"u{p}" for p in expect)})
        bad = [p.upper() for p in expect if vals.get(f"u{p}", 0.0) < PROBE_V * 0.5]
        if bad:
            try:                                  # 先降压再报错，避免持续过载
                self.source.set_point(self._keepalive_point(wt), settle_s=1)
            except Exception as e:
                logging.warning("自检失败后降压失败: %s", e)
            raise RuntimeError(
                f"{wt} 接线自检未通过：源在 {'/'.join(bad)} 相输出 {PROBE_V:g}V，"
                f"但电表读到 {', '.join(f'U{p}={vals.get(f'u{p.lower()}', 0):.1f}V' for p in bad)}。"
                f"⇒ 这些相的电压端子没接到电表（台面疑似仍是 1E2W/2E3W1P 接法，B/C 并接或接 N）。"
                f"请按 {wt} 正确换线后重跑——继续加压到测点电压会让源过载。")
        self.on_status(f"{wt} 接线自检通过（{', '.join(f'U{p.upper()}≈{vals[f'u{p}']:.1f}V' for p in expect)}）")

    def _guard_point(self, s: dict) -> tuple[dict, bool]:
        """A 相供电护栏（最高优先级）：A 相低于工作下限时强制抬到保活电压。

        返回 (下发点, 是否被抬压)。被抬压时期望值按抬压后的实际输出重算，报告记 [guard]。
        """
        min_v, keep_v = cfg.supply_guard()
        if s["ua"] < min_v:
            g = dict(s)
            g["ua"] = keep_v
            self.on_status(f"⚠ A相 {s['ua']}V < 工作下限 {min_v}V → 强制 {keep_v}V（自供电保护）")
            return g, True
        return s, False

    _AF_KEYS = ("qua", "qub", "quc", "qia", "qib", "qic", "freq")

    @staticmethod
    def _zero_current(src_pt: dict) -> dict:
        """同一测点的"零流版"：电压/角度/频率照目标值，三相电流全 0。

        🔑 只要 Va 还在，电表就不会掉电；源电流输出 0 本身无任何风险（2026-07-27 用户澄清）。
        因此所有切换都放在 0A 上做，代价仅是每点多一次幅值帧。
        """
        z = dict(src_pt)
        z["ia"] = z["ib"] = z["ic"] = 0.0
        return z

    def _settle_at_zero_current(self, src_pt: dict, pt: dict):
        """测点切换第一步：在 0A 上把电压/角度/频率切到目标值，并确认源仍在出力。

        🔴 会掉源输出的**只有频率切换**（2026-07-27 用户澄清；角度帧带 freq 字段，所以
        早前记的"角度帧掉输出"实为频率重设所致）。因此频率有变时：
          先重发档位模式帧（切频率的前提）→ 在 0A 上切 → 验证 → 掉了就重试/冷重臂。
        电压/电流档位切换不掉源，0A 下加降电流更是零风险，所以其余切换直接做。
        """
        zero = self._zero_current(src_pt)
        freq_changed = self._last_freq is not None and self._last_freq != src_pt["freq"]
        tries = FREQ_SWITCH_TRIES if freq_changed or self._last_freq is None else 1
        if freq_changed:
            self.on_status(f"频率切换 {self._last_freq:.0f}→{src_pt['freq']:.0f}Hz："
                           f"档位模式 + 0A 上切换（唯一会掉源输出的动作）")

        for attempt in range(1, tries + 1):
            try:
                if freq_changed or attempt > 1:
                    self.source.ensure_gear_mode()      # 切频率前提；该帧不影响输出
                self.source.set_point(zero, settle_s=1.0, force=attempt > 1)
            except Exception as e:
                logging.warning("零流切换失败（第 %d 次）: %s", attempt, e)
            if self._ensure_meter_ready(zero) == "ok":
                self._last_freq = src_pt["freq"]
                return
            if attempt < tries:
                self.on_status(f"⚠ 切到 {src_pt['freq']:.0f}Hz 后源掉输出 → 冷重臂后重试"
                               f"（{attempt}/{tries - 1}）")
                try:
                    self.source.reinit_output()
                except Exception as e:
                    logging.warning("冷重臂失败: %s", e)
        raise RuntimeError(
            f"{pt['case_id']}: 在 0A 上切到 {src_pt['ua']:g}V/{src_pt['freq']:.0f}Hz 后源未恢复出力"
            f"（{tries} 轮尝试，其中 {max(tries - 1, 0)} 轮含冷重臂）。"
            f"零电流输出本身无风险、电压/电流档位切换也不掉源 ⇒ "
            f"问题在频率切换或源已锁存保护，请到 CL3021 面板确认输出与保护状态（锁存需手动复位）。")

    def _rescue_source(self, point: dict):
        """救源：先在 0A 上把输出拉回来，电表活了再把幅值加回目标点；不行则冷重臂重来。

        🔴 带载重发容易再次触发保护/瞬断，而零电流输出本身无风险（只要 Va 在电表就不掉）
        ⇒ 探活一律用 0A：先试"目标点的零流版"（电压/角度不动，改动最小），
        再退到保活点（A 相降到供电下限，排除高压档位问题）。
        """
        zero = self._zero_current(point)
        ka = self._keepalive_point()
        ka.update({k: point[k] for k in self._AF_KEYS})   # 角度/频率沿用目标点，少发角度帧
        for cold in (False, True):
            if cold:
                try:
                    self.source.reinit_output()
                except Exception as e:
                    logging.warning("源冷重臂失败: %s", e)
                    continue
            for tag, probe in (("目标点零流版", zero), ("保活点", ka)):
                try:
                    self.source.set_point(probe, settle_s=3, force=True)
                except Exception as e:
                    logging.warning("以%s探活失败: %s", tag, e)
                    continue
                if self.meter.read_float(self.addrs["freq"]) is None:
                    continue
                self.on_status(f"↻ 源已在 0A 恢复输出（{tag}{'，冷重臂后' if cold else ''}）")
                if max(abs(point["ia"]), abs(point["ib"]), abs(point["ic"])) > 0:
                    try:
                        self.source.set_point(point, settle_s=3, force=True)
                    except Exception as e:
                        logging.warning("恢复后重加幅值失败: %s", e)
                return

    # ── 主流程 ────────────────────────────────────────────────

    def _run(self):
        try:
            self._run_inner()
        except Exception as e:
            logging.exception("测试中止")
            self.on_status(f"⚠ 测试中止: {e}")

    def _run_inner(self):
        self.on_status("测试开始")
        sheet_name = SHEET_MAP.get(self.ct_type, "test_case_mA")

        # 统计总测试点数（用于进度条）+ 阈值一致性复核
        total_pts = 0
        pts_by_wire: dict[str, list[dict]] = {}
        for wt in self.wire_types:
            pts = load_testpoints(self.testpoint_file, sheet_name, wt)
            pts_by_wire[wt] = pts
            total_pts += len(pts)
        if not total_pts:
            self.on_status(f"⚠ 测点文件 sheet '{sheet_name}' 中未找到所选接线方式的测点")
            return
        diffs = check_thresholds([p for v in pts_by_wire.values() for p in v],
                                 cfg.accuracy_thresholds())
        if diffs:
            self.on_status(f"⚠ 测点表阈值与 config.yaml 不一致（{len(diffs)} 处），已按测点表执行")
            for d in diffs[:10]:
                logging.warning("阈值不一致: %s", d)
        done = 0

        # 源初始化：电表在线（源已在出力）→ 暖初始化不切屏；否则冷初始化重开输出
        if self.meter.read_float(self.addrs["freq"]) is not None:
            self.on_status("源暖初始化（电表在线，不切屏以免瞬断）")
            self.source.warm_init()
        else:
            self.on_status("源冷初始化（联机→切AC界面→档位模式→钉档）")
            self.source.init()

        for wt in self.wire_types:
            if self._stop_event.is_set():
                break

            pts = pts_by_wire[wt]
            if not pts:
                self.on_status(f"{wt}：无测试点，跳过")
                continue
            self._wire_current = wt          # 保活点据此决定哪几相允许加压

            # 换线提示（换线期间保持保活输出，电表不断电）
            if wt in REWIRE_REQUIRED:
                ev = threading.Event()
                self.on_rewire(wt, ev)
                self.on_status(f"等待用户完成 {wt} 换线并确认…")
                ev.wait()

            # 接线自检：30V 低压确认该接线要带电的相真的接到电表，避免高压下过载
            first_src, _ = self._guard_point(self._source_point(pts[0], wt))
            self._probe_wiring(wt, pts[0], first_src)

            # 结果 Excel
            result_file = self._make_result_file(wt)
            wb = Workbook()
            ws = wb.active
            ws.title = wt
            self._write_header(ws)

            for idx, pt in enumerate(pts):
                if self._stop_event.is_set():
                    break

                self.on_status(f"{wt} [{idx+1}/{len(pts)}] {pt['case_id']}")
                src_pt, guarded = self._guard_point(self._source_point(pt, wt))

                # 逐相能力过滤：本台源某相出不到额定电流时跳过该点（换正常源改 config 即恢复）
                over = cfg.phase_cap_violations(src_pt)
                if over:
                    reason = "；".join(over)
                    if cfg.over_phase_cap() == "abort":
                        raise RuntimeError(f"{pt['case_id']}: {reason}"
                                           f"（config source.over_phase_cap=abort）")
                    self.on_status(f"⏭ 跳过 {pt['case_id']}：{reason}（本台源能力所限）")
                    self.all_results.append({"wire": wt, "case_id": pt["case_id"],
                                             "skipped": reason})
                    self.on_result(wt, pt["case_id"], {"skipped": reason})
                    self._write_row(ws, idx + 2, pt, src_pt, self._calc_expected(pt, src_pt, wt),
                                    {}, wt, guarded, skipped=reason)
                    wb.save(result_file)
                    done += 1
                    self.on_progress(done, total_pts)
                    continue

                # ① 先在 0A 上把电压/角度/频率切到目标值（Va 保持，电表不掉电）
                self._settle_at_zero_current(src_pt, pt)

                # ② 电表配置（接线/CT类型/CT Primary/频率选择）幂等同步
                #    此时源已在目标频率、且处于 0A —— 重启窗口内电表对着正确输入判稳
                try:
                    self._sync_meter_cfg(wt, pt, self._zero_current(src_pt))
                except RuntimeError as e:
                    self.on_status(f"⚠ {e}")
                    raise

                # ③ 最后把电流加到测点值（只变幅值帧，角度帧命中同值跳发）
                try:
                    self.source.set_point(src_pt, settle_s=self.settle_s)
                except Exception as src_err:
                    logging.error("源输出失败 [%s]: %s", pt['case_id'], src_err)
                    self.on_status(f"⚠ 源命令失败，跳过 {pt['case_id']}: {src_err}")
                    done += 1
                    self.on_progress(done, total_pts)
                    continue

                # 电表存活 + 测量就绪（源掉输出 = 自供电表直接断电）
                ready = self._ensure_meter_ready(src_pt)
                if ready == "src_dead":
                    # 🔴 停批而不是跳点：电表一直没电，继续跑只会把剩余测点刷成空壳
                    raise RuntimeError(
                        f"{pt['case_id']}: 电表持续无 Modbus 应答，即源输出未恢复"
                        f"（已重发测点+冷重臂 {RESCUE_TRIES} 轮）。请到 CL3021 面板确认是否显示"
                        f"过载/保护——保护锁存只能手动复位，程序侧的切屏重臂清不掉；"
                        f"并核对本接线方式下未接线相是否被误加压（1E2W 的 B/C、2E3W1P 的 B）。")
                if ready == "measure_dead":
                    raise RuntimeError(
                        f"{pt['case_id']}: 电表在线（Va 有电）但测量恒读 0/无效频率。"
                        f"对照 knowledge context 疑似缺陷记录：异常掉电重启后测量恒 0，需对电表"
                        f"执行【恢复出厂设置】才能恢复；也请顺带确认源确实在按测点出力。")

                # 期望值（含源侧电流 → 电表读数换算）
                expected = self._calc_expected(pt, src_pt, wt)
                self.on_expected(pt, expected)

                # 采样（优先使用界面输入值，否则用测点 Excel 中的值）
                n   = self._sample_cnt    if self._sample_cnt    is not None else pt["sample_cnt"]
                ims = self._sample_int_ms if self._sample_int_ms is not None else pt["sample_int_ms"]
                samples = self._sample_guarded(pt, src_pt, n, ims)
                if samples is None:
                    self.on_status(f"⚠ {pt['case_id']}：采样期间测量反复中断，跳过本点")
                    done += 1
                    self.on_progress(done, total_pts)
                    continue

                # 精度判定
                row_data = self._calc_row(pt, expected, samples, wt)
                self.all_results.append({"wire": wt, "case_id": pt["case_id"], **row_data})
                self.on_result(wt, pt["case_id"], row_data)

                # 写入 Excel
                self._write_row(ws, idx + 2, pt, src_pt, expected, row_data, wt, guarded)
                wb.save(result_file)

                done += 1
                self.on_progress(done, total_pts)

            # 本接线跑完：回保活（🔴 不关源——关源即给自供电电表断电）
            # 保活点沿用当前频率，不做无谓的频率切换（下一接线首点若需换频率会自行处理）
            try:
                self.source.set_point(self._keepalive_point(), settle_s=2)
            except Exception as e:
                logging.warning("回保活失败: %s", e)

        self.on_status("✅ 测试完成" if not self._stop_event.is_set() else "⚠ 测试已停止")

    # ── 电表配置（幂等 + 铅封门禁 + 延迟重启窗口）────────────

    def _sync_meter_cfg(self, wt: str, pt: dict, src_pt: dict):
        """按接线方式/CT/频率同步电表配置；只写有差异项，全一致则整组跳过。"""
        ct_val = CT_TYPE_MAP.get(self.ct_type)
        if ct_val is None:
            raise RuntimeError(f"CT 类型 '{self.ct_type}' 无寄存器映射（80mA/RCT 待固件确认）")
        freq_hz = int(pt["freq"])
        if freq_hz not in FREQ_SEL_MAP:
            raise RuntimeError(f"频率 {freq_hz}Hz 无寄存器映射（仅 50/60Hz）")
        primary = int(pt["ct_primary"])

        desired: dict[int, int] = {
            self.addrs["wire_mode"]: WIRE_MODE_MAP[wt],
            self.addrs["freq_set"]: FREQ_SEL_MAP[freq_hz],
        }
        for ch in (1, 2, 3):
            desired[self.addrs[f"ct_ch{ch}"]] = ct_val
            desired[self.addrs[f"ct_pri_ch{ch}"]] = primary

        diffs = {addr: val for addr, val in desired.items()
                 if self.meter.read_uint16(addr) != val}
        if not diffs:
            return

        seal = self.meter.read_seal(self.addrs["seal"])
        if seal is None or seal != SEAL_UNLOCKED:
            raise RuntimeError(
                f"电表铅封锁定（Sealing Status@{hex(self.addrs['seal'])}={seal}），"
                f"接线/CT/频率等测量配置只读，写入会被拒（exception 1）。"
                f"请拨开电表背面端子侧 Dip Switch 解锁（LCD 应弹 'Remote Configuration Mode'）后重跑。")

        self.on_status(f"写入电表配置 {len(diffs)}/{len(desired)} 项"
                       f"（接线={wt}, CT={self.ct_type}/{primary}A, 频率={freq_hz}Hz）")
        for addr, val in diffs.items():
            if not self.meter.write_uint16(addr, val):
                raise RuntimeError(f"配置寄存器 {hex(addr)} 写入失败（值 {val}）")
        time.sleep(1)
        # 写频率选择会触发延迟重启（实测 30~60s）；其余配置只需等测量稳定
        expect_restart = self.addrs["freq_set"] in diffs
        self._wait_cfg_stable(expect_restart, src_pt)

    def _wait_cfg_stable(self, expect_restart: bool, src_pt: dict):
        """吃掉配置写入后的（延迟）重启窗口：等测量连续稳定；失联超时补拉保活救源。

        写频率选择后电表会自行延迟重启（实测 30~60s），必须等它重启完并连续稳定
        CFG_STABLE_HOLD_S 才能采样，否则重启会砸在首个测点上。等待期间每
        CFG_PROGRESS_S 秒播报一次进度——否则这段静默会被误认为程序卡死。
        """
        t0 = time.time()
        restarted = False
        offline_since = None
        stable_since = None
        next_tick = CFG_PROGRESS_S
        window = cfg.cfg_restart_window_s()
        if expect_restart:
            self.on_status(f"等电表延迟重启（写了频率选择；窗口 {window:.0f}s + 连续稳定 "
                           f"{CFG_STABLE_HOLD_S}s，重启若砸在采样期会自动重采）")
        while time.time() - t0 < CFG_STABILIZE_S and not self._stop_event.is_set():
            waited = time.time() - t0
            if waited >= next_tick:
                next_tick += CFG_PROGRESS_S
                state = ("已跨过重启，等稳定" if restarted else
                         "等重启出现" if expect_restart else "等测量稳定")
                self.on_status(f"配置生效等待中：已等 {waited:.0f}s（{state}；"
                               f"源在 0A 保持输出，电表未断电）")
            val = self.meter.read_float(self.addrs["freq"])
            ok = val is not None and val > 10.0
            now = time.time()
            if ok:
                if offline_since is not None:
                    restarted = True          # 曾失联又回来 = 延迟重启已发生并恢复
                offline_since = None
                stable_since = stable_since or now
                awaiting = expect_restart and not restarted and (now - t0 < window)
                if now - stable_since >= CFG_STABLE_HOLD_S and not awaiting:
                    self.on_status(f"配置写入后测量已稳定（等待 {now - t0:.0f}s"
                                   + ("，已跨过延迟重启）" if restarted else "）"))
                    return
            else:
                stable_since = None
                offline_since = offline_since or now
                if now - offline_since > SRC_REARM_OFFLINE_S:
                    self.on_status("配置重启窗口内电表失联 → 补拉保活/重臂源输出")
                    self._rescue_source(src_pt)
                    offline_since = now
            time.sleep(1)
        self.on_status(f"⚠ 配置写入后 {CFG_STABILIZE_S}s 内测量未达连续稳定，继续执行并留证")

    def _ensure_meter_ready(self, src_pt: dict) -> str:
        """设点后确认电表在线且测量就绪。返回 "ok" / "src_dead" / "measure_dead"。

        🔑 AcuRev-100 **只能由 Va/Vn 供电**（USB 口不供电）：
          - Modbus 无应答 ⇒ 电表没电 ⇒ 源输出掉 0 ⇒ 救源（强制重发本点 → 冷重臂），
            最多 RESCUE_TRIES 轮；冷重臂清不掉源侧锁存的过载/保护，需人工复位。
          - Modbus 有应答但频率读不到有效值 ⇒ **电表有电**，是测量块还在启动窗口，
            或撞上"异常掉电重启后测量恒 0"的固件缺陷（见 knowledge context 疑似缺陷记录，
            需恢复出厂设置）——此时救源是错的处置，只延长等待再判。
        """
        for attempt in range(1, RESCUE_TRIES + 1):
            t0 = time.time()
            alive = False
            while time.time() - t0 < DEAD_DETECT_S:
                if self.meter.read_float(self.addrs["freq"]) is not None:
                    alive = True
                    break
                time.sleep(0.5)

            if alive:
                # 电表有电: 给足测量就绪窗口(首轮短、后续放宽到测量块冷启动量级)
                timeout = MEASURE_READY_S if attempt == 1 else MEASURE_RETRY_S
                if self.meter.wait_measure_ready(self.addrs["freq"], timeout_s=timeout):
                    return "ok"
                self.on_status(f"⚠ 电表在线但测量读不到有效频率（第 {attempt}/{RESCUE_TRIES} 轮，"
                               f"已等 {timeout:.0f}s）——测量块启动中或测量恒0缺陷，继续等")
                continue

            self.on_status(f"⚠ 电表无 Modbus 应答（只能 Va/Vn 供电 ⇒ 源输出掉0），"
                           f"第 {attempt}/{RESCUE_TRIES} 次救源")
            self._rescue_source(src_pt)
            try:
                self.meter.wait_alive(self.addrs["freq"], timeout_s=REARM_BOOT_S)
            except RuntimeError as e:
                logging.warning("%s", e)
        return "src_dead" if not alive else "measure_dead"

    def _sample_guarded(self, pt: dict, src_pt: dict, n: int, ims: int):
        """采样并校验采样期间测量没中断；中断则等就绪后重采。返回 None = 反复失败。

        🔑 有了这道网，"等电表延迟重启"的窗口才敢压短：万一重启真砸在采样中间，
        代价只是本点重采一次，而不是把它判成 FAIL（频率样本里会出现 ≤10Hz 的坑）。
        """
        addr_map = get_measure_addr_map(self.addrs)
        for attempt in range(1, SAMPLE_TRIES + 1):
            samples = self.meter.sample_n_times(addr_map, n, ims)
            freqs = samples.get("freq") or []
            if freqs and min(freqs) > 10.0:
                return samples
            got = min(freqs) if freqs else 0.0
            self.on_status(f"⚠ {pt['case_id']} 采样期间测量中断（频率样本最小 {got:.1f}Hz，"
                           f"疑似电表重启）→ 等就绪后重采（{attempt}/{SAMPLE_TRIES}）")
            if self._ensure_meter_ready(src_pt) != "ok":
                return None
        return None

    # ── 源测点 / 期望值 ───────────────────────────────────────

    @staticmethod
    def _source_point(pt: dict, wt: str) -> dict:
        """测点 → 源下发点：**未接线相的电压与电流幅值一律清零**。

        🔴 2026-07-27 实机教训：1E2W 下给 B/C 相加压，源报 "Ub 过载"（台面未接线相的
        电压端子并接/接 N，加压即近似短路），随后进保护锁存把输出打掉、冷重臂拉不回。
        1320 老测点表（同台面跑过）1E2W 行本就是 U=(69,0,0)、2E3W1P 行是 U=(69,0,69)，
        即未接线相不出压——本方法按接线方式强制这条铁律，与测点表内容无关。

        六个相角照测点表下发（幅值为 0 时角度无意义，但保持角度帧内容稳定 → 命中
        SourceUdp 同值跳发，少一次角度帧就少一次输出瞬断/电表掉电重启）。
        """
        s = dict(qua=pt["ua_p"], qub=pt["ub_p"], quc=pt["uc_p"],
                 qia=pt["ia_p"], qib=pt["ib_p"], qic=pt["ic_p"],
                 ua=pt["ua"], ub=pt["ub"], uc=pt["uc"],
                 ia=pt["ia"], ib=pt["ib"], ic=pt["ic"],
                 freq=float(pt["freq"]))
        if wt == "1E2W":                     # 只有 A 相接线：B/C 相不出压不出流
            s["ub"] = s["uc"] = s["ib"] = s["ic"] = 0.0
        elif wt == "2E3W1P":                 # A+C 两相：B 相不出压不出流
            s["ub"] = s["ib"] = 0.0
        return s

    def _calc_expected(self, pt: dict, src_pt: dict, wt: str) -> dict:
        """按**实际下发的源输出**算期望值；电流按台体 CT 换算成电表读数量级。"""
        scale = cfg.current_scale(self.ct_type, pt["ct_primary"])
        ua, ub, uc = src_pt["ua"], src_pt["ub"], src_pt["uc"]
        ua_p, ub_p, uc_p = src_pt["qua"], src_pt["qub"], src_pt["quc"]
        ia_p, ib_p, ic_p = src_pt["qia"], src_pt["qib"], src_pt["qic"]
        ia, ib, ic = src_pt["ia"] * scale, src_pt["ib"] * scale, src_pt["ic"] * scale

        uab, ubc, uca = pc.line_to_line_voltage(ua, ub, uc, ua_p, ub_p, uc_p)
        pa, qa, sa = pc.phase_power(ua, ia, ua_p, ia_p)
        pb, qb, sb = pc.phase_power(ub, ib, ub_p, ib_p)
        pcp, qc, sc = pc.phase_power(uc, ic, uc_p, ic_p)

        valid = VALID_QTY.get(wt, set(_ALL_QTY))
        p_terms = [v for k, v in (("pa", pa), ("pb", pb), ("pc", pcp)) if k in valid]
        q_terms = [v for k, v in (("qa", qa), ("qb", qb), ("qc", qc)) if k in valid]
        s_terms = [v for k, v in (("sa", sa), ("sb", sb), ("sc", sc)) if k in valid]

        return {
            "i_scale": scale,
            "ua": ua, "ub": ub, "uc": uc,
            "ia": ia, "ib": ib, "ic": ic,
            "ua_angle": ua_p, "ub_angle": ub_p, "uc_angle": uc_p,
            "ia_angle": ia_p, "ib_angle": ib_p, "ic_angle": ic_p,
            "pa": pa, "pb": pb, "pc": pcp, "p_sys": pc.sys_power(*p_terms),
            "qa": qa, "qb": qb, "qc": qc, "q_sys": pc.sys_power(*q_terms),
            "sa": sa, "sb": sb, "sc": sc, "s_sys": pc.sys_power(*s_terms),
            "uab": uab, "ubc": ubc, "uca": uca,
        }

    # ── 精度汇总 ──────────────────────────────────────────────

    def _calc_row(self, pt: dict, exp: dict, samples: dict[str, list], wt: str) -> dict:
        """逐量判定；未接线相的量返回 None（报告置空、不判定）。"""
        m = self.meter
        valid = VALID_QTY.get(wt, set(_ALL_QTY))
        # 相对量的参照视在功率（用于期望值≈0 时改绝对带，见 config accuracy 口径）
        s_ref = {
            "pa": exp["sa"], "pb": exp["sb"], "pc": exp["sc"], "p_sys": exp["s_sys"],
            "qa": exp["sa"], "qb": exp["sb"], "qc": exp["sc"], "q_sys": exp["s_sys"],
            "sa": exp["sa"], "sb": exp["sb"], "sc": exp["sc"], "s_sys": exp["s_sys"],
        }
        thr = {
            "ua": pt["v_acc"], "ub": pt["v_acc"], "uc": pt["v_acc"],
            "uab": pt["v_acc"], "ubc": pt["v_acc"], "uca": pt["v_acc"],
            "ia": pt["i_acc"], "ib": pt["i_acc"], "ic": pt["i_acc"],
            "pa": pt["p_acc"], "pb": pt["p_acc"], "pc": pt["p_acc"], "p_sys": pt["p_acc"],
            "qa": pt["q_acc"], "qb": pt["q_acc"], "qc": pt["q_acc"], "q_sys": pt["q_acc"],
            "sa": pt["s_acc"], "sb": pt["s_acc"], "sc": pt["s_acc"], "s_sys": pt["s_acc"],
        }

        row: dict[str, dict | None] = {}
        for key in _ALL_QTY:
            if key not in valid or key not in samples:
                row[key] = None
                continue
            std = exp[key]
            if key.endswith("_angle"):
                res = m.calc_phase_angle_accuracy(std, samples[key], pt["angle_acc"])
                res["thr"] = pt["angle_acc"]          # 绝对门限（°）
                row[key] = res
                continue
            pct = thr[key]
            ref = s_ref.get(key)
            if std == 0 or (ref and abs(std) < 0.01 * abs(ref)):
                # 期望≈0（如 PF=1 时的无功）：改用以 |S| 为参照的绝对带
                if not ref:
                    row[key] = None
                    continue
                band = pct * abs(ref)
                res = m.calc_accuracy_abs(std, samples[key], band)
                res["thr"] = band                     # 绝对门限（同量单位）
            else:
                res = m.calc_accuracy(std, samples[key], pct)
                res["thr"] = pct                      # 相对门限（比值）
            row[key] = res
        return row

    # ── Excel 输出 ────────────────────────────────────────────

    def _make_result_file(self, wire_type: str) -> str:
        os.makedirs(self.results_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = f"Precision_Measure_{wire_type}_{ts}.xlsx"
        return os.path.join(self.results_dir, name)

    @staticmethod
    def _input_headers() -> list[str]:
        return [
            "用例编号", "接线方式", "CT类型", "CT Primary(A)", "频率(Hz)", "电流换算系数",
            "voltage_accuracy", "current_accuracy", "phase_angle_accuracy(°)",
            "active_power_accuracy", "reactive_power_accuracy", "apparent_power_accuracy",
            "源A相电压(V)", "源B相电压(V)", "源C相电压(V)",
            "源A相电流(A)", "源B相电流(A)", "源C相电流(A)",
            "源A相电压相角(°)", "源B相电压相角(°)", "源C相电压相角(°)",
            "源A相电流相角(°)", "源B相电流相角(°)", "源C相电流相角(°)",
            "备注",
        ]

    def _write_header(self, ws):
        headers = self._input_headers()
        for _key, label in REPORT_BLOCKS:
            headers += [f"{label}期望值",
                        f"{label}最小值", f"{label}最小误差",
                        f"{label}最大值", f"{label}最大误差",
                        f"{label}平均值", f"{label}平均误差",
                        f"{label}精度测试结果"]

        hdr_fill = PatternFill("solid", fgColor="4F81BD")
        hdr_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    def _write_row(self, ws, row_num: int, pt: dict, src_pt: dict, exp: dict,
                   row_data: dict, wire_type: str, guarded: bool = False,
                   skipped: str = ""):
        """写一行数据；超门限着色，未接线相的量整块置空。

        skipped 非空 = 本测点未执行（源能力所限），各量整块留空、备注写原因。
        """
        red    = PatternFill("solid", fgColor="FFC7CE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        green  = PatternFill("solid", fgColor="C6EFCE")

        col = 1

        def wc(value, fill=None):
            nonlocal col
            cell = ws.cell(row=row_num, column=col, value=value)
            if fill:
                cell.fill = fill
            col += 1

        remark = pt.get("remark", "")
        if guarded:
            remark = ("[guard] A相被抬到供电下限，期望值按实际输出重算；" + remark).strip("；")
        if skipped:
            remark = f"[SKIPPED] 未执行：{skipped}；" + remark

        # ── 输入参数列 ──
        wc(pt["case_id"]); wc(wire_type); wc(self.ct_type)
        wc(pt["ct_primary"]); wc(pt["freq"]); wc(round(exp["i_scale"], 6))
        wc(pt["v_acc"]); wc(pt["i_acc"]); wc(pt["angle_acc"])
        wc(pt["p_acc"]); wc(pt["q_acc"]); wc(pt["s_acc"])
        wc(src_pt["ua"]); wc(src_pt["ub"]); wc(src_pt["uc"])
        wc(src_pt["ia"]); wc(src_pt["ib"]); wc(src_pt["ic"])
        wc(src_pt["qua"]); wc(src_pt["qub"]); wc(src_pt["quc"])
        wc(src_pt["qia"]); wc(src_pt["qib"]); wc(src_pt["qic"])
        wc(remark, PatternFill("solid", fgColor="D9D9D9") if skipped else None)

        # ── 各物理量块（8 列）──
        for key, _label in REPORT_BLOCKS:
            d = row_data.get(key)
            if d is None:
                for _ in range(8):
                    wc("")
                continue
            thr = d.get("thr", 0.0)
            mn_over = abs(d["min_err"]) > thr
            mx_over = abs(d["max_err"]) > thr
            av_over = abs(d["avg_err"]) > thr
            wc(round(exp[key], 5))
            wc(d["min"],     yellow if mn_over else None)
            wc(d["min_err"], yellow if mn_over else None)
            wc(d["max"],     yellow if mx_over else None)
            wc(d["max_err"], yellow if mx_over else None)
            wc(d["avg"],     red    if av_over else None)
            wc(d["avg_err"], red    if av_over else None)
            wc("Passed" if d["pass"] else "Failed", green if d["pass"] else red)
