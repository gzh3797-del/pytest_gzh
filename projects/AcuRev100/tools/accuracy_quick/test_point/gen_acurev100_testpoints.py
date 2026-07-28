"""生成 AcuRev-100(ACmeter) 快速精度测试的测点文件 acurev100_test_case.xlsx。

xlsx 是二进制、改动不可 review，故矩阵以本脚本为权威定义，改点位请改本脚本再重跑：
    python projects/AcuRev100/tools/accuracy_quick/test_point/gen_acurev100_testpoints.py

矩阵口径（精简矩阵，2026-07-27 用户定）
  电压: 120 / 220 / 277V —— 277V 为 PRS「电压测量 CAT III 277V/480V」的 VLN 上限；
        1320 老表的 347/480V 作 VLN 已超 AcuRev-100 规格，不再使用。
  电流: 以 CT 一次侧额定 200A 为基准，覆盖 额定100% / 50% / 10% / Imin(1%) / Ist(0.1%)。
        表中列 4~6 存**源侧输出电流**，电表读数 = 源电流 × Primary ÷ 台体CT额定
        （mA 型台体 20A/100mA → ×10；mV 型台体 5A/333mV → ×40，见 config source.bench_ct_a）。
  PF:   1.0 / 0.5L / 0.8L / 0.5C / 0.8C（L=感性电流滞后, C=容性电流超前）。
  频率: 50Hz 为主 + 2 个 60Hz 点（写频率选择会触发电表延迟重启 30~60s，故不逐点切）。
  阈值: 取自 projects/AcuRev100/config.yaml accuracy 段（V/I 0.2%、P 0.2%、Q/S 0.5%、
        相角 ±0.5°）；Ist 点低于 Imin 不在精度等级适用区间，按 config accuracy.default_pct
        ±10% 定性口径，并在备注标"宽口径"（testpoint_reader.check_thresholds 据此跳过复核）。
"""
from __future__ import annotations

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.abspath(os.path.join(_HERE, "..")))

from core import project_config as cfg          # noqa: E402
from core.testpoint_reader import WIRE_TYPES    # noqa: E402

OUT_FILE = os.path.join(_HERE, "acurev100_test_case.xlsx")

CT_PRIMARY = 200.0          # CT 一次侧额定电流(A)：CTEP 场景可选 200/400，取 200 与 case_map 一致
I_RATED = CT_PRIMARY        # 额定电流(电表侧 A)

# 各接线方式的电压相角（Va/Vb/Vc）
#   1E2W / 3E4WY：标准三相 0/240/120
#   2E3W1P：**分相 A+C 反相 180°**（已实跑用例 009_02 与 1320 老表均为 0/240/180；
#           120° 是 case_map 里未执行的 mV 用例的沿用值，跑出来 Uca 期望全错）
V_ANGLES = {
    "1E2W":   (0.0, 240.0, 120.0),
    "2E3W1P": (0.0, 240.0, 180.0),
    "3E4WY":  (0.0, 240.0, 120.0),
}
# 高压档取值：受 PRS「电压测量 CAT III 277V/480V」双上限约束
#   1E2W 只有 A 相无线电压 → 277V；3E4WY 线电压 √3×277≈480V → 277V；
#   2E3W1P 分相线电压 = 2×相电压 → 取 240V（2×240=480 恰在上限；277V 会到 554V 超规格）
V_HIGH = {"1E2W": 277.0, "2E3W1P": 240.0, "3E4WY": 277.0}

HEADERS = [
    "test_case", "Phase_A_Voltage(V)", "Phase_B_Voltage(V)", "Phase_C_Voltage(V)",
    "Phase_A_Current_src(A)", "Phase_B_Current_src(A)", "Phase_C_Current_src(A)",
    "Va_angle", "Vb_angle", "Vc_angle", "Ia_angle", "Ib_angle", "Ic_angle",
    "等待时间(h)", "PF", "接线方式", "频率(Hz)",
    "voltage_accuracy", "current_accuracy", "phase_angle_accuracy(°)",
    "active_power_accuracy", "reactive_power_accuracy", "apparent_power_accuracy",
    "采样次数", "采样间隔(s)", "CT_Primary(A)", "备注",
]

# (电压档, 电表侧电流A, PF标签, 频率Hz, 备注)
#   电压 "HIGH" = 按接线方式取 V_HIGH
#
# 🔴 矩阵按**完整覆盖**编写：满载点就是 100% 额定（mA 型对应源侧 20A），
#    不把某台源的个体能力烧进测点文件。当前台面这台 CL3021 的 Ic 出不到 20A，
#    由引擎在运行时按 config source.max_current_a_phase / over_phase_cap 跳过并留证；
#    换正常源把该项改回 20A 即自动全跑，测点文件不用重生成。
#    为保证受限源下仍有近满载覆盖，另置一个 75% 额定点（源侧 15A，本台源可跑）。
# 排序原则：**同频率的点连续**（频率切换是唯一会掉源输出的动作，60Hz 排末尾使每种接线只切一次），
#   同 PF 的点也连续（角度帧带 freq 字段，同值可跳发，少一次无谓的频率重设）。
MATRIX: list[tuple[float | str, float, str, int, str]] = [
    (120.0,  I_RATED,         "1.0",  50, "额定电流 100%"),
    (120.0,  I_RATED * 0.75,  "1.0",  50, "额定电流 75%（受限源下的近满载点）"),
    (120.0,  I_RATED * 0.5,   "1.0",  50, "额定电流 50%"),
    (120.0,  I_RATED * 0.1,   "1.0",  50, "额定电流 10%"),
    (120.0,  I_RATED * 0.01,  "1.0",  50, "Imin = 1%×额定"),
    (120.0,  I_RATED * 0.001, "1.0",  50, "Ist = 0.1%×额定；低于 Imin 不适用精度等级，宽口径"),
    (220.0,  I_RATED,         "1.0",  50, "额定电流 100%"),
    (220.0,  I_RATED * 0.5,   "1.0",  50, ""),
    (220.0,  I_RATED * 0.01,  "1.0",  50, "Imin"),
    ("HIGH", I_RATED,         "1.0",  50, "电压上限 + 额定电流 100%"),
    ("HIGH", I_RATED * 0.01,  "1.0",  50, "电压上限 + Imin"),
    (220.0,  I_RATED * 0.5,   "0.5L", 50, "感性"),
    (220.0,  I_RATED * 0.5,   "0.8L", 50, "感性"),
    (220.0,  I_RATED * 0.5,   "0.5C", 50, "容性"),
    (220.0,  I_RATED * 0.5,   "0.8C", 50, "容性"),
    # 60Hz 点排末尾：写频率选择寄存器要吃 90s 延迟重启窗口，排末尾则每种接线只切一次
    (220.0,  I_RATED,         "1.0",  60, "60Hz 额定电流 100%"),
    (220.0,  I_RATED * 0.5,   "0.5L", 60, "60Hz 感性"),
]

PF_PHI = {"1.0": 0.0, "0.5": 60.0, "0.8": 36.87}


def _current_angles(wire: str, pf_label: str) -> tuple[float, float, float]:
    """PF 标签 → 三相电流相角（跟随该接线的电压相角）。

    L=电流滞后(θ_V−θ_I>0 ⇒ Q>0)，C=电流超前(Q<0)。
    """
    core = pf_label.rstrip("LC")
    phi = PF_PHI[core]
    sign = -1.0 if pf_label.endswith("L") else 1.0        # 滞后 = 相角减小
    return tuple((va + sign * phi) % 360.0 for va in V_ANGLES[wire])


def _rows(bench_ct_a: float, thr: dict, wide_thr: float) -> list[list]:
    scale = CT_PRIMARY / bench_ct_a               # 源侧电流 → 电表读数 的换算系数
    rows: list[list] = []
    idx = 0
    for wire in WIRE_TYPES:
        va_ang, vb_ang, vc_ang = V_ANGLES[wire]
        for volt_spec, i_meter, pf, freq, remark in MATRIX:
            idx += 1
            volt = V_HIGH[wire] if volt_spec == "HIGH" else float(volt_spec)
            i_src = round(float(i_meter) / scale, 6)
            ia_ang, ib_ang, ic_ang = _current_angles(wire, pf)
            # 🔴 未接线相不出压不出流（1E2W 仅 A；2E3W1P 为 A+C，B 相端子不接线）——
            #    2026-07-27 实机：1E2W 给 B/C 加压导致源报 "Ub 过载" 并进保护
            ub = 0.0 if wire in ("1E2W", "2E3W1P") else volt
            uc = 0.0 if wire == "1E2W" else volt
            ib = 0.0 if wire in ("1E2W", "2E3W1P") else i_src
            ic = 0.0 if wire == "1E2W" else i_src
            wide = "宽口径" in remark
            rows.append([
                f"RACG_case{idx}", volt, ub, uc,
                i_src, ib, ic,
                va_ang, vb_ang, vc_ang,
                round(ia_ang, 2), round(ib_ang, 2), round(ic_ang, 2),
                0, pf, wire, freq,
                thr["v_acc"],
                wide_thr if wide else thr["i_acc"],
                thr["angle_acc"],
                wide_thr if wide else thr["p_acc"],
                wide_thr if wide else thr["q_acc"],
                wide_thr if wide else thr["s_acc"],
                10, 0.2, CT_PRIMARY,
                f"电表侧电流 {i_meter:g}A（源侧 {i_src:g}A）"
                + (f"；{remark}" if remark else ""),
            ])
    return rows


def _write_sheet(ws, rows: list[list]):
    hdr_fill = PatternFill("solid", fgColor="4F81BD")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions[get_column_letter(len(HEADERS))].width = 42   # 备注列


def main():
    thr = cfg.accuracy_thresholds()
    wide = cfg.default_pct()
    wb = Workbook()

    ws = wb.active
    ws.title = "test_case_mA"
    _write_sheet(ws, _rows(cfg.bench_ct_a("100mA"), thr, wide))

    ws_mv = wb.create_sheet("test_case_mV")
    _write_sheet(ws_mv, _rows(cfg.bench_ct_a("333mV"), thr, wide))

    ws_doc = wb.create_sheet("说明")
    doc = [
        ("测点矩阵来源", "由 test_point/gen_acurev100_testpoints.py 生成，改点位请改脚本后重跑"),
        ("电流列语义", "列 5~7 = 源侧输出电流(A)；电表期望读数 = 源电流 × CT_Primary ÷ 台体CT额定"),
        ("台体CT额定", f"mA 型 {cfg.bench_ct_a('100mA'):g}A(20A/100mA) / "
                       f"mV 型 {cfg.bench_ct_a('333mV'):g}A(5A/333mV)，见 config.yaml source.bench_ct_a"),
        ("阈值来源", "projects/AcuRev100/config.yaml accuracy 段；相角为绝对量(°)，其余为比值"),
        ("满载点", "矩阵按 100%额定编写（mA 型源侧 20A）；某台源某相出不到额定时由引擎运行时"
                    "跳过并在报告标 [SKIPPED]（config source.max_current_a_phase / over_phase_cap），"
                    "换正常源改配置即全跑，本文件不用重生成。另置 75%额定点保证受限源下的近满载覆盖"),
        ("宽口径行", f"Ist(0.1%×额定) 低于 Imin，不在精度等级适用区间，按 ±{wide * 100:g}% 定性判定"),
        ("电压上限", "1E2W/3E4WY 取 277V（CAT III 277V，3E4WY 线电压 √3×277≈480V 恰在上限）；"
                     "2E3W1P 分相线电压=2×相电压，故取 240V（2×240=480V）"),
        ("2E3W1P 相角", "分相 A+C 反相：Vc 相角 180°（不是 120°）——已实跑用例 009_02 与 1320 老表口径"),
        ("未接线相", "🔴 未接线相不出压不出流：1E2W 的 B/C、2E3W1P 的 B 全置 0"
                     "（2026-07-27 实机：1E2W 给 B 相加压 → 源报 Ub 过载并进保护锁存）"),
        ("排序原则", "同频率的点连续（频率切换是唯一会掉源输出的动作，60Hz 排末尾每接线只切一次）；"
                      "同 PF 的点也连续，角度帧同值可跳发"),
        ("频率点", "50Hz 为主；60Hz 仅 2 点——写频率选择寄存器会触发电表延迟重启 30~60s"),
    ]
    ws_doc.column_dimensions["A"].width = 16
    ws_doc.column_dimensions["B"].width = 110
    for r, (k, v) in enumerate(doc, 1):
        ws_doc.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_doc.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True)

    wb.save(OUT_FILE)
    print(f"已生成 {OUT_FILE}")
    print(f"  test_case_mA: {ws.max_row - 1} 行 / test_case_mV: {ws_mv.max_row - 1} 行 "
          f"（{len(WIRE_TYPES)} 种接线 × {len(MATRIX)} 点）")


if __name__ == "__main__":
    main()
