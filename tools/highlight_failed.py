#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @File     :highlight_failed.py
# @Author   :lcs
# @Time     :2025/8/5
# @Desc     :

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def highlight_failed_cells(input_file, output_file):
    """
    传入Excel文件，检查单元格值是否包含 "Failed"，如果包含则高亮显示（背景标黄色，字体标记红色）。
    """
    # 加载工作簿
    wb = load_workbook(input_file)

    # 定义样式
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_font = Font(color='FF0000')

    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 遍历工作表中的所有单元格
        for row in ws.iter_rows():
            for cell in row:
                # 检查单元格值是否包含 "Failed" (不区分大小写)
                if cell.value and isinstance(cell.value, str) and "Failed" in cell.value:
                    cell.fill = yellow_fill
                    cell.font = red_font

    # 保存修改后的工作簿
    wb.save(output_file)
    print(f"处理完成，结果已保存到: {output_file}")

# 使用方法举例
# input_excel = "E3_3W_Y_Precision_Measure_20250715—RCT16-500.xlsx"  # 输入文件路径
# output_excel = "E3_3W_Y_Precision_Measure_20250715—RCT16-500.xlsx"  # 输出文件路径
# highlight_failed_cells(input_excel, output_excel)
