import openpyxl
import xlwt
from tools.log import root_path


def dcpara_addr_get(file_path, sheet_name):
    workbook = openpyxl.load_workbook(root_path + file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []
    data_dict = {}
    i = 0
    for row in sheet.rows:
        data.append([])
        for cell in row:
            data[i].append(cell.value)
        i += 1
    for i in range(1, len(data)):
        if sheet_name == 'System Infomation':
            if data[i][data[0].index('Descrption')] is None:
                continue
            data_dict[data[i][data[0].index('Descrption')]] = {}
            data_dict[data[i][data[0].index('Descrption')]]['Start(Dec)'] = data[i][data[0].index('Start(Dec)')]
            data_dict[data[i][data[0].index('Descrption')]]['Reg'] = data[i][data[0].index('Reg')]
        elif data[i][data[0].index('Descrption')] is None or data[i][data[0].index('Data type')] is None:
            continue
        else:
            data_dict[data[i][data[0].index('Descrption')] + ' ' + data[i][data[0].index('Data type')]] = {}
            data_dict[data[i][data[0].index('Descrption')] + ' ' + data[i][data[0].index('Data type')]]['Start(Dec)'] = \
            data[i][data[0].index('Start(Dec)')]
            data_dict[data[i][data[0].index('Descrption')] + ' ' + data[i][data[0].index('Data type')]]['Reg'] = \
                data[i][data[0].index('Reg')]
    # print(data_dict)
    return data_dict


def excel_write():
    my_workbook = xlwt.Workbook()
    sheet = my_workbook.add_sheet('AcuDC300 SYS TIME')
    sheet.write(0, 0, '0')
    my_workbook.save('TestCase_StabiLity_0809_case1_2.xlsx')


def data_read(file_path, sheet_name):
    workbook = openpyxl.load_workbook(root_path + file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []
    i = 0
    for row in sheet.rows:
        data.append([])
        for cell in row:
            data[i].append(cell.value)
        i += 1
    return data

def dcpara_4100addr_get(file_path, sheet_name):
    """
    AcuRev4100读取地址表中的地址

    """
    workbook = openpyxl.load_workbook(root_path + file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []
    data_dict = {}
    i = 0
    for row in sheet.rows:
        data.append([])
        for cell in row:
            data[i].append(cell.value)
        i += 1
    for i in range(1, len(data)):
        if data[i][data[0].index('Description')] is None or data[i][data[0].index('Data type')] is None:
            continue
        else:
            data_dict[data[i][data[0].index('Description')]] = {}
            data_dict[data[i][data[0].index('Description')]]['Start(Dec)'] = data[i][data[0].index('Start(Dec)')]
            data_dict[data[i][data[0].index('Description')]]['Reg'] = data[i][data[0].index('Reg')]
    return data_dict
