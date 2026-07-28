# -*- coding: utf-8 -*-
"""
精度测试引擎（无 GUI 层）

把控源页已连接、且线程安全的 XL9600 实例，适配成 fast_accuracy_test 框架要求的
统一控源接口 Source，复用框架的读表(pymodbus)/用例执行/精度算法/报告全部逻辑。

对外主要函数：
  - load_cases(path=None)                 读 xlsx 用例
  - save_cases_xlsx(cases, path=None)      把(可能在界面改过的)用例写回 xlsx
  - run_cases(dev, ..., cases, ...)        跑整批用例，出 Excel 报告
  - run_point(dev, ..., voltage, current)  单点快测，返回 CaseResult
"""

from __future__ import annotations

import dataclasses
import os
import sys
import time
import threading

import openpyxl

from fast_accuracy_test.core.config import load_config
from fast_accuracy_test.core.excel_input import read_cases, Case
from fast_accuracy_test.core.modbus_reader import make_client, ModbusReader
from fast_accuracy_test.core.case_runner import run_case
from fast_accuracy_test.core.report import write_report
from fast_accuracy_test.core.source_iface import Source

from xl9600 import OutputPoint


class StopRequested(Exception):
    """用户点了“停止”时，从 sleep 钩子抛出以中断当前用例。"""


# 界面用例表的列定义（字段名, 中文表头, 类型）；optfloat = 可空浮点
CASE_FIELDS = [
    ("test_case", "用例号", "str"),
    ("voltage", "电压(V)", "float"),
    ("current_1", "电流1(A)", "float"),
    ("current_2", "电流2(A)", "float"),
    ("wait_h", "等待(h)", "float"),
    ("voltage_accuracy", "压精度", "float"),
    ("current_accuracy", "流/能精度", "float"),
    ("power_accuracy", "功率精度", "float"),
    ("sample_cnt", "采样次数", "int"),
    ("sample_interval", "采样间隔", "float"),
    ("pulse_const", "脉冲常数", "optfloat"),
    ("pulse_error_acc", "脉冲误差", "optfloat"),
]

# 字段名 -> xlsx 列名（写回 xlsx 时用）
_FIELD_TO_XLSX = {
    "test_case": "test_case", "voltage": "Voltage",
    "current_1": "Current_1", "current_2": "Current_2", "wait_h": "等待时间(h)",
    "voltage_accuracy": "voltage_accuracy", "current_accuracy": "current_accuracy",
    "power_accuracy": "power_accuracy", "sample_cnt": "采样次数", "sample_interval": "采样间隔",
    "pulse_const": "能量脉冲常数", "pulse_error_acc": "能量脉冲误差",
}


# --------------------------------------------------------------------------- #
# 路径解析：打包成 exe 时数据文件放 exe 同目录，源码运行时放本文件目录
# --------------------------------------------------------------------------- #
def data_root() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def default_config_path() -> str:
    return os.path.join(data_root(), "config.json")


def resolve(path: str) -> str:
    return path if os.path.isabs(path) else os.path.normpath(os.path.join(data_root(), path))


# --------------------------------------------------------------------------- #
# 配置构建：以 config.json 为默认，再用界面读表设置覆盖
# --------------------------------------------------------------------------- #
def build_cfg(meter: dict, base_config_path: str | None = None) -> dict:
    cfg = load_config(base_config_path or default_config_path())
    for k in ("conn_mode", "device_model", "word_order", "settle_s", "read_retries"):
        if meter.get(k) not in (None, ""):
            cfg[k] = meter[k]
    if meter.get("rtu"):
        cfg["rtu"] = {**cfg.get("rtu", {}), **meter["rtu"]}
    if meter.get("tcp"):
        cfg["tcp"] = {**cfg.get("tcp", {}), **meter["tcp"]}
    cfg["is_dual"] = (cfg["device_model"] == "260")  # device_model 可能被覆盖，重新推导
    return cfg


# --------------------------------------------------------------------------- #
# XL9600 -> Source 适配器
# --------------------------------------------------------------------------- #
def _open_reader(cfg):
    """建读表器；连接失败时给出带连接信息和排查提示的清晰错误。"""
    try:
        return ModbusReader(cfg, make_client(cfg))
    except Exception as e:  # noqa: BLE001
        if cfg.get("conn_mode") == "rtu":
            s = cfg.get("rtu", {})
            where = f"串口 {s.get('port')} @ {s.get('baudrate')} 从站{s.get('slaveid')}"
        else:
            s = cfg.get("tcp", {})
            where = f"网口 {s.get('ip')}:{s.get('port')} 从站{s.get('slaveid')}"
        raise IOError(f"读表连接失败（{where}）：{e}。"
                      "请检查被检表是否上电/接线、串口号或IP、从站号是否正确") from e


def make_source(dev, rated_v: float, rated_i: float,
                base_params=None, pulse_read_count: int = 5) -> Source:
    """把控源页的 XL9600 包成框架 Source。

    - sour_output 按 额定/标定 基准换算成检定点%；电流为负 → 电能方向:反向。
    - set_pulse_const：把脉冲常数写进 XL9600 参数配置(基于 base_params 改 脉冲常数 后下发)。
    - measure_pulse_error：调 XL9600「误差读取」，返回均值(%)作为能量脉冲误差测量值，
      原始值供报告逐次列；读不到(nan)返回 None。
    """
    if rated_v <= 0 or rated_i <= 0:
        raise ValueError("额定电压/标定电流必须大于 0")
    state = {"samples": None}

    def out(voltage, current):
        if abs(voltage) < 1e-9 and abs(current) < 1e-9:
            dev.source_stop()
            return
        pv = abs(voltage) / rated_v * 100.0
        pi = abs(current) / rated_i * 100.0
        point = OutputPoint(电压检定点="%g%%" % pv, 电流检定点="%g%%" % pi,
                            电能方向="正向" if current >= 0 else "反向")
        dev.source_output(point)

    def set_pulse_const(const):
        if base_params is not None and const is not None:
            p = dataclasses.replace(base_params, **{"脉冲常数": "%g" % float(const)})
            dev.config(p)

    def measure_pulse_error(meter_const):
        # 设备测完全部 N 次误差才回应答，每次≥1个校验周期；给足等待时间：
        # 每次按 60s 预算 + 60s 裕量（低功率点校验一圈可能很慢）
        wait_s = pulse_read_count * 60.0 + 60.0
        res = dev.read_error(统计次数=pulse_read_count, recv_timeout=wait_s)
        state["samples"] = res.原始值
        mean = res.均值
        return None if mean != mean else mean   # nan → None(未测到)

    def last_pulse_samples():
        return state["samples"]

    return Source(out, dev.source_stop, pulse_fn=measure_pulse_error,
                  pulse_samples_fn=last_pulse_samples, set_pulse_const_fn=set_pulse_const)


# --------------------------------------------------------------------------- #
# 停止感知的 sleep：点“停止”后从 sleep 处抛 StopRequested 中断
# --------------------------------------------------------------------------- #
def _make_sleep(stop_event: threading.Event):
    def s(sec):
        end = time.time() + max(0.0, sec)
        while True:
            if stop_event.is_set():
                raise StopRequested()
            remain = end - time.time()
            if remain <= 0:
                return
            time.sleep(min(0.2, remain))
    return s


# --------------------------------------------------------------------------- #
# 用例读 / 写
# --------------------------------------------------------------------------- #
def load_cases(path: str | None = None) -> list:
    if path is None:
        path = resolve(load_config(default_config_path())["input_xlsx"])
    return read_cases(path)


def input_xlsx_path() -> str:
    return resolve(load_config(default_config_path())["input_xlsx"])


def export_cases_xlsx(cases: list, path: str) -> str:
    """把 cases 另存为一个全新的 xlsx（标准列头，含可选脉冲两列，可被重新加载）。

    与 save_cases_xlsx 的区别：本函数新建文件、写全部标准列，不依赖原表结构；
    适合导出当前(可能改过的)用例做模板或留档。返回写入路径。
    """
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    fields = [fld for fld, _, _ in CASE_FIELDS]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cases"
    ws.append([_FIELD_TO_XLSX[f] for f in fields])     # 表头用 xlsx 列名
    for c in cases:
        ws.append([getattr(c, f, None) for f in fields])
    wb.save(path)
    return path


def save_cases_xlsx(cases: list, path: str | None = None) -> str:
    """把 cases 写回 xlsx（按列名就地更新，保留其它列与格式）。返回写入路径。"""
    path = path or input_xlsx_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows())
    if not rows:
        raise ValueError(f"{path}: 空表")
    header = {cell.value: idx for idx, cell in enumerate(rows[0])}

    ci = 0
    for r in rows[1:]:
        tc_idx = header.get("test_case")
        if tc_idx is None or r[tc_idx].value in (None, ""):
            continue
        if ci >= len(cases):
            break
        c = cases[ci]
        ci += 1
        for field, col in _FIELD_TO_XLSX.items():
            if col in header:
                r[header[col]].value = getattr(c, field, None)
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# 跑整批用例
# --------------------------------------------------------------------------- #
def run_cases(dev, rated_v, rated_i, meter, cases, *,
              source_params=None, progress=None, stop_event=None) -> str:
    """跑全部用例并出报告。

    progress(kind, payload) 回调（在工作线程里调用，GUI 端需切回主线程）：
      'total'(n) / 'case_start'(name) / 'case_done'(CaseResult) /
      'stopped'() / 'report'(path) / 'warn'(msg)
    返回报告路径；无结果返回 ""。
    """
    stop_event = stop_event or threading.Event()

    def emit(kind, payload=None):
        if progress:
            progress(kind, payload)

    cfg = build_cfg(meter)
    result_dir = resolve(cfg["result_dir"])
    sleep_fn = _make_sleep(stop_event)

    reader = _open_reader(cfg)
    source = make_source(dev, rated_v, rated_i, base_params=source_params)

    # 测试开始先发一次参数配置（额定基准/接入方式），设备会记住
    if source_params is not None:
        dev.config(source_params)

    # 填了脉冲常数即测能量脉冲误差(本轮任一用例填了，则各行都带脉冲列，不填的行 N/A)
    include_pulse = any(c.pulse_const is not None for c in cases)
    results = []
    emit("total", len(cases))
    try:
        for c in cases:
            if stop_event.is_set():
                emit("stopped")
                break
            emit("case_start", c.test_case)
            try:
                res = run_case(c, reader, source, is_dual=cfg["is_dual"],
                               settle_s=float(cfg.get("settle_s", 5.0)),
                               include_pulse=include_pulse, sleep_fn=sleep_fn)
                results.append(res)
                emit("case_done", res)
            except StopRequested:
                emit("stopped")
                break
            except Exception as e:  # noqa: BLE001
                emit("warn", f"用例 {c.test_case} 执行失败: {e}")
        path = write_report(results, result_dir, cfg["device_model"]) if results else ""
        if path:
            emit("report", path)
        return path
    finally:
        try:
            source.output(0, 0)
            source.stop()
        except Exception:  # noqa: BLE001
            pass
        try:
            reader.close()
        except Exception:  # noqa: BLE001
            pass


# --------------------------------------------------------------------------- #
# 单点快测
# --------------------------------------------------------------------------- #
def run_point(dev, rated_v, rated_i, meter, voltage, current, *,
              vacc=0.001, iacc=0.075, pacc=0.075, source_params=None, stop_event=None):
    """只测一个电压/电流点，返回 CaseResult（不出 Excel）。"""
    stop_event = stop_event or threading.Event()
    cfg = build_cfg(meter)
    sleep_fn = _make_sleep(stop_event)

    reader = _open_reader(cfg)
    source = make_source(dev, rated_v, rated_i, base_params=source_params)
    if source_params is not None:
        dev.config(source_params)

    case = Case(test_case="单点快测", voltage=voltage, current_1=current, current_2=0.0,
                wait_h=0.0, voltage_accuracy=vacc, current_accuracy=iacc, power_accuracy=pacc,
                sample_cnt=20, sample_interval=0.1)
    try:
        return run_case(case, reader, source, is_dual=cfg["is_dual"],
                        settle_s=float(cfg.get("settle_s", 5.0)), sleep_fn=sleep_fn)
    finally:
        try:
            source.output(0, 0)
            source.stop()
        except Exception:  # noqa: BLE001
            pass
        try:
            reader.close()
        except Exception:  # noqa: BLE001
            pass
