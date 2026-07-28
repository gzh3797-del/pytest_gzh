"""excel_writer.py — webtestcase_manual_to_auto skill 的 Excel 辅助工具

适配真实共享用例集 schema（sheet「测试用例」）：
    模块 | 子模块 | FTS编号 | FTS名称(用例编号) | 用例编号 | 用例标题 | 预置条件 |
    测试步骤 | 预期结果 | 用例级别 | 测试是否通过 | 测试负责人 |
    自动化(是/否) | 自动化类型 | 是否通过连跑 | 备注

按列名（而非列序）匹配，兼容列顺序变化。--init 幂等追加 3 个 Claude 列，
不改动任何现有列。

用法：
  python excel_writer.py --list-modules <excel_path>
  python excel_writer.py --init <excel_path>
  python excel_writer.py --read <excel_path> --module <模块名>
  python excel_writer.py --read <excel_path> --case <用例编号>
  python excel_writer.py --write <excel_path> --data-file <json_path>
  python excel_writer.py --write-result <excel_path> --results-file <json_path>
  python excel_writer.py --status <excel_path>
"""

import argparse
import io
import json
import sys
from pathlib import Path

# 统一 stdout 为 UTF-8，避免 Windows 终端乱码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)

# ── 现有列名（真实 schema，只读，绝不修改）──────────────────────────────────
COL_MODULE = "模块"
COL_SUBMODULE = "子模块"
COL_CASE_ID = "用例编号"
COL_TITLE = "用例标题"
COL_PRECOND = "预置条件"
COL_STEPS = "测试步骤"
COL_EXPECTED = "预期结果"
COL_LEVEL = "用例级别"
COL_AUTO = "自动化(是/否)"
COL_AUTO_TYPE = "自动化类型"

# ── 新增 Claude 列（--init 幂等追加）────────────────────────────────────────
COL_CLAUDE_NOTE = "需补充信息(claude识别回填)"
COL_USER_REPLY = "用户答复(澄清信息)"
COL_DEBUG_RESULT = "自动化脚本调试结果"

CLAUDE_COLS = [COL_CLAUDE_NOTE, COL_USER_REPLY, COL_DEBUG_RESULT]

COLOR_GREEN = "006400"   # 深绿：已理解，可生成
COLOR_RED = "FF0000"     # 红色：信息缺失/步骤不清，需澄清（附根因）
COLOR_ORANGE = "FFA500"  # 橙色：半自动化，部分步骤需人工介入


# ── 工具函数 ────────────────────────────────────────────────────────────────

def _load(path: str):
    p = Path(path)
    if not p.exists():
        _exit_error(f"文件不存在: {path}")
    return openpyxl.load_workbook(str(p))


def _save(wb, path: str) -> None:
    """先写入内存缓冲再落盘，避免半写文件损坏。"""
    buf = io.BytesIO()
    wb.save(buf)
    with open(path, "wb") as f:
        f.write(buf.getvalue())


def _normalize_col(s: str) -> str:
    """规范化列名：去换行、去首尾空格、统一中英文逗号。"""
    return s.replace("\n", "").replace("\r", "").replace("，", ",").strip()


def _col_map(ws) -> dict:
    """返回 {规范化列名: 列索引(1-based)} 的映射。"""
    return {
        _normalize_col(cell.value): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def _exit_error(msg: str) -> None:
    print(json.dumps({"error": msg}, ensure_ascii=False))
    sys.exit(1)


def _str(v):
    return str(v).strip() if v is not None else None


def _font_color(cell):
    """读取单元格字体色（末 6 位去 alpha），无则 None。"""
    if cell.font and cell.font.color and cell.font.color.rgb:
        rgb = cell.font.color.rgb
        if isinstance(rgb, str):
            return rgb[-6:]
    return None


# ── 子命令实现 ───────────────────────────────────────────────────────────────

def cmd_list_modules(excel_path: str) -> None:
    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)
    if COL_MODULE not in cm:
        _exit_error(f"列 '{COL_MODULE}' 不存在")
    mod_idx = cm[COL_MODULE]
    auto_idx = cm.get(COL_AUTO)

    counts: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        module = _str(row[mod_idx - 1])
        if not module:
            continue
        rec = counts.setdefault(module, {"module": module, "total": 0, "auto": 0})
        rec["total"] += 1
        if auto_idx and _str(row[auto_idx - 1]) == "是":
            rec["auto"] += 1

    result = sorted(counts.values(), key=lambda x: -x["auto"])
    print(json.dumps(result, ensure_ascii=False, indent=2))


def cmd_init(excel_path: str) -> None:
    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)
    added = {}

    # 追加位置紧跟最后一个「有表头文字」的列，而非 ws.max_column
    # （真实用例集存在大量无表头的幻影空列，max_column 会高达 16384）。
    next_col = (max(cm.values()) if cm else 0) + 1

    for col_name in CLAUDE_COLS:
        if col_name in cm:
            added[col_name] = {"col": cm[col_name], "created": False}
        else:
            cell = ws.cell(row=1, column=next_col, value=col_name)
            cell.font = Font(bold=True)
            cm[col_name] = next_col
            added[col_name] = {"col": next_col, "created": True}
            next_col += 1

    _save(wb, excel_path)
    print(json.dumps(added, ensure_ascii=False, indent=2))


def _read_rows(ws, cm: dict, predicate) -> list:
    def _get(row_vals, col_name):
        if col_name not in cm:
            return None
        return _str(row_vals[cm[col_name] - 1])

    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not predicate(row, _get):
            continue

        note_color = None
        if COL_CLAUDE_NOTE in cm:
            note_color = _font_color(ws.cell(row=row_idx, column=cm[COL_CLAUDE_NOTE]))

        results.append({
            "row": row_idx,
            "module": _get(row, COL_MODULE),
            "submodule": _get(row, COL_SUBMODULE),
            "case_id": _get(row, COL_CASE_ID),
            "title": _get(row, COL_TITLE),
            "precondition": _get(row, COL_PRECOND),
            "steps": _get(row, COL_STEPS),
            "expected": _get(row, COL_EXPECTED),
            "level": _get(row, COL_LEVEL),
            "auto": _get(row, COL_AUTO),
            "auto_type": _get(row, COL_AUTO_TYPE),
            "claude_note": _get(row, COL_CLAUDE_NOTE),
            "claude_color": note_color,
            "user_reply": _get(row, COL_USER_REPLY),
            "debug_result": _get(row, COL_DEBUG_RESULT),
        })
    return results


def cmd_read(excel_path: str, module: str = None, case: str = None) -> None:
    if module is None and case is None:
        _exit_error("--read 需提供 --module 或 --case")

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)

    def pred(row, get):
        if module is not None:
            return get(row, COL_MODULE) == module
        return get(row, COL_CASE_ID) == case

    print(json.dumps(_read_rows(ws, cm, pred), ensure_ascii=False, indent=2))


def cmd_write(excel_path: str, data_json: str) -> None:
    """写「需补充信息(claude识别回填)」列（分析结论 + 字体色）。
    data: [{"row": int, "text": str, "color": "006400/FF0000/FFA500"}, ...]
    """
    try:
        data = json.loads(data_json)
    except json.JSONDecodeError as e:
        _exit_error(f"JSON 解析失败: {e}")
        return

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)
    if COL_CLAUDE_NOTE not in cm:
        _exit_error(f"列 '{COL_CLAUDE_NOTE}' 不存在，请先运行 --init")

    col_idx = cm[COL_CLAUDE_NOTE]
    written = 0
    for item in data:
        row_num = item.get("row")
        if not row_num:
            continue
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = item.get("text", "")
        cell.font = Font(color=item.get("color", COLOR_RED))
        cell.alignment = Alignment(wrap_text=True)
        written += 1

    _save(wb, excel_path)
    print(json.dumps({"written": written, "file": excel_path}, ensure_ascii=False))


def cmd_write_result(excel_path: str, results_json: str) -> None:
    """写「自动化脚本调试结果」列（按 用例编号 匹配，支持一号多行）。
    results: [{"case_id": "TestCase_...", "value": "是 / 跳过|原因 / 调试失败|原因"}, ...]
    """
    try:
        results = json.loads(results_json)
    except json.JSONDecodeError as e:
        _exit_error(f"JSON 解析失败: {e}")
        return

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)
    if COL_CASE_ID not in cm:
        _exit_error(f"列 '{COL_CASE_ID}' 不存在")
    if COL_DEBUG_RESULT not in cm:
        _exit_error(f"列 '{COL_DEBUG_RESULT}' 不存在，请先运行 --init")

    case_id_col = cm[COL_CASE_ID]
    result_col = cm[COL_DEBUG_RESULT]

    case_row_map: dict = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cid = _str(row[case_id_col - 1])
        if cid:
            case_row_map.setdefault(cid, []).append(row_idx)

    written = 0
    not_found = []
    for item in results:
        case_id = (item.get("case_id") or "").strip()
        value = item.get("value") or item.get("result") or ""
        if case_id not in case_row_map:
            not_found.append(case_id)
            continue
        for row_num in case_row_map[case_id]:
            cell = ws.cell(row=row_num, column=result_col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True)
            written += 1

    _save(wb, excel_path)
    print(json.dumps(
        {"written": written, "not_found": not_found, "file": excel_path},
        ensure_ascii=False,
    ))


def cmd_status(excel_path: str) -> None:
    """统计各模块中 自动化(是/否)=是 的用例按 Claude 列字体色的分布。"""
    from collections import defaultdict

    wb = _load(excel_path)
    ws = wb.active
    cm = _col_map(ws)

    mod_col = cm.get(COL_MODULE)
    auto_col = cm.get(COL_AUTO)
    note_col = cm.get(COL_CLAUDE_NOTE)
    reply_col = cm.get(COL_USER_REPLY)
    if not all([mod_col, auto_col, note_col, reply_col]):
        _exit_error("必要列不存在，请先运行 --init")

    stats = defaultdict(lambda: {
        "green": 0, "red_with_reply": 0, "red_no_reply": 0,
        "orange": 0, "no_color": 0,
    })

    for row in ws.iter_rows(min_row=2):
        module = _str(row[mod_col - 1].value)
        auto = _str(row[auto_col - 1].value)
        if not module or auto != "是":
            continue

        color = _font_color(row[note_col - 1])
        reply = _str(row[reply_col - 1].value)
        s = stats[module]
        if color == COLOR_GREEN:
            s["green"] += 1
        elif color == COLOR_RED:
            s["red_with_reply" if reply else "red_no_reply"] += 1
        elif color == COLOR_ORANGE:
            s["orange"] += 1
        else:
            s["no_color"] += 1

    result = []
    for mod, s in sorted(stats.items(),
                         key=lambda x: x[1]["red_with_reply"] + x[1]["red_no_reply"],
                         reverse=True):
        result.append({
            "module": mod,
            **s,
            "total_auto": sum(s.values()),
            "pending_reanalyze": s["red_with_reply"],
            "pending_user_input": s["red_no_reply"],
        })

    print(json.dumps(result, ensure_ascii=False, indent=2))


# ── CLI 入口 ─────────────────────────────────────────────────────────────────

def _read_json_arg(inline: str, file_path: str) -> str:
    if file_path:
        return Path(file_path).read_text(encoding="utf-8")
    if inline:
        return inline
    _exit_error("需提供 --data/--results 或 --data-file/--results-file")
    return ""


def main() -> None:
    argv = sys.argv[1:]
    if not argv:
        print("用法: excel_writer.py --list-modules <path>")
        print("      excel_writer.py --init <path>")
        print("      excel_writer.py --read <path> --module <模块名> | --case <用例编号>")
        print("      excel_writer.py --write <path> --data-file <json_path>")
        print("      excel_writer.py --write-result <path> --results-file <json_path>")
        print("      excel_writer.py --status <path>")
        return

    cmd, rest = argv[0], argv[1:]

    if cmd == "--list-modules":
        if not rest:
            _exit_error("缺少 excel_path 参数")
        cmd_list_modules(rest[0])

    elif cmd == "--init":
        if not rest:
            _exit_error("缺少 excel_path 参数")
        cmd_init(rest[0])

    elif cmd == "--read":
        p = argparse.ArgumentParser()
        p.add_argument("excel_path")
        p.add_argument("--module")
        p.add_argument("--case")
        a = p.parse_args(rest)
        cmd_read(a.excel_path, a.module, a.case)

    elif cmd == "--write":
        p = argparse.ArgumentParser()
        p.add_argument("excel_path")
        p.add_argument("--data")
        p.add_argument("--data-file")
        a = p.parse_args(rest)
        cmd_write(a.excel_path, _read_json_arg(a.data, a.data_file))

    elif cmd == "--write-result":
        p = argparse.ArgumentParser()
        p.add_argument("excel_path")
        p.add_argument("--results")
        p.add_argument("--results-file")
        a = p.parse_args(rest)
        cmd_write_result(a.excel_path, _read_json_arg(a.results, a.results_file))

    elif cmd == "--status":
        if not rest:
            _exit_error("缺少 excel_path 参数")
        cmd_status(rest[0])

    else:
        _exit_error(
            f"未知命令: {cmd}。支持："
            "--list-modules / --init / --read / --write / --write-result / --status"
        )


if __name__ == "__main__":
    main()
