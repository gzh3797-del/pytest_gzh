"""手工用例 xlsx 的读取与回填 —— "手工用例 → 自动化"流程的 I/O 端。

模板列(Sheet1, 第 1 行表头; 一律按*表头名*定位, 不依赖列号):
  模块 | 子模块 | 用例编号 | 用例标题 | 预置条件 | 测试步骤 | 预期结果 | 是否可自动化 | 调试结果 | 测试结果

  read_cases(xlsx, only_automatable=True) -> list[dict]   读用例(默认只收"是否可自动化==是")
  write_back(xlsx, 用例编号, debug=None, result=None)     回填「调试结果」/「测试结果」列

CLI:
  python -m comm.ctl_acuview.manual_xlsx <xlsx>                            # 列出可自动化用例
  python -m comm.ctl_acuview.manual_xlsx <xlsx> --case <编号> --debug "问题" --result 未通过
"""
from __future__ import annotations

import sys

import openpyxl

CASE_ID_COL = "用例编号"
DEBUG_COL = "调试结果"
RESULT_COL = "测试结果"
AUTO_COL = "是否可自动化"
META_KEYS = ("用例编号", "用例标题", "预置条件", "测试步骤", "预期结果")


def _headers(ws) -> dict[str, int]:
    """表头名 -> 列号(1 起)。"""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            out[str(v).strip()] = c
    return out


def _open(xlsx: str):
    try:
        return openpyxl.load_workbook(xlsx)
    except PermissionError as exc:
        raise RuntimeError(
            f"无法打开 {xlsx}(可能正在 Excel 中打开 / 存在 ~$ 锁文件)。请关闭后重试。") from exc


def read_cases(xlsx: str, only_automatable: bool = True, sheet: str | None = None) -> list[dict]:
    """读手工用例为结构化列表; 每条含表头各列 + 'row'(行号)。"""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    hdr = _headers(ws)
    if CASE_ID_COL not in hdr:
        raise ValueError(f"{xlsx} 缺表头列「{CASE_ID_COL}」; 现有表头: {list(hdr)}")
    cases = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr[CASE_ID_COL]).value
        if cid is None or not str(cid).strip():
            continue
        rec = {"row": r}
        for name, col in hdr.items():
            v = ws.cell(r, col).value
            rec[name] = str(v).strip() if v is not None else ""
        if only_automatable:
            auto = rec.get(AUTO_COL, "")
            if str(auto).strip() not in ("是", "Y", "y", "yes", "Yes", "TRUE", "True", "1"):
                continue
        cases.append(rec)
    return cases


def case_meta(rec: dict) -> dict:
    """从一行用例提取引擎需要的 5 项备注(编号/标题/预置条件/测试步骤/预期结果)。"""
    return {
        "编号": rec.get("用例编号", ""),
        "标题": rec.get("用例标题", ""),
        "预置条件": rec.get("预置条件", ""),
        "测试步骤": rec.get("测试步骤", ""),
        "预期结果": rec.get("预期结果", ""),
    }


def write_back(xlsx: str, case_id: str, debug: str | None = None,
               result: str | None = None, sheet: str | None = None) -> bool:
    """按用例编号定位行, 回填「调试结果」(debug) / 「测试结果」(result)。返回是否写到。"""
    wb = _open(xlsx)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    hdr = _headers(ws)
    if CASE_ID_COL not in hdr:
        raise ValueError(f"{xlsx} 缺表头「{CASE_ID_COL}」")
    cid_col = hdr[CASE_ID_COL]
    target_row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, cid_col).value or "").strip() == str(case_id).strip():
            target_row = r
            break
    if target_row is None:
        raise ValueError(f"未在 {xlsx} 找到用例编号 {case_id}")

    if debug is not None:
        if DEBUG_COL not in hdr:
            raise ValueError(f"{xlsx} 缺表头「{DEBUG_COL}」")
        ws.cell(target_row, hdr[DEBUG_COL]).value = debug
    if result is not None:
        if RESULT_COL not in hdr:
            raise ValueError(f"{xlsx} 缺表头「{RESULT_COL}」")
        ws.cell(target_row, hdr[RESULT_COL]).value = result

    try:
        wb.save(xlsx)
    except PermissionError as exc:
        raise RuntimeError(f"保存失败: {xlsx} 正被占用(请关闭 Excel)。") from exc
    return True


def main(argv):
    sys.stdout.reconfigure(encoding="utf-8")
    if not argv:
        print("用法: python -m comm.ctl_acuview.manual_xlsx <xlsx> [--case 编号 --debug 文本 --result 通过/未通过]")
        return 2
    xlsx = argv[0]
    rest = argv[1:]
    if "--case" in rest:
        i = rest.index("--case")
        cid = rest[i + 1]
        debug = result = None
        if "--debug" in rest:
            debug = rest[rest.index("--debug") + 1]
        if "--result" in rest:
            result = rest[rest.index("--result") + 1]
        write_back(xlsx, cid, debug=debug, result=result)
        print(f"已回填 {cid}: 调试结果={debug!r} 测试结果={result!r}")
        return 0
    cases = read_cases(xlsx, only_automatable=False)
    autoc = sum(1 for c in cases if str(c.get(AUTO_COL, "")).strip() == "是")
    print(f"{xlsx}: 共 {len(cases)} 条, 可自动化 {autoc} 条")
    for c in cases:
        print(f"  [{c.get(AUTO_COL,'?')}] {c.get('用例编号')} | {c.get('用例标题','')[:50]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
