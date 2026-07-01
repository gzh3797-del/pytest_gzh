"""统一模型构建器。

把两份素材合并成机器可用的模型：
  1) AcuRev-1320 Modbus 地址表 (Excel, 33 sheet) -> registers.json
       以 *地址(十进制)* 为主键(读写最可靠)，名称为辅。
  2) Default_1.01 1.json (上位机控件模型)     -> pages.json
       每个页面: Page_Detail / Read_Sequence(地址块) / Read_List(寄存器名) / 控件(类型+相对坐标+data名)

设计要点:
  - 读取页面绑定寄存器靠 **地址**(Read_Sequence)，不依赖名称匹配 —— 因为 Excel 用人类描述
    ("System Frequency")，JSON 用 UPPER_SNAKE("PHASE_A_VOLTAGE")，两者命名风格不同。
  - 名称索引仅作辅助/写目标定位，匹配不上的记录到 unmatched 供人工补别名。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

from .config import consume_config_arg, get_config

# Excel 表头列名(行0)
HEADER_KEYS = {"Start(Dec)", "Description", "Data type"}
# 非寄存器 sheet(无地址表)
SKIP_SHEETS = {"Version history", "Overview"}


def _norm_name(s: str | None) -> str:
    """名称规范化：用于跨表匹配。大写、非字母数字转下划线、压缩。"""
    if not s:
        return ""
    s = re.sub(r"\s+", " ", str(s)).strip()
    key = re.sub(r"[^0-9a-zA-Z]+", "_", s).strip("_").upper()
    return key


def _parse_range_enum(range_raw: str | None) -> tuple[dict | None, list | None]:
    """尝试把 Range 文本解析为枚举。

    返回 (code_label_map, value_options):
      - "0: even 1: odd 2: none"       -> ({0:'even',1:'odd',2:'none'}, None)  码->标签
      - "9600\\19200\\38400\\57600..." -> (None, ['9600','19200',...])         可选值列表
    两者均解析不出返回 (None, None)。
    """
    if not range_raw:
        return None, None
    text = str(range_raw).strip()
    # 形式1: "code: label"(冒号映射，最常见的设置项)
    pairs = re.findall(r"(\d+)\s*[:：]\s*([^\n\r]+?)(?=\s+\d+\s*[:：]|$)", text)
    if len(pairs) >= 2:
        return {int(k): v.strip() for k, v in pairs}, None
    # 形式2: 反斜杠/竖线分隔的取值列表(去掉尾部括号说明)
    if "\\" in text or "|" in text:
        head = re.split(r"[(（]", text, maxsplit=1)[0]
        parts = [p.strip() for p in re.split(r"[\\|]", head) if p.strip()]
        if len(parts) >= 2:
            return None, parts
    return None, None


def _find_header_row(ws) -> int | None:
    """在前 5 行内定位表头行。"""
    for r in range(min(5, ws.max_row or 0)):
        vals = {str(c.value).strip() for c in ws[r + 1] if c.value is not None}
        if HEADER_KEYS.issubset(vals):
            return r + 1
    return None


def parse_excel(excel_path: Path) -> dict[str, Any]:
    """解析 Excel -> {by_addr, by_name, unmatched_names, sheets}。"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    by_addr: dict[int, dict] = {}
    by_name: dict[str, list[int]] = {}
    sheet_summaries: dict[str, int] = {}

    for sheet in wb.sheetnames:
        if sheet in SKIP_SHEETS:
            continue
        ws = wb[sheet]
        hdr = _find_header_row(ws)
        if hdr is None:
            continue
        # 列索引(基于表头)
        header = {str(c.value).strip(): i for i, c in enumerate(ws[hdr]) if c.value is not None}
        ci_block = 0
        ci = {
            "hex": header.get("Start(Hex)"),
            "end_hex": header.get("End(Hex)"),
            "dec": header.get("Start(Dec)"),
            "end_dec": header.get("End(Dec)"),
            "desc": header.get("Description"),
            "dtype": header.get("Data type"),
            "rw": header.get("RW"),
            "regnum": header.get("Reg Num"),
            "range": header.get("Range") if "Range" in header else header.get("Energy Parameter"),
        }
        if ci["dec"] is None or ci["desc"] is None:
            continue

        cur_block = ""
        count = 0
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            def cell(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            blk = cell(ci_block)
            if blk not in (None, ""):
                cur_block = str(blk).strip()

            dec = cell(ci["dec"])
            desc = cell(ci["desc"])
            if dec in (None, "") or desc in (None, ""):
                continue
            try:
                addr = int(dec)
            except (TypeError, ValueError):
                continue

            rw = str(cell(ci["rw"]) or "").strip()
            range_raw = cell(ci["range"])
            enum_map, enum_options = _parse_range_enum(range_raw)
            entry = {
                "addr": addr,
                "addr_hex": str(cell(ci["hex"]) or "").strip(),
                "end_dec": int(cell(ci["end_dec"])) if str(cell(ci["end_dec"]) or "").strip().isdigit() else addr,
                "name": _norm_name(desc),
                "description": re.sub(r"\s+", " ", str(desc)).strip(),
                "dtype": str(cell(ci["dtype"]) or "").strip(),
                "rw": rw,
                "reg_num": int(cell(ci["regnum"])) if str(cell(ci["regnum"]) or "").strip().isdigit() else 1,
                "range": re.sub(r"\s+", " ", str(range_raw)).strip() if range_raw not in (None, "") else "",
                "enum": enum_map,
                "enum_options": enum_options,
                "sheet": sheet,
                "block": cur_block,
            }
            # 同地址若已存在(不同 sheet 重复定义)，保留首个，记录在 dup
            if addr not in by_addr:
                by_addr[addr] = entry
                by_name.setdefault(entry["name"], []).append(addr)
            count += 1
        sheet_summaries[sheet] = count

    return {
        "by_addr": by_addr,
        "by_name": by_name,
        "sheets": sheet_summaries,
    }


def _is_widget(v: Any) -> bool:
    return isinstance(v, dict) and "widgetType" in v and "basicAttribute" in v


def parse_json_pages(json_path: Path) -> dict[str, Any]:
    """解析控件 JSON -> {nav, pages}。"""
    with open(json_path, "r", encoding="utf-8") as fh:
        d = json.load(fh)

    nav = d.get("Select_Tree", {})
    pages: dict[str, Any] = {}

    for page_name, page in d.items():
        if page_name == "Select_Tree" or not isinstance(page, dict):
            continue
        if "Page_Detail" not in page:
            continue

        widgets = {}
        for wname, wval in page.items():
            if _is_widget(wval):
                ba = wval.get("basicAttribute", {})
                widgets[wname] = {
                    "type": wval.get("widgetType"),
                    "data": ba.get("data"),
                    "x": ba.get("x"),
                    "y": ba.get("y"),
                    "w": ba.get("width"),
                    "h": ba.get("height"),
                }

        # 收集所有 Read_Sequence / Read_List 变体
        read_seqs = {k: v for k, v in page.items() if k.endswith("Read_Sequence")}
        read_lists = {k: v for k, v in page.items() if k.endswith("Read_List")}

        pages[page_name] = {
            "page_detail": page.get("Page_Detail", {}),
            "read_sequence": read_seqs,
            "read_list": read_lists,
            "widgets": widgets,
        }

    return {"nav": nav, "pages": pages}


def iter_read_list_names(read_list_block: dict):
    """展平 Read_List。

    结构形如 {"Read_List": {"Communication_Setting": [names...], ...}} 或
             {"Read_List": [names...]}。
    逐个 yield (section, name)。
    """
    for list_key, val in read_list_block.items():
        if isinstance(val, list):
            for nm in val:
                yield (list_key, nm)
        elif isinstance(val, dict):
            for section, names in val.items():
                if isinstance(names, list):
                    for nm in names:
                        yield (section, nm)


def cross_reference(excel: dict, jpages: dict) -> dict:
    """把 JSON 各页 Read_List 的寄存器名解析到 Excel 地址；统计匹配率。"""
    by_name = excel["by_name"]
    unmatched = {}
    total = 0
    hit = 0
    for page_name, page in jpages["pages"].items():
        for _section, nm in iter_read_list_names(page["read_list"]):
            total += 1
            key = _norm_name(nm)
            if key in by_name:
                hit += 1
            else:
                unmatched.setdefault(page_name, []).append(nm)
    return {
        "total_read_list_names": total,
        "matched_count": hit,
        "match_rate": round(hit / total, 4) if total else 0.0,
        "unmatched": unmatched,
    }


def build(verbose: bool = True) -> dict:
    cfg = get_config()
    excel = parse_excel(cfg.excel_path)
    jpages = parse_json_pages(cfg.json_path)
    xref = cross_reference(excel, jpages)

    cfg.spec_dir.mkdir(parents=True, exist_ok=True)

    registers = {
        "source": cfg.excel_path.name,
        "count": len(excel["by_addr"]),
        "by_addr": {str(a): e for a, e in sorted(excel["by_addr"].items())},
        "by_name": excel["by_name"],
        "sheets": excel["sheets"],
    }
    with open(cfg.registers_json, "w", encoding="utf-8") as fh:
        json.dump(registers, fh, ensure_ascii=False, indent=1)

    pages_out = {
        "source": cfg.json_path.name,
        "nav": jpages["nav"],
        "pages": jpages["pages"],
        "xref": xref,
    }
    with open(cfg.pages_json, "w", encoding="utf-8") as fh:
        json.dump(pages_out, fh, ensure_ascii=False, indent=1)

    if verbose:
        print(f"[spec] registers: {registers['count']}  ->  {cfg.registers_json}")
        print(f"[spec] pages    : {len(jpages['pages'])}  ->  {cfg.pages_json}")
        print(f"[spec] Read_List 名称匹配率: {xref['matched_count']}/{xref['total_read_list_names']} "
              f"= {xref['match_rate']*100:.1f}%")
        # 打印 Communication 页寄存器映射作为抽检
        comm = jpages["pages"].get("Communication")
        if comm:
            print("\n[抽检] Communication 页 Read_List -> 地址:")
            cur_sec = None
            for section, nm in iter_read_list_names(comm["read_list"]):
                if section != cur_sec:
                    print(f"  <{section}>")
                    cur_sec = section
                addrs = excel["by_name"].get(_norm_name(nm))
                if addrs:
                    e = excel["by_addr"][addrs[0]]
                    print(f"    {nm:<46} addr={e['addr']:<6}(0x{e['addr']:04X}) "
                          f"{e['dtype']:<9} {e['rw']:<4} range={e['range'][:28]}")
                else:
                    print(f"    {nm:<46} [未匹配到地址]")
    return {"registers": registers, "pages": pages_out}


# ---- 运行期便捷加载(供其它模块使用) ----
def load_spec() -> tuple[dict, dict]:
    """加载已生成的 registers.json + pages.json。"""
    cfg = get_config()
    if not cfg.registers_json.exists() or not cfg.pages_json.exists():
        build(verbose=False)
    with open(cfg.registers_json, encoding="utf-8") as fh:
        registers = json.load(fh)
    with open(cfg.pages_json, encoding="utf-8") as fh:
        pages = json.load(fh)
    return registers, pages


def reg_by_addr(registers: dict, addr: int) -> dict | None:
    return registers["by_addr"].get(str(addr))


def reg_by_name(registers: dict, name: str) -> dict | None:
    addrs = registers["by_name"].get(_norm_name(name))
    if not addrs:
        return None
    return registers["by_addr"].get(str(addrs[0]))


if __name__ == "__main__":
    _args = consume_config_arg(sys.argv[1:])   # 取出 --config <项目>/config_acuview.yaml
    if "--build" in _args or not _args:
        build(verbose=True)
