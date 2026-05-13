# -*- coding: utf-8 -*-
"""
enip_comparator.py — EtherNet/IP vs Modbus TCP 数值比对

流程：
  1. 解析 EDS + 模板 SNMP 列，执行参数范围检查
  2. 并发读取 EtherNet/IP Assembly 与 Modbus TCP
  3. 逐参数按容差比对
  4. 输出控制台摘要 + HTML 报告（范围检查 + 数值比对）

用法：
  python EtherNetIP/enip_comparator.py
  python EtherNetIP/enip_comparator.py --quick          # 只比对前 30 个参数
  python EtherNetIP/enip_comparator.py --keys FREQ_Hz VLN_a_V
"""
from __future__ import annotations

import asyncio
import html
import logging
import sys
import time
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

import config
from enip_reader import EnipReader, EnipResult, parse_eds
from modbus_reader import ModbusReader, ModbusResult, get_reader

log = logging.getLogger(__name__)

# ─── EDS 路径 ────────────────────────────────────────────────────────────────
EDS_PATH_MAP: dict[str, str] = {
    "AcuRev4100": r"C:\Users\ZihanGao\Desktop\web2sprint2\BacnetIP\Datas\awsdatas\eds\AcuRev-4100.eds",
}

# ─── EtherNet/IP 网关配置 ────────────────────────────────────────────────────
ENIP_HOST_MAP: dict[str, str] = {
    "AcuRev4100": "192.168.2.63",
}
ENIP_SLOT = 0


# ─────────────────────────────────────────────────────────────────────────────
# 模板 SNMP 参数范围读取
# ─────────────────────────────────────────────────────────────────────────────

def get_snmp_params(template_dir: str, device_name: str) -> list[str]:
    """从模板 blockParams 读取 SNMP 列非空的参数名列表。"""
    needle = device_name.lower().replace("-", "").replace("_", "")
    for p in Path(template_dir).rglob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        if needle in p.stem.lower().replace("-", "").replace("_", ""):
            break
    else:
        raise FileNotFoundError(f"未找到 {device_name} 的模板文件")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(p), data_only=True)
    ws = wb["blockParams"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    snmp_col  = headers.index("SNMP")
    param_col = headers.index("paramType")
    params = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[snmp_col] and str(row[snmp_col]).strip() and row[param_col]:
            params.append(str(row[param_col]).strip())
    wb.close()
    return params


# ─────────────────────────────────────────────────────────────────────────────
# 数据结构
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ScopeReport:
    template_count:   int
    eds_count:        int
    matched_keys:     list[str]
    missing_from_eds: list[str]   # 模板有但 EDS 没有
    extra_in_eds:     list[str]   # EDS 有但模板未标注 SNMP

    @property
    def scope_ok(self) -> bool:
        return not self.missing_from_eds and not self.extra_in_eds


@dataclass
class CompareResult:
    param_key:    str
    enip_value:   Optional[float] = None
    modbus_value: Optional[float] = None
    enip_error:   str = ""
    modbus_error: str = ""
    diff_abs:     Optional[float] = None
    diff_pct:     Optional[float] = None
    status:       str = ""   # PASS | FAIL | ENIP_ERR | MODBUS_ERR | BOTH_ERR

    @property
    def ok(self) -> bool:
        return self.status == "PASS"


# ─────────────────────────────────────────────────────────────────────────────
# 比对逻辑
# ─────────────────────────────────────────────────────────────────────────────

def _compare_one(
    param_key: str,
    er: EnipResult,
    mr: ModbusResult,
) -> CompareResult:
    cr = CompareResult(param_key=param_key)

    if not er.ok:
        cr.enip_error = er.error
    if not mr.ok:
        cr.modbus_error = mr.error

    if not er.ok and not mr.ok:
        cr.status = "BOTH_ERR"; return cr
    if not er.ok:
        cr.status = "ENIP_ERR"; return cr
    if not mr.ok:
        cr.status = "MODBUS_ERR"; return cr

    ev, mv = er.value, mr.value
    cr.enip_value   = ev
    cr.modbus_value = mv

    diff = abs(ev - mv)
    ref  = max(abs(ev), abs(mv))
    cr.diff_abs = diff
    cr.diff_pct = (diff / ref * 100) if ref > 1e-12 else 0.0

    tol = max(config.TOLERANCE_ABSOLUTE, ref * config.TOLERANCE_PERCENT / 100)
    cr.status = "PASS" if diff <= tol else "FAIL"
    return cr


def compare_all(
    enip_map:   dict[str, EnipResult],
    modbus_map: dict[str, ModbusResult],
    keys:       list[str],
) -> list[CompareResult]:
    results = []
    for key in keys:
        er = enip_map.get(key)
        mr = modbus_map.get(key)
        if er is None or mr is None:
            continue
        results.append(_compare_one(key, er, mr))
    return results


# ─────────────────────────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────────────────────────

async def run_comparison(
    param_keys: Optional[list[str]] = None,
) -> tuple[ScopeReport, list[CompareResult]]:
    t0 = time.time()

    device    = config.DEVICE_NAME
    eds_path  = EDS_PATH_MAP.get(device)
    enip_host = ENIP_HOST_MAP.get(device)

    if not eds_path or not enip_host:
        raise ValueError(f"设备 '{device}' 未在 EDS_PATH_MAP / ENIP_HOST_MAP 中配置")

    # ── 1. 模板 SNMP 范围 ────────────────────────────────────────────────────
    tmpl_snmp  = get_snmp_params(config.TEMPLATE_DIR, device)
    tmpl_set   = set(tmpl_snmp)

    # ── 2. EDS Assembly 范围 ─────────────────────────────────────────────────
    eds_params = parse_eds(eds_path)
    eds_set    = {p.name for p in eds_params}
    eds_order  = [p.name for p in eds_params]   # 保持 Assembly 顺序

    # ── 3. 参数范围检查 ──────────────────────────────────────────────────────
    matched         = sorted(tmpl_set & eds_set)
    missing_from_eds = sorted(tmpl_set - eds_set)
    extra_in_eds    = sorted(eds_set - tmpl_set)
    scope_report = ScopeReport(
        template_count   = len(tmpl_set),
        eds_count        = len(eds_set),
        matched_keys     = matched,
        missing_from_eds = missing_from_eds,
        extra_in_eds     = extra_in_eds,
    )
    log.info("范围检查：模板SNMP=%d  EDS=%d  匹配=%d  缺失=%d  多余=%d",
             len(tmpl_set), len(eds_set), len(matched),
             len(missing_from_eds), len(extra_in_eds))

    # ── 4. 确定数值比对目标 ──────────────────────────────────────────────────
    async with EnipReader(enip_host, eds_path, ENIP_SLOT) as enip, \
               get_reader() as modbus:

        modbus_keys_set = set(modbus.known_params())
        matched_set     = set(matched)

        if param_keys:
            target = [k for k in param_keys
                      if k in matched_set and k in modbus_keys_set]
        else:
            # 按 Assembly 顺序，只取模板 SNMP & Modbus 均有的参数
            target = [k for k in eds_order
                      if k in matched_set and k in modbus_keys_set]

        log.info("数值比对参数数：%d", len(target))

        # ── 5. 并发读取 ──────────────────────────────────────────────────────
        enip_task   = asyncio.create_task(enip.read_all())
        modbus_task = asyncio.create_task(modbus.read_params(target))
        enip_map, modbus_results = await asyncio.gather(enip_task, modbus_task)

    modbus_map = {r.param_key: r for r in modbus_results}
    results    = compare_all(enip_map, modbus_map, target)

    log.info("比对完成，耗时 %.1f 秒", time.time() - t0)
    return scope_report, results


# ─────────────────────────────────────────────────────────────────────────────
# 统计 & 控制台输出
# ─────────────────────────────────────────────────────────────────────────────

def summary(results: list[CompareResult]) -> dict:
    by_status: dict[str, int] = {}
    for r in results:
        by_status[r.status] = by_status.get(r.status, 0) + 1
    total  = len(results)
    passed = by_status.get("PASS", 0)
    failed = by_status.get("FAIL", 0)
    errors = total - passed - failed
    fail_list = sorted([r for r in results if r.status == "FAIL"],
                       key=lambda r: r.diff_pct or 0, reverse=True)
    return {
        "total": total, "pass": passed, "fail": failed, "error": errors,
        "pass_rate": f"{passed/total*100:.1f}%" if total else "N/A",
        "by_status": by_status, "worst_fails": fail_list[:10],
    }


def print_summary(scope: ScopeReport, results: list[CompareResult]) -> None:
    s = summary(results)
    print("\n" + "=" * 70)
    print(f"  EtherNet/IP vs Modbus 比对摘要")
    print(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    print(f"  【参数范围】模板SNMP={scope.template_count}  EDS={scope.eds_count}  "
          f"匹配={len(scope.matched_keys)}  "
          f"缺失={len(scope.missing_from_eds)}  多余={len(scope.extra_in_eds)}")
    if scope.missing_from_eds:
        print(f"  缺失（前10）: {scope.missing_from_eds[:10]}")
    if scope.extra_in_eds:
        print(f"  多余（前10）: {scope.extra_in_eds[:10]}")
    print(f"  【数值比对】总={s['total']}  PASS={s['pass']} ({s['pass_rate']})  "
          f"FAIL={s['fail']}  ERR={s['error']}")
    if s["worst_fails"]:
        print("\n  差异最大参数（Top 10）：")
        for r in s["worst_fails"]:
            print(f"    {r.param_key:40s}  EIP={r.enip_value:.6g}  "
                  f"Modbus={r.modbus_value:.6g}  Δ%={r.diff_pct:.2f}%")
    print("=" * 70)


# ─────────────────────────────────────────────────────────────────────────────
# HTML 报告
# ─────────────────────────────────────────────────────────────────────────────

_STATUS_COLOR = {
    "PASS":       "#d4edda",
    "FAIL":       "#f8d7da",
    "ENIP_ERR":   "#fff3cd",
    "MODBUS_ERR": "#fff3cd",
    "BOTH_ERR":   "#e2e3e5",
}
_STATUS_LABEL = {
    "PASS":       "通过",
    "FAIL":       "失败",
    "ENIP_ERR":   "EIP异常",
    "MODBUS_ERR": "Modbus异常",
    "BOTH_ERR":   "双路异常",
}


def _fmt(v: Optional[float], d: int = 6) -> str:
    return "—" if v is None else f"{v:.{d}g}"


def generate_html_report(
    scope: ScopeReport,
    results: list[CompareResult],
    output_path: Optional[str] = None,
) -> str:
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = Path(config.REPORT_DIR)
        report_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(report_dir / f"enip_compare_{config.DEVICE_NAME}_{ts}.html")

    s = summary(results)
    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    enip_host = ENIP_HOST_MAP.get(config.DEVICE_NAME, "?")

    # ── Section 1: 参数范围检查 HTML ─────────────────────────────────────────
    scope_ok    = scope.scope_ok
    scope_color = "#d4edda" if scope_ok else "#f8d7da"
    scope_label = "一致" if scope_ok else "不一致"

    def _list_rows(keys: list[str], bg: str) -> str:
        return "".join(
            f'<tr style="background:{bg}"><td class="num">{i+1}</td>'
            f'<td class="key">{html.escape(k)}</td></tr>'
            for i, k in enumerate(keys)
        )

    missing_html = (
        f'<h3 style="color:#721c24;margin-top:12px">模板有但 EDS 缺少（{len(scope.missing_from_eds)} 条）</h3>'
        f'<table><colgroup><col style="width:48px"><col></colgroup>'
        f'<thead><tr><th>#</th><th>param_key</th></tr></thead>'
        f'<tbody>{_list_rows(scope.missing_from_eds, "#f8d7da")}</tbody></table>'
    ) if scope.missing_from_eds else ""

    extra_html = (
        f'<h3 style="color:#856404;margin-top:12px">EDS 有但模板未标 SNMP（{len(scope.extra_in_eds)} 条）</h3>'
        f'<table><colgroup><col style="width:48px"><col></colgroup>'
        f'<thead><tr><th>#</th><th>param_key</th></tr></thead>'
        f'<tbody>{_list_rows(scope.extra_in_eds, "#fff3cd")}</tbody></table>'
    ) if scope.extra_in_eds else ""

    scope_badge = (f'<span class="badge err-badge">缺失 {len(scope.missing_from_eds)}</span>'
                   if scope.missing_from_eds else "") + \
                  (f'<span class="badge warn-badge">多余 {len(scope.extra_in_eds)}</span>'
                   if scope.extra_in_eds else "") + \
                  ('<span class="badge ok-badge">一致</span>' if scope_ok else "")

    scope_section = f"""
<details open class="section">
<summary>一、参数范围检查（模板 SNMP 列 vs EDS Assembly）
  <span class="sum-info">模板 {scope.template_count} / EDS {scope.eds_count} / 匹配 {len(scope.matched_keys)}</span>
  {scope_badge}
</summary>
<div class="section-body">
<div class="cards">
  <div class="card total"><div class="num">{scope.template_count}</div><div class="lbl">模板 SNMP 参数</div></div>
  <div class="card total"><div class="num">{scope.eds_count}</div><div class="lbl">EDS Assembly</div></div>
  <div class="card pass"> <div class="num">{len(scope.matched_keys)}</div><div class="lbl">匹配</div></div>
  <div class="card fail"> <div class="num">{len(scope.missing_from_eds)}</div><div class="lbl">模板有/EDS缺</div></div>
  <div class="card err">  <div class="num">{len(scope.extra_in_eds)}</div><div class="lbl">EDS多/模板无</div></div>
  <div class="card" style="background:{scope_color}"><div class="num" style="font-size:18px">{scope_label}</div><div class="lbl">范围结论</div></div>
</div>
{missing_html}{extra_html}
</div>
</details>"""

    # ── Section 2: 数值比对 HTML ─────────────────────────────────────────────
    rows_html = []
    for i, r in enumerate(results):
        bg   = _STATUS_COLOR.get(r.status, "#fff")
        stat = _STATUS_LABEL.get(r.status, r.status)
        err  = ""
        if r.enip_error:   err += f"EIP: {html.escape(r.enip_error)}"
        if r.modbus_error: err += f"{'<br>' if err else ''}Modbus: {html.escape(r.modbus_error)}"
        rows_html.append(f"""
        <tr style="background:{bg}">
          <td class="num">{i+1}</td>
          <td class="key">{html.escape(r.param_key)}</td>
          <td class="val">{_fmt(r.enip_value)}</td>
          <td class="val">{_fmt(r.modbus_value)}</td>
          <td class="val">{_fmt(r.diff_abs, 4) if r.diff_abs is not None else "—"}</td>
          <td class="val">{f"{r.diff_pct:.3f}%" if r.diff_pct is not None else "—"}</td>
          <td class="stat">{stat}</td>
          <td class="err">{err}</td>
        </tr>""")

    val_badge = (f'<span class="badge err-badge">失败 {s["fail"]}</span>' if s["fail"] else "") + \
                (f'<span class="badge warn-badge">异常 {s["error"]}</span>' if s["error"] else "") + \
                ('<span class="badge ok-badge">全部通过</span>' if not s["fail"] and not s["error"] else "")

    val_section = f"""
<details open class="section">
<summary>二、数值比对（EtherNet/IP Assembly vs Modbus TCP）
  <span class="sum-info">共 {s['total']} 项 · 通过率 {s['pass_rate']}</span>
  {val_badge}
</summary>
<div class="section-body">
<div class="cards">
  <div class="card total"><div class="num">{s['total']}</div><div class="lbl">总参数</div></div>
  <div class="card pass"> <div class="num">{s['pass']}</div><div class="lbl">通过 ({s['pass_rate']})</div></div>
  <div class="card fail"> <div class="num">{s['fail']}</div><div class="lbl">失败</div></div>
  <div class="card err">  <div class="num">{s['error']}</div><div class="lbl">异常</div></div>
</div>
<table>
<colgroup>
  <col class="c-idx"><col class="c-key">
  <col class="c-val"><col class="c-val">
  <col class="c-diff"><col class="c-pct">
  <col class="c-stat"><col class="c-err">
</colgroup>
<thead>
  <tr>
    <th>#</th><th>参数名 (param_key)</th>
    <th>EtherNet/IP 值</th><th>Modbus 值</th>
    <th>绝对差值</th><th>相对差值</th>
    <th>结果</th><th>错误信息</th>
  </tr>
</thead>
<tbody>{"".join(rows_html)}</tbody>
</table>
</div>
</details>"""

    content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>EtherNet/IP vs Modbus 比对报告 — {html.escape(config.DEVICE_NAME)}</title>
<style>
  body {{ font-family:"Microsoft YaHei",Arial,sans-serif; font-size:13px; margin:20px; background:#f5f5f5; color:#333; }}
  h1   {{ font-size:20px; margin-bottom:4px; }}
  .device-name {{ font-size:15px; font-weight:bold; color:#0056b3; margin-bottom:4px; }}
  .meta {{ color:#666; font-size:12px; margin-bottom:16px; }}
  .cards {{ display:flex; gap:12px; margin-bottom:16px; flex-wrap:wrap; }}
  .card {{ background:#fff; border-radius:6px; padding:14px 20px; box-shadow:0 1px 4px rgba(0,0,0,.1); min-width:110px; text-align:center; }}
  .card .num {{ font-size:28px; font-weight:bold; }}
  .card .lbl {{ font-size:11px; color:#888; margin-top:2px; }}
  .card.pass .num {{ color:#28a745; }} .card.fail .num {{ color:#dc3545; }}
  .card.err  .num {{ color:#ffc107; }} .card.total .num {{ color:#007bff; }}
  table {{ border-collapse:collapse; width:100%; background:#fff; box-shadow:0 1px 4px rgba(0,0,0,.1); border-radius:6px; overflow:hidden; table-layout:fixed; margin-bottom:24px; }}
  colgroup col.c-idx {{ width:48px; }} colgroup col.c-key {{ width:280px; }}
  colgroup col.c-val {{ width:110px; }} colgroup col.c-diff {{ width:90px; }}
  colgroup col.c-pct {{ width:80px; }} colgroup col.c-stat {{ width:90px; }}
  colgroup col.c-err {{ width:auto; }}
  thead tr {{ background:#343a40; color:#fff; }}
  th {{ padding:9px 8px; text-align:center; font-size:12px; white-space:nowrap; }}
  td {{ padding:5px 8px; border-bottom:1px solid #eee; vertical-align:middle; }}
  td.num  {{ text-align:right; color:#999; font-size:11px; }}
  td.key  {{ font-family:monospace; font-size:12px; word-break:break-all; }}
  td.val  {{ text-align:right; font-family:monospace; font-size:12px; }}
  td.stat {{ text-align:center; font-weight:bold; font-size:12px; }}
  td.err  {{ font-size:11px; color:#666; word-break:break-all; }}
  tr:hover td {{ filter:brightness(0.96); }}
  thead th {{ position:sticky; top:0; z-index:1; }}
  details.section {{ background:#fff; border-radius:8px; margin-bottom:20px; box-shadow:0 1px 4px rgba(0,0,0,.1); overflow:hidden; }}
  details.section > summary {{ list-style:none; cursor:pointer; padding:12px 16px;
    background:#f0f4f8; border-left:4px solid #0056b3;
    font-size:15px; font-weight:bold; color:#0056b3;
    display:flex; align-items:center; gap:8px; user-select:none; }}
  details.section > summary::-webkit-details-marker {{ display:none; }}
  details.section > summary::before {{ content:"▶"; font-size:10px; margin-right:4px; }}
  details[open].section > summary::before {{ content:"▼"; }}
  .sum-info {{ font-size:12px; font-weight:normal; color:#666; margin-left:4px; }}
  .badge {{ display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:bold; margin-left:4px; }}
  .ok-badge   {{ background:#d4edda; color:#155724; }}
  .err-badge  {{ background:#f8d7da; color:#721c24; }}
  .warn-badge {{ background:#fff3cd; color:#856404; }}
  .section-body {{ padding:16px; }}
</style>
</head>
<body>
<h1>EtherNet/IP vs Modbus TCP 比对报告</h1>
<div class="device-name">设备：{html.escape(config.DEVICE_NAME)}</div>
<div class="meta">
  生成时间：{now_str} &nbsp;|&nbsp;
  EtherNet/IP：{enip_host}:44818 &nbsp;|&nbsp;
  Modbus：{config.MODBUS_HOST}:{config.MODBUS_PORT} Unit={config.MODBUS_UNIT} &nbsp;|&nbsp;
  容差：±{config.TOLERANCE_PERCENT}% / ±{config.TOLERANCE_ABSOLUTE}
</div>
{scope_section}
{val_section}
</body>
</html>"""

    Path(output_path).write_text(content, encoding="utf-8")
    log.info("HTML 报告已保存：%s", output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# 命令行入口
# ─────────────────────────────────────────────────────────────────────────────

async def _main(param_keys: Optional[list[str]], quick: bool) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )

    if quick and param_keys is None:
        from modbus_reader import get_param_map
        param_keys = list(get_param_map().keys())[:30]
        log.info("快速模式：仅比对前 %d 个参数", len(param_keys))

    # 初始化 Modbus 连接参数
    if config.DEVICE_NAME in config.MODBUS_DEVICE_MAP:
        config.MODBUS_HOST, config.MODBUS_PORT, config.MODBUS_UNIT = \
            config.MODBUS_DEVICE_MAP[config.DEVICE_NAME]

    scope, results = await run_comparison(param_keys)
    print_summary(scope, results)
    report_path = generate_html_report(scope, results)
    print(f"\n  HTML 报告：{report_path}\n")


if __name__ == "__main__":
    quick = "--quick" in sys.argv
    if "--keys" in sys.argv:
        idx  = sys.argv.index("--keys")
        keys = sys.argv[idx + 1:]
    else:
        keys = None

    asyncio.run(_main(keys, quick))
