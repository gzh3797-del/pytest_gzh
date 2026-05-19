能不能从产出# -*- coding: utf-8 -*-
"""
template_reader.py — 读取设备参数模板文件（Template/*.xlsx）

blockParams sheet 关键列：
  paramType        — 参数标识（如 FREQ_Hz）
  descrption       — 参数描述（模板原始拼写，如 System Frequency）
  unit             — 工程单位（如 Hz）
  BACnetIP         — 非空 → 该参数在 BACnet/IP 北向发布
  AcuCloud         — 非空 → 通用模板中的 AcuCloud 列（⚠️ 范围不准，仅供参考）

AcuCloud 比对参数范围请使用专用模板目录（AcuCloud 模板适配/）：
  paramType_AcuCloud — 各设备 AcuCloud 实际比对参数标识
  对于 AcuRev-4100，该列在 Excel 中为第二个名为 paramType 的列。
  使用 get_cloud_acucloud_params() 读取，不要使用 get_cloud_params()。
"""
from __future__ import annotations

import re
import warnings
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class TemplateParam:
    param_key:   str   # paramType 列
    description: str   # descrption 列（模板原文拼写）
    unit:        str   # unit 列（空字符串表示无单位）
    bacnet_ip:   str   # BACnetIP 列原始值（非空 → BACnet 发布）
    acucloud:    str   # AcuCloud 列原始值（非空 → AcuCloud 同步）
    mqtt:        str = ""     # MQTT 列原始值（非空 → MQTT 发布）
    datalog:     str = ""     # DataLog 列原始值（非空 → Datalog 记录）
    snmp:        str = ""     # SNMP 列原始值（非空 → SNMP 发布）


def _norm(s: str) -> str:
    """小写并去除连字符/下划线/空格，用于模糊匹配文件名。"""
    return s.lower().replace("-", "").replace("_", "").replace(" ", "")


def natural_sort_key(s: str) -> list:
    """自然排序键，使参数名中的数字按数值比较（如 DI_2 < DI_10）。"""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]


def find_template_file(template_dir: str, device_name: str) -> str:
    """
    在 template_dir 下递归查找设备模板文件（大小写/连字符/下划线不敏感）。
    例：device_name='AcuRev4100' → 匹配 'AcuRev-4100_v1.01_20260427.xlsx'
    """
    needle = _norm(device_name)
    for p in Path(template_dir).rglob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        if needle in _norm(p.stem):
            return str(p)
    raise FileNotFoundError(
        f"在 {template_dir} 中未找到设备 '{device_name}' 的模板文件"
    )


def load_template(xlsx_path: str) -> list[TemplateParam]:
    """解析模板 blockParams sheet，返回全部参数列表。"""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = wb["blockParams"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def _col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            raise ValueError(f"模板缺少列 '{name}'，当前列：{headers}")

    ci_key     = _col("paramType")
    ci_desc    = _col("descrption")   # 模板原始拼写（少一个 i）
    ci_unit    = _col("unit")
    ci_bacnet  = _col("BACnetIP")
    ci_cloud   = _col("AcuCloud")
    ci_mqtt    = headers.index("MQTT")    if "MQTT"    in headers else -1
    ci_datalog = headers.index("DataLog") if "DataLog" in headers else -1
    ci_snmp    = headers.index("SNMP")    if "SNMP"    in headers else -1

    def _str(row: tuple, ci: int) -> str:
        return str(row[ci]).strip() if ci >= 0 and row[ci] is not None else ""

    params: list[TemplateParam] = []
    for row in rows[1:]:
        key = row[ci_key]
        if not key:
            continue
        params.append(TemplateParam(
            param_key   = str(key).strip(),
            description = _str(row, ci_desc),
            unit        = _str(row, ci_unit),
            bacnet_ip   = _str(row, ci_bacnet),
            acucloud    = _str(row, ci_cloud),
            mqtt        = _str(row, ci_mqtt),
            datalog     = _str(row, ci_datalog),
            snmp        = _str(row, ci_snmp),
        ))

    return params


def get_bacnet_params(xlsx_path: str) -> list[TemplateParam]:
    """返回 BACnetIP 列非空的参数（BACnet/IP 发布范围）。"""
    return [p for p in load_template(xlsx_path) if p.bacnet_ip]


def get_cloud_params(xlsx_path: str) -> list[TemplateParam]:
    """返回 AcuCloud 列非空的参数（AcuCloud 同步范围）。"""
    return [p for p in load_template(xlsx_path) if p.acucloud]


def get_mqtt_params(xlsx_path: str) -> list[TemplateParam]:
    """返回 MQTT 列非空的参数（MQTT 发布范围）。"""
    return [p for p in load_template(xlsx_path) if p.mqtt]


def get_datalog_params(xlsx_path: str) -> list[TemplateParam]:
    """返回 DataLog 列非空的参数（Datalog 记录范围）。"""
    return [p for p in load_template(xlsx_path) if p.datalog]


def get_snmp_params(xlsx_path: str) -> list[TemplateParam]:
    """返回 SNMP 列非空的参数（SNMP / EtherNet/IP 发布范围）。"""
    return [p for p in load_template(xlsx_path) if p.snmp]


def get_bacnet_params_by_range(xlsx_path: str, range_marker: str) -> list[TemplateParam]:
    """
    返回 range 列包含 range_marker 的参数（AcuIOM 设备 BACnet 参数范围）。

    用于没有 BACnetIP 列的设备模板（如 AcuIOM），以 range 列中的协议标识号来确定
    BACnet 发布范围。AcuIOM-01/02 使用 range_marker="8"，AcuIOM-03/04 使用 "10"。
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = wb["blockParams"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    try:
        ci_key   = headers.index("paramType")
        ci_desc  = headers.index("descrption")
        ci_unit  = headers.index("unit")
        ci_range = headers.index("range")
    except ValueError as exc:
        raise ValueError(
            f"AcuIOM 模板 {xlsx_path} 缺少必要列：{exc}，当前列：{headers}"
        )

    def _str(row: tuple, ci: int) -> str:
        return str(row[ci]).strip() if ci >= 0 and row[ci] is not None else ""

    params: list[TemplateParam] = []
    for row in rows[1:]:
        key = row[ci_key]
        if not key:
            continue
        range_val = str(row[ci_range]).strip() if row[ci_range] is not None else ""
        if range_marker not in range_val:
            continue
        params.append(TemplateParam(
            param_key   = str(key).strip(),
            description = _str(row, ci_desc),
            unit        = _str(row, ci_unit),
            bacnet_ip   = "bacnet",  # 非空占位，表示属于 BACnet 范围
            acucloud    = "",
        ))
    return params


def get_cloud_acucloud_params(xlsx_path: str) -> list[TemplateParam]:
    """
    从 AcuCloud 专用模板文件（AcuCloud 模板适配/）读取 AcuCloud 比对参数范围。

    要求文件含 paramType_AcuCloud 列；若不存在则抛出 ValueError。
    调用方可捕获异常并回退到 get_cloud_params()（如 AcuRev-4100 暂未迁移时）。

    ⚠️ 不要用 get_cloud_params() 替代本函数——通用模板的 AcuCloud 列范围不准确。
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = wb["blockParams"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    if "paramType_AcuCloud" not in headers:
        raise ValueError(
            f"AcuCloud 模板 {xlsx_path} 中未找到 paramType_AcuCloud 列；"
            f"该设备可能尚未迁移，请回退到通用模板的 AcuCloud 列。"
        )
    ci_ac = headers.index("paramType_AcuCloud")

    params: list[TemplateParam] = []
    for row in rows[1:]:
        if ci_ac >= len(row):
            continue
        key = row[ci_ac]
        if not key:
            continue
        params.append(TemplateParam(
            param_key   = str(key).strip(),
            description = "",
            unit        = "",
            bacnet_ip   = "",
            acucloud    = "acucloud",   # 非空占位，表示属于 AcuCloud 范围
        ))

    return params
